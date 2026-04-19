"""
GST COMPLETE SUITE v11 — AY 2025-26
====================================
Based on ACTUAL portal screenshots + official GST documentation

CONFIRMED PORTAL FLOW (from screenshots + docs):
=================================================
LOGIN:
  www.gst.gov.in → Click LOGIN → username + password + CAPTCHA → Submit

RETURNS DASHBOARD:
  Services → Returns → Returns Dashboard
  Select FY + Quarter + Period → SEARCH
  Tiles appear:
    GSTR-1  → Status: Filed → VIEW | DOWNLOAD → JSON only
    GSTR-1A → PREPARE ONLINE (only if supplier amended their GSTR-1)
    GSTR-2B → VIEW | DOWNLOAD → Excel (INSTANT — no wait, same as GSTR-3B)
    GSTR-2A → VIEW | DOWNLOAD → JSON + EXCEL (generates in ~15-20 mins — wait required)
    GSTR-3B → Status: Filed → VIEW GSTR3B | DOWNLOAD → PDF only

DOWNLOAD TYPE CLASSIFICATION:
  ⚡ DIRECT / INSTANT (No 20-min generate wait):
      GSTR-2B → Click DOWNLOAD tile → Click GENERATE EXCEL → File ready INSTANTLY
      GSTR-3B → Click DOWNLOAD tile → PDF downloads directly

  ⏳ GENERATE-FIRST (15-20 min wait required — then download link appears):
      GSTR-1  → Click DOWNLOAD tile → Click GENERATE JSON → Wait → Collect link
      GSTR-1A → Click DOWNLOAD tile → Click GENERATE JSON → Wait → Collect link
      GSTR-2A → Click DOWNLOAD tile → Click GENERATE EXCEL → Wait → Collect link

PHASED DOWNLOAD FLOW (Optimised — all 12 months, one session, no logout):
  Phase 1 → Trigger GENERATE for GSTR-1 + GSTR-2A (all 12 months) — click & move on
  Phase 2 → Direct download GSTR-2B (all 12 months) — INSTANT, keeps session alive
  Phase 3 → Direct download GSTR-3B (all 12 months) — INSTANT, keeps session alive
  Phase 4 → Collect DOWNLOAD LINK for GSTR-1 + GSTR-2A (now ready after ~20 min)
  Total   → ~20 minutes for all 12 months × all 4 returns = 48 files

FAST CASES (No generate at all — GSTR-2B + GSTR-3B both instant):
  Option 15 → GSTR-2B + GSTR-3B only → 24 tabs → ~10 min
  Option 4  → GSTR-2B only           → 12 tabs → ~5  min
  Option 6  → GSTR-3B only           → 12 tabs → ~5  min

RETRY LOGIC — T1 / T2 / T3 (per month per return):
  T1 → immediate attempt
  T2 → wait 30s → retry
  T3 → wait 60s → retry
  All 3 fail → logged, move to next month

SESSION SAFETY:
  Portal timeout = 20 mins idle
  Phase 2 + Phase 3 (instant downloads) keep session alive during generate wait
  Not Found page → auto F5 reload up to 3 times
  Batch size = max 6 months per batch (avoids portal Not Found overload)

ALL RESULTS SAVED TO:
  Downloads/GST_Automation/AY2025-26_YYYYMMDD_HHMM/ClientName/
    GSTR1_April_2024.zip
    GSTR2B_April_2024.xlsx
    GSTR2A_April_2024.xlsx
    GSTR3B_April_2024.pdf
  MASTER_REPORT_YYYYMMDD_HHMM.xlsx  ← summary of all clients
"""

import os, sys, time, json, logging, zipfile
from datetime import datetime
from pathlib import Path

MISSING = []
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.edge.service import Service as EdgeService
    from selenium.webdriver.edge.options import Options as EdgeOptions
    from selenium.webdriver.chrome.service import Service as ChromeService
    from selenium.webdriver.chrome.options import Options as ChromeOptions
except ImportError:
    MISSING.append("selenium")

try:
    from webdriver_manager.chrome import ChromeDriverManager
    CHROME_MGR = True
except:
    CHROME_MGR = False

try:
    import pandas as pd
except ImportError:
    MISSING.append("pandas")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    MISSING.append("openpyxl")

# -- Constants ----------------------------------------------
PAGE_WAIT       = 4    # was 8  — halved; smart waits replace fixed sleeps
SHORT_WAIT      = 1    # was 3
ACTION_WAIT     = 1    # was 2
CLIENT_GAP      = 5    # was 10
FILE_GEN_WAIT   = 300  # was 600 — 5 min max per file (portal usually <90s)
FILE_GEN_RETRY  = 10   # was 20
FY_LABEL        = "2025-26"

MONTHS = [
    ("April","04","2025"),    ("May","05","2025"),      ("June","06","2025"),
    ("July","07","2025"),     ("August","08","2025"),   ("September","09","2025"),
    ("October","10","2025"),  ("November","11","2025"), ("December","12","2025"),
    ("January","01","2026"),  ("February","02","2026"), ("March","03","2026"),
]

QUARTER_MAP = {
    "April":"Quarter 1 (Apr - Jun)", "May":"Quarter 1 (Apr - Jun)",   "June":"Quarter 1 (Apr - Jun)",
    "July":"Quarter 2 (Jul - Sep)",  "August":"Quarter 2 (Jul - Sep)","September":"Quarter 2 (Jul - Sep)",
    "October":"Quarter 3 (Oct - Dec)","November":"Quarter 3 (Oct - Dec)","December":"Quarter 3 (Oct - Dec)",
    "January":"Quarter 4 (Jan - Mar)","February":"Quarter 4 (Jan - Mar)","March":"Quarter 4 (Jan - Mar)",
}

# Excel styling
DARK_BLUE="1F3864"; MED_BLUE="2E75B6"
GREEN_BG="C6EFCE";  GREEN_FG="276221"
RED_BG="FFC7CE";    RED_FG="9C0006"
YELLOW_BG="FFEB9C"; YELLOW_FG="9C6500"
GREY_BG="F2F2F2";   WHITE="FFFFFF"

def fill(h): return PatternFill("solid", fgColor=h)
def fnt(b=False,c="000000",s=10): return Font(name="Arial",bold=b,color=c,size=s)
def bdr(): s=Side(style="thin"); return Border(left=s,right=s,top=s,bottom=s)
def aln(h="center",wrap=True): return Alignment(horizontal=h,vertical="center",wrap_text=wrap)
def sc(cell,bold=False,fg="000000",bg=None,size=10,h="center"):
    cell.font=fnt(bold,fg,size)
    if bg: cell.fill=fill(bg)
    cell.alignment=aln(h=h); cell.border=bdr()

def clean_str(s):
    if s is None or str(s).strip() in ["nan","None",""]: return ""
    return str(s).strip()

def clean_num(s):
    try: return float(str(s).replace(",","").replace("₹","").strip())
    except: return 0.0

def get_latest_file(folder, extensions):
    """Return most recently modified file with given extensions."""
    files = []
    for ext in extensions:
        files.extend(Path(folder).glob(f"*{ext}"))
    if not files: return None
    return max(files, key=lambda f: f.stat().st_mtime)

def rename_latest(folder, new_name, extensions, log):
    try:
        f = get_latest_file(folder, extensions)
        if f:
            dest = Path(folder)/new_name
            if not dest.exists():
                f.rename(dest)
                if log: log.info(f"      Saved: {new_name}")
            return True
    except Exception as e:
        if log: log.warning(f"      Rename failed: {e}")
    return False


# ==========================================================
# LOGGER
# ==========================================================
def setup_logger(log_dir):
    log_path = os.path.join(log_dir, f"run_log_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)-8s | %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ]
    )
    return logging.getLogger("gst_v5")


# ==========================================================
# FIND EDGE DRIVER
# ==========================================================
def find_edge_driver():
    import shutil
    p = shutil.which("msedgedriver")
    if p: return p
    for path in [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "msedgedriver.exe"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "msedgedriver.EXE"),
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedgedriver.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedgedriver.exe",
    ]:
        if os.path.exists(path): return path
    return None


# ==========================================================
# BROWSER SETUP
# ==========================================================
def make_driver(download_dir):
    dl_dir = str(download_dir)
    edge_path = find_edge_driver()
    if edge_path:
        try:
            opts = EdgeOptions()
            opts.add_experimental_option("prefs", {
                "download.default_directory":       dl_dir,
                "download.prompt_for_download":     False,
                "download.directory_upgrade":       True,
                "safebrowsing.enabled":             True,
                "credentials_enable_service":       False,
                "profile.password_manager_enabled": False,
            })
            opts.add_argument("--start-maximized")
            opts.add_argument("--disable-blink-features=AutomationControlled")
            opts.add_argument("--disable-save-password-bubble")
            opts.add_argument("--disable-features=msEdgeEnhancedSecurityMode")
            opts.add_experimental_option("excludeSwitches", ["enable-automation","enable-logging"])
            opts.add_experimental_option("useAutomationExtension", False)
            opts.add_argument(
                "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/146.0.0.0 Safari/537.36 Edg/146.0.3856.78"
            )
            driver = webdriver.Edge(service=EdgeService(edge_path), options=opts)
            driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                "source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
            })
            print("  Browser: Microsoft Edge ✓")
            return driver
        except Exception as e:
            print(f"  Edge failed: {e} — trying Chrome...")

    try:
        opts = ChromeOptions()
        opts.add_experimental_option("prefs", {
            "download.default_directory":       dl_dir,
            "download.prompt_for_download":     False,
            "download.directory_upgrade":       True,
            "credentials_enable_service":       False,
            "profile.password_manager_enabled": False,
        })
        opts.add_argument("--start-maximized")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation","enable-logging"])
        opts.add_experimental_option("useAutomationExtension", False)
        svc = ChromeService(ChromeDriverManager().install()) if CHROME_MGR else ChromeService()
        driver = webdriver.Chrome(service=svc, options=opts)
        driver.execute_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
        print("  Browser: Google Chrome ✓")
        return driver
    except Exception as e:
        print(f"\n  Both Edge and Chrome failed: {e}")
        sys.exit(1)


# ==========================================================
# HELPERS
# ==========================================================
def try_click(driver, xpaths, timeout=8, log=None):
    for xp in xpaths:
        try:
            el = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.XPATH, xp)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            time.sleep(0.3)
            try: el.click()
            except: driver.execute_script("arguments[0].click();", el)
            if log: log.info(f"    Clicked: {xp[:60]}")
            return True
        except: continue
    return False

def human_type(driver, by, val, text, log=None):
    try:
        el = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((by, val)))
        driver.execute_script("arguments[0].scrollIntoView(true);", el)
        time.sleep(0.3); el.click(); time.sleep(0.2)
        el.clear(); time.sleep(0.2)
        for ch in str(text): el.send_keys(ch); time.sleep(0.03)
        time.sleep(0.3)
        if el.get_attribute("value") or "":
            if log: log.info(f"    Typed via {val} ✓")
            return True
        # JS fallback
        driver.execute_script(
            "arguments[0].value=arguments[1];"
            "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));"
            "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
            el, text)
        return True
    except Exception as e:
        if log: log.warning(f"    Type failed {val}: {e}")
        return False

def select_by_text(driver, text, log=None):
    """Find any dropdown containing text and select it."""
    for sel_el in driver.find_elements(By.TAG_NAME, "select"):
        try:
            s = Select(sel_el)
            for opt in s.options:
                if text.lower() in opt.text.lower():
                    s.select_by_visible_text(opt.text)
                    if log: log.info(f"    Dropdown selected: {opt.text} ✓")
                    return True
        except: continue
    return False

# ==========================================================
# ── NEW: SESSION KEEP-ALIVE ────────────────────────────────
# Prevents 20-min GST portal timeout during GENERATE waits
# ==========================================================
def keep_session_alive(driver, log=None):
    """
    Ping the portal by scrolling the current page.
    Call this every 5-7 mins during GENERATE wait phase
    so session does not expire.
    """
    try:
        driver.execute_script("window.scrollBy(0, 100);")
        time.sleep(0.5)
        driver.execute_script("window.scrollBy(0, -100);")
        if log: log.info("    [KeepAlive] Session pinged ✓")
        return True
    except Exception as e:
        if log: log.warning(f"    [KeepAlive] Failed: {e}")
        return False


# ==========================================================
# ── NEW: NOT-FOUND PAGE HANDLER ───────────────────────────
# Handles "Not Found" / blank page due to fast tab opening
# ==========================================================
def handle_not_found_page(driver, log=None, max_reload=3):
    """
    If portal shows 'Not Found' / blank / error page, reload.
    Waits for full page load after each reload.
    Returns True if page recovered, False if all reloads failed.
    """
    not_found_indicators = [
        "not found", "404", "page not available",
        "server error", "bad gateway", "service unavailable"
    ]
    for attempt in range(1, max_reload + 1):
        try:
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            if any(ind in body for ind in not_found_indicators) or len(body.strip()) < 30:
                if log: log.warning(f"    [NotFound] Attempt {attempt}/{max_reload} — reloading page...")
                driver.refresh()
                # Wait for full page load
                WebDriverWait(driver, 20).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
                time.sleep(PAGE_WAIT)
                body = driver.find_element(By.TAG_NAME, "body").text.lower()
                if not any(ind in body for ind in not_found_indicators) and len(body.strip()) > 30:
                    if log: log.info(f"    [NotFound] Page recovered on attempt {attempt} ✓")
                    return True
            else:
                return True  # Page is fine
        except Exception as e:
            if log: log.warning(f"    [NotFound] Reload error: {e}")
            time.sleep(2)
    if log: log.error("    [NotFound] Page did not recover after all reloads")
    return False


# ==========================================================
# ── NEW: RETRY WRAPPER — T1 / T2 / T3 ────────────────────
# Wraps any download action with 3 retries + backoff
# Usage discussed: try 1 → wait 30s → try 2 → wait 1min → try 3
# ==========================================================
def retry_download(action_fn, label="download", log=None, max_retries=3):
    """
    Retry an action up to 3 times (T1, T2, T3).
    action_fn: callable that returns True on success, False on failure.
    Returns (success: bool, attempts_used: int)
    """
    wait_times = [0, 30, 60]   # T1: no wait, T2: 30s, T3: 60s
    for attempt in range(1, max_retries + 1):
        wait = wait_times[attempt - 1]
        if wait > 0:
            if log: log.info(f"    [Retry T{attempt}] Waiting {wait}s before retry for {label}...")
            time.sleep(wait)
        try:
            if log: log.info(f"    [Retry T{attempt}] Attempting {label}...")
            result = action_fn()
            if result:
                if log: log.info(f"    [Retry T{attempt}] ✓ Success: {label}")
                return True, attempt
            else:
                if log: log.warning(f"    [Retry T{attempt}] ✗ Failed: {label}")
        except Exception as e:
            if log: log.warning(f"    [Retry T{attempt}] ✗ Error: {e}")
    if log: log.error(f"    All {max_retries} retries exhausted for {label}. Try early morning 7-9 AM.")
    return False, max_retries


# ==========================================================
# ── NEW: PHASED DOWNLOAD COORDINATOR ─────────────────────
# Implements the exact flow we designed:
#
#  Phase 1: Click GENERATE on GSTR-1 & GSTR-2A (all months)
#  Phase 2: Direct download GSTR-2B (keeps session alive)
#  Phase 3: Direct download GSTR-3B (keeps session alive)
#  Phase 4: Come back, click DOWNLOAD LINK for GSTR-1 & GSTR-2A
#
# This way session never times out and all downloads happen
# in ~20 minutes total for 12 months.
#
# BATCH RULE: Open max 6 months per batch to avoid Not Found
# ==========================================================
def phased_download_coordinator(driver, client, client_dir, log, returns_todo,
                                 navigate_to_month_fn, download_direct_fn,
                                 generate_fn, collect_generated_fn,
                                 batch_size=6):
    """
    Phased download flow:
      Phase 1 → Trigger GENERATE for all months (GSTR-1, GSTR-2A) in batches
      Phase 2 → Direct download GSTR-2B all months (session stays alive)
      Phase 3 → Direct download GSTR-3B all months (session stays alive)
      Phase 4 → Collect GENERATE links for GSTR-1, GSTR-2A (now ready)

    Parameters:
      navigate_to_month_fn(month_tuple) → navigates to returns dashboard for that month
      download_direct_fn(rtype, month_tuple) → downloads a direct-download return (2B, 3B)
      generate_fn(rtype, month_tuple) → clicks GENERATE button (1, 2A, 1A)
      collect_generated_fn(rtype, month_tuple) → downloads the generated file link (1, 2A, 1A)
      batch_size → max months to process before pausing (default 6)

    Returns dict: {rtype: {month_name: "ok"|"failed"|"skipped"}}
    """
    results = {r: {} for r in returns_todo}

    GENERATE_RETURNS = [r for r in ("GSTR1", "GSTR1A", "GSTR2A") if r in returns_todo]
    DIRECT_RETURNS   = [r for r in ("GSTR2B", "GSTR3B") if r in returns_todo]

    if log:
        log.info("  ═══ PHASED DOWNLOAD START ═══")
        log.info(f"  Generate-first returns : {GENERATE_RETURNS}")
        log.info(f"  Direct download returns: {DIRECT_RETURNS}")
        log.info(f"  Batch size             : {batch_size} months per batch")

    # ── PHASE 1: Trigger GENERATE for all months ──────────────────
    if GENERATE_RETURNS:
        if log: log.info("\n  ── Phase 1: Triggering GENERATE for all months ──")
        for batch_start in range(0, len(MONTHS), batch_size):
            batch = MONTHS[batch_start : batch_start + batch_size]
            if log: log.info(f"    Batch {batch_start//batch_size + 1}: {[m[0] for m in batch]}")
            for month_tuple in batch:
                month_name = month_tuple[0]
                for rtype in GENERATE_RETURNS:
                    try:
                        navigate_to_month_fn(month_tuple)
                        handle_not_found_page(driver, log)
                        ok = generate_fn(rtype, month_tuple)
                        results[rtype][month_name] = "generating" if ok else "gen_failed"
                        if log: log.info(f"    Generate {rtype} {month_name}: {'triggered ✓' if ok else '✗ failed'}")
                        time.sleep(SHORT_WAIT)
                    except Exception as e:
                        results[rtype][month_name] = "gen_error"
                        if log: log.warning(f"    Generate {rtype} {month_name} error: {e}")
            time.sleep(ACTION_WAIT)  # small gap between batches

    # ── PHASE 2+3: Direct download 2B and 3B (keeps session alive) ─
    if DIRECT_RETURNS:
        if log: log.info("\n  ── Phase 2/3: Direct downloading (keeps session alive) ──")
        for rtype in DIRECT_RETURNS:
            if log: log.info(f"    Downloading {rtype} — all months...")
            for batch_start in range(0, len(MONTHS), batch_size):
                batch = MONTHS[batch_start : batch_start + batch_size]
                for month_tuple in batch:
                    month_name = month_tuple[0]
                    def _do_direct(rt=rtype, mt=month_tuple):
                        navigate_to_month_fn(mt)
                        handle_not_found_page(driver, log)
                        return download_direct_fn(rt, mt)
                    ok, tries = retry_download(_do_direct, f"{rtype} {month_name}", log)
                    results[rtype][month_name] = "ok" if ok else f"failed(T{tries})"
                    keep_session_alive(driver, log)   # ping after each month
                    time.sleep(SHORT_WAIT)

    # ── PHASE 4: Collect generated files (GSTR-1, 2A now ready) ───
    if GENERATE_RETURNS:
        if log: log.info("\n  ── Phase 4: Collecting GENERATE download links ──")
        # Check if enough time has passed — if not, wait with keep-alive pings
        gen_wait_start = time.time()
        gen_wait_max   = FILE_GEN_WAIT   # 5 mins max (portal usually <2 min)
        for month_tuple in MONTHS:
            month_name = month_tuple[0]
            for rtype in GENERATE_RETURNS:
                if results.get(rtype, {}).get(month_name) not in ("generating",):
                    continue  # skip if generate was not triggered

                def _do_collect(rt=rtype, mt=month_tuple):
                    navigate_to_month_fn(mt)
                    handle_not_found_page(driver, log)
                    return collect_generated_fn(rt, mt)

                # Keep pinging session while waiting
                elapsed = time.time() - gen_wait_start
                if elapsed < 60:
                    if log: log.info(f"    Waiting for generate to complete... ({int(60-elapsed)}s remaining)")
                    for _ in range(int(60 - elapsed) // 10):
                        keep_session_alive(driver, log)
                        time.sleep(10)

                ok, tries = retry_download(_do_collect, f"{rtype} {month_name} link", log)
                results[rtype][month_name] = "ok" if ok else f"failed(T{tries})"
                keep_session_alive(driver, log)
                time.sleep(SHORT_WAIT)

    # ── Summary ───────────────────────────────────────────────────
    if log:
        log.info("\n  ═══ PHASED DOWNLOAD COMPLETE ═══")
        for rtype, months_res in results.items():
            ok_count   = sum(1 for v in months_res.values() if v == "ok")
            fail_count = sum(1 for v in months_res.values() if "failed" in v)
            log.info(f"    {rtype}: {ok_count} ok | {fail_count} failed")

    return results


def wait_for_download_link(driver, timeout_seconds, log):
    """
    After clicking GENERATE JSON/EXCEL, the portal shows:
      - First: "Request is being processed" or spinning icon
      - Then:  A table row with a DOWNLOAD link (href ending in .zip or .json)

    Key: we must REFRESH the page every 30s to check if file is ready.
    A real download link has href containing 'filedownload' or ends in .zip/.json
    We must NOT pick up the GENERATE button itself as a link.

    Portal behaviour (confirmed from logs):
      - File usually ready in 30s - 2 mins
      - Page shows "Click here" or a table with download icon/link
    """
    log.info(f"    Waiting for file to be generated (refreshing every 30s, max {timeout_seconds}s)...")

    def find_real_download_link():
        """
        Return a real download link element — NOT the generate button.
        Real links have href with 'filedownload', '.zip', '.json' or 'download' path.
        """
        try:
            # Look for links with actual file hrefs
            all_links = driver.find_elements(By.TAG_NAME, "a")
            for el in all_links:
                try:
                    if not el.is_displayed():
                        continue
                    href = el.get_attribute("href") or ""
                    text = el.text.strip().lower()
                    # Must have a real file URL — not just a page link
                    is_file_link = (
                        "filedownload" in href.lower() or
                        href.lower().endswith(".zip") or
                        href.lower().endswith(".json") or
                        href.lower().endswith(".xlsx") or
                        ("download" in href.lower() and "gst" in href.lower() and len(href) > 60)
                    )
                    is_download_text = any(x in text for x in [
                        "click here", "download", "here to download"
                    ])
                    if is_file_link or (is_download_text and len(href) > 40):
                        return el
                except: continue
        except: pass
        return None

    def page_still_processing():
        """Check if portal is still generating (spinner or 'processing' text)."""
        try:
            page = driver.page_source.lower()
            return any(x in page for x in [
                "request is being processed",
                "being processed",
                "processing",
                "please wait",
                "generating",
            ])
        except:
            return False

    elapsed = 0
    refresh_interval = 30  # refresh page every 30 seconds

    while elapsed < timeout_seconds:
        # Check immediately after generate click
        time.sleep(3)
        elapsed += 3

        link = find_real_download_link()
        if link:
            log.info(f"    ✅ Download link ready after {elapsed}s")
            return link

        if page_still_processing():
            log.info(f"    Portal still processing... ({elapsed}s)")

        # Wait in 30s intervals, refreshing page each time
        wait_chunk = min(refresh_interval, timeout_seconds - elapsed)
        for _ in range(wait_chunk):
            time.sleep(1)
            elapsed += 1

        # Refresh page to check if file is ready
        log.info(f"    Refreshing page to check status... ({elapsed}s elapsed)")
        try:
            driver.refresh()
            time.sleep(4)
        except: pass

        link = find_real_download_link()
        if link:
            log.info(f"    ✅ Download link ready after {elapsed}s (found after refresh)")
            return link

    log.warning(f"    ⚠ File not ready after {timeout_seconds}s — moving on")
    return None


# ==========================================================
# SESSION STATE — stores current client creds for auto re-login
# ==========================================================
CURRENT_CLIENT = {}   # populated in process_client before phases start


def is_session_lost(driver):
    """Return True if portal has logged us out or shown access denied."""
    try:
        url = driver.current_url.lower()
        if "accessdenied" in url:
            return True
        if "login" in url and "fowelcome" not in url and "gst.gov.in" in url:
            return True
        # Check page text for session-expired messages
        body = driver.find_element(By.TAG_NAME, "body").text.lower()
        for phrase in ["session expired", "session has expired",
                       "you are not logged in", "please login again",
                       "access denied"]:
            if phrase in body:
                return True
    except:
        pass
    return False


def relogin_if_needed(driver, log):
    """
    Full re-login from www.gst.gov.in when session is lost or Access Denied.
    Flow: www.gst.gov.in → LOGIN button → username + password + CAPTCHA → Submit
    Returns True if re-login succeeded, False if failed.
    """
    if not is_session_lost(driver):
        return True   # session still valid

    log.warning("  ⚠ Session expired / Access Denied — going back to Login page...")

    username = CURRENT_CLIENT.get("username", "")
    password = CURRENT_CLIENT.get("password", "")

    # If no credentials stored, ask the user to type them right now
    if not username:
        import getpass
        print("\n  No stored username found. Please enter credentials manually.")
        log.warning("  No credentials in CURRENT_CLIENT — prompting user to enter manually")
        username = input("  Username: ").strip()
        password = getpass.getpass("  Password: ").strip()
        CURRENT_CLIENT["username"] = username
        CURRENT_CLIENT["password"] = password

    print()
    print("  " + "="*56)
    print("  ⚠  ACCESS DENIED / SESSION EXPIRED")
    print("  " + "-"*56)
    print(f"  Client  : {username}")
    print("  Action  : Navigating to www.gst.gov.in → Login page")
    print("  Please type the CAPTCHA in the browser when prompted.")
    print("  " + "="*56)

    # do_login navigates to www.gst.gov.in, clicks LOGIN, fills username +
    # password, waits for CAPTCHA input, then clicks LOGIN button.
    # It has its own 3-attempt retry loop with credential re-entry on denial.
    return do_login(driver, username, password, log)


def safe_go_to_dashboard(driver, log):
    """
    Navigate to Returns Dashboard with full re-login on Access Denied / session loss.
    Flow:
      1. If session is already lost  → full re-login FIRST, then navigate
      2. Try normal navigation        (Services → Returns → Returns Dashboard)
      3. If navigation fails/denied   → full re-login again, then retry once
    """
    # ── Step 1: Re-login BEFORE navigating if session already lost ──
    if is_session_lost(driver):
        log.warning("  Session lost before dashboard nav — re-logging in first...")
        print("\n" + "!"*56)
        print("  SESSION EXPIRED — Re-login required before continuing")
        print("!"*56)
        if not relogin_if_needed(driver, log):
            log.error("  Re-login failed — cannot navigate to dashboard")
            return False

    # ── Step 2: Attempt normal navigation ──────────────────────────
    result = go_to_returns_dashboard(driver, log)

    # ── Step 3: If navigation failed or Access Denied → re-login + retry ──
    if not result or is_session_lost(driver):
        log.warning("  Dashboard navigation failed/Access Denied — re-logging in...")
        print("\n" + "!"*56)
        print("  ACCESS DENIED during navigation — going back to Login page")
        print("!"*56)
        if relogin_if_needed(driver, log):
            result = go_to_returns_dashboard(driver, log)
        else:
            result = False

    return result


# ==========================================================
# GENERATE + INSTANT DOWNLOAD (GSTR-2B — same flow as GSTR-3B)
# Portal flow:  tile DOWNLOAD  →  Generate page
#               Click GENERATE EXCEL  →  file downloads INSTANTLY
#               No link appears. No page polling. Pure folder watch.
# ==========================================================
def generate_then_download_immediate(driver, client_dir, save_name, log,
                                      gen_xpaths=None, max_wait=60):
    """
    GSTR-2B INSTANT DOWNLOAD — identical flow to GSTR-3B:
      1. Click tile DOWNLOAD  →  lands on Generate page
      2. Click GENERATE EXCEL →  file downloads INSTANTLY (no link, no polling page)
      3. Poll folder every 0.5s until file appears  →  rename  →  done

    No link scanning. No page polling. Pure file-system watch like GSTR-3B.
    max_wait: max seconds to wait for the file after clicking Generate (default 60s).
    """
    if gen_xpaths is None:
        gen_xpaths = [
            "//button[contains(text(),'GENERATE EXCEL FILE TO DOWNLOAD')]",
            "//button[contains(text(),'GENERATE EXCEL')]",
            "//button[contains(text(),'Generate Excel')]",
            "//button[contains(text(),'GENERATE JSON FILE TO DOWNLOAD')]",
            "//button[contains(text(),'Generate JSON')]",
        ]

    client_path = Path(client_dir)

    def _snap(exts):
        """Snapshot current files in folder."""
        return {str(f): f.stat().st_mtime
                for f in client_path.iterdir()
                if f.suffix.lower() in exts}

    def _first_new(before, exts):
        """Return first complete new file that appeared since snapshot."""
        for f in client_path.iterdir():
            if f.suffix.lower() not in exts: continue
            if f.name.endswith((".crdownload", ".tmp", ".part")): continue
            prev = before.get(str(f))
            if (prev is None or f.stat().st_mtime > prev + 0.1) and f.stat().st_size > 500:
                return f
        return None

    EXTS = {".xlsx", ".zip", ".json"}

    # 1. Wait briefly for Generate page to render (replaces old sleep(SHORT_WAIT))
    time.sleep(0.8)
    log.info(f"    GSTR-2B generate page: {driver.current_url}")

    # 2. Snapshot BEFORE clicking Generate
    before = _snap(EXTS)

    # 3. Click GENERATE EXCEL — triggers instant browser download
    gen_clicked = try_click(driver, gen_xpaths, timeout=8, log=log)
    if gen_clicked:
        log.info("    ✅ GENERATE EXCEL clicked — watching folder for instant file...")
    else:
        log.warning("    ⚠ GENERATE EXCEL button not found on page")

    # 4. Fast-poll folder every 0.5s (same as GSTR-3B fast poll)
    deadline = time.time() + max_wait
    while time.time() < deadline:
        time.sleep(0.5)
        new_f = _first_new(before, EXTS)
        if new_f:
            log.info(f"    ⚡ File received instantly: {new_f.name}")
            time.sleep(0.5)   # let Chrome finish flushing
            rename_latest(client_dir, save_name, [".xlsx", ".zip", ".json"], log)
            return True

    log.warning(f"    ⚠ GSTR-2B file not received within {max_wait}s for {save_name}")
    return False



def do_login(driver, username, password, log):
    """
    Login to GST portal with up to 3 attempts.
    On Access Denied or wrong credentials, prompts user to retype
    username and password before retrying.
    """
    MAX_ATTEMPTS = 3

    for attempt in range(1, MAX_ATTEMPTS + 1):

        # ── Navigate to portal & click LOGIN ──────────────────────
        log.info(f"    Opening www.gst.gov.in (attempt {attempt}/{MAX_ATTEMPTS})...")
        driver.get("https://www.gst.gov.in")
        time.sleep(2)  # was 4

        log.info("    Clicking LOGIN button...")
        try_click(driver, [
            "//a[normalize-space()='LOGIN']",
            "//a[normalize-space()='Login']",
            "//button[normalize-space()='LOGIN']",
            "//a[contains(@href,'login')]",
        ], timeout=8, log=log)
        time.sleep(PAGE_WAIT)  # wait for login page
        log.info(f"    Login page: {driver.current_url}")

        # ── Fill username ──────────────────────────────────────────
        log.info(f"    Filling username: {username}")
        filled = False
        for by, val in [
            (By.ID,   "username"),
            (By.NAME, "username"),
            (By.ID,   "user_name"),
            (By.NAME, "user_name"),
            (By.CSS_SELECTOR, "input[placeholder*='sername']"),
            (By.CSS_SELECTOR, "input[type='text']:not([readonly])"),
        ]:
            if human_type(driver, by, val, username, log):
                filled = True; break
        if not filled:
            log.error("    Cannot fill username field!"); return False

        time.sleep(ACTION_WAIT)

        # ── Fill password ──────────────────────────────────────────
        log.info("    Filling password...")
        filled = False
        for by, val in [
            (By.ID,   "user_pass"),
            (By.NAME, "user_pass"),
            (By.ID,   "password"),
            (By.NAME, "password"),
            (By.CSS_SELECTOR, "input[type='password']"),
        ]:
            if human_type(driver, by, val, password, log):
                filled = True; break
        if not filled:
            log.error("    Cannot fill password field!"); return False

        time.sleep(ACTION_WAIT)

        # ── CAPTCHA — manual ───────────────────────────────────────
        print()
        print("  " + "="*52)
        print(f"  CLIENT  : {username}")
        if attempt > 1:
            print(f"  ATTEMPT : {attempt} of {MAX_ATTEMPTS}  (previous login was denied)")
        print("  " + "-"*52)
        print("  1. Look at the browser window")
        print("  2. Type CAPTCHA letters in the browser")
        print("  3. Come back here and press ENTER")
        print("  NOTE: Do NOT click Login — script will do it")
        print("  " + "="*52)
        input("  >> Press ENTER after typing CAPTCHA: ")

        # ── Click LOGIN ────────────────────────────────────────────
        log.info("    Clicking LOGIN button...")
        try_click(driver, [
            "//button[@id='btnlogin']",
            "//button[normalize-space()='LOGIN']",
            "//button[normalize-space()='Login']",
            "//button[@type='submit']",
            "//input[@type='submit']",
            "//button[contains(text(),'LOGIN')]",
            "//button[contains(text(),'Login')]",
        ], timeout=8, log=log)

        # Smart wait: poll for dashboard/home after login (max 6s)
        try:
            from selenium.webdriver.support.ui import WebDriverWait
            WebDriverWait(driver, 6).until(
                lambda d: "login" not in d.current_url.lower() or "fowelcome" in d.current_url.lower()
            )
        except: time.sleep(3)

        # ── OTP if needed ──────────────────────────────────────────
        try:
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            if "otp" in body and ("enter" in body or "verify" in body):
                print("\n  OTP REQUIRED — Enter OTP in browser then press ENTER here")
                input("  >> Press ENTER after OTP entered: ")
                time.sleep(2)
        except: pass

        # ── Check result ───────────────────────────────────────────
        cur = driver.current_url.lower()
        log.info(f"    Post-login URL: {driver.current_url}")

        login_failed = (
            "accessdenied" in cur or
            ("login" in cur and "fowelcome" not in cur)
        )

        if not login_failed:
            log.info("    Login SUCCESSFUL ✓")
            return True

        # ── Login denied — prompt user to retype credentials ───────
        log.warning(f"    Login DENIED (attempt {attempt}/{MAX_ATTEMPTS})")
        print()
        print("  " + "!"*52)
        print("  ✗  ACCESS DENIED — Username or Password is wrong")
        print("  " + "!"*52)

        if attempt < MAX_ATTEMPTS:
            print(f"\n  Attempt {attempt} of {MAX_ATTEMPTS} failed.")
            print("  Please retype the credentials below.")
            print("  (Press ENTER to keep the current value)\n")

            import getpass
            new_user = input(f"  Username [{username}]: ").strip()
            new_pass = getpass.getpass(f"  Password (leave blank to keep current): ").strip()

            if new_user:
                username = new_user
            if new_pass:
                password = new_pass

            # Update global CURRENT_CLIENT so relogin_if_needed also uses new creds
            global CURRENT_CLIENT
            CURRENT_CLIENT["username"] = username
            CURRENT_CLIENT["password"] = password

            log.info(f"    Retrying with username: {username}")
            print(f"\n  Retrying login for: {username} ...\n")
        else:
            print(f"\n  All {MAX_ATTEMPTS} login attempts failed for: {username}")
            print("  Skipping this client. Check credentials in clients.xlsx/csv.")
            log.error(f"    Login FAILED after {MAX_ATTEMPTS} attempts — skipping client")

    return False


# ==========================================================
# GO TO RETURNS DASHBOARD
# Services → Returns → Returns Dashboard
# ==========================================================
def go_to_returns_dashboard(driver, log):
    """
    Navigate: Services --> Returns --> Returns Dashboard
    Always follows the proper menu path after login.
    Returns False immediately if session is lost — caller must handle re-login.
    """
    cur = driver.current_url

    # Already on dashboard — nothing to do
    if "return.gst.gov.in" in cur and "dashboard" in cur:
        log.info("    Already on Returns Dashboard +")
        return True

    # ── EARLY EXIT: session already lost — do NOT try Services navigation ──
    # Caller (safe_go_to_dashboard) will trigger full re-login.
    if is_session_lost(driver):
        log.warning("    Session/Access Denied detected at nav start — returning False for re-login")
        return False

    log.info("    Navigating: Services --> Returns --> Returns Dashboard")

    # ── Step 0: If browser drifted off the portal, use browser Back / logo
    # to get back — NEVER use a direct URL (direct URLs trigger Access Denied).
    if "gst.gov.in" not in cur:
        log.info("    Not on GST portal — pressing Back to return to portal...")
        try:
            driver.back()
            time.sleep(2)
        except: pass
        # If still off-portal after Back, the session is likely gone
        if "gst.gov.in" not in driver.current_url:
            log.warning("    Still off portal after Back — treating as session loss")
            return False
        if is_session_lost(driver):
            log.warning("    Session lost after Back — returning False for re-login")
            return False

    def _click_services():
        """Click Services in the top nav — try XPath then JS scan."""
        for attempt in range(3):
            clicked = try_click(driver, [
                "//a[normalize-space(text())='Services']",
                "//li[contains(@class,'nav')]//a[normalize-space()='Services']",
                "//nav//a[normalize-space()='Services']",
                "//ul[contains(@class,'nav')]//a[contains(text(),'Services')]",
            ], timeout=6, log=log)
            if clicked:
                log.info("    Services menu clicked +"); return True
            try:
                driver.execute_script("""
                    var links=document.querySelectorAll('a,li');
                    for(var i=0;i<links.length;i++){
                        var t=(links[i].innerText||links[i].textContent||'').trim();
                        if(t==='Services'){
                            links[i].dispatchEvent(new MouseEvent('mouseover',{bubbles:true}));
                            links[i].click(); break;
                        }
                    }
                """)
                time.sleep(0.5)
            except: pass
            time.sleep(0.5)  # was 1.5
        return False

    def _click_returns_menu():
        """Click Returns in the Services dropdown."""
        for attempt in range(3):
            clicked = try_click(driver, [
                "//a[normalize-space(text())='Returns']",
                "//*[contains(@class,'dropdown-menu')]//a[normalize-space()='Returns']",
                "//*[contains(@class,'open')]//a[normalize-space()='Returns']",
                "//ul[contains(@class,'open')]//a[contains(text(),'Returns')]",
            ], timeout=6, log=log)
            if clicked:
                log.info("    Returns clicked +"); return True
            try:
                driver.execute_script("""
                    var links=document.querySelectorAll('a');
                    for(var i=0;i<links.length;i++){
                        var t=(links[i].innerText||'').trim();
                        if(t==='Returns'&&links[i].offsetParent!==null){links[i].click();break;}
                    }
                """)
                time.sleep(0.5)
            except: pass
            time.sleep(0.5)  # was 1.5
        return False

    def _click_dashboard_link():
        """Click Returns Dashboard in the Returns submenu."""
        for attempt in range(3):
            clicked = try_click(driver, [
                "//a[contains(normalize-space(text()),'Returns Dashboard')]",
                "//li//a[contains(@href,'dashboard')]",
            ], timeout=6, log=log)
            if clicked:
                log.info("    Returns Dashboard clicked +"); return True
            try:
                for el in driver.find_elements(By.TAG_NAME, "a"):
                    try:
                        if "Returns Dashboard" in (el.text or "") and el.is_displayed():
                            driver.execute_script("arguments[0].click();", el)
                            log.info("    Returns Dashboard clicked (scan) +")
                            return True
                    except: continue
            except: pass
            time.sleep(0.5)  # was 1.5
        return False

    # ── Steps 1–3: Services → Returns → Returns Dashboard ────────────
    # Attempt the full menu path up to 2 times before giving up.
    # On failure we return False — caller (safe_go_to_dashboard) will
    # trigger a full re-login; we never fall back to a direct URL.
    for nav_attempt in range(1, 3):
        log.info(f"    Step 1: Clicking Services menu (nav attempt {nav_attempt})...")
        _click_services()
        time.sleep(0.5)  # was 1.5

        log.info("    Step 2: Clicking Returns...")
        _click_returns_menu()
        time.sleep(0.5)  # was 1.5

        log.info("    Step 3: Clicking Returns Dashboard...")
        _click_dashboard_link()

        # Smart wait: poll for dashboard URL (max 6s) instead of fixed PAGE_WAIT
        try:
            WebDriverWait(driver, 6).until(
                lambda d: "return.gst.gov.in" in d.current_url and "dashboard" in d.current_url
            )
        except: time.sleep(2)

        final_url = driver.current_url
        log.info(f"    URL after nav attempt {nav_attempt}: {final_url}")

        if "accessdenied" in final_url.lower():
            log.warning(f"    Access Denied after nav attempt {nav_attempt} — returning False for re-login")
            return False

        if "dashboard" in final_url.lower() and "return.gst.gov.in" in final_url.lower():
            log.info("    Returns Dashboard loaded OK +")
            return True

        log.warning(f"    Nav attempt {nav_attempt} did not land on dashboard — retrying menu path...")
        time.sleep(1)  # was 2

    log.warning("    Could not reach Returns Dashboard via menu — returning False for re-login")
    return False


def select_and_search(driver, month_name, log):
    """Select FY/Quarter/Period dropdowns then click SEARCH. Uses smart waits — no fixed sleeps."""
    log.info(f"    Setting: FY={FY_LABEL}  Quarter={QUARTER_MAP.get(month_name,'')}  Period={month_name}")

    # Wait for dropdowns to appear (max 4s)
    try:
        WebDriverWait(driver, 4).until(
            lambda d: len(d.find_elements(By.TAG_NAME, "select")) >= 1
        )
    except: pass

    all_sels = driver.find_elements(By.TAG_NAME, "select")

    # FY
    for sel_el in all_sels:
        try:
            s = Select(sel_el)
            opts = [o.text.strip() for o in s.options]
            if any("-" in o and len(o) <= 9 for o in opts):
                for opt in s.options:
                    if FY_LABEL in opt.text:
                        s.select_by_visible_text(opt.text)
                        log.info(f"    FY: {opt.text} ✓")
                        break
                break
        except: continue

    # Quarter — re-fetch selects after FY change
    all_sels = driver.find_elements(By.TAG_NAME, "select")
    qtr = QUARTER_MAP.get(month_name, "")
    for sel_el in all_sels:
        try:
            s = Select(sel_el)
            opts = [o.text.strip() for o in s.options]
            if any("quarter" in o.lower() for o in opts):
                for opt in s.options:
                    if qtr[:9].lower() in opt.text.lower():
                        s.select_by_visible_text(opt.text)
                        log.info(f"    Quarter: {opt.text} ✓")
                        break
                break
        except: continue

    # Period/Month
    all_sels = driver.find_elements(By.TAG_NAME, "select")
    month_names_lower = ["january","february","march","april","may","june",
                         "july","august","september","october","november","december"]
    for sel_el in all_sels:
        try:
            s = Select(sel_el)
            opts = [o.text.strip() for o in s.options]
            if any(m in " ".join(opts).lower() for m in month_names_lower):
                for opt in s.options:
                    if month_name.lower() in opt.text.lower():
                        s.select_by_visible_text(opt.text)
                        log.info(f"    Period: {opt.text} ✓")
                        break
                break
        except: continue

    # SEARCH
    clicked = try_click(driver, [
        "//button[normalize-space()='SEARCH']",
        "//button[normalize-space()='Search']",
        "//button[contains(text(),'SEARCH')]",
        "//input[@value='SEARCH']",
    ], timeout=8, log=log)

    if not clicked:
        try:
            driver.execute_script("""
                var btns=document.querySelectorAll('button,input[type=submit]');
                for(var i=0;i<btns.length;i++){
                    if((btns[i].innerText||btns[i].value||'').toUpperCase().includes('SEARCH')){
                        btns[i].click(); break;
                    }
                }
            """)
            log.info("    SEARCH clicked via JS")
        except: pass

    # Smart wait: stop as soon as GSTR tiles appear (max 6s)
    try:
        WebDriverWait(driver, 6).until(
            lambda d: any(t in d.find_element(By.TAG_NAME,"body").text
                          for t in ["GSTR-1","GSTR-2","GSTR-3","GSTR1","GSTR2","GSTR3"])
        )
    except: time.sleep(2)
    log.info("    Tiles loaded ✓")


# ==========================================================
# CLICK TILE + DOWNLOAD BUTTON
# After SEARCH, tiles appear: GSTR-1, GSTR-2B, GSTR-2A, GSTR-3B
# Each tile has VIEW and DOWNLOAD buttons
# ==========================================================
def click_tile_download(driver, tile_name, log):
    """
    Find the tile for tile_name and click its DOWNLOAD button.
    
    From actual portal screenshot (Image 2), tile subtitle texts are:
      GSTR-1  tile subtitle = "GSTR1"   (no dash)
      GSTR-2B tile subtitle = "GSTR2B"  (no dash)
      GSTR-2A tile subtitle = "GSTR2A"  (no dash)  
      GSTR-3B tile subtitle = "GSTR-3B" (with dash)
    
    Strategy: find the subtitle element, walk UP to tile container,
    then find DOWNLOAD button INSIDE that container only.
    This prevents clicking wrong tile's button.
    """
    log.info(f"    Finding {tile_name} tile DOWNLOAD button...")
    # No fixed sleep — tiles should already be loaded from select_and_search smart wait

    # Exact subtitle texts seen in portal screenshots
    # Map input name → possible subtitle texts on page
    name_variants = {
        "GSTR1":  ["GSTR1", "GSTR-1"],
        "GSTR1A": ["GSTR1A", "GSTR-1A"],
        "GSTR2B": ["GSTR2B", "GSTR-2B"],
        "GSTR2A": ["GSTR2A", "GSTR-2A"],
        "GSTR3B": ["GSTR3B", "GSTR-3B"],
    }
    variants = name_variants.get(tile_name.upper().replace("-",""), [tile_name])

    for variant in variants:
        try:
            # Find the subtitle text element
            subtitle_els = driver.find_elements(By.XPATH,
                f"//*[normalize-space(text())='{variant}']"
            )
            for subtitle_el in subtitle_els:
                if not subtitle_el.is_displayed():
                    continue
                # Walk UP parent tree to find the tile container
                # (the container that has both the title and the buttons)
                parent = subtitle_el
                for level in range(6):
                    try:
                        parent = driver.execute_script(
                            "return arguments[0].parentElement;", parent)
                        if parent is None:
                            break
                        # Check if this parent has a DOWNLOAD button
                        btns = parent.find_elements(By.XPATH,
                            ".//button[contains(translate(text(),'download','DOWNLOAD'),'DOWNLOAD')] | "
                            ".//a[contains(translate(text(),'download','DOWNLOAD'),'DOWNLOAD')]"
                        )
                        for btn in btns:
                            if btn.is_displayed():
                                driver.execute_script(
                                    "arguments[0].scrollIntoView({block:'center'});", btn)
                                time.sleep(0.2)
                                driver.execute_script("arguments[0].click();", btn)
                                log.info(f"    {tile_name} ({variant}) DOWNLOAD clicked at level {level} ✓")
                                return True
                    except: break
        except: continue

    log.warning(f"    {tile_name} tile DOWNLOAD button not found on page")
    return False


# ==========================================================
# GENERATE FILE AND DOWNLOAD
# After clicking DOWNLOAD on tile → new page shows:
#   GENERATE JSON FILE TO DOWNLOAD
#   GENERATE EXCEL FILE TO DOWNLOAD  (for 2B/2A)
# Click generate → wait → download link appears → click it
# ==========================================================
def click_generate_only(driver, log):
    """
    Phase 1: Click GENERATE JSON button and move on immediately.
    Do NOT wait for file — portal generates in background.
    Returns True if button was clicked.
    """
    time.sleep(0.5)  # was SHORT_WAIT=3
    log.info(f"    Generate page: {driver.current_url}")
    clicked = try_click(driver, [
        "//button[contains(text(),'GENERATE JSON FILE TO DOWNLOAD')]",
        "//button[contains(text(),'GENERATE JSON')]",
        "//button[contains(text(),'Generate JSON')]",
        "//a[contains(text(),'GENERATE JSON')]",
    ], timeout=8, log=log)
    if clicked:
        log.info("    GENERATE JSON clicked — moving on (file generates in background)")
        time.sleep(0.5)  # was 2
    else:
        log.warning("    GENERATE JSON button not found")
    return clicked

def download_ready_file(driver, client_dir, save_name, log):
    """
    Phase 2: Find 'Click here to download' link and download the ready ZIP.
    Uses fast 0.5s file polling instead of fixed sleep(8).
    Returns True if downloaded successfully.
    """
    time.sleep(0.5)  # was SHORT_WAIT=3
    log.info(f"    Checking for download link on: {driver.current_url}")

    client_path = Path(client_dir)
    before = {str(f): f.stat().st_mtime for f in client_path.iterdir()
              if f.suffix.lower() in {".zip", ".json", ".xlsx"}}

    for xp in [
        "//a[contains(text(),'Click here to download')]",
        "//a[contains(text(),'click here to download')]",
        "//a[contains(text(),'File 1')]",
        "//a[contains(text(),'File 2')]",
        "//a[contains(@href,'.zip')]",
        "//a[contains(@href,'filedownload')]",
        "//a[contains(@href,'download') and string-length(@href) > 50]",
    ]:
        try:
            els = driver.find_elements(By.XPATH, xp)
            for el in els:
                if el.is_displayed():
                    href = el.get_attribute("href") or ""
                    txt  = el.text.strip()
                    log.info(f"    Found link: '{txt}' → {href[:60]}")
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    driver.execute_script("arguments[0].click();", el)
                    log.info(f"    Download clicked: {save_name} ✓")
                    # Fast poll instead of fixed sleep(8)
                    deadline = time.time() + 15
                    while time.time() < deadline:
                        time.sleep(0.5)
                        for f in client_path.iterdir():
                            if f.suffix.lower() not in {".zip", ".json", ".xlsx"}: continue
                            if f.name.endswith((".crdownload", ".tmp")): continue
                            prev = before.get(str(f))
                            if (prev is None or f.stat().st_mtime > prev + 0.1) and f.stat().st_size > 500:
                                time.sleep(0.5)
                                rename_latest(client_dir, save_name, [".zip", ".json", ".xlsx"], log)
                                return True
                    rename_latest(client_dir, save_name, [".zip", ".json", ".xlsx"], log)
                    return True
        except: continue

    log.warning(f"    No download link found for {save_name} — file may not be ready yet")
    return False

def generate_and_download(driver, file_type, client_dir, save_name, log):
    """Legacy wrapper — kept for compatibility."""
    return click_generate_only(driver, log)


# Legacy function — replaced by phase1_trigger_all + phase2_download_all
def download_month(driver, month_name, month_num, year, client_dir, log):
    return {"GSTR1":"SKIPPED","GSTR2B":"SKIPPED","GSTR2A":"SKIPPED","GSTR3B":"SKIPPED"}


# ==========================================================
# PHASE 1 — TRIGGER FILE GENERATION FOR ALL MONTHS
# For each month: navigate dashboard → select period → SEARCH
#   GSTR-1  : click DOWNLOAD → click GENERATE JSON → move on
#   GSTR-2B : click DOWNLOAD → click GENERATE EXCEL → move on
#   GSTR-2A : click DOWNLOAD → click GENERATE EXCEL → move on
#   GSTR-3B : click DOWNLOAD → PDF downloads directly → rename
# Returns dict  key="{Month}_{Year}_{RetType}" → status string
# ==========================================================
def phase1_trigger_all(driver, client_dir, log, returns_todo=None):
    """
    returns_todo: set of return types to trigger e.g. {"GSTR1","GSTR2B","GSTR3B"}
                  None means trigger all.
    """
    if returns_todo is None:
        returns_todo = {"GSTR1", "GSTR1A", "GSTR2B", "GSTR2A", "GSTR3B"}

    triggered = {}

    EXCEL_GENERATE_XPATHS = [
        "//button[contains(text(),'GENERATE EXCEL FILE TO DOWNLOAD')]",
        "//button[contains(text(),'GENERATE EXCEL')]",
        "//button[contains(text(),'Generate Excel')]",
        "//a[contains(text(),'GENERATE EXCEL')]",
        # Fall back to JSON generate if no Excel option
        "//button[contains(text(),'GENERATE JSON FILE TO DOWNLOAD')]",
        "//button[contains(text(),'GENERATE JSON')]",
        "//button[contains(text(),'Generate JSON')]",
    ]

    # Only loop through months for per-month returns (not TAX_LIABILITY which is FY-level)
    MONTHLY_RETURNS = {"GSTR1", "GSTR1A", "GSTR2B", "GSTR2A", "GSTR3B"}
    months_to_run = MONTHS if (returns_todo & MONTHLY_RETURNS) else []

    for month_name, month_num, year in months_to_run:
        key = f"{month_name}_{year}"
        log.info(f"\n  -- Phase 1 [{month_name} {year}] --")

        # -- GSTR-3B: PDF — fast 0.5s poll instead of fixed sleep(PAGE_WAIT+3) --
        if "GSTR3B" in returns_todo:
            try:
                safe_go_to_dashboard(driver, log)
                select_and_search(driver, month_name, log)
                save_name = f"GSTR3B_{month_name}_{year}.pdf"
                client_path = Path(client_dir)
                before_3b = {str(f): f.stat().st_mtime for f in client_path.iterdir()
                             if f.suffix.lower() == ".pdf"}
                if click_tile_download(driver, "GSTR3B", log):
                    # Fast poll: 0.5s intervals, max 30s
                    got_3b = False
                    deadline_3b = time.time() + 30
                    while time.time() < deadline_3b:
                        time.sleep(0.5)
                        for f in client_path.iterdir():
                            if f.suffix.lower() != ".pdf": continue
                            if f.name.endswith((".crdownload", ".tmp")): continue
                            prev = before_3b.get(str(f))
                            if (prev is None or f.stat().st_mtime > prev + 0.1) and f.stat().st_size > 500:
                                time.sleep(0.8)
                                rename_latest(client_dir, save_name, [".pdf"], log)
                                triggered[f"{key}_GSTR3B"] = "OK"
                                got_3b = True
                                break
                        if got_3b: break
                    if not got_3b:
                        triggered[f"{key}_GSTR3B"] = "NOT_FOUND"
                else:
                    triggered[f"{key}_GSTR3B"] = "TILE_FAIL"
            except Exception as e:
                log.warning(f"    GSTR3B trigger failed [{month_name}]: {e}")
                triggered[f"{key}_GSTR3B"] = f"ERR:{e}"

        # -- GSTR-1: trigger JSON generate ------------------
        if "GSTR1" in returns_todo:
            try:
                safe_go_to_dashboard(driver, log)
                select_and_search(driver, month_name, log)
                if click_tile_download(driver, "GSTR1", log):
                    time.sleep(0.5)  # was PAGE_WAIT
                    if click_generate_only(driver, log):
                        triggered[f"{key}_GSTR1"] = "TRIGGERED"
                    else:
                        triggered[f"{key}_GSTR1"] = "GEN_FAIL"
                else:
                    triggered[f"{key}_GSTR1"] = "TILE_FAIL"
            except Exception as e:
                log.warning(f"    GSTR1 trigger failed [{month_name}]: {e}")
                triggered[f"{key}_GSTR1"] = f"ERR:{e}"

        # -- GSTR-1A: trigger JSON generate -----------------
        if "GSTR1A" in returns_todo:
            try:
                safe_go_to_dashboard(driver, log)
                select_and_search(driver, month_name, log)
                if click_tile_download(driver, "GSTR1A", log):
                    time.sleep(0.5)  # was PAGE_WAIT
                    if click_generate_only(driver, log):
                        triggered[f"{key}_GSTR1A"] = "TRIGGERED"
                    else:
                        triggered[f"{key}_GSTR1A"] = "GEN_FAIL"
                else:
                    triggered[f"{key}_GSTR1A"] = "TILE_NOT_FOUND"
            except Exception as e:
                log.warning(f"    GSTR1A trigger failed [{month_name}]: {e}")
                triggered[f"{key}_GSTR1A"] = f"ERR:{e}"

        # -- GSTR-2B: ALWAYS click GENERATE EXCEL then fast-poll for instant download --
        # ROOT CAUSE FIX: Portal ALWAYS shows generate page after tile click.
        # Never wait for direct download first — always click Generate immediately.
        if "GSTR2B" in returns_todo:
            try:
                safe_go_to_dashboard(driver, log)
                select_and_search(driver, month_name, log)
                save_name = f"GSTR2B_{month_name}_{year}.xlsx"
                if click_tile_download(driver, "GSTR2B", log):
                    ok = generate_then_download_immediate(
                        driver, client_dir, save_name, log,
                        gen_xpaths=EXCEL_GENERATE_XPATHS, max_wait=120)
                    triggered[f"{key}_GSTR2B"] = "OK" if ok else "NOT_FOUND"
                else:
                    triggered[f"{key}_GSTR2B"] = "TILE_FAIL"
            except Exception as e:
                log.warning(f"    GSTR2B trigger failed [{month_name}]: {e}")
                triggered[f"{key}_GSTR2B"] = f"ERR:{e}"

        # -- GSTR-2A: trigger Excel generate ----------------
        if "GSTR2A" in returns_todo:
            try:
                safe_go_to_dashboard(driver, log)
                select_and_search(driver, month_name, log)
                if click_tile_download(driver, "GSTR2A", log):
                    time.sleep(0.5)  # was PAGE_WAIT
                    clicked = try_click(driver, EXCEL_GENERATE_XPATHS, timeout=8, log=log)
                    if clicked:
                        log.info("    GSTR-2A GENERATE clicked — moving on")
                        triggered[f"{key}_GSTR2A"] = "TRIGGERED"
                        time.sleep(0.5)  # was 2
                    else:
                        triggered[f"{key}_GSTR2A"] = "GEN_FAIL"
                else:
                    triggered[f"{key}_GSTR2A"] = "TILE_FAIL"
            except Exception as e:
                log.warning(f"    GSTR2A trigger failed [{month_name}]: {e}")
                triggered[f"{key}_GSTR2A"] = f"ERR:{e}"


    # ======================================================
    # TAX LIABILITY & ITC COMPARISON
    #
    # CONFIRMED EXACT PORTAL FLOW (from user):
    #   1. Click Services
    #   2. Move cursor → Returns (hover to open submenu)
    #   3. Move cursor → "Tax liabilities and ITC comparison" → CLICK
    #   4. Page loads: dropdown shows EMPTY → click dropdown → years appear
    #   5. Select year (2024-25) → Click SEARCH
    #   6. Scroll to BOTTOM → Click "DOWNLOAD COMPARISON REPORTS (EXCEL)"
    #   7. File downloads instantly
    #   8. Repeat for second year (2023-24)
    # ======================================================
    if "TAX_LIABILITY" in returns_todo:

        # ==================================================
        # TAX LIABILITY & ITC COMPARISON DOWNLOAD
        #
        # EXACT FLOW — identical pattern to go_to_returns_dashboard:
        #   CLICK Services  →  CLICK Returns  →  CLICK "Tax liabilities and ITC comparison"
        # Then on the /comparison page:
        #   Select FY from dropdown → SEARCH → scroll bottom → DOWNLOAD COMPARISON REPORTS (EXCEL)
        #
        # NO hover, NO direct URL (causes Access Denied)
        # Same try_click() calls used for every other menu navigation
        # ==================================================

        TAX_FY_YEARS = ["2024-25", "2025-26"]   # both FY years

        def go_to_tax_liabilities():
            """
            CONFIRMED WORKING FLOW (Selenium ActionChains — NO pyautogui needed):
              Step 1: Click Services
              Step 2: move_to_element(Returns) — CSS :hover opens submenu
              Step 3: click Tax liabilities and ITC comparison

            Uses only selenium.webdriver.common.action_chains.ActionChains
            Exactly as user confirmed: Option 1 / Visible hover inside browser.
            """
            from selenium.webdriver.common.action_chains import ActionChains
            log.info("    [TAX] Services → hover Returns → click Tax liabilities")

            # -- Step 1: Click Services ------------------------
            svc_ok = try_click(driver, [
                "//a[normalize-space(text())='Services']",
                "//nav//a[normalize-space()='Services']",
                "//ul[contains(@class,'nav')]//a[contains(text(),'Services')]",
                "//*[@id='main-nav']//a[contains(text(),'Services')]",
                "//li//a[normalize-space()='Services']",
            ], timeout=10, log=log)
            if not svc_ok:
                log.warning("    [TAX] Services not found"); return False
            time.sleep(1.5)
            log.info("    [TAX] Services clicked ✓")

            # -- Step 2: ActionChains.move_to_element(Returns) -
            # This triggers CSS :hover on the <li> → submenu slides out
            # Per user confirmed method: move cursor to Returns, pause 2s
            log.info("    [TAX] Hovering Returns via ActionChains...")
            returns_menu = None
            wait = WebDriverWait(driver, 8)
            for xp in [
                "//a[normalize-space(text())='Returns']",
                "//li[.//a[normalize-space(text())='Returns']]",
                "//*[contains(@class,'open')]//a[normalize-space()='Returns']",
                "//ul[contains(@class,'dropdown-menu')]//a[normalize-space()='Returns']",
                "//li//a[normalize-space()='Returns']",
            ]:
                try:
                    returns_menu = wait.until(
                        EC.visibility_of_element_located((By.XPATH, xp)))
                    if returns_menu.is_displayed():
                        break
                except: continue

            if returns_menu:
                actions = ActionChains(driver)
                actions.move_to_element(returns_menu).perform()
                time.sleep(1)   # allow submenu to open
                log.info("    [TAX] Hovered Returns ✓  submenu should be open")
            else:
                log.warning("    [TAX] Returns element not found — trying JS mouseenter")
                try:
                    driver.execute_script("""
                        var lis=document.querySelectorAll('li');
                        for(var li of lis){
                            var a=li.querySelector('a');
                            if(a && a.textContent.trim()==='Returns'){
                                li.dispatchEvent(new MouseEvent('mouseenter',{bubbles:true}));
                                li.dispatchEvent(new MouseEvent('mouseover',{bubbles:true}));
                                li.classList.add('open'); break;
                            }
                        }
                    """)
                    time.sleep(1.5)
                    log.info("    [TAX] JS mouseenter triggered ✓")
                except Exception as _je:
                    log.warning(f"    [TAX] JS fallback failed: {_je}")

            # -- Step 3: Click Tax liabilities ----------------
            # Exactly per user: wait.until(EC.element_to_be_clickable(...))
            log.info("    [TAX] Clicking Tax liabilities...")
            tax_xpaths = [
                "//a[normalize-space(text())='Tax liabilities and ITC comparison']",
                "//a[contains(text(),'Tax liabilities and ITC')]",
                "//a[contains(text(),'Tax liabilities')]",
                "//a[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'tax liabilities')]",
                "//li[contains(@class,'open')]//a[contains(text(),'Tax')]",
                "//*[contains(@class,'dropdown-menu')]//a[contains(text(),'Tax')]",
            ]
            for xp in tax_xpaths:
                try:
                    tax_el = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, xp)))
                    tax_el.click()
                    log.info(f"    [TAX] Clicked: {tax_el.text.strip()} ✓")
                    time.sleep(2)
                    log.info(f"    [TAX] URL: {driver.current_url}")
                    if "comparison" in driver.current_url.lower():
                        return True
                    break
                except: continue

            # -- Scan all links as fallback --------------------
            for el in driver.find_elements(By.TAG_NAME, "a"):
                try:
                    if "tax liabilities" in el.text.lower() and el.is_displayed():
                        el.click()
                        log.info(f"    [TAX] Scan click: {el.text.strip()}")
                        time.sleep(2)
                        if "comparison" in driver.current_url.lower():
                            return True
                except: continue

            # -- Retry once more -------------------------------
            log.info("    [TAX] Retry from Services...")
            for _ in range(2):
                try:
                    svc_el = WebDriverWait(driver, 6).until(
                        EC.element_to_be_clickable((By.XPATH, "//a[normalize-space(text())='Services']")))
                    svc_el.click()
                    time.sleep(0.5)
                    for xp in ["//a[normalize-space(text())='Returns']",
                               "//li//a[normalize-space()='Returns']"]:
                        try:
                            ret_el = WebDriverWait(driver, 4).until(
                                EC.visibility_of_element_located((By.XPATH, xp)))
                            ActionChains(driver).move_to_element(ret_el).perform()
                            time.sleep(2)
                            break
                        except: continue
                    for el in driver.find_elements(By.TAG_NAME, "a"):
                        try:
                            if "tax liabilities" in el.text.lower() and el.is_displayed():
                                el.click()
                                log.info(f"    [TAX] Retry click: {el.text.strip()}")
                                time.sleep(2)
                                if "comparison" in driver.current_url.lower():
                                    return True
                        except: continue
                except Exception as _re:
                    log.warning(f"    [TAX] Retry error: {_re}")

            log.warning("    [TAX] ✗ Could not reach Tax Liability page")
            try:
                vis=[e.text.strip() for e in driver.find_elements(By.TAG_NAME,"a")
                     if e.is_displayed() and e.text.strip()]
                log.warning(f"    [TAX] Visible links: {vis[:20]}")
            except: pass
            return False

        def download_tax_fy(fy_label, file_suffix):
            """
            On /comparison page:
              1. Select FY from dropdown  (dropdown starts EMPTY)
              2. Click SEARCH  →  wait for all 6 sections to load
              3. Scroll through page to find DOWNLOAD button
              4. Click DOWNLOAD COMPARISON REPORTS (EXCEL)
              5. Wait for file → rename
            """
            log.info(f"\n    -- Tax Liability Download: FY {fy_label} --")
            log.info(f"    URL: {driver.current_url}")
            time.sleep(2)

            # -- 1. Select FY ----------------------------------
            fy_set = False
            for sel_el in driver.find_elements(By.TAG_NAME, "select"):
                try:
                    driver.execute_script("arguments[0].click();", sel_el)
                    time.sleep(0.8)
                    s = Select(sel_el)
                    opts = [o.text.strip() for o in s.options if o.text.strip()]
                    log.info(f"    FY dropdown options: {opts}")
                    for opt in s.options:
                        if fy_label in opt.text.strip():
                            s.select_by_visible_text(opt.text)
                            log.info(f"    FY selected: {opt.text} ✓")
                            fy_set = True; break
                    if fy_set: break
                except: continue

            if not fy_set:
                try:
                    driver.execute_script(f"""
                        var sels=document.querySelectorAll('select');
                        for(var s of sels){{
                            for(var o of s.options){{
                                if(o.text && o.text.trim().indexOf('{fy_label}')>-1){{
                                    s.value=o.value;
                                    s.dispatchEvent(new Event('change',{{bubbles:true}}));
                                    s.dispatchEvent(new Event('input',{{bubbles:true}}));
                                    break;
                                }}
                            }}
                        }}
                    """)
                    time.sleep(1.5); log.info("    FY set via JS ✓"); fy_set = True
                except Exception as _jse:
                    log.warning(f"    JS FY select failed: {_jse}")

            if not fy_set:
                log.warning(f"    ✗ Could not select FY {fy_label}"); return False
            time.sleep(1)

            # -- 2. Click SEARCH -------------------------------
            searched = try_click(driver, [
                "//button[normalize-space()='SEARCH']",
                "//button[normalize-space()='Search']",
                "//button[contains(text(),'SEARCH')]",
                "//input[@value='SEARCH' or @value='Search']",
                "//button[@type='submit']",
            ], timeout=8, log=log)
            if not searched:
                driver.execute_script("""
                    var btns=document.querySelectorAll('button,input[type=submit]');
                    for(var b of btns){
                        if((b.innerText||b.value||'').toUpperCase().includes('SEARCH')){b.click();break;}
                    }
                """)
            time.sleep(4)
            log.info(f"    After SEARCH — URL: {driver.current_url}")

            # -- 3. Find and click DOWNLOAD button ------------
            log.info("    Scanning page for DOWNLOAD button...")
            dl_clicked = False
            save_ext   = ".xlsx"

            DOWNLOAD_XPATHS = [
                "//button[contains(text(),'DOWNLOAD COMPARISON REPORTS (EXCEL)')]",
                "//button[contains(text(),'DOWNLOAD COMPARISON REPORTS')]",
                "//button[contains(text(),'Download Comparison Reports')]",
                "//button[contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'DOWNLOAD COMPARISON')]",
                "//a[contains(text(),'DOWNLOAD COMPARISON')]",
                "//button[contains(text(),'DOWNLOAD') and contains(text(),'EXCEL')]",
            ]

            # Scroll to bottom first
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)

            # Log ALL buttons on page (critical for diagnosis)
            try:
                all_btns = [b.text.strip() for b in driver.find_elements(By.TAG_NAME,"button") if b.text.strip()]
                log.info(f"    ALL buttons: {all_btns}")
                vis_btns = [b.text.strip() for b in driver.find_elements(By.TAG_NAME,"button") if b.is_displayed() and b.text.strip()]
                log.info(f"    VISIBLE buttons: {vis_btns}")
            except: pass

            # Try xpaths at current position (bottom)
            dl_clicked = try_click(driver, DOWNLOAD_XPATHS, timeout=5, log=log)

            # Scroll 90%, 75%, 50% and try again
            if not dl_clicked:
                for pct in [0.9, 0.75, 0.5, 1.0]:
                    driver.execute_script(f"window.scrollTo(0, document.body.scrollHeight*{pct});")
                    time.sleep(1)
                    dl_clicked = try_click(driver, DOWNLOAD_XPATHS, timeout=3, log=log)
                    if dl_clicked:
                        log.info(f"    Button found at {int(pct*100)}% scroll ✓")
                        break

            # JS: deep scan including hidden/ng-click buttons
            if not dl_clicked:
                log.info("    JS deep scan for DOWNLOAD button...")
                try:
                    found_info = driver.execute_script("""
                        var elems = Array.from(document.querySelectorAll('button, a, input[type=button]'));
                        var found = [];
                        for(var el of elems){
                            var t = (el.innerText||el.textContent||el.value||el.getAttribute('title')||'')
                                      .toUpperCase().replace(/[\n\r\t]+/g,' ').trim();
                            if(t.length > 2) found.push(t.substring(0,60));
                            if(t.includes('DOWNLOAD') && (t.includes('COMPARISON')||t.includes('EXCEL')||t.includes('REPORT'))){
                                el.scrollIntoView({behavior:'instant',block:'center'});
                                el.click();
                                return {clicked:true, text:t};
                            }
                        }
                        return {clicked:false, all_texts: found};
                    """)
                    log.info(f"    JS scan result: {found_info}")
                    if found_info and found_info.get("clicked"):
                        dl_clicked = True; log.info("    JS DOWNLOAD click ✓")
                        time.sleep(3)
                except Exception as _je:
                    log.warning(f"    JS deep scan error: {_je}")

            # CSV fallback
            if not dl_clicked:
                log.info("    Trying CSV fallback...")
                dl_clicked = try_click(driver, [
                    "//button[contains(text(),'DOWNLOAD (CSV)')]",
                    "//button[contains(text(),'CSV')]",
                    "//a[contains(text(),'CSV')]",
                ], timeout=5, log=log)
                if dl_clicked: save_ext = ".csv"

            if not dl_clicked:
                log.warning(f"    ✗ No download button found. URL: {driver.current_url}")
                return False

            log.info("    Download button clicked ✓ — waiting 15s for file...")
            time.sleep(3)
            save_name = f"TaxLiability_{file_suffix}{save_ext}"
            if rename_latest(client_dir, save_name, [".xlsx",".xls",".csv"], log):
                log.info(f"    ✓ Saved: {save_name}"); return True
            log.warning("    Clicked but no file found in output folder")
            return False

        # -- Download Tax Liability for each FY ----------
        # Navigate ONCE to Tax Liability page, then loop FY years on same page
        tl_results = {}
        log.info("    Navigating to Tax Liability page (once)...")
        relogin_if_needed(driver, log)
        ok_nav = go_to_tax_liabilities()
        if not ok_nav:
            log.warning("    Tax Liability navigation failed — skipping")
            triggered["TAX_LIABILITY"] = "NAV_FAIL"
        else:
            for fy in TAX_FY_YEARS:
                try:
                    log.info(f"\n    Downloading Tax Liability: FY {fy}")
                    suffix = fy.replace("-","_")
                    ok = download_tax_fy(fy, suffix)
                    tl_results[fy] = "OK" if ok else "DOWNLOAD_FAIL"
                    # For multiple FYs: stay on same page, just re-select FY and SEARCH
                    # download_tax_fy already handles FY selection each call
                except Exception as e:
                    log.warning(f"    Tax Liability FY {fy} error: {e}")
                    tl_results[fy] = f"ERR:{e}"

        ok_tl = sum(1 for v in tl_results.values() if v == "OK")
        log.info(f"\n  Tax Liability: {ok_tl}/{len(TAX_FY_YEARS)} downloaded")
        for fy, st in tl_results.items():
            log.info(f"    FY {fy}: {st}")
        triggered["TAX_LIABILITY"] = "OK" if ok_tl > 0 else "FAIL"

        ok_count = sum(1 for v in triggered.values() if v in ("TRIGGERED", "OK"))
    log.info(f"\n  Phase 1 done — {ok_count} items triggered/downloaded")
    return triggered


# ==========================================================
# PHASE 2 — DOWNLOAD GENERATED FILES (after portal wait)
#
# CORRECT PORTAL FLOW (confirmed from logs + screenshots):
#   1. Navigate to dashboard → select month → SEARCH
#   2. Click tile DOWNLOAD  → lands on offlinedownload page
#   3. Click GENERATE again (JSON for GSTR1/1A, EXCEL for 2B/2A)
#      → portal shows "Your request acknowledged..." banner
#      → after a few seconds the "Click here to download" link appears
#   4. Click that download link → file downloads
#
# GSTR-3B is already downloaded in Phase 1 (PDF is immediate).
# Returns dict  key="{Month}_{Year}_{RetType}" → status string
# ==========================================================
def phase2_download_all(driver, client_dir, triggered, log, returns_todo=None):
    """
    returns_todo: set of return types to process e.g. {"GSTR1","GSTR2B","GSTR2A"}
                  None means process all triggered ones.
    """
    if returns_todo is None:
        returns_todo = {"GSTR1", "GSTR1A", "GSTR2B", "GSTR2A"}

    dl_results = {}

    # XPaths to click GENERATE on the offline-download page
    GENERATE_JSON_XP = [
        "//button[contains(text(),'GENERATE JSON FILE TO DOWNLOAD')]",
        "//button[contains(text(),'GENERATE JSON')]",
        "//button[contains(text(),'Generate JSON')]",
        "//a[contains(text(),'GENERATE JSON')]",
    ]
    GENERATE_EXCEL_XP = [
        "//button[contains(text(),'GENERATE EXCEL FILE TO DOWNLOAD')]",
        "//button[contains(text(),'GENERATE EXCEL')]",
        "//button[contains(text(),'Generate Excel')]",
        "//a[contains(text(),'GENERATE EXCEL')]",
        # fallback to JSON if no Excel button
        "//button[contains(text(),'GENERATE JSON FILE TO DOWNLOAD')]",
        "//button[contains(text(),'Generate JSON')]",
    ]

    # XPaths to find the actual download link after generation
    DOWNLOAD_LINK_XP = [
        "//a[contains(text(),'Click here to download')]",
        "//a[contains(text(),'click here to download')]",
        "//a[contains(text(),'File 1')]",
        "//a[contains(text(),'File 2')]",
        "//a[contains(@href,'.zip')]",
        "//a[contains(@href,'filedownload')]",
        "//a[contains(@href,'download') and string-length(@href) > 50]",
    ]

    def find_download_link():
        """Return the first visible download link, or None."""
        for xp in DOWNLOAD_LINK_XP:
            try:
                els = driver.find_elements(By.XPATH, xp)
                for el in els:
                    if el.is_displayed():
                        href = el.get_attribute("href") or ""
                        if len(href) > 20:   # ignore empty / anchor-only hrefs
                            return el
            except:
                continue
        return None

    def click_generate_and_wait(gen_xpaths, save_name, max_wait=120):
        """
        Click GENERATE, then fast-poll every 0.5s for 30s (instant downloads),
        then slower 3s poll for download link. Returns True if downloaded.
        """
        gen_clicked = try_click(driver, gen_xpaths, timeout=10, log=log)
        if gen_clicked:
            log.info(f"    GENERATE clicked — fast-polling for download...")
        else:
            log.warning(f"    GENERATE button not found — checking for link: {driver.current_url}")

        client_path = Path(client_dir)

        def _snap():
            exts = {".xlsx", ".zip", ".json"}
            return {str(f): f.stat().st_mtime for f in client_path.iterdir()
                    if f.suffix.lower() in exts}

        def _check_new(before):
            for f in client_path.iterdir():
                if f.suffix.lower() not in {".xlsx", ".zip", ".json"}: continue
                if f.name.endswith((".crdownload", ".tmp")): continue
                prev = before.get(str(f))
                if (prev is None or f.stat().st_mtime > prev + 0.1) and f.stat().st_size > 500:
                    return f
            return None

        before = _snap()

        # Fast poll 30s for instant download
        deadline_fast = time.time() + 30
        while time.time() < deadline_fast:
            time.sleep(0.5)
            new_f = _check_new(before)
            if new_f:
                log.info(f"    ⚡ Instant download: {new_f.name}")
                time.sleep(0.8)
                rename_latest(client_dir, save_name, [".zip", ".json", ".xlsx"], log)
                return True

        # Slower poll: link-based
        elapsed = 0
        before2 = _snap()
        while elapsed < max_wait:
            new_f = _check_new(before2)
            if new_f:
                log.info(f"    📥 File arrived: {new_f.name}")
                time.sleep(0.8)
                rename_latest(client_dir, save_name, [".zip", ".json", ".xlsx"], log)
                return True

            link = find_download_link()
            if link:
                txt = link.text.strip() or link.get_attribute("href")[:60]
                log.info(f"    Download link: '{txt}'")
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", link)
                driver.execute_script("arguments[0].click();", link)
                time.sleep(4)
                rename_latest(client_dir, save_name, [".zip", ".json", ".xlsx"], log)
                return True

            if elapsed > 0 and elapsed % 30 == 0:
                log.info(f"    Link not found — refreshing ({elapsed}s)...")
                try:
                    driver.refresh()
                    time.sleep(2)
                    if is_session_lost(driver):
                        log.warning("    Session lost — re-logging in...")
                        if not relogin_if_needed(driver, log):
                            return False
                except: pass
            else:
                time.sleep(3)
            elapsed += 3

        log.warning(f"    Download link not found for {save_name}")
        return False

    # -- per-return config — GSTR-2B excluded (downloaded in Phase 1) --
    ret_config = {
        "GSTR1":  ("GSTR1",  GENERATE_JSON_XP,  ".zip"),
        "GSTR1A": ("GSTR1A", GENERATE_JSON_XP,  ".zip"),
        "GSTR2A": ("GSTR2A", GENERATE_EXCEL_XP, ".zip"),
        # GSTR-2B: handled entirely in Phase 1 (immediate download)
        # GSTR-3B: handled entirely in Phase 1 (PDF, no generate needed)
    }

    for month_name, month_num, year in MONTHS:
        key = f"{month_name}_{year}"
        active = [r for r in ret_config if r in returns_todo]
        if not active:
            continue
        log.info(f"\n  -- Phase 2 [{month_name} {year}] --")

        for ret_type in active:
            tile_name, gen_xp, save_ext = ret_config[ret_type]
            tkey = f"{key}_{ret_type}"

            p1_status = triggered.get(tkey, "SKIPPED")
            if p1_status not in ("TRIGGERED", "OK"):
                dl_results[tkey] = p1_status
                log.info(f"    {ret_type}: skip (Phase1={p1_status})")
                continue

            save_name = f"{ret_type}_{month_name}_{year}{save_ext}"

            try:
                safe_go_to_dashboard(driver, log)
                select_and_search(driver, month_name, log)

                if not click_tile_download(driver, tile_name, log):
                    dl_results[tkey] = "TILE_FAIL"
                    continue

                time.sleep(0.5)  # was PAGE_WAIT=8

                if click_generate_and_wait(gen_xp, save_name):
                    dl_results[tkey] = "OK"
                    log.info(f"    {ret_type}: DOWNLOADED ✓")
                else:
                    dl_results[tkey] = "NOT_FOUND"

            except Exception as e:
                log.warning(f"    {ret_type} download failed [{month_name}]: {e}")
                dl_results[tkey] = f"ERR:{e}"

    ok_count = sum(1 for v in dl_results.values() if v == "OK")
    log.info(f"\n  Phase 2 complete — {ok_count} files downloaded")
    return dl_results


# ==========================================================
# PHASE 3 — EXTRACT JSON ZIPs → EXCEL INVOICE SUMMARIES
# Finds GSTR1_*.zip files, extracts JSON, writes invoice rows.
# ==========================================================
def extract_json_to_excel(client_dir, client_name, log):
    """
    Extract GSTR-1 JSON ZIPs → multi-sheet Excel workbooks.

    Per-ZIP workbook sheets:
      1. B2B_Invoices        — invoice-level B2B detail
      2. B2CS_Summary        — B2C Small (state+rate level, no invoice number)
      3. B2CL_Invoices       — B2C Large inter-state invoices
      4. CDNR_CreditNotes    — Credit Notes to registered buyers (ntty=C)
      5. CDNR_DebitNotes     — Debit Notes to registered buyers  (ntty=D)
      6. CDNUR_Notes         — Credit/Debit Notes to unregistered buyers
      7. Exports             — Export invoices (WOPAY / WPAY)
      8. GSTR1A_Amendments   — Amended invoices (b2ba, cdnra etc.) if present
      9. Summary             — Grand totals per type for this month
    """
    log.info("\n  Extracting GSTR-1 JSON → Excel (multi-sheet)...")
    extracted = 0

    # Style helpers (module-level constants already defined)
    def _mk_sheet(wb, title, headers, widths, title_cols, bg_color=None):
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        ncols = int(title_cols) if not isinstance(title_cols, int) else title_cols
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        tc = ws["A1"]
        tc.value = title.replace("_"," ")
        tc.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        hdr_color = bg_color if bg_color else DARK_BLUE
        tc.fill = fill(hdr_color); tc.alignment = aln()
        ws.row_dimensions[1].height = 26
        for ci,(h,w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=2, column=ci, value=h)
            sc(c, bold=True, fg="FFFFFF", bg=MED_BLUE, size=9)
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[2].height = 20
        ws.freeze_panes = "A3"
        return ws, [3]   # mutable row counter

    def _wr(ws, ri_ref, vals, type_col=None, type_bg=None):
        ri = ri_ref[0]
        bg = GREY_BG if ri%2==0 else WHITE
        for ci,v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = Font(name="Arial", size=9)
            cell.fill = fill(type_bg if (type_col and ci==type_col) else bg)
            cell.alignment = aln(h="right" if isinstance(v,(int,float)) else "left")
            cell.border = bdr()
            if isinstance(v,(int,float)):
                cell.number_format = "#,##0.00"
        ws.row_dimensions[ri].height = 15
        ri_ref[0] += 1

    def _totrow(ws, ri_ref, vals):
        ri = ri_ref[0]
        for ci,v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", bold=True, size=9)
            c.fill = fill("D6DCE4")
            c.alignment = aln(h="right" if isinstance(v,(int,float)) else "left")
            c.border = bdr()
            if isinstance(v,(int,float)): c.number_format = "#,##0.00"
        ws.row_dimensions[ri].height = 17
        ri_ref[0] += 1

    for zip_path in sorted(Path(client_dir).glob("GSTR1_*.zip")):
        try:
            extract_dir = Path(str(zip_path).replace(".zip","_ex"))
            extract_dir.mkdir(exist_ok=True)
            with zipfile.ZipFile(zip_path, "r") as zf:
                zf.extractall(extract_dir)

            json_files = list(extract_dir.glob("*.json")) + list(extract_dir.glob("**/*.json"))
            if not json_files:
                log.warning(f"    No JSON inside {zip_path.name}"); continue

            with open(json_files[0], "r", encoding="utf-8") as f:
                d = json.load(f)

            stem      = zip_path.stem
            month_lbl = stem.replace("GSTR1_","").replace("_"," ")
            wb = Workbook(); wb.remove(wb.active)
            log.info(f"    Processing: {stem}")

            # --- Counters for Summary sheet ----------------
            totals = {}  # type → {tx,ig,cg,sg,cnt}
            def acc(t, tx, ig, cg, sg, n=1):
                if t not in totals: totals[t]={"tx":0.,"ig":0.,"cg":0.,"sg":0.,"cnt":0}
                totals[t]["tx"]+=tx; totals[t]["ig"]+=ig
                totals[t]["cg"]+=cg; totals[t]["sg"]+=sg; totals[t]["cnt"]+=n

            # -- Sheet 1: B2B Invoices ---------------------
            HDR_B2B = ["GSTIN Receiver","Trade Name","Invoice No","Invoice Date",
                       "Invoice Value ₹","Place of Supply","Reverse Charge",
                       "Rate %","Taxable Value ₹","IGST ₹","CGST ₹","SGST ₹"]
            WID_B2B = [22,26,18,12,15,16,10,8,15,12,12,12]
            ws1, ri1 = _mk_sheet(wb,"B2B_Invoices",HDR_B2B,WID_B2B,len(HDR_B2B))
            ws1["A1"].value = f"GSTR-1 B2B Invoices — {client_name} — {month_lbl}"

            for entry in d.get("b2b",[]):
                ctin = entry.get("ctin",""); trdnm = entry.get("trdnm","")
                for inv in entry.get("inv",[]):
                    rc = inv.get("rchrg","N")
                    for it in inv.get("itms",[]):
                        det = it.get("itm_det",{})
                        rt  = det.get("rt",0)
                        tv  = float(det.get("txval",0) or 0)
                        ig  = float(det.get("iamt",0) or 0)
                        cg  = float(det.get("camt",0) or 0)
                        sg  = float(det.get("samt",0) or 0)
                        _wr(ws1, ri1, [ctin, trdnm,
                            inv.get("inum",""), inv.get("idt",""),
                            round(float(inv.get("val",0) or 0),2),
                            inv.get("pos",""), rc, rt,
                            round(tv,2), round(ig,2), round(cg,2), round(sg,2)])
                        acc("B2B", tv, ig, cg, sg)

            # -- Sheet 2: B2CS Summary ---------------------
            HDR_B2CS = ["Place of Supply","Supply Type","Rate %",
                        "Taxable Value ₹","IGST ₹","CGST ₹","SGST ₹"]
            ws2, ri2 = _mk_sheet(wb,"B2CS_Summary",HDR_B2CS,[16,14,8,15,12,12,12],len(HDR_B2CS))
            ws2["A1"].value = f"GSTR-1 B2CS Summary — {client_name} — {month_lbl}"
            for rec in d.get("b2cs",[]):
                tv=float(rec.get("txval",0) or 0); ig=float(rec.get("iamt",0) or 0)
                cg=float(rec.get("camt",0) or 0);  sg=float(rec.get("samt",0) or 0)
                _wr(ws2, ri2, [rec.get("pos",""), rec.get("sply_ty","INTRA"),
                               rec.get("rt",0), round(tv,2),round(ig,2),round(cg,2),round(sg,2)])
                acc("B2CS", tv, ig, cg, sg)

            # -- Sheet 3: B2CL Invoices --------------------
            HDR_B2CL = ["Place of Supply","Invoice No","Invoice Date",
                        "Invoice Value ₹","Rate %","Taxable Value ₹","IGST ₹"]
            ws3, ri3 = _mk_sheet(wb,"B2CL_Invoices",HDR_B2CL,[16,18,12,15,8,15,12],len(HDR_B2CL))
            ws3["A1"].value = f"GSTR-1 B2CL Invoices — {client_name} — {month_lbl}"
            for rec in d.get("b2cl",[]):
                pos=rec.get("pos","")
                for inv in rec.get("inv",[]):
                    for it in inv.get("itms",[]):
                        det=it.get("itm_det",{})
                        tv=float(det.get("txval",0) or 0); ig=float(det.get("iamt",0) or 0)
                        _wr(ws3, ri3, [pos, inv.get("inum",""), inv.get("idt",""),
                                       round(float(inv.get("val",0) or 0),2),
                                       det.get("rt",0), round(tv,2), round(ig,2)])
                        acc("B2CL", tv, ig, 0, 0)

            # -- Sheet 4+5: CDNR Credit Notes / Debit Notes -
            HDR_CDN = ["Note Type","GSTIN Receiver","Trade Name",
                       "Note Number","Note Date","Note Value ₹",
                       "Place of Supply","Pre-GST","Rate %",
                       "Taxable Value ₹","IGST ₹","CGST ₹","SGST ₹"]
            WID_CDN = [10,22,24,18,12,14,14,8,8,15,12,12,12]
            wsCR, riCR = _mk_sheet(wb,"CDNR_CreditNotes",HDR_CDN,WID_CDN,len(HDR_CDN))
            wsCR["A1"].value = f"GSTR-1 CDNR Credit Notes — {client_name} — {month_lbl}"
            wsDR, riDR = _mk_sheet(wb,"CDNR_DebitNotes", HDR_CDN,WID_CDN,len(HDR_CDN))
            wsDR["A1"].value = f"GSTR-1 CDNR Debit Notes — {client_name} — {month_lbl}"

            for entry in d.get("cdnr",[]):
                ctin=entry.get("ctin",""); trdnm=entry.get("trdnm","")
                for note in entry.get("nt",[]):
                    ntty = note.get("ntty","C")   # C=Credit  D=Debit
                    for it in note.get("itms",[]):
                        det=it.get("itm_det",{})
                        tv=float(det.get("txval",0) or 0); ig=float(det.get("iamt",0) or 0)
                        cg=float(det.get("camt",0) or 0);  sg=float(det.get("samt",0) or 0)
                        row=[ntty, ctin, trdnm,
                             note.get("nt_num",""), note.get("nt_dt",""),
                             round(float(note.get("val",0) or 0),2),
                             note.get("pos",""), note.get("p_gst","N"),
                             det.get("rt",0), round(tv,2),round(ig,2),round(cg,2),round(sg,2)]
                        if ntty=="D": _wr(wsDR, riDR, row); acc("CDNR-Debit", tv,ig,cg,sg)
                        else:         _wr(wsCR, riCR, row); acc("CDNR-Credit",tv,ig,cg,sg)

            # -- Sheet 6: CDNUR (Unregistered CDN) --------
            HDR_CDNUR = ["Note Type","Supply Type","Note Number","Note Date",
                         "Note Value ₹","Rate %","Taxable Value ₹","IGST ₹"]
            wsCU, riCU = _mk_sheet(wb,"CDNUR_Unregistered",HDR_CDNUR,
                                   [10,12,18,12,14,8,15,12],len(HDR_CDNUR))
            wsCU["A1"].value = f"GSTR-1 CDNUR (Unregistered) — {client_name} — {month_lbl}"
            for note in d.get("cdnur",[]):
                ntty=note.get("ntty","C"); spty=note.get("typ","")
                tv=float(note.get("txval",0) or 0); ig=float(note.get("iamt",0) or 0)
                _wr(wsCU, riCU, [ntty, spty,
                    note.get("nt_num",""), note.get("nt_dt",""),
                    round(float(note.get("val",0) or 0),2),
                    note.get("rt",0), round(tv,2), round(ig,2)])
                acc("CDNUR", tv, ig, 0, 0)

            # -- Sheet 7: Exports --------------------------
            HDR_EXP = ["Export Type","Invoice No","Invoice Date",
                       "Invoice Value ₹","Port Code","Shipping Bill No",
                       "Shipping Bill Date","Rate %",
                       "Taxable Value ₹","IGST ₹"]
            wsEX, riEX = _mk_sheet(wb,"Exports",HDR_EXP,
                                   [14,18,12,15,12,18,16,8,15,12],len(HDR_EXP))
            wsEX["A1"].value = f"GSTR-1 Exports — {client_name} — {month_lbl}"
            for exp in d.get("exp",[]):
                etype = exp.get("exp_typ","")
                for inv in exp.get("inv",[]):
                    for it in inv.get("itms",[]):
                        det=it.get("itm_det",{})
                        tv=float(det.get("txval",0) or 0); ig=float(det.get("iamt",0) or 0)
                        _wr(wsEX, riEX, [etype,
                            inv.get("inum",""), inv.get("idt",""),
                            round(float(inv.get("val",0) or 0),2),
                            inv.get("pcode",""), inv.get("sbnum",""),
                            inv.get("sbdt",""), det.get("rt",0),
                            round(tv,2), round(ig,2)])
                        acc("Exports", tv, ig, 0, 0)

            # -- Sheet 8: GSTR-1A Amendments --------------
            # Portal returns amended invoices in b2ba, b2cla, cdnra, expа
            has_amendments = any(d.get(k) for k in ["b2ba","cdnra","b2cla","expa"])
            if has_amendments:
                HDR_AMD = ["Section","GSTIN","Original Inv/Note","Amended Inv/Note",
                           "Date","Value ₹","Rate %","Taxable ₹","IGST ₹","CGST+SGST ₹"]
                wsAM, riAM = _mk_sheet(wb,"GSTR1A_Amendments",HDR_AMD,
                                       [12,22,18,18,12,14,8,14,12,14],len(HDR_AMD))
                wsAM["A1"].value = f"GSTR-1A Amendments — {client_name} — {month_lbl}"

                for entry in d.get("b2ba",[]):
                    ctin=entry.get("ctin","")
                    for inv in entry.get("inv",[]):
                        for it in inv.get("itms",[]):
                            det=it.get("itm_det",{})
                            tv=float(det.get("txval",0) or 0); ig=float(det.get("iamt",0) or 0)
                            csgst=float(det.get("camt",0) or 0)+float(det.get("samt",0) or 0)
                            _wr(wsAM, riAM, ["B2BA", ctin,
                                inv.get("oinum",""), inv.get("inum",""),
                                inv.get("idt",""),
                                round(float(inv.get("val",0) or 0),2),
                                det.get("rt",0), round(tv,2), round(ig,2), round(csgst,2)])
                            acc("Amended-B2B", tv, ig, 0, 0)

                for entry in d.get("cdnra",[]):
                    ctin=entry.get("ctin","")
                    for note in entry.get("nt",[]):
                        for it in note.get("itms",[]):
                            det=it.get("itm_det",{})
                            tv=float(det.get("txval",0) or 0); ig=float(det.get("iamt",0) or 0)
                            csgst=float(det.get("camt",0) or 0)+float(det.get("samt",0) or 0)
                            _wr(wsAM, riAM, ["CDNRA", ctin,
                                note.get("ont_num",""), note.get("nt_num",""),
                                note.get("nt_dt",""),
                                round(float(note.get("val",0) or 0),2),
                                0, round(tv,2), round(ig,2), round(csgst,2)])
                            acc("Amended-CDN", tv, ig, 0, 0)

            # -- Sheet 8b: Nil Rated / Exempt -------------
            HDR_NIL = ["Supply Type","Nil Rated ₹","Exempt ₹","Non-GST ₹","Total ₹"]
            wsNL, riNL = _mk_sheet(wb,"Nil_Rated",HDR_NIL,[20,16,16,16,14],len(HDR_NIL))
            wsNL["A1"].value = f"Nil/Exempt/Non-GST — {client_name} — {month_lbl}"
            nil_recs_xj = []
            ns_xj = d.get("nil_sup", d.get("nil", None))
            if isinstance(ns_xj, list): nil_recs_xj = ns_xj
            elif isinstance(ns_xj, dict): nil_recs_xj = ns_xj.get("inv", ns_xj.get("details",[]))
            nil_total_xj = 0.0
            for rec in nil_recs_xj:
                if not isinstance(rec, dict): continue
                stype=rec.get("sply_ty","")
                nil_v_xj =float(rec.get("nil_amt", rec.get("nil",  0)) or 0)
                expt_v_xj=float(rec.get("expt_amt",rec.get("expt", 0)) or 0)
                ngsup_xj =float(rec.get("ngsup_amt",rec.get("ngsup",0)) or 0)
                tot_xj = nil_v_xj+expt_v_xj+ngsup_xj
                if tot_xj == 0: continue
                _wr(wsNL, riNL, [stype, round(nil_v_xj,2), round(expt_v_xj,2), round(ngsup_xj,2), round(tot_xj,2)])
                acc("NIL/EXEMPT", tot_xj, 0, 0, 0)
                nil_total_xj += tot_xj

            # -- Sheet 9: Summary --------------------------
            HDR_SUM = ["Type","Count","Taxable Value ₹",
                       "IGST ₹","CGST ₹","SGST ₹","Total Tax ₹"]
            wsS, riS = _mk_sheet(wb,"Summary",HDR_SUM,[20,8,16,12,12,12,14],len(HDR_SUM))
            wsS["A1"].value = f"GSTR-1 Summary — {client_name} — {month_lbl}"
            tot_tx=tot_ig=tot_cg=tot_sg=tot_cnt=0
            for t,v in sorted(totals.items()):
                _wr(wsS, riS, [t, v["cnt"],
                    round(v["tx"],2), round(v["ig"],2),
                    round(v["cg"],2), round(v["sg"],2),
                    round(v["ig"]+v["cg"]+v["sg"],2)])
                tot_tx+=v["tx"]; tot_ig+=v["ig"]
                tot_cg+=v["cg"]; tot_sg+=v["sg"]; tot_cnt+=v["cnt"]
            _totrow(wsS, riS, ["GRAND TOTAL", _fsum("B",3,riS-1),
                               _fsum("C",3,riS-1), _fsum("D",3,riS-1),
                               _fsum("E",3,riS-1), _fsum("F",3,riS-1),
                               f"=SUM(D{riS}:F{riS})"])

            xl_name = stem + "_detail.xlsx"
            xl_path = Path(client_dir) / xl_name
            wb.save(str(xl_path))
            total_rows = sum(v["cnt"] for v in totals.values())
            log.info(f"    ✓ Saved: {xl_name}  ({total_rows} records across all types)")
            extracted += 1

        except zipfile.BadZipFile:
            log.warning(f"    Bad ZIP: {zip_path.name}")
        except Exception as e:
            log.warning(f"    JSON extract error [{zip_path.name}]: {e}")
            import traceback; log.warning(traceback.format_exc())

    log.info(f"  GSTR-1 extraction done — {extracted} Excel file(s) created")


# ==========================================================
# ANNUAL RECONCILIATION REPORT
# Based on actual GST portal report format (from sample files):
#   Sheet 1: GSTR-1 Sales Summary (B2B + B2CS monthwise)
#   Sheet 2: GSTR-2B ITC Summary (confirmed ITC from purchases)
#   Sheet 3: GSTR-2A Purchase Summary (all inward supplies)
#   Sheet 4: GSTR-3B Monthly Summary (tax paid)
#   Sheet 5: GSTR-3B vs GSTR-1 Reconciliation (R1 vs 3B)
#   Sheet 6: Tax Liability & ITC Comparison (final reconciliation)
# ==========================================================
# ================================================================
# RECONCILIATION — EXACT FORMAT MATCHING PORTAL DOWNLOADED FILES
#
# FILE 1: GSTR-3B vs GSTR-2A  (GSTR3BR2A_RECONCILED)
#   Sheets: Read me | Q1-APR-JUN | Q2-JUL-SEP | Q3-OCT-DEC | Q4-JAN-MAR | Annual-APR-MAR
#   Section A: GSTR-3B ITC Details   (4A(5) All other ITC, 4B ITC Reversed, Total A)
#   Section B: GSTR-2A ITC Details   (B2B, CDNR, TDS, TCS, Total B)
#   Row:       Difference (A - B)
#   Columns per month: IGST | CGST | SGST | Cess | Total
#
# FILE 2: GSTR-3B vs GSTR-1  (GSTR3BR1_RECONCILED)
#   Same sheet structure
#   Section A: GSTR-3B Supply Details (3.1(a)…3.2, Total A)
#   Section B: GSTR-1 Supply Details  (B2B, B2CS, B2CL, Exports, CDN-R, CDN-U, Total B)
#   Row:       Difference (A - B)
#   Columns per month: Taxable | IGST | CGST | SGST | Cess | Total
# ================================================================


# Module-level GSTR-3B PDF extractor
def extract_3b_pdf(pdf_path):
    """
    Robust GSTR-3B PDF extraction — line-by-line scan.
    Handles portal PDF quirks: 'L0', 'I0', split rows, stray chars.
    Returns ALL fields from Tables 3.1, 3.1(c), 4, 5.1, 6.1.
    """
    result = {
        # Meta
        "gstin":"","legal_name":"","trade_name":"","period":"","year":"","arn":"","arn_date":"",
        # 3.1(a) Outward taxable
        "taxable":0.,"o_igst":0.,"o_cgst":0.,"o_sgst":0.,"o_cess":0.,
        # 3.1(b) Zero rated
        "zero_taxable":0.,"zero_igst":0.,
        # 3.1(c) Nil / Exempt
        "nil_exempt":0.,
        # 3.1(d) RCM inward
        "rcm_taxable":0.,"rcm_igst":0.,"rcm_cgst":0.,"rcm_sgst":0.,
        # 3.1(e) Non-GST
        "non_gst":0.,
        # 4A ITC available
        "itc_import_goods":0.,"itc_import_svc":0.,"itc_rcm":0.,"itc_isd":0.,
        "itc_igst":0.,"itc_cgst":0.,"itc_sgst":0.,"itc_cess":0.,
        # 4B ITC reversed
        "rev_igst":0.,"rev_cgst":0.,"rev_sgst":0.,
        # 4C Net ITC
        "net_itc_igst":0.,"net_itc_cgst":0.,"net_itc_sgst":0.,
        # 5.1
        "interest_igst":0.,"interest_cgst":0.,"interest_sgst":0.,
        "late_fee_cgst":0.,"late_fee_sgst":0.,
        # 6.1 Tax paid
        "tax_paid_igst":0.,"tax_paid_cgst":0.,"tax_paid_sgst":0.,
    }
    if not pdf_path.exists():
        return result
    import re
    text = ""
    try:
        import pdfplumber
        with pdfplumber.open(str(pdf_path)) as pdf2:
            text = "\n".join(p.extract_text() or "" for p in pdf2.pages)
    except ImportError:
        try:
            import PyPDF2
            with open(str(pdf_path),"rb") as f2:
                r2 = PyPDF2.PdfReader(f2)
                text = "\n".join(p.extract_text() or "" for p in r2.pages)
        except: pass
    except Exception: pass
    if not text:
        return result

    def _n(s):
        """Convert string to float, handling L0/I0 portal artifacts."""
        try:
            return float(re.sub(r"[^\d.\-]","",str(s).replace("L","").replace("I","").replace(",","")))
        except: return 0.0

    def nums_on_line(line):
        """Extract all decimal numbers from a single PDF line."""
        clean = re.sub(r"[LI](\d)", r"\1", line)
        return [_n(n) for n in re.findall(r"-?\d[\d,]*\.\d*", clean)]

    lines = text.split("\n")
    # State flags
    in_section_4  = False   # inside Table 4 (ITC)
    in_section_6  = False   # inside Table 6.1
    past_51_header = False  # seen "5.1 Interest and Late fee" heading

    for i, line in enumerate(lines):
        lo   = line.lower().strip()
        nums = nums_on_line(line)

        # -- Meta fields --------------------------------------------
        if "gstin of the supplier" in lo:
            parts = re.split(r"supplier\s*", line, flags=re.I)
            if len(parts) > 1: result["gstin"] = parts[-1].strip()
        elif "2(a)." in line and "legal name" in lo:
            result["legal_name"] = re.sub(r"2\(a\)\..*?person\s*","",line,flags=re.I).strip()
        elif "2(b)." in line and "trade name" in lo:
            result["trade_name"] = re.sub(r"2\(b\)\.\s*Trade name,?\s*if any\s*","",line,flags=re.I).strip()
        elif re.match(r"Year\s+\d{4}", line):
            result["year"] = line.replace("Year","").strip()
        elif re.match(r"Period\s+[A-Za-z]", line) and "supply" not in lo:
            result["period"] = line.replace("Period","").strip()
        elif "2(c)." in line and "arn" in lo and "nil" not in lo and "exempt" not in lo:
            result["arn"] = re.sub(r"2\(c\)\.\s*ARN\s*","",line,flags=re.I).strip()
        elif "2(d)." in line and "date of arn" in lo:
            result["arn_date"] = re.sub(r"2\(d\)\.\s*Date of ARN\s*","",line,flags=re.I).strip()

        # -- Section flags ------------------------------------------
        elif re.search(r"4\.?\s*eligible itc|table.*4|^4\.\s", lo):
            in_section_4 = True; in_section_6 = False
        elif re.search(r"6\.?1.*payment of tax|payment of tax", lo):
            in_section_6 = True; in_section_4 = False
        elif re.search(r"5\.?1.*interest.*late fee|interest.*late fee.*previous", lo):
            past_51_header = True   # next "late fee" line will be actual data

        # -- 3.1(a) Outward taxable (other than zero/nil/exempt) ----
        elif re.search(r"\(a\).*outward taxable", lo) and "(b)" not in lo:
            # Line may be split: numbers on same line or next
            combined = line
            if len(nums) < 4 and i+1 < len(lines):
                combined = line + " " + lines[i+1]
                nums = nums_on_line(combined)
            if len(nums) >= 4:
                result["taxable"] = nums[0]
                result["o_igst"]  = nums[1]
                result["o_cgst"]  = nums[2]
                result["o_sgst"]  = nums[3]
                if len(nums) >= 5: result["o_cess"] = nums[4]

        # -- 3.1(b) Zero rated --------------------------------------
        elif re.search(r"\(b\).*zero rated", lo):
            if len(nums) >= 1: result["zero_taxable"] = nums[0]
            if len(nums) >= 2: result["zero_igst"]    = nums[1]

        # -- 3.1(c) Nil / Exempt ------------------------------------
        # CRITICAL: Must NOT match "2(c). ARN" line.
        # Correct line starts with "(c " or "(c)" and contains nil/exempt
        elif re.search(r"^\(c[\s\)]", lo) and re.search(r"nil|exempt", lo):
            if nums: result["nil_exempt"] = nums[0]

        # -- 3.1(d) Inward supplies liable to RCM -------------------
        elif re.search(r"\(d\).*inward.*reverse charge|\(d\).*reverse charge.*inward", lo):
            if len(nums) >= 4:
                result["rcm_taxable"] = nums[0]
                result["rcm_igst"]    = nums[1]
                result["rcm_cgst"]    = nums[2]
                result["rcm_sgst"]    = nums[3]

        # -- 3.1(e) Non-GST -----------------------------------------
        elif re.search(r"\(e\).*non.?gst", lo):
            if nums: result["non_gst"] = nums[0]

        # -- 3.1.1 ECO supplies — skip, just track -----------------
        # (i) ECO operator pays: adds to output tax but already in 3.1(a) aggregate

        # -- Table 4: ITC -------------------------------------------
        elif re.search(r"\(1\).*import of goods", lo):
            if len(nums) >= 1: result["itc_import_goods"] = nums[0]
        elif re.search(r"\(2\).*import of services", lo):
            if len(nums) >= 1: result["itc_import_svc"] = nums[0]
        elif re.search(r"\(3\).*reverse charge.*other than.*1.*2|\(3\).*inward.*reverse charge.*other", lo):
            if len(nums) >= 1: result["itc_rcm"] = nums[0]
        elif re.search(r"\(4\).*isd|\(4\).*inward.*isd", lo):
            if len(nums) >= 1: result["itc_isd"] = nums[0]
        elif re.search(r"\(5\).*all other itc", lo):
            # Columns: IGST | CGST | SGST | Cess
            if len(nums) >= 3:
                result["itc_igst"] = nums[0]
                result["itc_cgst"] = nums[1]
                result["itc_sgst"] = nums[2]
                if len(nums) >= 4: result["itc_cess"] = nums[3]

        elif re.search(r"\(1\).*rules 38.*42|\(1\).*as per rules", lo):
            if len(nums) >= 3:
                result["rev_igst"] = nums[0]
                result["rev_cgst"] = nums[1]
                result["rev_sgst"] = nums[2]

        elif re.search(r"net itc available|^c\..*net itc", lo):
            if len(nums) >= 3:
                result["net_itc_igst"] = nums[0]
                result["net_itc_cgst"] = nums[1]
                result["net_itc_sgst"] = nums[2]

        # -- 5.1 Interest Paid --------------------------------------
        elif re.search(r"^interest paid|interest paid\s", lo):
            # "Interest Paid 1.15 8.31 8.31 0.00"
            if len(nums) >= 3:
                result["interest_igst"] = nums[0]
                result["interest_cgst"] = nums[1]
                result["interest_sgst"] = nums[2]
            elif len(nums) == 2:
                result["interest_cgst"] = nums[0]
                result["interest_sgst"] = nums[1]

        # -- 5.1 Late fee --------------------------------------------
        # CRITICAL: Skip the section heading "5.1 Interest and Late fee..."
        # Only capture the data row "Late fee  -  25.00  25.00  -"
        elif re.search(r"late fee", lo) and "5.1" not in lo and "interest and late" not in lo:
            # Filter out the heading — only take lines with actual decimal values
            late_nums = [n for n in nums if n > 0]
            if len(late_nums) >= 2:
                result["late_fee_cgst"] = late_nums[0]
                result["late_fee_sgst"] = late_nums[1]
            elif len(late_nums) == 1:
                result["late_fee_cgst"] = late_nums[0]

        # -- 6.1 Tax paid rows --------------------------------------
        elif in_section_6:
            # Match "Central tax  20372.00  0.00  20372.00  0.00  6094.00  -  -  14278.00  8.00  25.00"
            if re.match(r"central\s*(tax)?", lo) and len(nums) >= 3:
                tp = nums[0]
                candidates = [n for n in nums[2:] if 0 < n < tp*2]
                if candidates:
                    result["tax_paid_cgst"] = candidates[-2] if len(candidates)>=2 else candidates[0]
            elif re.match(r"state|ut\s*(tax)?", lo) and len(nums) >= 3:
                tp = nums[0]
                candidates = [n for n in nums[2:] if 0 < n < tp*2]
                if candidates:
                    result["tax_paid_sgst"] = candidates[-2] if len(candidates)>=2 else candidates[0]
            elif re.match(r"integrated\s*(tax)?", lo) and len(nums) >= 3:
                tp = nums[0]
                candidates = [n for n in nums[2:] if 0 < n < tp*2]
                if candidates:
                    result["tax_paid_igst"] = candidates[-2] if len(candidates)>=2 else candidates[0]

    # -- Fallback: derive ITC from net if 4A(5) was zero ------------
    if result["itc_cgst"] == 0 and result["net_itc_cgst"] > 0:
        result["itc_cgst"] = result["net_itc_cgst"]
        result["itc_igst"] = result["net_itc_igst"]
        result["itc_sgst"] = result["net_itc_sgst"]

    return result


# ── Excel formula helpers ──────────────────────────────────────────
def _is_formula(v): return isinstance(v, str) and v.startswith("=")

def _fsum(col_letter, row_start, row_end):
    """=SUM(B3:B14)"""
    return f"=SUM({col_letter}{row_start}:{col_letter}{row_end})"

def _fdiff(col_a, col_b, row):
    """=D14-C14"""
    return f"={col_a}{row}-{col_b}{row}"

def write_annual_reconciliation(client_dir, client_name, gstin, log):
    """
    Builds ANNUAL_RECONCILIATION Excel with 7 sheets modelled on
    BHAVANI ELECTRICAL 2024-25.xlsm  Summary Report format.

    Sheets:
      1. Summary_Report       — GSTR-3B monthwise (matches xlsm Summary Report)
      2. GSTR1_Sales          — B2B + B2CS invoice detail from JSON ZIPs
      3. GSTR2B_ITC           — ITC detail from GSTR-2B ZIPs
      4. GSTR2A_Purchases     — Purchase detail from GSTR-2A ZIPs
      5. GSTR3B_Status        — PDF download status for each month
      6. R1_vs_3B_Recon       — GSTR-1 output tax vs 3B liability comparison
      7. Tax_Liability_Compare— Portal Tax Liabilities & ITC Comparison report
    """
    log.info(f"\n  Building Annual Reconciliation for {client_name}...")

    wb = Workbook()
    wb.remove(wb.active)

    # -- colour/style helpers ------------------------------
    HDR_BG   = "1F3864"; HDR_FG = "FFFFFF"
    SEC_BG   = "2E75B6"; SEC_FG = "FFFFFF"
    TOT_BG   = "D6DCE4"; TOT_FG = "000000"
    IGST_BG  = "DEEAF1"; CGST_BG = "E2EFDA"; SGST_BG = "FFF2CC"
    ALT1     = "FFFFFF"; ALT2 = "F2F2F2"
    NUM_FMT  = "#,##0.00"
    NUM_FMT0 = "#,##0"

    def _f(h): return PatternFill("solid", fgColor=h)
    def _font(bold=False, color="000000", size=9):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def _bdr():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)
    def _aln(h="left", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    def title(ws, text, ncols, bg=HDR_BG):
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c = ws["A1"]
        c.value = text
        c.font = Font(name="Arial", bold=True, color=HDR_FG, size=12)
        c.fill = _f(bg); c.alignment = _aln("center"); c.border = _bdr()
        ws.row_dimensions[1].height = 28

    def hdr(ws, labels_widths, row=2, bg=SEC_BG):
        for ci, (lbl, w) in enumerate(labels_widths, 1):
            c = ws.cell(row=row, column=ci, value=lbl)
            c.font = _font(True, HDR_FG, 9)
            c.fill = _f(bg)
            c.alignment = _aln("center")
            c.border = _bdr()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[row].height = 20

    def cell(ws, r, c, v, bg=ALT1, bold=False, fg="000000", numfmt=None, align="left"):
        cl = ws.cell(row=r, column=c, value=v)
        cl.font = _font(bold, fg, 9)
        cl.fill = _f(bg)
        cl.alignment = _aln(align)
        cl.border = _bdr()
        is_num = isinstance(v, (int, float))
        is_fml = isinstance(v, str) and v.startswith("=")
        if numfmt and (is_num or is_fml):
            cl.number_format = numfmt
        elif is_fml and not numfmt:
            cl.number_format = NUM_FMT  # default fmt for formula cells
        return cl

    def totrow(ws, r, vals, bg=TOT_BG, fg=TOT_FG):
        for ci, v in enumerate(vals, 1):
            cl = ws.cell(row=r, column=ci, value=v)
            cl.font = _font(True, fg, 9)
            cl.fill = _f(bg)
            is_num = isinstance(v, (int, float))
            is_fml = isinstance(v, str) and v.startswith("=")
            cl.alignment = _aln("right" if (is_num or is_fml) else "left")
            cl.border = _bdr()
            if is_num or is_fml:
                cl.number_format = NUM_FMT
        ws.row_dimensions[r].height = 18

    def secrow(ws, r, label, ncols, bg=SEC_BG):
        ws.merge_cells(f"A{r}:{get_column_letter(ncols)}{r}")
        c = ws.cell(row=r, column=1, value=label)
        c.font = _font(True, SEC_FG, 9)
        c.fill = _f(bg)
        c.alignment = _aln("left")
        c.border = _bdr()
        ws.row_dimensions[r].height = 16

    # -- DATA READERS -------------------------------------

    def _find_col(cl_dict, patterns):
        for p in patterns:
            for k, v in cl_dict.items():
                if p in k: return v
        return None

    def _read_xl(path, prefix):
        """
        Read GST portal Excel with multi-row merged header support.
        Portal GSTR-2A/2B Excel format:
          Row N  : GSTIN | Trade Name | Invoice details (merged) | Taxable Value | IGST | CGST | SGST
          Row N+1: (blank)| (blank)   | Inv No | Date | Inv Value | (same cols cont.)
        Combines both rows to get flat unique column names.
        """
        try:
            xl = pd.ExcelFile(path)
            preferred = ["b2b","gstr","data","sheet1"]
            sheet = next((s for s in xl.sheet_names
                          if any(p in s.lower() for p in preferred)), xl.sheet_names[0])
            log.info(f"    {prefix} sheet='{sheet}'")

            raw = pd.read_excel(path, sheet_name=sheet, header=None, dtype=str)

            # Find row containing GSTIN/supplier
            hdr_idx = None
            for i in range(min(25, len(raw))):
                rvals = [str(v).lower().strip() for v in raw.iloc[i]
                         if pd.notna(v) and str(v).strip() not in ("nan","")]
                if any("gstin" in v or "ctin" in v or "supplier" in v for v in rvals):
                    hdr_idx = i
                    break

            if hdr_idx is None:
                log.warning(f"    {prefix} GSTIN header not found")
                return None, {}

            # Build column names from row hdr_idx, forward-filling merged cells,
            # then combine with sub-header row hdr_idx+1 for tax column names
            def row_to_list(ridx):
                if ridx >= len(raw): return []
                return [str(v).strip() if pd.notna(v) and str(v).strip() not in ("nan","") else ""
                        for v in raw.iloc[ridx]]

            top  = row_to_list(hdr_idx)
            sub  = row_to_list(hdr_idx + 1)

            # Forward-fill merged cells in top row
            last = ""
            ff_top = []
            for v in top:
                if v and not v.startswith("Unnamed"):
                    last = v
                ff_top.append(last)

            # Combine: prefer sub if it has content and looks like a column name
            combined = []
            for parent, s in zip(ff_top, sub + [""]*(len(ff_top)-len(sub))):
                s_clean = s if s and not s.startswith("Unnamed") else ""
                if s_clean and s_clean.lower() != parent.lower():
                    col = f"{parent} {s_clean}".strip() if parent else s_clean
                else:
                    col = parent or s_clean or f"col_{len(combined)}"
                combined.append(col)

            log.info(f"    {prefix} header_row={hdr_idx}, cols={combined[:8]}")

            # Data starts after sub-header row
            data_start = hdr_idx + 2
            df = raw.iloc[data_start:].copy().reset_index(drop=True)
            # Assign combined column names
            ncols = len(df.columns)
            df.columns = (combined[:ncols] +
                          [f"_extra_{i}" for i in range(max(0, ncols - len(combined)))])

            cl = {str(c).lower().strip(): c for c in df.columns}
            return df, cl

        except Exception as e:
            log.warning(f"    {prefix} Excel read error: {e}")
        return None, {}

    def read_gstr1(month_name, year):
        """Returns (totals_dict, rows_list).
        rows_list: each row = (inv_type, gstin, receiver_name, inv_no, inv_date,
                               inv_value, place_of_supply, rate, taxable, igst, cgst, sgst)
        """
        t = {"inv":0,"b2b_tx":0.0,"b2cs_tx":0.0,"b2cl_tx":0.0,"igst":0.0,"cgst":0.0,"sgst":0.0,"val":0.0,
                 "cdn_cr":0.0,"cdn_dr":0.0,"cdn_ig":0.0,"cdn_cg":0.0,"cdn_sg":0.0,
                 "cdn_amend_cr":0.0,"cdn_amend_dr":0.0,  # CDNRA amendments
                 "exp_tx":0.0,"exp_igst":0.0,
                 "rcm_tx":0.0,"rcm_igst":0.0,"rcm_cgst":0.0,"rcm_sgst":0.0,
                 "rate_0":0.0,"rate_3":0.0,"rate_5":0.0,"rate_12":0.0,
                 "rate_18":0.0,"rate_28":0.0,"rate_other":0.0,
                 "nil_exempt":0.0,"nil_rated":0.0,"exempted":0.0,"non_gst":0.0}
        rows = []
        cdn_rows = []
        zp = Path(client_dir)/f"GSTR1_{month_name}_{year}.zip"
        if not zp.exists(): return t, rows, cdn_rows
        try:
            ed = Path(client_dir)/f"GSTR1_{month_name}_{year}_ex"
            ed.mkdir(exist_ok=True)
            with zipfile.ZipFile(zp) as z: z.extractall(ed)
            jf = list(ed.glob("*.json")) + list(ed.glob("**/*.json"))
            if not jf: return t, rows, cdn_rows
            with open(jf[0], encoding="utf-8") as f: d = json.load(f)

            # B2B invoices
            for p in d.get("b2b",[]):
                ctin = p.get("ctin","")
                rname = p.get("trdnm","")
                for inv in p.get("inv",[]):
                    t["inv"]+=1
                    inv_val = float(inv.get("val",0) or 0)
                    t["val"]+=inv_val
                    inv_tx = inv_ig = inv_cg = inv_sg = 0.0
                    rates_seen = []   # collect ALL rates in this invoice
                    for it in inv.get("itms",[]):
                        x=it.get("itm_det",{})
                        rt_v = x.get("rt", 0)
                        if rt_v and rt_v not in rates_seen:
                            rates_seen.append(rt_v)
                        inv_tx+=float(x.get("txval",0) or 0)
                        inv_ig+=float(x.get("iamt",0) or 0)
                        inv_cg+=float(x.get("camt",0) or 0)
                        inv_sg+=float(x.get("samt",0) or 0)
                    # Build rate label: "18" or "5/18" or "0/5/18"
                    if len(rates_seen)==0:
                        rate_label = 0
                    elif len(rates_seen)==1:
                        rate_label = rates_seen[0]
                    else:
                        rate_label = "/".join(str(int(r) if r==int(r) else r)
                                              for r in sorted(rates_seen))
                    t["b2b_tx"]+=inv_tx; t["igst"]+=inv_ig
                    t["cgst"]+=inv_cg; t["sgst"]+=inv_sg
                    # Accumulate rate-wise
                    for it_r in inv.get("itms",[]):
                        x_r=it_r.get("itm_det",{})
                        rt_r=x_r.get("rt",0); tx_r=float(x_r.get("txval",0) or 0)
                        key_r = {0:"rate_0",3:"rate_3",5:"rate_5",12:"rate_12",
                                 18:"rate_18",28:"rate_28"}.get(rt_r,"rate_other")
                        t[key_r]+=tx_r
                    rows.append(("B2B", ctin, rname,
                                 inv.get("inum",""), inv.get("idt",""),
                                 round(inv_val,2), inv.get("pos",""), rate_label,
                                 round(inv_tx,2), round(inv_ig,2),
                                 round(inv_cg,2), round(inv_sg,2)))

            # B2CS (small consumers — unregistered, no invoice detail)
            for r in d.get("b2cs",[]):
                tv=float(r.get("txval",0) or 0)
                ig=float(r.get("iamt",0) or 0)
                cg=float(r.get("camt",0) or 0)
                sg=float(r.get("samt",0) or 0)
                t["b2cs_tx"]+=tv; t["igst"]+=ig; t["cgst"]+=cg; t["sgst"]+=sg
                rt_b2cs=r.get("rt",0)
                key_r={0:"rate_0",3:"rate_3",5:"rate_5",12:"rate_12",
                       18:"rate_18",28:"rate_28"}.get(rt_b2cs,"rate_other")
                t[key_r]+=tv
                rows.append(("B2CS","","(Small Consumers)",
                             "-","-", round(tv+ig+cg+sg,2),
                             r.get("pos",""), r.get("rt",0),
                             round(tv,2), round(ig,2), round(cg,2), round(sg,2)))

            # B2CL (large consumers — unregistered, inter-state)
            for r in d.get("b2cl",[]):
                for inv in r.get("inv",[]):
                    t["inv"]+=1
                    inv_val=float(inv.get("val",0) or 0)
                    t["val"]+=inv_val
                    inv_tx=inv_ig=0.0; b2cl_rates=[]
                    for it in inv.get("itms",[]):
                        x=it.get("itm_det",{})
                        rt_v=x.get("rt",0)
                        if rt_v and rt_v not in b2cl_rates: b2cl_rates.append(rt_v)
                        inv_tx+=float(x.get("txval",0) or 0)
                        inv_ig+=float(x.get("iamt",0) or 0)
                    b2cl_rate = (b2cl_rates[0] if len(b2cl_rates)==1
                                 else "/".join(str(int(r) if r==int(r) else r)
                                               for r in sorted(b2cl_rates)) if b2cl_rates else 0)
                    t["b2cl_tx"]+=inv_tx; t["igst"]+=inv_ig  # B2CL is separate from B2B
                    rows.append(("B2CL","","(Large Consumer)",
                                 inv.get("inum",""),inv.get("idt",""),
                                 round(inv_val,2),r.get("pos",""),b2cl_rate,
                                 round(inv_tx,2),round(inv_ig,2),0.0,0.0))

            # CDNR — Credit/Debit Notes (Registered)
            for p in d.get("cdnr",[]):
                ctin=p.get("ctin",""); rname=p.get("trdnm","")
                for note in p.get("nt",[]):
                    nv=float(note.get("val",0) or 0)
                    nt_tx=nt_ig=nt_cg=nt_sg=0.0
                    for it in note.get("itms",[]):
                        x=it.get("itm_det",{})
                        nt_tx+=float(x.get("txval",0) or 0)
                        nt_ig+=float(x.get("iamt",0) or 0)
                        nt_cg+=float(x.get("camt",0) or 0)
                        nt_sg+=float(x.get("samt",0) or 0)
                    ntype=note.get("ntty","CDN")
                    rows.append((f"CDN-{ntype}", ctin, rname,
                                 note.get("nt_num",""), note.get("nt_dt",""),
                                 round(nv,2), note.get("pos",""), 0,
                                 round(nt_tx,2),round(nt_ig,2),round(nt_cg,2),round(nt_sg,2)))

            log.info(f"    GSTR1 {month_name}: {t['inv']} invoices, taxable ₹{t['b2b_tx']+t['b2cs_tx']:,.0f}")
            # -- CDNR — Credit/Debit Notes (Registered buyers) --
            for entry in d.get("cdnr",[]):
                ctin = entry.get("ctin","")
                rnm  = entry.get("trdnm","")
                for note in entry.get("nt",[]):
                    ntype = note.get("ntty","")   # C=Credit, D=Debit
                    nt_num = note.get("nt_num","")
                    nt_dt  = note.get("nt_dt","")
                    val    = float(note.get("val",0) or 0)
                    ig=cg=sg=tv=0.0
                    for it in note.get("itms",[]):
                        x=it.get("itm_det",{})
                        tv+=float(x.get("txval",0) or 0)
                        ig+=float(x.get("iamt",0) or 0)
                        cg+=float(x.get("camt",0) or 0)
                        sg+=float(x.get("samt",0) or 0)
                    label = "CDNR-Credit" if ntype in ("C","") else "CDNR-Debit"
                    if ntype == "D":
                        t["cdn_dr"] += tv
                    else:
                        t["cdn_cr"] += tv
                    t["cdn_ig"]+=ig; t["cdn_cg"]+=cg; t["cdn_sg"]+=sg
                    cdn_rows.append((label, ctin, rnm, nt_num, nt_dt,
                                     round(val,2), "", 0,
                                     round(tv,2), round(ig,2), round(cg,2), round(sg,2)))
                    rows.append((label, ctin, rnm, nt_num, nt_dt,
                                 round(val,2), "", 0,
                                 round(tv,2), round(ig,2), round(cg,2), round(sg,2)))

            # -- CDNUR — Credit/Debit Notes (Unregistered buyers) --
            for entry in d.get("cdnur",[]):
                ntype  = entry.get("ntty","")
                nt_num = entry.get("nt_num","")
                nt_dt  = entry.get("nt_dt","")
                val    = float(entry.get("val",0) or 0)
                ig=cg=sg=tv=0.0
                for it in entry.get("itms",[]):
                    x=it.get("itm_det",{})
                    tv+=float(x.get("txval",0) or 0)
                    ig+=float(x.get("iamt",0) or 0)
                    cg+=float(x.get("camt",0) or 0)
                    sg+=float(x.get("samt",0) or 0)
                label = "CDNUR-Cr" if ntype in ("C","") else "CDNUR-Dr"
                if ntype == "D":
                    t["cdn_dr"] += tv
                else:
                    t["cdn_cr"] += tv
                t["cdn_ig"]+=ig; t["cdn_cg"]+=cg; t["cdn_sg"]+=sg
                cdn_rows.append((label,"(Unregistered)","(B2CL unregistered)",
                                 nt_num, nt_dt, round(val,2),"",0,
                                 round(tv,2),round(ig,2),round(cg,2),round(sg,2)))
                rows.append((label,"(Unregistered)","(B2CL unregistered)",
                             nt_num, nt_dt, round(val,2),"",0,
                             round(tv,2),round(ig,2),round(cg,2),round(sg,2)))

            # -- CDNRA — Amended Credit/Debit Notes (GSTR-1A) ------
            for entry in d.get("cdnra",[]):
                ctin  = entry.get("ctin","")
                rnm   = gstin_name_map.get(ctin,"") or entry.get("trdnm","")
                for note in entry.get("nt",[]):
                    ntype  = note.get("ntty","")
                    nt_num = note.get("nt_num","")
                    nt_dt  = note.get("nt_dt","")
                    val    = float(note.get("val",0) or 0)
                    ig=cg=sg=tv=0.0
                    for it in note.get("itms",[]):
                        x=it.get("itm_det",{})
                        tv+=float(x.get("txval",0) or 0)
                        ig+=float(x.get("iamt",0) or 0)
                        cg+=float(x.get("camt",0) or 0)
                        sg+=float(x.get("samt",0) or 0)
                    label_a = "CDNRA-Credit" if ntype in ("C","") else "CDNRA-Debit"
                    if ntype == "D":
                        t["cdn_amend_dr"] += tv
                    else:
                        t["cdn_amend_cr"] += tv
                    t["cdn_ig"]+=ig; t["cdn_cg"]+=cg; t["cdn_sg"]+=sg
                    cdn_rows.append((label_a, ctin, rnm, nt_num, nt_dt,
                                     round(val,2), "", 0,
                                     round(tv,2), round(ig,2), round(cg,2), round(sg,2)))
                    rows.append((label_a, ctin, rnm, nt_num, nt_dt,
                                 round(val,2), "", 0,
                                 round(tv,2), round(ig,2), round(cg,2), round(sg,2)))

            # -- SEZ Supplies (with/without tax payment) ------------------
            # JSON keys: sezwp = SEZ with payment, sezwop = SEZ without payment
            for sez_key in ("sezwp", "sezwop"):
                for sez_entry in d.get(sez_key, []):
                    for sez_inv in sez_entry.get("inv", []):
                        sez_val = float(sez_inv.get("val", 0) or 0)
                        sez_tx = sez_ig = sez_cg = sez_sg = 0.0
                        for it in sez_inv.get("itms", []):
                            det = it.get("itm_det", {})
                            sez_tx += float(det.get("txval", 0) or 0)
                            sez_ig += float(det.get("iamt", 0) or 0)
                            sez_cg += float(det.get("camt", 0) or 0)
                            sez_sg += float(det.get("samt", 0) or 0)
                        ctin_s = sez_entry.get("ctin", "")
                        nm_s   = gstin_name_map.get(ctin_s, "") or sez_entry.get("trdnm", "")
                        t["b2b_tx"] += sez_tx   # SEZ is IGST-liable outward supply
                        t["igst"]   += sez_ig
                        t["val"]    += sez_val
                        t["inv"]    += 1
                        rows.append((f"SEZ-{sez_key.upper()}", ctin_s, nm_s,
                                     sez_inv.get("inum",""), sez_inv.get("idt",""),
                                     round(sez_val,2), sez_inv.get("pos",""), 0,
                                     round(sez_tx,2), round(sez_ig,2),
                                     round(sez_cg,2), round(sez_sg,2)))

            # -- Nil-rated, Exempted, Non-GST supplies ----------------------
            # GST portal JSON uses multiple possible structures:
            #   {"nil_sup": [{"sply_ty":"INTRB2B","nil_amt":0,"expt_amt":0,"ngsup_amt":0}]}
            #   {"nil": {"inv": [{"sply_ty":"INTRB2B","nil":0,"expt":0,"ngsup":0}]}}
            #   {"exp": [...]}  for exports
            # Collect nil records from all known keys
            nil_records = []
            # Form 1: nil_sup as list
            ns = d.get("nil_sup", d.get("nil", None))
            if isinstance(ns, list):
                nil_records = ns
            elif isinstance(ns, dict):
                nil_records = ns.get("inv", ns.get("details", []))
            # Also check alternate keys used by different portal versions
            for alt_key in ["expt","nil_details","nil_rated"]:
                alt = d.get(alt_key)
                if isinstance(alt, list): nil_records += alt
                elif isinstance(alt, dict): nil_records += alt.get("inv",[])

            for rec in nil_records:
                if not isinstance(rec, dict): continue
                stype = rec.get("sply_ty","")
                # Field names vary: nil_amt vs nil, expt_amt vs expt, ngsup_amt vs ngsup
                nil_v  = float(rec.get("nil_amt", rec.get("nil",  0)) or 0)
                expt_v = float(rec.get("expt_amt",rec.get("expt", 0)) or 0)
                ngsup_v= float(rec.get("ngsup_amt",rec.get("ngsup",0)) or 0)
                total_nil = nil_v + expt_v + ngsup_v
                if total_nil == 0: continue
                t["nil_exempt"] += nil_v + expt_v
                t["nil_rated"]  += nil_v
                t["exempted"]   += expt_v
                t["non_gst"]    += ngsup_v
                rows.append(("NIL/EXEMPT", stype, f"Nil={nil_v:.0f} Expt={expt_v:.0f} NonGST={ngsup_v:.0f}",
                             "-", "-", round(total_nil,2), "", 0,
                             round(total_nil,2), 0, 0, 0))
            log.info(f"    GSTR1 {month_name}: nil_exempt={t['nil_exempt']:.0f} non_gst={t['non_gst']:.0f}")

        except Exception as e:
            log.warning(f"    GSTR1 read [{month_name}]: {e}")
        return t, rows, cdn_rows

    def read_gstr2b(month_name, year):
        """Read GSTR-2B Excel.
        Portal downloads as direct .xlsx per month.
        Confirmed column layout from sample file:
          Col 0=GSTIN, 1=TradeName, 2=InvNo, 3=InvType, 4=InvDate,
          5=InvValue, 6=POS, 7=ReverseCharge, 8=Rate, 9=TaxableValue,
          10=IGST, 11=CGST, 12=SGST, 13=Cess, 14=Period, 15=FilingDate,
          16=ITCAvailability
        Header rows: row index 4 (main) + row index 5 (sub-headers), data from row 6.
        """
        itc={"igst":0.0,"cgst":0.0,"sgst":0.0}; rows=[]
        # GSTR-2B downloads as direct Excel (.xlsx) OR inside a ZIP
        direct_xl = Path(client_dir)/f"GSTR2B_{month_name}_{year}.xlsx"
        zp         = Path(client_dir)/f"GSTR2B_{month_name}_{year}.zip"

        # Also scan for portal-named files like: 082024_33AAHFE3141K1ZN_GSTR2B_*.xlsx
        if not direct_xl.exists():
            # Try month number pattern
            mnum = {"April":"04","May":"05","June":"06","July":"07","August":"08",
                    "September":"09","October":"10","November":"11","December":"12",
                    "January":"01","February":"02","March":"03"}.get(month_name,"")
            yr2 = year[2:]  # 2024→24
            matches = (list(Path(client_dir).glob(f"{mnum}{year[-4:]}*GSTR2B*.xlsx")) +
                       list(Path(client_dir).glob(f"*GSTR2B*{mnum}{year[-4:]}*.xlsx")) +
                       list(Path(client_dir).glob(f"*{mnum}{yr2}*GSTR*2B*.xlsx")))
            if matches:
                direct_xl = matches[0]
                log.info(f"    GSTR2B {month_name}: found portal-named file {direct_xl.name}")

        # If direct Excel exists, use it directly
        if direct_xl.exists():
            try:
                log.info(f"    GSTR2B {month_name}: reading {direct_xl.name}")
                # Read raw to handle multi-row merged headers
                raw = pd.read_excel(direct_xl, sheet_name="B2B", header=None, dtype=str)
                # Data starts at row 6 (index 6) based on confirmed sample layout
                # Header rows are at 4 and 5 (merged) — use fixed column positions
                data_rows = raw.iloc[6:].reset_index(drop=True)
                for _, row in data_rows.iterrows():
                    sup = clean_str(row.iloc[0])
                    if not sup or sup.lower() in ("nan","none","","-","gstin"): continue
                    try:
                        nm   = clean_str(row.iloc[1])
                        inum = clean_str(row.iloc[2])
                        idt  = clean_str(row.iloc[4])
                        iv   = clean_num(row.iloc[5])
                        pos  = clean_str(row.iloc[6])
                        rate = clean_num(row.iloc[8])
                        tv   = clean_num(row.iloc[9])
                        ig   = clean_num(row.iloc[10])
                        cg   = clean_num(row.iloc[11])
                        sg   = clean_num(row.iloc[12])
                        itc_avail = clean_str(row.iloc[16]) if len(row) > 16 else "Yes"
                        rows.append((sup, nm, inum, idt, round(iv,2), pos, rate,
                                     round(tv,2), round(ig,2), round(cg,2), round(sg,2),
                                     itc_avail))
                        itc["igst"]+=ig; itc["cgst"]+=cg; itc["sgst"]+=sg
                    except: continue
                log.info(f"    GSTR2B {month_name}: {len(rows)} records, ITC ₹{sum(itc.values()):,.0f}")
                return itc, rows
            except Exception as e:
                log.warning(f"    GSTR2B direct Excel read [{month_name}]: {e}")
                # Fallback to column-name detection
                try:
                    df,cl=_read_xl(direct_xl, f"GSTR2B {month_name}")
                    if df is not None:
                        gc=lambda ps: _find_col(cl,ps)
                        gstin_c=gc(["gstin of supplier","gstin","ctin"])
                        nm_c=gc(["trade/legal name","trade name","supplier name"])
                        inv_c=gc(["invoice number","invoice no"])
                        dt_c=gc(["invoice date","doc date"])
                        tv_c=gc(["taxable value","taxable value (₹)"])
                        ig_c=gc(["integrated tax","igst","integrated tax(₹)"])
                        cg_c=gc(["central tax","cgst","central tax(₹)"])
                        sg_c=gc(["state/ut tax","sgst","state/ut tax(₹)"])
                        for _,row in df.iterrows():
                            sup=clean_str(row.get(gstin_c,"") if gstin_c else "")
                            if not sup or sup.lower() in ("nan","none","","-"): continue
                            ig=clean_num(row.get(ig_c,0) if ig_c else 0)
                            cg=clean_num(row.get(cg_c,0) if cg_c else 0)
                            sg=clean_num(row.get(sg_c,0) if sg_c else 0)
                            tv=clean_num(row.get(tv_c,0) if tv_c else 0)
                            rows.append((sup,
                                clean_str(row.get(nm_c,"") if nm_c else ""),
                                clean_str(row.get(inv_c,"") if inv_c else ""),
                                clean_str(row.get(dt_c,"") if dt_c else ""),
                                0.0,"",0,round(tv,2),round(ig,2),round(cg,2),round(sg,2),"Yes"))
                            itc["igst"]+=ig; itc["cgst"]+=cg; itc["sgst"]+=sg
                        log.info(f"    GSTR2B {month_name}: {len(rows)} records (fallback)")
                        return itc, rows
                except: pass

        if not zp.exists():
            log.info(f"    GSTR2B {month_name}: no file (xlsx or zip)")
            return itc, rows
        try:
            ed = Path(client_dir)/f"GSTR2B_{month_name}_{year}_ex"
            ed.mkdir(exist_ok=True)
            with zipfile.ZipFile(zp) as z: z.extractall(ed)
            jf = list(ed.glob("*.json"))+list(ed.glob("**/*.json"))
            if jf:
                with open(jf[0], encoding="utf-8") as f: d=json.load(f)
                tmp=d
                for kp in [["data","docdata"],["data"],["docdata"]]:
                    t2=d
                    for k in kp:
                        if isinstance(t2,dict): t2=t2.get(k,t2)
                    if isinstance(t2,dict) and "b2b" in t2: tmp=t2; break
                for p in tmp.get("b2b",[]):
                    stin=p.get("ctin",""or p.get("gstin",""))
                    nm=p.get("trdnm","")
                    for doc in p.get("docs",p.get("inv",[])):
                        ig=float(doc.get("igst",doc.get("iamt",0)) or 0)
                        cg=float(doc.get("cgst",doc.get("camt",0)) or 0)
                        sg=float(doc.get("sgst",doc.get("samt",0)) or 0)
                        tv=float(doc.get("txval",doc.get("val",0)) or 0)
                        rows.append((stin,nm,doc.get("inum",""),doc.get("idt",""),
                                     round(tv,2),round(ig,2),round(cg,2),round(sg,2)))
                        itc["igst"]+=ig; itc["cgst"]+=cg; itc["sgst"]+=sg
            else:
                xlf=(list(ed.glob("*.xlsx"))+list(ed.glob("**/*.xlsx"))+
                     list(ed.glob("*.xls"))+list(ed.glob("**/*.xls")))
                if xlf:
                    log.info(f"    GSTR2B {month_name}: reading Excel {xlf[0].name}")
                    df,cl=_read_xl(xlf[0], f"GSTR2B {month_name}")
                    if df is not None:
                        gc=lambda ps: _find_col(cl,ps)
                        gstin_c=gc(["gstin of supplier","gstin","ctin"])
                        nm_c=gc(["trade/legal name","trade name","supplier name","legal name"])
                        inv_c=gc(["invoice number","invoice no","doc no"])
                        dt_c=gc(["invoice date","doc date"])
                        tv_c=gc(["taxable value","taxable"])
                        ig_c=gc(["integrated tax","igst","iamt"])
                        cg_c=gc(["central tax","cgst","camt"])
                        sg_c=gc(["state/ut tax","state tax","sgst","utgst"])
                        log.info(f"    GSTR2B cols: gstin={gstin_c} tv={tv_c} ig={ig_c}")
                        for _,row in df.iterrows():
                            sup=clean_str(row.get(gstin_c,"") if gstin_c else "")
                            if not sup or sup.lower() in ("nan","none","","-"): continue
                            ig=clean_num(row.get(ig_c,0) if ig_c else 0)
                            cg=clean_num(row.get(cg_c,0) if cg_c else 0)
                            sg=clean_num(row.get(sg_c,0) if sg_c else 0)
                            tv=clean_num(row.get(tv_c,0) if tv_c else 0)
                            rows.append((sup,
                                clean_str(row.get(nm_c,"") if nm_c else ""),
                                clean_str(row.get(inv_c,"") if inv_c else ""),
                                clean_str(row.get(dt_c,"") if dt_c else ""),
                                round(tv,2),round(ig,2),round(cg,2),round(sg,2)))
                            itc["igst"]+=ig; itc["cgst"]+=cg; itc["sgst"]+=sg
            log.info(f"    GSTR2B {month_name}: {len(rows)} records, ITC ₹{sum(itc.values()):,.0f}")
        except Exception as e:
            log.warning(f"    GSTR2B read [{month_name}]: {e}")
        return itc, rows

    def read_gstr2a(month_name, year):
        """
        Read GSTR-2A data. Portal downloads as:
          a) GSTR2A_{month}_{year}.zip  (our named file after download)
          b) Direct .xlsx files: 33AAHFE3141K1ZN_042024_R2A.xlsx (portal naming)
          c) ZIP extracted already in _ex folder
        Handles multi-row merged headers via _read_xl.
        """
        tot={"tv":0.0,"igst":0.0,"cgst":0.0,"sgst":0.0}; rows=[]
        cdir_path = Path(client_dir)

        # -- Find the source file --------------------------
        # 1. Our standard named ZIP
        zp = cdir_path / f"GSTR2A_{month_name}_{year}.zip"
        # 2. Direct xlsx files — portal naming or any *R2A* file for this month/year
        mn2 = f"{month_name[:3].lower()}"  # apr, may, ...
        yr2 = year[2:]                      # 24, 25
        direct_xls = (
            list(cdir_path.glob(f"*{year[-2:]}{MONTHS[[m[0] for m in MONTHS].index(month_name)+1 if month_name in [m[0] for m in MONTHS] else 0][1] if month_name in [m[0] for m in MONTHS] else ''}*R2A*"))
            if False else  # skip complex logic — use broad glob below
            list(cdir_path.glob("*_R2A*.xlsx")) +
            list(cdir_path.glob("*_R2A*.xls"))  +
            list(cdir_path.glob("GSTR2A_*.xlsx")) +
            list(cdir_path.glob("GSTR2A_*.xls"))
        )
        # Filter to files matching this month+year
        month_num_map = {"April":"04","May":"05","June":"06","July":"07","August":"08",
                         "September":"09","October":"10","November":"11","December":"12",
                         "January":"01","February":"02","March":"03"}
        mnum = month_num_map.get(month_name, "")
        yr_short = year[2:]  # 2024 → 24
        # FIX: use exact regex match to avoid "02" matching inside "042026"
        import re as _re2a
        def _exact_month_match(stem, mnum, year, yr_short):
            # Pattern: MMYYYY or MM_YYYY or MMYY where MM is exact 2-digit month
            full_yr = "20" + yr_short  # "2026"
            patterns = [
                rf"[_\-\.]{mnum}{full_yr}[_\-\.]",   # _022026_
                rf"[_\-\.]{mnum}{full_yr}R2A",         # _022026R2A
                rf"[_\-\.]{mnum}{full_yr}$",            # ends with _022026
                rf"[_\-\.]{mnum}{yr_short}[_\-\.]",    # _0226_
                rf"{mnum}{full_yr}_R2A",                 # 022026_R2A
            ]
            return any(_re2a.search(p, stem, _re2a.IGNORECASE) for p in patterns)
        matched_direct = [
            f for f in direct_xls
            if _exact_month_match(f.stem, mnum, year, yr_short)
        ]

        def _parse_xl_to_rows(xl_path, src_label):
            nonlocal tot, rows
            df, cl = _read_xl(xl_path, f"GSTR2A {month_name}")
            if df is None: return False
            gc = lambda ps: _find_col(cl, ps)
            gstin_c = gc(["gstin of supplier","gstin","ctin","supplier gstin"])
            nm_c    = gc(["trade/legal name","trade name","supplier name","legal name"])
            inv_c   = gc(["invoice number","invoice no","doc no"])
            dt_c    = gc(["invoice date","doc date"])
            tv_c    = gc(["taxable value","taxable value (₹)","taxable"])
            ig_c    = gc(["integrated tax (₹)","integrated tax","igst","iamt"])
            cg_c    = gc(["central tax (₹)","central tax","cgst","camt"])
            sg_c    = gc(["state/ut tax (₹)","state/ut tax","state tax","sgst","utgst"])
            log.info(f"    GSTR2A cols: gstin={gstin_c} tv={tv_c} ig={ig_c} cg={cg_c} sg={sg_c}")
            found = 0
            for _, row in df.iterrows():
                sup = clean_str(row.get(gstin_c,"") if gstin_c else "")
                if not sup or len(sup) < 5 or sup.lower() in ("nan","none","","-"): continue
                tv = clean_num(row.get(tv_c, 0) if tv_c else 0)
                ig = clean_num(row.get(ig_c, 0) if ig_c else 0)
                cg = clean_num(row.get(cg_c, 0) if cg_c else 0)
                sg = clean_num(row.get(sg_c, 0) if sg_c else 0)
                rows.append((sup,
                    clean_str(row.get(nm_c,"") if nm_c else ""),
                    clean_str(row.get(inv_c,"") if inv_c else ""),
                    clean_str(row.get(dt_c,"") if dt_c else ""),
                    round(tv,2), round(ig,2), round(cg,2), round(sg,2)))
                tot["tv"]+=tv; tot["igst"]+=ig; tot["cgst"]+=cg; tot["sgst"]+=sg
                found += 1
            log.info(f"    GSTR2A {month_name} ({src_label}): {found} records, tv ₹{tot['tv']:,.0f}")
            return found > 0

        # Try direct xlsx files first
        for xl_path in matched_direct[:1]:
            try:
                log.info(f"    GSTR2A {month_name}: reading direct Excel {xl_path.name}")
                if _parse_xl_to_rows(xl_path, "direct_xlsx"):
                    return tot, rows
            except Exception as e:
                log.warning(f"    GSTR2A direct Excel [{month_name}]: {e}")

        # Try ZIP
        if zp.exists():
            try:
                ed = cdir_path / f"GSTR2A_{month_name}_{year}_ex"
                ed.mkdir(exist_ok=True)
                with zipfile.ZipFile(zp) as z: z.extractall(ed)
                # JSON inside ZIP
                jf = list(ed.glob("*.json")) + list(ed.glob("**/*.json"))
                if jf:
                    with open(jf[0], encoding="utf-8") as f: d=json.load(f)
                    for p in d.get("b2b",[]):
                        stin=p.get("ctin",""); nm=p.get("trdnm","")
                        for inv in p.get("inv",[]):
                            ig=cg=sg=tv=0.0
                            for it in inv.get("itms",[]):
                                x=it.get("itm_det",{})
                                tv+=float(x.get("txval",0) or 0)
                                ig+=float(x.get("iamt",0) or 0)
                                cg+=float(x.get("camt",0) or 0)
                                sg+=float(x.get("samt",0) or 0)
                            rows.append((stin,nm,inv.get("inum",""),inv.get("idt",""),
                                         round(tv,2),round(ig,2),round(cg,2),round(sg,2)))
                            tot["tv"]+=tv; tot["igst"]+=ig; tot["cgst"]+=cg; tot["sgst"]+=sg
                    log.info(f"    GSTR2A {month_name} (ZIP/JSON): {len(rows)} records")
                    return tot, rows
                # Excel inside ZIP
                xlf = (list(ed.glob("*.xlsx")) + list(ed.glob("**/*.xlsx")) +
                       list(ed.glob("*.xls"))  + list(ed.glob("**/*.xls")))
                if xlf:
                    log.info(f"    GSTR2A {month_name}: reading ZIP Excel {xlf[0].name}")
                    _parse_xl_to_rows(xlf[0], "zip_xlsx")
                    return tot, rows
            except Exception as e:
                log.warning(f"    GSTR2A ZIP read [{month_name}]: {e}")
        else:
            if not matched_direct:
                log.info(f"    GSTR2A {month_name}: no ZIP or direct xlsx found")

        return tot, rows

    # -----------------------------------------------------
    # Collect all monthly data
    # -----------------------------------------------------
    N     = len(MONTHS)
    NCOLS = N + 3  # section + particulars + blank + 12 months + total
    g1          = {}
    g1_inv_rows = {}   # per-invoice detail: mk → list of rows
    g1_cdn_rows = {}   # CDNR/Debit Note detail: mk → list of rows
    g2b         = {}
    g2a         = {}

    # -- Build GSTIN→name lookup from GSTR-2A / GSTR-2B ------------------
    # GSTR-1 trdnm is often blank; GSTR-2A/2B Excel always has supplier names
    gstin_name_map = {}
    cdir_path = Path(client_dir)

    # -- Load customer/party name Excel (user-provided GSTIN→Name) ----------
    # Place any of these files in the script folder or client folder:
    #   customer_names.xlsx / customers.xlsx / party_master.xlsx etc.
    # Format: any sheet, GSTIN in any column, Name in the NEXT column
    _cust_files_loaded = 0
    for _sdir in [Path(os.path.dirname(os.path.abspath(__file__))),
                  cdir_path, cdir_path.parent]:
        for _pat in ["customer_names.xlsx","customers.xlsx","party_master.xlsx",
                     "customer_master.xlsx","GSTIN_Names.xlsx","gstin_names.xlsx",
                     "customer_list.xlsx","PartyMaster.xlsx","CustomerMaster.xlsx",
                     "name_list.xlsx","client_names.xlsx"]:
            _cf = _sdir / _pat
            if not _cf.exists(): continue
            try:
                _xl = pd.ExcelFile(_cf, engine="openpyxl")
                for _sn in _xl.sheet_names:
                    try:
                        _df = _xl.parse(_sn, header=None, dtype=str, nrows=20000)
                        for _, _rw in _df.iterrows():
                            for _ci in range(min(8, len(_rw)-1)):
                                _g = str(_rw.iloc[_ci] or "").strip().upper()
                                _n = str(_rw.iloc[_ci+1] or "").strip()
                                if (len(_g)==15 and _g[:2].isdigit()
                                        and _n and _n.lower() not in ("nan","none","")):
                                    gstin_name_map[_g] = _n
                                    _cust_files_loaded += 1
                    except: pass
            except Exception as _ce:
                log.warning(f"    Customer Excel [{_cf.name}] error: {_ce}")
    if _cust_files_loaded:
        log.info(f"    Customer Excel: {_cust_files_loaded} GSTIN->Name entries loaded")

    def _scan_xl_for_names(xl_path):
        """Scan all sheets of an Excel file for GSTIN→name pairs."""
        try:
            xl_f = pd.ExcelFile(xl_path, engine="openpyxl")
            for sname in xl_f.sheet_names:
                try:
                    df_s = xl_f.parse(sname, header=None, dtype=str, nrows=5000)
                    for _, row in df_s.iterrows():
                        for col_i in range(min(4, len(row)-1)):
                            g_v = str(row.iloc[col_i] or '').strip()
                            n_v = str(row.iloc[col_i+1] or '').strip()
                            if (len(g_v)==15 and g_v[:2].isdigit() and
                                    n_v and n_v.lower() not in ('nan','none','')):
                                gstin_name_map[g_v] = n_v
                except: pass
        except: pass

    # Scan GSTR-2A ZIPs
    for zpath in list(cdir_path.glob("GSTR2A_*.zip")) + list(cdir_path.glob("*_R2A*.zip")):
        try:
            ed_nm = cdir_path / (zpath.stem + "_nmex")
            ed_nm.mkdir(exist_ok=True)
            with zipfile.ZipFile(zpath) as z2: z2.extractall(ed_nm)
            for xlf in list(ed_nm.glob("*.xlsx")) + list(ed_nm.glob("*.xls")):
                _scan_xl_for_names(xlf)
            for jf in list(ed_nm.glob("*.json")) + list(ed_nm.glob("**/*.json")):
                try:
                    dj = json.load(open(jf, encoding="utf-8"))
                    for ent in dj.get("b2b",[]):
                        g2 = ent.get("ctin","").strip(); nm2 = ent.get("trdnm","").strip()
                        if len(g2)==15 and nm2 and nm2.lower() not in ("nan","none",""):
                            gstin_name_map[g2] = nm2
                except: pass
        except: pass

    # Scan direct GSTR-2A xlsx files
    for xlf in (list(cdir_path.glob("*_R2A*.xlsx")) + list(cdir_path.glob("GSTR2A_*.xlsx"))):
        _scan_xl_for_names(xlf)

    # Scan GSTR-2B Excel files (also has supplier names)
    for xlf in (list(cdir_path.glob("GSTR2B_*.xlsx")) + list(cdir_path.glob("*_R2B*.xlsx"))):
        _scan_xl_for_names(xlf)

    # Also extract names from GSTR-2B ZIPs
    for zpath in list(cdir_path.glob("GSTR2B_*.zip")):
        try:
            ed_nb = cdir_path / (zpath.stem + "_nmex")
            ed_nb.mkdir(exist_ok=True)
            with zipfile.ZipFile(zpath) as z3: z3.extractall(ed_nb)
            for xlf in list(ed_nb.glob("*.xlsx")):
                _scan_xl_for_names(xlf)
        except: pass

    log.info(f"    GSTIN→name map: {len(gstin_name_map)} entries (from GSTR-2A/2B)")

    # -- Try public GST search API for any missing names ----------
    # Collect all GSTINs from GSTR-1 invoice rows that have no name yet
    # We do a quick offline read first to find unknown GSTINs
    unknown_gstins = set()
    for zp in list(cdir_path.glob("GSTR1_*.zip")):
        try:
            import zipfile as _zf2, json as _j2
            ed2 = cdir_path / (zp.stem + "_ex2")
            ed2.mkdir(exist_ok=True)
            with _zf2.ZipFile(zp) as z2: z2.extractall(ed2)
            for jf2 in list(ed2.glob("*.json"))+list(ed2.glob("**/*.json")):
                d2 = _j2.load(open(jf2, encoding="utf-8"))
                for p2 in d2.get("b2b",[]):
                    g2 = p2.get("ctin","").strip()
                    if g2 and g2 not in gstin_name_map:
                        unknown_gstins.add(g2)
        except: pass
    log.info(f"    Unknown GSTINs needing API lookup: {len(unknown_gstins)}")

    if unknown_gstins:
        try:
            import urllib.request, json as _j3, time as _t3
            looked_up = 0
            for gstin_q in list(unknown_gstins)[:50]:  # max 50 API calls
                try:
                    # Public GST search API (no auth needed for basic lookup)
                    url = f"https://sheet.gstincheck.co.in/check/apikey/{gstin_q}"
                    # Fallback to knowyourgst endpoint
                    url2 = f"https://www.knowyourgst.com/developers/getsinglegstin/?gstin={gstin_q}&user=apitest"
                    req = urllib.request.Request(
                        f"https://api.knowyourgst.com/gstin/{gstin_q}",
                        headers={"User-Agent":"Mozilla/5.0","Accept":"application/json"})
                    try:
                        with urllib.request.urlopen(req, timeout=3) as resp:
                            data = _j3.loads(resp.read())
                            name = (data.get("lgnm") or data.get("tradeNam") or
                                    data.get("tradeName") or data.get("legal_name") or "")
                            if name and name.lower() not in ("null","none",""):
                                gstin_name_map[gstin_q] = name
                                looked_up += 1
                    except:
                        # Try alternate public endpoint
                        req2 = urllib.request.Request(
                            f"https://sheet.gstincheck.co.in/check/39592d0c-5c45-4f20-b7d7-7a0ee92b4e38/{gstin_q}",
                            headers={"User-Agent":"Mozilla/5.0"})
                        try:
                            with urllib.request.urlopen(req2, timeout=3) as resp2:
                                data2 = _j3.loads(resp2.read())
                                name2 = (data2.get("lgnm") or data2.get("tradeNam") or "")
                                if name2:
                                    gstin_name_map[gstin_q] = name2
                                    looked_up += 1
                        except: pass
                    _t3.sleep(0.2)  # polite delay between API calls
                except: pass
            log.info(f"    API lookup complete: {looked_up} names fetched")
        except Exception as e:
            log.warning(f"    API lookup failed: {e}")

    log.info(f"    GSTIN→name map final: {len(gstin_name_map)} entries")

    for mn, _, yr in MONTHS:
        mk      = f"{mn}_{yr}"
        g1_tot, g1_rows, g1_cdn = read_gstr1(mn, yr)
        g1[mk]  = g1_tot
        # -- Patch missing receiver names using lookup map --
        patched = []
        for r in g1_rows:
            inv_type,gstin_r,nm,inum,idt,iv,pos,rate,tv,ig,cg,sg = r
            if not nm or str(nm).strip() in ('','nan','None'):
                nm = gstin_name_map.get(str(gstin_r).strip(), '')
            patched.append((inv_type,gstin_r,nm,inum,idt,iv,pos,rate,tv,ig,cg,sg))
        g1_inv_rows[mk] = patched
        g1_cdn_rows[mk] = g1_cdn
        itc, r2b_rows = read_gstr2b(mn, yr)
        g2b[mk] = {"itc": itc, "rows": r2b_rows}
        tot2a, r2a_rows = read_gstr2a(mn, yr)
        g2a[mk] = {"tot": tot2a, "rows": r2a_rows}
        # Add 2A names to map for future use
        for r2 in r2a_rows:
            if len(r2)>=2 and r2[0] and len(str(r2[0]).strip())==15:
                gstin_name_map[str(r2[0]).strip()] = str(r2[1] or '').strip()

    log.info(f"  Data loaded: "
             f"R1={sum(1 for mk in g1 if g1[mk]['inv']>0)}/12  "
             f"R2B={sum(1 for mk in g2b if g2b[mk]['rows'])}/12  "
             f"R2A={sum(1 for mk in g2a if g2a[mk]['rows'])}/12 months")

    # ======================================================
    # PRE-EXTRACT ALL GSTR-3B PDFs IMMEDIATELY
    # MUST be here — before ANY sheet code — because Python's
    # scoping rules block ALL uses of g3b_monthly if there is
    # a later assignment anywhere in this function.
    # ======================================================
    g3b_monthly = {}
    for _mn, _dummy, _yr in MONTHS:
        _pdf = Path(client_dir) / f"GSTR3B_{_mn}_{_yr}.pdf"
        g3b_monthly[f"{_mn}_{_yr}"] = extract_3b_pdf(_pdf)
    _ok_3b = sum(1 for v in g3b_monthly.values()
                 if v.get("taxable",0)>0 or v.get("o_cgst",0)>0)
    log.info(f"  3B PDFs extracted: {_ok_3b}/12 months with data")

    # -----------------------------------------------------
    # SHEET 1 — Summary_Report (GSTR-3B Monthwise format)
    # -----------------------------------------------------
    ws1 = wb.create_sheet("Summary_Report")
    ws1.sheet_view.showGridLines = False

    # Build month headers
    month_labels = [mn + "\n" + yr for mn,_,yr in MONTHS]

    # Title
    ws1.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    tc = ws1["A1"]
    tc.value = f"GST ANNUAL SUMMARY REPORT — {client_name} ({gstin}) — FY {FY_LABEL}"
    tc.font = Font(name="Arial", bold=True, color=HDR_FG, size=12)
    tc.fill = fill(HDR_BG); tc.alignment = aln(); ws1.row_dimensions[1].height = 28

    # Taxpayer info row
    ws1.merge_cells(f"A2:{get_column_letter(NCOLS)}2")
    ic = ws1["A2"]
    ic.value = f"GSTIN: {gstin}   |   Client: {client_name}   |   FY: {FY_LABEL}   |   Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ic.font = Font(name="Arial", size=9, italic=True)
    ic.fill = fill("D6DCE4"); ic.alignment = aln(h="left"); ws1.row_dimensions[2].height = 14

    # Column headers row 3 — Section | Particulars | | Apr | May | ... | Total
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 48
    ws1.column_dimensions["C"].width = 4
    for ci2, (ml, (mn,_,yr)) in enumerate(zip(month_labels, MONTHS), 4):
        ws1.column_dimensions[get_column_letter(ci2)].width = 12
        ch = ws1.cell(row=3, column=ci2, value=ml)
        ch.font = Font(name="Arial", bold=True, color=HDR_FG, size=8)
        ch.fill = fill(HDR_BG); ch.alignment = aln(h="center"); ch.border = bdr()
    # Total col
    tot_col = ws1.cell(row=3, column=4+N, value="Total")
    tot_col.font = Font(name="Arial", bold=True, color=HDR_FG, size=9)
    tot_col.fill = fill(HDR_BG); tot_col.alignment = aln(); tot_col.border = bdr()
    ws1.column_dimensions[get_column_letter(4+N)].width = 14
    # Section / Particulars headers
    for ci2, hdr_label in [(1,"Section"),(2,"Particulars")]:
        hc = ws1.cell(row=3, column=ci2, value=hdr_label)
        hc.font = Font(name="Arial", bold=True, color=HDR_FG, size=9)
        hc.fill = fill(HDR_BG); hc.alignment = aln(h="left"); hc.border = bdr()
    ws1.cell(row=3, column=3).fill = fill(HDR_BG); ws1.cell(row=3, column=3).border = bdr()
    ws1.row_dimensions[3].height = 28
    ws1.freeze_panes = "D4"

    ri = 4  # ← ri declared HERE in write_annual_reconciliation scope

    def srow(section, particulars, vals_by_month, bold=False, bg=ALT1, secbg=None):
        """Write one data row. vals_by_month = list of 12 floats."""
        nonlocal ri
        row_bg = secbg or bg
        # Section cell (merged vertically is complex — just write per row)
        c = ws1.cell(row=ri, column=1, value=section)
        c.font = _font(bold, color="000000", size=9)
        c.fill = _f(row_bg); c.alignment = _aln("left", wrap=True); c.border = _bdr()
        c = ws1.cell(row=ri, column=2, value=particulars)
        c.font = _font(False, "000000", 9)
        c.fill = _f(row_bg); c.alignment = _aln("left", wrap=True); c.border = _bdr()
        c = ws1.cell(row=ri, column=3, value="")
        c.fill = _f(row_bg); c.border = _bdr()
        total = 0.0
        for ci2, v in enumerate(vals_by_month, 4):
            cv = ws1.cell(row=ri, column=ci2, value=round(v, 2) if v else 0)
            cv.font = _font(bold, "000000", 9)
            cv.fill = _f(row_bg); cv.alignment = _aln("right"); cv.border = _bdr()
            cv.number_format = NUM_FMT; total += (v or 0)
        tc2 = ws1.cell(row=ri, column=4+N, value=round(total, 2))
        tc2.font = _font(True, "000000", 9)
        tc2.fill = _f(TOT_BG); tc2.alignment = _aln("right"); tc2.border = _bdr()
        tc2.number_format = NUM_FMT
        ws1.row_dimensions[ri].height = 16; ri += 1

    def blank_row():
        nonlocal ri
        for ci2 in range(1, NCOLS+1):
            c = ws1.cell(row=ri, column=ci2, value="")
            c.fill = _f(ALT1); c.border = _bdr()
        ws1.row_dimensions[ri].height = 8; ri += 1

    def sec_header(label):
        nonlocal ri
        ws1.merge_cells(f"A{ri}:{get_column_letter(NCOLS)}{ri}")
        c = ws1.cell(row=ri, column=1, value=label)
        c.font = _font(True, SEC_FG, 9); c.fill = _f(SEC_BG)
        c.alignment = _aln("left"); c.border = _bdr()
        ws1.row_dimensions[ri].height = 16; ri += 1

    # -- Section: Sales Summary --
    sec_header("SALES SUMMARY (GSTR-1 / GSTR-3B Table 3.1)")
    srow("Sales Summary", "3.1(a) Outward taxable supplies (excl zero-rated, nil, exempt)",
         [round(g1[f"{mn}_{yr}"]["b2b_tx"]+g1[f"{mn}_{yr}"]["b2cs_tx"],2) for mn,_,yr in MONTHS], bg="E2EFDA")
    srow("", "3.1(b) Outward taxable — Zero rated",
         [round(g3b_monthly.get(f"{mn}_{yr}",{}).get("zero_taxable",0),2) for mn,_,yr in MONTHS], bg=ALT2)
    srow("", "3.1(c) Nil rated / Exempt supplies (GSTR-1)",
         [round(g1[f"{mn}_{yr}"].get("nil_exempt",0),2) for mn,_,yr in MONTHS], bg="EBF3FB")
    srow("", "3.1(c) Nil rated / Exempt (GSTR-3B PDF)",
         [round(g3b_monthly.get(f"{mn}_{yr}",{}).get("nil_exempt",0),2) for mn,_,yr in MONTHS], bg="FFF2CC")
    srow("", "3.1(e) Non-GST outward supplies",
         [round(g1[f"{mn}_{yr}"].get("non_gst",0),2) for mn,_,yr in MONTHS], bg=ALT2)
    srow("TOTAL Taxable Turnover (B2B+B2CS+Nil+NonGST)", "",
         [round(g1[f"{mn}_{yr}"]["b2b_tx"]+g1[f"{mn}_{yr}"]["b2cs_tx"]
                +g1[f"{mn}_{yr}"].get("nil_exempt",0)+g1[f"{mn}_{yr}"].get("non_gst",0),2) for mn,_,yr in MONTHS],
         bold=True, bg=TOT_BG)
    blank_row()

    # -- Section: Tax Liability --
    sec_header("TAX LIABILITY — NON-REVERSE CHARGE (GSTR-3B Table 3.1)")
    srow("Tax Liability", "IGST",
         [round(g1[f"{mn}_{yr}"]["igst"],2) for mn,_,yr in MONTHS], bg=IGST_BG)
    srow("", "CGST",
         [round(g1[f"{mn}_{yr}"]["cgst"],2) for mn,_,yr in MONTHS], bg=CGST_BG)
    srow("", "SGST",
         [round(g1[f"{mn}_{yr}"]["sgst"],2) for mn,_,yr in MONTHS], bg=SGST_BG)
    srow("", "CESS", [0.0]*N, bg=ALT2)
    srow("TOTAL Tax Liability", "",
         [round(g1[f"{mn}_{yr}"]["igst"]+g1[f"{mn}_{yr}"]["cgst"]+g1[f"{mn}_{yr}"]["sgst"],2)
          for mn,_,yr in MONTHS], bold=True, bg=TOT_BG)
    blank_row()

    # -- Section: ITC from GSTR-2B --
    sec_header("INPUT TAX CREDIT — NON-REVERSE CHARGE (GSTR-2B Confirmed ITC)")
    srow("ITC Available", "IGST",
         [round(g2b[f"{mn}_{yr}"]["itc"]["igst"],2) for mn,_,yr in MONTHS], bg=IGST_BG)
    srow("", "CGST",
         [round(g2b[f"{mn}_{yr}"]["itc"]["cgst"],2) for mn,_,yr in MONTHS], bg=CGST_BG)
    srow("", "SGST",
         [round(g2b[f"{mn}_{yr}"]["itc"]["sgst"],2) for mn,_,yr in MONTHS], bg=SGST_BG)
    srow("", "CESS", [0.0]*N, bg=ALT2)
    srow("TOTAL ITC", "",
         [round(sum(g2b[f"{mn}_{yr}"]["itc"].values()),2) for mn,_,yr in MONTHS],
         bold=True, bg=TOT_BG)
    blank_row()

    # -- Section: Net Tax --
    sec_header("NET TAX PAYABLE  (Tax Liability minus ITC)")
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"
        tl=g1[mk]["igst"]+g1[mk]["cgst"]+g1[mk]["sgst"]
        itc_t=sum(g2b[mk]["itc"].values())
    srow("Net Tax Payable", "Total Tax − Total ITC",
         [round((g1[f"{mn}_{yr}"]["igst"]+g1[f"{mn}_{yr}"]["cgst"]+g1[f"{mn}_{yr}"]["sgst"])
                - sum(g2b[f"{mn}_{yr}"]["itc"].values()), 2) for mn,_,yr in MONTHS],
         bold=True, bg="FFC7CE")
    blank_row()

    # -- Section: ITC 2A vs 2B diff --
    sec_header("ITC RECONCILIATION  (GSTR-2A vs GSTR-2B)")
    srow("GSTR-2A Total ITC", "",
         [round(g2a[f"{mn}_{yr}"]["tot"]["igst"]+g2a[f"{mn}_{yr}"]["tot"]["cgst"]+
                g2a[f"{mn}_{yr}"]["tot"]["sgst"],2) for mn,_,yr in MONTHS], bg="FFF2CC")
    srow("GSTR-2B Total ITC", "",
         [round(sum(g2b[f"{mn}_{yr}"]["itc"].values()),2) for mn,_,yr in MONTHS], bg="E2EFDA")
    srow("Difference (2A − 2B)", "Positive = unclaimed ITC; Negative = excess claim",
         [round((g2a[f"{mn}_{yr}"]["tot"]["igst"]+g2a[f"{mn}_{yr}"]["tot"]["cgst"]+
                 g2a[f"{mn}_{yr}"]["tot"]["sgst"])
                - sum(g2b[f"{mn}_{yr}"]["itc"].values()), 2) for mn,_,yr in MONTHS],
         bold=True, bg="FFC7CE")
    blank_row()

    ws1.sheet_properties.tabColor = "1F3864"

    # ====================================================
    # SHEET 2: GSTR-1 Sales Detail
    # ====================================================
    ws2 = wb.create_sheet("GSTR1_Sales_Detail")
    ws2.sheet_view.showGridLines = False
    cols2 = [("Month",12),("Invoices",9),("B2B Taxable ₹",16),("B2CS Taxable ₹",16),
             ("Total Taxable ₹",16),("IGST ₹",13),("CGST ₹",13),("SGST ₹",13),
             ("Total Tax ₹",14),("Invoice Value ₹",16)]
    title(ws2, f"GSTR-1 Annual Sales Summary — {client_name} ({gstin}) — FY {FY_LABEL}", len(cols2))
    hdr(ws2, cols2)
    ws2.freeze_panes = "A3"
    ri2=3
    ann2={"inv":0,"b2b":0.,"b2cs":0.,"ig":0.,"cg":0.,"sg":0.,"val":0.}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; d=g1[mk]
        bg2=ALT2 if ri2%2==0 else ALT1
        tx=d["b2b_tx"]+d["b2cs_tx"]+d.get("nil_exempt",0)+d.get("non_gst",0); tt=d["igst"]+d["cgst"]+d["sgst"]
        for ci2,v in enumerate([f"{mn} {yr}",d["inv"],round(d["b2b_tx"],2),round(d["b2cs_tx"],2),
                                 round(tx,2),round(d["igst"],2),round(d["cgst"],2),round(d["sgst"],2),
                                 round(tt,2),round(d["val"],2)],1):
            cell(ws2,ri2,ci2,v,bg2,numfmt=NUM_FMT if ci2>1 else None,align="right" if ci2>1 else "left")
        ws2.row_dimensions[ri2].height=15; ri2+=1
        ann2["inv"]+=d["inv"]; ann2["b2b"]+=d["b2b_tx"]; ann2["b2cs"]+=d["b2cs_tx"]
        ann2["ig"]+=d["igst"]; ann2["cg"]+=d["cgst"]; ann2["sg"]+=d["sgst"]; ann2["val"]+=d["val"]
    totrow(ws2,ri2,["ANNUAL TOTAL",_fsum("B",3,ri2-1),_fsum("C",3,ri2-1),_fsum("D",3,ri2-1),
                    _fsum("E",3,ri2-1),_fsum("F",3,ri2-1),_fsum("G",3,ri2-1),
                    _fsum("H",3,ri2-1),f"=F{ri2}+G{ri2}+H{ri2}",_fsum("J",3,ri2-1)])
    ws2.sheet_properties.tabColor = "2E75B6"

    # ====================================================
    # SHEET 2x: GSTR-1 Rate-wise Breakdown
    # Taxable values split by GST rate (0/3/5/12/18/28) per month
    # ====================================================
    ws2x = wb.create_sheet("GSTR1_Rate_Breakdown")
    ws2x.sheet_view.showGridLines = False
    cols2x = [("Month",12),("0% ₹",13),("3% ₹",13),("5% ₹",14),
              ("12% ₹",14),("18% ₹",14),("28% ₹",14),("Other% ₹",13),
              ("Total Taxable ₹",16),("IGST ₹",13),("CGST ₹",13),("SGST ₹",13),("Total Tax ₹",14)]
    title(ws2x, f"GSTR-1 Rate-wise Taxable Breakdown -- {client_name} ({gstin}) -- FY {FY_LABEL}", len(cols2x))
    hdr(ws2x, cols2x)
    ws2x.freeze_panes = "A3"; ri2x=3
    ann2x={k:0. for k in ["r0","r3","r5","r12","r18","r28","ro","tx","ig","cg","sg"]}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; d=g1[mk]
        bg2x=ALT2 if ri2x%2==0 else ALT1
        r0=round(d.get("rate_0",0),2);   r3=round(d.get("rate_3",0),2)
        r5=round(d.get("rate_5",0),2);   r12=round(d.get("rate_12",0),2)
        r18=round(d.get("rate_18",0),2); r28=round(d.get("rate_28",0),2)
        ro=round(d.get("rate_other",0),2)
        tx=round(d["b2b_tx"]+d["b2cs_tx"],2)
        ig=round(d["igst"],2); cg=round(d["cgst"],2); sg=round(d["sgst"],2)
        for ci2x,v in enumerate([f"{mn} {yr}",r0,r3,r5,r12,r18,r28,ro,tx,ig,cg,sg,round(ig+cg+sg,2)],1):
            cell(ws2x,ri2x,ci2x,v,bg2x,numfmt=NUM_FMT if ci2x>1 else None,
                 align="right" if ci2x>1 else "left")
        ws2x.row_dimensions[ri2x].height=15; ri2x+=1
        for k,v in [("r0",r0),("r3",r3),("r5",r5),("r12",r12),
                    ("r18",r18),("r28",r28),("ro",ro),("tx",tx),
                    ("ig",ig),("cg",cg),("sg",sg)]:
            ann2x[k]+=v
    totrow(ws2x,ri2x,["ANNUAL TOTAL — use SUM formulas below",_fsum("B",3,ri2x-1),
                       round(ann2x["r0"],2),round(ann2x["r3"],2),round(ann2x["r5"],2),
                       round(ann2x["r12"],2),round(ann2x["r18"],2),round(ann2x["r28"],2),
                       round(ann2x["ro"],2),round(ann2x["tx"],2),
                       round(ann2x["ig"],2),round(ann2x["cg"],2),round(ann2x["sg"],2),
                       round(ann2x["ig"]+ann2x["cg"]+ann2x["sg"],2)])
    ws2x.sheet_properties.tabColor = "843C0C"

    # ====================================================
    # SHEET 2b: GSTR-1 Invoice Detail (per-invoice rows)
    # All invoice types: B2B, B2CS, B2CL, CDNR for all 12 months
    # Then monthly summary at end
    # ====================================================
    ws2b = wb.create_sheet("GSTR1_Invoice_Detail")
    ws2b.sheet_view.showGridLines = False
    cols2b = [("Type",8),("GSTIN/UIN Receiver",22),("Receiver Name",26),
              ("Invoice No",16),("Invoice Date",13),("Invoice Value ₹",16),
              ("Place of Supply",16),("Rate %",7),
              ("Taxable Value ₹",16),("IGST ₹",12),("CGST ₹",12),("SGST ₹",12)]
    title(ws2b, f"GSTR-1 Invoice Detail (All Types) — {client_name} ({gstin}) — FY {FY_LABEL}", len(cols2b))
    hdr(ws2b, cols2b)
    ws2b.freeze_panes = "A3"
    ri2b = 3
    prev_mn = None
    ann2b = {"tx":0.,"ig":0.,"cg":0.,"sg":0.,"iv":0.}
    for mn,_,yr in MONTHS:
        mk = f"{mn}_{yr}"
        inv_rows = g1_inv_rows.get(mk, [])
        if not inv_rows:
            continue
        # Month separator row
        ws2b.merge_cells(f"A{ri2b}:{get_column_letter(len(cols2b))}{ri2b}")
        sep = ws2b.cell(row=ri2b, column=1, value=f"-- {mn} {yr} --  ({len(inv_rows)} records)")
        sep.font = _font(True, HDR_FG, 9); sep.fill = _f(SEC_BG)
        sep.alignment = _aln("left"); sep.border = _bdr()
        ws2b.row_dimensions[ri2b].height = 14; ri2b += 1

        for row_data in inv_rows:
            inv_type,gstin_r,nm,inum,idt,iv,pos,rate,tv,ig,cg,sg = row_data
            bg2b = ALT2 if ri2b%2==0 else ALT1
            for ci2b,v in enumerate([inv_type,gstin_r,nm,inum,idt,
                                      iv,pos,rate,tv,ig,cg,sg],1):
                cell(ws2b,ri2b,ci2b,v,bg2b,
                     numfmt=NUM_FMT if ci2b in (6,9,10,11,12) else None,
                     align="right" if ci2b in (6,7,8,9,10,11,12) else "left")
            ws2b.row_dimensions[ri2b].height=14; ri2b+=1
            ann2b["tx"]+=float(tv or 0); ann2b["ig"]+=float(ig or 0)
            ann2b["cg"]+=float(cg or 0); ann2b["sg"]+=float(sg or 0)
            ann2b["iv"]+=float(iv or 0)

    # Annual total row with SUM formulas
    _inv_count = sum(len(g1_inv_rows.get(f"{mn}_{yr}",[]))for mn,_,yr in MONTHS)
    totrow(ws2b, ri2b, ["ANNUAL TOTAL","","",
                         f"Total: {_inv_count} records",
                         "","", _fsum("G",3,ri2b-1),"",
                         _fsum("I",3,ri2b-1),_fsum("J",3,ri2b-1),
                         _fsum("K",3,ri2b-1),_fsum("L",3,ri2b-1)])
    ws2b.sheet_properties.tabColor = "1F3864"

    # ====================================================
    # SHEET 2c: GSTR-1 CDNR & Debit Note Detail
    # Invoice-level rows + Monthly Summary for Credit/Debit Notes
    # ====================================================
    ws2c = wb.create_sheet("GSTR1_CDNR_Debit_Note")
    ws2c.sheet_view.showGridLines = False
    cols2c = [("Month",10),("Note Type",14),("Counterparty GSTIN",22),
              ("Counterparty Name",26),("Note Number",16),("Note Date",13),
              ("Note Value ₹",15),("Rate %",7),
              ("Taxable Value ₹",16),("IGST ₹",12),("CGST ₹",12),("SGST ₹",12),
              ("Total Tax ₹",13)]
    title(ws2c, f"GSTR-1 CDNR, Debit Notes & CDNRA Amendments — {client_name} ({gstin}) — FY {FY_LABEL}", len(cols2c))
    ws2c.merge_cells(f"A2:{get_column_letter(len(cols2c))}2")
    sh2c_note=ws2c["A2"]
    sh2c_note.value=("CDNR-Credit = Credit notes to registered buyers (reduces outward supply)  |  "
                     "CDNR-Debit = Debit notes to registered buyers (increases outward supply)  |  "
                     "CDNRA = Amended credit/debit notes via GSTR-1A  |  "
                     "CDNUR = Notes to unregistered buyers  |  "
                     "Overall values include credit, debit AND amendments for full picture")
    sh2c_note.font=_font(False,"000000",8); sh2c_note.fill=_f("FFF2CC")
    sh2c_note.alignment=_aln("left"); sh2c_note.border=_bdr()
    ws2c.row_dimensions[2].height=14
    hdr(ws2c, cols2c, row=3)
    ws2c.freeze_panes = "A4"
    ri2c = 4
    ann2c = {"cr_tv":0.,"dr_tv":0.,"amend_cr":0.,"amend_dr":0.,"ig":0.,"cg":0.,"sg":0.}

    # -- Invoice-level rows ------------------------------
    for mn,_,yr in MONTHS:
        mk = f"{mn}_{yr}"
        cdn_data = g1_cdn_rows.get(mk, [])
        if not cdn_data:
            continue
        # Month separator
        cr_count  = sum(1 for r in cdn_data if "Credit" in str(r[0]) or (r[0] in ("CDN-C","CDNUR-Cr")))
        dr_count  = sum(1 for r in cdn_data if "Debit"  in str(r[0]) or (r[0] in ("CDN-D","CDNUR-Dr")))
        amd_count = sum(1 for r in cdn_data if "CDNRA"  in str(r[0]))
        ws2c.merge_cells(f"A{ri2c}:{get_column_letter(len(cols2c))}{ri2c}")
        sep = ws2c.cell(row=ri2c, column=1,
                        value=(f"── {mn} {yr} ──  Credit:{cr_count}  Debit:{dr_count}  "
                               f"Amendments(CDNRA):{amd_count}  Total:{len(cdn_data)} notes"))
        sep.font=_font(True,HDR_FG,9); sep.fill=_f(SEC_BG)
        sep.alignment=_aln("left"); sep.border=_bdr()
        ws2c.row_dimensions[ri2c].height=14; ri2c+=1

        for row_data in cdn_data:
            inv_type,gstin_r,nm,inum,idt,iv,pos,rate,tv,ig,cg,sg = row_data
            bg2c = ALT2 if ri2c%2==0 else ALT1
            # Colour coding: Debit=yellow, Credit=green, Amendment=blue
            if "Debit" in str(inv_type) or "Dr" in str(inv_type):
                type_bg = "FFF2CC"
            elif "CDNRA" in str(inv_type):
                type_bg = "DEEAF1"
            else:
                type_bg = "E2EFDA"
            for ci2c,v in enumerate([mn+" "+yr, inv_type,gstin_r,nm,inum,idt,
                                      iv,rate,tv,ig,cg,sg,round(float(ig or 0)+float(cg or 0)+float(sg or 0),2)],1):
                cl2c = ws2c.cell(row=ri2c,column=ci2c,value=v)
                cl2c.font = _font(False,"000000",9)
                cl2c.fill = _f(type_bg if ci2c==2 else bg2c)
                cl2c.alignment=_aln("right" if ci2c in (7,8,9,10,11,12,13) else "left")
                cl2c.border=_bdr()
                if ci2c in (7,9,10,11,12,13) and isinstance(v,(int,float)):
                    cl2c.number_format=NUM_FMT
            ws2c.row_dimensions[ri2c].height=14; ri2c+=1
            if "Debit" in str(inv_type) or "Dr" in str(inv_type):
                ann2c["dr_tv"]+=float(tv or 0)
            elif "CDNRA" in str(inv_type):
                if "Debit" in str(inv_type):
                    ann2c["amend_dr"]+=float(tv or 0)
                else:
                    ann2c["amend_cr"]+=float(tv or 0)
            else:
                ann2c["cr_tv"]+=float(tv or 0)
            ann2c["ig"]+=float(ig or 0); ann2c["cg"]+=float(cg or 0); ann2c["sg"]+=float(sg or 0)

    # Overall annual totals row
    totrow(ws2c, ri2c, ["ANNUAL TOTAL (All Types)",
                         _fsum("B",3,ri2c-1),_fsum("C",3,ri2c-1),
                         _fsum("D",3,ri2c-1),_fsum("E",3,ri2c-1),
                         _fsum("F",3,ri2c-1),_fsum("G",3,ri2c-1),
                         _fsum("H",3,ri2c-1),_fsum("I",3,ri2c-1),
                         _fsum("J",3,ri2c-1),f"=SUM(H{ri2c}:J{ri2c})"])
    ri2c += 2

    # -- Monthly Summary table with Amendments column ----
    ws2c.merge_cells(f"A{ri2c}:{get_column_letter(len(cols2c))}{ri2c}")
    sh2c = ws2c.cell(row=ri2c, column=1, value="MONTHLY SUMMARY — CDNR (Credit + Debit + Amendments) — Overall Values")
    sh2c.font=_font(True,HDR_FG,9); sh2c.fill=_f(SEC_BG)
    sh2c.alignment=_aln("left"); sh2c.border=_bdr()
    ws2c.row_dimensions[ri2c].height=16; ri2c+=1

    # Summary header — now includes amendments columns
    sum_headers2 = [("Month",10),
                    ("Credit Notes #",14),("Credit TV ₹",16),
                    ("Debit Notes #",14),("Debit TV ₹",16),
                    ("CDNRA Amend #",14),("Amend TV ₹",16),
                    ("Overall IGST ₹",14),("Overall CGST ₹",14),("Overall SGST ₹",14),
                    ("Overall Tax ₹",14),("Net CDN Impact ₹",16)]
    for ci2c,(h,w) in enumerate(sum_headers2,1):
        c2cs = ws2c.cell(row=ri2c,column=ci2c,value=h)
        c2cs.font=_font(True,HDR_FG,9); c2cs.fill=_f(MED_BLUE)
        c2cs.alignment=_aln("center"); c2cs.border=_bdr()
        ws2c.column_dimensions[get_column_letter(ci2c)].width=w
    ws2c.row_dimensions[ri2c].height=20; ri2c+=1

    # Summary rows
    ann_sum = {"cr_cnt":0,"cr_tv":0.,"dr_cnt":0,"dr_tv":0.,
               "amd_cnt":0,"amd_tv":0.,"ig":0.,"cg":0.,"sg":0.}
    for mn,_,yr in MONTHS:
        mk = f"{mn}_{yr}"
        cdn_data = g1_cdn_rows.get(mk, [])
        cr_cnt=dr_cnt=amd_cnt=0
        cr_tv=dr_tv=amd_tv=ig=cg=sg=0.0
        for row_data in cdn_data:
            inv_type,_2,_3,_4,_5,iv,_7,rate,tv,rig,rcg,rsg = row_data
            if "CDNRA" in str(inv_type):
                amd_cnt+=1; amd_tv+=float(tv or 0)
            elif "Debit" in str(inv_type) or "Dr" in str(inv_type):
                dr_cnt+=1; dr_tv+=float(tv or 0)
            else:
                cr_cnt+=1; cr_tv+=float(tv or 0)
            ig+=float(rig or 0); cg+=float(rcg or 0); sg+=float(rsg or 0)
        # Net CDN impact = Debit TV - Credit TV + Amendment TV (net effect on turnover)
        net_cdn = round(dr_tv - cr_tv + amd_tv, 2)
        net_bg = GREEN_BG if net_cdn >= 0 else "FFC7CE"
        bg_sum = ALT2 if ri2c%2==0 else ALT1
        row_sum_vals=[f"{mn} {yr}",cr_cnt,round(cr_tv,2),dr_cnt,round(dr_tv,2),
                      amd_cnt,round(amd_tv,2),
                      round(ig,2),round(cg,2),round(sg,2),round(ig+cg+sg,2),net_cdn]
        for ci2c,v in enumerate(row_sum_vals,1):
            c2cs=ws2c.cell(row=ri2c,column=ci2c,value=v)
            c2cs.font=_font(False,"000000",9)
            c2cs.fill=_f(net_bg if ci2c==12 else bg_sum)
            c2cs.alignment=_aln("right" if ci2c>1 else "left"); c2cs.border=_bdr()
            if ci2c in (3,5,7,8,9,10,11,12) and isinstance(v,(int,float)): c2cs.number_format=NUM_FMT
        ws2c.row_dimensions[ri2c].height=15; ri2c+=1
        ann_sum["cr_cnt"]+=cr_cnt; ann_sum["cr_tv"]+=cr_tv
        ann_sum["dr_cnt"]+=dr_cnt; ann_sum["dr_tv"]+=dr_tv
        ann_sum["amd_cnt"]+=amd_cnt; ann_sum["amd_tv"]+=amd_tv
        ann_sum["ig"]+=ig; ann_sum["cg"]+=cg; ann_sum["sg"]+=sg

    ann_net_cdn=round(ann_sum["dr_tv"]-ann_sum["cr_tv"]+ann_sum["amd_tv"],2)
    totrow(ws2c, ri2c, ["ANNUAL TOTAL",
                         _fsum("B",3,ri2c-1),_fsum("C",3,ri2c-1),
                         _fsum("D",3,ri2c-1),_fsum("E",3,ri2c-1),
                         _fsum("F",3,ri2c-1),_fsum("G",3,ri2c-1),
                         _fsum("H",3,ri2c-1),_fsum("I",3,ri2c-1),
                         _fsum("J",3,ri2c-1),f"=SUM(H{ri2c}:J{ri2c})"])
    ws2c.sheet_properties.tabColor = "FF0000"

    # ====================================================
    # SHEET 3: GSTR-2B ITC Detail
    # ====================================================
    ws3 = wb.create_sheet("GSTR2B_ITC_Detail")
    ws3.sheet_view.showGridLines = False
    cols3=[("Month",12),("Supplier GSTIN",22),("Supplier Name",28),("Invoice No",16),
           ("Invoice Date",13),("Taxable Value ₹",16),("IGST ₹",12),("CGST ₹",12),
           ("SGST ₹",12),("Total ITC ₹",14)]
    title(ws3, f"GSTR-2B ITC Details — {client_name} ({gstin}) — FY {FY_LABEL}", len(cols3))
    hdr(ws3, cols3)
    ws3.freeze_panes = "A3"
    ri3=3; ann3={"ig":0.,"cg":0.,"sg":0.}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"
        for row_data in g2b[mk]["rows"]:
            # Support both old 8-field and new 12-field format
            if len(row_data) == 12:
                stin,nm,inum,idt,iv,pos,rate,tv,ig,cg,sg,itc_avail = row_data
            else:
                stin,nm,inum,idt,tv,ig,cg,sg = row_data[:8]
                iv=0; pos=""; rate=0; itc_avail="Yes"
            bg3=ALT2 if ri3%2==0 else ALT1
            for ci3,v in enumerate([f"{mn} {yr}",stin,nm,inum,idt,tv,ig,cg,sg,round(ig+cg+sg,2)],1):
                cell(ws3,ri3,ci3,v,bg3,numfmt=NUM_FMT if ci3>5 else None,
                     align="right" if ci3>5 else "left")
            ws3.row_dimensions[ri3].height=15; ri3+=1
            ann3["ig"]+=ig; ann3["cg"]+=cg; ann3["sg"]+=sg
    totrow(ws3,ri3,["ANNUAL TOTAL","","","","","",_fsum("G",3,ri3-1),_fsum("H",3,ri3-1),
                    round(ann3["sg"],2),round(ann3["ig"]+ann3["cg"]+ann3["sg"],2)])
    ws3.sheet_properties.tabColor = "276221"

    # ====================================================
    # SHEET 4: GSTR-2A Purchase Detail
    # ====================================================
    ws4 = wb.create_sheet("GSTR2A_Purchase_Detail")
    ws4.sheet_view.showGridLines = False
    cols4=[("Month",12),("Supplier GSTIN",22),("Supplier Name",28),("Invoice No",16),
           ("Invoice Date",13),("Taxable Value ₹",16),("IGST ₹",12),("CGST ₹",12),
           ("SGST ₹",12),("Total Tax ₹",14)]
    title(ws4, f"GSTR-2A Purchase Details — {client_name} ({gstin}) — FY {FY_LABEL}", len(cols4))
    hdr(ws4, cols4)
    ws4.freeze_panes = "A3"
    ri4=3; ann4={"tv":0.,"ig":0.,"cg":0.,"sg":0.}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"
        for stin,nm,inum,idt,tv,ig,cg,sg in g2a[mk]["rows"]:
            bg4=ALT2 if ri4%2==0 else ALT1
            for ci4,v in enumerate([f"{mn} {yr}",stin,nm,inum,idt,tv,ig,cg,sg,round(ig+cg+sg,2)],1):
                cell(ws4,ri4,ci4,v,bg4,numfmt=NUM_FMT if ci4>5 else None,
                     align="right" if ci4>5 else "left")
            ws4.row_dimensions[ri4].height=15; ri4+=1
            ann4["tv"]+=tv; ann4["ig"]+=ig; ann4["cg"]+=cg; ann4["sg"]+=sg
    totrow(ws4,ri4,["ANNUAL TOTAL","","","","",_fsum("F",3,ri4-1),_fsum("G",3,ri4-1),
                    round(ann4["cg"],2),round(ann4["sg"],2),
                    round(ann4["ig"]+ann4["cg"]+ann4["sg"],2)])
    ws4.sheet_properties.tabColor = "9C6500"

    # ====================================================
    # SHEET 5: GSTR-3B Download Status
    # ====================================================
    ws5 = wb.create_sheet("GSTR3B_Status")
    ws5.sheet_view.showGridLines = False
    cols5=[("Month",14),("Status",12),
           ("Taxable Supply ₹",18),("Output IGST ₹",15),
           ("Output CGST ₹",15),("Output SGST ₹",15),("Total Tax ₹",15),
           ("ITC IGST ₹",13),("ITC CGST ₹",13),("ITC SGST ₹",13),("Total ITC ₹",14),
           ("Net Payable ₹",15),("File Name",36)]
    title(ws5, f"GSTR-3B Monthly Summary — {client_name} ({gstin}) — FY {FY_LABEL}", len(cols5))
    hdr(ws5, cols5)
    ws5.freeze_panes = "A3"
    ri5=3
    ann5={k:0.0 for k in ["taxable","nil","nongst","zr","rcm","oig","ocg","osg",
                            "iig","icg","isg","rev","nicg","nisg","int_cg","int_sg",
                            "lf_cg","lf_sg","net"]}


    for mn,_,yr in MONTHS:
        pdf=Path(client_dir)/f"GSTR3B_{mn}_{yr}.pdf"
        ok=pdf.exists()
        bg5=GREEN_BG if ok else RED_BG; fg5=GREEN_FG if ok else RED_FG
        status="Downloaded" if ok else "Missing"

        d3b = extract_3b_pdf(pdf)
        total_out = round(d3b["o_igst"]+d3b["o_cgst"]+d3b["o_sgst"],2)
        total_itc = round(d3b["itc_igst"]+d3b["itc_cgst"]+d3b["itc_sgst"],2)
        net       = round(total_out - total_itc, 2)

        row_vals = [f"{mn} {yr}", status,
                    round(d3b["taxable"],2),
                    round(d3b["o_igst"],2), round(d3b["o_cgst"],2), round(d3b["o_sgst"],2),
                    total_out,
                    round(d3b["itc_igst"],2), round(d3b["itc_cgst"],2), round(d3b["itc_sgst"],2),
                    total_itc, net,
                    pdf.name if ok else "Not downloaded"]

        bg_row = ALT2 if ri5%2==0 else ALT1
        for ci5,v in enumerate(row_vals,1):
            c5=ws5.cell(row=ri5,column=ci5,value=v)
            if ci5==2:
                c5.font=_font(True,fg5,9); c5.fill=_f(bg5)
            else:
                c5.font=_font(False,"000000",9); c5.fill=_f(bg_row)
                if isinstance(v,(int,float)) and v != 0:
                    c5.number_format="#,##0.00"
            c5.alignment=_aln("right" if isinstance(v,(int,float)) else "left")
            c5.border=_bdr()
        ws5.row_dimensions[ri5].height=16; ri5+=1

    # -- Annual total row --------------------------------------
    ann5_tot_out = round(ann5["oig"]+ann5["ocg"]+ann5["osg"],2)
    ann5_net_itc = round(ann5["iig"]+ann5["icg"]+ann5["isg"],2)
    totrow(ws5, ri5, ["ANNUAL TOTAL","12 Months",
                       _fsum("C",3,ri5-1), _fsum("D",3,ri5-1), _fsum("E",3,ri5-1),
                       _fsum("F",3,ri5-1), _fsum("G",3,ri5-1),
                       _fsum("H",3,ri5-1), _fsum("I",3,ri5-1), _fsum("J",3,ri5-1),
                       f"=SUM(H{ri5}:J{ri5})",
                       _fsum("L",3,ri5-1), _fsum("M",3,ri5-1), round(ann5["isg"],2),
                       round(ann5["rev"],2),
                       round(ann5["nicg"],2), round(ann5["nisg"],2),
                       round(ann5["int_cg"],2), round(ann5["int_sg"],2),
                       round(ann5["lf_cg"],2), round(ann5["lf_sg"],2),
                       round(ann5["net"],2), "Annual"])
    ws5.sheet_properties.tabColor = "9C0006"

    # ======================================================
    # SHEET: GSTR-3B EXTRACTED DETAIL — All fields per month
    # Shows every extracted value from each 3B PDF in a table
    # ======================================================
    ws_3bd = wb.create_sheet("GSTR3B_Extracted_Detail")
    ws_3bd.sheet_view.showGridLines = False

    # Define all columns with labels
    cols_3bd = [
        ("Month",          14), ("Period",         12), ("FY",             9),
        ("GSTIN",          22), ("Legal Name",      28), ("Trade Name",     26),
        ("ARN",            22), ("ARN Date",        13),
        # 3.1 Outward
        ("3.1(a) Taxable ₹",  18), ("3.1(a) IGST ₹", 14),
        ("3.1(a) CGST ₹",     14), ("3.1(a) SGST ₹", 14), ("3.1(a) Cess ₹", 12),
        ("3.1(b) Zero Rated ₹",16), ("3.1(c) Nil/Exempt ₹",18),
        ("3.1(d) RCM Taxable ₹",16),("3.1(e) Non-GST ₹",14),
        # 4 ITC
        ("4A(1) Imp.Goods ₹",  15), ("4A(2) Imp.Svc ₹",  15),
        ("4A(3) RCM ITC ₹",    14), ("4A(4) ISD ITC ₹",  14),
        ("4A(5) Other IGST ₹", 15), ("4A(5) Other CGST ₹",15),
        ("4A(5) Other SGST ₹", 15),
        ("4B Rev IGST ₹",      14), ("4B Rev CGST ₹",     14), ("4B Rev SGST ₹", 14),
        ("4C Net IGST ₹",      14), ("4C Net CGST ₹",     14), ("4C Net SGST ₹", 14),
        # 5.1
        ("Int IGST ₹",    12), ("Int CGST ₹",     12), ("Int SGST ₹",     12),
        ("Late CGST ₹",   12), ("Late SGST ₹",    12),
        # 6.1
        ("Paid IGST ₹",   14), ("Paid CGST ₹",    14), ("Paid SGST ₹",    14),
        # Computed
        ("Total Output ₹",16), ("Total ITC ₹",    14), ("Net Payable ₹",  16),
        ("Status",         12),
    ]

    title(ws_3bd,
          f"GSTR-3B Extracted Detail — {client_name} ({gstin}) — FY {FY_LABEL}",
          len(cols_3bd))
    ws_3bd.merge_cells(f"A2:{get_column_letter(len(cols_3bd))}2")
    s2=ws_3bd["A2"]
    s2.value = ("All values extracted directly from downloaded GSTR-3B PDFs | "
                "3.1=Outward Supplies | 4=ITC | 5.1=Interest/LateFee | 6.1=Cash Paid | "
                "Green=Extracted OK | Red=PDF Missing")
    s2.font=_font(False,"000000",8); s2.fill=_f("FFF2CC"); s2.alignment=_aln("left")
    ws_3bd.row_dimensions[2].height=13

    # Header row 3
    for ci,(h,w) in enumerate(cols_3bd, 1):
        hc = ws_3bd.cell(row=3, column=ci, value=h)
        hc.font = _font(True,"FFFFFF",8)
        hc.fill = _f("1F3864" if ci <= 8 else
                      "2E75B6" if ci <= 17 else
                      "375623" if ci <= 29 else
                      "843C0C" if ci <= 34 else
                      "9C0006" if ci <= 37 else HDR_BG)
        hc.alignment = _aln("center"); hc.border = _bdr()
        ws_3bd.column_dimensions[get_column_letter(ci)].width = w
    ws_3bd.row_dimensions[3].height = 20
    ws_3bd.freeze_panes = "A4"

    # Re-extract all PDFs to populate this sheet
    r_3bd = 4
    ann_3bd = {"taxable":0.,"o_igst":0.,"o_cgst":0.,"o_sgst":0.,
               "itc_igst":0.,"itc_cgst":0.,"itc_sgst":0.,"paid_cgst":0.,"paid_sgst":0.}

    for mn,_,yr in MONTHS:
        mk = f"{mn}_{yr}"
        pdf_p = Path(client_dir)/f"GSTR3B_{mn}_{yr}.pdf"
        d3b = extract_3b_pdf(pdf_p)
        ok = pdf_p.exists()
        bgr = ALT2 if r_3bd%2==0 else ALT1
        status_v = "✓ OK" if ok else "✗ Missing"
        sbg = GREEN_BG if ok else RED_BG
        sfg = GREEN_FG if ok else RED_FG

        tot_out = round(d3b["o_igst"]+d3b["o_cgst"]+d3b["o_sgst"],2)
        tot_itc = round(d3b["net_itc_igst"]+d3b["net_itc_cgst"]+d3b["net_itc_sgst"],2)
        net_pay = round(tot_out - tot_itc, 2)

        row_vals = [
            f"{mn} {yr}", d3b.get("period",mn), d3b.get("year",FY_LABEL),
            d3b.get("gstin",gstin), d3b.get("legal_name",client_name),
            d3b.get("trade_name",""),
            d3b.get("arn",""), d3b.get("arn_date",""),
            d3b["taxable"], d3b["o_igst"], d3b["o_cgst"], d3b["o_sgst"], d3b["o_cess"],
            d3b["zero_taxable"], d3b["nil_exempt"],
            d3b["rcm_taxable"], d3b["non_gst"],
            d3b["itc_import_goods"], d3b["itc_import_svc"],
            d3b["itc_rcm"], d3b["itc_isd"],
            d3b["itc_igst"], d3b["itc_cgst"], d3b["itc_sgst"],
            d3b["rev_igst"], d3b["rev_cgst"], d3b["rev_sgst"],
            d3b["net_itc_igst"], d3b["net_itc_cgst"], d3b["net_itc_sgst"],
            d3b["interest_igst"], d3b["interest_cgst"], d3b["interest_sgst"],
            d3b["late_fee_cgst"], d3b["late_fee_sgst"],
            d3b["tax_paid_igst"], d3b["tax_paid_cgst"], d3b["tax_paid_sgst"],
            tot_out, tot_itc, net_pay, status_v,
        ]
        for ci, v in enumerate(row_vals, 1):
            cv = ws_3bd.cell(row=r_3bd, column=ci, value=v)
            if ci == len(cols_3bd):           # Status
                cv.font = _font(True, sfg, 9); cv.fill = _f(sbg)
            elif ci <= 8:                      # Meta cols — grey
                cv.font = _font(False,"000000",9); cv.fill = _f(bgr)
            elif isinstance(v, float) and v != 0.:
                cv.font = _font(False,"000000",9); cv.fill = _f(bgr)
                cv.number_format = NUM_FMT
            elif isinstance(v, float):        # zero value
                cv.font = _font(False,"AAAAAA",9); cv.fill = _f(bgr)
                cv.number_format = NUM_FMT
            else:
                cv.font = _font(False,"000000",9); cv.fill = _f(bgr)
            cv.alignment = _aln("right" if isinstance(v,float) else "left")
            cv.border = _bdr()
        ws_3bd.row_dimensions[r_3bd].height = 16; r_3bd += 1

        ann_3bd["taxable"]  += d3b["taxable"]
        ann_3bd["o_igst"]   += d3b["o_igst"]
        ann_3bd["o_cgst"]   += d3b["o_cgst"]
        ann_3bd["o_sgst"]   += d3b["o_sgst"]
        ann_3bd["itc_igst"] += d3b["net_itc_igst"]
        ann_3bd["itc_cgst"] += d3b["net_itc_cgst"]
        ann_3bd["itc_sgst"] += d3b["net_itc_sgst"]
        ann_3bd["paid_cgst"]+= d3b["tax_paid_cgst"]
        ann_3bd["paid_sgst"]+= d3b["tax_paid_sgst"]

    # Annual total row
    ann_out=round(ann_3bd["o_igst"]+ann_3bd["o_cgst"]+ann_3bd["o_sgst"],2)
    ann_itc=round(ann_3bd["itc_igst"]+ann_3bd["itc_cgst"]+ann_3bd["itc_sgst"],2)
    ann_net=round(ann_out-ann_itc,2)
    ann_vals = [
        "APR-MAR TOTAL","","",gstin,client_name,"","","",
        round(ann_3bd["taxable"],2),round(ann_3bd["o_igst"],2),
        round(ann_3bd["o_cgst"],2),round(ann_3bd["o_sgst"],2),0,
        0,0,0,0, 0,0,0,0,
        round(ann_3bd["itc_igst"],2),round(ann_3bd["itc_cgst"],2),round(ann_3bd["itc_sgst"],2),
        0,0,0,
        round(ann_3bd["itc_igst"],2),round(ann_3bd["itc_cgst"],2),round(ann_3bd["itc_sgst"],2),
        0,0,0,0,0,
        0,round(ann_3bd["paid_cgst"],2),round(ann_3bd["paid_sgst"],2),
        ann_out,ann_itc,ann_net,"Annual"
    ]
    for ci,v in enumerate(ann_vals,1):
        cv=ws_3bd.cell(row=r_3bd,column=ci,value=v)
        cv.font=_font(True,"FFFFFF",9 if ci>3 else 10)
        cv.fill=_f("C00000" if ci>37 else "1F3864")
        cv.alignment=_aln("right" if isinstance(v,float) else "left"); cv.border=_bdr()
        if isinstance(v,float): cv.number_format=NUM_FMT
    ws_3bd.row_dimensions[r_3bd].height=20
    ws_3bd.sheet_properties.tabColor="9C0006"

    # (g3b_monthly already extracted above)

    # ====================================================
    # SHEET 6: GSTR-1 vs GSTR-3B Reconciliation
    # ====================================================
    ws6 = wb.create_sheet("R1_vs_3B_Recon")
    ws6.sheet_view.showGridLines = False
    cols6=[("Month",13),
           ("R1 Taxable ₹",16),("R1 IGST ₹",13),("R1 CGST ₹",13),("R1 SGST ₹",13),("R1 Total Tax ₹",15),
           ("2B ITC IGST ₹",14),("2B ITC CGST ₹",14),("2B ITC SGST ₹",14),("2B Total ITC ₹",15),
           ("Net Tax Payable ₹",17),("2A vs 2B Diff ₹",16),("Status",14)]
    title(ws6, f"R1 vs 3B Reconciliation — {client_name} ({gstin}) — FY {FY_LABEL}", len(cols6))
    ws6.merge_cells(f"A2:{get_column_letter(len(cols6))}2")
    sh6=ws6["A2"]
    sh6.value=("Output Tax from GSTR-1 vs ITC from GSTR-2B. "
               "Net Tax = R1 Tax – 2B ITC. 2A vs 2B diff shows unclaimed credit. "
               "Green = OK  |  Red = Action needed")
    sh6.font=Font(name="Arial",size=8,italic=True)
    sh6.fill=_f("FFF2CC"); sh6.alignment=_aln("left"); sh6.border=_bdr()
    ws6.row_dimensions[2].height=14
    hdr(ws6, cols6, row=3)
    ws6.freeze_panes = "A4"
    ri6=4
    ann6={"tx":0.,"r1ig":0.,"r1cg":0.,"r1sg":0.,"iig":0.,"icg":0.,"isg":0.,"a2":0.}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; d=g1[mk]
        r1tx=round(d["b2b_tx"]+d["b2cs_tx"],2)
        r1ig=round(d["igst"],2); r1cg=round(d["cgst"],2); r1sg=round(d["sgst"],2)
        r1tt=round(r1ig+r1cg+r1sg,2)
        iig=round(g2b[mk]["itc"]["igst"],2); icg=round(g2b[mk]["itc"]["cgst"],2)
        isg=round(g2b[mk]["itc"]["sgst"],2); itt=round(iig+icg+isg,2)
        net=round(r1tt-itt,2)
        a2t=round(g2a[mk]["tot"]["igst"]+g2a[mk]["tot"]["cgst"]+g2a[mk]["tot"]["sgst"],2)
        diff=round(a2t-itt,2)
        status=("✓ OK" if abs(diff)<=100 and net>=0
                else ("⚠ Excess ITC" if diff<-100 else ("⚠ Unclaimed ITC" if diff>100 else "⚠ Check")))
        rbg="FFC7CE" if "⚠" in status else ("C6EFCE" if status=="✓ OK" else "FFEB9C")
        rfg="9C0006" if "⚠" in status else ("276221" if status=="✓ OK" else "9C6500")
        bg6=ALT2 if ri6%2==0 else ALT1
        vals6=[f"{mn} {yr}",r1tx,r1ig,r1cg,r1sg,r1tt,iig,icg,isg,itt,net,diff,status]
        for ci6,v in enumerate(vals6,1):
            c6=ws6.cell(row=ri6,column=ci6,value=v)
            if ci6==13:
                c6.font=_font(True,rfg,9); c6.fill=_f(rbg)
            else:
                c6.font=_font(False,"000000",9); c6.fill=_f(bg6)
            c6.alignment=_aln("right" if isinstance(v,(int,float)) else "left")
            c6.border=_bdr()
            if isinstance(v,(int,float)): c6.number_format=NUM_FMT
        ws6.row_dimensions[ri6].height=15; ri6+=1
        ann6["tx"]+=r1tx; ann6["r1ig"]+=r1ig; ann6["r1cg"]+=r1cg; ann6["r1sg"]+=r1sg
        ann6["iig"]+=iig; ann6["icg"]+=icg; ann6["isg"]+=isg; ann6["a2"]+=a2t
    ann_r1tt=round(ann6["r1ig"]+ann6["r1cg"]+ann6["r1sg"],2)
    ann_itt=round(ann6["iig"]+ann6["icg"]+ann6["isg"],2)
    ann_diff=round(ann6["a2"]-ann_itt,2)
    totrow(ws6,ri6,["ANNUAL TOTAL",_fsum("B",3,ri6-1),_fsum("C",3,ri6-1),_fsum("D",3,ri6-1),
                    round(ann6["r1sg"],2),ann_r1tt,round(ann6["iig"],2),round(ann6["icg"],2),
                    round(ann6["isg"],2),ann_itt,round(ann_r1tt-ann_itt,2),ann_diff,
                    "✓ Balanced" if abs(ann_diff)<=500 else "⚠ Review"])
    ws6.sheet_properties.tabColor = "843C0C"

    # ====================================================
    # SHEET 7: Tax Liabilities & ITC Comparison (portal download)
    # Services → Returns → Tax liabilities and ITC comparison
    # Select FY 2024-25 → SEARCH → DOWNLOAD COMPARISON REPORTS (EXCEL)
    # ====================================================
    ws7 = wb.create_sheet("Tax_Liability_Compare")
    ws7.sheet_view.showGridLines = False
    ws7.column_dimensions["A"].width = 50
    ws7.column_dimensions["B"].width = 30

    # Find the downloaded file
    tl_file = (next(Path(client_dir).glob("TaxLiability_Comparison_FY*.xlsx"), None) or
               next(Path(client_dir).glob("TaxLiability_Comparison_FY*.xls"),  None) or
               next(Path(client_dir).glob("TaxLiability_FY*.xlsx"), None))
    tl_csv  = next(Path(client_dir).glob("TaxLiability_FY*.csv"),  None) if not tl_file else None

    # Title
    ws7.merge_cells("A1:B1")
    t7=ws7["A1"]
    t7.value = f"Tax Liabilities & ITC Comparison — {client_name} ({gstin}) — FY {FY_LABEL}"
    t7.font=Font(name="Arial",bold=True,color=HDR_FG,size=12)
    t7.fill=_f("7030A0"); t7.alignment=_aln("center"); t7.border=_bdr()
    ws7.row_dimensions[1].height=28

    # Sub-header explaining portal flow
    ws7.merge_cells("A2:B2")
    sub=ws7["A2"]
    sub.value=("Source: Services → Returns → Tax liabilities and ITC comparison → "
               f"Select FY {FY_LABEL} → SEARCH → DOWNLOAD COMPARISON REPORTS (EXCEL)")
    sub.font=Font(name="Arial",size=8,italic=True)
    sub.fill=_f("EAD1DC"); sub.alignment=_aln("left"); sub.border=_bdr()
    ws7.row_dimensions[2].height=14

    ri7=3
    if tl_file and tl_file.exists():
        try:
            xl7=pd.ExcelFile(tl_file)
            log.info(f"    Tax Liability file: {tl_file.name}, sheets: {xl7.sheet_names}")
            # Header row
            c=ws7.cell(row=ri7,column=1,value="Source File")
            c.font=_font(True,HDR_FG,9); c.fill=_f(SEC_BG); c.border=_bdr()
            c2=ws7.cell(row=ri7,column=2,value=tl_file.name)
            c2.font=_font(False,"000000",9); c2.fill=_f(ALT1); c2.border=_bdr()
            ri7+=1
            # Write all sheets from the downloaded Excel
            for sname in xl7.sheet_names:
                # Sheet sub-header
                ws7.merge_cells(f"A{ri7}:B{ri7}")
                sh=ws7.cell(row=ri7,column=1,value=f"— {sname} —")
                sh.font=_font(True,SEC_FG,9); sh.fill=_f(SEC_BG)
                sh.alignment=_aln("center"); sh.border=_bdr()
                ws7.row_dimensions[ri7].height=16; ri7+=1
                df7=pd.read_excel(tl_file,sheet_name=sname,header=None,dtype=str)
                for _,row7 in df7.iterrows():
                    vals=[str(v).strip() if pd.notna(v) and str(v).strip()!="nan" else "" for v in row7]
                    if not any(vals): continue   # skip blank rows
                    bg7=ALT2 if ri7%2==0 else ALT1
                    for ci7,rv in enumerate(vals[:2],1):
                        cv=ws7.cell(row=ri7,column=ci7,value=rv)
                        cv.font=_font(False,"000000",9); cv.fill=_f(bg7)
                        cv.alignment=_aln("left"); cv.border=_bdr()
                    ws7.row_dimensions[ri7].height=15; ri7+=1
            log.info(f"    Tax Liability sheet: {ri7-3} rows written")
        except Exception as e:
            log.warning(f"    Tax Liability file read error: {e}")
            ws7.merge_cells(f"A{ri7}:B{ri7}")
            c=ws7.cell(row=ri7,column=1,value=f"⚠ Error reading {tl_file.name}: {e}")
            c.font=_font(False,RED_FG,9); c.fill=_f(RED_BG); c.border=_bdr(); ri7+=1
    elif tl_csv and tl_csv.exists():
        try:
            df_csv=pd.read_csv(tl_csv,dtype=str)
            for ci7,col in enumerate(df_csv.columns[:2],1):
                c=ws7.cell(row=ri7,column=ci7,value=col)
                c.font=_font(True,HDR_FG,9); c.fill=_f(SEC_BG); c.border=_bdr()
            ri7+=1
            for _,row7 in df_csv.iterrows():
                bg7=ALT2 if ri7%2==0 else ALT1
                for ci7,v in enumerate(list(row7)[:2],1):
                    rv=str(v) if pd.notna(v) else ""
                    cv=ws7.cell(row=ri7,column=ci7,value=rv)
                    cv.font=_font(False,"000000",9); cv.fill=_f(bg7)
                    cv.alignment=_aln("left"); cv.border=_bdr()
                ri7+=1
        except Exception as e:
            log.warning(f"    CSV read error: {e}")
    else:
        # Not downloaded yet — show clear instructions
        ws7.merge_cells(f"A{ri7}:B{ri7}")
        c=ws7.cell(row=ri7,column=1,value="⚠ Tax Liability Comparison report NOT downloaded yet")
        c.font=_font(True,RED_FG,10); c.fill=_f(RED_BG); c.border=_bdr()
        ws7.row_dimensions[ri7].height=20; ri7+=1
        instructions=[
            "HOW TO DOWNLOAD MANUALLY:",
            "1. Login to GST portal → https://return.gst.gov.in",
            f"2. Click:  Services  →  Returns  →  Tax liabilities and ITC comparison",
            f"3. Select Financial Year:  {FY_LABEL}",
            "4. Click SEARCH button",
            "5. Scroll down → click  'DOWNLOAD COMPARISON REPORTS (EXCEL)'  button",
            f"6. Save the downloaded file as:  TaxLiability_Comparison_FY{FY_LABEL.replace('-','_')}.xlsx",
            f"7. Place it in:  {client_dir}",
            "8. Re-run the reconciliation report to populate this sheet",
            "",
            "OR: Run the script again and select option [10] Tax Liability to auto-download.",
        ]
        for inst in instructions:
            ws7.merge_cells(f"A{ri7}:B{ri7}")
            bg7=YELLOW_BG if inst.startswith(("HOW","OR:")) else ALT1
            bld=inst.startswith(("HOW","OR:"))
            c=ws7.cell(row=ri7,column=1,value=inst)
            c.font=_font(bld,"000000",9); c.fill=_f(bg7)
            c.alignment=_aln("left",wrap=True); c.border=_bdr()
            ws7.row_dimensions[ri7].height=16; ri7+=1
    ws7.sheet_properties.tabColor = "7030A0"

    # ====================================================
    # SHEET 8: GSTR-3B vs GSTR-1 Reconciliation
    # Matches GSTR3BR1_RECONCILED_Summary format exactly
    # ====================================================
    ws8 = wb.create_sheet("GSTR3B_vs_R1_Recon")
    ws8.sheet_view.showGridLines = False
    # Columns: Month | R1 Taxable | R1 IGST | R1 CGST | R1 SGST | R1 Total |
    #          3B Taxable | 3B IGST | 3B CGST | 3B SGST | 3B Total |
    #          Diff Taxable | Diff IGST | Diff CGST | Diff SGST | Diff Total | Status
    RCOLS8=[("Month",14),
            ("R1 Taxable ₹",16),("R1 IGST ₹",13),("R1 CGST ₹",13),("R1 SGST ₹",13),("R1 Total Tax ₹",16),
            ("3B Taxable ₹",14),("3B IGST ₹",13),("3B CGST ₹",13),("3B SGST ₹",13),("3B Total Tax ₹",16),
            ("Diff Taxable ₹",14),("Diff IGST ₹",13),("Diff CGST ₹",13),("Diff SGST ₹",13),("Diff Total ₹",14),
            ("Status",14)]
    title(ws8, f"GSTR-3B vs GSTR-1 Monthly Reconciliation — {client_name} ({gstin}) — FY {FY_LABEL}", len(RCOLS8))
    ws8.merge_cells(f"A2:{get_column_letter(len(RCOLS8))}2")
    sh8=ws8["A2"]
    sh8.value="R1=GSTR-1 Supply (B2B+B2CS) | 3B=GSTR-3B Output Tax (from PDF) | Diff=3B−R1 | Green=Match | Red=Mismatch (>₹100)"
    sh8.font=_font(False,"000000",8); sh8.fill=_f(YELLOW_BG); sh8.alignment=_aln("left")
    ws8.row_dimensions[2].height=14
    hdr(ws8, RCOLS8, row=3, bg=HDR_BG)
    ws8.freeze_panes = "A4"
    ri8=4

    # Ann accumulators for ALL R1 types + 3B
    ann8={"b2b_tx":0.,"b2cs_tx":0.,"b2cl_tx":0.,"nil_tx":0.,"exp_tx":0.,"cdn_cr":0.,"cdn_dr":0.,
          "r1ig":0.,"r1cg":0.,"r1sg":0.,
          "b3tx":0.,"b3ig":0.,"b3cg":0.,"b3sg":0.}

    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; d=g1[mk]; d3b=g3b_monthly.get(mk,{})
        # -- GSTR-1 breakdown by type --
        b2b_tx=round(d["b2b_tx"],2); b2cs_tx=round(d["b2cs_tx"],2)
        # Sum CDN credit/debit from totals
        cdn_cr=round(d.get("cdn_cr",0),2); cdn_dr=round(d.get("cdn_dr",0),2)
        r1_tx=round(b2b_tx+b2cs_tx,2)
        r1_ig=round(d["igst"],2); r1_cg=round(d["cgst"],2); r1_sg=round(d["sgst"],2)
        r1_tt=round(r1_ig+r1_cg+r1_sg,2)
        # Net after CDN
        net_r1_tx=round(r1_tx - cdn_cr + cdn_dr, 2)
        # -- GSTR-3B values from PDF --
        b3_tx=round(d3b.get("taxable",0),2)
        b3_ig=round(d3b.get("o_igst",0),2); b3_cg=round(d3b.get("o_cgst",0),2)
        b3_sg=round(d3b.get("o_sgst",0),2); b3_tt=round(b3_ig+b3_cg+b3_sg,2)
        # -- Differences --
        d_tx=round(b3_tx-r1_tx,2); d_ig=round(b3_ig-r1_ig,2)
        d_cg=round(b3_cg-r1_cg,2); d_sg=round(b3_sg-r1_sg,2); d_tt=round(d_ig+d_cg+d_sg,2)
        if abs(d_tt)<1:      status8="✓ Match";    sbg=GREEN_BG; sfg=GREEN_FG
        elif abs(d_tt)<1000: status8="⚠ Minor";    sbg="FFEB9C"; sfg="9C6500"
        else:                status8="✗ Mismatch"; sbg=RED_BG;   sfg=RED_FG

        bgr=ALT2 if ri8%2==0 else ALT1

        # -- Write month detail rows (all R1 types) --
        # Section header: Month name
        ws8.merge_cells(f"A{ri8}:{get_column_letter(len(RCOLS8))}{ri8}")
        mh=ws8.cell(row=ri8,column=1,value=f"{mn} {yr}")
        mh.font=_font(True,"FFFFFF",10); mh.fill=_f(SEC_BG)
        mh.alignment=_aln("left"); mh.border=_bdr(); ws8.row_dimensions[ri8].height=16; ri8+=1

        def wr8(lbl, r1_tx_v, r1_ig_v, r1_cg_v, r1_sg_v,
                b3_tx_v, b3_ig_v, b3_cg_v, b3_sg_v, bg=ALT1, bold=False):
            nonlocal ri8
            r1_tt2=round(r1_ig_v+r1_cg_v+r1_sg_v,2)
            b3_tt2=round(b3_ig_v+b3_cg_v+b3_sg_v,2)
            d_t2=round(r1_tx_v-b3_tx_v,2) if b3_tx_v else 0
            d_ig2=round(b3_ig_v-r1_ig_v,2); d_cg2=round(b3_cg_v-r1_cg_v,2)
            d_sg2=round(b3_sg_v-r1_sg_v,2); d_tt2=round(d_ig2+d_cg2+d_sg2,2)
            diff_bg=GREEN_BG if abs(d_tt2)<1 else (RED_BG if abs(d_tt2)>100 else "FFEB9C")
            row_vals=[lbl,
                      round(r1_tx_v,2),round(r1_ig_v,2),round(r1_cg_v,2),round(r1_sg_v,2),r1_tt2,
                      round(b3_tx_v,2),round(b3_ig_v,2),round(b3_cg_v,2),round(b3_sg_v,2),b3_tt2,
                      d_t2,d_ig2,d_cg2,d_sg2,d_tt2,""]
            for ci_w,v_w in enumerate(row_vals,1):
                cv=ws8.cell(row=ri8,column=ci_w,value=v_w)
                cv.font=_font(bold,"000000",9)
                if ci_w in (12,13,14,15,16): cv.fill=_f(diff_bg)
                else: cv.fill=_f(TOT_BG if bold else bg)
                cv.alignment=_aln("right" if ci_w>1 else "left"); cv.border=_bdr()
                if isinstance(v_w,(int,float)) and ci_w<17: cv.number_format=NUM_FMT
            ws8.row_dimensions[ri8].height=15; ri8+=1

        # GSTR-1 types in rows, 3B only on Total row
        wr8("3.1(a) B2B Supplies",     b2b_tx, round(r1_ig*b2b_tx/r1_tx if r1_tx else 0,2),
            round(r1_cg*b2b_tx/r1_tx if r1_tx else 0,2),
            round(r1_sg*b2b_tx/r1_tx if r1_tx else 0,2),
            0,0,0,0, bg="DEEAF1")
        wr8("3.1(a) B2CS Supplies",    b2cs_tx, 0,
            round(r1_cg*b2cs_tx/r1_tx if r1_tx else 0,2),
            round(r1_sg*b2cs_tx/r1_tx if r1_tx else 0,2),
            0,0,0,0, bg="E2EFDA")
        wr8("3.1(b) Zero Rated",       0,0,0,0, 0,0,0,0, bg=ALT2)
        wr8("3.1(c) Nil/Exempt",       0,0,0,0, 0,0,0,0, bg=ALT2)
        wr8("3.1(e) Non-GST",          0,0,0,0, 0,0,0,0, bg=ALT2)
        wr8("CDNR Credit",             -cdn_cr, 0,
            -round(r1_cg*cdn_cr/r1_tx if r1_tx else 0,2),
            -round(r1_sg*cdn_cr/r1_tx if r1_tx else 0,2),
            0,0,0,0, bg="FFF2CC")
        wr8("CDNR Debit",              cdn_dr, 0,
            round(r1_cg*cdn_dr/r1_tx if r1_tx else 0,2),
            round(r1_sg*cdn_dr/r1_tx if r1_tx else 0,2),
            0,0,0,0, bg="FFF2CC")
        # TOTAL row for this month (R1 vs 3B)
        row8=["TOTAL " + f"{mn[:3]} {yr}",
              r1_tx,r1_ig,r1_cg,r1_sg,r1_tt,
              b3_tx,b3_ig,b3_cg,b3_sg,b3_tt,
              d_tx,d_ig,d_cg,d_sg,d_tt,status8]
        for ci8,v in enumerate(row8,1):
            cv8=ws8.cell(row=ri8,column=ci8,value=v)
            if ci8==17: cv8.font=_font(True,sfg,9); cv8.fill=_f(sbg)
            elif ci8 in (12,13,14,15,16):
                cv8.font=_font(True,"000000",9)
                cv8.fill=_f(GREEN_BG if abs(float(v or 0))<1 else (RED_BG if abs(float(v or 0))>100 else "FFEB9C"))
            else: cv8.font=_font(True,"FFFFFF",9); cv8.fill=_f("1F3864")
            cv8.alignment=_aln("right" if ci8>1 and ci8<17 else "left"); cv8.border=_bdr()
            if isinstance(v,(int,float)) and ci8<17: cv8.number_format=NUM_FMT
        ws8.row_dimensions[ri8].height=18; ri8+=1; ri8+=1  # blank line after each month

        ann8["b2b_tx"]+=b2b_tx; ann8["b2cs_tx"]+=b2cs_tx
        ann8["cdn_cr"]+=cdn_cr; ann8["cdn_dr"]+=cdn_dr
        ann8["r1ig"]+=r1_ig; ann8["r1cg"]+=r1_cg; ann8["r1sg"]+=r1_sg
        ann8["b3tx"]+=b3_tx; ann8["b3ig"]+=b3_ig; ann8["b3cg"]+=b3_cg; ann8["b3sg"]+=b3_sg

    # -- APR-MAR ANNUAL TOTAL (cumulative all 12 months) ----------
    ann_r1tx=round(ann8["b2b_tx"]+ann8["b2cs_tx"],2)
    ann_r1tt=round(ann8["r1ig"]+ann8["r1cg"]+ann8["r1sg"],2)
    ann_b3tt=round(ann8["b3ig"]+ann8["b3cg"]+ann8["b3sg"],2)
    ann_d_tx=round(ann8["b3tx"]-ann_r1tx,2)
    ann_d_ig=round(ann8["b3ig"]-ann8["r1ig"],2)
    ann_d_cg=round(ann8["b3cg"]-ann8["r1cg"],2)
    ann_d_sg=round(ann8["b3sg"]-ann8["r1sg"],2)
    ann_d_tt=round(ann_b3tt-ann_r1tt,2)
    ann_status="✓ Balanced" if abs(ann_d_tt)<100 else "✗ Review"
    ann_sbg=GREEN_BG if abs(ann_d_tt)<100 else RED_BG
    ann_sfg=GREEN_FG if abs(ann_d_tt)<100 else RED_FG
    ann_row=["APR-MAR TOTAL (Full Year)",
             ann_r1tx,round(ann8["r1ig"],2),round(ann8["r1cg"],2),round(ann8["r1sg"],2),ann_r1tt,
             round(ann8["b3tx"],2),round(ann8["b3ig"],2),round(ann8["b3cg"],2),round(ann8["b3sg"],2),ann_b3tt,
             ann_d_tx,ann_d_ig,ann_d_cg,ann_d_sg,ann_d_tt,ann_status]
    for ci8,v in enumerate(ann_row,1):
        cv8=ws8.cell(row=ri8,column=ci8,value=v)
        if ci8==17: cv8.font=_font(True,ann_sfg,11); cv8.fill=_f(ann_sbg)
        elif ci8>11: cv8.font=_font(True,"000000",11); cv8.fill=_f(TOT_BG)
        else: cv8.font=_font(True,"FFFFFF",11); cv8.fill=_f("C00000")
        cv8.alignment=_aln("right" if ci8>1 and ci8<17 else "left"); cv8.border=_bdr()
        if isinstance(v,(int,float)) and ci8<17: cv8.number_format=NUM_FMT
    ws8.row_dimensions[ri8].height=22
    ws8.sheet_properties.tabColor = "1F3864"

    # ====================================================
    # SHEET 9: GSTR-3B vs GSTR-2A Reconciliation
    # Matches GSTR3BR2A_RECONCILED_Summary format exactly
    # ====================================================
    ws9 = wb.create_sheet("GSTR3B_vs_R2A_Recon")
    ws9.sheet_view.showGridLines = False
    RCOLS9=[("Section",24),("Particulars",44),
            ("IGST ₹",14),("CGST ₹",14),("SGST ₹",14),("Cess ₹",10),("Total ₹",16)]
    title(ws9, f"GSTR-3B vs GSTR-2A ITC Reconciliation — {client_name} ({gstin}) — FY {FY_LABEL}", len(RCOLS9))
    ws9.merge_cells(f"A2:{get_column_letter(len(RCOLS9))}2")
    sh9=ws9["A2"]
    sh9.value=f"GSTIN: {gstin}  |  Difference = GSTR-3B (A) minus GSTR-2A (B)  |  Negative = less ITC claimed vs available"
    sh9.font=_font(False,"000000",8); sh9.fill=_f(YELLOW_BG); sh9.alignment=_aln("left")
    ws9.row_dimensions[2].height=14
    hdr(ws9, RCOLS9, row=3, bg=HDR_BG)
    ws9.freeze_panes = "A4"
    ri9=4

    # Annual totals
    r2a_ig=r2a_cg=r2a_sg=0.0
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; tot=g2a[mk]["tot"]
        r2a_ig+=tot.get("igst",0); r2a_cg+=tot.get("cgst",0); r2a_sg+=tot.get("sgst",0)

    r2b_ig=r2b_cg=r2b_sg=0.0
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; itc=g2b[mk]["itc"]
        r2b_ig+=itc["igst"]; r2b_cg+=itc["cgst"]; r2b_sg+=itc["sgst"]

    r3b_itc_ig=r3b_itc_cg=r3b_itc_sg=0.0
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; d3b=g3b_monthly.get(mk,{})
        r3b_itc_ig+=d3b.get("itc_igst",0); r3b_itc_cg+=d3b.get("itc_cgst",0)
        r3b_itc_sg+=d3b.get("itc_sgst",0)

    def r9row(sect,part,ig,cg,sg,bg=ALT1,bold=False):
        nonlocal ri9
        tot=round(ig+cg+sg,2)
        for ci9,v in enumerate([sect,part,round(ig,2),round(cg,2),round(sg,2),0.0,tot],1):
            cell(ws9,ri9,ci9,v,bg,bold=bold,numfmt=NUM_FMT if ci9>2 else None,
                 align="right" if ci9>2 else "left")
        ws9.row_dimensions[ri9].height=16; ri9+=1
    def r9sec(label):
        nonlocal ri9
        ws9.merge_cells(f"A{ri9}:{get_column_letter(len(RCOLS9))}{ri9}")
        sc_=ws9.cell(row=ri9,column=1,value=label)
        sc_.font=_font(True,HDR_FG,9); sc_.fill=_f(SEC_BG)
        sc_.alignment=_aln("left"); sc_.border=_bdr()
        ws9.row_dimensions[ri9].height=16; ri9+=1

    r9sec("GSTR-3B ITC Details (Other than ISD)")
    r9row("GSTR-3B","4A(5) - All other ITC (A)",r3b_itc_ig,r3b_itc_cg,r3b_itc_sg,bg="E2EFDA")
    r9row("","4B - ITC Reversed",0,0,0,bg=ALT2)
    r9row("Total from GSTR-3B (A)","",r3b_itc_ig,r3b_itc_cg,r3b_itc_sg,bg=TOT_BG,bold=True)
    ri9+=1

    r9sec("GSTR-2A ITC Details (Other than ISD)")
    r9row("GSTR-2A","B2B",r2a_ig,r2a_cg,r2a_sg,bg="DEEAF1")
    r9row("","CDNR",0,0,0,bg=ALT2)
    r9row("","TDS",0,0,0,bg=ALT2)
    r9row("","TCS",0,0,0,bg=ALT2)
    r9row("Total from GSTR-2A (B)","",r2a_ig,r2a_cg,r2a_sg,bg=TOT_BG,bold=True)
    ri9+=1

    r9sec("Difference (A - B)")
    d9ig=round(r3b_itc_ig-r2a_ig,2); d9cg=round(r3b_itc_cg-r2a_cg,2); d9sg=round(r3b_itc_sg-r2a_sg,2)
    d9bg = GREEN_BG if abs(d9ig+d9cg+d9sg)<1 else RED_BG
    r9row("Difference (A - B)","Positive=excess ITC claimed vs available | Negative=ITC in 2A not claimed in 3B",
          d9ig,d9cg,d9sg,bg=d9bg,bold=True)
    ri9+=2

    r9sec("GSTR-2B vs GSTR-3B (Confirmed ITC)")
    r9row("GSTR-2B","B2B Confirmed ITC",r2b_ig,r2b_cg,r2b_sg,bg="FFF2CC")
    r9row("Total from GSTR-2B","",r2b_ig,r2b_cg,r2b_sg,bg=TOT_BG,bold=True)
    ri9+=1
    d2b_ig=round(r3b_itc_ig-r2b_ig,2); d2b_cg=round(r3b_itc_cg-r2b_cg,2); d2b_sg=round(r3b_itc_sg-r2b_sg,2)
    d2b_bg = GREEN_BG if abs(d2b_ig+d2b_cg+d2b_sg)<1 else RED_BG
    r9row("Difference (3B - 2B)","Positive=excess claimed | Negative=ITC available but not claimed",
          d2b_ig,d2b_cg,d2b_sg,bg=d2b_bg,bold=True)
    ws9.sheet_properties.tabColor = "9C0006"

    # ====================================================
    # SHEET 8: Month-Wise Reconciliation
    # ====================================================
    ws_mw = wb.create_sheet("Monthwise_Reconciliation")
    ws_mw.sheet_view.showGridLines = False

    cols8 = [
        ("Month",13),
        ("R1 Taxable ₹",16),("R1 IGST ₹",13),("R1 CGST ₹",13),("R1 SGST ₹",13),("R1 Tax Total ₹",16),
        ("3B Out IGST ₹",15),("3B Out CGST ₹",15),("3B Out SGST ₹",15),("3B Tax Total ₹",15),("R1 vs 3B ₹",15),
        ("2B ITC IGST ₹",14),("2B ITC CGST ₹",14),("2B ITC SGST ₹",14),("2B ITC Total ₹",16),
        ("2A ITC Total ₹",16),("2A vs 2B Diff ₹",16),
        ("CDN Credit TV ₹",16),("CDN Debit TV ₹",16),("CDN Tax Total ₹",15),
        ("Net Tax Payable ₹",18),("Status",16),
    ]
    title(ws8, f"Month-Wise Reconciliation — {client_name} ({gstin}) — FY {FY_LABEL}", len(cols8))
    ws_mw.merge_cells(f"A2:{get_column_letter(len(cols8))}2")
    sh8=ws_mw["A2"]
    sh8.value=("R1=GSTR-1  3B=GSTR-3B  2B=Confirmed ITC  2A=Auto-ITC  CDN=Credit/Debit Notes  "
               "Net=R1 Tax − 2B ITC  |  Green=OK  Yellow=Minor  Red=Review needed")
    sh8.font=Font(name="Arial",size=8,italic=True)
    sh8.fill=_f("FFF2CC"); sh8.alignment=_aln("left"); sh8.border=_bdr()
    ws_mw.row_dimensions[2].height=14
    hdr(ws8, cols8, row=3)
    ws_mw.freeze_panes="A4"
    ri8=4
    ann8={"r1tx":0.,"r1ig":0.,"r1cg":0.,"r1sg":0.,
          "b3ig":0.,"b3cg":0.,"b3sg":0.,
          "b2ig":0.,"b2cg":0.,"b2sg":0.,
          "a2t":0.,"cdn_cr":0.,"cdn_dr":0.,"cdn_tax":0.}

    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"
        d1=g1[mk]; d3b=g3b_monthly[mk]
        r1tx=round(d1["b2b_tx"]+d1["b2cs_tx"],2)
        r1ig=round(d1["igst"],2); r1cg=round(d1["cgst"],2); r1sg=round(d1["sgst"],2); r1tt=round(r1ig+r1cg+r1sg,2)
        b3ig=round(d3b["o_igst"],2); b3cg=round(d3b["o_cgst"],2); b3sg=round(d3b["o_sgst"],2); b3tt=round(b3ig+b3cg+b3sg,2)
        r1_3b=round(r1tt-b3tt,2)
        b2ig=round(g2b[mk]["itc"]["igst"],2); b2cg=round(g2b[mk]["itc"]["cgst"],2)
        b2sg=round(g2b[mk]["itc"]["sgst"],2); b2tt=round(b2ig+b2cg+b2sg,2)
        a2t=round(g2a[mk]["tot"]["igst"]+g2a[mk]["tot"]["cgst"]+g2a[mk]["tot"]["sgst"],2)
        a2_2b=round(a2t-b2tt,2)
        cdn_cr=round(d1.get("cdn_cr",0),2); cdn_dr=round(d1.get("cdn_dr",0),2)
        cdn_tax=round(d1.get("cdn_ig",0)+d1.get("cdn_cg",0)+d1.get("cdn_sg",0),2)
        net=round(r1tt-b2tt,2)
        # Status
        if abs(r1_3b)<=100 and abs(a2_2b)<=500:
            status8="✓ Balanced"; sbg="C6EFCE"; sfg="276221"
        elif abs(r1_3b)>1000 or abs(a2_2b)>5000:
            status8="⚠ Review"; sbg="FFC7CE"; sfg="9C0006"
        else:
            status8="⚠ Minor Diff"; sbg="FFEB9C"; sfg="9C6500"
        bg8=ALT2 if ri8%2==0 else ALT1
        vals8=[f"{mn} {yr}",r1tx,r1ig,r1cg,r1sg,r1tt,b3ig,b3cg,b3sg,b3tt,r1_3b,
               b2ig,b2cg,b2sg,b2tt,a2t,a2_2b,cdn_cr,cdn_dr,cdn_tax,net,status8]
        for ci8,v in enumerate(vals8,1):
            c8=ws_mw.cell(row=ri8,column=ci8,value=v)
            if ci8==22:  c8.font=_font(True,sfg,9);  c8.fill=_f(sbg)
            elif ci8==11:c8.fill=_f("FFC7CE" if abs(r1_3b)>100 else "C6EFCE"); c8.font=_font(False,"000000",9)
            elif ci8==17:c8.fill=_f("FFC7CE" if abs(a2_2b)>500 else "C6EFCE"); c8.font=_font(False,"000000",9)
            else:        c8.font=_font(False,"000000",9); c8.fill=_f(bg8)
            c8.alignment=_aln("right" if ci8>1 else "left"); c8.border=_bdr()
            if isinstance(v,(int,float)): c8.number_format=NUM_FMT
        ws_mw.row_dimensions[ri8].height=16; ri8+=1
        ann8["r1tx"]+=r1tx; ann8["r1ig"]+=r1ig; ann8["r1cg"]+=r1cg; ann8["r1sg"]+=r1sg
        ann8["b3ig"]+=b3ig; ann8["b3cg"]+=b3cg; ann8["b3sg"]+=b3sg
        ann8["b2ig"]+=b2ig; ann8["b2cg"]+=b2cg; ann8["b2sg"]+=b2sg
        ann8["a2t"]+=a2t; ann8["cdn_cr"]+=cdn_cr; ann8["cdn_dr"]+=cdn_dr; ann8["cdn_tax"]+=cdn_tax

    ann_r1tt=round(ann8["r1ig"]+ann8["r1cg"]+ann8["r1sg"],2)
    ann_b3tt=round(ann8["b3ig"]+ann8["b3cg"]+ann8["b3sg"],2)
    ann_b2tt=round(ann8["b2ig"]+ann8["b2cg"]+ann8["b2sg"],2)
    totrow(ws8,ri8,["ANNUAL TOTAL",_fsum("B",3,ri8-1),_fsum("C",3,ri8-1),_fsum("D",3,ri8-1),_fsum("E",3,ri8-1),f"=SUM(C{ri8}:E{ri8})",
                    round(ann8["b3ig"],2),round(ann8["b3cg"],2),round(ann8["b3sg"],2),ann_b3tt,round(ann_r1tt-ann_b3tt,2),
                    round(ann8["b2ig"],2),round(ann8["b2cg"],2),round(ann8["b2sg"],2),ann_b2tt,
                    round(ann8["a2t"],2),round(ann8["a2t"]-ann_b2tt,2),
                    round(ann8["cdn_cr"],2),round(ann8["cdn_dr"],2),round(ann8["cdn_tax"],2),
                    round(ann_r1tt-ann_b2tt,2),""])
    ws_mw.sheet_properties.tabColor="7030A0"

    # -----------------------------------------------------------------
    # SHEET 9 (Final): ANNUAL SUMMARY — one-page executive summary
    # -----------------------------------------------------------------
    ws_ann = wb.create_sheet("Annual_Summary")
    ws_ann.sheet_view.showGridLines = False
    ws_ann.column_dimensions["A"].width=40
    ws_ann.column_dimensions["B"].width=20
    ws_ann.column_dimensions["C"].width=20
    ws_ann.column_dimensions["D"].width=20

    ws_ann.merge_cells("A1:D1")
    ts=ws_ann["A1"]; ts.value=f"ANNUAL SUMMARY — {client_name} ({gstin}) — FY {FY_LABEL}"
    ts.font=Font(name="Arial",bold=True,color="FFFFFF",size=13)
    ts.fill=_f(HDR_BG); ts.alignment=_aln("center"); ts.border=_bdr()
    ws_ann.row_dimensions[1].height=30
    ws_ann.merge_cells("A2:D2")
    ts2=ws_ann["A2"]; ts2.value=f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}   |   Source: GSTR-1 JSON + GSTR-2B + GSTR-2A + GSTR-3B PDFs"
    ts2.font=Font(name="Arial",size=9,italic=True); ts2.fill=_f("D6DCE4")
    ts2.alignment=_aln("left"); ts2.border=_bdr(); ws_ann.row_dimensions[2].height=13

    def sum_header(r, label):
        ws_ann.merge_cells(f"A{r}:D{r}")
        c=ws_ann.cell(row=r,column=1,value=label)
        c.font=_font(True,"FFFFFF",10); c.fill=_f(SEC_BG)
        c.alignment=_aln("left"); c.border=_bdr(); ws_ann.row_dimensions[r].height=18

    def sum_row(r, label, igst, cgst, sgst, bold=False, bg=ALT1):
        total=round(igst+cgst+sgst,2)
        for ci,(v,al) in enumerate([(label,"left"),(round(igst,2),"right"),
                                     (round(cgst,2),"right"),(round(sgst,2),"right")],1):
            c=ws_ann.cell(row=r,column=ci,value=v)
            c.font=_font(bold,"000000",9); c.fill=_f(bg)
            c.alignment=_aln(al); c.border=_bdr()
            if ci>1 and isinstance(v,(int,float)): c.number_format=NUM_FMT
        ws_ann.row_dimensions[r].height=16

    def sum_total(r, label, igst, cgst, sgst, bg=TOT_BG):
        for ci,(v,al) in enumerate([(label,"left"),(round(igst,2),"right"),
                                     (round(cgst,2),"right"),(round(igst+cgst+sgst,2),"right")],1):
            c=ws_ann.cell(row=r,column=ci,value=v)
            c.font=_font(True,"000000",9); c.fill=_f(bg)
            c.alignment=_aln(al); c.border=_bdr()
            if ci>1 and isinstance(v,(int,float)): c.number_format=NUM_FMT
        ws_ann.row_dimensions[r].height=18

    # Column headers row 3
    for ci,(h,bg) in enumerate([("Particulars",HDR_BG),("IGST ₹",HDR_BG),
                                  ("CGST ₹",HDR_BG),("SGST / Total ₹",HDR_BG)],1):
        c=ws_ann.cell(row=3,column=ci,value=h)
        c.font=_font(True,"FFFFFF",9); c.fill=_f(bg)
        c.alignment=_aln("center"); c.border=_bdr()
    ws_ann.row_dimensions[3].height=20

    sr = 4  # start row
    # -- Annual totals from collected data --
    tot_r1ig=ann8["r1ig"]; tot_r1cg=ann8["r1cg"]; tot_r1sg=ann8["r1sg"]
    tot_b3ig=ann8["b3ig"]; tot_b3cg=ann8["b3cg"]; tot_b3sg=ann8["b3sg"]
    tot_b2ig=ann8["b2ig"]; tot_b2cg=ann8["b2cg"]; tot_b2sg=ann8["b2sg"]
    tot_a2  =ann8["a2t"]

    sum_header(sr,"1. OUTPUT TAX (GSTR-1 Reported)"); sr+=1
    sum_row(sr,"Total Sales (B2B + B2CS) Taxable Value",ann8["r1tx"],0,0,bg="E2EFDA"); sr+=1
    sum_row(sr,"Output IGST",tot_r1ig,0,0,bg=IGST_BG); sr+=1
    sum_row(sr,"Output CGST",0,tot_r1cg,0,bg=CGST_BG); sr+=1
    sum_row(sr,"Output SGST",0,0,tot_r1sg,bg=SGST_BG); sr+=1
    sum_row(sr,"Credit Notes (TV)",ann8["cdn_cr"],0,0,bg="E8F5E9"); sr+=1
    sum_row(sr,"Debit Notes (TV)", ann8["cdn_dr"],0,0,bg="FFF9C4"); sr+=1
    sum_total(sr,"TOTAL OUTPUT TAX",tot_r1ig,tot_r1cg,tot_r1sg); sr+=2

    sum_header(sr,"2. TAX FILED (GSTR-3B)"); sr+=1
    sum_row(sr,"Output IGST (filed)",tot_b3ig,0,0,bg=IGST_BG); sr+=1
    sum_row(sr,"Output CGST (filed)",0,tot_b3cg,0,bg=CGST_BG); sr+=1
    sum_row(sr,"Output SGST (filed)",0,0,tot_b3sg,bg=SGST_BG); sr+=1
    diff_r1_3b=round(tot_r1ig+tot_r1cg+tot_r1sg-tot_b3ig-tot_b3cg-tot_b3sg,2)
    bg_diff1="C6EFCE" if abs(diff_r1_3b)<100 else "FFC7CE"
    sum_row(sr,f"R1 vs 3B Difference (should be ≈0)",diff_r1_3b,0,0,bg=bg_diff1,bold=True); sr+=2

    sum_header(sr,"3. INPUT TAX CREDIT"); sr+=1
    sum_row(sr,"2B Confirmed ITC — IGST",tot_b2ig,0,0,bg=IGST_BG); sr+=1
    sum_row(sr,"2B Confirmed ITC — CGST",0,tot_b2cg,0,bg=CGST_BG); sr+=1
    sum_row(sr,"2B Confirmed ITC — SGST",0,0,tot_b2sg,bg=SGST_BG); sr+=1
    sum_row(sr,"2A Auto-drafted ITC",tot_a2,0,0,bg="FFF2CC"); sr+=1
    diff_2a_2b=round(tot_a2-tot_b2ig-tot_b2cg-tot_b2sg,2)
    bg_diff2="C6EFCE" if abs(diff_2a_2b)<500 else "FFC7CE"
    sum_row(sr,"2A vs 2B Difference",diff_2a_2b,0,0,bg=bg_diff2,bold=True); sr+=1
    sum_total(sr,"TOTAL 2B ITC",tot_b2ig,tot_b2cg,tot_b2sg); sr+=2

    sum_header(sr,"4. NET TAX PAYABLE"); sr+=1
    net_payable=round(tot_r1ig+tot_r1cg+tot_r1sg-tot_b2ig-tot_b2cg-tot_b2sg,2)
    bg_net="C6EFCE" if net_payable>=0 else "FFC7CE"
    sum_total(sr,"NET TAX (Output − 2B ITC)",
              round(tot_r1ig-tot_b2ig,2),round(tot_r1cg-tot_b2cg,2),round(tot_r1sg-tot_b2sg,2),
              bg=bg_net); sr+=2

    ws_ann.sheet_properties.tabColor="1F3864"

    # -- Save ----------------------------------------------
    safe = client_name.replace(" ","_").replace("/","_")[:20]
    rp = Path(client_dir)/f"ANNUAL_RECONCILIATION_{safe}_{FY_LABEL.replace('-','_')}.xlsx"

    # =======================================================
    # PER-MONTH SHEETS — one sheet per month with full detail
    # Sheet name: Apr-24, May-24, ... Mar-25
    # =======================================================
    MONTH_COLORS = ["1F3864","2E75B6","2F5496","17375E","243F60","17375E",
                    "375623","375623","4F6228","974706","843C0C","C00000"]

    for m_idx, (mn,_,yr) in enumerate(MONTHS):
        mk   = f"{mn}_{yr}"
        sname = f"{mn[:3]}-{yr[2:]}"   # Apr-24, May-24 ...
        wm   = wb.create_sheet(sname)
        wm.sheet_view.showGridLines = False
        mcol = MONTH_COLORS[m_idx % len(MONTH_COLORS)]

        # Title row
        wm.merge_cells("A1:K1")
        tm=wm["A1"]
        tm.value=f"{mn} {yr}  |  {client_name} ({gstin})  |  FY {FY_LABEL}"
        tm.font=Font(name="Arial",bold=True,color="FFFFFF",size=11)
        tm.fill=_f(mcol); tm.alignment=_aln("center"); tm.border=_bdr()
        wm.row_dimensions[1].height=26
        rm=2

        # -- GSTR-1 Summary ------------------------------
        wm.merge_cells(f"A{rm}:K{rm}")
        s=wm.cell(row=rm,column=1,value="GSTR-1 SALES SUMMARY")
        s.font=_font(True,"FFFFFF",9); s.fill=_f("2E75B6")
        s.alignment=_aln("left"); s.border=_bdr(); wm.row_dimensions[rm].height=14; rm+=1
        d1=g1[mk]
        nil_month = round(d1.get("nil_exempt",0),2)
        non_gst_month = round(d1.get("non_gst",0),2)
        for lbl,val in [("B2B Taxable ₹",d1["b2b_tx"]),("B2CS Taxable ₹",d1["b2cs_tx"]),
                         ("Nil/Exempt Taxable ₹",nil_month),("Non-GST Supplies ₹",non_gst_month),
                         ("Total Taxable ₹",round(d1["b2b_tx"]+d1["b2cs_tx"]+nil_month+non_gst_month,2)),
                         ("IGST ₹",d1["igst"]),("CGST ₹",d1["cgst"]),("SGST ₹",d1["sgst"]),
                         ("Total Tax ₹",round(d1["igst"]+d1["cgst"]+d1["sgst"],2)),
                         ("Invoice Count",d1["inv"])]:
            lc=wm.cell(row=rm,column=1,value=lbl)
            lc.font=_font(False,"000000",9); lc.fill=_f(ALT1); lc.border=_bdr()
            vc=wm.cell(row=rm,column=2,value=round(float(val),2) if isinstance(val,float) else val)
            vc.font=_font(True,"000000",9); vc.fill=_f(TOT_BG); vc.alignment=_aln("right"); vc.border=_bdr()
            if isinstance(val,float): vc.number_format=NUM_FMT
            rm+=1
        rm+=1

        # -- GSTR-1 Invoice Detail (B2B) -----------------
        wm.merge_cells(f"A{rm}:K{rm}")
        s=wm.cell(row=rm,column=1,value="GSTR-1 INVOICE DETAIL")
        s.font=_font(True,"FFFFFF",9); s.fill=_f("2E75B6")
        s.alignment=_aln("left"); s.border=_bdr(); wm.row_dimensions[rm].height=14; rm+=1
        inv_h=["Type","GSTIN Receiver","Receiver Name","Invoice No","Date","Value ₹","Rate%","Taxable ₹","IGST ₹","CGST ₹","SGST ₹"]
        inv_w=[7,22,26,14,12,14,6,14,12,12,12]
        for ci,(h,w) in enumerate(zip(inv_h,inv_w),1):
            hc=wm.cell(row=rm,column=ci,value=h)
            hc.font=_font(True,"FFFFFF",9); hc.fill=_f(HDR_BG)
            hc.alignment=_aln("center"); hc.border=_bdr()
            wm.column_dimensions[get_column_letter(ci)].width=w
        wm.row_dimensions[rm].height=16; rm+=1
        inv_rows_m = g1_inv_rows.get(mk,[])
        for row_d in inv_rows_m:
            inv_type,gstin_r,nm,inum,idt,iv,pos,rate,tv,ig,cg,sg = row_d
            bgm=ALT2 if rm%2==0 else ALT1
            for ci,v in enumerate([inv_type,gstin_r,nm,inum,idt,iv,rate,tv,ig,cg,sg],1):
                cell(wm,rm,ci,v,bgm,numfmt=NUM_FMT if ci in (6,8,9,10,11) else None,
                     align="right" if ci in (6,7,8,9,10,11) else "left")
            wm.row_dimensions[rm].height=14; rm+=1
        if not inv_rows_m:
            wm.cell(row=rm,column=1,value="No invoices").font=_font(False,"000000",9); rm+=1
        rm+=1

        # -- GSTR-2A Purchases ---------------------------
        wm.merge_cells(f"A{rm}:K{rm}")
        s=wm.cell(row=rm,column=1,value="GSTR-2A PURCHASES")
        s.font=_font(True,"FFFFFF",9); s.fill=_f("375623")
        s.alignment=_aln("left"); s.border=_bdr(); wm.row_dimensions[rm].height=14; rm+=1
        pur_h=["GSTIN Supplier","Supplier Name","Invoice No","Date","Taxable ₹","IGST ₹","CGST ₹","SGST ₹","Total ₹","",""]
        for ci,h in enumerate(pur_h[:9],1):
            hc=wm.cell(row=rm,column=ci,value=h)
            hc.font=_font(True,"FFFFFF",9); hc.fill=_f(HDR_BG)
            hc.alignment=_aln("center"); hc.border=_bdr()
        wm.row_dimensions[rm].height=16; rm+=1
        for row2a in g2a[mk]["rows"]:
            stin,nm,inum,idt,tv,ig,cg,sg = row2a
            bgm=ALT2 if rm%2==0 else ALT1
            for ci,v in enumerate([stin,nm,inum,idt,tv,ig,cg,sg,round(ig+cg+sg,2)],1):
                cell(wm,rm,ci,v,bgm,numfmt=NUM_FMT if ci>4 else None,align="right" if ci>4 else "left")
            wm.row_dimensions[rm].height=14; rm+=1
        if not g2a[mk]["rows"]:
            wm.cell(row=rm,column=1,value="No 2A data").font=_font(False,"000000",9); rm+=1
        rm+=1

        # -- GSTR-2B ITC Detail --------------------------
        wm.merge_cells(f"A{rm}:K{rm}")
        s=wm.cell(row=rm,column=1,value="GSTR-2B ITC DETAIL (Confirmed ITC)")
        s.font=_font(True,"FFFFFF",9); s.fill=_f("1F3864")
        s.alignment=_aln("left"); s.border=_bdr(); wm.row_dimensions[rm].height=14; rm+=1
        b2b_h=["GSTIN Supplier","Supplier Name","Invoice No","Date","Invoice Value ₹","Rate%","Taxable ₹","IGST ₹","CGST ₹","SGST ₹","ITC?"]
        for ci,h in enumerate(b2b_h,1):
            hc=wm.cell(row=rm,column=ci,value=h)
            hc.font=_font(True,"FFFFFF",9); hc.fill=_f(HDR_BG)
            hc.alignment=_aln("center"); hc.border=_bdr()
        wm.row_dimensions[rm].height=16; rm+=1
        b2b_rows_m=g2b[mk]["rows"]
        for row_d in b2b_rows_m:
            if len(row_d)==12:
                sup,nm,inum,idt,iv,pos,rate,tv,ig,cg,sg,itc_av=row_d
            else:
                sup,nm,inum,idt,tv,ig,cg,sg=row_d[:8]; iv=0; rate=0; itc_av="Yes"
            bgm=ALT2 if rm%2==0 else ALT1
            itc_bg="C6EFCE" if str(itc_av).lower() in ("yes","y","true","1") else "FFC7CE"
            for ci,v in enumerate([sup,nm,inum,idt,iv,rate,tv,ig,cg,sg,itc_av],1):
                cv=wm.cell(row=rm,column=ci,value=v)
                cv.font=_font(False,"000000",9)
                cv.fill=_f(itc_bg if ci==11 else bgm)
                cv.alignment=_aln("right" if ci in (5,6,7,8,9,10) else "left")
                cv.border=_bdr()
                if isinstance(v,float) and ci in (5,7,8,9,10): cv.number_format=NUM_FMT
            wm.row_dimensions[rm].height=14; rm+=1
        if not b2b_rows_m:
            wm.cell(row=rm,column=1,value="No 2B data this month").font=_font(False,"000000",9); rm+=1
        rm+=1

        # -- GSTR-3B Summary -----------------------------
        wm.merge_cells(f"A{rm}:K{rm}")
        s=wm.cell(row=rm,column=1,value="GSTR-3B SUMMARY (from PDF)")
        s.font=_font(True,"FFFFFF",9); s.fill=_f("9C0006")
        s.alignment=_aln("left"); s.border=_bdr(); wm.row_dimensions[rm].height=14; rm+=1
        d3b=g3b_monthly.get(mk,{})
        o_ig=d3b.get("o_igst",0); o_cg=d3b.get("o_cgst",0); o_sg=d3b.get("o_sgst",0)
        i_ig=d3b.get("itc_igst",0); i_cg=d3b.get("itc_cgst",0); i_sg=d3b.get("itc_sgst",0)
        for lbl,val in [("Output IGST ₹",o_ig),("Output CGST ₹",o_cg),("Output SGST ₹",o_sg),
                         ("Total Output Tax ₹",round(o_ig+o_cg+o_sg,2)),
                         ("ITC Total ₹",round(i_ig+i_cg+i_sg,2)),
                         ("Net Payable ₹",round(o_ig+o_cg+o_sg-i_ig-i_cg-i_sg,2))]:
            lc=wm.cell(row=rm,column=1,value=lbl)
            lc.font=_font(False,"000000",9); lc.fill=_f(ALT1); lc.border=_bdr()
            vc=wm.cell(row=rm,column=2,value=round(float(val),2))
            vc.font=_font(True,"000000",9); vc.fill=_f(TOT_BG); vc.alignment=_aln("right"); vc.border=_bdr()
            vc.number_format=NUM_FMT; rm+=1

        wm.freeze_panes="A2"
        wm.sheet_properties.tabColor=mcol

    # =======================================================
    # COMPANY OVERALL TOTAL SHEET
    # =======================================================
    ws_cot = wb.create_sheet("Company_Overall_Total")
    ws_cot.sheet_view.showGridLines = False
    for col,w in zip(["A","B","C","D","E","F"],[40,20,18,18,18,18]):
        ws_cot.column_dimensions[col].width=w
    ws_cot.merge_cells("A1:F1")
    tt=ws_cot["A1"]
    tt.value=f"COMPANY OVERALL TOTAL — {client_name} ({gstin}) — FY {FY_LABEL}"
    tt.font=Font(name="Arial",bold=True,color="FFFFFF",size=13)
    tt.fill=_f("1F3864"); tt.alignment=_aln("center"); tt.border=_bdr()
    ws_cot.row_dimensions[1].height=32
    rc=2

    def cot_sec(lbl):
        nonlocal rc
        ws_cot.merge_cells(f"A{rc}:F{rc}")
        s=ws_cot.cell(row=rc,column=1,value=lbl)
        s.font=_font(True,"FFFFFF",9); s.fill=_f("2E75B6")
        s.alignment=_aln("left"); s.border=_bdr(); ws_cot.row_dimensions[rc].height=16; rc+=1
    def cot_hdr(labels):
        nonlocal rc
        for ci,h in enumerate(labels,1):
            hc=ws_cot.cell(row=rc,column=ci,value=h)
            hc.font=_font(True,"FFFFFF",9); hc.fill=_f(HDR_BG)
            hc.alignment=_aln("center"); hc.border=_bdr()
        ws_cot.row_dimensions[rc].height=18; rc+=1
    def cot_row(lbl, vals, bold=False, bg=ALT1):
        nonlocal rc
        lc=ws_cot.cell(row=rc,column=1,value=lbl)
        lc.font=_font(bold,"000000",10); lc.fill=_f(bg); lc.border=_bdr(); lc.alignment=_aln("left")
        for ci,v in enumerate(vals,2):
            vc=ws_cot.cell(row=rc,column=ci,value=round(float(v),2) if isinstance(v,(int,float)) else v)
            vc.font=_font(bold,"000000",10); vc.fill=_f(TOT_BG if bold else bg)
            vc.alignment=_aln("right"); vc.border=_bdr()
            if isinstance(v,(int,float)): vc.number_format=NUM_FMT
        ws_cot.row_dimensions[rc].height=18; rc+=1

    # Compute annual totals
    t_b2b=t_b2cs=t_ig=t_cg=t_sg=t_inv=0.0
    t_3b_ig=t_3b_cg=t_3b_sg=t_itc_ig=t_itc_cg=t_itc_sg=0.0
    t_2b_ig=t_2b_cg=t_2b_sg=t_2a_ig=t_2a_cg=t_2a_sg=0.0
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; d=g1[mk]
        t_b2b+=d["b2b_tx"]; t_b2cs+=d["b2cs_tx"]
        t_ig+=d["igst"]; t_cg+=d["cgst"]; t_sg+=d["sgst"]; t_inv+=d["inv"]
        d3=g3b_monthly.get(mk,{})
        t_3b_ig+=d3.get("o_igst",0); t_3b_cg+=d3.get("o_cgst",0); t_3b_sg+=d3.get("o_sgst",0)
        t_itc_ig+=d3.get("itc_igst",0); t_itc_cg+=d3.get("itc_cgst",0); t_itc_sg+=d3.get("itc_sgst",0)
        t_2b_ig+=g2b[mk]["itc"]["igst"]; t_2b_cg+=g2b[mk]["itc"]["cgst"]; t_2b_sg+=g2b[mk]["itc"]["sgst"]
        t_2a_ig+=g2a[mk]["tot"].get("igst",0); t_2a_cg+=g2a[mk]["tot"].get("cgst",0); t_2a_sg+=g2a[mk]["tot"].get("sgst",0)

    t_tx=t_b2b+t_b2cs; t_tt=t_ig+t_cg+t_sg
    t_3b_tt=t_3b_ig+t_3b_cg+t_3b_sg; t_itc_tt=t_itc_ig+t_itc_cg+t_itc_sg
    t_2b_tt=t_2b_ig+t_2b_cg+t_2b_sg; t_2a_tt=t_2a_ig+t_2a_cg+t_2a_sg

    cot_sec("GSTR-1 ANNUAL SALES")
    cot_hdr(["Particulars","IGST ₹","CGST ₹","SGST ₹","Total Tax ₹","Taxable ₹"])
    cot_row("B2B Taxable",   [round(t_ig*t_b2b/(t_tx or 1),2),round(t_cg*t_b2b/(t_tx or 1),2),
                               round(t_sg*t_b2b/(t_tx or 1),2),round(t_tt*t_b2b/(t_tx or 1),2),t_b2b])
    cot_row("B2CS Taxable",  [0,round(t_cg*t_b2cs/(t_tx or 1),2),round(t_sg*t_b2cs/(t_tx or 1),2),
                               round(t_tt*t_b2cs/(t_tx or 1),2),t_b2cs])
    cot_row("GRAND TOTAL",   [t_ig,t_cg,t_sg,t_tt,t_tx],bold=True,bg=TOT_BG)
    cot_row(f"Total Invoices: {int(t_inv)}",[" "," "," "," "," "])
    rc+=1

    cot_sec("GSTR-3B ANNUAL SUMMARY (from PDF)")
    cot_hdr(["Particulars","IGST ₹","CGST ₹","SGST ₹","Total ₹",""])
    cot_row("Output Tax",    [t_3b_ig,t_3b_cg,t_3b_sg,t_3b_tt,""])
    cot_row("ITC Claimed",   [t_itc_ig,t_itc_cg,t_itc_sg,t_itc_tt,""])
    cot_row("Net Payable",   [t_3b_ig-t_itc_ig,t_3b_cg-t_itc_cg,
                               t_3b_sg-t_itc_sg,t_3b_tt-t_itc_tt,""],bold=True,bg=TOT_BG)
    rc+=1

    cot_sec("ITC ANNUAL SUMMARY")
    cot_hdr(["Particulars","IGST ₹","CGST ₹","SGST ₹","Total ₹",""])
    cot_row("GSTR-2B Confirmed ITC",[t_2b_ig,t_2b_cg,t_2b_sg,t_2b_tt,""])
    cot_row("GSTR-2A Available ITC",[t_2a_ig,t_2a_cg,t_2a_sg,t_2a_tt,""])
    cot_row("Difference (2A−2B)",   [t_2a_ig-t_2b_ig,t_2a_cg-t_2b_cg,
                                      t_2a_sg-t_2b_sg,t_2a_tt-t_2b_tt,""],bold=True,bg=TOT_BG)
    rc+=1

    cot_sec("RECONCILIATION")
    cot_hdr(["Check","IGST Diff","CGST Diff","SGST Diff","Total Diff","Status"])
    r1_3b_diff=t_tt-t_3b_tt; itc_diff=t_itc_tt-t_2a_tt
    cot_row("R1 Supply vs 3B Output",
            [round(t_ig-t_3b_ig,2),round(t_cg-t_3b_cg,2),round(t_sg-t_3b_sg,2),
             round(r1_3b_diff,2),"Balanced" if abs(r1_3b_diff)<100 else "Review"],
            bold=abs(r1_3b_diff)>100,
            bg="C6EFCE" if abs(r1_3b_diff)<100 else "FFC7CE")
    cot_row("3B ITC vs 2A ITC",
            [round(t_itc_ig-t_2a_ig,2),round(t_itc_cg-t_2a_cg,2),round(t_itc_sg-t_2a_sg,2),
             round(itc_diff,2),"Balanced" if abs(itc_diff)<500 else "Review"],
            bold=abs(itc_diff)>500,
            bg="C6EFCE" if abs(itc_diff)<500 else "FFC7CE")

    ws_cot.freeze_panes="A2"
    ws_cot.sheet_properties.tabColor="C00000"

    # ======================================================
    # GSTR-1 MONTHLY BREAKDOWN — per-return-type monthly sheet
    # Columns: Month | B2B Invoices | B2B Taxable | B2CS Taxable |
    #          IGST | CGST | SGST | Total Tax | CDN Credit | CDN Debit | Net Taxable | Tax% of Taxable
    # ======================================================
    ws_g1m = wb.create_sheet("GSTR1_Monthly_Breakdown")
    ws_g1m.sheet_view.showGridLines = False
    g1m_cols = [
        ("Month",12),("B2B Inv",8),("B2B Taxable ₹",16),("B2CS Taxable ₹",16),
        ("Total Taxable ₹",16),("IGST ₹",13),("CGST ₹",13),("SGST ₹",13),
        ("Total Tax ₹",14),("CDN Credit ₹",14),("CDN Debit ₹",13),
        ("Net Taxable ₹",16),("Tax Rate %",10),("Status",14)
    ]
    title(ws_g1m, f"GSTR-1 Monthly Breakdown — {client_name} ({gstin}) — FY {FY_LABEL}", len(g1m_cols))
    ws_g1m.merge_cells(f"A2:{get_column_letter(len(g1m_cols))}2")
    sub_g1m = ws_g1m["A2"]
    sub_g1m.value = "B2B=Business-to-Business | B2CS=Small Consumers | CDN=Credit/Debit Notes | Tax%=Effective Tax Rate on Taxable Supply"
    sub_g1m.font=_font(False,"000000",8); sub_g1m.fill=_f("FFF2CC"); sub_g1m.alignment=_aln("left")
    ws_g1m.row_dimensions[2].height=13
    hdr(ws_g1m, g1m_cols, row=3, bg=HDR_BG)
    ws_g1m.freeze_panes="A4"
    rg1m=4
    ann_g1m={"inv":0,"b2b":0.,"b2cs":0.,"ig":0.,"cg":0.,"sg":0.,
              "cdn_cr":0.,"cdn_dr":0.}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; d1=g1[mk]
        b2b_tx=round(d1["b2b_tx"],2); b2cs_tx=round(d1["b2cs_tx"],2)
        total_tx=round(b2b_tx+b2cs_tx,2)
        ig=round(d1["igst"],2); cg=round(d1["cgst"],2); sg=round(d1["sgst"],2)
        tt=round(ig+cg+sg,2)
        cdn_cr=round(d1.get("cdn_cr",0),2); cdn_dr=round(d1.get("cdn_dr",0),2)
        net_tx=round(total_tx-cdn_cr+cdn_dr,2)
        tax_pct=round(tt/total_tx*100,2) if total_tx else 0
        # Status: expected ~18% tax rate for most businesses
        if total_tx==0:   status="No Sales";  sbg="D9D9D9"
        elif tax_pct>=17: status="Normal";     sbg="C6EFCE"
        elif tax_pct>=10: status="Mixed Rate"; sbg="FFEB9C"
        else:             status="Low Rate";   sbg="FFC7CE"
        bgr=ALT2 if rg1m%2==0 else ALT1
        vals=[f"{mn} {yr}",d1["inv"],b2b_tx,b2cs_tx,total_tx,ig,cg,sg,tt,cdn_cr,cdn_dr,net_tx,tax_pct,status]
        for ci,v in enumerate(vals,1):
            cv=ws_g1m.cell(row=rg1m,column=ci,value=v)
            if ci==14: cv.font=_font(True,"000000",9); cv.fill=_f(sbg)
            elif ci==13: cv.font=_font(False,"000000",9); cv.fill=_f("DEEAF1")
            else: cv.font=_font(False,"000000",9); cv.fill=_f(bgr)
            cv.alignment=_aln("right" if ci>1 and ci<14 else "left")
            cv.border=_bdr()
            if isinstance(v,float) and ci not in (13,): cv.number_format=NUM_FMT
        ws_g1m.row_dimensions[rg1m].height=15; rg1m+=1
        ann_g1m["inv"]+=d1["inv"]; ann_g1m["b2b"]+=b2b_tx; ann_g1m["b2cs"]+=b2cs_tx
        ann_g1m["ig"]+=ig; ann_g1m["cg"]+=cg; ann_g1m["sg"]+=sg
        ann_g1m["cdn_cr"]+=cdn_cr; ann_g1m["cdn_dr"]+=cdn_dr
    ann_tx=ann_g1m["b2b"]+ann_g1m["b2cs"]; ann_tt=ann_g1m["ig"]+ann_g1m["cg"]+ann_g1m["sg"]
    ann_pct=round(ann_tt/ann_tx*100,2) if ann_tx else 0
    totrow(ws_g1m, rg1m, ["ANNUAL TOTAL", _fsum("B",3,rg1m-1),
                            round(ann_g1m["b2b"],2), round(ann_g1m["b2cs"],2),
                            round(ann_tx,2), round(ann_g1m["ig"],2),
                            round(ann_g1m["cg"],2), round(ann_g1m["sg"],2),
                            round(ann_tt,2), round(ann_g1m["cdn_cr"],2),
                            round(ann_g1m["cdn_dr"],2), round(ann_tx-ann_g1m["cdn_cr"]+ann_g1m["cdn_dr"],2),
                            ann_pct, "Annual"])
    ws_g1m.sheet_properties.tabColor="2E75B6"

    # ======================================================
    # GSTR-1A MONTHLY BREAKDOWN — Amendments per month
    # ======================================================
    ws_g1am = wb.create_sheet("GSTR1A_Monthly_Amendments")
    ws_g1am.sheet_view.showGridLines = False
    g1am_cols = [
        ("Month",12),("Amended Inv Count",16),("Amended B2B Taxable ₹",20),
        ("IGST ₹",13),("CGST ₹",13),("SGST ₹",13),("Total Tax ₹",14),
        ("Amendment Type",18),("Notes",30)
    ]
    title(ws_g1am, f"GSTR-1A Amendment Summary — {client_name} ({gstin}) — FY {FY_LABEL}", len(g1am_cols))
    ws_g1am.merge_cells(f"A2:{get_column_letter(len(g1am_cols))}2")
    sub_g1am=ws_g1am["A2"]
    sub_g1am.value="GSTR-1A = Amendments to previously filed GSTR-1 invoices (B2BA, CDNRA, B2CLA, EXPA etc.)"
    sub_g1am.font=_font(False,"000000",8); sub_g1am.fill=_f("FFF2CC"); sub_g1am.alignment=_aln("left")
    ws_g1am.row_dimensions[2].height=13
    hdr(ws_g1am, g1am_cols, row=3, bg=HDR_BG)
    ws_g1am.freeze_panes="A4"
    rg1am=4
    ann_g1am={"inv":0,"b2b":0.,"ig":0.,"cg":0.,"sg":0.}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; d1=g1[mk]
        # GSTR-1A amendments (b2ba, cdnra data if collected in g1_inv_rows)
        amend_rows=[r for r in g1_inv_rows.get(mk,[]) if r[0] in ("B2BA","CDNRA","B2CLA","EXPA")]
        a_inv=len(amend_rows)
        a_tx=round(sum(r[8] for r in amend_rows),2)
        a_ig=round(sum(r[9] for r in amend_rows),2)
        a_cg=round(sum(r[10] for r in amend_rows),2)
        a_sg=round(sum(r[11] for r in amend_rows),2)
        a_tt=round(a_ig+a_cg+a_sg,2)
        types_seen=list(set(r[0] for r in amend_rows)) if amend_rows else []
        note="Amended: "+", ".join(types_seen) if types_seen else "No amendments this month"
        bgr=ALT2 if rg1am%2==0 else ALT1
        hlt="FFEB9C" if a_inv>0 else bgr
        vals=[f"{mn} {yr}",a_inv,a_tx,a_ig,a_cg,a_sg,a_tt,",".join(types_seen) if types_seen else "-",note]
        for ci,v in enumerate(vals,1):
            cv=ws_g1am.cell(row=rg1am,column=ci,value=v)
            cv.font=_font(a_inv>0,"000000",9); cv.fill=_f(hlt)
            cv.alignment=_aln("right" if 2<ci<8 else "left"); cv.border=_bdr()
            if isinstance(v,float): cv.number_format=NUM_FMT
        ws_g1am.row_dimensions[rg1am].height=15; rg1am+=1
        ann_g1am["inv"]+=a_inv; ann_g1am["b2b"]+=a_tx
        ann_g1am["ig"]+=a_ig; ann_g1am["cg"]+=a_cg; ann_g1am["sg"]+=a_sg
    totrow(ws_g1am, rg1am, ["ANNUAL TOTAL", _fsum("B",3,rg1am-1),
                          _fsum("C",3,rg1am-1),_fsum("D",3,rg1am-1),
                          _fsum("E",3,rg1am-1),_fsum("F",3,rg1am-1),
                          _fsum("G",3,rg1am-1),_fsum("H",3,rg1am-1),
                          f"=SUM(D{rg1am}:H{rg1am})"])
    ws_g1am.sheet_properties.tabColor="9C6500"

    # ======================================================
    # COMPANY-WISE GSTR-1 SUMMARY — party-level aggregation
    # Shows: per-company annual taxable, IGST, CGST, SGST, invoice count, % of total
    # ======================================================
    ws_cw = wb.create_sheet("Company_Wise_Summary")
    ws_cw.sheet_view.showGridLines = False
    cw_cols=[
        ("GSTIN",22),("Company Name",32),("Invoice Count",13),
        ("Taxable Value ₹",18),("IGST ₹",13),("CGST ₹",13),("SGST ₹",13),
        ("Total Tax ₹",14),("Invoice Value ₹",17),
        ("% of Total Taxable",17),("% of Total Tax",16),("Avg Invoice ₹",16)
    ]
    title(ws_cw, f"Company-Wise Sales Summary — {client_name} ({gstin}) — FY {FY_LABEL}", len(cw_cols))
    ws_cw.merge_cells(f"A2:{get_column_letter(len(cw_cols))}2")
    sub_cw=ws_cw["A2"]
    sub_cw.value="Aggregated from all 12 months GSTR-1 B2B data | % shows contribution to total annual turnover"
    sub_cw.font=_font(False,"000000",8); sub_cw.fill=_f("FFF2CC"); sub_cw.alignment=_aln("left")
    ws_cw.row_dimensions[2].height=13
    hdr(ws_cw, cw_cols, row=3, bg=HDR_BG)
    ws_cw.freeze_panes="A4"
    rcw=4

    # Aggregate all B2B invoices by company across all months
    company_data={}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"
        for row_d in g1_inv_rows.get(mk,[]):
            if row_d[0] not in ("B2B","B2CL"): continue
            inv_type,gstin_r,nm,inum,idt,iv,pos,rate,tv,ig,cg,sg = row_d
            if not gstin_r or gstin_r=="-": continue
            key=gstin_r.strip()
            if key not in company_data:
                company_data[key]={"name":nm,"inv":0,"tv":0.,"ig":0.,"cg":0.,"sg":0.,"iv":0.}
            company_data[key]["name"]=nm or company_data[key]["name"]
            company_data[key]["inv"]+=1
            company_data[key]["tv"]+=float(tv or 0)
            company_data[key]["ig"]+=float(ig or 0)
            company_data[key]["cg"]+=float(cg or 0)
            company_data[key]["sg"]+=float(sg or 0)
            company_data[key]["iv"]+=float(iv or 0)

    # Sort by taxable value descending
    sorted_companies=sorted(company_data.items(), key=lambda x: x[1]["tv"], reverse=True)
    grand_tv=sum(d["tv"] for _,d in sorted_companies) or 1
    grand_tt=sum(d["ig"]+d["cg"]+d["sg"] for _,d in sorted_companies) or 1

    for gstin_r,(d) in sorted_companies:
        tv=round(d["tv"],2); ig=round(d["ig"],2); cg=round(d["cg"],2); sg=round(d["sg"],2)
        tt=round(ig+cg+sg,2); iv=round(d["iv"],2)
        pct_tv=round(d["tv"]/grand_tv*100,2)
        pct_tt=round((ig+cg+sg)/(grand_tt)*100,2) if grand_tt else 0
        avg_inv=round(iv/d["inv"],2) if d["inv"] else 0
        # Color top contributors
        if pct_tv>=10:   bgc="E2EFDA"
        elif pct_tv>=5:  bgc="EBF3FB"
        elif pct_tv>=1:  bgc=ALT1
        else:            bgc=ALT2
        vals=[gstin_r,d["name"],d["inv"],tv,ig,cg,sg,tt,iv,pct_tv,pct_tt,avg_inv]
        for ci,v in enumerate(vals,1):
            cv=ws_cw.cell(row=rcw,column=ci,value=v)
            cv.font=_font(pct_tv>=10,"000000",9); cv.fill=_f(bgc)
            cv.alignment=_aln("right" if ci>2 else "left"); cv.border=_bdr()
            if isinstance(v,float) and ci not in (10,11):
                cv.number_format=NUM_FMT
            elif ci in (10,11) and isinstance(v,float):
                cv.number_format="0.00%"
        ws_cw.row_dimensions[rcw].height=15; rcw+=1

    # Grand total
    grand_tv_r=sum(d["tv"] for _,d in sorted_companies)
    grand_ig=sum(d["ig"] for _,d in sorted_companies)
    grand_cg=sum(d["cg"] for _,d in sorted_companies)
    grand_sg=sum(d["sg"] for _,d in sorted_companies)
    grand_iv=sum(d["iv"] for _,d in sorted_companies)
    grand_inv=sum(d["inv"] for _,d in sorted_companies)
    totrow(ws_cw, rcw, ["GRAND TOTAL",f"All {len(sorted_companies)} companies",grand_inv,
                         round(grand_tv_r,2),round(grand_ig,2),round(grand_cg,2),
                         round(grand_sg,2),round(grand_ig+grand_cg+grand_sg,2),
                         round(grand_iv,2),100.0,100.0,
                         round(grand_iv/grand_inv,2) if grand_inv else 0])
    ws_cw.sheet_properties.tabColor="243F60"


    # ==================================================================
    # XLSM-STYLE EXTRA SHEETS
    # These replicate the analysis sheets from the reference XLSM file
    # ==================================================================

    # -- SHEET: Annual Report 3B (rate-wise breakdown) -----------------
    # Matches: ANNUAL REPORT 3B sheet in XLSM
    # Shows: Month | Rate-wise purchases (5%/12%/18%/28%) |
    #        Rate-wise sales | Tax amounts | ITC carried forward
    ws_ar3b = wb.create_sheet("Annual_Report_3B")
    ws_ar3b.sheet_view.showGridLines = False
    ar3b_cols = [
        ("Month",12),("Filing Date",14),
        ("Sales 5% ₹",13),("Sales 12% ₹",13),("Sales 18% ₹",14),("Sales 28% ₹",13),("Sales Total ₹",15),
        ("Purch 5% ₹",13),("Purch 12% ₹",13),("Purch 18% ₹",14),("Purch 28% ₹",13),("Purch Total ₹",15),
        ("Output CGST ₹",14),("Output SGST ₹",14),("Output IGST ₹",14),("Output Tax Total ₹",17),
        ("ITC CGST ₹",13),("ITC SGST ₹",13),("ITC IGST ₹",13),("ITC Total ₹",14),
        ("Net Payable ₹",15),("ITC Carried Fwd ₹",17),
    ]
    title(ws_ar3b, f"Annual Report 3B (Rate-wise) — {client_name} ({gstin}) — FY {FY_LABEL}", len(ar3b_cols))
    ws_ar3b.merge_cells(f"A2:{get_column_letter(len(ar3b_cols))}2")
    sub=ws_ar3b["A2"]
    sub.value="Rate-wise Sales & Purchase breakdown | Output Tax | ITC | Net Payable | Carried Forward — Source: GSTR-1 JSON + GSTR-3B PDF + GSTR-2B Excel"
    sub.font=_font(False,"000000",8); sub.fill=_f("FFF2CC"); sub.alignment=_aln("left")
    ws_ar3b.row_dimensions[2].height=13
    hdr(ws_ar3b, ar3b_cols, row=3, bg=HDR_BG)
    ws_ar3b.freeze_panes="A4"
    r_ar=4
    ann_ar={"s5":0.,"s12":0.,"s18":0.,"s28":0.,"p5":0.,"p12":0.,"p18":0.,"p28":0.,
            "ocg":0.,"osg":0.,"oig":0.,"icg":0.,"isg":0.,"iig":0.}
    itc_cf=0.0  # ITC carried forward running total
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; d1=g1[mk]; d3b=g3b_monthly.get(mk,{})
        # Sales by rate (from g1_inv_rows — group by rate)
        inv_rows_m=g1_inv_rows.get(mk,[])
        s_by_rate={5:0.,12:0.,18:0.,28:0.}
        for rd in inv_rows_m:
            rate_v=int(float(rd[7] or 0))
            tv_v=float(rd[8] or 0)
            if rate_v in s_by_rate: s_by_rate[rate_v]+=tv_v
            else: s_by_rate[18]+=tv_v  # default to 18
        s_tot=sum(s_by_rate.values())
        # Purchase by rate (from g2b rows)
        p_by_rate={5:0.,12:0.,18:0.,28:0.}
        for row_2b in g2b[mk]["rows"]:
            if len(row_2b)>=12:
                rate_2b=int(float(row_2b[6] or 0))
                tv_2b=float(row_2b[7] or 0)
                if rate_2b in p_by_rate: p_by_rate[rate_2b]+=tv_2b
                else: p_by_rate[18]+=tv_2b
        p_tot=sum(p_by_rate.values())
        # Output tax from 3B PDF
        ocg=round(d3b.get("o_cgst",0),2); osg=round(d3b.get("o_sgst",0),2); oig=round(d3b.get("o_igst",0),2)
        otot=round(ocg+osg+oig,2)
        # ITC
        icg=round(d3b.get("itc_cgst",0),2); isg=round(d3b.get("itc_sgst",0),2); iig=round(d3b.get("itc_igst",0),2)
        itot=round(icg+isg+iig,2)
        net=round(otot-itot,2)
        itc_cf=round(max(0,itot-otot)+itc_cf,2)
        bgr=ALT2 if r_ar%2==0 else ALT1
        vals=[f"{mn} {yr}","",
              round(s_by_rate[5],2),round(s_by_rate[12],2),round(s_by_rate[18],2),round(s_by_rate[28],2),round(s_tot,2),
              round(p_by_rate[5],2),round(p_by_rate[12],2),round(p_by_rate[18],2),round(p_by_rate[28],2),round(p_tot,2),
              ocg,osg,oig,otot,icg,isg,iig,itot,net,itc_cf]
        for ci,v in enumerate(vals,1):
            cv=ws_ar3b.cell(row=r_ar,column=ci,value=v)
            cv.font=_font(False,"000000",9); cv.fill=_f(bgr)
            cv.alignment=_aln("right" if ci>2 else "left"); cv.border=_bdr()
            if isinstance(v,float): cv.number_format=NUM_FMT
        ws_ar3b.row_dimensions[r_ar].height=15; r_ar+=1
        ann_ar["s5"]+=s_by_rate[5]; ann_ar["s12"]+=s_by_rate[12]
        ann_ar["s18"]+=s_by_rate[18]; ann_ar["s28"]+=s_by_rate[28]
        ann_ar["p5"]+=p_by_rate[5]; ann_ar["p12"]+=p_by_rate[12]
        ann_ar["p18"]+=p_by_rate[18]; ann_ar["p28"]+=p_by_rate[28]
        ann_ar["ocg"]+=ocg; ann_ar["osg"]+=osg; ann_ar["oig"]+=oig
        ann_ar["icg"]+=icg; ann_ar["isg"]+=isg; ann_ar["iig"]+=iig
    totrow(ws_ar3b,r_ar,["ANNUAL TOTAL","",
                          _fsum("C",3,r_ar-1),_fsum("D",3,r_ar-1),_fsum("E",3,r_ar-1),_fsum("F",3,r_ar-1),
                          _fsum("G",3,r_ar-1)])
    ws_ar3b.sheet_properties.tabColor="1F3864"

    # -- SHEET: GST vs CDNR Monthly -------------------------------------
    # Matches: GST vs CDNR All + CDNR_Monthly_Summary sheets in XLSM
    ws_cdnr = wb.create_sheet("GST_vs_CDNR_Monthly")
    ws_cdnr.sheet_view.showGridLines = False
    cdnr_cols=[
        ("Month",12),("GST Taxable ₹",16),("GST CGST ₹",13),("GST SGST ₹",13),("GST IGST ₹",13),
        ("CDNR Taxable ₹",16),("CDNR CGST ₹",13),("CDNR SGST ₹",13),
        ("Net Taxable ₹",16),("Net CGST ₹",13),("Net SGST ₹",13),("Net IGST ₹",13),("Net Tax ₹",14),
    ]
    title(ws_cdnr, f"GST vs CDNR Monthly — {client_name} ({gstin}) — FY {FY_LABEL}", len(cdnr_cols))
    ws_cdnr.merge_cells(f"A2:{get_column_letter(len(cdnr_cols))}2")
    sub2=ws_cdnr["A2"]
    sub2.value="GST Sales vs Credit/Debit Notes (CDNR) | Net = GST minus CDNR Credit plus CDNR Debit"
    sub2.font=_font(False,"000000",8); sub2.fill=_f("FFF2CC"); sub2.alignment=_aln("left")
    ws_cdnr.row_dimensions[2].height=13
    hdr(ws_cdnr, cdnr_cols, row=3, bg=HDR_BG)
    ws_cdnr.freeze_panes="A4"
    r_cd=4
    ann_cd={"gtx":0.,"gcg":0.,"gsg":0.,"gig":0.,
             "ctx":0.,"ccg":0.,"csg":0.}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; d1=g1[mk]
        gtx=round(d1["b2b_tx"]+d1["b2cs_tx"],2)
        gcg=round(d1["cgst"],2); gsg=round(d1["sgst"],2); gig=round(d1["igst"],2)
        ctx=round(d1.get("cdn_cr",0),2)
        ccg=round(d1.get("cdn_cg",0),2); csg=round(d1.get("cdn_sg",0),2)
        net_tx=round(gtx-ctx,2)
        net_cg=round(gcg-ccg,2); net_sg=round(gsg-csg,2); net_ig=round(gig,2)
        net_tax=round(net_cg+net_sg+net_ig,2)
        bgr=ALT2 if r_cd%2==0 else ALT1
        has_cdnr=ctx>0
        vals=[f"{mn} {yr}",gtx,gcg,gsg,gig,ctx,ccg,csg,net_tx,net_cg,net_sg,net_ig,net_tax]
        for ci,v in enumerate(vals,1):
            bgc="FFEB9C" if has_cdnr and ci>=6 else bgr
            cv=ws_cdnr.cell(row=r_cd,column=ci,value=v)
            cv.font=_font(has_cdnr and ci>=6,"000000",9); cv.fill=_f(bgc)
            cv.alignment=_aln("right" if ci>1 else "left"); cv.border=_bdr()
            if isinstance(v,float): cv.number_format=NUM_FMT
        ws_cdnr.row_dimensions[r_cd].height=15; r_cd+=1
        ann_cd["gtx"]+=gtx; ann_cd["gcg"]+=gcg; ann_cd["gsg"]+=gsg; ann_cd["gig"]+=gig
        ann_cd["ctx"]+=ctx; ann_cd["ccg"]+=ccg; ann_cd["csg"]+=csg
    ann_net_tx=ann_cd["gtx"]-ann_cd["ctx"]
    ann_net_cg=ann_cd["gcg"]-ann_cd["ccg"]; ann_net_sg=ann_cd["gsg"]-ann_cd["csg"]
    totrow(ws_cdnr,r_cd,["ANNUAL TOTAL",
                          _fsum("B",3,r_cd-1),_fsum("C",3,r_cd-1),_fsum("D",3,r_cd-1),_fsum("E",3,r_cd-1),
                          _fsum("F",3,r_cd-1),_fsum("G",3,r_cd-1),_fsum("H",3,r_cd-1)])
    ws_cdnr.sheet_properties.tabColor="9C0006"

    # -- SHEET: Purchase 2B Detail ---------------------------------------
    # Matches: PURCHASE 2B sheet in XLSM — full GSTR-2B invoice detail
    ws_p2b = wb.create_sheet("Purchase_2B_Detail")
    ws_p2b.sheet_view.showGridLines = False
    p2b_cols=[
        ("Month",12),("GSTIN Supplier",22),("Trade Name",28),("Inv No",14),
        ("Type",8),("Invoice Date",13),("Invoice Value ₹",16),("Rate %",8),
        ("Taxable Value ₹",16),("IGST ₹",12),("CGST ₹",12),("SGST ₹",12),
        ("ITC Avail?",11),
    ]
    title(ws_p2b, f"GSTR-2B Purchase Detail — {client_name} ({gstin}) — FY {FY_LABEL}", len(p2b_cols))
    hdr(ws_p2b, p2b_cols, row=2, bg=HDR_BG)
    ws_p2b.freeze_panes="A3"
    r_p2b=3
    ann_p2b={"tv":0.,"ig":0.,"cg":0.,"sg":0.}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"
        for row_d in g2b[mk]["rows"]:
            if len(row_d)>=12:
                sup,nm,inum,idt,iv,pos,rate,tv,ig,cg,sg,itc_av = row_d
            else:
                sup,nm,inum,idt,tv,ig,cg,sg = row_d[:8]
                iv=0; rate=0; itc_av="Yes"
            bgr=ALT2 if r_p2b%2==0 else ALT1
            bgc="C6EFCE" if str(itc_av).lower() in ("yes","y","true","1") else "FFC7CE"
            vals=[f"{mn} {yr}",sup,nm,inum,"R",idt,iv,rate,tv,ig,cg,sg,itc_av]
            for ci,v in enumerate(vals,1):
                cv=ws_p2b.cell(row=r_p2b,column=ci,value=v)
                cv.font=_font(False,"000000",9)
                cv.fill=_f(bgc if ci==13 else bgr)
                cv.alignment=_aln("right" if ci in (7,8,9,10,11,12) else "left"); cv.border=_bdr()
                if isinstance(v,float): cv.number_format=NUM_FMT
            ws_p2b.row_dimensions[r_p2b].height=14; r_p2b+=1
            ann_p2b["tv"]+=float(tv or 0); ann_p2b["ig"]+=float(ig or 0)
            ann_p2b["cg"]+=float(cg or 0); ann_p2b["sg"]+=float(sg or 0)
    totrow(ws_p2b,r_p2b,["ANNUAL TOTAL","","","","","","",
                          _fsum("H",3,r_p2b-1),_fsum("I",3,r_p2b-1),_fsum("J",3,r_p2b-1),_fsum("K",3,r_p2b-1),
                          _fsum("L",3,r_p2b-1)])
    ws_p2b.sheet_properties.tabColor="375623"

    # -- SHEET: GST Sales Detail (party-wise monthly) --------------------
    # Matches: GST_Monthwise_Summary_Sal sheet in XLSM
    ws_gsd = wb.create_sheet("GST_Sales_Detail")
    ws_gsd.sheet_view.showGridLines = False
    gsd_cols=[
        ("Month",12),("Rate %",8),("GSTIN",22),("Trade Name",30),
        ("Invoice Date",13),("Taxable ₹",16),("CGST ₹",12),("SGST ₹",12),("IGST ₹",12),("Total GST ₹",14),
    ]
    title(ws_gsd, f"GST Sales (Party-Wise Monthly) — {client_name} ({gstin}) — FY {FY_LABEL}", len(gsd_cols))
    hdr(ws_gsd, gsd_cols, row=2, bg=HDR_BG)
    ws_gsd.freeze_panes="A3"
    r_gsd=3
    ann_gsd={"tx":0.,"cg":0.,"sg":0.,"ig":0.}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"
        inv_rows_m=g1_inv_rows.get(mk,[])
        if inv_rows_m:
            ws_gsd.merge_cells(f"A{r_gsd}:{get_column_letter(len(gsd_cols))}{r_gsd}")
            sep=ws_gsd.cell(row=r_gsd,column=1,value=f"-- {mn} {yr}  ({len(inv_rows_m)} invoices) --")
            sep.font=_font(True,HDR_FG,9); sep.fill=_f(SEC_BG)
            sep.alignment=_aln("left"); sep.border=_bdr()
            ws_gsd.row_dimensions[r_gsd].height=14; r_gsd+=1
        m_tx=m_cg=m_sg=m_ig=0.
        for rd in inv_rows_m:
            inv_type,gstin_r,nm,inum,idt,iv,pos,rate,tv,ig,cg,sg = rd
            bgr=ALT2 if r_gsd%2==0 else ALT1
            tot_gst=round(float(ig or 0)+float(cg or 0)+float(sg or 0),2)
            vals=[f"{mn} {yr}",rate,gstin_r,nm,idt,tv,cg,sg,ig,tot_gst]
            for ci,v in enumerate(vals,1):
                cv=ws_gsd.cell(row=r_gsd,column=ci,value=v)
                cv.font=_font(False,"000000",9); cv.fill=_f(bgr)
                cv.alignment=_aln("right" if ci in (2,6,7,8,9,10) else "left"); cv.border=_bdr()
                if isinstance(v,float): cv.number_format=NUM_FMT
            ws_gsd.row_dimensions[r_gsd].height=14; r_gsd+=1
            m_tx+=float(tv or 0); m_cg+=float(cg or 0); m_sg+=float(sg or 0); m_ig+=float(ig or 0)
        if inv_rows_m:
            sub_r=ws_gsd.cell(row=r_gsd,column=1,value=f"Subtotal {mn} {yr}")
            sub_r.font=_font(True,"000000",9); sub_r.fill=_f(TOT_BG); sub_r.border=_bdr()
            for ci,v in enumerate(["","","","",round(m_tx,2),round(m_cg,2),round(m_sg,2),round(m_ig,2),round(m_cg+m_sg+m_ig,2)],2):
                sv=ws_gsd.cell(row=r_gsd,column=ci,value=v)
                sv.font=_font(True,"000000",9); sv.fill=_f(TOT_BG); sv.border=_bdr()
                sv.alignment=_aln("right")
                if isinstance(v,float): sv.number_format=NUM_FMT
            ws_gsd.row_dimensions[r_gsd].height=15; r_gsd+=1
            ann_gsd["tx"]+=m_tx; ann_gsd["cg"]+=m_cg; ann_gsd["sg"]+=m_sg; ann_gsd["ig"]+=m_ig
    totrow(ws_gsd,r_gsd,["ANNUAL TOTAL","","","","",
                          _fsum("F",3,r_gsd-1),_fsum("G",3,r_gsd-1),_fsum("H",3,r_gsd-1),_fsum("I",3,r_gsd-1)])
    ws_gsd.sheet_properties.tabColor="2E75B6"

    # -- SHEET: Turnover Match (AIS vs GST) -----------------------------
    # Matches: Turnover_Match_Summary sheet in XLSM
    ws_tm = wb.create_sheet("Turnover_AIS_Match")
    ws_tm.sheet_view.showGridLines = False
    tm_cols=[
        ("Month",14),("GST Turnover ₹",18),("R1 Taxable ₹",18),("Difference ₹",16),
        ("Match Status",14),("Variance %",12),("Notes",28),
    ]
    title(ws_tm, f"Turnover AIS vs GST Match — {client_name} ({gstin}) — FY {FY_LABEL}", len(tm_cols))
    ws_tm.merge_cells(f"A2:{get_column_letter(len(tm_cols))}2")
    sub_tm=ws_tm["A2"]
    sub_tm.value="Comparison: GSTR-1 Taxable Turnover vs AIS Reported Turnover | Match=within ₹1 | Mismatch=difference>₹1"
    sub_tm.font=_font(False,"000000",8); sub_tm.fill=_f("FFF2CC"); sub_tm.alignment=_aln("left")
    ws_tm.row_dimensions[2].height=13
    hdr(ws_tm, tm_cols, row=3, bg=HDR_BG)
    ws_tm.freeze_panes="A4"
    r_tm=4
    ann_tm={"r1":0.}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; d1=g1[mk]
        r1_tx=round(d1["b2b_tx"]+d1["b2cs_tx"],2)
        # AIS turnover not available from downloads — mark as To Be Verified
        diff=0.0; match_s="Verify vs AIS"; mbg="FFEB9C"
        var_pct=0.0
        note="Cross-check with AIS downloaded separately"
        bgr=ALT2 if r_tm%2==0 else ALT1
        vals=[f"{mn} {yr}",0,r1_tx,diff,match_s,var_pct,note]
        for ci,v in enumerate(vals,1):
            cv=ws_tm.cell(row=r_tm,column=ci,value=v)
            cv.font=_font(False,"000000",9)
            cv.fill=_f(mbg if ci==5 else bgr)
            cv.alignment=_aln("right" if ci in (2,3,4,6) else "left"); cv.border=_bdr()
            if isinstance(v,float) and ci not in (6,): cv.number_format=NUM_FMT
            elif ci==6: cv.number_format="0.00%"
        ws_tm.row_dimensions[r_tm].height=15; r_tm+=1
        ann_tm["r1"]+=r1_tx
    totrow(ws_tm,r_tm,["ANNUAL TOTAL",_fsum("B",3,r_tm-1),_fsum("C",3,r_tm-1),_fsum("D",3,r_tm-1),"Annual","","Full year GSTR-1 vs AIS"])
    ws_tm.sheet_properties.tabColor="974706"

    # ======================================================
    # COMPANY-WISE MONTHLY DETAIL SHEET
    # Each company as a row; 12 months + FY Total as columns
    # Values: Taxable, CGST, SGST, IGST, Total Tax per month
    # ======================================================
    ws_cwm = wb.create_sheet("Company_Month_Detail")
    ws_cwm.sheet_view.showGridLines = False
    month_short=[mn[:3]+"-"+yr[2:] for mn,_,yr in MONTHS]
    n_vals=5  # Taxable, IGST, CGST, SGST, Total Tax
    total_cols=3+len(MONTHS)*n_vals+n_vals  # GSTIN+Name+Type + 12months + FYtotal

    ws_cwm.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    tt=ws_cwm["A1"]
    tt.value=f"Company-Wise Monthly Detail — {client_name} ({gstin}) — FY {FY_LABEL}"
    tt.font=Font(name="Arial",bold=True,color="FFFFFF",size=12)
    tt.fill=_f("1F3864"); tt.alignment=_aln("center"); tt.border=_bdr()
    ws_cwm.row_dimensions[1].height=26
    # Row 2: month group headers
    col_s=4
    for ml in month_short:
        ws_cwm.merge_cells(start_row=2,start_column=col_s,end_row=2,end_column=col_s+n_vals-1)
        mh=ws_cwm.cell(row=2,column=col_s,value=ml)
        mh.font=_font(True,"FFFFFF",9); mh.fill=_f(HDR_BG)
        mh.alignment=_aln("center"); mh.border=_bdr()
        col_s+=n_vals
    ws_cwm.merge_cells(start_row=2,start_column=col_s,end_row=2,end_column=col_s+n_vals-1)
    ah=ws_cwm.cell(row=2,column=col_s,value="FY Total (APR-MAR)")
    ah.font=_font(True,"FFFFFF",9); ah.fill=_f("C00000")
    ah.alignment=_aln("center"); ah.border=_bdr()
    # Row 3: sub-headers
    for ci3,h3 in enumerate(["GSTIN","Company Name","Type"],1):
        hc=ws_cwm.cell(row=3,column=ci3,value=h3)
        hc.font=_font(True,"FFFFFF",9); hc.fill=_f(HDR_BG)
        hc.alignment=_aln("center"); hc.border=_bdr()
    col_s=4
    for _mi in range(len(MONTHS)+1):
        bg_s="C00000" if _mi==len(MONTHS) else HDR_BG
        for sv in ["Taxable","IGST","CGST","SGST","Tax Total"]:
            hc=ws_cwm.cell(row=3,column=col_s,value=sv)
            hc.font=_font(True,"FFFFFF",8); hc.fill=_f(bg_s)
            hc.alignment=_aln("center"); hc.border=_bdr()
            ws_cwm.column_dimensions[get_column_letter(col_s)].width=13
            col_s+=1
    ws_cwm.column_dimensions["A"].width=22
    ws_cwm.column_dimensions["B"].width=32
    ws_cwm.column_dimensions["C"].width=8
    ws_cwm.row_dimensions[2].height=20; ws_cwm.row_dimensions[3].height=18
    ws_cwm.freeze_panes="D4"
    rcwm=4

    # Aggregate invoice data by company and month
    comp_data={}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"
        for r in g1_inv_rows.get(mk,[]):
            inv_type,gstin_r,nm,inum,idt,iv,pos,rate,tv,ig,cg,sg=r
            if not gstin_r or gstin_r=="-": continue
            key=gstin_r.strip()
            if key not in comp_data:
                comp_data[key]={"name":nm or gstin_name_map.get(key,""),
                                "type":inv_type,
                                "months":{m:{"tx":0.,"ig":0.,"cg":0.,"sg":0.}
                                          for mm,__,yy in MONTHS for m in [f"{mm}_{yy}"]}}
            if not comp_data[key]["name"] and nm:
                comp_data[key]["name"]=nm
            comp_data[key]["months"][mk]["tx"]+=float(tv or 0)
            comp_data[key]["months"][mk]["ig"]+=float(ig or 0)
            comp_data[key]["months"][mk]["cg"]+=float(cg or 0)
            comp_data[key]["months"][mk]["sg"]+=float(sg or 0)

    # Sort by FY total descending
    sorted_keys=sorted(comp_data.keys(),
                       key=lambda k:sum(v["tx"] for v in comp_data[k]["months"].values()),
                       reverse=True)

    grand_m={mk:{"tx":0.,"ig":0.,"cg":0.,"sg":0.} for mn,_,yr in MONTHS for mk in [f"{mn}_{yr}"]}
    grand_tot={"tx":0.,"ig":0.,"cg":0.,"sg":0.}

    for key in sorted_keys:
        d=comp_data[key]
        atx=sum(v["tx"] for v in d["months"].values())
        bgc="E2EFDA" if atx>=500000 else ("EBF3FB" if atx>=100000 else (ALT2 if rcwm%2==0 else ALT1))
        for ci3,v3 in enumerate([key,d["name"] or gstin_name_map.get(key,""),d["type"]],1):
            cc=ws_cwm.cell(row=rcwm,column=ci3,value=v3)
            cc.font=_font(False,"000000",9); cc.fill=_f(bgc)
            cc.alignment=_aln("left"); cc.border=_bdr()
        col_d=4; r_ann={"tx":0.,"ig":0.,"cg":0.,"sg":0.}
        for mn,_,yr in MONTHS:
            mk=f"{mn}_{yr}"; md=d["months"][mk]
            tx=round(md["tx"],2); ig=round(md["ig"],2)
            cg=round(md["cg"],2); sg=round(md["sg"],2); tt2=round(ig+cg+sg,2)
            for v3 in [tx,ig,cg,sg,tt2]:
                cv=ws_cwm.cell(row=rcwm,column=col_d,value=v3 if v3 else "")
                cv.font=_font(False,"000000",9); cv.fill=_f(bgc)
                cv.alignment=_aln("right"); cv.border=_bdr()
                if v3 and isinstance(v3,float): cv.number_format=NUM_FMT
                col_d+=1
            r_ann["tx"]+=md["tx"]; r_ann["ig"]+=md["ig"]
            r_ann["cg"]+=md["cg"]; r_ann["sg"]+=md["sg"]
            grand_m[mk]["tx"]+=md["tx"]; grand_m[mk]["ig"]+=md["ig"]
            grand_m[mk]["cg"]+=md["cg"]; grand_m[mk]["sg"]+=md["sg"]
        for v3 in [round(r_ann["tx"],2),round(r_ann["ig"],2),round(r_ann["cg"],2),
                   round(r_ann["sg"],2),round(r_ann["ig"]+r_ann["cg"]+r_ann["sg"],2)]:
            cv=ws_cwm.cell(row=rcwm,column=col_d,value=v3)
            cv.font=_font(True,"000000",10); cv.fill=_f(TOT_BG)
            cv.alignment=_aln("right"); cv.border=_bdr()
            if v3: cv.number_format=NUM_FMT
            col_d+=1
        grand_tot["tx"]+=r_ann["tx"]; grand_tot["ig"]+=r_ann["ig"]
        grand_tot["cg"]+=r_ann["cg"]; grand_tot["sg"]+=r_ann["sg"]
        ws_cwm.row_dimensions[rcwm].height=16; rcwm+=1

    # Grand total row
    for ci3,v3 in enumerate([f"GRAND TOTAL ({len(sorted_keys)} companies)","",""],1):
        cc=ws_cwm.cell(row=rcwm,column=ci3,value=v3)
        cc.font=_font(True,"FFFFFF",10); cc.fill=_f("1F3864")
        cc.alignment=_aln("left"); cc.border=_bdr()
    col_d=4
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"; gm=grand_m[mk]
        for v3 in [round(gm["tx"],2),round(gm["ig"],2),round(gm["cg"],2),
                   round(gm["sg"],2),round(gm["ig"]+gm["cg"]+gm["sg"],2)]:
            cv=ws_cwm.cell(row=rcwm,column=col_d,value=v3)
            cv.font=_font(True,"FFFFFF",9); cv.fill=_f("1F3864")
            cv.alignment=_aln("right"); cv.border=_bdr()
            if v3: cv.number_format=NUM_FMT
            col_d+=1
    for v3 in [round(grand_tot["tx"],2),round(grand_tot["ig"],2),round(grand_tot["cg"],2),
               round(grand_tot["sg"],2),round(grand_tot["ig"]+grand_tot["cg"]+grand_tot["sg"],2)]:
        cv=ws_cwm.cell(row=rcwm,column=col_d,value=v3)
        cv.font=_font(True,"FFFFFF",9); cv.fill=_f("C00000")
        cv.alignment=_aln("right"); cv.border=_bdr()
        if v3: cv.number_format=NUM_FMT
        col_d+=1
    ws_cwm.row_dimensions[rcwm].height=20
    ws_cwm.sheet_properties.tabColor="1F3864"

    # ==========================================================
    # COMPANY-WISE VERTICAL SHEET
    # Layout: Company name as section header, then one ROW per month
    # Columns: Month | Taxable | IGST | CGST | SGST | Total Tax | Invoice Count
    # At bottom of each company: FY Total row
    # Then next company starts (separated by blank + header)
    # ==========================================================
    ws_vert = wb.create_sheet("Company_Wise_Vertical")
    ws_vert.sheet_view.showGridLines = False
    vert_cols=[("Month",14),("Taxable ₹",18),("IGST ₹",14),
               ("CGST ₹",14),("SGST ₹",14),("Total Tax ₹",16),("Invoices",10)]
    title(ws_vert, f"Company-Wise Monthly Detail (Vertical) — {client_name} ({gstin}) — FY {FY_LABEL}", len(vert_cols))
    ws_vert.merge_cells(f"A2:{get_column_letter(len(vert_cols))}2")
    sv2=ws_vert["A2"]
    sv2.value="Each company shown as a separate block | Rows=Months | FY Total at bottom of each block | Sorted by annual taxable"
    sv2.font=_font(False,"000000",8); sv2.fill=_f("FFF2CC"); sv2.alignment=_aln("left")
    ws_vert.row_dimensions[2].height=13
    for ci,(_,w) in enumerate(vert_cols,1):
        ws_vert.column_dimensions[get_column_letter(ci)].width=w
    ws_vert.freeze_panes="A3"
    rv=3

    # Build company→month data (same as Company_Month_Detail)
    comp_v={}
    for mn,_,yr in MONTHS:
        mk=f"{mn}_{yr}"
        for r in g1_inv_rows.get(mk,[]):
            inv_type,gstin_r,nm,inum,idt,iv,pos,rate,tv,ig,cg,sg=r
            if not gstin_r or gstin_r in ("-",""):  continue
            key=gstin_r.strip()
            if key not in comp_v:
                comp_v[key]={"name":nm or gstin_name_map.get(key,""),
                             "months":{f"{m}_{y}":{"tx":0.,"ig":0.,"cg":0.,"sg":0.,"inv":0}
                                       for m,__,y in MONTHS}}
            if not comp_v[key]["name"] and nm:
                comp_v[key]["name"]=nm
            comp_v[key]["months"][mk]["tx"]+=float(tv or 0)
            comp_v[key]["months"][mk]["ig"]+=float(ig or 0)
            comp_v[key]["months"][mk]["cg"]+=float(cg or 0)
            comp_v[key]["months"][mk]["sg"]+=float(sg or 0)
            comp_v[key]["months"][mk]["inv"]+=1

    # Sort companies by annual taxable descending
    sorted_comp=sorted(comp_v.keys(),
                       key=lambda k:sum(v["tx"] for v in comp_v[k]["months"].values()),
                       reverse=True)

    grand_v={"tx":0.,"ig":0.,"cg":0.,"sg":0.,"inv":0}

    for comp_key in sorted_comp:
        cd=comp_v[comp_key]
        comp_name=cd["name"] or gstin_name_map.get(comp_key,"")
        ann_tx=sum(v["tx"] for v in cd["months"].values())
        ann_ig=sum(v["ig"] for v in cd["months"].values())
        ann_cg=sum(v["cg"] for v in cd["months"].values())
        ann_sg=sum(v["sg"] for v in cd["months"].values())
        ann_inv=sum(v["inv"] for v in cd["months"].values())

        # Company header row (merged, coloured)
        ws_vert.merge_cells(f"A{rv}:{get_column_letter(len(vert_cols))}{rv}")
        ch=ws_vert.cell(row=rv,column=1,
                        value=f"  {comp_key}  |  {comp_name}  |  Annual Total: ₹{round(ann_tx,0):,.0f}")
        ch.font=_font(True,"FFFFFF",10); ch.fill=_f("2E75B6")
        ch.alignment=_aln("left"); ch.border=_bdr()
        ws_vert.row_dimensions[rv].height=18; rv+=1

        # Column headers for this company
        for ci,(h,_) in enumerate(vert_cols,1):
            hc=ws_vert.cell(row=rv,column=ci,value=h)
            hc.font=_font(True,"FFFFFF",9); hc.fill=_f(HDR_BG)
            hc.alignment=_aln("center"); hc.border=_bdr()
        ws_vert.row_dimensions[rv].height=16; rv+=1

        # One row per month
        for mn,_,yr in MONTHS:
            mk=f"{mn}_{yr}"; md=cd["months"][mk]
            tx=round(md["tx"],2); ig=round(md["ig"],2)
            cg=round(md["cg"],2); sg=round(md["sg"],2); tt=round(ig+cg+sg,2)
            bg_v=ALT2 if rv%2==0 else ALT1
            for ci,v in enumerate([f"{mn[:3]}-{yr[2:]}",tx,ig,cg,sg,tt,md["inv"]],1):
                cv=ws_vert.cell(row=rv,column=ci,value=v if v else "")
                cv.font=_font(False,"000000",9); cv.fill=_f(bg_v)
                cv.alignment=_aln("right" if ci>1 else "left"); cv.border=_bdr()
                if isinstance(v,float) and v: cv.number_format=NUM_FMT
            ws_vert.row_dimensions[rv].height=15; rv+=1

        # FY Total row for this company
        fy_bg="E2EFDA" if ann_tx>=500000 else ("EBF3FB" if ann_tx>=100000 else TOT_BG)
        for ci,v in enumerate([f"FY TOTAL — {comp_name[:20]}",
                                round(ann_tx,2),round(ann_ig,2),round(ann_cg,2),
                                round(ann_sg,2),round(ann_ig+ann_cg+ann_sg,2),ann_inv],1):
            cv=ws_vert.cell(row=rv,column=ci,value=v)
            cv.font=_font(True,"000000",10); cv.fill=_f(fy_bg)
            cv.alignment=_aln("right" if ci>1 else "left"); cv.border=_bdr()
            if isinstance(v,float): cv.number_format=NUM_FMT
        ws_vert.row_dimensions[rv].height=18; rv+=1; rv+=1  # blank row between companies

        grand_v["tx"]+=ann_tx; grand_v["ig"]+=ann_ig
        grand_v["cg"]+=ann_cg; grand_v["sg"]+=ann_sg; grand_v["inv"]+=ann_inv

    # Grand total all companies
    ws_vert.merge_cells(f"A{rv}:{get_column_letter(len(vert_cols))}{rv}")
    gc=ws_vert.cell(row=rv,column=1,
                    value=f"GRAND TOTAL — All {len(sorted_comp)} Companies — FY {FY_LABEL}")
    gc.font=_font(True,"FFFFFF",11); gc.fill=_f("C00000")
    gc.alignment=_aln("left"); gc.border=_bdr(); ws_vert.row_dimensions[rv].height=20; rv+=1
    for ci,v in enumerate(["APR-MAR",round(grand_v["tx"],2),round(grand_v["ig"],2),
                             round(grand_v["cg"],2),round(grand_v["sg"],2),
                             round(grand_v["ig"]+grand_v["cg"]+grand_v["sg"],2),grand_v["inv"]],1):
        cv=ws_vert.cell(row=rv,column=ci,value=v)
        cv.font=_font(True,"FFFFFF",10); cv.fill=_f("C00000")
        cv.alignment=_aln("right" if ci>1 else "left"); cv.border=_bdr()
        if isinstance(v,float): cv.number_format=NUM_FMT
    ws_vert.row_dimensions[rv].height=20
    ws_vert.sheet_properties.tabColor="C00000"

    wb.save(rp)
    log.info(f"  Annual Reconciliation saved: {rp.name}")

    # ======================================================
    # SEPARATE WORKBOOK 1: GSTR3B vs R1 — Portal exact format
    # Sheets: Read me | Q1-APR-JUN | Q2-JUL-SEP | Q3-OCT-DEC | Q4-JAN-MAR | Annual-APR-MAR
    # ======================================================
    p_r1 = Path(client_dir)/f"{gstin}_GSTR3BR1_RECONCILED_Summary_{FY_LABEL}.xlsx"
    wb_r1 = Workbook()
    wb_r1.remove(wb_r1.active)

    def _r1_readme():
        ws = wb_r1.create_sheet("Read me")
        ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=34
        ws.column_dimensions["C"].width=18; ws.column_dimensions["D"].width=16
        data=[
            ["Taxpayer's GSTIN", gstin, "Financial year", FY_LABEL],
            ["Trade name", client_name, "Date of generation",
             __import__("datetime").datetime.now().strftime("%d-%m-%Y")],
            [f"Taxpayer GSTIN: {gstin}\nName: {client_name}\nFY: {FY_LABEL}\n"
             f"GSTR-3B Vs GSTR-1 RECONCILED\n\n"
             f"NOTE: Data is sourced from downloaded GSTR files."]
        ]
        for ri_rm, row in enumerate(data,1):
            for ci_rm, v in enumerate(row,1):
                ws.cell(row=ri_rm, column=ci_rm, value=v)
        ws.row_dimensions[3].height=80
    _r1_readme()

    # Quarter definitions
    QUARTERS_R1=[
        ("Q1 - APR-JUN", MONTHS[0:3]),
        ("Q2 - JUL-SEP", MONTHS[3:6]),
        ("Q3 - OCT-DEC", MONTHS[6:9]),
        ("Q4 - JAN-MAR", MONTHS[9:12]),
    ]

    def _build_r1_sheet(ws_name, months_list, is_annual=False):
        ws = wb_r1.create_sheet(ws_name)
        # Column widths
        ws.column_dimensions["A"].width=44
        ws.column_dimensions["B"].width=54
        ws.column_dimensions["C"].width=3
        col=4
        if is_annual:
            labels=["Annual (APR-MAR)"]
            cols_per_period=6  # Taxable IGST CGST SGST Cess Total
        else:
            labels=[m[0] for m in months_list]+[f"Quarter - {ws_name[1]}"]
            cols_per_period=6
        for lbl in labels:
            ws.merge_cells(start_row=4,start_column=col,end_row=4,end_column=col+cols_per_period-2)
            c4=ws.cell(row=4,column=col,value=lbl)
            c4.font=Font(name="Arial",bold=True,size=9)
            c4.alignment=Alignment(horizontal="center")
            col+=cols_per_period
        # Header row 5
        col=4
        for _ in labels:
            for hh in ["Taxable","IGST","CGST","SGST","Cess","Total"]:
                h5=ws.cell(row=5,column=col,value=hh)
                h5.font=Font(name="Arial",bold=True,size=8)
                h5.alignment=Alignment(horizontal="center")
                col+=1

        # Title row
        ws.merge_cells("A1:B1")
        t1=ws["A1"]
        t1.value=f"{FY_LABEL}   Name: {client_name}   GSTIN: {gstin}   GSTR-3B Vs GSTR-1 RECONCILED"
        t1.font=Font(name="Arial",bold=True,size=10)

        # Data rows
        ROWS_R1=[
            ("GSTR-3B Supply Details","3.1(a) - Outward taxable supplies (other than zero rated, nil rated and exempted)","3b_31a"),
            ("GSTR-3B Supply Details","3.1(b) - Outward taxable supplies (zero rated)","3b_zero"),
            ("GSTR-3B Supply Details","3.1(c) - Other outward supplies (nil rated, exempted)","3b_nil"),
            ("GSTR-3B Supply Details","3.1(e) - Non GST outward supplies","3b_nongst"),
            ("GSTR-3B Supply Details","3.1.1(i) - Taxable supplies where tax is paid by recipient on RCM basis","zero"),
            ("GSTR-3B Supply Details","3.1.1(ii) - Taxable supplies between distinct persons where tax is paid on RCM","zero"),
            ("GSTR-3B Supply Details","3.2 - Inter-state supplies to unregistered persons, comp.dealers, UIN holders","zero"),
            (None,"Total from GSTR-3B (A)","tot_3b"),
            (None,None,None),
            (None,None,None),
            (None,None,None),
            ("GSTR-1 Supply Details","B2B Supplies","r1_b2b"),
            ("GSTR-1 Supply Details","B2C Small Supplies","r1_b2cs"),
            ("GSTR-1 Supply Details","B2C Large Supplies","r1_b2cl"),
            ("GSTR-1 Supply Details","Exports","r1_exp"),
            ("GSTR-1 Supply Details","Credit/Debit Notes - Registered (Net)","r1_cdncr"),
            ("GSTR-1 Supply Details","Credit/Debit Notes - Debit Notes","r1_cdndr"),
            ("GSTR-1 Supply Details","Advance Receipts","zero"),
            ("GSTR-1 Supply Details","Tax Adjustments","zero"),
            ("GSTR-1 Supply Details","Nil Rated / Exempt Supplies","r1_nil"),
            ("GSTR-1 Supply Details","Non-GST Supplies","r1_nongst"),
            (None,"Total from GSTR-1 (B)","tot_r1"),
            (None,None,None),
            (None,None,None),
            (None,None,None),
            (None,"Difference (A - B)","diff"),
        ]

        def get_vals(key, months):
            result=[]
            for mn,_,yr in months:
                mk=f"{mn}_{yr}"; d=g1[mk]; d3b=g3b_monthly.get(mk,{})
                tx_3b=round(d3b.get("taxable",0),2)
                ig_3b=round(d3b.get("o_igst",0),2); cg_3b=round(d3b.get("o_cgst",0),2); sg_3b=round(d3b.get("o_sgst",0),2)
                tot_3b_v=round(ig_3b+cg_3b+sg_3b,2)
                tx_r1=round(d["b2b_tx"]+d["b2cs_tx"],2)
                ig_r1=round(d["igst"],2); cg_r1=round(d["cgst"],2); sg_r1=round(d["sgst"],2)
                tot_r1_v=round(ig_r1+cg_r1+sg_r1,2)
                b2b_ig=round(d["igst"]*d["b2b_tx"]/(tx_r1 or 1),2)
                b2b_cg=round(d["cgst"]*d["b2b_tx"]/(tx_r1 or 1),2)
                b2b_sg=round(d["sgst"]*d["b2b_tx"]/(tx_r1 or 1),2)
                b2cs_cg=round(d["cgst"]*d["b2cs_tx"]/(tx_r1 or 1),2)
                b2cs_sg=round(d["sgst"]*d["b2cs_tx"]/(tx_r1 or 1),2)
                # CDN values from g1 dict
                cdn_cr_v = round(d.get("cdn_cr",0),2)
                cdn_dr_v = round(d.get("cdn_dr",0),2)
                cdn_cr_cg= round(d.get("cdn_cg",0)*cdn_cr_v/(cdn_cr_v+cdn_dr_v+0.001),2) if cdn_cr_v else 0
                cdn_cr_sg= round(d.get("cdn_sg",0)*cdn_cr_v/(cdn_cr_v+cdn_dr_v+0.001),2) if cdn_cr_v else 0
                cdn_cr_ig= round(d.get("cdn_ig",0)*cdn_cr_v/(cdn_cr_v+cdn_dr_v+0.001),2) if cdn_cr_v else 0
                cdn_dr_cg= round(d.get("cdn_cg",0)*cdn_dr_v/(cdn_cr_v+cdn_dr_v+0.001),2) if cdn_dr_v else 0
                cdn_dr_sg= round(d.get("cdn_sg",0)*cdn_dr_v/(cdn_cr_v+cdn_dr_v+0.001),2) if cdn_dr_v else 0
                cdn_dr_ig= round(d.get("cdn_ig",0)*cdn_dr_v/(cdn_cr_v+cdn_dr_v+0.001),2) if cdn_dr_v else 0
                nil_v = round(d.get("nil_exempt",0),2)
                non_gst_v=round(d.get("non_gst",0),2)
                # --- extra values ---
                zero_tx_3b = round(d3b.get("zero_taxable",0),2)
                zero_ig_3b = round(d3b.get("zero_igst",0),2)
                nongst_3b  = round(d3b.get("non_gst",0),2)
                exp_r1_tx  = round(d.get("exp_tx",0),2)
                exp_r1_ig  = round(d.get("exp_igst",0),2)
                b2cl_r1_tx = round(d.get("b2cl_tx",0),2)
                # Grand total from 3B = 3.1(a) + 3.1(b) zero + 3.1(c) nil
                grand_3b_tx = round(tx_3b + zero_tx_3b + round(d3b.get("nil_exempt",0),2),2)
                # Grand total from R1 = B2B + B2CS + B2CL + Exp + Nil + NonGST - CDN_cr + CDN_dr
                grand_r1_tx = round(d["b2b_tx"]+d["b2cs_tx"]+b2cl_r1_tx+exp_r1_tx+nil_v+non_gst_v-cdn_cr_v+cdn_dr_v,2)
                if   key=="3b_31a":   v=[tx_3b,ig_3b,cg_3b,sg_3b,0,tot_3b_v]
                elif key=="3b_zero":  v=[zero_tx_3b,zero_ig_3b,0,0,0,zero_ig_3b]
                elif key=="3b_nil":   v=[round(d3b.get("nil_exempt",0),2),0,0,0,0,0]
                elif key=="3b_nongst":v=[nongst_3b,0,0,0,0,0]
                elif key=="tot_3b":
                    # Grand total 3B = 3.1(a) taxable+tax + 3.1(b) zero + 3.1(c) nil
                    v=[grand_3b_tx, round(ig_3b+zero_ig_3b,2), cg_3b, sg_3b, 0, round(tot_3b_v+zero_ig_3b,2)]
                elif key=="r1_b2b":   v=[d["b2b_tx"],b2b_ig,b2b_cg,b2b_sg,0,round(b2b_ig+b2b_cg+b2b_sg,2)]
                elif key=="r1_b2cs":  v=[d["b2cs_tx"],0,b2cs_cg,b2cs_sg,0,round(b2cs_cg+b2cs_sg,2)]
                elif key=="r1_b2cl":  v=[b2cl_r1_tx,0,0,0,0,0]
                elif key=="r1_exp":   v=[exp_r1_tx,exp_r1_ig,0,0,0,exp_r1_ig]
                elif key=="r1_nil":   v=[nil_v,0,0,0,0,nil_v]
                elif key=="r1_nongst":v=[non_gst_v,0,0,0,0,non_gst_v]
                elif key=="r1_cdncr": v=[-cdn_cr_v,-cdn_cr_ig,-cdn_cr_cg,-cdn_cr_sg,0,-round(cdn_cr_ig+cdn_cr_cg+cdn_cr_sg,2)]
                elif key=="r1_cdndr": v=[cdn_dr_v,cdn_dr_ig,cdn_dr_cg,cdn_dr_sg,0,round(cdn_dr_ig+cdn_dr_cg+cdn_dr_sg,2)]
                elif key=="tot_r1":
                    v=[grand_r1_tx,ig_r1,cg_r1,sg_r1,0,tot_r1_v]
                elif key=="diff":
                    v=[round(grand_3b_tx-grand_r1_tx,2),
                       round(ig_3b-ig_r1,2),round(cg_3b-cg_r1,2),
                       round(sg_3b-sg_r1,2),0,round(tot_3b_v-tot_r1_v,2)]
                else:                 v=[0,0,0,0,0,0]
                result.extend(v)
            # Quarter total (appended as extra column after month columns)
            if not is_annual:
                for i in range(6):
                    result.append(round(sum(result[i::6][:len(months)]),2))
            return result

        def get_annual_vals(key):
            """For annual sheet: sum the given key across ALL 12 months."""
            # Sum each of the 6 columns across all 12 months
            totals = [0.0]*6
            for mn,_,yr in MONTHS:
                mk=f"{mn}_{yr}"; d=g1[mk]; d3b=g3b_monthly.get(mk,{})
                tx_3b=round(d3b.get("taxable",0),2)
                ig_3b=round(d3b.get("o_igst",0),2); cg_3b=round(d3b.get("o_cgst",0),2); sg_3b=round(d3b.get("o_sgst",0),2)
                tot_3b_v=round(ig_3b+cg_3b+sg_3b,2)
                tx_r1=round(d["b2b_tx"]+d["b2cs_tx"],2)
                ig_r1=round(d["igst"],2); cg_r1=round(d["cgst"],2); sg_r1=round(d["sgst"],2)
                tot_r1_v=round(ig_r1+cg_r1+sg_r1,2)
                b2b_ig=round(d["igst"]*d["b2b_tx"]/(tx_r1 or 1),2)
                b2b_cg=round(d["cgst"]*d["b2b_tx"]/(tx_r1 or 1),2)
                b2b_sg=round(d["sgst"]*d["b2b_tx"]/(tx_r1 or 1),2)
                b2cs_cg=round(d["cgst"]*d["b2cs_tx"]/(tx_r1 or 1),2)
                b2cs_sg=round(d["sgst"]*d["b2cs_tx"]/(tx_r1 or 1),2)
                cdn_cr_v=round(d.get("cdn_cr",0),2); cdn_dr_v=round(d.get("cdn_dr",0),2)
                cdn_ig=round(d.get("cdn_ig",0),2); cdn_cg=round(d.get("cdn_cg",0),2); cdn_sg=round(d.get("cdn_sg",0),2)
                nil_v=round(d.get("nil_exempt",0),2); non_gst_v=round(d.get("non_gst",0),2)
                zero_tx_3b = round(d3b.get("zero_taxable",0),2)
                zero_ig_3b = round(d3b.get("zero_igst",0),2)
                nongst_3b  = round(d3b.get("non_gst",0),2)
                exp_r1_tx  = round(d.get("exp_tx",0),2)
                exp_r1_ig  = round(d.get("exp_igst",0),2)
                b2cl_r1_tx = round(d.get("b2cl_tx",0),2)
                grand_3b_tx = round(tx_3b+zero_tx_3b+round(d3b.get("nil_exempt",0),2),2)
                grand_r1_tx = round(d["b2b_tx"]+d["b2cs_tx"]+b2cl_r1_tx+exp_r1_tx+nil_v+non_gst_v-cdn_cr_v+cdn_dr_v,2)
                if   key=="3b_31a":   v=[tx_3b,ig_3b,cg_3b,sg_3b,0,tot_3b_v]
                elif key=="3b_zero":  v=[zero_tx_3b,zero_ig_3b,0,0,0,zero_ig_3b]
                elif key=="3b_nil":   v=[round(d3b.get("nil_exempt",0),2),0,0,0,0,0]
                elif key=="3b_nongst":v=[nongst_3b,0,0,0,0,0]
                elif key=="tot_3b":   v=[grand_3b_tx,round(ig_3b+zero_ig_3b,2),cg_3b,sg_3b,0,round(tot_3b_v+zero_ig_3b,2)]
                elif key=="r1_b2b":   v=[d["b2b_tx"],b2b_ig,b2b_cg,b2b_sg,0,round(b2b_ig+b2b_cg+b2b_sg,2)]
                elif key=="r1_b2cs":  v=[d["b2cs_tx"],0,b2cs_cg,b2cs_sg,0,round(b2cs_cg+b2cs_sg,2)]
                elif key=="r1_b2cl":  v=[b2cl_r1_tx,0,0,0,0,0]
                elif key=="r1_exp":   v=[exp_r1_tx,exp_r1_ig,0,0,0,exp_r1_ig]
                elif key=="r1_nil":   v=[nil_v,0,0,0,0,nil_v]
                elif key=="r1_nongst":v=[non_gst_v,0,0,0,0,non_gst_v]
                elif key=="r1_cdncr": v=[-cdn_cr_v,0,0,0,0,0]
                elif key=="r1_cdndr": v=[cdn_dr_v,0,0,0,0,0]
                elif key=="tot_r1":   v=[grand_r1_tx,ig_r1,cg_r1,sg_r1,0,tot_r1_v]
                elif key=="diff":     v=[round(grand_3b_tx-grand_r1_tx,2),round(ig_3b-ig_r1,2),
                                        round(cg_3b-cg_r1,2),round(sg_3b-sg_r1,2),0,round(tot_3b_v-tot_r1_v,2)]
                else:                 v=[0,0,0,0,0,0]
                for i in range(6): totals[i]+=float(v[i] if i<len(v) else 0)
            return [round(x,2) for x in totals]

        data_r=6
        for sect,part,key in ROWS_R1:
            if key is None:
                data_r+=1; continue
            ws.cell(row=data_r,column=1,value=sect).font=Font(name="Arial",size=9)
            ws.cell(row=data_r,column=2,value=part).font=Font(name="Arial",size=9)
            vals = get_annual_vals(key) if is_annual else get_vals(key, months_list)
            bold_row = key in ("tot_3b","tot_r1","diff")
            col=4
            for v in vals:
                cv=ws.cell(row=data_r,column=col,value=v)
                cv.font=Font(name="Arial",bold=bold_row,size=9)
                if isinstance(v,(int,float)) and v!=0:
                    cv.number_format="#,##0.00"
                col+=1
            data_r+=1

    for q_name, q_months in QUARTERS_R1:
        _build_r1_sheet(q_name, q_months)
    _build_r1_sheet("Annual - APR-MAR", MONTHS, is_annual=True)

    wb_r1.save(str(p_r1))
    log.info(f"  Saved R1  Recon: {p_r1.name}")

    # ======================================================
    # SEPARATE WORKBOOK 2: GSTR3B vs R2A — Portal exact format
    # ======================================================
    p_r2a = Path(client_dir)/f"{gstin}_GSTR3BR2A_RECONCILED_Summary_{FY_LABEL}.xlsx"
    wb_r2a = Workbook()
    wb_r2a.remove(wb_r2a.active)

    def _r2a_readme():
        ws = wb_r2a.create_sheet("Read me")
        ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=34
        ws.column_dimensions["C"].width=18; ws.column_dimensions["D"].width=16
        data=[
            ["Taxpayer's GSTIN", gstin, "Financial year", FY_LABEL],
            ["Trade name", client_name, "Date of generation",
             __import__("datetime").datetime.now().strftime("%d-%m-%Y")],
            [f"Taxpayer GSTIN: {gstin}\nName: {client_name}\nFY: {FY_LABEL}\n"
             f"GSTR-3B Vs GSTR-2A RECONCILED\n\n"
             f"NOTE: Data is sourced from downloaded GSTR files."]
        ]
        for ri_rm, row in enumerate(data,1):
            for ci_rm, v in enumerate(row,1):
                ws.cell(row=ri_rm, column=ci_rm, value=v)
        ws.row_dimensions[3].height=80
    _r2a_readme()

    QUARTERS_R2A=[
        ("Q1 - APR-JUN", MONTHS[0:3]),
        ("Q2 - JUL-SEP", MONTHS[3:6]),
        ("Q3 - OCT-DEC", MONTHS[6:9]),
        ("Q4 - JAN-MAR", MONTHS[9:12]),
    ]

    def _build_r2a_sheet(ws_name, months_list, is_annual=False):
        ws = wb_r2a.create_sheet(ws_name)
        ws.column_dimensions["A"].width=44
        ws.column_dimensions["B"].width=54
        ws.column_dimensions["C"].width=3
        col=4
        if is_annual:
            labels=["Annual (APR-MAR)"]
        else:
            labels=[m[0] for m in months_list]+[f"Quarter - {ws_name[1]}"]
        for lbl in labels:
            ws.merge_cells(start_row=4,start_column=col,end_row=4,end_column=col+4)
            c4=ws.cell(row=4,column=col,value=lbl)
            c4.font=Font(name="Arial",bold=True,size=9); c4.alignment=Alignment(horizontal="center")
            col+=5
        col=4
        for _ in labels:
            for hh in ["IGST","CGST","SGST","Cess","Total"]:
                h5=ws.cell(row=5,column=col,value=hh)
                h5.font=Font(name="Arial",bold=True,size=8); h5.alignment=Alignment(horizontal="center")
                col+=1
        ws.merge_cells("A1:B1")
        t1=ws["A1"]
        t1.value=f"{FY_LABEL}   Name: {client_name}   GSTIN: {gstin}   GSTR-3B Vs GSTR-2A RECONCILED"
        t1.font=Font(name="Arial",bold=True,size=10)

        ROWS_R2A=[
            ("GSTR-3B ITC Details (Table 4)","4A(5) - All other ITC (excluding ISD, Import of Services)","3b_itc"),
            (None,"4B - ITC Reversed","zero"),
            (None,"Total from GSTR-3B (A)","tot_3b"),
            (None,None,None),(None,None,None),(None,None,None),
            ("GSTR-2A ITC Details","B2B","r2a_b2b"),
            (None,"CDNR","zero"),(None,"TDS","zero"),(None,"TCS","zero"),
            (None,"Total from GSTR-2A (B)","tot_r2a"),
            (None,None,None),(None,None,None),(None,None,None),
            (None,"Difference (A - B)","diff_itc"),
            (None,None,None),(None,None,None),(None,None,None),(None,None,None),
            ("ISD Details",None,None),
            ("GSTR-3B ISD Details","4A(4) - Inward supplies from ISD","zero"),
            ("GSTR-2A ISD Details","Total from GSTR-2A (D)","zero"),
            (None,"Difference (C - D)","zero"),
        ]

        def get_r2a_vals(key, months):
            """For quarterly sheets: returns per-month values + quarter total."""
            result=[]
            for mn,_,yr in months:
                mk=f"{mn}_{yr}"
                d3b=g3b_monthly.get(mk,{})
                ig_3b=round(d3b.get("itc_igst",0),2); cg_3b=round(d3b.get("itc_cgst",0),2)
                sg_3b=round(d3b.get("itc_sgst",0),2); tot_3b_v=round(ig_3b+cg_3b+sg_3b,2)
                ig_2a=round(g2a[mk]["tot"].get("igst",0),2); cg_2a=round(g2a[mk]["tot"].get("cgst",0),2)
                sg_2a=round(g2a[mk]["tot"].get("sgst",0),2); tot_2a=round(ig_2a+cg_2a+sg_2a,2)
                if   key=="3b_itc":  v=[ig_3b,cg_3b,sg_3b,0,tot_3b_v]
                elif key=="tot_3b":  v=[ig_3b,cg_3b,sg_3b,0,tot_3b_v]
                elif key=="r2a_b2b": v=[ig_2a,cg_2a,sg_2a,0,tot_2a]
                elif key=="tot_r2a": v=[ig_2a,cg_2a,sg_2a,0,tot_2a]
                elif key=="diff_itc":v=[round(ig_3b-ig_2a,2),round(cg_3b-cg_2a,2),round(sg_3b-sg_2a,2),0,round(tot_3b_v-tot_2a,2)]
                else:                v=[0,0,0,0,0]
                result.extend(v)
            # Quarter total column
            for i in range(5):
                result.append(round(sum(result[i::5][:len(months)]),2))
            return result

        def get_annual_r2a_vals(key):
            """For annual sheet: SUM all 12 months into a single 5-value list."""
            totals=[0.0]*5
            for mn,_,yr in MONTHS:   # always iterate ALL 12 months
                mk=f"{mn}_{yr}"
                d3b=g3b_monthly.get(mk,{})
                ig_3b=round(d3b.get("itc_igst",0),2); cg_3b=round(d3b.get("itc_cgst",0),2)
                sg_3b=round(d3b.get("itc_sgst",0),2); tot_3b_v=round(ig_3b+cg_3b+sg_3b,2)
                ig_2a=round(g2a[mk]["tot"].get("igst",0),2); cg_2a=round(g2a[mk]["tot"].get("cgst",0),2)
                sg_2a=round(g2a[mk]["tot"].get("sgst",0),2); tot_2a=round(ig_2a+cg_2a+sg_2a,2)
                if   key=="3b_itc":  v=[ig_3b,cg_3b,sg_3b,0,tot_3b_v]
                elif key=="tot_3b":  v=[ig_3b,cg_3b,sg_3b,0,tot_3b_v]
                elif key=="r2a_b2b": v=[ig_2a,cg_2a,sg_2a,0,tot_2a]
                elif key=="tot_r2a": v=[ig_2a,cg_2a,sg_2a,0,tot_2a]
                elif key=="diff_itc":v=[round(ig_3b-ig_2a,2),round(cg_3b-cg_2a,2),round(sg_3b-sg_2a,2),0,round(tot_3b_v-tot_2a,2)]
                else:                v=[0,0,0,0,0]
                for i in range(5): totals[i]+=float(v[i] if i<len(v) else 0)
            return [round(x,2) for x in totals]

        data_r=6
        for sect,part,key in ROWS_R2A:
            if sect=="ISD Details" and part is None:
                ws.cell(row=data_r,column=1,value="ISD Details").font=Font(name="Arial",bold=True,size=9)
                data_r+=1; continue
            if key is None:
                data_r+=1; continue
            ws.cell(row=data_r,column=1,value=sect).font=Font(name="Arial",size=9)
            ws.cell(row=data_r,column=2,value=part).font=Font(name="Arial",size=9)
            # KEY FIX: annual uses get_annual_r2a_vals (sums all 12); quarterly uses get_r2a_vals
            vals = get_annual_r2a_vals(key) if is_annual else get_r2a_vals(key, months_list)
            bold_row=key in ("tot_3b","tot_r2a","diff_itc")
            col=4
            for v in vals:
                cv=ws.cell(row=data_r,column=col,value=v)
                cv.font=Font(name="Arial",bold=bold_row,size=9)
                if isinstance(v,(int,float)) and v!=0: cv.number_format="#,##0.00"
                col+=1
            data_r+=1

    for q_name,q_months in QUARTERS_R2A:
        _build_r2a_sheet(q_name, q_months)
    _build_r2a_sheet("Annual - APR-MAR", MONTHS, is_annual=True)

    wb_r2a.save(str(p_r2a))
    log.info(f"  Saved R2A Recon: {p_r2a.name}")

    return str(rp)

def zip_all_outputs(base_dir, log=None):
    """
    After all processing is done, collect every generated output file
    (ANNUAL_RECONCILIATION, GSTR3BR1_RECONCILED, GSTR3BR2A_RECONCILED,
     MASTER_REPORT, GSTR1_*.xlsx, TaxLiability_*, GSTR2B_*.xlsx,
     GSTR2A_*.xlsx, GSTR3B_*.pdf, GSTR1_*.zip, GSTR2A_*.zip)
    from the entire base_dir tree, bundle them into a single ZIP, and
    open the containing folder in Windows Explorer so the user can
    download / copy everything in one click.
    """
    import zipfile as _zf, shutil as _sh

    _log = log.info if log else print

    # ── Patterns to collect ──────────────────────────────────────────
    INCLUDE_PATTERNS = [
        "ANNUAL_RECONCILIATION_*.xlsx",
        "*_GSTR3BR1_RECONCILED_*.xlsx",
        "*_GSTR3BR2A_RECONCILED_*.xlsx",
        "MASTER_REPORT_*.xlsx",
        "GSTR1_*.xlsx",          # per-client monthly invoice Excel
        "TaxLiability_*.xlsx",
        "TaxLiability_*.xls",
        "TaxLiability_*.csv",
        "GSTR2B_*.xlsx",
        "GSTR2A_*.xlsx",
        "*_R2A*.xlsx",
        "*_R2B*.xlsx",
        "GSTR3B_*.pdf",
        "GSTR1_*.zip",
        "GSTR1A_*.zip",
        "GSTR2A_*.zip",
        "GSTR2B_*.zip",
    ]
    EXCLUDE_DIRS = {"_MERGED_"}   # temp merge dirs we created

    base_path = Path(base_dir)
    collected = []

    for pattern in INCLUDE_PATTERNS:
        for f in base_path.rglob(pattern):
            # Skip files inside temp _MERGED_ dirs
            if any(part.startswith("_MERGED_") for part in f.parts):
                continue
            # Skip extraction sub-dirs (e.g. GSTR1_April_2025_ex/)
            if any(part.endswith("_ex") for part in f.parts):
                continue
            if f.is_file() and f not in collected:
                collected.append(f)

    if not collected:
        _log("  zip_all_outputs: no output files found to zip")
        return None

    zip_name = f"GST_ALL_OUTPUTS_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
    zip_path = base_path / zip_name

    _log(f"\n  Bundling {len(collected)} file(s) into {zip_name} ...")
    with _zf.ZipFile(zip_path, "w", _zf.ZIP_DEFLATED) as zout:
        for f in sorted(collected):
            # Keep relative path from base_dir so folder structure is preserved
            arcname = str(f.relative_to(base_path))
            zout.write(str(f), arcname)
            _log(f"    + {arcname}")

    size_mb = zip_path.stat().st_size / (1024*1024)
    _log(f"  ✓ ZIP created: {zip_path.name}  ({size_mb:.1f} MB)")

    # ── Open folder in Windows Explorer ─────────────────────────────
    try:
        import subprocess
        # /select, highlights the zip file itself in Explorer
        subprocess.Popen(["explorer", "/select,", str(zip_path)])
        _log(f"  ✓ Opened folder in Explorer: {base_dir}")
    except Exception as ex:
        _log(f"  (Could not open Explorer: {ex})")

    return str(zip_path)


def write_master_report(all_results, base_dir):
    wb = Workbook()
    ws = wb.active; ws.title = "Master Dashboard"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = f"GST DOWNLOAD MASTER REPORT — AY 2025-26 — {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
    c.font = Font(name="Arial",bold=True,color="FFFFFF",size=13)
    c.fill = fill(DARK_BLUE); c.alignment = aln()
    ws.row_dimensions[1].height = 38

    hdrs = ["Client","GSTIN","Month","GSTR-1","GSTR-2B","GSTR-2A","GSTR-3B"]
    wdts = [28,22,14,12,12,12,12]
    for ci,(h,w) in enumerate(zip(hdrs,wdts),1):
        c = ws.cell(row=2,column=ci,value=h)
        sc(c,bold=True,fg="FFFFFF",bg=MED_BLUE,size=10)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    ri = 3
    for res in all_results:
        # ── If login failed entirely, write a single highlighted row ──
        if res.get("login_failed"):
            msg = "LOGIN FAILED — credentials rejected after 3 attempts"
            for ci, val in enumerate([
                res.get("name", ""), res.get("gstin", ""), "ALL MONTHS",
                msg, msg, msg, msg,
            ], 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                sc(cell, bold=True, fg=RED_FG, bg=RED_BG, size=10,
                   h="left" if ci <= 3 else "center")
            ws.row_dimensions[ri].height = 20
            ri += 1
            continue

        for month_res in res.get("months", []):
            row_bg = GREY_BG if ri%2==0 else WHITE
            for ci,val in enumerate([
                res.get("name",""), res.get("gstin",""),
                month_res.get("month",""),
                month_res.get("GSTR1",""),
                month_res.get("GSTR2B",""),
                month_res.get("GSTR2A",""),
                month_res.get("GSTR3B",""),
            ],1):
                cell = ws.cell(row=ri,column=ci,value=val)
                if ci >= 4:
                    ok = val == "OK"
                    bg = GREEN_BG if ok else (YELLOW_BG if "NOT_FOUND" in str(val) else RED_BG)
                    fg = GREEN_FG if ok else (YELLOW_FG if "NOT_FOUND" in str(val) else RED_FG)
                    sc(cell,bold=True,fg=fg,bg=bg,size=10)
                else:
                    sc(cell,fg="000000",bg=row_bg,size=10,h="left")
            ws.row_dimensions[ri].height = 20
            ri += 1

    report_path = os.path.join(base_dir,
        f"MASTER_REPORT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    wb.save(report_path)
    return report_path


# ==========================================================
# LOAD CLIENTS
# ==========================================================
def load_clients(script_dir):
    # ── Try Excel files first ──────────────────────────────────────
    for fname in ["Client_Manager_Secure_AY2025-26.xlsx","clients_manager.xlsx","clients.xlsx"]:
        p = os.path.join(script_dir, fname)
        if not os.path.exists(p): continue
        try:
            df = pd.read_excel(p, sheet_name="🔐 Client Credentials", header=2, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
            col_map = {}
            for col in df.columns:
                cl = col.lower().replace("\n","").strip()
                if "client name" in cl:  col_map["name"]     = col
                elif "gstin" in cl:      col_map["gstin"]    = col
                elif "entity" in cl:     col_map["entity"]   = col
                elif "username" in cl:   col_map["username"] = col
                elif "password" in cl:   col_map["password"] = col
                elif "active" in cl:     col_map["active"]   = col
            clients = []
            for _, row in df.iterrows():
                name   = clean_str(row.get(col_map.get("name",""),""))
                user   = clean_str(row.get(col_map.get("username",""),""))
                pwd    = clean_str(row.get(col_map.get("password",""),""))
                active = str(row.get(col_map.get("active",""),"Yes")).strip().upper()
                if not name or "sample" in name.lower(): continue
                if active == "NO": continue
                if not user or not pwd:
                    print(f"  ⚠  WARNING: '{name}' has missing username/password — skipped."
                          f" Fix in {fname} → '🔐 Client Credentials' sheet.")
                    continue
                clients.append({
                    "name":name,
                    "gstin":  clean_str(row.get(col_map.get("gstin",""),"")),
                    "entity": clean_str(row.get(col_map.get("entity",""),"")),
                    "username": user,
                    "password": pwd,
                })
            if clients:
                print(f"  Loaded {len(clients)} client(s) from {fname}")
                return clients
            else:
                print(f"  ⚠  {fname} found but no valid client rows loaded.")
        except Exception as e:
            print(f"  Excel read error ({fname}): {e}")

    # ── CSV fallback ───────────────────────────────────────────────
    for fname in ["clients.csv"]:
        p = os.path.join(script_dir, fname)
        if not os.path.exists(p): continue
        try:
            import csv
            clients = []
            with open(p, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    row = {k.strip().lower(): v.strip() for k, v in row.items()}
                    name = row.get("name","") or row.get("client name","") or row.get("client","")
                    user = row.get("username","") or row.get("user","")
                    pwd  = row.get("password","") or row.get("pwd","")
                    if not name or "sample" in name.lower(): continue
                    if str(row.get("active","yes")).strip().upper() == "NO": continue
                    if not user or not pwd:
                        print(f"  ⚠  WARNING: '{name}' in clients.csv has no username/password — skipped.")
                        continue
                    clients.append({
                        "name": name,
                        "gstin": row.get("gstin",""),
                        "entity": row.get("entity",""),
                        "username": user,
                        "password": pwd,
                    })
            if clients:
                print(f"  Loaded {len(clients)} client(s) from {fname}")
                return clients
        except Exception as e:
            print(f"  CSV read error ({fname}): {e}")

    # ── Nothing found ──────────────────────────────────────────────
    print()
    print("  " + "="*60)
    print("  ✗  NO CLIENT FILE FOUND")
    print("  " + "-"*60)
    print("  Please place ONE of these files in the same folder as this script:")
    print("    • Client_Manager_Secure_AY2025-26.xlsx  (sheet: 🔐 Client Credentials)")
    print("    • clients_manager.xlsx                  (sheet: 🔐 Client Credentials)")
    print("    • clients.xlsx                          (sheet: 🔐 Client Credentials)")
    print("    • clients.csv  (columns: name, username, password, gstin, entity, active)")
    print()
    print("  Required columns: Client Name, Username, Password")
    print("  Optional columns: GSTIN, Entity, Active (YES/NO)")
    print("  " + "="*60)
    input("  Press Enter to close...")
    sys.exit(1)


# ==========================================================
# PROCESS ONE CLIENT
# ==========================================================
def process_client(client, base_dir, log, returns_todo=None):
    global CURRENT_CLIENT
    name = client["name"]
    log.info(f"\n{'='*60}")
    log.info(f"CLIENT: {name}  |  GSTIN: {client.get('gstin','')}")
    log.info(f"{'='*60}")

    safe = name.replace(" ","_").replace("/","_")
    cdir = os.path.join(base_dir, safe)
    os.makedirs(cdir, exist_ok=True)

    result = {
        "name":   name,
        "gstin":  client.get("gstin",""),
        "entity": client.get("entity",""),
        "months": [],
    }

    driver = None
    try:
        driver = make_driver(cdir)

        # Store credentials globally so relogin_if_needed can use them
        CURRENT_CLIENT = {
            "username": client["username"],
            "password": client["password"],
        }

        if not do_login(driver, client["username"], client["password"], log):
            print()
            print("  " + "="*56)
            print(f"  ✗  SKIPPING CLIENT: {name}")
            print("     Login failed after all attempts.")
            print("     Check credentials in clients.xlsx / clients.csv")
            print("     and use Option 12 (Retry Failed) to re-run this client.")
            print("  " + "="*56)
            log.error(f"  SKIPPED {name} — login failed after all attempts")
            result["login_failed"] = True
            return result

        # -- PHASE 1: Trigger/download for selected returns × all months --
        log.info(f"\n  Returns selected: {sorted(returns_todo)}")
        triggered = phase1_trigger_all(driver, cdir, log, returns_todo)

        # No fixed wait — Phase 2 polls until download link appears

        # -- PHASE 2: Download GSTR-1 / GSTR-1A / GSTR-2A --
        dl_results = phase2_download_all(driver, cdir, triggered, log, returns_todo)

        # -- AUTO-RETRY: Re-attempt failed downloads (up to 2 rounds) --
        FAIL_TAGS = {"TILE_FAIL","NOT_FOUND","TILE_NOT_FOUND","GEN_FAIL","ERR","MISS"}
        for _retry_round in range(1, 3):  # Round 1, Round 2
            failed_months = []
            for month_name, month_num, year in MONTHS:
                key = f"{month_name}_{year}"
                for rt in returns_todo:
                    if rt not in ("GSTR1","GSTR1A","GSTR2B","GSTR2A","GSTR3B"): continue
                    status = (dl_results.get(f"{key}_{rt}","")
                              or triggered.get(f"{key}_{rt}",""))
                    status = str(status).upper()
                    file_ok = False
                    suf = {rt: {
                        "GSTR1":".zip","GSTR1A":".zip","GSTR2A":".zip",
                        "GSTR2B":".xlsx","GSTR3B":".pdf"
                    }.get(rt,".zip")}.get(rt,".zip")
                    fpath = os.path.join(cdir, f"{rt}_{month_name}_{year}{suf}")
                    file_ok = os.path.exists(fpath) and os.path.getsize(fpath) > 1000
                    if not file_ok and (any(t in status for t in FAIL_TAGS) or not status or status=="SKIP"):
                        failed_months.append((month_name, month_num, year, rt))

            if not failed_months:
                log.info(f"  Auto-retry round {_retry_round}: nothing to retry - all OK ✓")
                break

            log.info(f"  Auto-retry round {_retry_round}: {len(failed_months)} missing file(s)")
            for month_name, month_num, year, rt in failed_months:
                key = f"{month_name}_{year}"
                log.info(f"    Retrying {rt} {month_name} {year}...")
                try:
                    safe_go_to_dashboard(driver, log)
                    select_and_search(driver, month_name, log)
                    time.sleep(2)
                    if rt == "GSTR3B":
                        save = f"GSTR3B_{month_name}_{year}.pdf"
                        if click_tile_download(driver, "GSTR3B", log):
                            time.sleep(2)
                            ok = rename_latest(cdir, save, [".pdf"], log)
                            if ok: triggered[f"{key}_GSTR3B"] = "OK"
                    elif rt in ("GSTR1","GSTR1A"):
                        save = f"{rt}_{month_name}_{year}.zip"
                        if click_tile_download(driver, rt, log):
                            time.sleep(2)
                            click_generate_only(driver, log)
                            # Poll for download
                            ok = poll_and_download(driver, cdir, save, log, max_wait=180)
                            if ok: dl_results[f"{key}_{rt}"] = "OK"
                    elif rt == "GSTR2B":
                        save = f"GSTR2B_{month_name}_{year}.xlsx"
                        if click_tile_download(driver, "GSTR2B", log):
                            ok = generate_then_download_immediate(driver, cdir, save, log, max_wait=120)
                            if ok: triggered[f"{key}_GSTR2B"] = "OK"
                    elif rt == "GSTR2A":
                        save = f"GSTR2A_{month_name}_{year}.zip"
                        if click_tile_download(driver, "GSTR2A", log):
                            time.sleep(2)
                            ok = poll_and_download(driver, cdir, save, log, max_wait=180)
                            if ok: dl_results[f"{key}_GSTR2A"] = "OK"
                except Exception as _re:
                    log.warning(f"    Retry failed [{rt} {month_name}]: {_re}")
            time.sleep(3)

        # -- PHASE 3: Extract GSTR-1 JSON → monthly invoice Excel --
        if "GSTR1" in returns_todo:
            extract_json_to_excel(cdir, client["name"], log)

        # -- PHASE 4: Annual Reconciliation Report ----------
        try:
            write_annual_reconciliation(
                cdir, client["name"], client.get("gstin",""), log)
        except Exception as e:
            log.warning(f"  Reconciliation report error: {e}")

        # Build month results for master report
        for month_name, month_num, year in MONTHS:
            key = f"{month_name}_{year}"
            result["months"].append({
                "month":  f"{month_name} {year}",
                "GSTR1":  dl_results.get(f"{key}_GSTR1",  triggered.get(f"{key}_GSTR1",  "SKIP")),
                "GSTR1A": dl_results.get(f"{key}_GSTR1A", triggered.get(f"{key}_GSTR1A", "SKIP")),
                "GSTR2B": triggered.get(f"{key}_GSTR2B", "SKIP"),   # Phase 1 result
                "GSTR2A": dl_results.get(f"{key}_GSTR2A", triggered.get(f"{key}_GSTR2A", "SKIP")),
                "GSTR3B": triggered.get(f"{key}_GSTR3B", "SKIP"),   # Phase 1 result
            })

        with open(os.path.join(base_dir, "progress.json"), "w") as f:
            json.dump(result, f, indent=2, default=str)

    except Exception as e:
        log.error(f"  Client error: {e}")
    finally:
        CURRENT_CLIENT = {}
        if driver:
            try: driver.quit()
            except: pass
        log.info(f"  Client done. Waiting {CLIENT_GAP}s...")
        time.sleep(CLIENT_GAP)

    return result



# ==========================================================
# OFFLINE RECONCILIATION — Run from already-downloaded files
# No browser, no login, no CAPTCHA.
# Reads existing GSTR1_*.zip / GSTR2B_*.zip / GSTR2A_*.zip /
#   GSTR3B_*.pdf from a client folder and generates the
#   ANNUAL_RECONCILIATION_*.xlsx report.
# ==========================================================
def run_offline_reconciliation(log=None):
    """
    Ask user for the folder where files are already downloaded,
    then generate the reconciliation Excel for each client sub-folder.
    No browser required.
    """
    if log is None:
        logging.basicConfig(level=logging.INFO,
            format="%(asctime)s | %(levelname)-8s | %(message)s",
            handlers=[logging.StreamHandler(sys.stdout)])
        log = logging.getLogger("gst_offline")

    print("\n" + "="*60)
    print("  OFFLINE RECONCILIATION — Use Already Downloaded Files")
    print("="*60)
    print("\n  This option reads your existing GST download folder and")
    print("  generates the Annual Reconciliation Excel for each client.")
    print("  No browser / login / CAPTCHA needed.")
    print()

    # -- Find the base folder ------------------------------
    default_base = os.path.join(os.path.expanduser("~"), "Downloads", "GST_Automation")

    print(f"  Default folder: {default_base}")
    choice = input("  Press ENTER to use default, or type a different folder path: ").strip()
    if choice:
        base_dir = choice
    else:
        base_dir = default_base

    if not os.path.isdir(base_dir):
        print(f"\n  ✗ Folder not found: {base_dir}")
        input("  Press Enter to exit..."); return

    # -- List available run-folders (AY2025-26_YYYYMMDD_HHMM) --
    run_folders = sorted([
        d for d in Path(base_dir).iterdir()
        if d.is_dir() and (d.name.startswith("AY") or d.name.startswith("20"))
    ], reverse=True)

    if not run_folders:
        # Maybe the user pointed directly at a run folder
        run_folders = [Path(base_dir)]

    print(f"\n  Found {len(run_folders)} run folder(s):")
    for i, rf in enumerate(run_folders[:10], 1):
        print(f"    [{i}] {rf.name}")
    if len(run_folders) > 10:
        print(f"    ... and {len(run_folders)-10} more")

    if len(run_folders) == 1:
        selected_run = run_folders[0]
    else:
        idx = input("\n  Enter number to select (ENTER = most recent [1]): ").strip()
        try:
            selected_run = run_folders[int(idx)-1] if idx else run_folders[0]
        except:
            selected_run = run_folders[0]

    print(f"\n  Selected: {selected_run}")

    # -- Find client sub-folders ---------------------------
    # Accept folders with ZIPs, PDFs, or direct xlsx files (portal naming)
    def _has_gst_files(d):
        return (any(d.glob("*.zip")) or any(d.glob("*.pdf")) or
                any(d.glob("*_R2A*.xlsx")) or any(d.glob("*_R2B*.xlsx")) or
                any(d.glob("GSTR*.xlsx")))

    raw_dirs = [d for d in Path(selected_run).iterdir()
                if d.is_dir() and _has_gst_files(d)]

    if not raw_dirs:
        if _has_gst_files(Path(selected_run)):
            raw_dirs = [Path(selected_run)]

    if not raw_dirs:
        print(f"\n  \u2717 No client folders with GST files found in: {selected_run}")
        print("  Expected files: GSTR1_*.zip  GSTR2B_*.xlsx  *_R2A*.xlsx  GSTR3B_*.pdf")
        input("  Press Enter to exit..."); return

    # ── GSTIN-grouping: detect per-month folders for the same GSTIN ──────────
    # When the portal downloads one month per folder (e.g. GSTR2A_April_2025_nm/
    # contains only 33AYNPV2214Q1ZU_042025_R2A.xlsx), we group all such folders
    # by GSTIN and merge their files into a single temp folder so that
    # write_annual_reconciliation can see all 12 months at once.

    def _extract_gstin_from_dir(d):
        """Try to find the 15-char GSTIN from any GST file in folder d."""
        import zipfile as _zf
        for zp in d.glob("GSTR1_*.zip"):
            try:
                with _zf.ZipFile(zp) as z:
                    for n in z.namelist():
                        if n.endswith(".json"):
                            with z.open(n) as jf:
                                data = json.load(jf)
                                g = (data.get("gstin","") or
                                     data.get("data",{}).get("gstin","") or "")
                                if g: return g
            except: pass
        for xl in list(d.glob("*_R2A*.xlsx")) + list(d.glob("*_R2B*.xlsx")) + \
                  list(d.glob("GSTR2B_*.xlsx")) + list(d.glob("GSTR1_*.zip")):
            parts = xl.stem.split("_")
            if parts and len(parts[0]) == 15:
                return parts[0]
        return ""

    # Group raw_dirs by GSTIN
    gstin_to_dirs   = {}   # gstin -> [list of source dirs]
    gstin_to_name   = {}   # gstin -> friendly client name
    no_gstin_dirs   = []   # dirs where GSTIN could not be determined

    for d in raw_dirs:
        g = _extract_gstin_from_dir(d)
        if g:
            gstin_to_dirs.setdefault(g, []).append(d)
            # Build a clean client name: strip month/year suffix from folder name
            # e.g. "GSTR2A_April_2025_nm" -> try to strip month words to get base name
            import re as _re
            month_words = r"(January|February|March|April|May|June|July|August|" \
                          r"September|October|November|December|" \
                          r"Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Oct|Nov|Dec)"
            base = _re.sub(r"(?i)[_\s]*" + month_words + r"[_\s]*\d{4}[_\s]*", "_", d.name)
            base = _re.sub(r"[_\s]+", " ", base).strip(" _")
            # Only update name if we haven't stored one, or if this one is shorter/cleaner
            if g not in gstin_to_name:
                gstin_to_name[g] = base.title()
        else:
            no_gstin_dirs.append(d)

    # Build the effective client list:
    # For each GSTIN group with >1 folder, create a merged temp dir inside selected_run
    import shutil as _shutil, tempfile as _tempfile

    client_jobs = []  # list of (cdir_path, client_name, gstin)

    for gstin, dirs in gstin_to_dirs.items():
        if len(dirs) == 1:
            # Single folder — use as-is (original behaviour)
            client_jobs.append((dirs[0], dirs[0].name.replace("_"," ").title(), gstin))
        else:
            # Multiple per-month folders → merge into one temp folder
            merge_dir = Path(selected_run) / f"_MERGED_{gstin}"
            merge_dir.mkdir(exist_ok=True)
            copied = 0
            for src_dir in dirs:
                for f in src_dir.iterdir():
                    if f.is_file():
                        dst = merge_dir / f.name
                        if not dst.exists():
                            _shutil.copy2(str(f), str(dst))
                            copied += 1
            name = gstin_to_name.get(gstin, gstin)
            log.info(f"  Merged {len(dirs)} month-folders for GSTIN {gstin} "
                     f"into {merge_dir.name}  ({copied} files copied)")
            client_jobs.append((merge_dir, name, gstin))

    # Dirs with unknown GSTIN — fall back to old behaviour (process individually)
    for d in no_gstin_dirs:
        client_jobs.append((d, d.name.replace("_"," ").title(), ""))

    # ── Display summary ───────────────────────────────────────────────────────
    print(f"\n  Found {len(client_jobs)} client(s) to process "
          f"(from {len(raw_dirs)} folder(s)):")
    for cdir, cname, cgstin in client_jobs:
        xls2a = list(cdir.glob("*_R2A*.xlsx"))
        xls2b = list(cdir.glob("*_R2B*.xlsx")) + list(cdir.glob("GSTR2B_*.xlsx"))
        zips  = list(cdir.glob("*.zip"))
        pdfs  = list(cdir.glob("*.pdf"))
        print(f"    \u2022 {cname}  ({cgstin or 'GSTIN?'})")
        print(f"      ZIPs={len(zips)}  PDFs={len(pdfs)}  "
              f"2A_xlsx={len(xls2a)}  2B_xlsx={len(xls2b)}")

    if input("\n  Generate reconciliation for ALL? (YES/no): ").strip().upper() not in ("YES","Y",""):
        print("  Cancelled."); return

    # -- Process each client -------------------------------
    generated = 0
    for cdir, client_name, gstin in client_jobs:

        print(f"\n  Processing: {client_name}  ({gstin or 'GSTIN unknown'})")
        log.info(f"\n{'='*55}")
        log.info(f"OFFLINE RECON: {client_name}  GSTIN: {gstin}")

        g1  = len(list(cdir.glob("GSTR1_*.zip")))
        g2b = len(list(cdir.glob("GSTR2B_*.xlsx")) + list(cdir.glob("GSTR2B_*.zip")) +
                  list(cdir.glob("*_R2B*.xlsx")))
        g2a = len(list(cdir.glob("GSTR2A_*.zip")) + list(cdir.glob("*_R2A*.xlsx")))
        g3b = len(list(cdir.glob("GSTR3B_*.pdf")))
        print(f"    GSTR1={g1}  GSTR2B={g2b}  GSTR2A={g2a}  GSTR3B={g3b}")

        try:
            rp = write_annual_reconciliation(str(cdir), client_name, gstin, log)
            print(f"    \u2713 Saved: {Path(rp).name}")
            generated += 1
        except Exception as e:
            print(f"    \u2717 Error: {e}")
            log.error(f"    Offline recon error [{client_name}]: {e}")
            import traceback
            log.error(traceback.format_exc())

    print(f"\n{'='*60}")
    print(f"  DONE \u2014 {generated}/{len(client_jobs)} reconciliation reports generated")
    print(f"  Folder: {selected_run}")
    print("="*60)

    # ── Auto-bundle ALL output files and open folder ───────────────
    if generated > 0:
        print("\n  Bundling all output files into a single ZIP for easy download...")
        zip_path = zip_all_outputs(str(selected_run), log)
        if zip_path:
            print(f"  ZIP: {os.path.basename(zip_path)}  ← all files bundled here")
            print(f"  Windows Explorer has been opened at: {selected_run}")

    input("\n  Press Enter to close...")



# ==========================================================
# OPTION 12: RETRY FAILED — Read Master Report Excel
# Parses MASTER_REPORT*.xlsx, finds TILE_FAIL / NOT_FOUND rows,
# asks user to confirm, then re-downloads just those months.
# ==========================================================
def retry_from_master_excel(log=None):
    """
    Read a MASTER_REPORT*.xlsx and retry all failed downloads.
    No need to re-download months that already have OK files.
    """
    import glob as _glob
    if log is None:
        logging.basicConfig(level=logging.INFO,
                            format="%(asctime)s | %(levelname)-8s | %(message)s",
                            datefmt="%Y-%m-%d %H:%M:%S")
        log = logging.getLogger(__name__)

    script_dir  = os.path.dirname(os.path.abspath(__file__))
    base_dir_def= os.path.join(os.path.expanduser("~"), "Downloads", "GST_Automation")

    print("\n" + "="*60)
    print("  RETRY FAILED — Read Master Report Excel")
    print("="*60)
    print(f"  Default folder: {base_dir_def}")
    user_dir = input("  Press ENTER for default, or type folder path: ").strip()
    base_dir = user_dir if user_dir else base_dir_def

    # Find most recent MASTER_REPORT*.xlsx
    master_files = sorted(_glob.glob(os.path.join(base_dir, "MASTER_REPORT*.xlsx")), reverse=True)
    if not master_files:
        master_files = sorted(_glob.glob(os.path.join(base_dir, "*", "MASTER_REPORT*.xlsx")), reverse=True)
    if not master_files:
        print("  No MASTER_REPORT*.xlsx found. Please enter path:")
        custom = input("  Path: ").strip().strip('"')
        if os.path.exists(custom):
            master_files = [custom]
        else:
            print("  File not found. Exiting."); return

    if len(master_files)==1:
        master_path = master_files[0]
        print(f"  Reading: {os.path.basename(master_path)}")
    else:
        print("  Found multiple Master Reports:")
        for i,f in enumerate(master_files[:10],1):
            print(f"    [{i}] {os.path.basename(f)}")
        idx = input("  Select (ENTER=1): ").strip()
        try: master_path = master_files[int(idx)-1] if idx else master_files[0]
        except: master_path = master_files[0]

    # Parse the Master Excel
    try:
        df_master = pd.read_excel(master_path, sheet_name="Master Dashboard",
                                   header=1, dtype=str)
        df_master.columns = [str(c).strip() for c in df_master.columns]
    except Exception as e:
        print(f"  Error reading master: {e}"); return

    FAIL_TAGS = ("TILE_FAIL","NOT_FOUND","TILE_NOT_FOUND","ERR")
    failed_items = []
    for _, row in df_master.iterrows():
        client = str(row.get("Client","")).strip()
        gstin  = str(row.get("GSTIN","")).strip()
        month_str = str(row.get("Month","")).strip()
        if not client or client=="nan" or not month_str or month_str=="nan": continue
        parts  = month_str.split()
        if len(parts) < 2: continue
        mn_name, mn_year = parts[0], parts[1]
        bad_returns = []
        for ret_key in ("GSTR-1","GSTR-2B","GSTR-2A","GSTR-3B","GSTR1","GSTR2B","GSTR2A","GSTR3B"):
            val = str(row.get(ret_key, row.get(ret_key.replace("-",""),""))).strip()
            if any(t in val for t in FAIL_TAGS):
                # Normalize return name
                rkey = ret_key.replace("-","")
                if rkey not in bad_returns: bad_returns.append(rkey)
        if bad_returns:
            failed_items.append({"client":client,"gstin":gstin,
                                  "month":mn_name,"year":mn_year,
                                  "returns":bad_returns})

    if not failed_items:
        print("  ✓ No failed items found in master report! All OK."); return

    print(f"\n  Found {len(failed_items)} failed item(s):")
    print(f"  {'#':3} {'Client':28} {'Month':12} {'Returns'}")
    print("  " + "-"*65)
    for i,fi in enumerate(failed_items,1):
        print(f"  [{i:2d}] {fi['client']:28} {fi['month']} {fi['year']}  → {', '.join(fi['returns'])}")

    print("\n  Options: YES=retry all  |  no=cancel  |  1 3 5=select by number")
    ans = input("  Retry? ").strip().lower()
    items_to_retry = []
    if ans in ("","yes","y","all"):
        items_to_retry = list(failed_items)
    elif ans in ("no","n","cancel"):
        return
    else:
        for ns in ans.split():
            try:
                idx = int(ns)-1
                if 0<=idx<len(failed_items): items_to_retry.append(failed_items[idx])
            except: pass

    if not items_to_retry:
        print("  Nothing selected."); return

    # Find the client folders
    run_dirs = sorted([d for d in _glob.glob(os.path.join(base_dir,"AY*"))
                       if os.path.isdir(d)], reverse=True)

    # Load clients for credentials
    clients = load_clients(os.path.dirname(os.path.abspath(__file__)))
    client_map = {c["name"].strip().lower(): c for c in clients}

    # Determine which run folder to use
    if run_dirs:
        print(f"\n  Available run folders:")
        for i,d in enumerate(run_dirs[:5],1):
            print(f"    [{i}] {os.path.basename(d)}")
        sel = input("  Select run folder (ENTER=1): ").strip()
        try: chosen_run = run_dirs[int(sel)-1] if sel else run_dirs[0]
        except: chosen_run = run_dirs[0]
    else:
        chosen_run = base_dir
    print(f"  Run folder: {chosen_run}")

    # Group by client
    from collections import defaultdict
    by_client = defaultdict(list)
    for fi in items_to_retry:
        by_client[fi["client"]].append(fi)

    for cname, mlist in by_client.items():
        client = client_map.get(cname.strip().lower())
        if not client:
            # Try partial match
            for k,v in client_map.items():
                if cname.lower() in k or k in cname.lower():
                    client = v; break
        if not client:
            print(f"  ✗ Credentials not found for {cname}"); continue

        cdir_r = os.path.join(chosen_run, cname.replace(" ","_").replace("/","_"))
        os.makedirs(cdir_r, exist_ok=True)
        print(f"\n  Retrying {cname} ({len(mlist)} month(s))...")

        driver_r = None
        try:
            driver_r = make_driver(cdir_r)
            global CURRENT_CLIENT
            CURRENT_CLIENT = {"username":client["username"],"password":client["password"]}
            if not do_login(driver_r, client["username"], client["password"], log):
                print()
                print("  " + "="*56)
                print(f"  ✗  SKIPPING: {cname}")
                print("     Login failed after all attempts.")
                print("     Fix credentials in clients.xlsx then retry again.")
                print("  " + "="*56)
                driver_r.quit(); continue

            for fi in mlist:
                mn_r = fi["month"]; yr_r = fi["year"]
                print(f"    {mn_r} {yr_r}: {fi['returns']}...")
                try:
                    # CORRECT ORDER: navigate to dashboard first, THEN set FY/Period
                    safe_go_to_dashboard(driver_r, log)
                    select_and_search(driver_r, mn_r, log)
                    time.sleep(2)
                    for rtype in fi["returns"]:
                        suf = {"GSTR3B":".pdf","GSTR2B":".xlsx"}.get(rtype,".zip")
                        save = f"{rtype}_{mn_r}_{yr_r}{suf}"
                        save_path = os.path.join(cdir_r, save)
                        if os.path.exists(save_path):
                            print(f"      → {save} already exists, skipping")
                            continue
                        if click_tile_download(driver_r, rtype, log):
                            if rtype in ("GSTR3B", "GSTR2B"):
                                # Both GSTR-3B and GSTR-2B: single-level direct download
                                time.sleep(2)
                                ok = rename_latest(cdir_r, save, [suf], log)
                                if not ok and rtype == "GSTR2B":
                                    # Fallback: portal showed generate page
                                    log.info(f"    GSTR-2B retry [{mn_r}] — generate page fallback...")
                                    _GEN_XP = [
                                        "//button[contains(text(),'GENERATE EXCEL FILE TO DOWNLOAD')]",
                                        "//button[contains(text(),'GENERATE EXCEL')]",
                                        "//button[contains(text(),'Generate Excel')]",
                                        "//a[contains(text(),'GENERATE EXCEL')]",
                                        "//button[contains(text(),'GENERATE JSON FILE TO DOWNLOAD')]",
                                        "//button[contains(text(),'Generate JSON')]",
                                    ]
                                    ok = generate_then_download_immediate(
                                        driver_r, cdir_r, save, log,
                                        gen_xpaths=_GEN_XP, max_wait=120)
                            else:
                                time.sleep(2)
                                ok = generate_then_download_immediate(
                                    driver_r, cdir_r, save, log, max_wait=120)
                            print(f"      {'✓' if ok else '✗'} {save}")
                        else:
                            print(f"      ✗ Tile not found for {rtype}")
                except Exception as _re:
                    print(f"      ✗ Error: {_re}")

        except Exception as _me:
            print(f"  ✗ Session error: {_me}")
        finally:
            if driver_r:
                try: driver_r.quit()
                except: pass
            CURRENT_CLIENT = {}

    print("\n  Retry complete. Run option 11 (Offline Recon) to rebuild reconciliation.")
    input("  Press Enter to close...")


def ask_returns_menu():
    """
    Show a numbered menu so the user can choose which returns to download.
    Returns a set of return-type strings, or the string "OFFLINE_RECON".
    """
    ALL = {"GSTR1", "GSTR1A", "GSTR2B", "GSTR2A", "GSTR3B"}

    OPTIONS = [
        # ── FULL SUITE ─────────────────────────────────────────────
        ("1",  "FULL SUITE — GSTR-1 + GSTR-1A + GSTR-2B + GSTR-2A + GSTR-3B + Tax Liability"
               "  [Phased: Generate 1&2A → Download 2B&3B → Collect 1&2A links]",
                ALL | {"TAX_LIABILITY"}),

        # ── FAST CASES (Direct download only — no generate wait) ───
        ("15", "⚡ FAST — GSTR-2B + GSTR-3B only  (Direct download, no generate, ~10 min)",
                {"GSTR2B", "GSTR3B"}),
        ("4",  "⚡ FAST — GSTR-2B only  (Excel, direct download, ~5 min)",
                {"GSTR2B"}),
        ("6",  "⚡ FAST — GSTR-3B only  (PDF, direct download, ~5 min)",
                {"GSTR3B"}),

        # ── GENERATE-FIRST CASES ──────────────────────────────────
        ("2",  "GSTR-1  only  (JSON — generate first, then download)",
                {"GSTR1"}),
        ("3",  "GSTR-1A only  (JSON — generate first, only if available)",
                {"GSTR1A"}),
        ("5",  "GSTR-2A only  (Excel — generate first, then download)",
                {"GSTR2A"}),

        # ── COMBO CASES ───────────────────────────────────────────
        ("7",  "GSTR-1 + GSTR-3B",
                {"GSTR1", "GSTR3B"}),
        ("8",  "GSTR-2B + GSTR-2A  (both purchase returns)",
                {"GSTR2B", "GSTR2A"}),
        ("9",  "GSTR-1 + GSTR-2B + GSTR-2A  (no 3B)",
                {"GSTR1", "GSTR2B", "GSTR2A"}),
        ("13", "GSTR-1 + GSTR-2B + GSTR-3B  (NO GSTR-2A)"
               "  [Phased: Generate 1 → Download 2B&3B → Collect 1]",
                {"GSTR1", "GSTR2B", "GSTR3B"}),
        ("14", "GSTR-1 + GSTR-1A + GSTR-2B + GSTR-3B  (NO GSTR-2A, 1A if available)"
               "  [Phased flow]",
                {"GSTR1", "GSTR1A", "GSTR2B", "GSTR3B"}),
        ("10", "Tax Liability & ITC Comparison only  (downloads 2024-25 + 2023-24)",
                {"TAX_LIABILITY"}),

        # ── UTILITY OPTIONS ───────────────────────────────────────
        ("11", "OFFLINE — Make Reconciliation Excel from already-downloaded files (no browser)",
                "OFFLINE_RECON"),
        ("12", "RETRY FAILED — Read Master Report Excel and re-download failed months (T1/T2/T3)",
                "RETRY_FROM_MASTER"),
        ("C",  "Custom — type return names manually",
                None),
    ]

    print("\n" + "="*70)
    print("  GST PORTAL DOWNLOAD — SELECT CASE")
    print("="*70)
    print("  FAST CASES (No generate needed — direct download):")
    for num, label, _ in OPTIONS:
        if num in ("15","4","6"):
            print(f"  [{num}]  {label}")
    print()
    print("  FULL / COMBO CASES (Phased: Generate → Direct → Collect):")
    for num, label, _ in OPTIONS:
        if num not in ("15","4","6","11","12","C","10"):
            print(f"  [{num}]  {label}")
    print()
    print("  UTILITIES:")
    for num, label, _ in OPTIONS:
        if num in ("10","11","12","C"):
            print(f"  [{num}]  {label}")
    print("="*70)
    print("  PHASED FLOW: Phase1=Generate(1,2A) → Phase2=Download(2B) →")
    print("               Phase3=Download(3B) → Phase4=Collect links(1,2A)")
    print("  RETRY LOGIC: T1 → wait 30s → T2 → wait 60s → T3 → report fail")
    print("  SESSION TIP: Downloading 2B & 3B keeps session alive during generate wait")
    print("="*70)

    while True:
        choice = input("  Enter choice: ").strip().upper()
        for num, label, ret_set in OPTIONS:
            if choice == num:
                if ret_set == "OFFLINE_RECON":
                    return "OFFLINE_RECON"
                elif ret_set == "RETRY_FROM_MASTER":
                    return "RETRY_FROM_MASTER"
                elif ret_set is not None:
                    print(f"\n  Selected: {label}")
                    return ret_set
                else:
                    print("\n  Enter return names separated by commas.")
                    print("  Valid names: GSTR1  GSTR1A  GSTR2B  GSTR2A  GSTR3B  TAX_LIABILITY")
                    raw = input("  Your selection: ").strip().upper()
                    VALID = ALL | {"TAX_LIABILITY"}
                    chosen = {r.strip() for r in raw.replace(" ","").split(",") if r.strip() in VALID}
                    if chosen:
                        print(f"  Selected: {sorted(chosen)}")
                        return chosen
                    else:
                        print("  No valid names entered. Try again.")
        print(f"  Invalid choice '{choice}'. Enter 1-11 or C.")


# ==========================================================
# MAIN
# ==========================================================
def main():
    print("\n" + "="*70)
    print("  GST COMPLETE SUITE v11 — AY 2025-26 (FINAL)")
    print("  Downloads: GSTR-1 (JSON) + GSTR-1A (JSON)")
    print("             GSTR-2B (Excel) + GSTR-2A (Excel)")
    print("             GSTR-3B (PDF) + Tax Liability & ITC")
    print()
    print("  NEW IN v11:")
    print("  • Option 15 : ⚡ FAST — GSTR-2B + GSTR-3B only (~10 min, no generate)")
    print("  • Phased DL : Generate 1&2A → Download 2B&3B → Collect 1&2A links")
    print("  • Retry T1/T2/T3: 3 retries per failed download with backoff")
    print("  • Not-Found handler: auto page reload on portal errors")
    print("  • Keep-alive: session ping during generate wait (no logout!)")
    print("  • Batch control: max 6 months per batch to avoid portal overload")
    print("="*70)

    if MISSING:
        print(f"\n  Missing packages: pip install {' '.join(MISSING)}")
        input("  Press Enter..."); return

    # -- Ask what to do -------------------------------------
    action = ask_returns_menu()

    # -- Option 11: Offline reconciliation (no browser) ----
    if action == "OFFLINE_RECON":
        run_offline_reconciliation()
        return

    # -- Option 12: Retry failed from Master Excel ----------
    if action == "RETRY_FROM_MASTER":
        retry_from_master_excel()
        return

    returns_todo = action   # it's a set of return types

    script_dir = os.path.dirname(os.path.abspath(__file__))
    clients    = load_clients(script_dir)

    base_dir = os.path.join(
        os.path.expanduser("~"), "Downloads", "GST_Automation",
        f"AY2025-26_{datetime.now().strftime('%Y%m%d_%H%M')}")
    os.makedirs(base_dir, exist_ok=True)
    print(f"\n  Output folder: {base_dir}")

    log = setup_logger(base_dir)
    print(f"\n  Clients loaded: {len(clients)}")

    print("\n  HOW IT WORKS:")
    print("  1. Browser opens → username & password auto-filled")
    print("  2. CAPTCHA appears → type it in browser → press ENTER here")
    print(f"  3. Script downloads: {', '.join(sorted(returns_todo))}")
    if "TAX_LIABILITY" in returns_todo:
        print("     Tax Liability: hover Services→Returns→Tax liabilities")
        print("     → select FY 2024-25 → SEARCH → scroll bottom → DOWNLOAD EXCEL")
        print("     → repeat for FY 2023-24")
    print("  4. Annual Reconciliation Excel generated automatically")

    if input("\n  Type YES to start: ").strip().upper() != "YES":
        print("  Cancelled."); return

    all_results = []
    for i, client in enumerate(clients, 1):
        print(f"\n  [{i}/{len(clients)}] {client['name']}")
        result = process_client(client, base_dir, log, returns_todo)
        all_results.append(result)

    print("\n  Generating master report...")
    report = write_master_report(all_results, base_dir)

    # ── Auto-bundle ALL output files and open folder ───────────────
    print("\n  Bundling all output files into a single ZIP for easy download...")
    zip_path = zip_all_outputs(base_dir, log)

    # ── Count login failures ───────────────────────────────────────
    failed_logins = [r["name"] for r in all_results if r.get("login_failed")]

    print("\n" + "="*60)
    print("  ALL DONE!")
    print(f"  Returns  : {', '.join(sorted(returns_todo))}")
    print(f"  Clients  : {len(all_results)} processed")
    if failed_logins:
        print(f"  ✗ Login FAILED ({len(failed_logins)}): {', '.join(failed_logins)}")
        print("    → Fix credentials in clients.xlsx, then use Option 12 to retry.")
    else:
        print(f"  ✓ All {len(all_results)} client(s) logged in successfully")
    print(f"  Report   : {os.path.basename(report)}")
    if zip_path:
        print(f"  ZIP      : {os.path.basename(zip_path)}  ← all files bundled here")
    print(f"  Folder   : {base_dir}")
    print("="*60)
    input("\n  Press Enter to close...")

if __name__ == "__main__":
    main()
