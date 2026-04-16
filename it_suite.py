"""
================================================================================
  INCOME TAX COMPLETE SUITE v3 — AY 2025-26 (with PDF Unlock)
  ==========================================================
  Automates login to incometax.gov.in and downloads:
    - Form 26AS  (TRACES 2.0 — TDS/TCS certificate)
    - AIS        (Annual Information Statement — PDF)
    - TIS        (Taxpayer Information Summary — PDF)

  NEW FEATURES:
    ✓ Auto-extract PAN & DOB from My Profile
    ✓ Auto-unlock PDFs after download (password removed)
    ✓ Fixed multiple download permission issue
    ✓ Profile-based password generation for PDF unlock

  PORTAL FLOW (Updated April 2026 — IT Act 2025 transition):
    https://eportal.incometax.gov.in/iec/foservices/#/login
    -> Login -> PAN + Password + OTP (mobile/email)
    -> Services -> Annual Information Statement (AIS)
         -> Download AIS / TIS
    -> e-File -> Income Tax Returns -> View Form 26AS
         -> Download 26AS (via TRACES 2.0 redirect)

  PDF PASSWORD FORMAT:
    PAN (lowercase) + DOB (DDMMYYYY)
    Example: ahjpy5761e01031985

  FILES SAVED TO:
    Downloads/IT_Automation/AY2025-26_YYYYMMDD_HHMM/ClientName/
      26AS_<PAN>_AY2025-26.pdf          (unlocked)
      AIS_<PAN>_AY2025-26.pdf           (unlocked)
      TIS_<PAN>_AY2025-26.pdf           (unlocked)
      IT_RECONCILIATION_<n>_FY2024_25.xlsx
================================================================================
"""
import os, sys, time, json, logging, shutil, re, glob
from datetime import datetime
from pathlib import Path

# ── Optional imports ──────────────────────────────────────
MISSING = []
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.edge.service   import Service as EdgeService
    from selenium.webdriver.edge.options   import Options as EdgeOptions
    from selenium.webdriver.chrome.service import Service as ChromeService
    from selenium.webdriver.chrome.options import Options as ChromeOptions
except ImportError:
    MISSING.append("selenium")

try:
    from webdriver_manager.chrome import ChromeDriverManager
    CHROME_MGR = True
except:
    CHROME_MGR = False

try:
    import pandas as pd
except ImportError:
    MISSING.append("pandas")

try:
    from pypdf import PdfReader, PdfWriter
    PIKEPDF_AVAILABLE = True  # keep same flag name so rest of code unchanged
except ImportError:
    try:
        from PyPDF2 import PdfReader, PdfWriter
        PIKEPDF_AVAILABLE = True
    except ImportError:
        PIKEPDF_AVAILABLE = False
        MISSING.append("pypdf")

# ── Constants ─────────────────────────────────────────────
IT_PORTAL      = "https://www.incometax.gov.in/iec/foportal"
IT_EPORTAL     = "https://eportal.incometax.gov.in/iec/foservices/#/login"
IT_LOGIN_URL   = "https://eportal.incometax.gov.in/iec/foservices/#/login"
IT_DASHBOARD   = "https://eportal.incometax.gov.in/iec/foservices/#/dashboard"
IT_PROFILE_URL = "https://eportal.incometax.gov.in/iec/foservices/#/dashboard/myProfile/profileDetail"
TRACES_URL     = "https://traces.tdscpc.gov.in"
TRACES_URL_OLD = "https://www.tdscpc.gov.in"
PAGE_WAIT      = 10
SHORT_WAIT     = 4
ACTION_WAIT    = 1.5
CLIENT_GAP     = 10
FY_LABEL       = "2024-25"
AY_LABEL       = "2025-26"

# ── Logging ───────────────────────────────────────────────
def setup_logger(log_dir):
    log_file = os.path.join(log_dir, f"it_automation_{datetime.now().strftime('%Y%m%d_%H%M')}.log")
    fmt = "%(asctime)s | %(levelname)-8s | %(message)s"
    logging.basicConfig(
        level=logging.INFO, format=fmt,
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ]
    )
    return logging.getLogger("it_suite")


# ==========================================================
# BROWSER SETUP (with multiple download fix)
# ==========================================================
def _find_edge_driver():
    p = shutil.which("msedgedriver")
    if p: return p
    for path in [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "msedgedriver.exe"),
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedgedriver.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedgedriver.exe",
    ]:
        if os.path.exists(path): return path
    return None


def make_driver(download_dir):
    """Create Edge or Chrome driver with auto-download to download_dir."""
    dl = str(download_dir)
    prefs = {
        "download.default_directory":       dl,
        "download.prompt_for_download":     False,
        "download.directory_upgrade":       True,
        "safebrowsing.enabled":             True,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "plugins.always_open_pdf_externally": True,
        "credentials_enable_service":       False,
        "profile.password_manager_enabled": False,
    }

    STEALTH_JS = """
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(navigator, 'languages', {get: () => ['en-IN', 'en']});
        Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
        window.chrome = { runtime: {}, loadTimes: function(){}, csi: function(){}, app: {} };
        Object.defineProperty(navigator, 'permissions', {
            query: (p) => Promise.resolve({ state: 'granted', onchange: null })
        });
    """

    def _apply_stealth(drv):
        drv.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": STEALTH_JS})
        # Grant multiple downloads permission via CDP
        drv.execute_cdp_cmd("Browser.setDownloadBehavior", {
            "behavior": "allow",
            "downloadPath": dl,
            "eventsEnabled": True,
        })
        return drv

    # Auto-detect server/cloud environment for headless mode
    _IS_SERVER = bool(os.environ.get("RENDER") or os.environ.get("PORT") or os.environ.get("HEADLESS"))

    edge_path = _find_edge_driver()
    if edge_path and not _IS_SERVER:
        try:
            opts = EdgeOptions()
            opts.add_experimental_option("prefs", prefs)
            opts.add_argument("--start-maximized")
            opts.add_argument("--disable-blink-features=AutomationControlled")
            opts.add_argument("--disable-save-password-bubble")
            opts.add_argument("--disable-features=msEdgeEnhancedSecurityMode")
            opts.add_argument("--no-sandbox")
            opts.add_argument("--disable-dev-shm-usage")
            opts.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
            opts.add_experimental_option("useAutomationExtension", False)
            opts.add_argument(
                "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/146.0.0.0 Safari/537.36 Edg/146.0.3856.78"
            )
            driver = webdriver.Edge(service=EdgeService(edge_path), options=opts)
            _apply_stealth(driver)
            print("  Browser: Microsoft Edge ✓")
            return driver
        except Exception as e:
            print(f"  Edge failed: {e} — trying Chrome...")

    try:
        opts = ChromeOptions()
        opts.add_experimental_option("prefs", prefs)
        if _IS_SERVER:
            opts.add_argument("--headless=new")
            opts.add_argument("--disable-gpu")
            opts.add_argument("--window-size=1920,1080")
            opts.add_argument("--remote-debugging-port=9222")
        else:
            opts.add_argument("--start-maximized")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-extensions")
        opts.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        opts.add_experimental_option("useAutomationExtension", False)
        # On Render, use system chromium; locally use webdriver_manager
        if _IS_SERVER:
            import shutil as _sh
            _cb = (os.environ.get("CHROME_BIN")
                   or _sh.which("chromium") or _sh.which("chromium-browser")
                   or _sh.which("google-chrome"))
            _cd = (os.environ.get("CHROMEDRIVER_PATH")
                   or _sh.which("chromedriver"))
            if _cb:
                opts.binary_location = _cb
            svc = ChromeService(executable_path=_cd) if _cd else ChromeService()
        else:
            svc = ChromeService(ChromeDriverManager().install()) if CHROME_MGR else ChromeService()
        driver = webdriver.Chrome(service=svc, options=opts)
        _apply_stealth(driver)
        print("  Browser: Google Chrome ✓")
        return driver
    except Exception as e:
        print(f"\n  Both Edge and Chrome failed: {e}")
        sys.exit(1)


# ==========================================================
# SELENIUM HELPERS
# ==========================================================
def try_click(driver, xpaths, timeout=8, log=None):
    for xp in xpaths:
        try:
            el = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.XPATH, xp)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            time.sleep(0.3)
            try:   el.click()
            except: driver.execute_script("arguments[0].click();", el)
            if log: log.info(f"    Clicked: {xp[:70]}")
            return True
        except: continue
    return False


def human_type(driver, by, val, text, log=None):
    try:
        el = WebDriverWait(driver, 10).until(EC.presence_of_element_located((by, val)))
        driver.execute_script("arguments[0].scrollIntoView({block:'center',inline:'center'});", el)
        time.sleep(0.5)

        try:
            driver.execute_script("arguments[0].click(); arguments[0].focus();", el)
            time.sleep(0.2)
            el.clear()
            time.sleep(0.15)
            for ch in str(text): el.send_keys(ch); time.sleep(0.04)
            time.sleep(0.3)
            actual = (el.get_attribute("value") or "").strip()
            if actual == str(text).strip():
                if log: log.info(f"    Typed (JS-click+keys) ✓  [{val}]")
                return True
        except Exception as e1:
            if log: log.warning(f"    Strategy1 fail [{val}]: {e1}")

        try:
            driver.execute_script("""
                var el = arguments[0]; var v = arguments[1];
                el.focus();
                var setter = Object.getOwnPropertyDescriptor(
                    window.HTMLInputElement.prototype, 'value').set;
                setter.call(el, v);
                el.dispatchEvent(new Event('input',  {bubbles:true}));
                el.dispatchEvent(new Event('change', {bubbles:true}));
                el.dispatchEvent(new KeyboardEvent('keyup',{bubbles:true}));
                el.dispatchEvent(new Event('blur',   {bubbles:true}));
            """, el, str(text))
            time.sleep(0.4)
            actual = (el.get_attribute("value") or "").strip()
            if actual == str(text).strip():
                if log: log.info(f"    Typed (JS-inject) ✓  [{val}]")
                return True
        except Exception as e2:
            if log: log.warning(f"    Strategy2 fail [{val}]: {e2}")

        try:
            driver.execute_script("arguments[0].value=''; arguments[0].focus();", el)
            time.sleep(0.2)
            driver.execute_cdp_cmd("Input.insertText", {"text": str(text)})
            time.sleep(0.3)
            if log: log.info(f"    Typed (CDP) ✓  [{val}]")
            return True
        except Exception as e3:
            if log: log.warning(f"    Strategy3 fail [{val}]: {e3}")

        try:
            from selenium.webdriver.common.action_chains import ActionChains
            ActionChains(driver).move_to_element(el).click(el).send_keys(str(text)).perform()
            time.sleep(0.3)
            if log: log.info(f"    Typed (ActionChains) ✓  [{val}]")
            return True
        except Exception as e4:
            if log: log.warning(f"    Strategy4 fail [{val}]: {e4}")

        return False

    except Exception as e:
        if log: log.warning(f"    Type failed {val}: {e}")
        return False


def get_latest_file(folder, extensions, min_size_bytes=5000):
    files = []
    for ext in extensions:
        files.extend(Path(folder).glob(f"*{ext}"))
    files = [f for f in files if f.stat().st_size >= min_size_bytes]
    return max(files, key=lambda f: f.stat().st_mtime) if files else None


def wait_for_new_file(folder, extensions, before_files, timeout=120, log=None):
    deadline = time.time() + timeout
    if log: log.info(f"    Waiting up to {timeout}s for download...")
    while time.time() < deadline:
        time.sleep(2)
        current = set(Path(folder).glob("*"))
        new_files = [
            f for f in current
            if f not in before_files
            and f.suffix.lower() in extensions
            and not f.name.endswith(".crdownload")
            and not f.name.endswith(".tmp")
            and f.stat().st_size > 1000
        ]
        if new_files:
            newest = max(new_files, key=lambda f: f.stat().st_mtime)
            if log: log.info(f"    Download complete: {newest.name}")
            return newest
        partials = list(Path(folder).glob("*.crdownload")) + list(Path(folder).glob("*.tmp"))
        if partials and log:
            log.info(f"    Still downloading... ({partials[0].name})")
    if log: log.warning(f"    Download timed out after {timeout}s")
    return None


def _type_mat_input(driver, by, val, text, log=None):
    try:
        el = WebDriverWait(driver, 8).until(EC.presence_of_element_located((by, val)))
        driver.execute_script("arguments[0].scrollIntoView({block:'center',inline:'center'});", el)
        time.sleep(0.4)
        driver.execute_script("arguments[0].click(); arguments[0].focus();", el)
        time.sleep(0.3)
        driver.execute_script("""
            arguments[0].value = '';
            arguments[0].dispatchEvent(new Event('input',{bubbles:true}));
        """, el)
        time.sleep(0.2)
        driver.execute_cdp_cmd("Input.insertText", {"text": str(text)})
        time.sleep(0.3)
        driver.execute_script("""
            arguments[0].dispatchEvent(new Event('input',  {bubbles:true}));
            arguments[0].dispatchEvent(new Event('change', {bubbles:true}));
            arguments[0].dispatchEvent(new KeyboardEvent('keyup',{bubbles:true}));
        """, el)
        time.sleep(0.2)
        actual = (el.get_attribute("value") or "").strip().upper()
        expected = str(text).strip().upper()
        if actual == expected or len(actual) > 0:
            if log: log.info(f"    ✓ Typed via _type_mat_input: '{actual}'  [{val}]")
            return True
        el.clear()
        for ch in str(text): el.send_keys(ch); time.sleep(0.04)
        time.sleep(0.2)
        actual = (el.get_attribute("value") or "").strip().upper()
        if actual == expected or len(actual) > 0:
            if log: log.info(f"    ✓ Typed via send_keys fallback: '{actual}'  [{val}]")
            return True
        return False
    except Exception as e:
        if log: log.warning(f"    _type_mat_input fail [{val}]: {e}")
        return False


# ==========================================================
# PROFILE EXTRACTION - Get PAN & DOB from My Profile
# ==========================================================
def get_profile_details(driver, pan_hint=None, log=None):
    """
    Navigate to My Profile page and extract PAN and Date of Birth.

    CONFIRMED PAGE STRUCTURE (screenshot April 2026):
    ─────────────────────────────────────────────────────────────
    URL: eportal.incometax.gov.in/iec/foservices/#/dashboard/myProfile/profileDetail
    Page shows a "Profile" card with:
      Name          : RAMANAND YADAV        (varies per client)
      Date of Birth : 01-Mar-1985           (format DD-MMM-YYYY)
      PAN           : AHJPY5761E            (format: 5L 4D 1L)
    ─────────────────────────────────────────────────────────────

    WORKS FOR ANY CLIENT — does NOT use hardcoded name text.
    Uses direct URL navigation (no dropdown clicking needed).

    pdf_password = PAN (lowercase) + DOB as DDMMYYYY
    Example:  PAN=AHJPY5761E  DOB=01-Mar-1985  →  ahjpy5761e01031985

    Returns dict: {pan, dob, pdf_password, day, month, year}
    Returns None if extraction fails.
    """
    try:
        if log: log.info("    Navigating to My Profile page (direct URL)...")

        # ── Direct navigation — most reliable, no dropdown needed ──
        driver.get(IT_PROFILE_URL)
        time.sleep(PAGE_WAIT)

        # Dismiss any logout/security popup that fires on navigation
        _dismiss_portal_popup(driver, log)
        time.sleep(2)

        # Wait for the profile card to load (look for "Date of Birth" label)
        for _ in range(20):
            try:
                body_text = driver.find_element(By.TAG_NAME, "body").text
                if "Date of Birth" in body_text or "date of birth" in body_text.lower():
                    break
            except Exception:
                pass
            time.sleep(1)

        if log: log.info(f"    Profile page URL: {driver.current_url}")

        # ── Extract full page text ─────────────────────────────────
        body = driver.find_element(By.TAG_NAME, "body")
        page_text = body.text

        # ── Extract PAN ────────────────────────────────────────────
        # PAN format: 5 uppercase letters + 4 digits + 1 uppercase letter
        # Strategy 1: look near the label "PAN" in the profile card
        pan = None
        pan_pattern = re.compile(r'\b([A-Z]{5}\d{4}[A-Z])\b')

        # Try to find the PAN value that appears after the "PAN" label
        # The profile card text looks like: "PAN\nAHJPY5761E\n..."
        pan_label_match = re.search(r'PAN\s*\n?\s*([A-Z]{5}\d{4}[A-Z])', page_text)
        if pan_label_match:
            pan = pan_label_match.group(1)
            if log: log.info(f"    PAN found near label: {pan}")

        # Strategy 2: if pan_hint provided and matches the page, use it
        if not pan and pan_hint:
            if re.fullmatch(r'[A-Z]{5}\d{4}[A-Z]', pan_hint.strip().upper()):
                if pan_hint.upper() in page_text:
                    pan = pan_hint.upper()
                    if log: log.info(f"    PAN confirmed from hint: {pan}")

        # Strategy 3: general regex scan — pick the first PAN match that appears on page
        if not pan:
            matches = pan_pattern.findall(page_text)
            if matches:
                # Filter out common false positives (e.g. masked Aadhaar prefix)
                valid = [m for m in matches if not m.startswith("XXXXX")]
                if valid:
                    pan = valid[0]
                    if log: log.info(f"    PAN found via regex scan: {pan}")

        # Strategy 4: try JavaScript extraction from DOM elements
        if not pan:
            try:
                pan = driver.execute_script("""
                    var pattern = /\\b([A-Z]{5}\\d{4}[A-Z])\\b/g;
                    var els = document.querySelectorAll('p,span,td,div,h1,h2,h3,h4,h5,h6,li');
                    for (var el of els) {
                        var t = (el.innerText || el.textContent || '').trim();
                        var m = t.match(pattern);
                        if (m && m[0] && !m[0].startsWith('XXXXX')) return m[0];
                    }
                    return null;
                """)
                if pan: log.info(f"    PAN found via JS DOM scan: {pan}") if log else None
            except Exception:
                pass

        # ── Extract Date of Birth ──────────────────────────────────
        # DOB format on profile page: DD-MMM-YYYY  (e.g. 01-Mar-1985)
        # Also handle DD/MM/YYYY or DD-MM-YYYY variants just in case
        dob = None
        dob_match = None

        # Strategy 1: look near the label "Date of Birth" in the card
        dob_label_match = re.search(
            r'Date of Birth\s*\n?\s*(\d{2}[-/](?:[A-Za-z]{3}|\d{2})[-/]\d{4})',
            page_text)
        if dob_label_match:
            dob = dob_label_match.group(1)
            if log: log.info(f"    DOB found near label: {dob}")

        # Strategy 2: general scan for DD-MMM-YYYY (most common on portal)
        if not dob:
            dob_match = re.search(r'(\d{2})-([A-Za-z]{3})-(\d{4})', page_text)
            if dob_match:
                dob = dob_match.group(0)
                if log: log.info(f"    DOB found via regex (DD-MMM-YYYY): {dob}")

        # Strategy 3: DD/MM/YYYY or DD-MM-YYYY numeric format
        if not dob:
            dob_num_match = re.search(r'(\d{2})[-/](\d{2})[-/](\d{4})', page_text)
            if dob_num_match:
                dob = dob_num_match.group(0)
                if log: log.info(f"    DOB found via regex (numeric): {dob}")

        # Strategy 4: JS DOM scan for date near DOB label
        if not dob:
            try:
                dob = driver.execute_script("""
                    var pat1 = /\\d{2}-[A-Za-z]{3}-\\d{4}/;
                    var pat2 = /\\d{2}[\\/-]\\d{2}[\\/-]\\d{4}/;
                    var els = document.querySelectorAll('p,span,td,div');
                    for (var el of els) {
                        var t = (el.innerText || el.textContent || '').trim();
                        var m = t.match(pat1) || t.match(pat2);
                        if (m) return m[0];
                    }
                    return null;
                """)
                if dob: log.info(f"    DOB found via JS DOM scan: {dob}") if log else None
            except Exception:
                pass

        # ── Build PDF password from PAN + DOB ─────────────────────
        if not pan or not dob:
            if log: log.warning(
                f"    Profile extraction incomplete — PAN: {pan}, DOB: {dob}")
            # Partial return: at least pass what we have
            if pan and not dob:
                if log: log.warning("    PAN found but DOB missing — cannot build password")
            return None

        month_map = {
            'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04',
            'MAY': '05', 'JUN': '06', 'JUL': '07', 'AUG': '08',
            'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12'
        }

        # Parse the DOB string into day, month_num, year
        day, month_num, year = None, None, None

        # DD-MMM-YYYY
        m = re.match(r'(\d{2})-([A-Za-z]{3})-(\d{4})', dob)
        if m:
            day = m.group(1)
            month_num = month_map.get(m.group(2).upper(), '01')
            year = m.group(3)

        # DD/MM/YYYY or DD-MM-YYYY (numeric)
        if not day:
            m = re.match(r'(\d{2})[-/](\d{2})[-/](\d{4})', dob)
            if m:
                day = m.group(1)
                month_num = m.group(2)
                year = m.group(3)

        if not (day and month_num and year):
            if log: log.warning(f"    Could not parse DOB: {dob}")
            return None

        pdf_password = pan.lower() + day + month_num + year

        result = {
            "pan":          pan,
            "dob":          dob,
            "pdf_password": pdf_password,
            "day":          day,
            "month":        month_num,
            "year":         year
        }

        if log: log.info(f"    ✓ Profile extracted — PAN: {pan}  DOB: {dob}")
        if log: log.info(f"    ✓ PDF Password: {pdf_password}")
        return result

    except Exception as e:
        if log: log.warning(f"    Profile extraction error: {e}")
        import traceback
        if log: log.warning(traceback.format_exc())
        return None



# ==========================================================
# PDF UNLOCK FUNCTIONS
# ==========================================================
def unlock_pdf(input_path, output_path, password, log=None):
    """Remove password from PDF using pypdf."""
    if not PIKEPDF_AVAILABLE:
        if log: log.warning("    pypdf not installed — cannot unlock PDF")
        return False

    try:
        reader = PdfReader(input_path)
        if reader.is_encrypted:
            reader.decrypt(password)
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)
        with open(output_path, "wb") as f:
            writer.write(f)
        if log: log.info(f"    ✓ PDF unlocked: {os.path.basename(output_path)}")
        return True
    except Exception as e:
        if log: log.warning(f"    PDF unlock failed: {e}")
        return False


def unlock_all_pdfs(client_dir, pdf_password, log=None):
    """
    Unlock AIS and TIS PDFs in client directory (in-place).

    Only AIS and TIS are password-protected on the IT portal.
    26AS (from TRACES) is NOT password-protected — skip it.

    Replaces the locked file with the unlocked version in-place
    so the final filename stays clean (no _unlocked suffix).
    """
    if not PIKEPDF_AVAILABLE:
        print("\n  ⚠  pypdf not installed. Install with: pip install pypdf")
        print("     PDFs will remain password-protected.")
        return {}

    unlocked_files = {}

    for pdf_file in sorted(Path(client_dir).glob("*.pdf")):
        name_lower = pdf_file.name.lower()

        # Only process AIS and TIS — skip 26AS and any already-processed files
        if not (name_lower.startswith("ais_") or name_lower.startswith("tis_")):
            if log: log.info(f"    Skipping (not AIS/TIS): {pdf_file.name}")
            continue

        if log: log.info(f"    Unlocking: {pdf_file.name}  (password: {pdf_password})")

        # Write to a temp file first, then replace original
        tmp_path = pdf_file.parent / (pdf_file.stem + "_tmp_unlock.pdf")
        try:
            reader = PdfReader(str(pdf_file))
            if reader.is_encrypted:
                result = reader.decrypt(pdf_password)
                if result == 0:
                    if log: log.warning(
                        f"    Wrong password for {pdf_file.name} — "
                        f"tried: {pdf_password}  (check PAN case + DOB format)")
                    continue
            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            with open(str(tmp_path), "wb") as f_out:
                writer.write(f_out)
            # Replace original with unlocked version
            tmp_path.replace(pdf_file)
            if log: log.info(f"    ✓ Unlocked in-place: {pdf_file.name}")
            unlocked_files[pdf_file.name] = pdf_file.name
        except ValueError:
            if log: log.warning(
                f"    Wrong password for {pdf_file.name} — "
                f"tried: {pdf_password}  (check PAN case + DOB format)")
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception as e:
            if log: log.warning(f"    Unlock error for {pdf_file.name}: {e}")
            if tmp_path.exists():
                tmp_path.unlink()

    return unlocked_files



# ==========================================================
# INCOME TAX PORTAL LOGIN
# ==========================================================
def it_login(driver, pan, password, log):
    MAX_ATTEMPTS = 3

    for attempt in range(1, MAX_ATTEMPTS + 1):
        log.info(f"    IT Portal login attempt {attempt}/{MAX_ATTEMPTS} — PAN: {pan}")

        driver.get(IT_LOGIN_URL)
        log.info(f"    Opened: {IT_LOGIN_URL}")
        time.sleep(PAGE_WAIT)

        for _ in range(15):
            inputs = driver.find_elements(By.CSS_SELECTOR, "input:not([type='hidden'])")
            if inputs:
                log.info(f"    Login form ready — {len(inputs)} input(s) visible")
                break
            time.sleep(2)
        time.sleep(1)
        log.info(f"    Page 1 URL: {driver.current_url}")

        log.info(f"    Entering PAN: {pan}")
        filled = False
        filled = _type_mat_input(driver, By.ID, "panAdhaarUserId", pan, log)

        if not filled:
            for by, val in [
                (By.NAME, "panAdhaarUserId"),
                (By.CSS_SELECTOR, "#panAdhaarUserId"),
                (By.CSS_SELECTOR, "input[name='panAdhaarUserId']"),
                (By.CSS_SELECTOR, "input[placeholder*='PAN']"),
                (By.CSS_SELECTOR, "input[placeholder*='Aadhaar']"),
                (By.CSS_SELECTOR, "input[placeholder*='User ID']"),
                (By.CSS_SELECTOR, "input[formcontrolname='userId']"),
                (By.CSS_SELECTOR, "input.mat-mdc-input-element:not([type='password'])"),
                (By.CSS_SELECTOR, "input[type='text']:not([readonly]):not([disabled])"),
            ]:
                if _type_mat_input(driver, by, val, pan, log):
                    filled = True; break

        if not filled:
            log.error("    PAN field not found — portal may be slow")
            print("\n  ✗ PAN field not found. Waiting 15s for page to load...")
            time.sleep(15)
            filled = (_type_mat_input(driver, By.ID, "panAdhaarUserId", pan, log) or
                      _type_mat_input(driver, By.CSS_SELECTOR, "input[type='text']:not([readonly])", pan, log))
            if not filled: continue

        time.sleep(ACTION_WAIT)

        log.info("    Clicking Continue (after PAN)...")
        try_click(driver, [
            "//button[contains(normalize-space(),'Continue')]",
            "//button[@type='submit']",
            "//input[@value='Continue']",
            "//button[contains(@class,'btn')][not(@disabled)]",
        ], timeout=8, log=log)

        log.info("    Waiting for password page to load...")
        for _ in range(15):
            cur = driver.current_url
            if "password" in cur or "login/pass" in cur:
                log.info(f"    Password page loaded: {cur}")
                break
            time.sleep(2)
        time.sleep(SHORT_WAIT)
        log.info(f"    Page 2 URL: {driver.current_url}")

        log.info("    Ticking 'Secure Access Message' checkbox...")
        checkbox_clicked = False
        for xp in [
            "//input[@type='checkbox']",
            "//input[contains(@id,'secure')]",
            "//input[contains(@name,'secure')]",
            "//input[contains(@id,'confirm')]",
            "//*[contains(text(),'secure access')]/preceding-sibling::input[@type='checkbox']",
            "//*[contains(text(),'secure access')]/..//input[@type='checkbox']",
            "//*[contains(text(),'confirm')]/preceding-sibling::input[@type='checkbox']",
            "//*[contains(@class,'checkbox')]//input[@type='checkbox']",
        ]:
            try:
                cb = WebDriverWait(driver, 4).until(EC.presence_of_element_located((By.XPATH, xp)))
                if not cb.is_selected():
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cb)
                    time.sleep(0.3)
                    try:   cb.click()
                    except: driver.execute_script("arguments[0].click();", cb)
                log.info(f"    Checkbox ticked ✓  ({xp[:50]})")
                checkbox_clicked = True
                break
            except: continue

        if not checkbox_clicked:
            try:
                driver.execute_script("""
                    var cbs = document.querySelectorAll('input[type="checkbox"]');
                    cbs.forEach(function(cb){
                        if(!cb.checked){
                            cb.click();
                            cb.dispatchEvent(new Event('change',{bubbles:true}));
                        }
                    });
                """)
                log.info("    Checkbox ticked via JS bulk-click ✓")
                checkbox_clicked = True
            except: pass

        if not checkbox_clicked:
            log.warning("    Checkbox not found — portal may still accept login")
        time.sleep(ACTION_WAIT)

        log.info("    Entering password...")
        filled = (
            _type_mat_input(driver, By.CSS_SELECTOR, "input[type='password']", password, log) or
            _type_mat_input(driver, By.CSS_SELECTOR, "input[formcontrolname='password']", password, log) or
            _type_mat_input(driver, By.CSS_SELECTOR, "input[formcontrolname='currentPassword']", password, log) or
            _type_mat_input(driver, By.ID, "password", password, log) or
            _type_mat_input(driver, By.NAME, "password", password, log)
        )

        if not filled:
            log.error("    Password field not found!")
            continue
        time.sleep(ACTION_WAIT)

        log.info("    Clicking Continue (submit password)...")
        try_click(driver, [
            "//button[contains(normalize-space(),'Continue')]",
            "//button[@type='submit']",
            "//input[@value='Continue']",
            "//button[contains(@class,'btn')][not(@disabled)]",
        ], timeout=8, log=log)
        time.sleep(PAGE_WAIT + 3)

        _handle_otp(driver, pan, log)
        _handle_remember_device(driver, pan, log)

        cur = driver.current_url.lower()
        log.info(f"    Post-login URL: {driver.current_url}")

        success = (
            "dashboard" in cur or
            "profile"   in cur or
            "myaccount" in cur or
            ("foservices" in cur and "login" not in cur) or
            ("eportal"   in cur and "login" not in cur)
        )
        if success:
            log.info(f"    ✓ IT Login SUCCESSFUL — PAN: {pan}")
            return True

        try:
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            if "invalid" in body or "incorrect" in body or "wrong password" in body:
                log.warning(f"    Invalid credentials for {pan}")
                print(f"\n  ✗  INVALID CREDENTIALS — PAN: {pan}")
                if attempt < MAX_ATTEMPTS:
                    import getpass
                    new_pwd = getpass.getpass(f"  New password for {pan} (ENTER to keep): ").strip()
                    if new_pwd: password = new_pwd
            elif "locked" in body or "blocked" in body:
                log.error(f"    Account LOCKED for {pan}!")
                print(f"\n  ✗  ACCOUNT LOCKED: {pan}")
                print("  Unlock at incometax.gov.in before retrying.")
                return False
            elif "login" in cur:
                log.warning(f"    Still on login page — attempt {attempt} failed")
        except: pass

    log.error(f"    IT Login FAILED after {MAX_ATTEMPTS} attempts — {pan}")
    return False


def _handle_otp(driver, pan, log):
    time.sleep(3)
    try:
        body = driver.find_element(By.TAG_NAME, "body").text.lower()
        otp_present = ("otp" in body and ("enter" in body or "verify" in body or "sent" in body))
        if not otp_present:
            return

        log.info("    OTP screen detected")
        print()
        print("  " + "="*54)
        print(f"  OTP REQUIRED — PAN: {pan}")
        print("  " + "-"*54)
        print("  An OTP has been sent to your registered")
        print("  mobile number / email address.")
        print()
        print("  1. Look at the browser window")
        print("  2. Enter the OTP in the browser field")
        print("  3. Click Validate / Submit in the browser")
        print("  4. Come back here and press ENTER")
        print("  " + "="*54)
        input("  >> Press ENTER after OTP submitted: ")
        time.sleep(PAGE_WAIT)
        log.info("    OTP submitted by user ✓")

    except Exception as e:
        log.warning(f"    OTP check error: {e}")


def _handle_remember_device(driver, pan, log):
    time.sleep(3)
    try:
        body = driver.find_element(By.TAG_NAME, "body").text.lower()
        remember_present = any(w in body for w in [
            "remember", "register this device", "trust this device",
            "don't ask again", "secure access", "add device"
        ])
        if not remember_present:
            return

        log.info("    'Remember Device' prompt detected — clicking YES...")
        print()
        print("  " + "="*54)
        print(f"  DEVICE REGISTRATION — PAN: {pan}")
        print("  " + "-"*54)
        print("  Portal is asking to register this device.")
        print("  Clicking YES so future logins won't need OTP.")
        print("  " + "="*54)

        clicked = try_click(driver, [
            "//button[normalize-space()='Yes']",
            "//button[normalize-space()='YES']",
            "//button[contains(text(),'Yes')]",
            "//button[contains(text(),'Register')]",
            "//button[contains(text(),'Trust')]",
            "//a[normalize-space()='Yes']",
            "//input[@value='Yes']",
        ], timeout=6, log=log)

        if clicked:
            log.info("    Device registered ✓ — future logins will be OTP-free")
            print("  ✓ Device registered — OTP will NOT be needed next time")
            time.sleep(3)
        else:
            print()
            print("  Could not auto-click YES. Please click 'Yes'/'Register'")
            print("  in the browser, then press ENTER here.")
            input("  >> Press ENTER after clicking YES: ")
            time.sleep(3)
            log.info("    User manually registered device")

    except Exception as e:
        log.warning(f"    Remember-device check error: {e}")


# ==========================================================
# DOWNLOAD 26AS
# ==========================================================
def download_26as(driver, client_dir, pan, log):
    log.info(f"    Downloading 26AS for {pan}...")
    before = set(Path(client_dir).iterdir())

    success = _navigate_to_26as(driver, log)
    if not success:
        log.warning("    Could not navigate to 26AS page")
        return None

    time.sleep(PAGE_WAIT)

    cur = driver.current_url
    log.info(f"    Current URL: {cur}")

    if "tdscpc" in cur or "traces" in cur.lower() or "traces2" in cur.lower():
        log.info("    On TRACES portal — proceeding to download...")
        result = _download_26as_from_traces(driver, client_dir, before, log)
    else:
        log.info("    On IT portal — looking for 26AS download link...")
        result = _download_26as_direct(driver, client_dir, before, log)

    if result:
        std_name = Path(client_dir) / f"26AS_{pan}_AY{AY_LABEL.replace('-','_')}.pdf"
        if result != std_name and result.exists():
            result.rename(std_name)
            log.info(f"    26AS saved: {std_name.name} ✓")
            return str(std_name)
        log.info(f"    26AS saved: {result.name} ✓")
        return str(result)

    log.warning("    26AS download failed or timed out")
    return None


def _efile_hover_submenu(driver, log):
    from selenium.webdriver.common.action_chains import ActionChains
    log.info("    Step 1: Clicking e-File menu...")

    efile_clicked = try_click(driver, [
        "//a[normalize-space()='e-File']",
        "//span[normalize-space()='e-File']",
        "//li[normalize-space()='e-File']",
        "//button[normalize-space()='e-File']",
        "//*[normalize-space()='e-File']",
        "//nav//*[contains(normalize-space(),'e-File')]",
        "//*[contains(@class,'nav')]//*[contains(normalize-space(),'e-File')]",
    ], timeout=8, log=log)

    if not efile_clicked:
        try:
            driver.execute_script("""
                var tags = ['a','span','li','button','div'];
                for(var tag of tags){
                    var els = document.querySelectorAll(tag);
                    for(var el of els){
                        var t = (el.innerText||el.textContent||'').trim();
                        if(t === 'e-File'){
                            el.click(); return;
                        }
                    }
                }
            """)
            efile_clicked = True
            log.info("    e-File clicked via JS text scan ✓")
        except Exception as je:
            log.warning(f"    JS e-File click failed: {je}")

    if not efile_clicked:
        log.warning("    e-File menu not found!")
        return False

    time.sleep(SHORT_WAIT)
    log.info(f"    e-File clicked ✓  URL: {driver.current_url}")

    log.info("    Step 2: Hovering over 'Income Tax Returns' to open flyout...")
    itr_el = None

    for xp in [
        "//*[normalize-space()='Income Tax Returns']",
        "//*[contains(normalize-space(),'Income Tax Returns')]",
        "//a[contains(normalize-space(),'Income Tax Return')]",
        "//span[contains(normalize-space(),'Income Tax Return')]",
        "//li[contains(normalize-space(),'Income Tax Return')]",
        "//div[contains(normalize-space(),'Income Tax Return')]",
    ]:
        try:
            el = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, xp)))
            if el.is_displayed():
                itr_el = el
                log.info(f"    Found 'Income Tax Returns' via: {xp[:60]}")
                break
        except:
            continue

    if itr_el:
        ActionChains(driver).move_to_element(itr_el).perform()
        time.sleep(2)
        log.info("    Hovered over 'Income Tax Returns' ✓ — flyout should be open")
    else:
        log.warning("    'Income Tax Returns' not found — trying JS mouseenter...")
        try:
            driver.execute_script("""
                var tags = ['a','span','li','div','button'];
                for(var tag of tags){
                    var els = document.querySelectorAll(tag);
                    for(var el of els){
                        var t = (el.innerText||el.textContent||'').trim();
                        if(t.includes('Income Tax Return')){
                            el.dispatchEvent(new MouseEvent('mouseenter',{bubbles:true}));
                            el.dispatchEvent(new MouseEvent('mouseover',{bubbles:true}));
                            el.classList.add('open');
                            el.classList.add('active');
                            break;
                        }
                    }
                }
            """)
            time.sleep(2)
            log.info("    JS mouseenter on 'Income Tax Returns' ✓")
        except Exception as je:
            log.warning(f"    JS hover fallback failed: {je}")

    return True


def _navigate_to_26as(driver, log):
    from selenium.webdriver.common.action_chains import ActionChains
    log.info("    Nav: e-File → Income Tax Returns → View Form 26AS")
    original_handles = set(driver.window_handles)

    if not _efile_hover_submenu(driver, log):
        return False

    log.info("    Step 3: Clicking 'View Form 26AS' in flyout...")
    clicked = try_click(driver, [
        "//*[normalize-space()='View Form 26AS']",
        "//a[normalize-space()='View Form 26AS']",
        "//span[normalize-space()='View Form 26AS']",
        "//li[normalize-space()='View Form 26AS']",
        "//*[contains(normalize-space(),'View Form 26AS')]",
        "//a[contains(normalize-space(),'26AS')]",
        "//*[contains(normalize-space(),'26AS')]",
        "//a[contains(normalize-space(),'View Tax Credit')]",
        "//a[contains(normalize-space(),'Form 168')]",
    ], timeout=8, log=log)

    if not clicked:
        log.warning("    XPath failed — JS text scan for 'View Form 26AS'...")
        try:
            result = driver.execute_script("""
                var tags = ['a','button','span','li','div','p'];
                for(var tag of tags){
                    var els = document.querySelectorAll(tag);
                    for(var el of els){
                        var t = (el.innerText||el.textContent||'').trim();
                        if(t === 'View Form 26AS' && el.offsetParent !== null){
                            el.scrollIntoView({block:'center'});
                            el.click();
                            return 'clicked exact: ' + t;
                        }
                    }
                }
                for(var tag of tags){
                    var els = document.querySelectorAll(tag);
                    for(var el of els){
                        var t = (el.innerText||el.textContent||'').trim();
                        if(t.includes('26AS') && el.offsetParent !== null){
                            el.scrollIntoView({block:'center'});
                            el.click();
                            return 'clicked partial: ' + t;
                        }
                    }
                }
                return 'not found';
            """)
            log.info(f"    JS scan result: {result}")
            if result and "not found" not in str(result):
                clicked = True
        except Exception as je:
            log.warning(f"    JS text scan failed: {je}")

    if not clicked:
        log.warning("    'View Form 26AS' not found — logging all visible menu items...")
        try:
            vis = driver.execute_script("""
                var items = [];
                document.querySelectorAll('a,span,li,button').forEach(function(el){
                    var t = (el.innerText||el.textContent||'').trim();
                    if(t && t.length < 80 && el.offsetParent !== null) items.push(t);
                });
                return items;
            """)
            log.warning(f"    Visible clickable items: {vis[:30]}")
        except: pass
        return False

    log.info("    'View Form 26AS' clicked ✓ — waiting for TRACES...")
    time.sleep(PAGE_WAIT)

    new_handles = set(driver.window_handles) - original_handles
    if new_handles:
        driver.switch_to.window(list(new_handles)[0])
        log.info(f"    Switched to TRACES tab: {driver.current_url}")
    else:
        log.info(f"    26AS on same tab: {driver.current_url}")

    time.sleep(PAGE_WAIT)
    return True


def _download_26as_from_traces(driver, client_dir, before, log):
    log.info("    TRACES — starting 26AS download...")
    time.sleep(SHORT_WAIT)

    log.info("    Stage A: ticking disclaimer checkbox...")
    _tick_all_checkboxes(driver, log)
    time.sleep(ACTION_WAIT)

    log.info("    Stage B: clicking Proceed...")
    try_click(driver, [
        "//button[normalize-space()='Proceed']",
        "//input[@value='Proceed']",
        "//button[contains(normalize-space(),'Proceed')]",
        "//a[normalize-space()='Proceed']",
    ], timeout=8, log=log)
    time.sleep(PAGE_WAIT + 2)
    log.info(f"    After Proceed: {driver.current_url}")

    log.info("    Stage C: clicking 'View Tax Credit (Form 26AS/Annual Tax Statement)' link...")
    clicked = try_click(driver, [
        "//a[contains(normalize-space(),'View Tax Credit (Form 26AS/Annual Tax Statement)')]",
        "//a[contains(normalize-space(),'View Tax Credit')]",
        "//a[contains(normalize-space(),'Annual Tax Statement')]",
        "//a[contains(normalize-space(),'Form 26AS')]",
        "//a[contains(normalize-space(),'26AS')]",
        "//a[contains(@href,'view26AS')]",
        "//a[contains(@href,'form26AS')]",
        "//a[contains(@href,'annualTaxStatement')]",
    ], timeout=10, log=log)
    if clicked:
        time.sleep(PAGE_WAIT + 2)
        log.info(f"    After 'View Tax Credit': {driver.current_url}")

    log.info("    Stage D: selecting Financial Year 2025-26...")
    fy_selected = False
    try:
        from selenium.webdriver.support.ui import Select
        all_selects = driver.find_elements(By.TAG_NAME, "select")
        log.info(f"    Found {len(all_selects)} dropdown(s) on TRACES page")
        for sel_el in all_selects:
            try:
                s = Select(sel_el)
                opts_text = [o.text.strip() for o in s.options]
                log.info(f"    Dropdown options: {opts_text}")
                for opt in s.options:
                    if FY_LABEL in opt.text or "2025-26" in opt.text or "2025" in opt.text:
                        s.select_by_visible_text(opt.text)
                        log.info(f"    FY selected: {opt.text} ✓")
                        fy_selected = True
                        break
                if fy_selected:
                    break
            except:
                continue
    except Exception as e:
        log.warning(f"    FY dropdown error: {e}")
    if not fy_selected:
        log.warning("    FY dropdown not found — proceeding with default")
    time.sleep(SHORT_WAIT)

    log.info("    Stage E: selecting 'View As = HTML'...")
    try:
        from selenium.webdriver.support.ui import Select
        for sel_el in driver.find_elements(By.TAG_NAME, "select"):
            try:
                s = Select(sel_el)
                opts_text = [o.text.lower() for o in s.options]
                if any("html" in o for o in opts_text):
                    for opt in s.options:
                        if "html" in opt.text.lower():
                            s.select_by_visible_text(opt.text)
                            log.info(f"    View As = {opt.text} ✓")
                            break
                    break
            except:
                continue
    except Exception as e:
        log.warning(f"    View As dropdown error: {e}")
    time.sleep(SHORT_WAIT)

    log.info("    Stage F: clicking View / Download button...")
    try_click(driver, [
        "//input[@value='View / Download']",
        "//button[normalize-space()='View / Download']",
        "//input[@value='View/Download']",
        "//button[normalize-space()='View/Download']",
        "//input[@value='View']",
        "//button[normalize-space()='View']",
        "//a[normalize-space()='View']",
        "//button[contains(normalize-space(),'View')]",
        "//input[contains(@value,'View')]",
    ], timeout=10, log=log)
    time.sleep(PAGE_WAIT + 3)

    log.info("    Stage G: clicking 'Export as PDF'...")
    clicked = try_click(driver, [
        "//a[contains(normalize-space(),'Export as PDF')]",
        "//button[contains(normalize-space(),'Export as PDF')]",
        "//input[contains(@value,'Export as PDF')]",
        "//a[contains(normalize-space(),'Export')]",
        "//button[contains(normalize-space(),'Export')]",
        "//a[contains(normalize-space(),'Download')]",
        "//button[contains(normalize-space(),'Download')]",
        "//a[contains(@href,'.pdf')]",
    ], timeout=10, log=log)

    if not clicked:
        log.warning("    Export as PDF not found — 26AS download failed")
        return None

    time.sleep(SHORT_WAIT)
    try_click(driver, [
        "//button[normalize-space()='OK']",
        "//button[normalize-space()='Confirm']",
    ], timeout=3, log=log)

    return wait_for_new_file(client_dir, {".pdf"}, before, timeout=120, log=log)


def _download_26as_direct(driver, client_dir, before, log):
    time.sleep(SHORT_WAIT)
    try_click(driver, [
        "//button[contains(text(),'Download')]",
        "//a[contains(text(),'Download')]",
        "//a[contains(@href,'.pdf')]",
    ], timeout=10, log=log)
    time.sleep(SHORT_WAIT)
    return wait_for_new_file(client_dir, {".pdf"}, before, timeout=90, log=log)


def _tick_all_checkboxes(driver, log):
    ticked = 0
    try:
        for cb in driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox']"):
            try:
                if not cb.is_selected():
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cb)
                    time.sleep(0.2)
                    try:   cb.click()
                    except: driver.execute_script("arguments[0].click();", cb)
                    ticked += 1
            except: continue
    except: pass
    if ticked == 0:
        try:
            driver.execute_script("""
                document.querySelectorAll('input[type="checkbox"]').forEach(function(cb){
                    if(!cb.checked){ cb.click();
                        cb.dispatchEvent(new Event('change',{bubbles:true})); }
                });
            """)
        except: pass
    log.info(f"    Ticked {ticked} checkbox(es) ✓")


# ==========================================================
# DOWNLOAD AIS & TIS (with multiple download fix)
# ==========================================================
def _is_popup_open(driver):
    try:
        btns = driver.find_elements(By.XPATH, "//button[normalize-space()='Download']")
        visible = [b for b in btns if b.is_displayed()]
        return len(visible) >= 3
    except Exception:
        return False


def _close_download_popup(driver, log):
    from selenium.webdriver.common.keys import Keys
    closed = try_click(driver, [
        "//*[@role='dialog']//button[contains(@aria-label,'close') or contains(@aria-label,'Close')]",
        "//*[@role='dialog']//button[normalize-space()='×' or normalize-space()='✕' or normalize-space()='Close']",
        "//*[contains(@class,'modal') or contains(@class,'dialog')]"
        "//button[contains(@class,'close') or contains(normalize-space(),'close')]",
        "//button[@mat-dialog-close]",
        "//button[contains(@class,'close')]",
    ], timeout=3, log=log)
    if not closed:
        try:
            driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
            time.sleep(0.5)
            log.info("    Popup dismissed via Escape key")
        except Exception as e:
            log.warning(f"    Popup close fallback failed: {e}")
    else:
        log.info("    Popup closed via X/Close button ✓")
    time.sleep(1)


def download_ais_tis(driver, client_dir, pan, log):
    log.info(f"    Downloading AIS & TIS for {pan}...")

    success = _navigate_to_ais(driver, log)
    if not success:
        log.warning("    Could not navigate to AIS page")
        return None, None

    time.sleep(PAGE_WAIT)
    log.info(f"    AIS page URL: {driver.current_url}")

    before = set(Path(client_dir).iterdir())
    ais_path = None
    tis_path = None

    log.info("    Opening Download AIS/TIS popup (single session for both)...")
    if not _open_download_popup(driver, log):
        log.warning("    Could not open Download popup — cannot download AIS or TIS")
        return None, None

    log.info("    Clicking AIS PDF [Download] button...")
    ais_clicked = _click_popup_download_btn(driver, "ais", log)

    if ais_clicked:
        time.sleep(SHORT_WAIT)
        log.info("    Waiting for AIS PDF download...")
        ais_result = wait_for_new_file(client_dir, {".pdf"}, before, timeout=120, log=log)
        if ais_result:
            std = Path(client_dir) / f"AIS_{pan}_AY{AY_LABEL.replace('-','_')}.pdf"
            try:
                if ais_result.exists() and ais_result != std:
                    ais_result.rename(std)
                    ais_result = std
            except Exception as e:
                log.warning(f"    AIS rename failed: {e}")
            ais_path = str(ais_result)
            log.info(f"    AIS saved: {ais_result.name} ✓")
        else:
            log.warning("    AIS download timed out")
    else:
        log.warning("    AIS PDF Download button not found — skipping AIS")

    before_after_ais = set(Path(client_dir).iterdir())
    time.sleep(SHORT_WAIT)

    popup_still_open = _is_popup_open(driver)
    log.info(f"    Popup still open after AIS? {popup_still_open}")

    if not popup_still_open:
        log.info("    Popup closed after AIS — waiting before reopening for TIS...")
        _close_download_popup(driver, log)
        time.sleep(SHORT_WAIT + 2)

        log.info("    Reopening Download popup for TIS...")
        if not _open_download_popup(driver, log):
            log.warning("    Could not reopen popup for TIS — skipping TIS")
            return ais_path, None
    else:
        log.info("    Popup still open — clicking TIS directly (no reopen needed) ✓")

    log.info("    Clicking TIS PDF [Download] button...")
    tis_clicked = _click_popup_download_btn(driver, "tis", log)

    if tis_clicked:
        time.sleep(SHORT_WAIT)
        log.info("    Waiting for TIS PDF download...")
        tis_result = wait_for_new_file(client_dir, {".pdf"}, before_after_ais, timeout=120, log=log)
        if tis_result:
            std = Path(client_dir) / f"TIS_{pan}_AY{AY_LABEL.replace('-','_')}.pdf"
            try:
                if tis_result.exists() and tis_result != std:
                    tis_result.rename(std)
                    tis_result = std
            except Exception as e:
                log.warning(f"    TIS rename failed: {e}")
            tis_path = str(tis_result)
            log.info(f"    TIS saved: {tis_result.name} ✓")
        else:
            log.warning("    TIS download timed out")
    else:
        log.warning("    TIS PDF Download button not found — skipping TIS")

    return ais_path, tis_path


def _navigate_to_ais(driver, log):
    log.info("    Nav: clicking AIS directly (no e-File menu)")

    cur = driver.current_url
    if "traces" in cur.lower() or "insight.gov" in cur.lower():
        log.warning("    Still on TRACES/AIS tab — switching to IT portal tab...")
        for h in driver.window_handles:
            driver.switch_to.window(h)
            url = driver.current_url
            if "eportal.incometax" in url or "incometax.gov" in url:
                log.info(f"    Switched to IT portal tab: {url}")
                break
        time.sleep(SHORT_WAIT)

    _dismiss_portal_popup(driver, log)

    original_handles = set(driver.window_handles)
    log.info(f"    Starting from: {driver.current_url}  (tabs: {len(original_handles)})")

    log.info("    Step 1: clicking 'AIS' link/button directly...")
    clicked = try_click(driver, [
        "//a[normalize-space()='AIS']",
        "//li/a[normalize-space()='AIS']",
        "//nav//a[normalize-space()='AIS']",
        "//button[normalize-space()='AIS']",
        "//*[normalize-space()='AIS']",
        "//a[contains(normalize-space(),'Annual Information Statement')]",
        "//*[contains(@class,'ais')]//a",
        "//*[contains(@id,'ais')]",
    ], timeout=10, log=log)

    if not clicked:
        log.warning("    Direct AIS click failed — trying Services → Annual Information Statement...")
        try_click(driver, [
            "//a[normalize-space()='Services']",
            "//span[normalize-space()='Services']",
            "//*[normalize-space()='Services']",
        ], timeout=8, log=log)
        time.sleep(SHORT_WAIT)
        clicked = try_click(driver, [
            "//*[normalize-space()='Annual Information Statement (AIS)']",
            "//a[contains(normalize-space(),'Annual Information Statement')]",
            "//*[contains(normalize-space(),'Annual Information Statement')]",
            "//*[normalize-space()='AIS']",
        ], timeout=8, log=log)

    if not clicked:
        log.warning("    XPath failed — JS full-page scan for AIS...")
        try:
            result = driver.execute_script("""
                var tags = ['a','button','span','li','div'];
                for(var tag of tags){
                    var els = document.querySelectorAll(tag);
                    for(var el of els){
                        var t = (el.innerText||el.textContent||'').trim();
                        if(t === 'AIS' && el.offsetParent !== null){
                            el.scrollIntoView({block:'center'});
                            el.click();
                            return 'clicked exact: AIS';
                        }
                    }
                }
                for(var tag of tags){
                    var els = document.querySelectorAll(tag);
                    for(var el of els){
                        var t = (el.innerText||el.textContent||'').trim();
                        if(t.includes('Annual Information Statement') && el.offsetParent !== null){
                            el.scrollIntoView({block:'center'});
                            el.click();
                            return 'clicked partial: ' + t;
                        }
                    }
                }
                return 'not found';
            """)
            log.info(f"    JS scan result: {result}")
            if result and "not found" not in str(result):
                clicked = True
        except Exception as je:
            log.warning(f"    JS AIS scan failed: {je}")

    if not clicked:
        log.warning("    AIS navigation completely failed — logging visible items...")
        try:
            vis = driver.execute_script("""
                var items = [];
                document.querySelectorAll('a,span,li,button').forEach(function(el){
                    var t = (el.innerText||el.textContent||'').trim();
                    if(t && t.length < 80 && el.offsetParent !== null) items.push(t);
                });
                return items;
            """)
            log.warning(f"    Visible items: {vis[:30]}")
        except: pass
        return False

    time.sleep(2)
    _dismiss_portal_popup(driver, log)

    log.info("    Waiting for AIS tab to open (insight.gov.in)...")
    ais_handle = None
    for _ in range(20):
        new_handles = set(driver.window_handles) - original_handles
        if new_handles:
            ais_handle = list(new_handles)[0]
            break
        _dismiss_portal_popup(driver, log)
        time.sleep(1)

    if ais_handle:
        driver.switch_to.window(ais_handle)
        log.info(f"    Switched to AIS tab: {driver.current_url}")

        try:
            driver.execute_script("""
                Object.defineProperty(document, 'visibilityState',
                    {get: () => 'visible', configurable: true});
                Object.defineProperty(document, 'hidden',
                    {get: () => false, configurable: true});
                document.dispatchEvent(new Event('visibilitychange'));
                window.dispatchEvent(new Event('focus'));
                window.dispatchEvent(new Event('pageshow'));
            """)
            log.info("    AIS tab focus events dispatched ✓")
        except Exception as fe:
            log.warning(f"    Tab focus dispatch failed (non-critical): {fe}")
    else:
        log.info(f"    AIS on same tab: {driver.current_url}")

    log.info("    Waiting for AIS JWT redirect to complete...")
    for i in range(30):
        cur_url = driver.current_url or ""
        if "access?param" not in cur_url and "insight.gov.in" in cur_url:
            log.info(f"    JWT redirect complete — AIS dashboard: {cur_url[:80]}")
            break
        if i % 5 == 0:
            log.info(f"    Still redirecting... ({i}s)  URL: {cur_url[:60]}")
        time.sleep(2)
    else:
        log.warning("    JWT redirect may not have completed — proceeding anyway")

    log.info("    Polling for 'Download AIS/TIS' button to appear (up to 90s)...")
    deadline = time.time() + 90
    download_btn_found = False
    while time.time() < deadline:
        try:
            btns = driver.find_elements(By.XPATH,
                "//*[contains(normalize-space(),'Download AIS') or "
                "contains(normalize-space(),'Download AIS/TIS')]")
            visible = [b for b in btns if b.is_displayed()]
            if visible:
                log.info(f"    Download AIS/TIS button visible ✓ — page fully loaded: '{visible[0].text.strip()}'")
                download_btn_found = True
                break
        except Exception:
            pass
        time.sleep(3)

    if not download_btn_found:
        log.warning("    'Download AIS/TIS' button did not appear within 90s")
        try:
            vis = driver.execute_script("""
                var items = [];
                document.querySelectorAll('button,a').forEach(function(el){
                    var t = (el.innerText||el.textContent||'').trim();
                    if(t && t.length < 80 && el.offsetParent !== null) items.push(t);
                });
                return items.slice(0,20);
            """)
            log.warning(f"    Visible buttons/links on AIS page: {vis}")
        except: pass

    log.info("    Selecting FY 2024-25 tab on AIS portal...")
    _select_ais_fy_tab(driver, log)
    time.sleep(SHORT_WAIT + 1)

    log.info(f"    AIS page URL: {driver.current_url}")
    return True


def _select_ais_fy_tab(driver, log):
    FY_LABELS = ["2024-25", "FY 2024-25", "FY2024-25", "2024 - 25",
                 "AY 2025-26", "AY2025-26", "2025-26"]
    try:
        for label in FY_LABELS:
            clicked = try_click(driver, [
                f"//*[contains(normalize-space(),'{label}') and ("
                f"self::button or self::a or self::li or "
                f"contains(@class,'tab') or contains(@role,'tab'))]",
                f"//*[@role='tab'][contains(normalize-space(),'{label}')]",
                f"//button[contains(normalize-space(),'{label}')]",
                f"//a[contains(normalize-space(),'{label}')]",
                f"//li[contains(normalize-space(),'{label}')]",
            ], timeout=4, log=log)
            if clicked:
                log.info(f"    AIS FY tab '{label}' selected ✓")
                time.sleep(SHORT_WAIT)
                return

        result = driver.execute_script("""
            var labels = ['2024-25','FY 2024-25','AY 2025-26','2025-26'];
            var roles  = ['tab','button','a','li'];
            for (var label of labels) {
                var all = document.querySelectorAll('[role="tab"],button,a,li');
                for (var el of all) {
                    var t = (el.innerText || el.textContent || '').trim();
                    if (t.includes(label) && el.offsetParent !== null) {
                        el.scrollIntoView({block:'center'});
                        el.click();
                        return 'clicked: ' + t;
                    }
                }
            }
            return 'not found';
        """)
        if result and "not found" not in str(result):
            log.info(f"    AIS FY tab selected via JS ✓ ({result})")
            time.sleep(SHORT_WAIT)
        else:
            log.warning("    AIS FY tab not found — proceeding with current selection")
    except Exception as e:
        log.warning(f"    _select_ais_fy_tab error: {e}")


def _click_popup_download_btn(driver, target, log):
    log.info(f"    Strategy 1: scoped modal search for {target.upper()} PDF button...")
    try:
        result = driver.execute_script("""
            var target = arguments[0];
            var modalSelectors = [
                '[role="dialog"]',
                '.modal', '.dialog', '.popup',
                '.cdk-overlay-pane', '.cdk-dialog-container',
                '[class*="modal"]', '[class*="dialog"]', '[class*="popup"]',
                '[class*="overlay-container"]', '[class*="download"]'
            ];
            var modal = null;
            for (var sel of modalSelectors) {
                var els = document.querySelectorAll(sel);
                for (var el of els) {
                    var btns = el.querySelectorAll('button');
                    var dlBtns = Array.from(btns).filter(function(b){
                        return (b.innerText||b.textContent||'').trim() === 'Download'
                               && b.offsetParent !== null;
                    });
                    if (dlBtns.length >= 3) {
                        modal = el;
                        break;
                    }
                }
                if (modal) break;
            }

            if (!modal) {
                return 'modal_not_found';
            }

            var rows = modal.querySelectorAll('tr, li, .row, [class*="row"], [class*="item"], [class*="list"]');
            if (rows.length === 0) {
                rows = Array.from(modal.children);
            }

            var aisKeywords  = ['annual information', 'ais'];
            var tisKeywords  = ['taxpayer information summary', 'tis'];
            var skipKeywords = ['json', 'utility'];

            for (var row of rows) {
                var txt = (row.innerText || row.textContent || '').toLowerCase();
                var btn = row.querySelector('button');
                if (!btn || (btn.innerText||btn.textContent||'').trim() !== 'Download') continue;
                if (!btn.offsetParent) continue;

                if (target === 'ais') {
                    var isAis = aisKeywords.some(function(k){ return txt.includes(k); });
                    var isJson = skipKeywords.some(function(k){ return txt.includes(k); });
                    var isTis = tisKeywords.some(function(k){ return txt.includes(k); });
                    if (isAis && !isJson && !isTis) {
                        btn.scrollIntoView({block:'center'});
                        btn.click();
                        return 'clicked_ais_by_text: ' + txt.substring(0,60);
                    }
                } else {
                    var isTisRow = tisKeywords.some(function(k){ return txt.includes(k); });
                    var isJsonRow = skipKeywords.some(function(k){ return txt.includes(k); });
                    if (isTisRow && !isJsonRow) {
                        btn.scrollIntoView({block:'center'});
                        btn.click();
                        return 'clicked_tis_by_text: ' + txt.substring(0,60);
                    }
                }
            }

            var allDlBtns = Array.from(modal.querySelectorAll('button')).filter(function(b){
                return (b.innerText||b.textContent||'').trim() === 'Download'
                       && b.offsetParent !== null;
            });

            var idx = (target === 'ais') ? 0 : 2;
            if (allDlBtns.length > idx) {
                allDlBtns[idx].scrollIntoView({block:'center'});
                allDlBtns[idx].click();
                return 'clicked_by_index_' + idx + ' (total=' + allDlBtns.length + ')';
            }

            return 'not_found (modal buttons=' + allDlBtns.length + ')';
        """, target)

        log.info(f"    Modal-scoped click result: {result}")
        if result and "not_found" not in str(result) and "modal_not_found" not in str(result):
            return True

    except Exception as e:
        log.warning(f"    Strategy 1 error: {e}")

    log.info(f"    Strategy 2: XPath scoped modal search for {target.upper()}...")
    if target == "ais":
        row_xpaths = [
            "//*[@role='dialog']//*[contains(normalize-space(),'Annual Information Statement') "
            "and contains(normalize-space(),'PDF') "
            "and not(contains(normalize-space(),'JSON')) "
            "and not(contains(normalize-space(),'Taxpayer'))]"
            "//button[normalize-space()='Download']",

            "//*[@role='dialog']//*[contains(normalize-space(),'Annual Information Statement') "
            "and contains(normalize-space(),'PDF')]"
            "/following-sibling::*//button[normalize-space()='Download']",

            "//*[contains(@class,'modal') or contains(@class,'dialog') or contains(@class,'popup')]"
            "//*[contains(normalize-space(),'Annual Information Statement') "
            "and contains(normalize-space(),'PDF') "
            "and not(contains(normalize-space(),'JSON'))]"
            "//button[normalize-space()='Download']",
        ]
    else:
        row_xpaths = [
            "//*[@role='dialog']//*[contains(normalize-space(),'Taxpayer Information Summary')]"
            "//button[normalize-space()='Download']",

            "//*[@role='dialog']//*[contains(normalize-space(),'Taxpayer Information Summary')]"
            "/following-sibling::*//button[normalize-space()='Download']",

            "//*[contains(@class,'modal') or contains(@class,'dialog') or contains(@class,'popup')]"
            "//*[contains(normalize-space(),'Taxpayer Information Summary')]"
            "//button[normalize-space()='Download']",
        ]

    clicked = try_click(driver, row_xpaths, timeout=5, log=log)
    if clicked:
        return True

    log.info(f"    Strategy 3: positional fallback (skip page button at [0])...")
    try:
        all_btns = driver.find_elements(By.XPATH,
            "//button[contains(normalize-space(),'Download')]")
        visible = [b for b in all_btns if b.is_displayed()]
        log.info(f"    Total visible Download buttons on page: {len(visible)}")

        popup_btns = visible[1:]
        log.info(f"    Popup-only buttons (after slicing [0]): {len(popup_btns)}")

        idx = 0 if target == "ais" else 2
        if len(popup_btns) > idx:
            btn = popup_btns[idx]
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", btn)
            log.info(f"    {target.upper()} clicked at popup_btns[{idx}] ✓")
            return True
        else:
            log.warning(f"    Not enough popup buttons ({len(popup_btns)}) for index {idx}")
    except Exception as e:
        log.warning(f"    Strategy 3 error: {e}")

    return False


def _open_download_popup(driver, log):
    log.info("    Opening 'Download AIS/TIS' popup...")
    popup_opened = try_click(driver, [
        "//button[contains(normalize-space(),'Download AIS/TIS (F.Y.')]",
        "//a[contains(normalize-space(),'Download AIS/TIS (F.Y.')]",
        "//button[contains(normalize-space(),'Download AIS/TIS (FY')]",
        "//a[contains(normalize-space(),'Download AIS/TIS (FY')]",
        "//button[contains(normalize-space(),'Download AIS/TIS')]",
        "//a[contains(normalize-space(),'Download AIS/TIS')]",
        "//button[contains(normalize-space(),'Download AIS')]",
        "//*[contains(normalize-space(),'Download AIS/TIS')]",
    ], timeout=15, log=log)

    if not popup_opened:
        log.warning("    'Download AIS/TIS' button not found — cannot open popup")
        return False

    log.info("    Waiting for popup to show 3 Download buttons...")
    for attempt in range(25):
        try:
            exact_btns = driver.find_elements(By.XPATH,
                "//button[normalize-space()='Download']")
            visible_exact = [b for b in exact_btns if b.is_displayed()]
            log.info(f"    Attempt {attempt+1}: {len(visible_exact)} exact 'Download' button(s) visible")
            if len(visible_exact) >= 3:
                log.info("    Popup fully loaded — 3 Download buttons visible ✓")
                time.sleep(0.3)
                return True
        except Exception:
            pass
        time.sleep(1)

    log.warning("    Popup did not show 3 Download buttons within 25s — proceeding anyway")
    return True


# ==========================================================
# PORTAL HELPERS
# ==========================================================
def _dismiss_portal_popup(driver, log):
    try:
        try:
            alert = driver.switch_to.alert
            log.warning(f"    Native alert detected: '{alert.text[:80]}' — dismissing...")
            alert.dismiss()
            time.sleep(0.5)
        except Exception:
            pass

        no_clicked = try_click(driver, [
            "//button[normalize-space()='No']",
            "//button[normalize-space()='NO']",
            "//button[contains(normalize-space(),'No')]",
            "//*[@role='dialog']//button[contains(normalize-space(),'No')]",
            "//mat-dialog-container//button[contains(normalize-space(),'No')]",
            "//*[contains(@class,'modal')]//button[contains(normalize-space(),'No')]",
            "//*[contains(@class,'dialog')]//button[contains(normalize-space(),'No')]",
            "//button[contains(normalize-space(),'Cancel')]",
            "//button[contains(normalize-space(),'Stay')]",
        ], timeout=3, log=None)

        if no_clicked:
            log.info("    Portal security popup dismissed (clicked 'No') ✓")
            time.sleep(1)
        return no_clicked
    except Exception as e:
        log.warning(f"    _dismiss_portal_popup error: {e}")
        return False


def _go_to_dashboard(driver, log):
    try:
        all_handles = driver.window_handles
        if len(all_handles) > 1:
            it_portal_handle = all_handles[0]
            for h in all_handles[1:]:
                try:
                    driver.switch_to.window(h)
                    log.info(f"    Closing extra tab: {driver.current_url}")
                    driver.close()
                except Exception as _ce:
                    log.warning(f"    Could not close tab: {_ce}")
            driver.switch_to.window(it_portal_handle)
            log.info("    Switched back to IT portal tab ✓")
            time.sleep(1)

        _dismiss_portal_popup(driver, log)

        driver.get(IT_DASHBOARD)
        time.sleep(SHORT_WAIT)

        _dismiss_portal_popup(driver, log)

        log.info(f"    Dashboard URL: {driver.current_url}")
    except Exception as e:
        log.warning(f"    _go_to_dashboard error: {e}")
        try:
            _dismiss_portal_popup(driver, log)
            try_click(driver, [
                "//a[contains(@href,'dashboard')]",
                "//a[normalize-space()='Home']",
                "//a[@title='Home']",
                "//img[contains(@class,'logo')]/..",
            ], timeout=5, log=log)
            time.sleep(SHORT_WAIT)
        except: pass


# ==========================================================
# LOAD CLIENTS
# ==========================================================
def load_it_clients(script_dir):
    clients = []

    for fname in ["Client_Manager_Secure_AY2025-26.xlsx","clients_manager.xlsx","clients.xlsx"]:
        p = os.path.join(script_dir, fname)
        if not os.path.exists(p): continue
        try:
            df = pd.read_excel(p, sheet_name="🔐 Client Credentials", header=2, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
            col = {}
            for c in df.columns:
                cl = c.lower().replace("\n","").strip()
                if "client name" in cl:  col["name"]        = c
                elif cl == "pan":        col["pan"]         = c
                elif "it_username" in cl or "it username" in cl: col["it_user"] = c
                elif "it_password" in cl or "it password" in cl: col["it_pass"] = c
                elif "it_active"   in cl or "it active"   in cl: col["it_active"] = c
                elif "gstin" in cl:      col["gstin"]       = c
                elif "entity" in cl:     col["entity"]      = c
                elif "username" in cl:   col["gst_user"]    = c
                elif "password" in cl:   col["gst_pass"]    = c
                elif "active" in cl:     col["active"]      = c

            for _, row in df.iterrows():
                name  = _cs(row.get(col.get("name",""),""))
                if not name or "sample" in name.lower(): continue

                it_active = str(row.get(col.get("it_active",""),"YES")).strip().upper()
                if it_active == "NO": continue

                pan = _cs(row.get(col.get("pan",""),""))
                if not pan:
                    gstin = _cs(row.get(col.get("gstin",""),""))
                    if gstin and len(gstin) >= 12:
                        pan = gstin[2:12]
                if not pan:
                    print(f"  ⚠  '{name}' has no PAN column — skipped for IT.")
                    continue

                it_user = (_cs(row.get(col.get("it_user",""),"")) or pan)
                it_pass = (_cs(row.get(col.get("it_pass",""),"")) or
                           _cs(row.get(col.get("gst_pass",""),"")))

                if not it_pass:
                    print(f"  ⚠  '{name}' has no IT password — skipped.")
                    continue

                clients.append({
                    "name":     name,
                    "pan":      pan.upper(),
                    "it_user":  it_user.upper(),
                    "it_pass":  it_pass,
                    "gstin":    _cs(row.get(col.get("gstin",""),"")),
                    "entity":   _cs(row.get(col.get("entity",""),"")),
                })

            if clients:
                print(f"  Loaded {len(clients)} IT client(s) from {fname}")
                return clients

        except Exception as e:
            print(f"  Excel read error ({fname}): {e}")

    for fname in ["clients.csv"]:
        p = os.path.join(script_dir, fname)
        if not os.path.exists(p): continue
        try:
            import csv
            with open(p, newline="", encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    row = {k.strip().lower(): v.strip() for k,v in row.items()}
                    name = row.get("name","") or row.get("client name","")
                    if not name or "sample" in name.lower(): continue
                    if row.get("it_active","yes").upper() == "NO": continue
                    pan  = row.get("pan","") or (row.get("gstin","")[2:12] if len(row.get("gstin",""))>=12 else "")
                    if not pan: continue
                    it_pass = row.get("it_password","") or row.get("password","")
                    if not it_pass: continue
                    clients.append({
                        "name":    name,
                        "pan":     pan.upper(),
                        "it_user": pan.upper(),
                        "it_pass": it_pass,
                        "gstin":   row.get("gstin",""),
                        "entity":  row.get("entity",""),
                    })
            if clients:
                print(f"  Loaded {len(clients)} IT client(s) from {fname}")
                return clients
        except Exception as e:
            print(f"  CSV read error ({fname}): {e}")

    print()
    print("  " + "="*60)
    print("  ✗  NO CLIENT FILE FOUND for IT Automation")
    print("  " + "-"*60)
    print("  Please add these columns to your clients.xlsx")
    print("  (sheet: 🔐 Client Credentials):")
    print()
    print("    PAN          — 10-character PAN  (e.g. AABCP1234C)")
    print("    IT_Password  — incometax.gov.in password")
    print("    IT_Active    — YES or NO  (optional, default YES)")
    print()
    print("  The IT_Username defaults to PAN (no separate column needed).")
    print("  IT_Password is separate from GST password — they may differ.")
    print("  " + "="*60)
    input("  Press Enter to close...")
    sys.exit(1)


def _cs(v):
    s = str(v).strip() if v is not None else ""
    return "" if s.lower() in ("nan","none","") else s


# ==========================================================
# PROCESS ONE CLIENT (with PDF unlock)
# ==========================================================
def process_it_client(client, base_dir, log):
    name    = client["name"]
    pan     = client["pan"]
    it_user = client["it_user"]
    it_pass = client["it_pass"]

    log.info(f"\n{'='*60}")
    log.info(f"IT CLIENT: {name}  |  PAN: {pan}")
    log.info(f"{'='*60}")

    safe = name.replace(" ","_").replace("/","_")
    cdir = Path(base_dir) / safe
    cdir.mkdir(parents=True, exist_ok=True)

    result = {
        "name":   name,
        "pan":    pan,
        "gstin":  client.get("gstin",""),
        "entity": client.get("entity",""),
        "26AS":   "SKIP",
        "AIS":    "SKIP",
        "TIS":    "SKIP",
        "recon":  "SKIP",
        "login_failed": False,
        "pdf_password": None,
    }

    driver = None
    profile_details = None
    
    try:
        driver = make_driver(str(cdir))

        print(f"\n  {'='*52}")
        print(f"  IT LOGIN: {name}  |  PAN: {pan}")
        print(f"  {'='*52}")
        print(f"  1. Browser will open incometax.gov.in")
        print(f"  2. PAN and password are auto-filled")
        print(f"  3. Enter OTP in browser when asked")
        print(f"  4. Click 'Yes'/'Register Device' when asked (one-time only)")
        print(f"  5. Script will auto-download 26AS, AIS, TIS")
        print(f"  6. PDFs will be automatically unlocked!")

        if not it_login(driver, it_user, it_pass, log):
            result["login_failed"] = True
            log.error(f"  SKIPPED {name} — IT login failed")
            print(f"\n  ✗  SKIPPED: {name} — login failed")
            return result

        # ── Extract Profile Details (PAN & DOB for PDF unlock) ──────
        # Navigate to My Profile → read PAN + DOB → build PDF password.
        # We do this RIGHT AFTER login while on dashboard, before any
        # downloads, so the session is clean.
        print(f"\n  [{name}] Extracting profile details for PDF unlock...")
        profile_details = get_profile_details(driver, pan_hint=pan, log=log)
        if profile_details:
            result["pdf_password"] = profile_details["pdf_password"]
            print(f"  ✓  Profile: PAN={profile_details['pan']}  DOB={profile_details['dob']}")
            print(f"  ✓  PDF Password: {profile_details['pdf_password']}")
        else:
            # Cannot auto-unlock without a confirmed DOB — warn and continue
            print(f"  ⚠  Could not extract DOB from profile — PDFs will stay locked")
            print(f"     (Open My Profile manually, check Date of Birth, re-run)")

        # Go back to dashboard after profile extraction
        _go_to_dashboard(driver, log)
        time.sleep(SHORT_WAIT)
        _dismiss_portal_popup(driver, log)
        time.sleep(1)

        # ── Download 26AS ──────────────────────────────────────────
        print(f"\n  [{name}] Downloading Form 26AS...")
        p26 = download_26as(driver, str(cdir), pan, log)
        result["26AS"] = "OK" if p26 else "FAIL"
        if p26:
            print(f"  ✓  26AS saved: {Path(p26).name}")
        else:
            print(f"  ✗  26AS download failed — you may download it manually")
            print(f"     Path: {cdir}")

        # ── Download AIS + TIS ─────────────────────────────────────
        print(f"\n  [{name}] Downloading AIS & TIS...")
        _go_to_dashboard(driver, log)
        time.sleep(SHORT_WAIT)
        _dismiss_portal_popup(driver, log)
        time.sleep(1)

        ais_p, tis_p = download_ais_tis(driver, str(cdir), pan, log)
        result["AIS"] = "OK" if ais_p else "FAIL"
        result["TIS"] = "OK" if tis_p else "FAIL"
        if ais_p: print(f"  ✓  AIS saved: {Path(ais_p).name}")
        else:     print(f"  ✗  AIS download failed")
        if tis_p: print(f"  ✓  TIS saved: {Path(tis_p).name}")
        else:     print(f"  ✗  TIS download failed")

        # ── Unlock PDFs ────────────────────────────────────────────
        if PIKEPDF_AVAILABLE and result["pdf_password"]:
            print(f"\n  [{name}] Unlocking PDFs...")
            unlocked = unlock_all_pdfs(str(cdir), result["pdf_password"], log)
            if unlocked:
                print(f"  ✓  {len(unlocked)} PDF(s) unlocked successfully")
            else:
                print(f"  ⚠  No PDFs were unlocked (may already be unlocked)")
        elif not PIKEPDF_AVAILABLE:
            print(f"\n  ⚠  pypdf not installed — PDFs remain password-protected")
            print(f"     Install with: pip install pypdf")

        # ── IT Reconciliation ─────────────────────────────────────
        print(f"\n  [{name}] Running IT Reconciliation...")
        recon_path = _run_it_recon(
            str(cdir), name, pan,
            client.get("gstin",""),
            FY_LABEL, log
        )
        result["recon"] = "OK" if recon_path else "FAIL"
        if recon_path:
            print(f"  ✓  IT Recon Excel: {Path(recon_path).name}")
        else:
            print(f"  ✗  IT Recon failed (PDFs may be missing)")

    except Exception as e:
        log.error(f"  Client error [{name}]: {e}")
        import traceback
        log.error(traceback.format_exc())
    finally:
        if driver:
            try: driver.quit()
            except: pass
        log.info(f"  Done: {name}. Waiting {CLIENT_GAP}s...")
        time.sleep(CLIENT_GAP)

    return result


def _run_it_recon(client_dir, name, pan, gstin, fy, log):
    script_folder = os.path.dirname(os.path.abspath(__file__))
    engine_paths = [
        os.path.join(script_folder, "it_recon_engine.py"),
        os.path.join(os.path.dirname(script_folder), "it_recon_engine.py"),
    ]

    engine_path = next((p for p in engine_paths if os.path.exists(p)), None)
    if not engine_path:
        log.warning("    it_recon_engine.py not found — skipping IT Recon")
        return None

    try:
        import importlib.util
        spec = importlib.util.spec_from_file_location("it_recon_engine", engine_path)
        engine = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(engine)

        def _log_fn(msg, t="info"):
            if t == "warn":   log.warning(f"    {msg}")
            elif t == "error": log.error(f"    {msg}")
            else:              log.info(f"    {msg}")

        out = engine.write_it_reconciliation(
            client_dir, name, pan, gstin, fy, log=_log_fn
        )
        return out
    except Exception as e:
        log.error(f"    IT Recon error: {e}")
        import traceback
        log.error(traceback.format_exc())
        return None


# ==========================================================
# MASTER REPORT
# ==========================================================
def write_it_master_report(all_results, base_dir):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl not installed — skipping master report")
        return None

    def _f(h): return PatternFill("solid", fgColor=h)
    def _fn(b=False,c="000000",s=10): return Font(name="Arial",bold=b,color=c,size=s)
    def _bd():
        x=Side(style="thin"); return Border(left=x,right=x,top=x,bottom=x)
    def _al(h="left"): return Alignment(horizontal=h,vertical="center")
    def _sc(cell,bold=False,fg="000000",bg=None,size=10,h="center"):
        cell.font=_fn(bold,fg,size)
        if bg: cell.fill=_f(bg)
        cell.alignment=_al(h); cell.border=_bd()

    wb = Workbook()
    ws = wb.active; ws.title = "IT Master Dashboard"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"INCOME TAX DOWNLOAD MASTER REPORT — AY {AY_LABEL} — {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
    _sc(c, bold=True, fg="FFFFFF", bg="1F3864", size=13, h="center")
    ws.row_dimensions[1].height = 36

    hdrs = ["Client","PAN","GSTIN","26AS","AIS","TIS","IT Recon","PDF Password"]
    wdts = [28, 14, 20, 10, 10, 10, 14, 25]
    for ci,(h,w) in enumerate(zip(hdrs,wdts),1):
        c = ws.cell(row=2,column=ci,value=h)
        _sc(c, bold=True, fg="FFFFFF", bg="2E75B6", size=10, h="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 26
    ws.freeze_panes = "A3"

    STATUS_COLORS = {
        "OK":   ("276221","C6EFCE"),
        "FAIL": ("9C0006","FFC7CE"),
        "SKIP": ("9C6500","FFEB9C"),
    }

    ri = 3
    for res in all_results:
        row_bg = "F2F2F2" if ri%2==0 else "FFFFFF"
        for ci, val in enumerate([
            res.get("name",""),
            res.get("pan",""),
            res.get("gstin",""),
        ],1):
            cl = ws.cell(row=ri,column=ci,value=val)
            _sc(cl, fg="000000", bg=row_bg, size=10, h="left")
        for ci, key in enumerate(["26AS","AIS","TIS","recon"],4):
            status = res.get(key,"SKIP")
            if res.get("login_failed"):
                status = "LOGIN FAIL"
                fg,bg = "9C0006","FFC7CE"
            else:
                fg,bg = STATUS_COLORS.get(status, ("000000","FFFFFF"))
            cl = ws.cell(row=ri,column=ci,value=status)
            _sc(cl, bold=True, fg=fg, bg=bg, size=10, h="center")
        # PDF Password column
        pwd = res.get("pdf_password", "")
        cl = ws.cell(row=ri,column=8,value=pwd)
        _sc(cl, fg="000000", bg=row_bg, size=10, h="left")
        ws.row_dimensions[ri].height = 20
        ri += 1

    ri += 1
    ok_counts = {k: sum(1 for r in all_results if r.get(k)=="OK") for k in ["26AS","AIS","TIS","recon"]}
    total = len(all_results)
    ws.merge_cells(f"A{ri}:C{ri}")
    sum_c = ws.cell(row=ri,column=1,value=f"TOTAL: {total} clients")
    _sc(sum_c, bold=True, fg="FFFFFF", bg="1F3864", size=10, h="left")
    for ci,key in enumerate(["26AS","AIS","TIS","recon"],4):
        cl = ws.cell(row=ri,column=ci,value=f"{ok_counts[key]}/{total} OK")
        _sc(cl, bold=True, fg="276221" if ok_counts[key]==total else "9C0006",
            bg="C6EFCE" if ok_counts[key]==total else "FFC7CE", size=10, h="center")
    ws.row_dimensions[ri].height = 22

    rp = os.path.join(base_dir, f"IT_MASTER_REPORT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    wb.save(rp)
    return rp


# ==========================================================
# ZIP ALL OUTPUTS
# ==========================================================
def zip_it_outputs(base_dir, log):
    import zipfile as _zf
    PATTERNS = ["26AS_*.pdf","AIS_*.pdf","TIS_*.pdf",
                "IT_RECONCILIATION_*.xlsx","IT_MASTER_REPORT_*.xlsx"]
    base_path = Path(base_dir)
    collected = []
    for pat in PATTERNS:
        for f in base_path.rglob(pat):
            if f.is_file() and f not in collected:
                collected.append(f)

    if not collected:
        log.info("  zip_it_outputs: no files to bundle")
        return None

    zip_name = f"IT_ALL_OUTPUTS_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
    zip_path = base_path / zip_name
    log.info(f"\n  Bundling {len(collected)} IT file(s) into {zip_name}...")
    with _zf.ZipFile(zip_path,"w",_zf.ZIP_DEFLATED) as zout:
        for f in sorted(collected):
            arcname = str(f.relative_to(base_path))
            zout.write(str(f), arcname)
            log.info(f"    + {arcname}")
    size_mb = zip_path.stat().st_size / (1024*1024)
    log.info(f"  ✓ ZIP: {zip_path.name}  ({size_mb:.1f} MB)")
    try:
        import subprocess
        subprocess.Popen(["explorer", "/select,", str(zip_path)])
    except: pass
    return str(zip_path)


# ==========================================================
# MENU
# ==========================================================
def ask_menu():
    print("\n" + "="*60)
    print("  INCOME TAX COMPLETE SUITE v3 — AY 2025-26 (with PDF Unlock)")
    print("  Downloads: Form 26AS + AIS + TIS (PDF)")
    print("  Auto-extracts profile & unlocks PDFs automatically!")
    print("="*60)
    print()
    print("  SELECT WHAT TO DO")
    print("  " + "-"*56)
    print("  [1]  ALL  — 26AS + AIS + TIS + IT Recon Excel + PDF Unlock")
    print("  [2]  26AS only")
    print("  [3]  AIS + TIS only")
    print("  [4]  IT Recon Excel only  (from already-downloaded PDFs)")
    print("  " + "-"*56)
    print()
    while True:
        c = input("  Enter choice [1/2/3/4]: ").strip()
        if c in ("1","2","3","4"): return c
        print("  Invalid choice. Enter 1, 2, 3, or 4.")


# ==========================================================
# MAIN
# ==========================================================
def main():
    print("\n" + "="*60)
    print("  INCOME TAX COMPLETE SUITE v3 — AY 2025-26")
    print("  Auto-downloads 26AS, AIS, TIS from incometax.gov.in")
    print("  Auto-extracts profile & unlocks PDFs automatically!")
    print("="*60)

    if MISSING:
        print(f"\n  Missing packages: pip install {' '.join(MISSING)}")
        input("  Press Enter..."); return

    choice = ask_menu()

    script_dir = os.path.dirname(os.path.abspath(__file__))

    if choice == "4":
        _run_offline_it_recon(script_dir)
        return

    clients = load_it_clients(script_dir)

    base_dir = os.path.join(
        os.path.expanduser("~"), "Downloads", "IT_Automation",
        f"AY{AY_LABEL}_{datetime.now().strftime('%Y%m%d_%H%M')}")
    os.makedirs(base_dir, exist_ok=True)
    print(f"\n  Output folder: {base_dir}")
    log = setup_logger(base_dir)
    print(f"  Clients loaded: {len(clients)}")

    print()
    print("  HOW IT WORKS:")
    print("  1. Browser opens → PAN + password auto-filled")
    print("  2. OTP arrives on mobile/email → type it in browser → ENTER")
    print("  3. Click 'Yes/Register Device' → no OTP needed next time ✓")
    print("  4. Script auto-extracts PAN & DOB from My Profile")
    print("  5. Script auto-downloads 26AS, AIS, TIS PDFs")
    print("  6. PDFs are automatically unlocked using profile password!")
    print()
    print("  NOTE: First login for each client needs OTP.")
    print("        After device is registered, future runs are fully automatic.")
    print()

    if input("  Type YES to start: ").strip().upper() != "YES":
        print("  Cancelled."); return

    all_results = []
    for i, client in enumerate(clients, 1):
        print(f"\n  [{i}/{len(clients)}] {client['name']}  |  PAN: {client['pan']}")
        result = process_it_client(client, base_dir, log)
        all_results.append(result)

    print("\n  Generating IT Master Report...")
    report = write_it_master_report(all_results, base_dir)

    print("  Bundling all output files...")
    zip_path = zip_it_outputs(base_dir, log)

    failed = [r["name"] for r in all_results if r.get("login_failed")]
    ok_26  = sum(1 for r in all_results if r.get("26AS")=="OK")
    ok_ais = sum(1 for r in all_results if r.get("AIS")=="OK")
    ok_tis = sum(1 for r in all_results if r.get("TIS")=="OK")
    ok_rec = sum(1 for r in all_results if r.get("recon")=="OK")
    total  = len(all_results)

    print()
    print("="*60)
    print("  IT AUTOMATION — ALL DONE!")
    print(f"  Clients : {total} processed")
    print(f"  26AS    : {ok_26}/{total} downloaded")
    print(f"  AIS     : {ok_ais}/{total} downloaded")
    print(f"  TIS     : {ok_tis}/{total} downloaded")
    print(f"  IT Recon: {ok_rec}/{total} Excel reports generated")
    if failed:
        print(f"  ✗ Login FAILED ({len(failed)}): {', '.join(failed)}")
        print("    → Fix IT credentials in clients.xlsx and re-run")
    else:
        print(f"  ✓ All {total} client(s) logged in successfully")
    if report:
        print(f"  Report  : {os.path.basename(report)}")
    if zip_path:
        print(f"  ZIP     : {os.path.basename(zip_path)}  ← all files here")
    print(f"  Folder  : {base_dir}")
    print("="*60)
    input("\n  Press Enter to close...")


def _run_offline_it_recon(script_dir):
    import glob as _glob

    print("\n" + "="*60)
    print("  OFFLINE IT RECONCILIATION — Already Downloaded PDFs")
    print("="*60)
    print()

    base = os.path.join(os.path.expanduser("~"), "Downloads", "IT_Automation")
    print(f"  Default folder: {base}")
    choice = input("  Press ENTER to use default, or type folder path: ").strip()
    if choice: base = choice

    if not os.path.isdir(base):
        print(f"  ✗ Folder not found: {base}"); return

    run_folders = sorted([
        Path(base)/d for d in os.listdir(base)
        if os.path.isdir(os.path.join(base,d)) and d.startswith("AY")
    ], reverse=True)
    if not run_folders:
        run_folders = [Path(base)]

    print(f"\n  Found {len(run_folders)} run folder(s):")
    for i,rf in enumerate(run_folders[:8],1):
        print(f"    [{i}] {rf.name}")
    idx = input("\n  Select folder (ENTER = most recent): ").strip()
    try:    sel = run_folders[int(idx)-1] if idx else run_folders[0]
    except: sel = run_folders[0]

    client_dirs = [d for d in sel.iterdir()
                   if d.is_dir() and any(d.glob("*.pdf"))]
    if not client_dirs:
        print(f"  ✗ No folders with PDF files found in {sel}"); return

    log_file = sel / f"it_offline_{datetime.now().strftime('%Y%m%d_%H%M')}.log"
    logging.basicConfig(level=logging.INFO,
        format="%(asctime)s | %(levelname)-8s | %(message)s",
        handlers=[logging.FileHandler(str(log_file),encoding="utf-8"),
                  logging.StreamHandler(sys.stdout)])
    log = logging.getLogger("it_offline")

    clients = load_it_clients(script_dir)
    pan_lookup = {c["name"].replace(" ","_"): c for c in clients}

    print(f"\n  Found {len(client_dirs)} client folder(s) with PDFs:")
    generated = 0
    for cdir in client_dirs:
        client = pan_lookup.get(cdir.name)
        if not client:
            for k,v in pan_lookup.items():
                if k.lower() in cdir.name.lower() or cdir.name.lower() in k.lower():
                    client = v; break
        if not client:
            print(f"  ⚠  No matching client found for folder: {cdir.name} — skipping")
            continue

        name  = client["name"]
        pan   = client["pan"]
        gstin = client.get("gstin","")
        print(f"\n  Processing: {name}  (PAN: {pan})")

        pdfs = list(cdir.glob("*.pdf"))
        print(f"    PDFs found: {[p.name for p in pdfs]}")

        rp = _run_it_recon(str(cdir), name, pan, gstin, FY_LABEL, log)
        if rp:
            print(f"  ✓ Saved: {Path(rp).name}")
            generated += 1
        else:
            print(f"  ✗ Recon failed for {name}")

    print(f"\n  DONE — {generated}/{len(client_dirs)} IT Recon reports generated")
    print(f"  Folder: {sel}")
    input("\n  Press Enter to close...")


if __name__ == "__main__":
    main()
