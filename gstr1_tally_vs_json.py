"""
gstr1_tally_vs_json.py  — v1.0
═══════════════════════════════════════════════════════════════════════════════
Compare Tally GSTR-1 Sales Register vs GST Portal JSON (gstr1_fy_v5 output)

USAGE:
  python gstr1_tally_vs_json.py
      --tally   "Tally_2424-25_AP.xls"
      --json    "path/to/GSTR1_ZIPs_folder"   (or GSTR1_FY2025-26_*.xlsx)
      --name    "ARUN ENTERPRISES"
      --fy      2024-25

OUTPUT:
  GSTR1_TALLY_VS_JSON_<name>_FY<fy>_<ts>.xlsx
  Sheets:
    01_SUMMARY        — Month-wise totals: Tally vs JSON side by side + diff
    02_B2B_MATCH      — Invoice-level B2B: matched / amount diff / missing
    03_B2B_TALLY_ONLY — B2B invoices in Tally but NOT in JSON (missing from portal)
    04_B2B_JSON_ONLY  — B2B invoices in JSON but NOT in Tally (portal extras)
    05_B2CS_COMPARE   — B2CS month+rate comparison
    06_RATE_WISE      — Rate-wise taxable value: Tally vs JSON
═══════════════════════════════════════════════════════════════════════════════
"""

import os, sys, json, re, zipfile, glob, argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter

# ── Colours ──────────────────────────────────────────────────────────────────
C_HEADER   = "1F4E79"   # dark blue  — section headers
C_MATCH    = "E2EFDA"   # light green — matched rows
C_DIFF     = "FFF2CC"   # yellow      — amount difference
C_TONLY    = "FCE4D6"   # orange-red  — Tally only (missing from JSON)
C_JONLY    = "DDEBF7"   # light blue  — JSON only (extra in portal)
C_TOTAL    = "BDD7EE"   # blue        — totals
C_SUBHDR   = "D6DCE4"   # grey        — sub-headers
C_OK       = "00B050"   # green text  — ✓
C_ERR      = "FF0000"   # red text    — ✗

MONTH_ORDER = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]
MONTH_MAP   = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
               7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

# ── Helper: apply fill ────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr_font(bold=True, white=True, size=10):
    return Font(bold=bold, color="FFFFFF" if white else "000000", size=size)

def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header_row(ws, row_num, values, bg=C_HEADER, white_text=True, bold=True):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.fill = fill(bg)
        c.font = Font(bold=bold, color="FFFFFF" if white_text else "000000", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_thin()

def write_data_row(ws, row_num, values, bg=None, num_cols=None, bold=False):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        if bg:
            c.fill = fill(bg)
        c.border = border_thin()
        c.font   = Font(bold=bold, size=10)
        if num_cols and col in num_cols:
            c.alignment = Alignment(horizontal="right")
            if isinstance(val, (int, float)) and not pd.isna(val):
                c.number_format = '#,##0.00'
        else:
            c.alignment = Alignment(horizontal="left", wrap_text=False)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ═══════════════════════════════════════════════════════════════════════════════
# PART 1 — READ TALLY XLS
# ═══════════════════════════════════════════════════════════════════════════════
def read_tally(path):
    """
    Returns dict with:
      b2b   : list of dicts  {month, gstin, inv_no, date, gross_total,
                               tv_28, cgst_14, sgst_14,
                               tv_18, cgst_9,  sgst_9,
                               igst_28, igst_18, igst_12,
                               tv_12, cgst_6, sgst_6,
                               sez, nil, exempt, taxable_total}
      b2c   : list of dicts  (same but no gstin)
      monthly: {month -> {tv, cgst, sgst, igst, total_tax, gross}}
    """
    df = pd.read_excel(path, sheet_name="Sales Register",
                       header=10, engine="openpyxl")
    df = df.iloc[:-1]  # drop Grand Total
    df.columns = [str(c).strip() for c in df.columns]

    # Rename columns for easy access
    rename = {
        "Date":                 "date",
        "Particulars":          "party",
        "Voucher Type":         "vtype",
        "Voucher No.":          "inv_no",
        "GSTIN/UIN":            "gstin",
        "Value":                "value",
        "Gross Total":          "gross",
        "GST SALES @ 28%":      "tv_28",
        "CGST OUTWARD @ 14%":   "cgst_14",
        "SGST OUTWARD @ 14%":   "sgst_14",
        "GST SALES @ 18%":      "tv_18",
        "CGST OUTWARD @ 9%":    "cgst_9",
        "SGST OUTWARD @ 9%":    "sgst_9",
        "IGST SALES @ 28%":     "igst_tv_28",
        "IGST OUTWARD @ 28%":   "igst_28",
        "IGST SALES @ 18%":     "igst_tv_18",
        "IGST OUTWARD @ 18%":   "igst_18",
        "SALES GST":            "sales_gst",
        "GST EXEMPTED SALE @ 0%": "exempt",
        "GST SALES @12%":       "tv_12",
        "CGST OUTWARD6%":       "cgst_6",
        "SGST OUTWARD@6%":      "sgst_6",
        "IGST SALES @ 12%":     "igst_tv_12",
        "IGST OUTWARD @12%":    "igst_12",
        "SEZ SALES @ 0%":       "sez",
        "NILL SALE":            "nil",
    }
    df = df.rename(columns={k:v for k,v in rename.items() if k in df.columns})

    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].notna()]  # drop rows with no date

    # Skip cancelled vouchers
    df = df[~df["party"].astype(str).str.lower().str.contains("cancel")]

    # Numeric fill
    num_cols = [c for c in df.columns if c not in ("date","party","vtype","inv_no","gstin")]
    for c in num_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    df["month"] = df["date"].dt.month.map(MONTH_MAP)

    # Compute taxable total per row
    df["taxable_total"] = (df.get("tv_28",0) + df.get("tv_18",0) + df.get("tv_12",0) +
                           df.get("igst_tv_28",0) + df.get("igst_tv_18",0) + df.get("igst_tv_12",0) +
                           df.get("exempt",0) + df.get("nil",0) + df.get("sez",0))
    df["total_cgst"] = df.get("cgst_14",0) + df.get("cgst_9",0) + df.get("cgst_6",0)
    df["total_sgst"] = df.get("sgst_14",0) + df.get("sgst_9",0) + df.get("sgst_6",0)
    df["total_igst"] = df.get("igst_28",0) + df.get("igst_18",0) + df.get("igst_12",0)
    df["total_tax"]  = df["total_cgst"] + df["total_sgst"] + df["total_igst"]

    b2b = df[df["gstin"].notna() & (df["gstin"].astype(str).str.strip() != "")]
    b2c = df[df["gstin"].isna() | (df["gstin"].astype(str).str.strip() == "")]

    return df, b2b, b2c


# ═══════════════════════════════════════════════════════════════════════════════
# PART 2 — READ GSTR1 JSON (from ZIPs or already-extracted Excel)
# ═══════════════════════════════════════════════════════════════════════════════
def load_json_from_zip(zip_path):
    with zipfile.ZipFile(zip_path) as zf:
        jsons = [n for n in zf.namelist() if n.endswith(".json")]
        if not jsons:
            return None
        with zf.open(jsons[0]) as f:
            return json.load(f)

def parse_period(period_str):
    """'042024' or 'Apr' or '04/2024' -> month name"""
    s = str(period_str).strip()
    if len(s) == 6 and s.isdigit():
        return MONTH_MAP.get(int(s[:2]), s)
    if len(s) == 3:
        return s.title()
    # Try MM/YYYY
    m = re.match(r"(\d{1,2})[/-](\d{4})", s)
    if m:
        return MONTH_MAP.get(int(m.group(1)), s)
    return s

def read_json_folder(folder_path):
    """
    Read all GSTR1 ZIP files from folder. Returns:
      b2b_rows : list of dicts per invoice-item
      b2cs_rows: list of dicts per B2CS record
      monthly  : {month -> {tv, cgst, sgst, igst, total_tax}}
    """
    folder = Path(folder_path)
    zips = sorted(folder.glob("GSTR1_*.zip")) or sorted(folder.glob("*.zip"))
    if not zips:
        # Try subfolders
        zips = sorted(folder.rglob("GSTR1_*.zip"))

    if not zips:
        print(f"  ✗ No ZIP files found in {folder}")
        return [], [], {}

    b2b_rows  = []
    b2cs_rows = []
    monthly   = {}

    seen_periods = set()
    for zp in zips:
        data = load_json_from_zip(zp)
        if not data:
            continue
        period = parse_period(data.get("fp", zp.stem[:6]))
        if period in seen_periods:
            continue
        seen_periods.add(period)

        # B2B
        for ent in data.get("b2b", []):
            gstin = ent.get("ctin", "")
            for inv in ent.get("inv", []):
                inv_no  = inv.get("inum","")
                inv_dt  = inv.get("idt","")
                inv_val = float(inv.get("val", 0) or 0)
                inv_typ = inv.get("inv_typ","R")
                for it in inv.get("itms", []):
                    det = it.get("itm_det", {})
                    b2b_rows.append({
                        "month":     period,
                        "gstin":     gstin,
                        "inv_no":    inv_no,
                        "inv_date":  inv_dt,
                        "inv_type":  inv_typ,
                        "rate":      float(det.get("rt", 0) or 0),
                        "taxable":   float(det.get("txval", 0) or 0),
                        "igst":      float(det.get("iamt", 0) or 0),
                        "cgst":      float(det.get("camt", 0) or 0),
                        "sgst":      float(det.get("samt", 0) or 0),
                        "inv_val":   inv_val,
                    })

        # B2CS
        for rec in data.get("b2cs", []):
            b2cs_rows.append({
                "month":   period,
                "type":    rec.get("typ",""),
                "pos":     rec.get("pos",""),
                "rate":    float(rec.get("rt", 0) or 0),
                "taxable": float(rec.get("txval", 0) or 0),
                "igst":    float(rec.get("iamt", 0) or 0),
                "cgst":    float(rec.get("camt", 0) or 0),
                "sgst":    float(rec.get("samt", 0) or 0),
            })

    return b2b_rows, b2cs_rows


# ═══════════════════════════════════════════════════════════════════════════════
# PART 3 — BUILD COMPARISON
# ═══════════════════════════════════════════════════════════════════════════════
def normalize_inv(s):
    """Normalize invoice number for matching.
    Handles: AP/24-25/10001 → AP242510001, AP-24-25-10001 → AP242510001
    Also tries numeric suffix only for fallback.
    """
    s = str(s).upper().strip()
    # Remove all separators
    return re.sub(r"[\s\-/\\]", "", s)

def build_comparison(tally_path, json_path, name, fy):
    print(f"\n  Reading Tally: {tally_path}")
    df_all, t_b2b, t_b2c = read_tally(tally_path)

    print(f"  Reading JSON ZIPs: {json_path}")
    j_b2b_rows, j_b2cs_rows = read_json_folder(json_path)

    # Convert to DataFrames
    jb2b = pd.DataFrame(j_b2b_rows) if j_b2b_rows else pd.DataFrame(
        columns=["month","gstin","inv_no","inv_date","rate","taxable","igst","cgst","sgst","inv_val"])
    jb2cs = pd.DataFrame(j_b2cs_rows) if j_b2cs_rows else pd.DataFrame(
        columns=["month","type","pos","rate","taxable","igst","cgst","sgst"])

    # ── SUMMARY: month-wise totals ────────────────────────────────────────────
    months = MONTH_ORDER

    def month_totals(df, tax_col="total_tax", tv_col="taxable_total",
                     cgst_col="total_cgst", sgst_col="total_sgst", igst_col="total_igst"):
        result = {}
        for m in months:
            sub = df[df["month"] == m]
            result[m] = {
                "taxable": sub[tv_col].sum() if tv_col in sub.columns else 0,
                "cgst":    sub[cgst_col].sum() if cgst_col in sub.columns else 0,
                "sgst":    sub[sgst_col].sum() if sgst_col in sub.columns else 0,
                "igst":    sub[igst_col].sum() if igst_col in sub.columns else 0,
                "tax":     sub[tax_col].sum() if tax_col in sub.columns else 0,
            }
        return result

    def json_month_totals(df):
        result = {}
        for m in months:
            sub = df[df["month"] == m] if len(df) else pd.DataFrame()
            result[m] = {
                "taxable": sub["taxable"].sum() if len(sub) else 0,
                "cgst":    sub["cgst"].sum() if len(sub) else 0,
                "sgst":    sub["sgst"].sum() if len(sub) else 0,
                "igst":    sub["igst"].sum() if len(sub) else 0,
                "tax":     (sub["cgst"].sum()+sub["sgst"].sum()+sub["igst"].sum()) if len(sub) else 0,
            }
        return result

    t_monthly = month_totals(df_all)
    j_b2b_monthly  = json_month_totals(jb2b)
    j_b2cs_monthly = json_month_totals(jb2cs)

    def j_monthly(m):
        b = j_b2b_monthly.get(m, {})
        c = j_b2cs_monthly.get(m, {})
        return {k: b.get(k,0)+c.get(k,0) for k in ["taxable","cgst","sgst","igst","tax"]}

    # ── B2B invoice matching ──────────────────────────────────────────────────
    # Group JSON B2B by invoice (sum items per invoice)
    if len(jb2b):
        j_inv = jb2b.groupby(["month","gstin","inv_no"]).agg(
            taxable=("taxable","sum"), igst=("igst","sum"),
            cgst=("cgst","sum"), sgst=("sgst","sum"),
            inv_val=("inv_val","first")
        ).reset_index()
        j_inv["key"] = j_inv["gstin"].str.upper().str.strip() + "||" + j_inv["inv_no"].apply(normalize_inv)
    else:
        j_inv = pd.DataFrame(columns=["month","gstin","inv_no","taxable","igst","cgst","sgst","inv_val","key"])

    # Tally B2B per invoice (sum rates)
    t_b2b = t_b2b.copy()
    t_b2b["key"] = t_b2b["gstin"].astype(str).str.upper().str.strip() + "||" + t_b2b["inv_no"].apply(normalize_inv)
    t_inv = t_b2b.groupby(["month","key","gstin","inv_no"]).agg(
        taxable=("taxable_total","sum"),
        igst=("total_igst","sum"),
        cgst=("total_cgst","sum"),
        sgst=("total_sgst","sum"),
        gross=("gross","sum"),
    ).reset_index()

    j_keys = set(j_inv["key"].tolist()) if len(j_inv) else set()
    t_keys = set(t_inv["key"].tolist())

    matched_keys = t_keys & j_keys

    # Debug: if no matches, show sample keys from both sides
    if not matched_keys and t_keys and j_keys:
        print("  ⚠  0 matches found. Checking key format...")
        t_sample = sorted(t_keys)[:3]
        j_sample = sorted(j_keys)[:3]
        print(f"  Tally keys (sample): {t_sample}")
        print(f"  JSON  keys (sample): {j_sample}")

        # Fallback: try matching on invoice number ONLY (ignore GSTIN prefix)
        # Some ZIPs use buyer GSTIN differently
        def inv_only(key):
            return key.split("||")[-1] if "||" in key else key

        t_inv_map = {inv_only(k): k for k in t_keys}
        j_inv_map = {inv_only(k): k for k in j_keys}
        common_inv = set(t_inv_map) & set(j_inv_map)

        if common_inv:
            print(f"  ✓ Found {len(common_inv)} matches by invoice number only (GSTIN prefix differs)")
            # Rebuild keys using invoice-only match
            matched_keys = set()
            for inv in common_inv:
                tk = t_inv_map[inv]; jk = j_inv_map[inv]
                # Add to j_inv with tally key so lookup works
                matched_keys.add(tk)
                if tk != jk:
                    j_inv.loc[j_inv["key"]==jk, "key"] = tk
            j_keys = set(j_inv["key"].tolist())
        else:
            print("  ✗ Still no matches — invoice numbers may differ between Tally and JSON")

    tonly_keys   = t_keys - (j_keys if not matched_keys else set(j_inv["key"].tolist()))
    jonly_keys   = (j_keys if not matched_keys else set(j_inv["key"].tolist())) - t_keys
    # Recalculate after possible key remap
    matched_keys = t_keys & set(j_inv["key"].tolist())
    tonly_keys   = t_keys - matched_keys
    jonly_keys   = set(j_inv["key"].tolist()) - t_keys

    # Build matched comparison rows
    matched_rows = []
    for key in sorted(matched_keys):
        tr = t_inv[t_inv["key"]==key].iloc[0]
        jr = j_inv[j_inv["key"]==key].iloc[0]
        diff_tv  = round(tr["taxable"] - jr["taxable"], 2)
        diff_tax = round((tr["igst"]+tr["cgst"]+tr["sgst"]) - (jr["igst"]+jr["cgst"]+jr["sgst"]), 2)
        status   = "✓ MATCH" if abs(diff_tv) < 1 and abs(diff_tax) < 1 else "⚠ AMT DIFF"
        matched_rows.append({
            "Month":          tr["month"],
            "GSTIN":          tr["gstin"],
            "Invoice No.":    tr["inv_no"],
            "Status":         status,
            "Tally Taxable":  tr["taxable"],
            "JSON Taxable":   jr["taxable"],
            "Diff Taxable":   diff_tv,
            "Tally IGST":     tr["igst"],
            "JSON IGST":      jr["igst"],
            "Tally CGST":     tr["cgst"],
            "JSON CGST":      jr["cgst"],
            "Tally SGST":     tr["sgst"],
            "JSON SGST":      jr["sgst"],
            "Diff Tax":       diff_tax,
        })

    tonly_rows = []
    for key in sorted(tonly_keys):
        r = t_inv[t_inv["key"]==key].iloc[0]
        tonly_rows.append({
            "Month":       r["month"],
            "GSTIN":       r["gstin"],
            "Invoice No.": r["inv_no"],
            "Taxable":     r["taxable"],
            "IGST":        r["igst"],
            "CGST":        r["cgst"],
            "SGST":        r["sgst"],
            "Gross Total": r["gross"],
            "Remark":      "In Tally, NOT in GST Portal JSON",
        })

    jonly_rows = []
    for key in sorted(jonly_keys):
        r = j_inv[j_inv["key"]==key].iloc[0]
        jonly_rows.append({
            "Month":       r["month"],
            "GSTIN":       r["gstin"],
            "Invoice No.": r["inv_no"],
            "Taxable":     r["taxable"],
            "IGST":        r["igst"],
            "CGST":        r["cgst"],
            "SGST":        r["sgst"],
            "Inv Value":   r["inv_val"],
            "Remark":      "In GST Portal JSON, NOT in Tally",
        })

    return {
        "months":        months,
        "t_monthly":     t_monthly,
        "j_monthly_fn":  j_monthly,
        "matched_rows":  matched_rows,
        "tonly_rows":    tonly_rows,
        "jonly_rows":    jonly_rows,
        "jb2cs":         jb2cs,
        "t_b2c":         t_b2c,
        "t_b2b":         t_b2b,
        "jb2b":          jb2b,
        "df_all":        df_all,
        "name":          name,
        "fy":            fy,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# PART 4 — WRITE EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
def write_excel(data, out_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    name = data["name"]
    fy   = data["fy"]
    months = data["months"]

    # ── Sheet 1: SUMMARY ─────────────────────────────────────────────────────
    ws = wb.create_sheet("01_SUMMARY")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = f"GSTR-1 COMPARISON — Tally vs GST Portal JSON  |  {name}  |  FY {fy}"
    c.fill  = fill(C_HEADER)
    c.font  = Font(bold=True, color="FFFFFF", size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")

    hdr = ["Month",
           "Tally Taxable","JSON Taxable","Diff Taxable",
           "Tally CGST","JSON CGST","Diff CGST",
           "Tally SGST","JSON SGST","Diff SGST",
           "Tally IGST","JSON IGST","Diff IGST",
           "Status"]
    apply_header_row(ws, 2, hdr)
    set_col_widths(ws, [10,16,16,14,14,14,12,14,14,12,14,14,12,12])
    freeze(ws, "B3")

    NUM = set(range(2, 14))
    r = 3
    ann = {k: 0.0 for k in ["t_tv","j_tv","t_cg","j_cg","t_sg","j_sg","t_ig","j_ig"]}

    for m in months:
        tm = data["t_monthly"].get(m, {})
        jm = data["j_monthly_fn"](m)
        t_tv=tm.get("taxable",0); j_tv=jm.get("taxable",0)
        t_cg=tm.get("cgst",0);    j_cg=jm.get("cgst",0)
        t_sg=tm.get("sgst",0);    j_sg=jm.get("sgst",0)
        t_ig=tm.get("igst",0);    j_ig=jm.get("igst",0)
        d_tv=round(t_tv-j_tv,2); d_cg=round(t_cg-j_cg,2)
        d_sg=round(t_sg-j_sg,2); d_ig=round(t_ig-j_ig,2)
        ok = all(abs(x)<1 for x in [d_tv,d_cg,d_sg,d_ig])
        status = "✓ OK" if ok else ("⚠ DIFF" if max(abs(d_tv),abs(d_cg),abs(d_sg),abs(d_ig))<5000 else "✗ CHECK")
        row_vals = [m, t_tv,j_tv,d_tv, t_cg,j_cg,d_cg, t_sg,j_sg,d_sg, t_ig,j_ig,d_ig, status]
        bg = C_MATCH if ok else (C_DIFF if "DIFF" in status else C_TONLY)
        write_data_row(ws, r, row_vals, bg=bg if not ok else None, num_cols=NUM)
        # Color the diff cells
        for col_idx in [4,7,10,13]:
            cell = ws.cell(row=r, column=col_idx)
            val  = row_vals[col_idx-1]
            if abs(val) >= 1:
                cell.fill = fill(C_TONLY if val > 0 else C_JONLY)
                cell.font = Font(bold=True, size=10, color=C_ERR if abs(val)>5000 else "000000")
        # Status cell color
        sc = ws.cell(row=r, column=14)
        if ok:
            sc.font = Font(bold=True, size=10, color=C_OK)
        elif "CHECK" in status:
            sc.font = Font(bold=True, size=10, color=C_ERR)
        else:
            sc.font = Font(bold=True, size=10)
        for k,v in [("t_tv",t_tv),("j_tv",j_tv),("t_cg",t_cg),("j_cg",j_cg),
                    ("t_sg",t_sg),("j_sg",j_sg),("t_ig",t_ig),("j_ig",j_ig)]:
            ann[k] += v
        r += 1

    # Annual total row
    d_tv=round(ann["t_tv"]-ann["j_tv"],2)
    d_cg=round(ann["t_cg"]-ann["j_cg"],2)
    d_sg=round(ann["t_sg"]-ann["j_sg"],2)
    d_ig=round(ann["t_ig"]-ann["j_ig"],2)
    ok_all = all(abs(x)<1 for x in [d_tv,d_cg,d_sg,d_ig])
    totrow = ["★ FY TOTAL",
              ann["t_tv"],ann["j_tv"],d_tv,
              ann["t_cg"],ann["j_cg"],d_cg,
              ann["t_sg"],ann["j_sg"],d_sg,
              ann["t_ig"],ann["j_ig"],d_ig,
              "✓ MATCHED" if ok_all else "✗ MISMATCH"]
    write_data_row(ws, r, totrow, bg=C_TOTAL, num_cols=NUM, bold=True)

    # ── Sheet 2: B2B MATCHED ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("02_B2B_MATCH")
    ws2.merge_cells("A1:N1")
    c = ws2["A1"]
    c.value = f"B2B Invoice Match — {name}  FY {fy}   ({len(data['matched_rows'])} invoices)"
    c.fill  = fill(C_HEADER); c.font = hdr_font(); c.alignment = Alignment(horizontal="center",vertical="center")
    hdr2 = ["Month","GSTIN","Invoice No.","Status",
            "Tally Taxable","JSON Taxable","Diff Taxable",
            "Tally IGST","JSON IGST","Tally CGST","JSON CGST",
            "Tally SGST","JSON SGST","Diff Tax"]
    apply_header_row(ws2, 2, hdr2, bg=C_SUBHDR, white_text=False)
    set_col_widths(ws2, [10,20,18,12,15,15,13,13,13,13,13,13,13,13])
    freeze(ws2, "D3")
    NUM2 = set(range(5, 15))
    for i, row in enumerate(sorted(data["matched_rows"], key=lambda x:(MONTH_ORDER.index(x["Month"]) if x["Month"] in MONTH_ORDER else 99, x["GSTIN"], x["Invoice No."])), 3):
        ok_row = row["Status"] == "✓ MATCH"
        vals = [row["Month"],row["GSTIN"],row["Invoice No."],row["Status"],
                row["Tally Taxable"],row["JSON Taxable"],row["Diff Taxable"],
                row["Tally IGST"],row["JSON IGST"],row["Tally CGST"],row["JSON CGST"],
                row["Tally SGST"],row["JSON SGST"],row["Diff Tax"]]
        write_data_row(ws2, i, vals, bg=C_MATCH if ok_row else C_DIFF, num_cols=NUM2)
        sc = ws2.cell(row=i, column=4)
        sc.font = Font(bold=True, size=10, color=C_OK if ok_row else C_ERR)

    # ── Sheet 3: TALLY ONLY (missing from portal) ─────────────────────────────
    ws3 = wb.create_sheet("03_B2B_TALLY_ONLY")
    ws3.merge_cells("A1:I1")
    c = ws3["A1"]
    c.value = f"B2B In Tally, NOT in GST Portal JSON — {name}  ({len(data['tonly_rows'])} invoices)  ← Need to upload/check"
    c.fill  = fill(C_TONLY[:6] if len(C_TONLY)==6 else "FF6B35")
    c.fill  = fill("C55A11"); c.font = hdr_font()
    c.alignment = Alignment(horizontal="center",vertical="center")
    hdr3 = ["Month","GSTIN","Invoice No.","Taxable Value","IGST","CGST","SGST","Gross Total","Remark"]
    apply_header_row(ws3, 2, hdr3, bg="C55A11")
    set_col_widths(ws3, [10,20,18,16,14,14,14,16,35])
    NUM3 = {4,5,6,7,8}
    for i, row in enumerate(sorted(data["tonly_rows"], key=lambda x:(MONTH_ORDER.index(x["Month"]) if x["Month"] in MONTH_ORDER else 99, x["Invoice No."])), 3):
        vals = [row["Month"],row["GSTIN"],row["Invoice No."],row["Taxable"],
                row["IGST"],row["CGST"],row["SGST"],row["Gross Total"],row["Remark"]]
        write_data_row(ws3, i, vals, bg="FCE4D6", num_cols=NUM3)

    # ── Sheet 4: JSON ONLY (extra in portal) ──────────────────────────────────
    ws4 = wb.create_sheet("04_B2B_JSON_ONLY")
    ws4.merge_cells("A1:I1")
    c = ws4["A1"]
    c.value = f"B2B In GST Portal JSON, NOT in Tally — {name}  ({len(data['jonly_rows'])} invoices)  ← Check if missing in books"
    c.fill  = fill("1F4E79"); c.font = hdr_font()
    c.alignment = Alignment(horizontal="center",vertical="center")
    hdr4 = ["Month","GSTIN","Invoice No.","Taxable Value","IGST","CGST","SGST","Inv Value","Remark"]
    apply_header_row(ws4, 2, hdr4)
    set_col_widths(ws4, [10,20,18,16,14,14,14,16,35])
    NUM4 = {4,5,6,7,8}
    for i, row in enumerate(sorted(data["jonly_rows"], key=lambda x:(MONTH_ORDER.index(x["Month"]) if x["Month"] in MONTH_ORDER else 99, x["Invoice No."])), 3):
        vals = [row["Month"],row["GSTIN"],row["Invoice No."],row["Taxable"],
                row["IGST"],row["CGST"],row["SGST"],row["Inv Value"],row["Remark"]]
        write_data_row(ws4, i, vals, bg="DDEBF7", num_cols=NUM4)

    # ── Sheet 5: B2CS COMPARE ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("05_B2CS_COMPARE")
    ws5.merge_cells("A1:K1")
    c = ws5["A1"]
    c.value = f"B2CS Comparison (Unregistered Sales) — {name}  FY {fy}"
    c.fill = fill(C_HEADER); c.font = hdr_font()
    c.alignment = Alignment(horizontal="center",vertical="center")
    hdr5 = ["Month","Rate %","Tally Taxable","JSON Taxable","Diff Taxable",
            "Tally CGST","JSON CGST","Tally SGST","JSON SGST","Tally IGST","JSON IGST"]
    apply_header_row(ws5, 2, hdr5, bg=C_SUBHDR, white_text=False)
    set_col_widths(ws5, [10,8,16,16,14,14,14,14,14,14,14])
    NUM5 = set(range(3,12))

    # Tally B2C by month+rate
    t_b2c_copy = data["t_b2c"].copy()
    t_b2c_copy["rate"] = 0.0  # Tally doesn't split by rate cleanly in B2C
    # Use known rates from column names
    rates = [28, 18, 12]
    r5 = 3
    for m in months:
        tm_b2c = t_b2c_copy[t_b2c_copy["month"]==m]
        jm_b2cs = data["jb2cs"][data["jb2cs"]["month"]==m] if len(data["jb2cs"]) else pd.DataFrame()
        for rate in rates:
            # Tally: sum rows with that rate's columns
            if rate == 28:
                t_tv = tm_b2c["tv_28"].sum() if "tv_28" in tm_b2c.columns else 0
                t_cg = tm_b2c["cgst_14"].sum() if "cgst_14" in tm_b2c.columns else 0
                t_sg = tm_b2c["sgst_14"].sum() if "sgst_14" in tm_b2c.columns else 0
                t_ig = tm_b2c.get("igst_tv_28", pd.Series([0]*len(tm_b2c))).sum()
            elif rate == 18:
                t_tv = tm_b2c["tv_18"].sum() if "tv_18" in tm_b2c.columns else 0
                t_cg = tm_b2c["cgst_9"].sum() if "cgst_9" in tm_b2c.columns else 0
                t_sg = tm_b2c["sgst_9"].sum() if "sgst_9" in tm_b2c.columns else 0
                t_ig = tm_b2c.get("igst_tv_18", pd.Series([0]*len(tm_b2c))).sum()
            else:
                t_tv = tm_b2c["tv_12"].sum() if "tv_12" in tm_b2c.columns else 0
                t_cg = tm_b2c["cgst_6"].sum() if "cgst_6" in tm_b2c.columns else 0
                t_sg = tm_b2c["sgst_6"].sum() if "sgst_6" in tm_b2c.columns else 0
                t_ig = tm_b2c.get("igst_tv_12", pd.Series([0]*len(tm_b2c))).sum()

            # JSON B2CS for this month+rate
            jrec = jm_b2cs[jm_b2cs["rate"]==rate] if len(jm_b2cs) else pd.DataFrame()
            j_tv = jrec["taxable"].sum() if len(jrec) else 0
            j_cg = jrec["cgst"].sum() if len(jrec) else 0
            j_sg = jrec["sgst"].sum() if len(jrec) else 0
            j_ig = jrec["igst"].sum() if len(jrec) else 0

            if t_tv == 0 and j_tv == 0:
                continue
            d_tv = round(t_tv - j_tv, 2)
            vals = [m, f"{rate}%", t_tv, j_tv, d_tv, t_cg, j_cg, t_sg, j_sg, t_ig, j_ig]
            bg   = C_MATCH if abs(d_tv) < 1 else C_DIFF
            write_data_row(ws5, r5, vals, bg=bg, num_cols=NUM5)
            if abs(d_tv) >= 1:
                ws5.cell(row=r5, column=5).fill  = fill(C_TONLY if d_tv>0 else C_JONLY)
                ws5.cell(row=r5, column=5).font  = Font(bold=True, size=10)
            r5 += 1

    # ── Sheet 6: RATE-WISE SUMMARY ────────────────────────────────────────────
    ws6 = wb.create_sheet("06_RATE_WISE")
    ws6.merge_cells("A1:G1")
    c = ws6["A1"]
    c.value = f"Rate-wise Taxable Value — {name}  FY {fy}  (B2B + B2CS combined)"
    c.fill = fill(C_HEADER); c.font = hdr_font()
    c.alignment = Alignment(horizontal="center",vertical="center")
    hdr6 = ["Rate %","Tally Taxable","JSON Taxable","Difference",
            "Tally Tax","JSON Tax","Diff Tax"]
    apply_header_row(ws6, 2, hdr6, bg=C_SUBHDR, white_text=False)
    set_col_widths(ws6, [10,18,18,15,15,15,15])
    NUM6 = set(range(2,8))

    df_all = data["df_all"]
    for r6_i, rate in enumerate([28,18,12,0], 3):
        # Tally
        def _col(df, c, idx=None):
            if c in df.columns: return df[c]
            return pd.Series(0, index=df.index if idx is None else idx)
        if rate == 28:
            t_tv = _col(df_all,"tv_28").sum() + _col(df_all,"igst_tv_28").sum()
            t_tax= _col(df_all,"cgst_14").sum()*2 + _col(df_all,"igst_28").sum()
        elif rate == 18:
            t_tv = _col(df_all,"tv_18").sum() + _col(df_all,"igst_tv_18").sum()
            t_tax= _col(df_all,"cgst_9").sum()*2 + _col(df_all,"igst_18").sum()
        elif rate == 12:
            t_tv = _col(df_all,"tv_12").sum() + _col(df_all,"igst_tv_12").sum()
            t_tax= _col(df_all,"cgst_6").sum()*2 + _col(df_all,"igst_12").sum()
        else:
            t_tv = _col(df_all,"exempt").sum() + _col(df_all,"nil").sum()
            t_tax= 0

        # JSON
        j_b_rate = data["jb2b"][data["jb2b"]["rate"]==rate] if len(data["jb2b"]) else pd.DataFrame()
        j_c_rate = data["jb2cs"][data["jb2cs"]["rate"]==rate] if len(data["jb2cs"]) else pd.DataFrame()
        j_tv = (j_b_rate["taxable"].sum() if len(j_b_rate) else 0) + \
               (j_c_rate["taxable"].sum() if len(j_c_rate) else 0)
        j_tax= ((j_b_rate["cgst"].sum()+j_b_rate["sgst"].sum()+j_b_rate["igst"].sum()) if len(j_b_rate) else 0) + \
               ((j_c_rate["cgst"].sum()+j_c_rate["sgst"].sum()+j_c_rate["igst"].sum()) if len(j_c_rate) else 0)

        d_tv  = round(t_tv-j_tv,2)
        d_tax = round(t_tax-j_tax,2)
        vals = [f"{rate}%", t_tv, j_tv, d_tv, t_tax, j_tax, d_tax]
        bg = C_MATCH if abs(d_tv)<1 else C_DIFF
        write_data_row(ws6, r6_i, vals, bg=bg, num_cols=NUM6)


    # ── Sheet 7: B2C TALLY ONLY (individual invoices in Tally, not in JSON) ──
    ws7 = wb.create_sheet("07_B2C_TALLY_ONLY")
    ws7.merge_cells("A1:J1")
    c = ws7["A1"]
    c.value = f"B2C Invoices in Tally, NOT in GST Portal B2CS — {name}  FY {fy}"
    c.fill = fill("C55A11"); c.font = hdr_font()
    c.alignment = Alignment(horizontal="center", vertical="center")
    hdr7 = ["Month","Date","Invoice No.","Particulars",
            "Gross Total","TV @28%","TV @18%","TV @12%",
            "CGST","SGST","IGST","Remark"]
    ws7.merge_cells("A1:L1")
    apply_header_row(ws7, 2, hdr7, bg="C55A11")
    set_col_widths(ws7, [8,12,18,22,14,13,13,13,12,12,12,30])
    freeze(ws7, "E3")
    NUM7 = set(range(5,12))

    t_b2c_data = data["t_b2c"].copy()
    t_b2c_data["Date"] = pd.to_datetime(t_b2c_data["date"], errors="coerce")
    t_b2c_data = t_b2c_data.sort_values(["month","Date"])

    for i, (_, row) in enumerate(t_b2c_data.iterrows(), 3):
        # Determine which rate buckets are populated
        tv28 = row.get("tv_28", 0) or 0
        tv18 = row.get("tv_18", 0) or 0
        tv12 = row.get("tv_12", 0) or 0
        cgst = (row.get("cgst_14",0) or 0) + (row.get("cgst_9",0) or 0) + (row.get("cgst_6",0) or 0)
        sgst = (row.get("sgst_14",0) or 0) + (row.get("sgst_9",0) or 0) + (row.get("sgst_6",0) or 0)
        igst = (row.get("igst_28",0) or 0) + (row.get("igst_18",0) or 0) + (row.get("igst_12",0) or 0)
        gross = row.get("gross", 0) or 0
        date_str = row["Date"].strftime("%d-%m-%Y") if pd.notna(row["Date"]) else ""
        remark = "B2C — Not individually tracked in JSON (combined as B2CS)"
        vals = [row.get("month",""), date_str, row.get("inv_no",""), row.get("party",""),
                gross, tv28, tv18, tv12, cgst, sgst, igst, remark]
        write_data_row(ws7, i, vals, bg="FCE4D6", num_cols=NUM7)

    # ── Sheet 8: B2C JSON ONLY (B2CS records from JSON with no Tally match) ──
    ws8 = wb.create_sheet("08_B2C_JSON_ONLY")
    ws8.merge_cells("A1:J1")
    c = ws8["A1"]
    c.value = f"B2CS in GST Portal JSON, NOT matched in Tally — {name}  FY {fy}"
    c.fill = fill("1F4E79"); c.font = hdr_font()
    c.alignment = Alignment(horizontal="center", vertical="center")
    hdr8 = ["Month","Type","Place of Supply","Rate %",
            "Taxable Value","IGST","CGST","SGST","Total Tax","Remark"]
    apply_header_row(ws8, 2, hdr8)
    set_col_widths(ws8, [8,10,20,8,16,14,14,14,14,35])
    freeze(ws8, "E3")
    NUM8 = set(range(5,10))

    jb2cs_data = data["jb2cs"].copy()
    if len(jb2cs_data):
        jb2cs_data = jb2cs_data.sort_values(["month","rate"],
            key=lambda col: col.map(lambda x: MONTH_ORDER.index(x) if x in MONTH_ORDER else 99)
                            if col.name == "month" else col)
        for i, (_, row) in enumerate(jb2cs_data.iterrows(), 3):
            total_tax = (row.get("igst",0) or 0) + (row.get("cgst",0) or 0) + (row.get("sgst",0) or 0)
            vals = [row.get("month",""), row.get("type",""), row.get("pos",""),
                    f'{row.get("rate",0):.0f}%',
                    row.get("taxable",0), row.get("igst",0), row.get("cgst",0),
                    row.get("sgst",0), total_tax,
                    "B2CS from GST Portal — verify all invoices are in Tally"]
            write_data_row(ws8, i, vals, bg="DDEBF7", num_cols=NUM8)

    # ── Sheet 9: B2C MONTH-WISE BREAKUP ──────────────────────────────────────
    ws9 = wb.create_sheet("09_B2C_MONTHWISE")
    ws9.merge_cells("A1:P1")
    c = ws9["A1"]
    c.value = f"B2C Month-wise Breakup — Tally vs JSON  |  {name}  FY {fy}"
    c.fill = fill(C_HEADER); c.font = hdr_font()
    c.alignment = Alignment(horizontal="center", vertical="center")

    hdr9 = ["Month",
            "Tally Invoices","Tally Gross","Tally TV @28%","Tally TV @18%","Tally TV @12%",
            "Tally CGST","Tally SGST","Tally IGST","Tally Total Tax",
            "JSON Taxable","JSON CGST","JSON SGST","JSON IGST","JSON Total Tax",
            "Diff Taxable","Diff Tax","Status"]
    apply_header_row(ws9, 2, hdr9, bg=C_SUBHDR, white_text=False)
    set_col_widths(ws9, [8,10,14,13,13,13,12,12,12,12,13,12,12,12,12,13,12,10])
    freeze(ws9, "B3")
    NUM9 = set(range(2,18))

    r9 = 3
    ann9 = {k:0.0 for k in ["t_gross","t_tv28","t_tv18","t_tv12","t_cg","t_sg","t_ig","t_tax",
                              "j_tv","j_cg","j_sg","j_ig","j_tax","n_inv"]}

    for m in data["months"]:
        tm = t_b2c_data[t_b2c_data["month"]==m]
        jm = data["jb2cs"][data["jb2cs"]["month"]==m] if len(data["jb2cs"]) else pd.DataFrame()

        n_inv  = len(tm)
        t_gross= tm["gross"].sum() if "gross" in tm.columns else 0
        t_tv28 = tm["tv_28"].sum() if "tv_28" in tm.columns else 0
        t_tv18 = tm["tv_18"].sum() if "tv_18" in tm.columns else 0
        t_tv12 = tm["tv_12"].sum() if "tv_12" in tm.columns else 0
        t_cg   = (tm.get("cgst_14", pd.Series(0,index=tm.index)).sum() +
                  tm.get("cgst_9",  pd.Series(0,index=tm.index)).sum() +
                  tm.get("cgst_6",  pd.Series(0,index=tm.index)).sum())
        t_sg   = (tm.get("sgst_14", pd.Series(0,index=tm.index)).sum() +
                  tm.get("sgst_9",  pd.Series(0,index=tm.index)).sum() +
                  tm.get("sgst_6",  pd.Series(0,index=tm.index)).sum())
        t_ig   = (tm.get("igst_28", pd.Series(0,index=tm.index)).sum() +
                  tm.get("igst_18", pd.Series(0,index=tm.index)).sum() +
                  tm.get("igst_12", pd.Series(0,index=tm.index)).sum())
        t_tax  = t_cg + t_sg + t_ig

        j_tv  = jm["taxable"].sum() if len(jm) else 0
        j_cg  = jm["cgst"].sum()    if len(jm) else 0
        j_sg  = jm["sgst"].sum()    if len(jm) else 0
        j_ig  = jm["igst"].sum()    if len(jm) else 0
        j_tax = j_cg + j_sg + j_ig

        t_tv_total = t_tv28 + t_tv18 + t_tv12
        d_tv  = round(t_tv_total - j_tv, 2)
        d_tax = round(t_tax - j_tax, 2)
        ok    = abs(d_tv) < 1 and abs(d_tax) < 1
        status= "✓ OK" if ok else ("⚠ DIFF" if abs(d_tv)<5000 else "✗ CHECK")

        if n_inv == 0 and j_tv == 0:
            continue  # skip empty months

        vals = [m, n_inv, t_gross, t_tv28, t_tv18, t_tv12,
                t_cg, t_sg, t_ig, t_tax,
                j_tv, j_cg, j_sg, j_ig, j_tax,
                d_tv, d_tax, status]
        bg = C_MATCH if ok else (C_DIFF if "DIFF" in status else C_TONLY)
        write_data_row(ws9, r9, vals, bg=bg if not ok else None, num_cols=NUM9)

        # Color diff cells
        for col_idx in [16, 17]:
            cell = ws9.cell(row=r9, column=col_idx)
            val  = vals[col_idx-1]
            if isinstance(val, float) and abs(val) >= 1:
                cell.fill = fill(C_TONLY if val > 0 else C_JONLY)
                cell.font = Font(bold=True, size=10)
        sc = ws9.cell(row=r9, column=18)
        sc.font = Font(bold=True, size=10,
                       color=C_OK if ok else (C_ERR if "CHECK" in status else "000000"))

        for k,v in [("n_inv",n_inv),("t_gross",t_gross),("t_tv28",t_tv28),
                    ("t_tv18",t_tv18),("t_tv12",t_tv12),("t_cg",t_cg),
                    ("t_sg",t_sg),("t_ig",t_ig),("t_tax",t_tax),
                    ("j_tv",j_tv),("j_cg",j_cg),("j_sg",j_sg),("j_ig",j_ig),("j_tax",j_tax)]:
            ann9[k] += v
        r9 += 1

    # Annual total
    t_tv_ann = ann9["t_tv28"] + ann9["t_tv18"] + ann9["t_tv12"]
    d_tv_ann  = round(t_tv_ann - ann9["j_tv"], 2)
    d_tax_ann = round(ann9["t_tax"] - ann9["j_tax"], 2)
    tot9 = ["★ FY TOTAL", int(ann9["n_inv"]), ann9["t_gross"],
            ann9["t_tv28"], ann9["t_tv18"], ann9["t_tv12"],
            ann9["t_cg"], ann9["t_sg"], ann9["t_ig"], ann9["t_tax"],
            ann9["j_tv"], ann9["j_cg"], ann9["j_sg"], ann9["j_ig"], ann9["j_tax"],
            d_tv_ann, d_tax_ann,
            "✓ MATCHED" if abs(d_tv_ann)<1 else "✗ MISMATCH"]
    write_data_row(ws9, r9, tot9, bg=C_TOTAL, num_cols=NUM9, bold=True)

    # ── Sheet 10: B2C RATE-WISE DETAIL ───────────────────────────────────────
    ws10 = wb.create_sheet("10_B2C_RATEWISE")
    ws10.merge_cells("A1:M1")
    c = ws10["A1"]
    c.value = f"B2C Rate-wise Month Detail — {name}  FY {fy}  (Tally invoice-level vs JSON B2CS)"
    c.fill = fill(C_HEADER); c.font = hdr_font()
    c.alignment = Alignment(horizontal="center", vertical="center")
    hdr10 = ["Month","Rate",
             "Tally Inv Count","Tally Taxable","Tally CGST","Tally SGST","Tally IGST",
             "JSON Taxable","JSON CGST","JSON SGST","JSON IGST",
             "Diff Taxable","Status"]
    apply_header_row(ws10, 2, hdr10, bg=C_SUBHDR, white_text=False)
    set_col_widths(ws10, [8,7,12,15,13,13,13,14,13,13,13,14,10])
    freeze(ws10, "C3")
    NUM10 = set(range(4,13))

    r10 = 3
    for m in data["months"]:
        tm = t_b2c_data[t_b2c_data["month"]==m]
        jm = data["jb2cs"][data["jb2cs"]["month"]==m] if len(data["jb2cs"]) else pd.DataFrame()

        for rate, tv_col, cg_col, sg_col, ig_col in [
            (28, "tv_28",  "cgst_14", "sgst_14", "igst_28"),
            (18, "tv_18",  "cgst_9",  "sgst_9",  "igst_18"),
            (12, "tv_12",  "cgst_6",  "sgst_6",  "igst_12"),
        ]:
            t_tv = tm[tv_col].sum() if tv_col in tm.columns else 0
            t_cg = tm[cg_col].sum() if cg_col in tm.columns else 0
            t_sg = tm[sg_col].sum() if sg_col in tm.columns else 0
            t_ig = (tm["igst_" + str(rate)].sum()
                    if "igst_" + str(rate) in tm.columns else 0)

            jr = jm[jm["rate"]==rate] if len(jm) else pd.DataFrame()
            j_tv = jr["taxable"].sum() if len(jr) else 0
            j_cg = jr["cgst"].sum()    if len(jr) else 0
            j_sg = jr["sgst"].sum()    if len(jr) else 0
            j_ig = jr["igst"].sum()    if len(jr) else 0

            n_inv = int((tm[tv_col] > 0).sum()) if tv_col in tm.columns else 0

            if t_tv == 0 and j_tv == 0:
                continue

            d_tv   = round(t_tv - j_tv, 2)
            ok_row = abs(d_tv) < 1
            status = "✓" if ok_row else ("⚠" if abs(d_tv)<5000 else "✗")

            vals = [m, f"{rate}%", n_inv,
                    t_tv, t_cg, t_sg, t_ig,
                    j_tv, j_cg, j_sg, j_ig,
                    d_tv, status]
            bg = C_MATCH if ok_row else C_DIFF
            write_data_row(ws10, r10, vals, bg=bg, num_cols=NUM10)
            if not ok_row:
                ws10.cell(row=r10, column=12).fill = fill(C_TONLY if d_tv>0 else C_JONLY)
                ws10.cell(row=r10, column=12).font = Font(bold=True, size=10)
            ws10.cell(row=r10, column=13).font = Font(
                bold=True, size=10,
                color=C_OK if ok_row else (C_ERR if "✗" in status else "FF6600"))
            r10 += 1


    # ── Sheet 11: B2B MONTH-WISE BREAKUP ─────────────────────────────────────
    ws11 = wb.create_sheet("11_B2B_MONTHWISE")
    ws11.merge_cells("A1:R1")
    c = ws11["A1"]
    c.value = f"B2B Month-wise Breakup — Tally vs JSON  |  {name}  FY {fy}"
    c.fill = fill(C_HEADER); c.font = hdr_font()
    c.alignment = Alignment(horizontal="center", vertical="center")
    hdr11 = ["Month",
             "Tally Inv","Tally Taxable","Tally CGST","Tally SGST","Tally IGST","Tally Total Tax","Tally Inv Value",
             "JSON Inv","JSON Taxable","JSON CGST","JSON SGST","JSON IGST","JSON Total Tax","JSON Inv Value",
             "Diff Taxable","Diff Tax","Status"]
    apply_header_row(ws11, 2, hdr11, bg=C_SUBHDR, white_text=False)
    set_col_widths(ws11, [8,8,15,13,13,13,13,14,8,14,13,13,13,13,14,14,12,10])
    freeze(ws11, "B3")
    NUM11 = set(range(2,18))

    r11 = 3
    ann11 = {k:0.0 for k in ["t_inv","t_tv","t_cg","t_sg","t_ig","t_tax","t_val",
                               "j_inv","j_tv","j_cg","j_sg","j_ig","j_tax","j_val"]}

    t_b2b_data = data["t_b2b"].copy()
    jb2b_data  = data["jb2b"].copy()

    for m in data["months"]:
        tm = t_b2b_data[t_b2b_data["month"]==m]
        jm = jb2b_data[jb2b_data["month"]==m] if len(jb2b_data) else pd.DataFrame()

        t_inv = len(tm)
        t_tv  = tm["taxable_total"].sum() if "taxable_total" in tm.columns else 0
        t_cg  = tm["total_cgst"].sum()    if "total_cgst"    in tm.columns else 0
        t_sg  = tm["total_sgst"].sum()    if "total_sgst"    in tm.columns else 0
        t_ig  = tm["total_igst"].sum()    if "total_igst"    in tm.columns else 0
        t_tax = t_cg + t_sg + t_ig
        t_val = tm["gross"].sum()         if "gross"         in tm.columns else 0

        j_inv = len(jm)
        j_tv  = jm["taxable"].sum() if len(jm) else 0
        j_cg  = jm["cgst"].sum()    if len(jm) else 0
        j_sg  = jm["sgst"].sum()    if len(jm) else 0
        j_ig  = jm["igst"].sum()    if len(jm) else 0
        j_tax = j_cg + j_sg + j_ig
        j_val = jm["inv_val"].sum() if len(jm) else 0

        if t_inv == 0 and j_inv == 0:
            continue

        d_tv  = round(t_tv - j_tv, 2)
        d_tax = round(t_tax - j_tax, 2)
        ok    = abs(d_tv) < 1 and abs(d_tax) < 1
        status= "✓ OK" if ok else ("⚠ DIFF" if abs(d_tv)<5000 else "✗ CHECK")

        vals = [m, t_inv, t_tv, t_cg, t_sg, t_ig, t_tax, t_val,
                   j_inv, j_tv, j_cg, j_sg, j_ig, j_tax, j_val,
                   d_tv, d_tax, status]
        bg = C_MATCH if ok else (C_DIFF if "DIFF" in status else C_TONLY)
        write_data_row(ws11, r11, vals, bg=bg if not ok else None, num_cols=NUM11)
        for col_idx in [16,17]:
            cell = ws11.cell(row=r11, column=col_idx)
            val  = vals[col_idx-1]
            if isinstance(val,(int,float)) and abs(val)>=1:
                cell.fill = fill(C_TONLY if val>0 else C_JONLY)
                cell.font = Font(bold=True, size=10)
        sc = ws11.cell(row=r11, column=18)
        sc.font = Font(bold=True, size=10,
                       color=C_OK if ok else (C_ERR if "CHECK" in status else "FF6600"))
        for k,v in [("t_inv",t_inv),("t_tv",t_tv),("t_cg",t_cg),("t_sg",t_sg),
                    ("t_ig",t_ig),("t_tax",t_tax),("t_val",t_val),
                    ("j_inv",j_inv),("j_tv",j_tv),("j_cg",j_cg),("j_sg",j_sg),
                    ("j_ig",j_ig),("j_tax",j_tax),("j_val",j_val)]:
            ann11[k] += v
        r11 += 1

    # Annual total
    d_tv_a  = round(ann11["t_tv"]  - ann11["j_tv"],  2)
    d_tax_a = round(ann11["t_tax"] - ann11["j_tax"], 2)
    tot11 = ["★ FY TOTAL",
             int(ann11["t_inv"]), ann11["t_tv"], ann11["t_cg"], ann11["t_sg"],
             ann11["t_ig"], ann11["t_tax"], ann11["t_val"],
             int(ann11["j_inv"]), ann11["j_tv"], ann11["j_cg"], ann11["j_sg"],
             ann11["j_ig"], ann11["j_tax"], ann11["j_val"],
             d_tv_a, d_tax_a,
             "✓ MATCHED" if abs(d_tv_a)<1 else "✗ MISMATCH"]
    write_data_row(ws11, r11, tot11, bg=C_TOTAL, num_cols=NUM11, bold=True)

    # ── Sheet 12: B2B + B2C COMBINED MONTH-WISE ──────────────────────────────
    ws12 = wb.create_sheet("12_COMBINED_MONTHWISE")
    ws12.merge_cells("A1:T1")
    c = ws12["A1"]
    c.value = f"B2B + B2C Combined Month-wise — Tally vs JSON  |  {name}  FY {fy}"
    c.fill = fill(C_HEADER); c.font = hdr_font()
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Sub-header rows: row2=category labels, row3=column labels
    # Category labels spanning
    cats = [("",1),("TALLY — B2B",7),("TALLY — B2C",5),("TALLY TOTAL",2),
            ("JSON TOTAL",5),("VARIANCE",2),("",1)]
    col = 1
    for cat, span in cats:
        if cat:
            ws12.merge_cells(start_row=2, start_column=col,
                             end_row=2, end_column=col+span-1)
            c2 = ws12.cell(row=2, column=col, value=cat)
            bg_map = {"TALLY — B2B":"2E75B6","TALLY — B2C":"70AD47",
                      "TALLY TOTAL":"375623","JSON TOTAL":"833C00","VARIANCE":"C55A11"}
            c2.fill = fill(bg_map.get(cat, C_SUBHDR))
            c2.font = Font(bold=True, color="FFFFFF", size=10)
            c2.alignment = Alignment(horizontal="center", vertical="center")
        col += span

    hdr12 = ["Month",
             "B2B Inv","B2B Taxable","B2B CGST","B2B SGST","B2B IGST","B2B Tax","B2B Inv Val",
             "B2C Inv","B2C Taxable","B2C CGST+SGST","B2C IGST",  "B2C Tax",
             "Total Taxable","Total Tax",
             "JSON Taxable","JSON CGST","JSON SGST","JSON IGST","JSON Tax",
             "Diff Taxable","Diff Tax","Status"]
    apply_header_row(ws12, 3, hdr12, bg=C_SUBHDR, white_text=False)
    set_col_widths(ws12, [8,7,14,12,12,12,12,13,7,13,13,13,12,14,12,14,12,12,12,12,13,11,10])
    ws12.freeze_panes = "B4"
    NUM12 = set(range(2,23))

    r12 = 4
    ann12 = {k:0.0 for k in [
        "b2b_inv","b2b_tv","b2b_cg","b2b_sg","b2b_ig","b2b_tax","b2b_val",
        "b2c_inv","b2c_tv","b2c_cg_sg","b2c_ig","b2c_tax",
        "t_total_tv","t_total_tax",
        "j_tv","j_cg","j_sg","j_ig","j_tax"]}

    jb2cs_data2 = data["jb2cs"].copy()

    for m in data["months"]:
        # Tally B2B
        tb2b = t_b2b_data[t_b2b_data["month"]==m]
        b2b_inv = len(tb2b)
        b2b_tv  = tb2b["taxable_total"].sum() if "taxable_total" in tb2b.columns else 0
        b2b_cg  = tb2b["total_cgst"].sum()    if "total_cgst"    in tb2b.columns else 0
        b2b_sg  = tb2b["total_sgst"].sum()    if "total_sgst"    in tb2b.columns else 0
        b2b_ig  = tb2b["total_igst"].sum()    if "total_igst"    in tb2b.columns else 0
        b2b_tax = b2b_cg + b2b_sg + b2b_ig
        b2b_val = tb2b["gross"].sum()         if "gross"         in tb2b.columns else 0

        # Tally B2C
        tb2c = t_b2c_data[t_b2c_data["month"]==m]
        b2c_inv = len(tb2c)
        b2c_tv  = (tb2c.get("tv_28", pd.Series(0,index=tb2c.index)).sum() +
                   tb2c.get("tv_18", pd.Series(0,index=tb2c.index)).sum() +
                   tb2c.get("tv_12", pd.Series(0,index=tb2c.index)).sum())
        b2c_cg  = (tb2c.get("cgst_14",pd.Series(0,index=tb2c.index)).sum() +
                   tb2c.get("cgst_9", pd.Series(0,index=tb2c.index)).sum() +
                   tb2c.get("cgst_6", pd.Series(0,index=tb2c.index)).sum())
        b2c_sg  = (tb2c.get("sgst_14",pd.Series(0,index=tb2c.index)).sum() +
                   tb2c.get("sgst_9", pd.Series(0,index=tb2c.index)).sum() +
                   tb2c.get("sgst_6", pd.Series(0,index=tb2c.index)).sum())
        b2c_ig  = (tb2c.get("igst_28",pd.Series(0,index=tb2c.index)).sum() +
                   tb2c.get("igst_18",pd.Series(0,index=tb2c.index)).sum() +
                   tb2c.get("igst_12",pd.Series(0,index=tb2c.index)).sum())
        b2c_tax = b2c_cg + b2c_sg + b2c_ig

        # Tally totals
        t_total_tv  = b2b_tv  + b2c_tv
        t_total_tax = b2b_tax + b2c_tax

        # JSON totals (B2B + B2CS)
        jb2b_m = jb2b_data[jb2b_data["month"]==m]   if len(jb2b_data)   else pd.DataFrame()
        jb2cs_m= jb2cs_data2[jb2cs_data2["month"]==m] if len(jb2cs_data2) else pd.DataFrame()
        j_tv  = (jb2b_m["taxable"].sum() if len(jb2b_m) else 0) + (jb2cs_m["taxable"].sum() if len(jb2cs_m) else 0)
        j_cg  = (jb2b_m["cgst"].sum()    if len(jb2b_m) else 0) + (jb2cs_m["cgst"].sum()    if len(jb2cs_m) else 0)
        j_sg  = (jb2b_m["sgst"].sum()    if len(jb2b_m) else 0) + (jb2cs_m["sgst"].sum()    if len(jb2cs_m) else 0)
        j_ig  = (jb2b_m["igst"].sum()    if len(jb2b_m) else 0) + (jb2cs_m["igst"].sum()    if len(jb2cs_m) else 0)
        j_tax = j_cg + j_sg + j_ig

        if b2b_inv == 0 and b2c_inv == 0 and j_tv == 0:
            continue

        d_tv  = round(t_total_tv  - j_tv,  2)
        d_tax = round(t_total_tax - j_tax, 2)
        ok    = abs(d_tv) < 1 and abs(d_tax) < 1
        status= "✓ OK" if ok else ("⚠ DIFF" if abs(d_tv)<5000 else "✗ CHECK")

        vals = [m,
                b2b_inv, b2b_tv, b2b_cg, b2b_sg, b2b_ig, b2b_tax, b2b_val,
                b2c_inv, b2c_tv, b2c_cg+b2c_sg, b2c_ig, b2c_tax,
                t_total_tv, t_total_tax,
                j_tv, j_cg, j_sg, j_ig, j_tax,
                d_tv, d_tax, status]
        bg = C_MATCH if ok else (C_DIFF if "DIFF" in status else C_TONLY)
        write_data_row(ws12, r12, vals, bg=bg if not ok else None, num_cols=NUM12)

        for col_idx in [21,22]:
            cell = ws12.cell(row=r12, column=col_idx)
            val  = vals[col_idx-1]
            if isinstance(val,(int,float)) and abs(val)>=1:
                cell.fill = fill(C_TONLY if val>0 else C_JONLY)
                cell.font = Font(bold=True, size=10)
        sc = ws12.cell(row=r12, column=23)
        sc.font = Font(bold=True, size=10,
                       color=C_OK if ok else (C_ERR if "CHECK" in status else "FF6600"))

        for k,v in [
            ("b2b_inv",b2b_inv),("b2b_tv",b2b_tv),("b2b_cg",b2b_cg),("b2b_sg",b2b_sg),
            ("b2b_ig",b2b_ig),("b2b_tax",b2b_tax),("b2b_val",b2b_val),
            ("b2c_inv",b2c_inv),("b2c_tv",b2c_tv),("b2c_cg_sg",b2c_cg+b2c_sg),
            ("b2c_ig",b2c_ig),("b2c_tax",b2c_tax),
            ("t_total_tv",t_total_tv),("t_total_tax",t_total_tax),
            ("j_tv",j_tv),("j_cg",j_cg),("j_sg",j_sg),("j_ig",j_ig),("j_tax",j_tax)]:
            ann12[k] += v
        r12 += 1

    # Annual total row
    d_tv_12  = round(ann12["t_total_tv"]  - ann12["j_tv"],  2)
    d_tax_12 = round(ann12["t_total_tax"] - ann12["j_tax"], 2)
    tot12 = ["★ FY TOTAL",
             int(ann12["b2b_inv"]), ann12["b2b_tv"], ann12["b2b_cg"], ann12["b2b_sg"],
             ann12["b2b_ig"], ann12["b2b_tax"], ann12["b2b_val"],
             int(ann12["b2c_inv"]), ann12["b2c_tv"], ann12["b2c_cg_sg"],
             ann12["b2c_ig"], ann12["b2c_tax"],
             ann12["t_total_tv"], ann12["t_total_tax"],
             ann12["j_tv"], ann12["j_cg"], ann12["j_sg"], ann12["j_ig"], ann12["j_tax"],
             d_tv_12, d_tax_12,
             "✓ MATCHED" if abs(d_tv_12)<1 else "✗ MISMATCH"]
    write_data_row(ws12, r12, tot12, bg=C_TOTAL, num_cols=NUM12, bold=True)

    wb.save(out_path)
    print(f"\n  ✓ Saved: {out_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def _find_output_base():
    from pathlib import Path
    home = Path.home()
    candidates = [
        home / "OneDrive" / "Desktop" / "OUTPUT",
        home / "OneDrive - Personal" / "Desktop" / "OUTPUT",
        home / "Desktop" / "OUTPUT",
        home / "Downloads",
    ]
    try:
        candidates[2:2] = [p / "Desktop" / "OUTPUT"
                            for p in home.glob("OneDrive*") if p.is_dir()]
    except Exception:
        pass
    for c in candidates:
        if c.exists():
            return c
    return home / "Downloads"

def _pick_file_interactive():
    """
    Interactive file/folder picker — no command-line args needed.
    Uses tkinter file dialog if available, otherwise numbered list fallback.
    """
    import tkinter as tk
    from tkinter import filedialog, simpledialog, messagebox

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    print("\n  ════════════════════════════════════════════════════════")
    print("   GSTR-1 Tally vs JSON Comparison — FILE SELECTOR")
    print("  ════════════════════════════════════════════════════════")

    # ── Step 1: Tally XLS ─────────────────────────────────────────
    print("\n  Step 1 of 4 — Select your Tally Sales Register file")
    print("  (File picker window will open...)")
    tally_path = filedialog.askopenfilename(
        title="Select Tally GSTR-1 Sales Register",
        filetypes=[("Excel files", "*.xls *.xlsx *.xlsm"), ("All files", "*.*")],
        initialdir=str(_find_output_base())
    )
    if not tally_path:
        print("  ✗ No file selected. Exiting.")
        sys.exit(0)
    print(f"  ✓ Tally file : {tally_path}")

    # ── Step 2: JSON ZIPs folder ──────────────────────────────────
    print("\n  Step 2 of 4 — Select folder containing GSTR1 ZIP files")
    print("  (Folder picker window will open...)")
    json_folder = filedialog.askdirectory(
        title="Select folder with GSTR1_*.zip files (from GST portal download)",
        initialdir=str(_find_output_base())
    )
    if not json_folder:
        print("  ✗ No folder selected. Exiting.")
        sys.exit(0)
    # Check ZIPs present
    zips = list(Path(json_folder).glob("*.zip")) + list(Path(json_folder).rglob("GSTR1_*.zip"))
    print(f"  ✓ JSON folder: {json_folder}  ({len(zips)} ZIP files found)")
    if not zips:
        print("  ⚠  No ZIP files found in this folder!")
        messagebox.showwarning("No ZIPs found",
            f"No ZIP files found in:\n{json_folder}\n\nPlease select the folder that contains GSTR1_*.zip files.")

    # ── Step 3: Client name ───────────────────────────────────────
    print("\n  Step 3 of 4 — Enter client name")
    name = simpledialog.askstring(
        "Client Name",
        "Enter client / company name\n(e.g. ARUN ENTERPRISES):",
        initialvalue="ARUN ENTERPRISES",
        parent=root
    ) or "CLIENT"
    print(f"  ✓ Client name: {name}")

    # ── Step 4: Financial Year ────────────────────────────────────
    print("\n  Step 4 of 4 — Select Financial Year")
    fy = simpledialog.askstring(
        "Financial Year",
        "Enter Financial Year (e.g. 2024-25):",
        initialvalue="2024-25",
        parent=root
    ) or "2024-25"
    print(f"  ✓ Financial year: {fy}")

    root.destroy()
    return tally_path, json_folder, name, fy


def _pick_file_console():
    """Fallback: numbered list picker when tkinter is not available."""
    base = _find_output_base()

    print("\n  ════════════════════════════════════════════════════════")
    print("   GSTR-1 Tally vs JSON Comparison — CONSOLE SELECTOR")
    print("  ════════════════════════════════════════════════════════")

    # Find Tally XLS files
    print("\n  Step 1 of 4 — Select Tally Sales Register file")
    tally_files = (sorted(base.rglob("*.xls")) +
                   sorted(base.rglob("*.xlsx")) +
                   sorted(Path.home().glob("Downloads/*.xls*")))
    # deduplicate
    seen = set(); tally_files = [f for f in tally_files if str(f) not in seen and not seen.add(str(f))]

    if tally_files:
        for i, f in enumerate(tally_files, 1):
            print(f"    {i:2d}.  {f}")
        print(f"     0.  Type path manually")
        raw = input("  Enter number (or 0): ").strip()
        if raw == "0" or not raw.isdigit():
            tally_path = input("  Paste full path to Tally XLS: ").strip().strip('"')
        else:
            tally_path = str(tally_files[int(raw)-1])
    else:
        tally_path = input("  Paste full path to Tally XLS: ").strip().strip('"')
    print(f"  ✓ Tally file: {tally_path}")

    # Find GSTR1 folders
    print("\n  Step 2 of 4 — Select GSTR1 ZIP folder")
    gstr1_folders = sorted(base.rglob("GSTR1"))
    if not gstr1_folders:
        gstr1_folders = [d for d in base.rglob("*") if d.is_dir() and
                         any(d.glob("GSTR1_*.zip"))]

    if gstr1_folders:
        for i, f in enumerate(gstr1_folders, 1):
            zc = len(list(f.glob("*.zip")))
            print(f"    {i:2d}.  {f}  ({zc} ZIPs)")
        print(f"     0.  Type path manually")
        raw = input("  Enter number (or 0): ").strip()
        if raw == "0" or not raw.isdigit():
            json_folder = input("  Paste full path to GSTR1 ZIPs folder: ").strip().strip('"')
        else:
            json_folder = str(gstr1_folders[int(raw)-1])
    else:
        json_folder = input("  Paste full path to GSTR1 ZIPs folder: ").strip().strip('"')
    print(f"  ✓ JSON folder: {json_folder}")

    name = input("\n  Step 3 of 4 — Client name [ARUN ENTERPRISES]: ").strip() or "ARUN ENTERPRISES"
    fy   = input("  Step 4 of 4 — Financial Year [2024-25]: ").strip() or "2024-25"

    return tally_path, json_folder, name, fy


def main():
    ap = argparse.ArgumentParser(description="Compare Tally GSTR-1 vs GST Portal JSON")
    ap.add_argument("--tally", default=None, help="Path to Tally XLS file (skip picker)")
    ap.add_argument("--json",  default=None, help="Folder with GSTR1 ZIPs (skip picker)")
    ap.add_argument("--name",  default="",   help="Client name")
    ap.add_argument("--fy",    default="",   help="Financial year e.g. 2024-25")
    ap.add_argument("--out",   default=None, help="Output folder")
    args = ap.parse_args()

    # If args not provided, use interactive picker
    if not args.tally or not args.json:
        try:
            tally_path, json_folder, name, fy = _pick_file_interactive()
        except Exception as e:
            print(f"  (tkinter not available: {e}) — using console picker")
            tally_path, json_folder, name, fy = _pick_file_console()
    else:
        tally_path  = args.tally
        json_folder = args.json
        name        = args.name or "CLIENT"
        fy          = args.fy   or "2024-25"

    out_dir = Path(args.out) if args.out else (Path.home() / "Downloads" / "GSTR1_Comparison")
    out_dir.mkdir(parents=True, exist_ok=True)

    ts        = datetime.now().strftime("%Y%m%d_%H%M")
    safe_name = re.sub(r'[\\/:*?"<>|]', "_", name)
    out_path  = out_dir / f"GSTR1_TALLY_VS_JSON_{safe_name}_FY{fy.replace('-','_')}_{ts}.xlsx"

    print(f"""
  ═══════════════════════════════════════════════════════
   GSTR-1 Tally vs JSON Comparison — RUNNING
   Client : {name}
   FY     : {fy}
   Tally  : {tally_path}
   JSON   : {json_folder}
   Output : {out_path}
  ═══════════════════════════════════════════════════════""")

    data = build_comparison(tally_path, json_folder, name, fy)

    print(f"\n  Summary:")
    print(f"    B2B Matched invoices  : {len(data['matched_rows'])}")
    print(f"    Tally only (missing)  : {len(data['tonly_rows'])}")
    print(f"    JSON only (extra)     : {len(data['jonly_rows'])}")

    write_excel(data, out_path)
    print(f"\n  Output saved to: {out_path}")
    print("  ✓ Done! Open the Excel file to review the comparison.")

if __name__ == "__main__":
    main()
