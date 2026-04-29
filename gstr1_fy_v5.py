"""
GSTR-1 Full Financial Year Extractor  v5
==========================================
FIXES & NEW FEATURES vs v4
  ✓ BUG FIX  : Duplicate month data removed (glob dedup)
  ✓ NEW      : Trader name shown in every sheet title
  ✓ NEW      : Month-wise SUBTOTAL row after each month's data
  ✓ NEW      : Per-GSTIN subtotal rows within B2B / CDNR sheets
  ✓ NEW      : Sheet 25 — FY GSTIN-wise Annual Summary (one row per GSTIN)
  ✓ NEW      : Sheet 26 — Master All-Invoices (B2B flat, all months, filterable)
  ✓ NEW      : FY Grand Total row at bottom of every sheet

USAGE
  python gstr1_fy_v5.py  --name "SOWMIYA ENTERPRISES"  /folder/with/zips
  python gstr1_fy_v5.py  --name "XYZ Co"  Apr.zip May.zip Jun.zip ...
  python gstr1_fy_v5.py  /folder          (name auto-read from GSTIN if omitted)
"""

import json, zipfile, sys, argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── GSTIN Name Cache (auto-fetches buyer names from GST portal) ───────────────
try:
    from gstin_name_cache import GSTINNameCache as _GSTINNameCache
    _NAME_CACHE    = _GSTINNameCache(log_fn=lambda m: print(m))
    _CACHE_ENABLED = True
except Exception:
    _NAME_CACHE    = None
    _CACHE_ENABLED = False

# ── Palette ──────────────────────────────────────────────────────────────
C_DARK   = "1F3864"   # navy  – main titles
C_MED    = "2E75B6"   # blue  – col headers
C_AMEND  = "7030A0"   # purple– amendment sheets
C_SEP    = "17375E"   # dark  – month separator
C_MSUB   = "2F5496"   # month subtotal
C_GSUB   = "4472C4"   # GSTIN subtotal (lighter blue)
C_ANN    = "1F3864"   # annual total
C_ORANGE = "C55A11"   # warning
C_LGREY  = "F2F2F2"
C_WHITE  = "FFFFFF"
C_YELLOW = "FFF2CC"
C_LTBLUE = "DEEAF1"
C_LGREEN = "E2EFDA"
C_LPUR   = "F3E8FF"

MONTH_ORDER = {"04":1,"05":2,"06":3,"07":4,"08":5,"09":6,
               "10":7,"11":8,"12":9,"01":10,"02":11,"03":12}
MONTH_NAME  = {"01":"January","02":"February","03":"March","04":"April",
               "05":"May","06":"June","07":"July","08":"August",
               "09":"September","10":"October","11":"November","12":"December"}
DOC_TYPES   = {1:"Invoices for outward supply",2:"Invoices for inward supply (unregistered)",
               3:"Revised Invoice",4:"Debit Note",5:"Credit Note",6:"Advance Receipt",
               7:"Payment Voucher",8:"Refund Voucher",9:"Delivery Challan – job work",
               10:"Delivery Challan – supply on approval",11:"Delivery Challan – liquid gas",
               12:"Delivery Challan – others"}
INV_TYPE    = {"R":"Regular","SEZWP":"SEZ with Payment","SEZWOP":"SEZ without Payment",
               "DE":"Deemed Export","CBW":"Customs Bonded Warehouse"}
NUM = "#,##0.00"; INT = "#,##0"

# ── Style helpers ─────────────────────────────────────────────────────────
def fill(c):   return PatternFill("solid", fgColor=c)
def font(b=False, c="000000", s=9): return Font(name="Arial", bold=b, color=c, size=s)
def bdr():
    sd = Side(style="thin")
    return Border(left=sd, right=sd, top=sd, bottom=sd)
def aln(h="left", w=False): return Alignment(horizontal=h, vertical="center", wrap_text=w)

_ri = {}
def _r(ws):      return _ri.setdefault(ws.title, 3)
def _next(ws):   r = _r(ws); _ri[ws.title] = r+1; return r
def _setr(ws,n): _ri[ws.title] = n

def wcell(ws, row, col, val, bold=False, bg=C_WHITE,
          ha="left", fmt=None, fc="000000", wrap=False, size=9):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = font(bold, fc, size)
    c.fill      = fill(bg)
    c.alignment = aln(ha, wrap)
    c.border    = bdr()
    if fmt: c.number_format = fmt
    return c

def sheet_title(ws, title, nc, bg=C_DARK):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    c = ws["A1"]; c.value = title
    c.font = font(True, "FFFFFF", 11); c.fill = fill(bg)
    c.alignment = aln("center"); c.border = bdr()
    ws.row_dimensions[1].height = 26

def trader_row(ws, trader_name, gstin, nc, row=2):
    """Second row showing trader name + GSTIN."""
    ws.merge_cells(f"A{row}:{get_column_letter(nc)}{row}")
    c = ws.cell(row=row, column=1)
    c.value = f"  {trader_name}   |   GSTIN: {gstin}"
    c.font  = font(True, "FFFF00", 10)
    c.fill  = fill(C_SEP)
    c.alignment = aln("left")
    c.border = bdr()
    ws.row_dimensions[row].height = 20

def col_headers(ws, headers, widths, row=3, bg=C_MED):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = font(True, "FFFFFF", 9); c.fill = fill(bg)
        c.alignment = aln("center", True); c.border = bdr()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 32
    ws.freeze_panes = f"A{row+1}"

def month_sep(ws, label, nc):
    r = _next(ws)
    ws.merge_cells(f"A{r}:{get_column_letter(nc)}{r}")
    c = ws.cell(row=r, column=1, value=f"  ▶  {label}")
    c.font = font(True, "FFFFFF", 9); c.fill = fill(C_SEP)
    c.alignment = aln("left"); c.border = bdr()
    ws.row_dimensions[r].height = 15

def no_data_row(ws, nc):
    r = _next(ws)
    ws.merge_cells(f"A{r}:{get_column_letter(nc)}{r}")
    c = ws.cell(row=r, column=1, value="    ✔  No records this period")
    c.font = font(False, "595959", 9); c.fill = fill(C_LGREEN)
    c.alignment = aln("left"); c.border = bdr()
    ws.row_dimensions[r].height = 14

def drow(ws, vals, fmts=None, bold=False, bg=None):
    r = _next(ws)
    bg = bg or (C_LGREY if r % 2 == 0 else C_WHITE)
    for i, v in enumerate(vals, 1):
        fmt = fmts[i-1] if fmts and i-1 < len(fmts) else None
        ha  = "right" if (isinstance(v, (int,float)) and fmt in (NUM,INT)) else "left"
        wcell(ws, r, i, v, bold=bold, bg=bg, ha=ha, fmt=fmt)
    ws.row_dimensions[r].height = 15
    return r

def subtotal_row(ws, label, vals, fmts=None, bg=C_MSUB, label_col=1, data_start_col=2):
    """Write a subtotal / month-total row."""
    r = _next(ws)
    nc_total = label_col + len(vals)
    # label spans from col 1 to data_start_col-1
    if data_start_col > 1:
        ws.merge_cells(f"{get_column_letter(label_col)}{r}:{get_column_letter(data_start_col-1)}{r}")
    c = ws.cell(row=r, column=label_col, value=label)
    c.font = font(True, "FFFFFF", 9); c.fill = fill(bg)
    c.alignment = aln("left"); c.border = bdr()
    for i, v in enumerate(vals, data_start_col):
        fmt = fmts[i-data_start_col] if fmts and (i-data_start_col) < len(fmts) else None
        ha  = "right" if isinstance(v, (int,float)) else "left"
        cc  = ws.cell(row=r, column=i, value=v)
        cc.font = font(True, "FFFFFF", 9); cc.fill = fill(bg)
        cc.alignment = aln(ha); cc.border = bdr()
        if fmt: cc.number_format = fmt
    ws.row_dimensions[r].height = 17

def annual_total_row(ws, label, vals, fmts=None, data_start_col=2):
    r = _next(ws)
    ws.merge_cells(f"A{r}:{get_column_letter(data_start_col-1)}{r}")
    c = ws.cell(row=r, column=1, value=label)
    c.font = font(True, "FFFFFF", 11); c.fill = fill(C_ANN)
    c.alignment = aln("left"); c.border = bdr()
    for i, v in enumerate(vals, data_start_col):
        fmt = fmts[i-data_start_col] if fmts and (i-data_start_col) < len(fmts) else None
        cc  = ws.cell(row=r, column=i, value=v)
        cc.font = font(True, "FFFFFF", 11); cc.fill = fill(C_ANN)
        cc.alignment = aln("right"); cc.border = bdr()
        if fmt: cc.number_format = fmt
    ws.row_dimensions[r].height = 22

# ════════════════════════════════════════════════════════════════════════
#  SHEET SETUP HELPERS
# ════════════════════════════════════════════════════════════════════════
def make_sheet(wb, name, title, headers, widths, bg=C_DARK,
               trader_name="", gstin=""):
    ws = wb.create_sheet(name)
    sheet_title(ws, title, len(headers), bg=bg)
    trader_row(ws, trader_name, gstin, len(headers), row=2)
    col_headers(ws, headers, widths, row=3, bg=(C_AMEND if bg == C_AMEND else C_MED))
    _setr(ws, 4)
    return ws

# ════════════════════════════════════════════════════════════════════════
#  COLUMN DEFINITIONS
# ════════════════════════════════════════════════════════════════════════
def cols_b2b():
    H = ["Period","Buyer GSTIN","Buyer Name","Invoice No.","Invoice Date","Invoice Type",
         "Place of Supply","Rev.Chg","Rate (%)","Taxable Value (₹)",
         "IGST (₹)","CGST (₹)","SGST/UTGST (₹)","Total Tax (₹)","Invoice Value (₹)",
         "CFS","Updated By","C-Flag"]
    W = [14,18,32,14,12,14,10,7,7,16,14,14,14,14,16,6,8,6]
    F = [None,None,None,None,None,None,None,None,NUM,NUM,NUM,NUM,NUM,NUM,NUM,None,None,None]
    return H, W, F

def cols_cdnr():
    H = ["Period","Buyer GSTIN","Buyer Name","Note Type","Note No.","Note Date",
         "Note Value (₹)","Pl.of Supply","Pre-GST","Rate (%)",
         "Taxable Value (₹)","IGST (₹)","CGST (₹)","SGST/UTGST (₹)","Total Tax (₹)"]
    W = [14,18,32,12,14,12,14,10,7,7,16,14,14,14,14]
    F = [None]*7+[None,None,NUM,NUM,NUM,NUM,NUM,NUM]
    return H, W, F

def cols_b2cs():
    H = ["Period","Type","Pl.of Supply","Rate (%)","Taxable Value (₹)",
         "IGST (₹)","CGST (₹)","SGST/UTGST (₹)","Total Tax (₹)","E-Comm GSTIN"]
    W = [14,10,12,7,16,14,14,14,14,22]
    F = [None,None,None,NUM,NUM,NUM,NUM,NUM,NUM,None]
    return H, W, F

def cols_b2cl():
    H = ["Period","Pl.of Supply","Invoice No.","Invoice Date",
         "Invoice Value (₹)","Rate (%)","Taxable Value (₹)","IGST (₹)","E-Comm GSTIN"]
    W = [14,12,14,12,16,7,16,14,22]
    F = [None,None,None,None,NUM,NUM,NUM,NUM,None]
    return H, W, F

def cols_exp():
    H = ["Period","Export Type","Invoice No.","Invoice Date","Invoice Value (₹)",
         "Port Code","Shipping Bill No.","Shipping Bill Date",
         "Rate (%)","Taxable Value (₹)","IGST (₹)"]
    W = [14,14,14,12,16,10,16,14,7,16,14]
    F = [None]*4+[NUM]+[None,None,None,NUM,NUM,NUM]
    return H, W, F

def cols_nil():
    H = ["Period","Supply Type","Nil Rated (₹)","Exempt (₹)","Non-GST (₹)","Total (₹)"]
    W = [14,20,16,16,16,16]; F = [None,None,NUM,NUM,NUM,NUM]
    return H, W, F

def cols_hsn():
    H = ["Period","HSN Code","Description","UOM","Quantity","Total Value (₹)",
         "Taxable Value (₹)","IGST (₹)","CGST (₹)","SGST/UTGST (₹)","Total Tax (₹)"]
    W = [14,12,28,8,12,15,16,14,14,14,14]
    F = [None,None,None,None,INT,NUM,NUM,NUM,NUM,NUM,NUM]
    return H, W, F

def cols_doc():
    H = ["Period","Doc Type","Description","From Sr.No.","To Sr.No.",
         "Total Issued","Cancelled","Net Issued"]
    W = [14,10,38,14,14,14,12,12]; F = [None,None,None,None,None,INT,INT,INT]
    return H, W, F

def cols_cdnur():
    H = ["Period","Note Type","Supply Type","Note No.","Note Date",
         "Note Value (₹)","Rate (%)","Taxable Value (₹)","IGST (₹)"]
    W = [14,12,14,14,12,14,7,16,14]; F = [None]*5+[NUM,NUM,NUM,NUM]
    return H, W, F

def cols_adv():
    H = ["Period","Pl.of Supply","Rate (%)","Amount (₹)",
         "IGST (₹)","CGST (₹)","SGST/UTGST (₹)","Total Tax (₹)","E-Comm GSTIN"]
    W = [14,12,7,16,14,14,14,14,22]; F = [None,None,NUM,NUM,NUM,NUM,NUM,NUM,None]
    return H, W, F

def cols_eco():
    H = ["Period","ECO GSTIN","Supply Type","Net Value (₹)",
         "IGST (₹)","CGST (₹)","SGST/UTGST (₹)","Total Tax (₹)"]
    W = [14,22,20,16,14,14,14,14]; F = [None,None,None,NUM,NUM,NUM,NUM,NUM]
    return H, W, F

def cols_sec():
    H = ["Period","ECO GSTIN","Rate (%)","Net Value (₹)",
         "IGST (₹)","CGST (₹)","SGST/UTGST (₹)","Total Tax (₹)"]
    W = [14,22,7,16,14,14,14,14]; F = [None,None,NUM,NUM,NUM,NUM,NUM,NUM]
    return H, W, F

def cols_amd_b2b():
    H = ["Period","Tbl","Buyer GSTIN","Orig Inv No.","Amend Inv No.",
         "Invoice Date","Invoice Value (₹)","Rate (%)","Taxable Value (₹)",
         "IGST (₹)","CGST (₹)","SGST/UTGST (₹)","Total Tax (₹)"]
    W = [14,6,18,16,16,12,16,7,16,14,14,14,14]
    F = [None]*6+[NUM,NUM,NUM,NUM,NUM,NUM,NUM]
    return H, W, F

def cols_amd_b2cl():
    H = ["Period","Pl.of Supply","Orig Inv No.","Amend Inv No.",
         "Invoice Date","Invoice Value (₹)","Rate (%)","Taxable Value (₹)","IGST (₹)"]
    W = [14,12,16,16,12,16,7,16,14]; F = [None]*5+[NUM,NUM,NUM,NUM]
    return H, W, F

def cols_amd_exp():
    H = ["Period","Export Type","Orig Inv No.","Amend Inv No.",
         "Invoice Date","Invoice Value (₹)","Rate (%)","Taxable Value (₹)","IGST (₹)"]
    W = [14,14,16,16,12,16,7,16,14]; F = [None]*5+[NUM,NUM,NUM,NUM]
    return H, W, F

def cols_amd_cdnr():
    H = ["Period","Buyer GSTIN","Note Type","Orig Note No.","Amend Note No.",
         "Note Date","Note Value (₹)","Rate (%)","Taxable Value (₹)",
         "IGST (₹)","CGST (₹)","SGST/UTGST (₹)","Total Tax (₹)"]
    W = [14,18,12,16,16,12,14,7,16,14,14,14,14]
    F = [None]*6+[NUM,NUM,NUM,NUM,NUM,NUM,NUM]
    return H, W, F

def cols_amd_cdnur():
    H = ["Period","Note Type","Supply Type","Orig Note No.","Amend Note No.",
         "Note Date","Note Value (₹)","Rate (%)","Taxable Value (₹)","IGST (₹)"]
    W = [14,12,14,16,16,12,14,7,16,14]; F = [None]*6+[NUM,NUM,NUM,NUM]
    return H, W, F

def cols_amd_b2cs():
    H = ["Period","Type","Orig POS","Orig Rate","Amend POS","Amend Rate",
         "Taxable Value (₹)","IGST (₹)","CGST (₹)","SGST/UTGST (₹)","Total Tax (₹)","E-Comm GSTIN"]
    W = [14,10,12,7,12,7,16,14,14,14,14,22]; F = [None]*6+[NUM,NUM,NUM,NUM,NUM,None]
    return H, W, F

# ════════════════════════════════════════════════════════════════════════
#  DATA WRITERS — return dict of totals
# ════════════════════════════════════════════════════════════════════════

def write_b2b(ws, d, pl, fmts, nc):
    """B2B: groups by GSTIN, writes per-GSTIN subtotal, then month total."""
    # Collect all rows grouped by buyer GSTIN
    by_gstin = defaultdict(list)
    for ent in d.get("b2b", []):
        ctin = ent.get("ctin", ""); cfs = ent.get("cfs", "")
        for inv in ent.get("inv", []):
            typ = INV_TYPE.get(inv.get("inv_typ","R"), inv.get("inv_typ",""))
            val = float(inv.get("val", 0) or 0)
            for it in inv.get("itms", []):
                det = it.get("itm_det", {})
                rt  = det.get("rt", 0)
                tv  = float(det.get("txval", 0) or 0)
                ig  = float(det.get("iamt",  0) or 0)
                cg  = float(det.get("camt",  0) or 0)
                sg  = float(det.get("samt",  0) or 0)
                by_gstin[ctin].append(
                    [pl, ctin, "",   # "" = Buyer Name placeholder (col 3)
                     inv.get("inum",""), inv.get("idt",""), typ,
                     inv.get("pos",""), inv.get("rchrg","N"), rt,
                     tv, ig, cg, sg, ig+cg+sg, val, cfs,
                     inv.get("updby",""), inv.get("cflag","")]
                )

    if not by_gstin:
        no_data_row(ws, nc)
        return {"tv":0,"ig":0,"cg":0,"sg":0,"val":0,"n":0}

    # ── Bulk-fetch buyer names (cache first, portal only for unknowns) ────────
    if _CACHE_ENABLED and _NAME_CACHE:
        name_map = _NAME_CACHE.get_bulk(list(by_gstin.keys()), show_progress=False)
        for rows_list in by_gstin.values():
            for row in rows_list:
                row[2] = name_map.get(row[1], "")   # col index 2 = Buyer Name

    m_tv=m_ig=m_cg=m_sg=m_val=0.0; m_n=0
    gstin_totals = []   # for FY GSTIN sheet

    for ctin in sorted(by_gstin):
        g_tv=g_ig=g_cg=g_sg=g_val=0.0
        for row_vals in by_gstin[ctin]:
            drow(ws, row_vals, fmts=fmts)
            tv,ig,cg,sg,val = row_vals[9],row_vals[10],row_vals[11],row_vals[12],row_vals[14]
            g_tv+=tv; g_ig+=ig; g_cg+=cg; g_sg+=sg; g_val+=val; m_n+=1

        # Per-GSTIN subtotal (now shifted right by 1 due to Buyer Name col)
        bname = name_map.get(ctin,"") if (_CACHE_ENABLED and _NAME_CACHE) else ""
        label = f"  Sub-Total: {ctin}{(' — '+bname) if bname else ''} ({pl})"
        subtotal_row(ws, label,
                     [g_tv, g_ig, g_cg, g_sg, g_ig+g_cg+g_sg, g_val],
                     fmts=[NUM,NUM,NUM,NUM,NUM,NUM],
                     bg=C_GSUB, label_col=1, data_start_col=10)
        m_tv+=g_tv; m_ig+=g_ig; m_cg+=g_cg; m_sg+=g_sg; m_val+=g_val
        gstin_totals.append((ctin, g_tv, g_ig, g_cg, g_sg, g_val))

    # Month total
    subtotal_row(ws,
                 f"  ✦ MONTH TOTAL — {pl}",
                 [m_tv, m_ig, m_cg, m_sg, m_ig+m_cg+m_sg, m_val],
                 fmts=[NUM,NUM,NUM,NUM,NUM,NUM],
                 bg=C_MSUB, label_col=1, data_start_col=10)

    return {"tv":m_tv,"ig":m_ig,"cg":m_cg,"sg":m_sg,"val":m_val,
            "n":m_n,"gstin_totals":gstin_totals}


def write_cdnr(ws, d, pl, fmts, nc):
    by_gstin = defaultdict(list)
    for ent in d.get("cdnr", []):
        ctin = ent.get("ctin","")
        for nt in ent.get("nt",[]):
            ntty=nt.get("ntty","C"); tv=ig=cg=sg=0.0; rt=0
            for it in nt.get("itms",[]):
                det=it.get("itm_det",{})
                tv+=float(det.get("txval",0)or 0); ig+=float(det.get("iamt",0)or 0)
                cg+=float(det.get("camt",0)or 0);  sg+=float(det.get("samt",0)or 0)
                rt=det.get("rt",rt)
            by_gstin[ctin].append(
                [pl, ctin, "",   # "" = Buyer Name placeholder (col 3)
                 "Credit Note" if ntty=="C" else "Debit Note",
                 nt.get("nt_num",""), nt.get("nt_dt",""),
                 float(nt.get("val",0)or 0), nt.get("pos",""), nt.get("p_gst","N"),
                 rt, tv, ig, cg, sg, ig+cg+sg]
            )

    if not by_gstin:
        no_data_row(ws,nc)
        return {"cr":0,"dr":0,"ig":0,"cg":0,"sg":0}

    # Bulk-fetch buyer names
    if _CACHE_ENABLED and _NAME_CACHE:
        name_map = _NAME_CACHE.get_bulk(list(by_gstin.keys()), show_progress=False)
        for rows_list in by_gstin.values():
            for row in rows_list:
                row[2] = name_map.get(row[1], "")

    m_tv=m_ig=m_cg=m_sg=0.0; m_cr=m_dr=0.0
    for ctin in sorted(by_gstin):
        g_tv=g_ig=g_cg=g_sg=0.0
        for row_vals in by_gstin[ctin]:
            drow(ws, row_vals, fmts=fmts)
            nt_type=row_vals[3]; tv=row_vals[10]; ig=row_vals[11]
            cg=row_vals[12]; sg=row_vals[13]
            g_tv+=tv; g_ig+=ig; g_cg+=cg; g_sg+=sg
            if "Credit" in nt_type: m_cr+=tv
            else: m_dr+=tv
        bname = name_map.get(ctin,"") if (_CACHE_ENABLED and _NAME_CACHE) else ""
        label = f"  Sub-Total: {ctin}{(' — '+bname) if bname else ''} ({pl})"
        subtotal_row(ws, label,
                     [g_tv,g_ig,g_cg,g_sg,g_ig+g_cg+g_sg],
                     fmts=[NUM,NUM,NUM,NUM,NUM],
                     bg=C_GSUB, label_col=1, data_start_col=11)
        m_tv+=g_tv; m_ig+=g_ig; m_cg+=g_cg; m_sg+=g_sg

    subtotal_row(ws, f"  ✦ MONTH TOTAL — {pl}",
                 [m_tv,m_ig,m_cg,m_sg,m_ig+m_cg+m_sg],
                 fmts=[NUM,NUM,NUM,NUM,NUM],
                 bg=C_MSUB, label_col=1, data_start_col=11)
    return {"cr":m_cr,"dr":m_dr,"ig":m_ig,"cg":m_cg,"sg":m_sg}


def _simple_write(ws, rows_list, fmts, nc, num_cols,
                  tv_col, ig_col, cg_col=None, sg_col=None,
                  month_label=""):
    """Generic writer for sheets without GSTIN grouping."""
    if not rows_list:
        no_data_row(ws, nc)
        tv=ig=cg=sg=0.0
    else:
        tv=ig=cg=sg=0.0
        for rv in rows_list:
            drow(ws, rv, fmts=fmts)
            tv += rv[tv_col] if len(rv)>tv_col else 0
            ig += rv[ig_col] if len(rv)>ig_col else 0
            if cg_col: cg += rv[cg_col] if len(rv)>cg_col else 0
            if sg_col: sg += rv[sg_col] if len(rv)>sg_col else 0
        # month subtotal: put values starting after label
        sub_vals = [tv, ig]
        sub_fmts = [NUM, NUM]
        if cg_col: sub_vals.append(cg); sub_fmts.append(NUM)
        if sg_col: sub_vals.append(sg); sub_fmts.append(NUM)
        sub_vals.append(ig+(cg or 0)+(sg or 0)); sub_fmts.append(NUM)
        subtotal_row(ws, f"  ✦ MONTH TOTAL — {month_label}",
                     sub_vals, fmts=sub_fmts,
                     bg=C_MSUB, label_col=1, data_start_col=tv_col+1)
    return tv, ig, cg, sg


def write_b2cs(ws, d, pl, fmts, nc):
    rows=[]
    for rec in d.get("b2cs",[]):
        tv=float(rec.get("txval",0)or 0); ig=float(rec.get("iamt",0)or 0)
        cg=float(rec.get("camt",0)or 0);  sg=float(rec.get("samt",0)or 0)
        rows.append([pl,rec.get("typ",""),rec.get("pos",""),rec.get("rt",0),
                     tv,ig,cg,sg,ig+cg+sg,rec.get("etin","")])
    tv,ig,cg,sg=_simple_write(ws,rows,fmts,nc,10,4,5,6,7,pl)
    return {"tv":tv,"ig":ig,"cg":cg,"sg":sg}

def write_b2cl(ws, d, pl, fmts, nc):
    rows=[]
    for rec in d.get("b2cl",[]):
        pos=rec.get("pos","")
        for inv in rec.get("inv",[]):
            tv=ig=0.0; rt=0
            for it in inv.get("itms",[]):
                det=it.get("itm_det",{})
                tv+=float(det.get("txval",0)or 0); ig+=float(det.get("iamt",0)or 0); rt=det.get("rt",rt)
            rows.append([pl,pos,inv.get("inum",""),inv.get("idt",""),
                         float(inv.get("val",0)or 0),rt,tv,ig,inv.get("etin","")])
    tv,ig,_,__=_simple_write(ws,rows,fmts,nc,9,6,7,month_label=pl)
    return {"tv":tv,"ig":ig}

def write_exp(ws, d, pl, fmts, nc):
    rows=[]
    for exp in d.get("exp",[]):
        et=exp.get("exp_typ","")
        for inv in exp.get("inv",[]):
            tv=ig=0.0; rt=0
            for it in inv.get("itms",[]):
                det=it.get("itm_det",{})
                tv+=float(det.get("txval",0)or 0); ig+=float(det.get("iamt",0)or 0); rt=det.get("rt",rt)
            rows.append([pl,et,inv.get("inum",""),inv.get("idt",""),
                         float(inv.get("val",0)or 0),inv.get("pcode",""),
                         inv.get("sbnum",""),inv.get("sbdt",""),rt,tv,ig])
    tv,ig,_,__=_simple_write(ws,rows,fmts,nc,11,9,10,month_label=pl)
    return {"tv":tv,"ig":ig}

def write_nil(ws, d, pl, fmts, nc):
    raw=d.get("nil_sup") or d.get("nil",{})
    nil_list=raw if isinstance(raw,list) else raw.get("inv",[])
    rows=[]
    for rec in nil_list:
        nv=float(rec.get("nil_amt",  rec.get("nil",0))  or 0)
        ev=float(rec.get("expt_amt", rec.get("expt",0)) or 0)
        gv=float(rec.get("ngsup_amt",rec.get("ngsup",0))or 0)
        if nv+ev+gv==0: continue
        rows.append([pl,rec.get("sply_ty",""),nv,ev,gv,nv+ev+gv])
    if not rows: no_data_row(ws,nc); return
    for rv in rows: drow(ws,rv,fmts=fmts)
    tot_n=sum(r[2] for r in rows); tot_e=sum(r[3] for r in rows); tot_g=sum(r[4] for r in rows)
    subtotal_row(ws, f"  ✦ MONTH TOTAL — {pl}",
                 [tot_n,tot_e,tot_g,tot_n+tot_e+tot_g], fmts=[NUM,NUM,NUM,NUM],
                 bg=C_MSUB, label_col=1, data_start_col=3)

def write_hsn(ws, d, pl, fmts, nc):
    raw=d.get("hsn",[])
    items=raw.get("data",[]) if isinstance(raw,dict) else (raw if isinstance(raw,list) else [])
    rows=[]
    for rec in items:
        tv=float(rec.get("txval",0)or 0); ig=float(rec.get("iamt",0)or 0)
        cg=float(rec.get("camt",0)or 0);  sg=float(rec.get("samt",0)or 0)
        val=float(rec.get("val",0)or 0);   qty=float(rec.get("qty",0)or 0)
        rows.append([pl,rec.get("hsn_sc",""),rec.get("desc",""),rec.get("uqc",""),
                     qty,val,tv,ig,cg,sg,ig+cg+sg])
    if not rows: no_data_row(ws,nc); return {"tv":0,"ig":0,"cg":0,"sg":0}
    for rv in rows: drow(ws,rv,fmts=fmts)
    tv=sum(r[6] for r in rows); ig=sum(r[7] for r in rows)
    cg=sum(r[8] for r in rows); sg=sum(r[9] for r in rows)
    val=sum(r[5] for r in rows)
    subtotal_row(ws, f"  ✦ MONTH TOTAL — {pl}",
                 [val,tv,ig,cg,sg,ig+cg+sg], fmts=[NUM,NUM,NUM,NUM,NUM,NUM],
                 bg=C_MSUB, label_col=1, data_start_col=6)
    return {"tv":tv,"ig":ig,"cg":cg,"sg":sg}

def write_docs(ws, d, pl, fmts, nc):
    doc_issue=d.get("doc_issue",{}); rows=0
    tot_i=tot_c=tot_n=0
    for det in doc_issue.get("doc_det",[]):
        dn=det.get("doc_num",det.get("ty_cd",""))
        desc=DOC_TYPES.get(int(dn) if str(dn).isdigit() else 0, f"Type {dn}")
        for doc in det.get("docs",[]):
            tn=int(doc.get("totnum",doc.get("num",0))or 0)
            ca=int(doc.get("cancel",0)or 0)
            nt=int(doc.get("net_issue",tn-ca)or(tn-ca))
            drow(ws,[pl,dn,desc,doc.get("from",""),doc.get("to",""),tn,ca,nt],fmts=fmts)
            tot_i+=tn; tot_c+=ca; tot_n+=nt; rows+=1
    if rows==0: no_data_row(ws,nc); return
    subtotal_row(ws, f"  ✦ MONTH TOTAL — {pl}",
                 [tot_i,tot_c,tot_n], fmts=[INT,INT,INT],
                 bg=C_MSUB, label_col=1, data_start_col=6)

def write_cdnur(ws, d, pl, fmts, nc):
    rows=[]
    for nt in d.get("cdnur",[]):
        tv=float(nt.get("txval",0)or 0); ig=float(nt.get("iamt",0)or 0)
        rows.append([pl,nt.get("ntty",""),nt.get("typ",""),
                     nt.get("nt_num",""),nt.get("nt_dt",""),
                     float(nt.get("val",0)or 0),nt.get("rt",0),tv,ig])
    tv,ig,_,__=_simple_write(ws,rows,fmts,nc,9,7,8,month_label=pl)
    return {"tv":tv,"ig":ig}

def write_adv(ws, items, pl, fmts, nc, month_label):
    rows=[]
    for rec in items:
        for it in rec.get("itms",[]):
            det=it.get("itm_det",{})
            amt=float(det.get("ad_amount",det.get("ad_amt",0))or 0)
            ig=float(det.get("iamt",0)or 0); cg=float(det.get("camt",0)or 0)
            sg=float(det.get("samt",0)or 0); rt=det.get("rt",0)
            rows.append([pl,rec.get("pos",""),rt,amt,ig,cg,sg,ig+cg+sg,rec.get("etin","")])
    _simple_write(ws,rows,fmts,nc,9,3,4,5,6,pl)

def write_eco(ws, items, pl, fmts, nc):
    rows=[]
    for rec in items:
        nv=float(rec.get("net_val",rec.get("txval",0))or 0)
        ig=float(rec.get("iamt",0)or 0); cg=float(rec.get("camt",0)or 0); sg=float(rec.get("samt",0)or 0)
        rows.append([pl,rec.get("etin",rec.get("gstin","")),
                     rec.get("sup_ty",rec.get("spty","")),nv,ig,cg,sg,ig+cg+sg])
    _simple_write(ws,rows,fmts,nc,8,3,4,5,6,pl)

def write_sec95(ws, items, pl, fmts, nc):
    rows=[]
    for rec in items:
        nv=float(rec.get("txval",rec.get("net_val",0))or 0)
        ig=float(rec.get("iamt",0)or 0); cg=float(rec.get("camt",0)or 0); sg=float(rec.get("samt",0)or 0)
        rows.append([pl,rec.get("etin",rec.get("gstin","")),rec.get("rt",0),nv,ig,cg,sg,ig+cg+sg])
    _simple_write(ws,rows,fmts,nc,8,3,4,5,6,pl)

def write_amd_b2b(ws, d, pl, fmts, nc):
    rows=0
    for ent in d.get("b2ba",[]):
        ctin=ent.get("ctin","")
        for inv in ent.get("inv",[]):
            tv=ig=cg=sg=0.0; rt=0
            for it in inv.get("itms",[]):
                det=it.get("itm_det",{})
                tv+=float(det.get("txval",0)or 0); ig+=float(det.get("iamt",0)or 0)
                cg+=float(det.get("camt",0)or 0);  sg+=float(det.get("samt",0)or 0); rt=det.get("rt",rt)
            drow(ws,[pl,"B2BA",ctin,inv.get("oinum",""),inv.get("inum",""),
                     inv.get("idt",""),float(inv.get("val",0)or 0),rt,tv,ig,cg,sg,ig+cg+sg],
                 fmts=fmts,bg=C_LPUR); rows+=1
    if rows==0: no_data_row(ws,nc)

def write_amd_b2cl(ws, d, pl, fmts, nc):
    rows=0
    for rec in d.get("b2cla",[]):
        for inv in rec.get("inv",[]):
            tv=ig=0.0; rt=0
            for it in inv.get("itms",[]):
                det=it.get("itm_det",{})
                tv+=float(det.get("txval",0)or 0); ig+=float(det.get("iamt",0)or 0); rt=det.get("rt",rt)
            drow(ws,[pl,rec.get("pos",""),inv.get("oinum",""),inv.get("inum",""),
                     inv.get("idt",""),float(inv.get("val",0)or 0),rt,tv,ig],
                 fmts=fmts,bg=C_LPUR); rows+=1
    if rows==0: no_data_row(ws,nc)

def write_amd_exp(ws, d, pl, fmts, nc):
    rows=0
    for exp in d.get("expa",[]):
        for inv in exp.get("inv",[]):
            tv=ig=0.0; rt=0
            for it in inv.get("itms",[]):
                det=it.get("itm_det",{})
                tv+=float(det.get("txval",0)or 0); ig+=float(det.get("iamt",0)or 0); rt=det.get("rt",rt)
            drow(ws,[pl,exp.get("exp_typ",""),inv.get("oinum",""),inv.get("inum",""),
                     inv.get("idt",""),float(inv.get("val",0)or 0),rt,tv,ig],
                 fmts=fmts,bg=C_LPUR); rows+=1
    if rows==0: no_data_row(ws,nc)

def write_amd_cdnr(ws, d, pl, fmts, nc):
    rows=0
    for ent in d.get("cdnra",[]):
        for nt in ent.get("nt",[]):
            tv=ig=cg=sg=0.0
            for it in nt.get("itms",[]):
                det=it.get("itm_det",{})
                tv+=float(det.get("txval",0)or 0); ig+=float(det.get("iamt",0)or 0)
                cg+=float(det.get("camt",0)or 0);  sg+=float(det.get("samt",0)or 0)
            drow(ws,[pl,ent.get("ctin",""),nt.get("ntty",""),nt.get("ont_num",""),
                     nt.get("nt_num",""),nt.get("nt_dt",""),float(nt.get("val",0)or 0),
                     0,tv,ig,cg,sg,ig+cg+sg],fmts=fmts,bg=C_LPUR); rows+=1
    if rows==0: no_data_row(ws,nc)

def write_amd_cdnur(ws, d, pl, fmts, nc):
    rows=0
    for nt in d.get("cdnura",[]):
        tv=float(nt.get("txval",0)or 0); ig=float(nt.get("iamt",0)or 0)
        drow(ws,[pl,nt.get("ntty",""),nt.get("typ",""),nt.get("ont_num",""),
                 nt.get("nt_num",""),nt.get("nt_dt",""),
                 float(nt.get("val",0)or 0),nt.get("rt",0),tv,ig],fmts=fmts,bg=C_LPUR); rows+=1
    if rows==0: no_data_row(ws,nc)

def write_amd_b2cs(ws, d, pl, fmts, nc):
    rows=0
    for rec in d.get("b2csa",[]):
        tv=float(rec.get("txval",0)or 0); ig=float(rec.get("iamt",0)or 0)
        cg=float(rec.get("camt",0)or 0);  sg=float(rec.get("samt",0)or 0)
        drow(ws,[pl,rec.get("typ",""),rec.get("opos",""),rec.get("ort",""),
                 rec.get("pos",""),rec.get("rt",""),tv,ig,cg,sg,ig+cg+sg,rec.get("etin","")],
             fmts=fmts,bg=C_LPUR); rows+=1
    if rows==0: no_data_row(ws,nc)

# ════════════════════════════════════════════════════════════════════════
#  SHEET 25 — FY GSTIN-WISE ANNUAL SUMMARY
# ════════════════════════════════════════════════════════════════════════
def build_gstin_sheet(wb, fy_gstin_data, trader_name, gstin, fy_label):
    ws = wb.create_sheet("25_GSTIN_Annual_Summary")
    H  = ["Buyer GSTIN","Jan-Mar Total",
          "April","May","June","July","August","September",
          "October","November","December","January","February","March",
          "FY Total Taxable (₹)","FY Total IGST (₹)","FY Total CGST (₹)",
          "FY Total SGST (₹)","FY Total Tax (₹)","FY Invoice Value (₹)"]
    W  = [20,14,13,13,13,13,13,13,13,13,13,13,13,13,18,17,17,17,17,18]

    sheet_title(ws, f"FY GSTIN-wise Annual Summary — {fy_label}", len(H))
    trader_row(ws, trader_name, gstin, len(H), row=2)
    col_headers(ws, H, W, row=3)
    _setr(ws, 4)
    ws.freeze_panes = "B4"

    month_order = ["April","May","June","July","August","September",
                   "October","November","December","January","February","March"]

    # Aggregate: {gstin: {month: {tv,ig,cg,sg,val}}}
    all_gstins = sorted(fy_gstin_data.keys())
    fy_totals  = [0.0]*6  # tv,ig,cg,sg,tax,val

    for ctin in all_gstins:
        months = fy_gstin_data[ctin]
        row_tv  = sum(months.get(m,{}).get("tv",0) for m in month_order)
        row_ig  = sum(months.get(m,{}).get("ig",0) for m in month_order)
        row_cg  = sum(months.get(m,{}).get("cg",0) for m in month_order)
        row_sg  = sum(months.get(m,{}).get("sg",0) for m in month_order)
        row_val = sum(months.get(m,{}).get("val",0) for m in month_order)
        row_tax = row_ig+row_cg+row_sg

        month_tvs = [months.get(m,{}).get("tv",0.0) for m in month_order]

        r  = _next(ws)
        bg = C_LGREY if r%2==0 else C_WHITE
        wcell(ws,r,1,ctin,bg=bg,ha="left")
        # Jan-Mar sub-total (months 10,11,12 = Jan,Feb,Mar)
        jan_mar=sum(months.get(m,{}).get("tv",0) for m in ["January","February","March"])
        wcell(ws,r,2,jan_mar,bg=bg,ha="right",fmt=NUM)
        for ci,v in enumerate(month_tvs,3):
            wcell(ws,r,ci,v,bg=bg,ha="right",fmt=NUM)
        for ci,v in enumerate([row_tv,row_ig,row_cg,row_sg,row_tax,row_val],15):
            wcell(ws,r,ci,v,bg=bg,ha="right",fmt=NUM,bold=True)
        ws.row_dimensions[r].height=16

        for i,v in enumerate([row_tv,row_ig,row_cg,row_sg,row_tax,row_val]):
            fy_totals[i]+=v

    # FY Grand Total
    r=_next(ws)
    wcell(ws,r,1,"FY GRAND TOTAL",bold=True,bg=C_ANN,fc="FFFFFF")
    wcell(ws,r,2,"",bg=C_ANN)
    for ci in range(3,15):
        ws.cell(row=r,column=ci).fill=fill(C_ANN); ws.cell(row=r,column=ci).border=bdr()
    for ci,v in enumerate(fy_totals,15):
        wcell(ws,r,ci,v,bold=True,bg=C_ANN,fc="FFFFFF",ha="right",fmt=NUM)
    ws.row_dimensions[r].height=20
    return ws

# ════════════════════════════════════════════════════════════════════════
#  SHEET 26 — MASTER ALL INVOICES (B2B flat, all months)
# ════════════════════════════════════════════════════════════════════════
def build_master_sheet(wb, master_rows, trader_name, gstin, fy_label):
    ws = wb.create_sheet("26_Master_All_Invoices")
    H  = ["Sr.No","Period","Supply Type","Buyer GSTIN","Invoice No.","Invoice Date",
          "Invoice Type","Pl.of Supply","Rev.Chg","Rate (%)","Taxable Value (₹)",
          "IGST (₹)","CGST (₹)","SGST/UTGST (₹)","Total Tax (₹)","Invoice Value (₹)",
          "Note Type","Note No.","Note Date"]
    W  = [6,14,14,20,14,12,14,10,7,7,16,14,14,14,14,16,12,14,12]

    sheet_title(ws, f"Master — All Invoices & Credit/Debit Notes ({fy_label})", len(H))
    trader_row(ws, trader_name, gstin, len(H), row=2)
    col_headers(ws, H, W, row=3)
    _setr(ws, 4)

    # master_rows columns (0-indexed):
    # 0=period, 1=supply_type, 2=buyer_gstin, 3=inv_no, 4=inv_dt,
    # 5=inv_type, 6=pos, 7=rev_chg, 8=rate, 9=taxable_val,
    # 10=igst, 11=cgst, 12=sgst, 13=total_tax, 14=inv_val,
    # 15=note_type, 16=note_no, 17=note_dt
    TV,IG,CG,SG,VAL = 9,10,11,12,14

    sr=0; prev_month=""; m_tv=m_ig=m_cg=m_sg=m_val=0.0
    fy_tv=fy_ig=fy_cg=fy_sg=fy_val=0.0

    # col numbers in Excel (1-indexed, sr adds 1):
    # Sr=1, period=2, supply=3, gstin=4, inv_no=5, inv_dt=6,
    # inv_type=7, pos=8, rchrg=9, rate=10, taxable=11,
    # igst=12, cgst=13, sgst=14, total_tax=15, inv_val=16,
    # note_type=17, note_no=18, note_dt=19
    NUM_COLS = {11,12,13,14,15,16}   # 1-indexed Excel columns that need NUM format

    for row_data in master_rows:
        month = row_data[0]   # period is index 0
        if prev_month and month != prev_month:
            subtotal_row(ws, f"  ✦ MONTH TOTAL — {prev_month}",
                         [m_tv, m_ig, m_cg, m_sg, m_ig+m_cg+m_sg, m_val],
                         fmts=[NUM,NUM,NUM,NUM,NUM,NUM],
                         bg=C_MSUB, label_col=1, data_start_col=11)
            m_tv=m_ig=m_cg=m_sg=m_val=0.0

        sr+=1; prev_month=month
        bg = C_LGREY if sr%2==0 else C_WHITE
        full_row = [sr] + list(row_data)   # Excel col 1=Sr, col 2=period ...
        r = _next(ws)
        for ci, v in enumerate(full_row, 1):
            fmt = NUM if ci in NUM_COLS else None
            # Force numeric where expected
            if fmt == NUM and not isinstance(v, (int, float)):
                try: v = float(v)
                except: v = 0.0
            ha = "right" if fmt == NUM else "left"
            wcell(ws, r, ci, v, bg=bg, ha=ha, fmt=fmt)
        ws.row_dimensions[r].height = 15

        tv  = float(row_data[TV]  or 0)
        ig  = float(row_data[IG]  or 0)
        cg  = float(row_data[CG]  or 0)
        sg  = float(row_data[SG]  or 0)
        val = float(row_data[VAL] or 0)
        m_tv+=tv; m_ig+=ig; m_cg+=cg; m_sg+=sg; m_val+=val
        fy_tv+=tv; fy_ig+=ig; fy_cg+=cg; fy_sg+=sg; fy_val+=val

    if prev_month:
        subtotal_row(ws, f"  ✦ MONTH TOTAL — {prev_month}",
                     [m_tv, m_ig, m_cg, m_sg, m_ig+m_cg+m_sg, m_val],
                     fmts=[NUM,NUM,NUM,NUM,NUM,NUM],
                     bg=C_MSUB, label_col=1, data_start_col=11)

    annual_total_row(ws, f"  ★ FY ANNUAL TOTAL — {fy_label}  ({sr} records)",
                     [fy_tv, fy_ig, fy_cg, fy_sg, fy_ig+fy_cg+fy_sg, fy_val],
                     fmts=[NUM,NUM,NUM,NUM,NUM,NUM], data_start_col=11)
    return ws

# ════════════════════════════════════════════════════════════════════════
#  RECON ANNUAL SUMMARY
# ════════════════════════════════════════════════════════════════════════
def build_recon(wb, fy_label, trader_name, gstin, months_data):
    ws = wb.create_sheet("24_RECON_Annual_Summary")
    ws.sheet_view.showGridLines = False
    nc = 16

    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    c=ws["A1"]; c.value=f"GSTR-1 Annual Reconciliation — {fy_label}"
    c.font=font(True,"FFFFFF",12); c.fill=fill(C_DARK)
    c.alignment=aln("center"); c.border=bdr(); ws.row_dimensions[1].height=28

    trader_row(ws, trader_name, gstin, nc, row=2)

    col_hdrs=["Month","B2B Taxable (₹)","B2B IGST (₹)","B2B CGST (₹)","B2B SGST (₹)",
              "B2CS Taxable (₹)","B2CL Taxable (₹)","Exports Taxable (₹)",
              "CDNR Credit (₹)","CDNR Debit (₹)","HSN Taxable (₹)",
              "Total IGST (₹)","Total CGST (₹)","Total SGST (₹)","Total Tax (₹)","Invoice Value (₹)"]
    widths=[16,15,13,13,13,15,15,15,14,13,15,13,13,13,13,15]

    for i,(h,w) in enumerate(zip(col_hdrs,widths),1):
        c=ws.cell(row=3,column=i,value=h)
        c.font=font(True,"FFFFFF",9); c.fill=fill(C_MED)
        c.alignment=aln("center",True); c.border=bdr()
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[3].height=32; ws.freeze_panes="A4"
    _setr(ws,4)

    month_order=["April","May","June","July","August","September",
                 "October","November","December","January","February","March"]
    ann={k:0.0 for k in["b2b_tv","b2b_ig","b2b_cg","b2b_sg","b2cs_tv","b2cl_tv",
                          "exp_tv","cdnr_cr","cdnr_dr","hsn_tv",
                          "tot_ig","tot_cg","tot_sg","tot_tax","inv_val"]}

    for mname in month_order:
        m=months_data.get(mname)
        if not m: continue
        r=_next(ws); bg=C_LGREY if r%2==0 else C_WHITE
        vals=[mname,m.get("b2b_tv",0),m.get("b2b_ig",0),m.get("b2b_cg",0),m.get("b2b_sg",0),
              m.get("b2cs_tv",0),m.get("b2cl_tv",0),m.get("exp_tv",0),
              m.get("cdnr_cr",0),m.get("cdnr_dr",0),m.get("hsn_tv",0),
              m.get("tot_ig",0),m.get("tot_cg",0),m.get("tot_sg",0),
              m.get("tot_tax",0),m.get("inv_val",0)]
        for i,v in enumerate(vals,1):
            wcell(ws,r,i,v,bg=bg,ha="right" if isinstance(v,float) else "left",
                  fmt=NUM if isinstance(v,float) else None)
        ws.row_dimensions[r].height=16
        for k in ann: ann[k]+=m.get(k,0)

    # Annual total
    r=_next(ws)
    ann_vals=["★ FY ANNUAL TOTAL",ann["b2b_tv"],ann["b2b_ig"],ann["b2b_cg"],ann["b2b_sg"],
              ann["b2cs_tv"],ann["b2cl_tv"],ann["exp_tv"],ann["cdnr_cr"],ann["cdnr_dr"],
              ann["hsn_tv"],ann["tot_ig"],ann["tot_cg"],ann["tot_sg"],ann["tot_tax"],ann["inv_val"]]
    for i,v in enumerate(ann_vals,1):
        c=ws.cell(row=r,column=i,value=v)
        c.font=font(True,"FFFFFF",10); c.fill=fill(C_ANN)
        c.alignment=aln("right" if isinstance(v,float) else "left")
        c.border=bdr()
        if isinstance(v,float): c.number_format=NUM
    ws.row_dimensions[r].height=22

    # HSN Variance
    _next(ws)
    rv=_next(ws)
    ws.merge_cells(f"A{rv}:{get_column_letter(nc)}{rv}")
    cv=ws.cell(row=rv,column=1,
               value="⚠  HSN vs Supply-Type Annual Variance  (must be ZERO for clean filing)")
    cv.font=font(True,"FFFFFF",10); cv.fill=fill(C_ORANGE)
    cv.alignment=aln("center"); cv.border=bdr(); ws.row_dimensions[rv].height=20

    rh=_next(ws)
    for i,(h,w) in enumerate(zip(["Comparison","Taxable Value (₹)","IGST (₹)",
                                   "CGST (₹)","SGST (₹)"],[40,18,16,16,16]),1):
        c=ws.cell(row=rh,column=i,value=h)
        c.font=font(True,"FFFFFF",9); c.fill=fill(C_MED)
        c.alignment=aln("center"); c.border=bdr()
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[rh].height=20

    sup_tv=ann["b2b_tv"]+ann["b2cs_tv"]+ann["b2cl_tv"]
    sup_ig=ann["b2b_ig"]; sup_cg=ann["b2b_cg"]; sup_sg=ann["b2b_sg"]
    hsn_tv=ann["hsn_tv"]

    for lbl,vals,bold in[
        ("Supply-Type Totals (B2B+B2CS+B2CL)",[sup_tv,sup_ig,sup_cg,sup_sg],False),
        ("HSN Summary Totals",[hsn_tv,0,0,0],False),
        ("Difference (Supply − HSN)",[sup_tv-hsn_tv,sup_ig,sup_cg,sup_sg],True),
    ]:
        rx=_next(ws)
        bg=C_YELLOW if bold else(C_LGREY if rx%2==0 else C_WHITE)
        wcell(ws,rx,1,lbl,bold=bold,bg=bg)
        for ci,v in enumerate(vals,2): wcell(ws,rx,ci,v,bold=bold,bg=bg,ha="right",fmt=NUM)
        ws.row_dimensions[rx].height=17
    return ws

# ════════════════════════════════════════════════════════════════════════
#  HEADER SHEET
# ════════════════════════════════════════════════════════════════════════
def build_header(wb, fy_label, trader_name, gstin, files_info, sheet_names):
    ws = wb.create_sheet("00_GSTR1_Header")
    ws.sheet_view.showGridLines = False
    nc=6

    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    c=ws["A1"]; c.value=f"GSTR-1 — {trader_name}  |  Financial Year {fy_label}  |  GSTIN: {gstin}"
    c.font=font(True,"FFFFFF",13); c.fill=fill(C_DARK)
    c.alignment=aln("center"); c.border=bdr(); ws.row_dimensions[1].height=30

    for col_i, w in enumerate([4,22,14,14,14,30],1):
        ws.column_dimensions[get_column_letter(col_i)].width=w

    # Filing summary
    r=2
    ws.merge_cells(f"A{r}:{get_column_letter(nc)}{r}")
    c=ws.cell(row=r,column=1,value="  ▶  Monthly Filing Details")
    c.font=font(True,"FFFFFF",9); c.fill=fill(C_MED)
    c.alignment=aln("left"); c.border=bdr(); ws.row_dimensions[r].height=18; r+=1

    for col,h in zip(range(2,7),["Month","Period Code","Filing Type","Filing Date","Source File"]):
        cc=ws.cell(row=r,column=col,value=h)
        cc.font=font(True,"FFFFFF",9); cc.fill=fill(C_MED)
        cc.alignment=aln("center"); cc.border=bdr()
    ws.row_dimensions[r].height=20; r+=1

    for i,info in enumerate(files_info):
        bg=C_LTBLUE if i%2==0 else C_WHITE
        ws.cell(r,1).fill=fill(bg); ws.cell(r,1).border=bdr()
        for col,val in zip(range(2,7),[info["month"],info["fp"],
                                        info["filing_typ"],info["fil_dt"],info["file"]]):
            wcell(ws,r,col,val,bg=bg)
        ws.row_dimensions[r].height=16; r+=1

    # Sheets index
    r+=1
    ws.merge_cells(f"A{r}:{get_column_letter(nc)}{r}")
    c=ws.cell(row=r,column=1,value="  ▶  Sheets in This Workbook")
    c.font=font(True,"FFFFFF",9); c.fill=fill(C_MED)
    c.alignment=aln("left"); c.border=bdr(); ws.row_dimensions[r].height=18; r+=1

    for sn in sheet_names:
        bg=C_LGREY if r%2==0 else C_WHITE
        ws.cell(r,1).fill=fill(bg); ws.cell(r,1).border=bdr()
        ws.merge_cells(f"B{r}:{get_column_letter(nc)}{r}")
        wcell(ws,r,2,sn,bg=bg); ws.row_dimensions[r].height=15; r+=1

    ws.merge_cells(f"A{r}:{get_column_letter(nc)}{r}")
    c=ws.cell(row=r,column=1,
              value=f"Generated: {datetime.now().strftime('%d-%m-%Y  %H:%M')}  |  gstr1_fy_v5.py")
    c.font=font(False,"595959",9); c.fill=fill(C_LGREY)
    c.alignment=aln("center"); c.border=bdr()
    return ws

# ════════════════════════════════════════════════════════════════════════
#  FILE DISCOVERY (BUG FIX: deduplicate)
# ════════════════════════════════════════════════════════════════════════
def find_zips(args):
    seen=set(); zips=[]
    for a in args:
        p=Path(a)
        candidates=[]
        if p.is_dir():
            candidates=list(p.glob("*.zip"))
        elif p.suffix.lower()==".zip" and p.exists():
            candidates=[p]
        for z in candidates:
            rz=z.resolve()
            if rz not in seen:
                seen.add(rz); zips.append(z)
    return zips

def fp_label(fp):
    return f"{MONTH_NAME.get(fp[:2],fp[:2])} {fp[2:]}"

def sort_key_fp(fp):
    return MONTH_ORDER.get(fp[:2], 99)

def read_zip(zp):
    with zipfile.ZipFile(zp) as zf:
        jsons=[n for n in zf.namelist() if n.endswith(".json")]
        if not jsons: return None, None
        with zf.open(jsons[0]) as f:
            return json.load(f), jsons[0]

# ════════════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════════════
def extract_fy(zip_paths, trader_name="", out_path=None):
    # Sort by FY month order
    def get_fp(zp):
        d,_=read_zip(zp); return d.get("fp","9999") if d else "9999"
    zip_paths=sorted(zip_paths, key=lambda z: sort_key_fp(get_fp(z)))

    if not zip_paths:
        print("  ✗ No ZIP files found."); return

    print(f"\n  Processing {len(zip_paths)} month(s)...")

    # ── Column definitions ────────────────────────────────────────────
    H_b2b,W_b2b,F_b2b         = cols_b2b()
    H_cdnr,W_cdnr,F_cdnr      = cols_cdnr()
    H_b2cs,W_b2cs,F_b2cs      = cols_b2cs()
    H_b2cl,W_b2cl,F_b2cl      = cols_b2cl()
    H_exp,W_exp,F_exp          = cols_exp()
    H_nil,W_nil,F_nil          = cols_nil()
    H_hsn,W_hsn,F_hsn         = cols_hsn()
    H_doc,W_doc,F_doc          = cols_doc()
    H_cdnur,W_cdnur,F_cdnur   = cols_cdnur()
    H_adv,W_adv,F_adv         = cols_adv()
    H_eco,W_eco,F_eco          = cols_eco()
    H_sec,W_sec,F_sec          = cols_sec()
    H_ab2b,W_ab2b,F_ab2b      = cols_amd_b2b()
    H_ab2cl,W_ab2cl,F_ab2cl   = cols_amd_b2cl()
    H_aexp,W_aexp,F_aexp      = cols_amd_exp()
    H_acdnr,W_acdnr,F_acdnr  = cols_amd_cdnr()
    H_acdnur,W_acdnur,F_acdnur=cols_amd_cdnur()
    H_ab2cs,W_ab2cs,F_ab2cs  = cols_amd_b2cs()

    # Detect GSTIN from first zip
    d0,_=read_zip(zip_paths[0])
    gstin=d0.get("gstin","") if d0 else ""
    if not trader_name: trader_name=f"GSTIN: {gstin}"

    wb=Workbook(); wb.remove(wb.active)

    def ms(name,title,H,W,bg=C_DARK):
        return make_sheet(wb,name,title,H,W,bg=bg,trader_name=trader_name,gstin=gstin)

    ws01=ms("01_B2B_Invoices",      "Table 4A/4B/6B/6C — B2B Invoices",          H_b2b,W_b2b)
    ws02=ms("02_CDNR_Registered",   "Table 9B — CDNR Registered",                 H_cdnr,W_cdnr)
    ws03=ms("03_B2CS_Others",        "Table 7 — B2CS Others",                      H_b2cs,W_b2cs)
    ws04=ms("04_B2CL_Large",         "Table 5 — B2CL Large Invoices",              H_b2cl,W_b2cl)
    ws05=ms("05_Exports",            "Table 6A — Exports",                         H_exp,W_exp)
    ws06=ms("06_Nil_Exempt",         "Tables 8A-8D — Nil / Exempt / Non-GST",      H_nil,W_nil)
    ws07=ms("07_HSN_Summary",        "Table 12 — HSN-wise Summary",                H_hsn,W_hsn)
    ws08=ms("08_Document_Summary",   "Table 13 — Document Issuance Summary",       H_doc,W_doc)
    ws09=ms("09_CDNUR_Unregistered", "Table 9B — CDNUR Unregistered",              H_cdnur,W_cdnur)
    ws10=ms("10_Advances_TaxLiab",   "Tables 11A(1/2) — Advances Tax Liability",   H_adv,W_adv)
    ws11=ms("11_Advances_Adjust",    "Tables 11B(1/2) — Adjustment of Advances",   H_adv,W_adv)
    ws12=ms("12_ECO_Supplies",       "Table 14 — ECO Supplies",                    H_eco,W_eco)
    ws13=ms("13_Section9_5",         "Table 15 — Supplies U/s 9(5)",               H_sec,W_sec)
    ws14=ms("14_AMD_B2B",            "Table 9A — Amended B2B Invoices",            H_ab2b,W_ab2b,bg=C_AMEND)
    ws15=ms("15_AMD_B2CL",           "Table 9A — Amended B2CL",                   H_ab2cl,W_ab2cl,bg=C_AMEND)
    ws16=ms("16_AMD_Exports",        "Table 9A — Amended Exports",                 H_aexp,W_aexp,bg=C_AMEND)
    ws17=ms("17_AMD_CDNR",           "Table 9C — Amended CDNR Registered",         H_acdnr,W_acdnr,bg=C_AMEND)
    ws18=ms("18_AMD_CDNUR",          "Table 9C — Amended CDNUR Unregistered",      H_acdnur,W_acdnur,bg=C_AMEND)
    ws19=ms("19_AMD_B2CS",           "Table 10 — Amended B2CS",                   H_ab2cs,W_ab2cs,bg=C_AMEND)
    ws20=ms("20_AMD_AdvTaxLiab",     "Table 11A — Amended Advances",               H_adv,W_adv,bg=C_AMEND)
    ws21=ms("21_AMD_AdvAdjust",      "Table 11B — Amended Adjustments",            H_adv,W_adv,bg=C_AMEND)
    ws22=ms("22_AMD_ECO",            "Table 14A — Amended ECO",                   H_eco,W_eco,bg=C_AMEND)
    ws23=ms("23_AMD_Sec9_5",         "Table 15A — Amended Sec 9(5)",              H_sec,W_sec,bg=C_AMEND)

    all_ws=[ws01,ws02,ws03,ws04,ws05,ws06,ws07,ws08,ws09,ws10,
            ws11,ws12,ws13,ws14,ws15,ws16,ws17,ws18,ws19,ws20,ws21,ws22,ws23]
    all_nc=[len(H_b2b),len(H_cdnr),len(H_b2cs),len(H_b2cl),len(H_exp),
            len(H_nil),len(H_hsn),len(H_doc),len(H_cdnur),len(H_adv),
            len(H_adv),len(H_eco),len(H_sec),len(H_ab2b),len(H_ab2cl),
            len(H_aexp),len(H_acdnr),len(H_acdnur),len(H_ab2cs),len(H_adv),
            len(H_adv),len(H_eco),len(H_sec)]

    months_data  = {}
    files_info   = []
    fy_fps       = []
    master_rows  = []           # for sheet 26
    fy_gstin_data= defaultdict(lambda: defaultdict(dict))  # {gstin:{month:{tv..}}}

    for zp in zip_paths:
        zp = Path(zp) if not hasattr(zp, 'name') else zp  # ensure Path object
        d,src_file=read_zip(zp)
        if not d: continue
        fp=d.get("fp",""); pl=fp_label(fp); mname=MONTH_NAME.get(fp[:2],"")
        fy_fps.append(fp)
        print(f"    ✓ {pl}  ({zp.name})")
        files_info.append({"month":pl,"fp":fp,"filing_typ":d.get("filing_typ",""),
                           "fil_dt":d.get("fil_dt",""),"file":src_file})

        # Month separator on every sheet
        for ws,nc in zip(all_ws,all_nc):
            month_sep(ws,pl,nc)

        # Write data
        sb2b  = write_b2b(ws01,  d, pl, F_b2b,  len(H_b2b))
        scdnr = write_cdnr(ws02, d, pl, F_cdnr, len(H_cdnr))
        sb2cs = write_b2cs(ws03, d, pl, F_b2cs, len(H_b2cs))
        sb2cl = write_b2cl(ws04, d, pl, F_b2cl, len(H_b2cl))
        sexp  = write_exp(ws05,  d, pl, F_exp,  len(H_exp))
        write_nil(ws06, d, pl, F_nil, len(H_nil))
        shsn  = write_hsn(ws07, d, pl, F_hsn, len(H_hsn))
        write_docs(ws08, d, pl, F_doc, len(H_doc))
        write_cdnur(ws09, d, pl, F_cdnur, len(H_cdnur))
        write_adv(ws10, d.get("at",[]),    pl, F_adv, len(H_adv), pl)
        write_adv(ws11, d.get("txpd",[]),  pl, F_adv, len(H_adv), pl)
        write_eco(ws12, d.get("eco",[]),   pl, F_eco, len(H_eco))
        write_sec95(ws13,d.get("supeco",[]),pl, F_sec, len(H_sec))
        write_amd_b2b(ws14,  d, pl, F_ab2b, len(H_ab2b))
        write_amd_b2cl(ws15, d, pl, F_ab2cl,len(H_ab2cl))
        write_amd_exp(ws16,  d, pl, F_aexp, len(H_aexp))
        write_amd_cdnr(ws17, d, pl, F_acdnr,len(H_acdnr))
        write_amd_cdnur(ws18,d, pl, F_acdnur,len(H_acdnur))
        write_amd_b2cs(ws19, d, pl, F_ab2cs,len(H_ab2cs))
        write_adv(ws20, d.get("ata",[]),   pl, F_adv, len(H_adv), pl)
        write_adv(ws21, d.get("txpda",[]), pl, F_adv, len(H_adv), pl)
        write_eco(ws22, d.get("ecoa",[]),  pl, F_eco, len(H_eco))
        write_sec95(ws23,d.get("supaeco",[]),pl,F_sec,len(H_sec))

        # Accumulate for recon
        tot_ig=(sb2b.get("ig",0)+sb2cs.get("ig",0)+sb2cl.get("ig",0)+
                sexp.get("ig",0)+scdnr.get("ig",0))
        tot_cg=sb2b.get("cg",0)+sb2cs.get("cg",0)+scdnr.get("cg",0)
        tot_sg=sb2b.get("sg",0)+sb2cs.get("sg",0)+scdnr.get("sg",0)
        months_data[mname]={
            "b2b_tv":sb2b.get("tv",0),"b2b_ig":sb2b.get("ig",0),
            "b2b_cg":sb2b.get("cg",0),"b2b_sg":sb2b.get("sg",0),
            "b2cs_tv":sb2cs.get("tv",0),"b2cl_tv":sb2cl.get("tv",0),
            "exp_tv":sexp.get("tv",0),"cdnr_cr":scdnr.get("cr",0),
            "cdnr_dr":scdnr.get("dr",0),"hsn_tv":shsn.get("tv",0),
            "tot_ig":tot_ig,"tot_cg":tot_cg,"tot_sg":tot_sg,
            "tot_tax":tot_ig+tot_cg+tot_sg,
            "inv_val":sb2b.get("val",0)+sb2cl.get("val",0)+sexp.get("val",0),
        }

        # Master sheet rows (B2B invoices)
        for ent in d.get("b2b",[]):
            ctin=ent.get("ctin","")
            for inv in ent.get("inv",[]):
                typ=INV_TYPE.get(inv.get("inv_typ","R"),inv.get("inv_typ",""))
                val=float(inv.get("val",0)or 0)
                for it in inv.get("itms",[]):
                    det=it.get("itm_det",{})
                    tv=float(det.get("txval",0)or 0); ig=float(det.get("iamt",0)or 0)
                    cg=float(det.get("camt",0)or 0);  sg=float(det.get("samt",0)or 0)
                    master_rows.append([pl,"B2B Invoice",ctin,inv.get("inum",""),
                                        inv.get("idt",""),typ,inv.get("pos",""),
                                        inv.get("rchrg","N"),det.get("rt",0),
                                        tv,ig,cg,sg,ig+cg+sg,val,"","",""])
        # CDNR rows
        for ent in d.get("cdnr",[]):
            ctin=ent.get("ctin","")
            for nt in ent.get("nt",[]):
                ntty=nt.get("ntty","C"); tv=ig=cg=sg=0.0
                for it in nt.get("itms",[]):
                    det=it.get("itm_det",{})
                    tv+=float(det.get("txval",0)or 0); ig+=float(det.get("iamt",0)or 0)
                    cg+=float(det.get("camt",0)or 0);  sg+=float(det.get("samt",0)or 0)
                master_rows.append([pl,"Credit/Debit Note",ctin,"","","",
                                    nt.get("pos",""),"",0,tv,ig,cg,sg,ig+cg+sg,
                                    float(nt.get("val",0)or 0),
                                    "Credit Note" if ntty=="C" else "Debit Note",
                                    nt.get("nt_num",""),nt.get("nt_dt","")])

        # Per-GSTIN accumulation
        for ctin, g_tv, g_ig, g_cg, g_sg, g_val in sb2b.get("gstin_totals",[]):
            fy_gstin_data[ctin][mname]={"tv":g_tv,"ig":g_ig,"cg":g_cg,"sg":g_sg,"val":g_val}

    # Annual totals on all sheets
    fy_tv_b2b=sum(m.get("b2b_tv",0) for m in months_data.values())
    fy_ig_b2b=sum(m.get("b2b_ig",0) for m in months_data.values())
    fy_cg_b2b=sum(m.get("b2b_cg",0) for m in months_data.values())
    fy_sg_b2b=sum(m.get("b2b_sg",0) for m in months_data.values())
    fy_val_b2b=sum(m.get("inv_val",0) for m in months_data.values())

    annual_total_row(ws01,
                     f"  ★ FY ANNUAL TOTAL — {trader_name}",
                     [fy_tv_b2b,fy_ig_b2b,fy_cg_b2b,fy_sg_b2b,
                      fy_ig_b2b+fy_cg_b2b+fy_sg_b2b,fy_val_b2b],
                     fmts=[NUM,NUM,NUM,NUM,NUM,NUM], data_start_col=9)

    # FY label
    fps_sorted=sorted(fy_fps, key=lambda f: MONTH_ORDER.get(f[:2],99))
    if fps_sorted:
        yr=int(fps_sorted[0][2:])
        fy_label=f"FY {yr}-{str(yr+1)[-2:]}"
    else: fy_label="FY"

    # Build special sheets
    build_recon(wb, fy_label, trader_name, gstin, months_data)
    build_gstin_sheet(wb, fy_gstin_data, trader_name, gstin, fy_label)
    build_master_sheet(wb, master_rows, trader_name, gstin, fy_label)
    build_header(wb, fy_label, trader_name, gstin, files_info,
                 [s.title for s in wb.worksheets])
    wb.move_sheet("00_GSTR1_Header",
                  offset=-(wb.sheetnames.index("00_GSTR1_Header")))

    if out_path is None:
        out_path=(Path(zip_paths[0]).parent /
                  f"GSTR1_{fy_label.replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

    wb.save(str(out_path))
    print(f"\n  ✓ Saved : {Path(out_path).name}")
    print(f"  Sheets  : {[s.title for s in wb.worksheets]}")

    # Persist any newly fetched GSTIN names to local cache
    if _CACHE_ENABLED and _NAME_CACHE:
        _NAME_CACHE.save()
        st = _NAME_CACHE.stats()
        print(f"  Cache   : {st['total']} GSTIN names stored "
              f"({st['portal']} from portal, {st['master']} from master, {st['manual']} manual)")

    return str(out_path)


# ════════════════════════════════════════════════════════════════════════
#  CLI
# ════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="GSTR-1 FY Extractor v5",
        epilog=(
            "Examples:\n"
            "  python gstr1_fy_v5.py                          "
            "# run in current folder (auto-detect ZIPs)\n"
            "  python gstr1_fy_v5.py .                        "
            "# same — explicit current folder\n"
            '  python gstr1_fy_v5.py --name "ARUN ENTERPRISES" .\n'
            "  python gstr1_fy_v5.py C:\\path\\to\\gstr1_folder\n"
            "  python gstr1_fy_v5.py Apr.zip May.zip Jun.zip\n"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    # paths is now OPTIONAL — defaults to current working directory
    parser.add_argument(
        "paths", nargs="*", default=["."],
        help="ZIP files or folder containing GSTR1_*.zip files "
             "(default: current directory)",
    )
    parser.add_argument("--name", "-n", default="",
        help='Trader/Company name e.g. --name "SOWMIYA ENTERPRISES"')
    parser.add_argument("--out", "-o", default=None,
        help="Output Excel path (optional)")
    args = parser.parse_args()

    # If user passed no paths, use current directory
    if not args.paths:
        args.paths = ["."]

    zips = find_zips(args.paths)
    if not zips:
        print(
            "No GSTR1_*.zip files found.\n"
            f"  Searched in: {', '.join(args.paths)}\n\n"
            "Usage:\n"
            "  python gstr1_fy_v5.py                   "
            "  ← run from inside the GSTIN folder\n"
            '  python gstr1_fy_v5.py --name "ARUN ENTERPRISES"\n'
            "  python gstr1_fy_v5.py C:\\path\\to\\folder\n"
            "  python gstr1_fy_v5.py Apr.zip May.zip Jun.zip"
        )
        sys.exit(1)

    extract_fy(zips, trader_name=args.name, out_path=args.out)
