"""
================================================================================
  GSTIN LOOKUP TOOL — Free, No API Key, No Login Required
  ========================================================
  Uses the GST Portal's own public search endpoint.

  MODES:
  ──────
  1. Interactive  → just run:  python gstin_lookup.py
  2. Single       → python gstin_lookup.py 33AABCT1234C1ZX
  3. Bulk Excel   → python gstin_lookup.py --file clients.xlsx
  4. Bulk text    → python gstin_lookup.py --file gstins.txt
  5. Paste list   → python gstin_lookup.py --paste

  OUTPUT:
  ───────
  • Prints results to console (colored)
  • Saves Excel → GSTIN_LOOKUP_<date>.xlsx  (same folder as this script)
  • Excel has two sheets: Results + Not Found

  REQUIREMENTS:
  ─────────────
  pip install openpyxl requests
  (requests is optional — falls back to urllib if not installed)

================================================================================
"""

import sys, os, re, time, json, ssl
from pathlib import Path
from datetime import datetime

# ── Try to import requests; fall back to urllib ───────────────────────────────
try:
    import requests as _req
    _USE_REQUESTS = True
except ImportError:
    import urllib.request, urllib.error
    _USE_REQUESTS = False

# ── Try openpyxl ─────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_EXCEL = True
except ImportError:
    _HAS_EXCEL = False

# ── Console colours (Windows-safe) ───────────────────────────────────────────
if sys.platform == "win32":
    os.system("color")   # enable ANSI on Windows terminal
C_RESET  = "\033[0m"
C_CYAN   = "\033[96m"
C_GREEN  = "\033[92m"
C_YELLOW = "\033[93m"
C_RED    = "\033[91m"
C_BOLD   = "\033[1m"
C_GREY   = "\033[90m"
C_WHITE  = "\033[97m"

# ── GST Portal endpoints (tried in order) ────────────────────────────────────
GST_ENDPOINTS = [
    "https://services.gst.gov.in/services/api/search/gstin?gstin={gstin}",
    "https://www.gst.gov.in/util/rest/toolkit/searchTax?gstin={gstin}",
]
HEADERS   = {
    "User-Agent":      ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/124.0.0.0 Safari/537.36"),
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "en-IN,en-GB;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection":      "keep-alive",
    "Referer":         "https://services.gst.gov.in/services/searchtp",
    "Origin":          "https://services.gst.gov.in",
    "sec-ch-ua":       '"Chromium";v="124","Google Chrome";v="124"',
    "sec-fetch-dest":  "empty",
    "sec-fetch-mode":  "cors",
    "sec-fetch-site":  "same-origin",
}

DELAY_SEC = 0.35   # polite delay between bulk lookups


# ══════════════════════════════════════════════════════════════════════════════
# CORE LOOKUP
# ══════════════════════════════════════════════════════════════════════════════

def _get_json(gstin):
    """
    Try all GST portal endpoints in order.
    Returns parsed JSON from first one that succeeds.
    """
    last_err = None
    for endpoint_tpl in GST_ENDPOINTS:
        url = endpoint_tpl.format(gstin=gstin)
        try:
            if _USE_REQUESTS:
                r = _req.get(url, headers=HEADERS, timeout=12, verify=False)
                r.raise_for_status()
                return r.json()
            else:
                ctx = ssl.create_default_context()
                ctx.check_hostname = False
                ctx.verify_mode = ssl.CERT_NONE
                req = urllib.request.Request(url, headers=HEADERS)
                with urllib.request.urlopen(req, timeout=12, context=ctx) as resp:
                    return json.loads(resp.read().decode("utf-8"))
        except Exception as e:
            last_err = e
            continue   # try next endpoint
    raise last_err


def _parse(raw, gstin):
    """Parse GST portal response into a clean dict."""
    info = raw.get("taxpayerInfo") or raw.get("data") or raw

    def _g(*keys):
        for k in keys:
            v = info.get(k, "")
            if v: return str(v).strip()
        return ""

    legal  = _g("lgnm",     "legalName",  "tradeName")
    trade  = _g("tradeNam", "tradeName",  "lgnm")
    return {
        "gstin":             gstin,
        "pan":               gstin[2:12],
        "legal_name":        legal,
        "trade_name":        trade if trade != legal else "",
        "status":            _g("sts",  "status"),
        "state":             _g("stj",  "state", "pradr"),
        "business_type":     _g("ctb",  "constitutionOfBusiness"),
        "registration_date": _g("rgdt", "registrationDate"),
        "cancellation_date": _g("cxdt", "cancellationDate"),
        "error":             "",
    }


def lookup_one(gstin):
    """
    Look up a single GSTIN.
    Returns dict with legal_name, trade_name, status, etc.
    On failure, returns dict with 'error' key set.
    """
    gstin = gstin.strip().upper()
    if len(gstin) != 15:
        return {"gstin": gstin, "error": "Must be exactly 15 characters", "legal_name": ""}

    url = GST_ENDPOINTS[0].format(gstin=gstin)   # for error reporting only
    try:
        raw = _get_json(gstin)
        result = _parse(raw, gstin)
        if not result["legal_name"]:
            result["error"] = "Parsed but name not found"
        return result
    except Exception as e:
        code = ""
        if hasattr(e, "code"): code = f" (HTTP {e.code})"
        elif hasattr(e, "response") and e.response is not None:
            code = f" (HTTP {e.response.status_code})"
        return {
            "gstin": gstin, "pan": gstin[2:12] if len(gstin)==15 else "",
            "legal_name": "", "trade_name": "", "status": "",
            "state": "", "business_type": "", "registration_date": "",
            "cancellation_date": "", "error": f"{type(e).__name__}{code}: {e}",
        }


def lookup_bulk(gstins, show_progress=True):
    """Look up a list of GSTINs. Returns list of result dicts."""
    results = []
    total   = len(gstins)
    for i, g in enumerate(gstins, 1):
        if show_progress:
            print(f"  {C_GREY}[{i}/{total}]{C_RESET} {g} … ", end="", flush=True)
        r = lookup_one(g)
        if r["error"]:
            if show_progress: print(f"{C_RED}✗ {r['error']}{C_RESET}")
        else:
            if show_progress:
                print(f"{C_GREEN}✓{C_RESET} {C_BOLD}{r['legal_name']}{C_RESET}"
                      + (f"  {C_GREY}({r['trade_name']}){C_RESET}" if r['trade_name'] else ""))
        results.append(r)
        if i < total:
            time.sleep(DELAY_SEC)
    return results


# ══════════════════════════════════════════════════════════════════════════════
# CONSOLE DISPLAY
# ══════════════════════════════════════════════════════════════════════════════

def print_result(r):
    """Pretty-print a single lookup result."""
    if r["error"] and not r["legal_name"]:
        print(f"\n  {C_RED}✗ {r['gstin']} — {r['error']}{C_RESET}")
        return

    sts_col = C_GREEN if (r["status"] or "").upper() == "ACTIVE" else C_YELLOW
    print(f"""
  {C_CYAN}{'─'*60}{C_RESET}
  {C_BOLD}{C_WHITE}GSTIN       :{C_RESET}  {C_CYAN}{r['gstin']}{C_RESET}
  {C_BOLD}PAN         :{C_RESET}  {r['pan']}
  {C_BOLD}Legal Name  :{C_RESET}  {C_BOLD}{C_WHITE}{r['legal_name'] or '—'}{C_RESET}
  {C_BOLD}Trade Name  :{C_RESET}  {r['trade_name'] or '(same as legal)'}
  {C_BOLD}Status      :{C_RESET}  {sts_col}{r['status'] or '—'}{C_RESET}
  {C_BOLD}State       :{C_RESET}  {r['state'] or '—'}
  {C_BOLD}Biz Type    :{C_RESET}  {r['business_type'] or '—'}
  {C_BOLD}Reg. Date   :{C_RESET}  {r['registration_date'] or '—'}""")
    if r.get("cancellation_date"):
        print(f"  {C_BOLD}Cancel Date :{C_RESET}  {C_RED}{r['cancellation_date']}{C_RESET}")
    print(f"  {C_CYAN}{'─'*60}{C_RESET}")


def print_bulk_summary(results):
    """Print summary table for bulk results."""
    found  = [r for r in results if r["legal_name"]]
    failed = [r for r in results if not r["legal_name"]]

    print(f"\n  {C_CYAN}{'═'*72}{C_RESET}")
    print(f"  {C_BOLD}{'GSTIN':<18} {'PAN':<12} {'LEGAL NAME':<35} {'STATUS'}{C_RESET}")
    print(f"  {C_CYAN}{'─'*72}{C_RESET}")
    for r in results:
        if r["legal_name"]:
            sts_col = C_GREEN if r["status"].upper()=="ACTIVE" else C_YELLOW
            print(f"  {C_CYAN}{r['gstin']:<18}{C_RESET} "
                  f"{r['pan']:<12} "
                  f"{C_BOLD}{r['legal_name'][:34]:<35}{C_RESET} "
                  f"{sts_col}{r['status']}{C_RESET}")
        else:
            print(f"  {C_RED}{r['gstin']:<18}{C_RESET} "
                  f"{r.get('pan',''):<12} "
                  f"{C_RED}✗ {r['error'][:34]}{C_RESET}")
    print(f"  {C_CYAN}{'═'*72}{C_RESET}")
    print(f"  ✅ Found: {C_GREEN}{len(found)}{C_RESET}   "
          f"❌ Failed: {C_RED}{len(failed)}{C_RESET}   "
          f"Total: {len(results)}\n")


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def save_excel(results, out_path):
    """Save results to a formatted Excel workbook."""
    if not _HAS_EXCEL:
        print(f"  {C_YELLOW}⚠ openpyxl not installed — skipping Excel export.{C_RESET}")
        print(f"  Run: pip install openpyxl")
        return None

    wb = openpyxl.Workbook()

    # ── Helpers ──────────────────────────────────────────────────────────────
    def _f(hex_):  return PatternFill("solid", fgColor=hex_)
    def _fn(bold=False, color="000000", size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    def _bd():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)
    def _al(h="left"):  return Alignment(horizontal=h, vertical="center", wrap_text=False)

    def _hdr(ws, row, col, val, bg="1F3864", fg="FFFFFF"):
        c = ws.cell(row=row, column=col, value=val)
        c.font = _fn(True, fg, 10); c.fill = _f(bg)
        c.alignment = _al("center"); c.border = _bd()
        return c

    def _cel(ws, row, col, val, bg="FFFFFF", bold=False, fg="000000", align="left"):
        c = ws.cell(row=row, column=col, value=val)
        c.font = _fn(bold, fg); c.fill = _f(bg)
        c.alignment = _al(align); c.border = _bd()
        return c

    COLS = ["GSTIN", "PAN", "Legal Name", "Trade Name",
            "Status", "State / Jurisdiction", "Business Type",
            "Reg. Date", "Cancel Date", "Error"]
    WIDTHS = [20, 13, 42, 32, 12, 28, 24, 14, 14, 40]

    found  = [r for r in results if r["legal_name"]]
    failed = [r for r in results if not r["legal_name"]]

    for ws_name, rows_data in [("Results", found), ("Not Found", failed)]:
        if ws_name == "Results":
            ws = wb.active
            ws.title = ws_name
        else:
            if not rows_data: continue
            ws = wb.create_sheet(ws_name)

        ws.freeze_panes = "A2"

        # Header
        for ci, (col, w) in enumerate(zip(COLS, WIDTHS), 1):
            _hdr(ws, 1, ci, col)
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Data rows
        for ri, r in enumerate(rows_data, 2):
            alt = "F2F2F2" if ri % 2 == 0 else "FFFFFF"
            sts = (r.get("status") or "").upper()
            sts_fg = "276221" if sts == "ACTIVE" else ("9C0006" if sts else "000000")

            vals = [
                r.get("gstin",""),
                r.get("pan",""),
                r.get("legal_name",""),
                r.get("trade_name",""),
                r.get("status",""),
                r.get("state",""),
                r.get("business_type",""),
                r.get("registration_date",""),
                r.get("cancellation_date",""),
                r.get("error",""),
            ]
            fgs = ["2E75B6","000000","000000","595959",
                   sts_fg,"000000","000000","000000","9C0006","9C0006"]
            bolds = [True,False,True,False,False,False,False,False,False,False]

            for ci, (v, fg, bold) in enumerate(zip(vals, fgs, bolds), 1):
                _cel(ws, ri, ci, v, alt, bold, fg)

        # Summary row at bottom
        if rows_data:
            sr = len(rows_data) + 2
            ws.cell(row=sr, column=1, value=f"Total: {len(rows_data)} record(s)").font = _fn(True)

    # Title / meta sheet
    ws_meta = wb.create_sheet("Info", 0)
    ws_meta["A1"] = "GSTIN Lookup Report"
    ws_meta["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F3864")
    ws_meta["A2"] = f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws_meta["A2"].font = Font(name="Calibri", size=10, color="595959")
    ws_meta["A3"] = f"Total Lookups: {len(results)}  |  Found: {len(found)}  |  Failed: {len(failed)}"
    ws_meta["A3"].font = Font(name="Calibri", size=10, color="1F3864", bold=True)
    ws_meta.column_dimensions["A"].width = 55

    wb.save(str(out_path))
    return out_path


# ══════════════════════════════════════════════════════════════════════════════
# INPUT READERS
# ══════════════════════════════════════════════════════════════════════════════

def _clean(g): return re.sub(r'[^A-Z0-9]', '', g.strip().upper())

def read_from_excel(fpath):
    """Read GSTINs from an Excel file — auto-detects the GSTIN column."""
    if not _HAS_EXCEL:
        print(f"{C_RED}openpyxl not installed. Run: pip install openpyxl{C_RESET}")
        sys.exit(1)
    wb  = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
    ws  = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows: return []

    # Find GSTIN column
    headers = [str(c or "").strip().upper() for c in (rows[0] or [])]
    gstin_col = -1
    for kw in ("GSTIN", "GSTIN/UIN", "GST NO", "GST NUMBER", "GSTIN NO"):
        if kw in headers: gstin_col = headers.index(kw); break

    gstins = []
    data_rows = rows[1:] if gstin_col != -1 else rows   # skip header if found
    for row in data_rows:
        for ci, cell in enumerate(row):
            if ci == gstin_col or gstin_col == -1:
                v = _clean(str(cell or ""))
                if len(v) == 15: gstins.append(v); break
    return list(dict.fromkeys(gstins))   # deduplicate, preserve order


def read_from_txt(fpath):
    """Read GSTINs from a plain-text file (one per line, comma or semicolon separated)."""
    text = Path(fpath).read_text(encoding="utf-8", errors="ignore")
    raw  = re.split(r'[\n,;|\t]+', text)
    return list(dict.fromkeys(
        _clean(g) for g in raw if len(_clean(g)) == 15
    ))


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def _banner():
    print(f"""
{C_CYAN}╔══════════════════════════════════════════════════════════╗
║      GSTIN LOOKUP — Free  |  No API Key  |  No Login     ║
╚══════════════════════════════════════════════════════════╝{C_RESET}
  Source  : services.gst.gov.in  (GST Portal public API)
  Output  : Excel file saved in this folder
""")

def _save_and_report(results):
    if not results: return
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = Path(__file__).parent / f"GSTIN_LOOKUP_{ts}.xlsx"
    saved = save_excel(results, out)
    if saved:
        print(f"\n  💾 Excel saved → {C_GREEN}{saved}{C_RESET}")
    else:
        # CSV fallback
        csv_path = out.with_suffix(".csv")
        lines = ["GSTIN,PAN,Legal Name,Trade Name,Status,State,Business Type,Reg Date,Error"]
        for r in results:
            lines.append(",".join(f'"{r.get(k,"")}"' for k in
                ["gstin","pan","legal_name","trade_name","status","state",
                 "business_type","registration_date","error"]))
        csv_path.write_text("\n".join(lines), encoding="utf-8")
        print(f"\n  💾 CSV saved → {C_GREEN}{csv_path}{C_RESET}")


def _interactive():
    """Simple interactive mode — type GSTINs one by one."""
    print(f"  {C_GREY}Type a GSTIN and press Enter. Type {C_WHITE}BULK{C_GREY} to switch to bulk mode.")
    print(f"  Type {C_WHITE}DONE{C_GREY} or {C_WHITE}EXIT{C_GREY} when finished.{C_RESET}\n")
    all_results = []
    while True:
        try:
            raw = input(f"  {C_CYAN}GSTIN >{C_RESET} ").strip().upper()
        except (EOFError, KeyboardInterrupt):
            break
        if not raw: continue
        if raw in ("EXIT", "QUIT", "DONE", "Q"):
            break
        if raw == "BULK":
            print(f"\n  {C_GREY}Paste GSTINs (one per line). Enter a blank line when done.{C_RESET}")
            lines = []
            while True:
                try: ln = input().strip()
                except (EOFError, KeyboardInterrupt): break
                if not ln: break
                lines.append(ln)
            gstins = [_clean(g) for g in lines if len(_clean(g)) == 15]
            if gstins:
                print(f"\n  {C_YELLOW}Looking up {len(gstins)} GSTINs …{C_RESET}\n")
                results = lookup_bulk(gstins)
                print_bulk_summary(results)
                all_results.extend(results)
            continue
        r = lookup_one(raw)
        print_result(r)
        all_results.append(r)
    return all_results


def main():
    import argparse, warnings
    warnings.filterwarnings("ignore")

    # Suppress InsecureRequestWarning from requests
    try:
        import urllib3; urllib3.disable_warnings()
    except: pass

    _banner()

    ap = argparse.ArgumentParser(add_help=False)
    ap.add_argument("gstin",    nargs="?",       help="Single GSTIN to look up")
    ap.add_argument("--file",   metavar="PATH",  help="Excel or TXT file with GSTINs")
    ap.add_argument("--paste",  action="store_true", help="Paste mode — enter GSTINs interactively")
    ap.add_argument("--no-excel", action="store_true", help="Skip Excel output (print only)")
    args, _ = ap.parse_known_args()

    results = []

    # ── Mode 1: Single GSTIN from command line ────────────────────────────────
    if args.gstin:
        g = _clean(args.gstin)
        print(f"  Looking up {C_CYAN}{g}{C_RESET} …")
        r = lookup_one(g)
        print_result(r)
        results = [r]

    # ── Mode 2: File input ────────────────────────────────────────────────────
    elif args.file:
        fpath = Path(args.file)
        if not fpath.exists():
            print(f"  {C_RED}File not found: {fpath}{C_RESET}"); sys.exit(1)
        ext = fpath.suffix.lower()
        if ext in (".xlsx", ".xls"):
            gstins = read_from_excel(fpath)
        else:
            gstins = read_from_txt(fpath)
        if not gstins:
            print(f"  {C_RED}No valid GSTINs found in file.{C_RESET}"); sys.exit(1)
        print(f"  {C_GREEN}Found {len(gstins)} unique GSTIN(s) in file.{C_RESET}")
        print(f"  {C_YELLOW}Looking up all GSTINs …{C_RESET}\n")
        results = lookup_bulk(gstins)
        print_bulk_summary(results)

    # ── Mode 3: Paste mode ────────────────────────────────────────────────────
    elif args.paste:
        print(f"  {C_GREY}Paste GSTINs (one per line). Enter a blank line when done.{C_RESET}")
        lines = []
        while True:
            try: ln = input().strip()
            except (EOFError, KeyboardInterrupt): break
            if not ln: break
            lines.append(ln)
        gstins = [_clean(g) for g in lines if len(_clean(g)) == 15]
        if not gstins:
            print(f"  {C_RED}No valid GSTINs found.{C_RESET}"); sys.exit(1)
        print(f"\n  {C_YELLOW}Looking up {len(gstins)} GSTIN(s) …{C_RESET}\n")
        results = lookup_bulk(gstins)
        print_bulk_summary(results)

    # ── Mode 4: Interactive (default) ────────────────────────────────────────
    else:
        results = _interactive()

    # ── Save output ───────────────────────────────────────────────────────────
    if results and not args.no_excel:
        _save_and_report(results)

    print(f"\n  {C_GREY}Done.{C_RESET}\n")


if __name__ == "__main__":
    main()
