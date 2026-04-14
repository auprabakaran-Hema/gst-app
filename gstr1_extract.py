"""
GSTR-1 JSON → Comprehensive Excel Extractor  v2
=================================================
Advanced extraction from all GSTR1_*.zip files.

Sheets:
  1.  B2B_Invoices       — Invoice detail with customer name
  2.  B2B_ItemDetails    — Every HSN/item row with all tax columns
  3.  B2CS_Summary       — B2C Small by state+rate
  4.  B2CL_Invoices      — B2C Large inter-state
  5.  CDNR_CreditNotes   — Credit Notes (ntty=C) to registered
  6.  CDNR_DebitNotes    — Debit Notes  (ntty=D) to registered
  7.  CDNUR_Unregistered — CDN to unregistered buyers
  8.  Exports            — Exports with port/shipping details
  9.  Nil_Rated          — Nil, Exempt, Non-GST values
  10. Document_Summary   — Invoice issue/cancellation summary
  11. GSTR1A_Amendments  — Amended invoices if present
  12. HSN_Summary        — HSN-wise taxable and tax summary
  13. Master_Summary     — Monthly totals across all types

GSTIN Name Lookup:
  - Reads from GSTR-2A/2B Excel/ZIP in same folder
  - Falls back to GST public search API (best-effort, no auth needed)
"""

import os, sys, json, zipfile, re, time
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing package: {e}\nRun: pip install openpyxl pandas")
    sys.exit(1)

# -- Calendar month sort order ---------------------------------
MONTH_ORDER = {
    "January":1,"February":2,"March":3,"April":4,"May":5,"June":6,
    "July":7,"August":8,"September":9,"October":10,"November":11,"December":12
}

def sort_key_month(zip_path):
    """Sort GSTR1_Month_Year.zip files in calendar order."""
    stem = zip_path.stem  # e.g. "GSTR1_April_2024"
    parts = stem.split("_")
    # parts: ["GSTR1","April","2024"]
    if len(parts) >= 3:
        year = int(parts[-1]) if parts[-1].isdigit() else 9999
        month = MONTH_ORDER.get(parts[-2], 99) if len(parts[-2])>2 else 99
        # GST FY: April=start → next March=end
        # Sort: Apr(4)..Dec(12) first, then Jan(1)..Mar(3)
        if month >= 4:
            sort_month = month - 4   # Apr→0, May→1 ... Dec→8
        else:
            sort_month = month + 8   # Jan→9, Feb→10, Mar→11
        return (year - (0 if month >= 4 else 1), sort_month)
    return (9999, 99)

# -- Styling --------------------------------------------------
DARK_BLUE="1F3864"; MED_BLUE="2E75B6"
GREY_BG="F2F2F2"; WHITE="FFFFFF"
NUM_FMT="#,##0.00"

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(b=False, c="000000", s=9): return Font(name="Arial",bold=b,color=c,size=s)
def _bdr(): s=Side(style="thin"); return Border(left=s,right=s,top=s,bottom=s)
def _aln(h="left",wrap=False): return Alignment(horizontal=h,vertical="center",wrap_text=wrap)

# ── Excel formula helpers ─────────────────────────────────────────
def _is_fml(v): return isinstance(v, str) and v.startswith("=")
def _fsum(col, r1, r2): return f"=SUM({col}{r1}:{col}{r2})"
def _fdiff(ca, cb, r):  return f"={ca}{r}-{cb}{r}"

def make_sheet(wb, title, headers, widths, tc=DARK_BLUE):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    nc = len(headers)
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    t = ws["A1"]
    t.value = title.replace("_"," ")
    t.font = _font(True,"FFFFFF",11); t.fill = _fill(tc)
    t.alignment = _aln("center"); t.border = _bdr()
    ws.row_dimensions[1].height = 26
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row=2,column=ci,value=h)
        c.font=_font(True,"FFFFFF",9); c.fill=_fill(MED_BLUE)
        c.alignment=_aln("center"); c.border=_bdr()
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[2].height=20; ws.freeze_panes="A3"
    return ws

ri_map = {}

def wr(ws, vals, bold=False, bg=None):
    ri = ri_map.get(ws.title, 3)
    bg_use = bg or (GREY_BG if ri%2==0 else WHITE)
    for ci,v in enumerate(vals,1):
        c = ws.cell(row=ri,column=ci,value=v)
        c.font=_font(bold,"000000",9); c.fill=_fill(bg_use)
        is_num = isinstance(v,(int,float))
        is_fml = _is_fml(v)
        c.alignment=_aln("right" if (is_num or is_fml) else "left")
        c.border=_bdr()
        if is_num or is_fml: c.number_format=NUM_FMT
    ws.row_dimensions[ri].height=15
    ri_map[ws.title]=ri+1

def tot_row(ws, vals, bg="D6DCE4"):
    ri = ri_map.get(ws.title, 3)
    for ci,v in enumerate(vals,1):
        c = ws.cell(row=ri,column=ci,value=v)
        c.font=_font(True,"000000",9); c.fill=_fill(bg)
        is_num = isinstance(v,(int,float))
        is_fml = _is_fml(v)
        c.alignment=_aln("right" if (is_num or is_fml) else "left")
        c.border=_bdr()
        if is_num or is_fml: c.number_format=NUM_FMT
    ws.row_dimensions[ri].height=18; ri_map[ws.title]=ri+1

def month_sep(ws, label, ncols, color=MED_BLUE):
    ri = ri_map.get(ws.title, 3)
    ws.merge_cells(f"A{ri}:{get_column_letter(ncols)}{ri}")
    c = ws.cell(row=ri,column=1,value=f"-- {label} --")
    c.font=_font(True,"FFFFFF",9); c.fill=_fill(color)
    c.alignment=_aln("left"); c.border=_bdr()
    ws.row_dimensions[ri].height=14; ri_map[ws.title]=ri+1

# -- GSTIN Name Lookup -----------------------------------------
def build_name_map(folder):
    """Build GSTIN→Name dict from GSTR-2A/2B files."""
    name_map = {}
    folder = Path(folder)

    def try_gstin_name(g, n):
        if (len(g)==15 and (g[:2].isdigit() or g[0].isalpha())
                and n and n.lower() not in ("nan","none","")):
            name_map[g] = n.strip()

    def scan_xl(path):
        try:
            xf = pd.ExcelFile(path, engine="openpyxl")
            for sn in xf.sheet_names:
                try:
                    df = xf.parse(sn, header=None, dtype=str, nrows=5000)
                    for _, row in df.iterrows():
                        for ci in range(min(5,len(row)-1)):
                            try_gstin_name(str(row.iloc[ci] or "").strip(),
                                           str(row.iloc[ci+1] or "").strip())
                except: pass
        except: pass

    # 2A/2B direct xlsx
    for pat in ["*_R2A*.xlsx","GSTR2A_*.xlsx","GSTR2B_*.xlsx","*_R2B*.xlsx","*2B*.xlsx","*2A*.xlsx"]:
        for xl in folder.glob(pat): scan_xl(xl)

    # 2A/2B ZIPs
    for zp in list(folder.glob("GSTR2A_*.zip"))+list(folder.glob("*_R2A*.zip"))+list(folder.glob("GSTR2B_*.zip")):
        try:
            ed = folder/(zp.stem+"_nm"); ed.mkdir(exist_ok=True)
            with zipfile.ZipFile(zp) as z: z.extractall(ed)
            for jf in list(ed.glob("*.json"))+list(ed.glob("**/*.json")):
                try:
                    d = json.load(open(jf, encoding="utf-8"))
                    for ent in d.get("b2b",[]):
                        try_gstin_name(ent.get("ctin",""), ent.get("trdnm",""))
                except: pass
            for xl in list(ed.glob("*.xlsx"))+list(ed.glob("**/*.xlsx")):
                scan_xl(xl)
        except: pass

    # GSTR-1 ZIPs themselves (trdnm field)
    for zp in sorted(folder.glob("GSTR1_*.zip"), key=sort_key_month):
        try:
            ed = folder/(zp.stem+"_ex"); ed.mkdir(exist_ok=True)
            with zipfile.ZipFile(zp) as z: z.extractall(ed)
            for jf in list(ed.glob("*.json"))+list(ed.glob("**/*.json")):
                try:
                    d = json.load(open(jf, encoding="utf-8"))
                    for ent in d.get("b2b",[]):
                        g=ent.get("ctin",""); n=ent.get("trdnm","")
                        if n: try_gstin_name(g, n)
                except: pass
        except: pass

    print(f"    Name map from files: {len(name_map)} entries")

    # -- Load customer name Excel (user-provided GSTIN→Name) ----------------
    # Place customer_names.xlsx / customers.xlsx / party_master.xlsx etc. in the same folder
    folder = Path(folder)
    for _pat in ["customer_names.xlsx","customers.xlsx","party_master.xlsx",
                 "customer_master.xlsx","GSTIN_Names.xlsx","gstin_names.xlsx",
                 "customer_list.xlsx","PartyMaster.xlsx","CustomerMaster.xlsx"]:
        _cf = folder / _pat
        if not _cf.exists():
            _cf = folder.parent / _pat  # also check parent folder
        if _cf.exists():
            try:
                _xl = pd.ExcelFile(_cf, engine="openpyxl")
                _n_added = 0
                for _sn in _xl.sheet_names:
                    try:
                        _df = _xl.parse(_sn, header=None, dtype=str, nrows=20000)
                        for _, _rw in _df.iterrows():
                            for _ci in range(min(8, len(_rw)-1)):
                                _g = str(_rw.iloc[_ci] or "").strip().upper()
                                _n = str(_rw.iloc[_ci+1] or "").strip()
                                if (len(_g)==15 and _g[:2].isdigit()
                                        and _n and _n.lower() not in ("nan","none","")):
                                    name_map[_g] = _n; _n_added += 1
                    except: pass
                if _n_added: print(f"    Customer Excel [{_cf.name}]: {_n_added} entries")
            except Exception as _ce:
                print(f"    Customer Excel error: {_ce}")

    # API fallback for remaining unknown GSTINs
    unknown = []
    for zp in sorted(folder.glob("GSTR1_*.zip"), key=sort_key_month):
        try:
            ed = folder/(zp.stem+"_ex")
            for jf in list(ed.glob("*.json"))+list(ed.glob("**/*.json")):
                d = json.load(open(jf, encoding="utf-8"))
                for ent in d.get("b2b",[]):
                    g=ent.get("ctin","").strip()
                    if g and g not in name_map and g not in unknown:
                        unknown.append(g)
        except: pass

    if unknown:
        print(f"    Trying API lookup for {len(unknown)} unknown GSTINs...")
        try:
            import urllib.request, json as _j
            for g in unknown[:30]:
                try:
                    # GST public search API
                    url = f"https://sheet.gstincheck.co.in/check/apikey/{g}"
                    req = urllib.request.Request(
                        f"https://api.knowyourgst.com/gstin/{g}",
                        headers={"User-Agent":"Mozilla/5.0","Accept":"application/json"})
                    with urllib.request.urlopen(req, timeout=4) as resp:
                        data = _j.loads(resp.read())
                        nm = (data.get("lgnm") or data.get("tradeNam") or
                              data.get("tradeName") or data.get("legal_name") or "")
                        if nm and nm.lower() not in ("null","none",""):
                            name_map[g] = nm
                            print(f"      {g} → {nm}")
                except: pass
                time.sleep(0.15)
        except Exception as e:
            print(f"    API lookup error: {e}")

    print(f"    Final name map: {len(name_map)} entries")
    return name_map


def extract_gstr1_to_excel(folder, output_path=None):
    folder = Path(folder)
    zips = sorted(folder.glob("GSTR1_*.zip"), key=sort_key_month)
    if not zips:
        print(f"  No GSTR1_*.zip files in {folder}")
        return

    print(f"  Processing {len(zips)} ZIP files in calendar order:")
    for z in zips:
        print(f"    {z.name}")

    print(f"\n  Building supplier name map...")
    name_map = build_name_map(folder)

    wb = Workbook(); wb.remove(wb.active)

    # -- Create all sheets -------------------------------------
    wsB2B  = make_sheet(wb,"B2B_Invoices",
        ["Month","Supplier GSTIN","Customer GSTIN","Customer Name",
         "Invoice No","Invoice Date","Invoice Value ₹","POS",
         "Rev Charge","Taxable Value ₹","IGST ₹","CGST ₹","SGST ₹","Total Tax ₹"],
        [10,22,22,32,18,13,16,6,6,16,12,12,12,14], DARK_BLUE)
    ri_map["B2B_Invoices"]=3

    wsITM  = make_sheet(wb,"B2B_ItemDetails",
        ["Month","Customer GSTIN","Invoice No","Item No","HSN Code","Description",
         "UOM","Rate %","Taxable Value ₹","IGST ₹","CGST ₹","SGST ₹","Cess ₹","Total Tax ₹"],
        [10,22,18,7,12,28,8,7,16,12,12,12,9,13], "2E4053")
    ri_map["B2B_ItemDetails"]=3

    wsHSN  = make_sheet(wb,"HSN_Summary",
        ["Month","HSN Code","Description","UOM","Total Qty","Taxable Value ₹",
         "IGST ₹","CGST ₹","SGST ₹","Cess ₹","Total Tax ₹"],
        [10,13,28,8,11,16,12,12,12,9,14], "4472C4")
    ri_map["HSN_Summary"]=3

    wsB2CS = make_sheet(wb,"B2CS_Summary",
        ["Month","State Code","Supply Type","Rate %",
         "Taxable Value ₹","IGST ₹","CGST ₹","SGST ₹","Total Tax ₹"],
        [10,12,14,8,16,12,12,12,14], "375623")
    ri_map["B2CS_Summary"]=3

    wsB2CL = make_sheet(wb,"B2CL_Invoices",
        ["Month","POS","Invoice No","Invoice Date","Invoice Value ₹",
         "Rate %","Taxable Value ₹","IGST ₹"],
        [10,8,18,13,15,8,15,12], "404040")
    ri_map["B2CL_Invoices"]=3

    wsCR   = make_sheet(wb,"CDNR_CreditNotes",
        ["Month","Customer GSTIN","Customer Name","Note No","Note Date","Note Value ₹",
         "POS","Pre-GST","Rate %","Taxable Value ₹","IGST ₹","CGST ₹","SGST ₹"],
        [10,22,28,18,13,14,6,7,7,15,12,12,12], "276221")
    ri_map["CDNR_CreditNotes"]=3

    wsDR   = make_sheet(wb,"CDNR_DebitNotes",
        ["Month","Customer GSTIN","Customer Name","Note No","Note Date","Note Value ₹",
         "POS","Pre-GST","Rate %","Taxable Value ₹","IGST ₹","CGST ₹","SGST ₹"],
        [10,22,28,18,13,14,6,7,7,15,12,12,12], "C00000")
    ri_map["CDNR_DebitNotes"]=3

    wsCU   = make_sheet(wb,"CDNUR_Unregistered",
        ["Month","Note Type","Supply Type","Note No","Note Date",
         "Note Value ₹","Rate %","Taxable Value ₹","IGST ₹"],
        [10,10,12,18,13,14,8,15,12], "7030A0")
    ri_map["CDNUR_Unregistered"]=3

    wsEXP  = make_sheet(wb,"Exports",
        ["Month","Export Type","Invoice No","Invoice Date","Invoice Value ₹",
         "Port Code","Shipping Bill No","Shipping Bill Date",
         "Rate %","Taxable Value ₹","IGST ₹"],
        [10,14,18,13,15,12,18,16,8,15,12], "843C0C")
    ri_map["Exports"]=3

    wsNIL  = make_sheet(wb,"Nil_Rated",
        ["Month","Supply Type","Nil Rated ₹","Exempt ₹","Non-GST ₹","Total ₹"],
        [10,16,16,16,16,16], "808080")
    ri_map["Nil_Rated"]=3

    wsAMD  = make_sheet(wb,"GSTR1A_Amendments",
        ["Month","Section","Customer GSTIN","Customer Name",
         "Original Ref No","Amended No","Date","Note Value ₹",
         "Rate %","Taxable ₹","IGST ₹","CGST ₹","SGST ₹"],
        [10,10,22,26,18,18,13,14,7,14,12,12,12], "9C6500")
    ri_map["GSTR1A_Amendments"]=3

    wsDS   = make_sheet(wb,"Document_Summary",
        ["Month","Document Type","From No","To No","Total Issued","Cancelled","Net Issued"],
        [10,22,14,14,14,12,12], "4472C4")
    ri_map["Document_Summary"]=3

    wsMS   = make_sheet(wb,"Master_Summary",
        ["Month","B2B Invoices","B2B Taxable ₹","B2CS Taxable ₹","B2CL Taxable ₹",
         "Exports ₹","Nil/Exempt ₹","CDN Credit TV ₹","CDN Debit TV ₹",
         "IGST ₹","CGST ₹","SGST ₹","Total Tax ₹","Total Invoice Value ₹"],
        [12,10,16,16,15,14,14,16,16,12,12,12,14,16], DARK_BLUE)
    ri_map["Master_Summary"]=3

    # Annual accumulators
    ann = {k:0.0 for k in ["inv","b2b_tx","b2cs_tx","b2cl_tx","exp_tx",
                             "nil_tx","cdn_cr","cdn_dr","ig","cg","sg","iv"]}

    for zp in zips:
        stem = zp.stem
        parts = stem.split("_")
        # Extract month label: "GSTR1_April_2024" → "April 2024"
        if len(parts) >= 3:
            month_lbl = f"{parts[-2]} {parts[-1]}"
        else:
            month_lbl = stem.replace("GSTR1_","").replace("_"," ")

        try:
            ed = folder/(stem+"_ex"); ed.mkdir(exist_ok=True)
            with zipfile.ZipFile(zp) as z: z.extractall(ed)
            jfiles = list(ed.glob("*.json"))+list(ed.glob("**/*.json"))
            if not jfiles:
                print(f"  No JSON in {zp.name}"); continue
            d = json.load(open(jfiles[0], encoding="utf-8"))
        except Exception as e:
            print(f"  Error {zp.name}: {e}"); continue

        print(f"  Processing {month_lbl}...")
        gstin_self = d.get("gstin","")
        m = {k:0.0 for k in ["inv","b2b_tx","b2cs_tx","b2cl_tx","exp_tx",
                               "nil_tx","cdn_cr","cdn_dr","ig","cg","sg","iv"]}

        # -- B2B ----------------------------------------------
        month_sep(wsB2B,  month_lbl, 14)
        month_sep(wsITM,  month_lbl, 14)
        for entry in d.get("b2b",[]):
            ctin = entry.get("ctin","").strip()
            nm   = name_map.get(ctin,"") or entry.get("trdnm","") or ""
            for inv in entry.get("inv",[]):
                rc   = inv.get("rchrg","N")
                inum = inv.get("inum",""); idt=inv.get("idt","")
                ival = float(inv.get("val",0) or 0)
                pos  = inv.get("pos","")
                inv_tx=inv_ig=inv_cg=inv_sg=0.0
                inv_rates=[]  # collect all rates in this invoice
                for it_no, it in enumerate(inv.get("itms",[]),1):
                    det  = it.get("itm_det",{})
                    rt   = det.get("rt",0)
                    if rt and rt not in inv_rates: inv_rates.append(rt)
                    tv   = float(det.get("txval",0) or 0)
                    ig   = float(det.get("iamt",0)  or 0)
                    cg   = float(det.get("camt",0)  or 0)
                    sg   = float(det.get("samt",0)  or 0)
                    cs   = float(det.get("csamt",0) or 0)
                    item_num = it.get("num", it_no)
                    inv_tx+=tv; inv_ig+=ig; inv_cg+=cg; inv_sg+=sg
                    wr(wsITM,[month_lbl,ctin,inum,item_num,"","",
                               "",rt,round(tv,2),round(ig,2),round(cg,2),round(sg,2),
                               round(cs,2),round(ig+cg+sg+cs,2)])
                # Rate label: "18" or "5/18" for mixed-rate invoices
                rate_lbl = (inv_rates[0] if len(inv_rates)==1
                            else "/".join(str(int(r) if r==int(r) else r)
                                         for r in sorted(inv_rates))
                            if inv_rates else 0)
                wr(wsB2B,[month_lbl,gstin_self,ctin,nm,inum,idt,round(ival,2),pos,rc,
                           round(inv_tx,2),round(inv_ig,2),round(inv_cg,2),round(inv_sg,2),
                           round(inv_ig+inv_cg+inv_sg,2)])
                m["b2b_tx"]+=inv_tx; m["ig"]+=inv_ig; m["cg"]+=inv_cg
                m["sg"]+=inv_sg; m["iv"]+=ival; m["inv"]+=1

        # -- HSN Summary --------------------------------------
        month_sep(wsHSN, month_lbl, 11)
        for hsn_entry in d.get("hsn",{}).get("data",[]):
            hsn_cd = str(hsn_entry.get("hsn_sc","") or "")
            desc   = hsn_entry.get("desc","") or ""
            uom    = hsn_entry.get("uqc","") or ""
            qty    = float(hsn_entry.get("qty",0) or 0)
            tv     = float(hsn_entry.get("val",0) or 0)
            ig     = float(hsn_entry.get("iamt",0) or 0)
            cg     = float(hsn_entry.get("camt",0) or 0)
            sg     = float(hsn_entry.get("samt",0) or 0)
            cs     = float(hsn_entry.get("csamt",0) or 0)
            wr(wsHSN,[month_lbl,hsn_cd,desc,uom,round(qty,3),round(tv,2),
                       round(ig,2),round(cg,2),round(sg,2),round(cs,2),round(ig+cg+sg+cs,2)])

        # -- B2CS ---------------------------------------------
        month_sep(wsB2CS, month_lbl, 9)
        for rec in d.get("b2cs",[]):
            tv=float(rec.get("txval",0) or 0); ig=float(rec.get("iamt",0) or 0)
            cg=float(rec.get("camt",0) or 0);  sg=float(rec.get("samt",0) or 0)
            wr(wsB2CS,[month_lbl,rec.get("pos",""),rec.get("sply_ty","INTRA"),
                        rec.get("rt",0),round(tv,2),round(ig,2),round(cg,2),round(sg,2),
                        round(ig+cg+sg,2)])
            m["b2cs_tx"]+=tv; m["ig"]+=ig; m["cg"]+=cg; m["sg"]+=sg

        # -- B2CL ---------------------------------------------
        month_sep(wsB2CL, month_lbl, 8)
        for rec in d.get("b2cl",[]):
            pos=rec.get("pos","")
            for inv in rec.get("inv",[]):
                inum=inv.get("inum",""); idt=inv.get("idt","")
                ival=float(inv.get("val",0) or 0)
                iv_tx=iv_ig=0.0; rt=0
                for it in inv.get("itms",[]):
                    det=it.get("itm_det",{})
                    iv_tx+=float(det.get("txval",0) or 0)
                    iv_ig+=float(det.get("iamt",0) or 0)
                    rt=det.get("rt",rt)
                wr(wsB2CL,[month_lbl,pos,inum,idt,round(ival,2),rt,round(iv_tx,2),round(iv_ig,2)])
                m["b2cl_tx"]+=iv_tx; m["ig"]+=iv_ig; m["iv"]+=ival

        # -- CDNR ---------------------------------------------
        month_sep(wsCR, month_lbl, 13)
        month_sep(wsDR, month_lbl, 13)
        for entry in d.get("cdnr",[]):
            ctin=entry.get("ctin","").strip()
            nm=name_map.get(ctin,"") or entry.get("trdnm","")
            for note in entry.get("nt",[]):
                ntty=note.get("ntty","C")
                nnum=note.get("nt_num",""); ndt=note.get("nt_dt","")
                nval=float(note.get("val",0) or 0)
                pos=note.get("pos",""); pgst=note.get("p_gst","N")
                tv=ig=cg=sg=0.0; rt=0
                for it in note.get("itms",[]):
                    det=it.get("itm_det",{})
                    tv+=float(det.get("txval",0) or 0); ig+=float(det.get("iamt",0) or 0)
                    cg+=float(det.get("camt",0) or 0);  sg+=float(det.get("samt",0) or 0)
                    rt=det.get("rt",rt)
                row=[month_lbl,ctin,nm,nnum,ndt,round(nval,2),pos,pgst,rt,
                      round(tv,2),round(ig,2),round(cg,2),round(sg,2)]
                if ntty=="D": wr(wsDR,row); m["cdn_dr"]+=tv
                else:         wr(wsCR,row); m["cdn_cr"]+=tv
                m["ig"]+=ig; m["cg"]+=cg; m["sg"]+=sg

        # -- CDNUR --------------------------------------------
        month_sep(wsCU, month_lbl, 9)
        for note in d.get("cdnur",[]):
            ntty=note.get("ntty","C"); spty=note.get("typ","")
            tv=float(note.get("txval",0) or 0); ig=float(note.get("iamt",0) or 0)
            wr(wsCU,[month_lbl,ntty,spty,note.get("nt_num",""),note.get("nt_dt",""),
                      round(float(note.get("val",0) or 0),2),
                      note.get("rt",0),round(tv,2),round(ig,2)])
            m["ig"]+=ig

        # -- Exports ------------------------------------------
        month_sep(wsEXP, month_lbl, 11)
        for exp in d.get("exp",[]):
            etype=exp.get("exp_typ","")
            for inv in exp.get("inv",[]):
                inum=inv.get("inum",""); idt=inv.get("idt","")
                ival=float(inv.get("val",0) or 0)
                tv=ig=0.0; rt=0
                for it in inv.get("itms",[]):
                    det=it.get("itm_det",{})
                    tv+=float(det.get("txval",0) or 0)
                    ig+=float(det.get("iamt",0) or 0); rt=det.get("rt",rt)
                wr(wsEXP,[month_lbl,etype,inum,idt,round(ival,2),
                           inv.get("pcode",""),inv.get("sbnum",""),inv.get("sbdt",""),
                           rt,round(tv,2),round(ig,2)])
                m["exp_tx"]+=tv; m["ig"]+=ig; m["iv"]+=ival

        # -- Nil / Exempt -------------------------------------
        month_sep(wsNIL, month_lbl, 6)
        # AY2024-25: "nil_sup" -> list | AY2025-26: "nil" -> {"inv":[...]}
        _nil_raw = d.get("nil_sup", None) or d.get("nil", {})
        _nil_list = _nil_raw if isinstance(_nil_raw, list) else _nil_raw.get("inv", [])
        for rec in _nil_list:
            spty=rec.get("sply_ty","")
            nil_v =float(rec.get("nil_amt",  rec.get("nil",  0)) or 0)
            expt_v=float(rec.get("expt_amt", rec.get("expt", 0)) or 0)
            ngsup_v=float(rec.get("ngsup_amt",rec.get("ngsup",0)) or 0)
            total_v=round(nil_v+expt_v+ngsup_v,2)
            if total_v == 0: continue
            wr(wsNIL,[month_lbl,spty,round(nil_v,2),round(expt_v,2),round(ngsup_v,2),total_v])
            m["nil_tx"]+=total_v

        # -- GSTR-1A Amendments -------------------------------
        has_amd = any(d.get(k) for k in ["b2ba","cdnra","b2cla","expa"])
        if has_amd:
            month_sep(wsAMD, month_lbl, 13)
            for entry in d.get("b2ba",[]):
                ctin=entry.get("ctin","").strip()
                nm=name_map.get(ctin,"") or entry.get("trdnm","")
                for inv in entry.get("inv",[]):
                    tv=ig=cg=sg=0.0; rt=0
                    for it in inv.get("itms",[]):
                        det=it.get("itm_det",{})
                        tv+=float(det.get("txval",0) or 0)
                        ig+=float(det.get("iamt",0) or 0)
                        cg+=float(det.get("camt",0) or 0)
                        sg+=float(det.get("samt",0) or 0); rt=det.get("rt",rt)
                    wr(wsAMD,[month_lbl,"B2BA",ctin,nm,
                               inv.get("oinum",""),inv.get("inum",""),inv.get("idt",""),
                               round(float(inv.get("val",0) or 0),2),rt,
                               round(tv,2),round(ig,2),round(cg,2),round(sg,2)])
            for entry in d.get("cdnra",[]):
                ctin=entry.get("ctin","").strip()
                nm=name_map.get(ctin,"") or entry.get("trdnm","")
                for note in entry.get("nt",[]):
                    tv=ig=cg=sg=0.0
                    for it in note.get("itms",[]):
                        det=it.get("itm_det",{})
                        tv+=float(det.get("txval",0) or 0)
                        ig+=float(det.get("iamt",0) or 0)
                        cg+=float(det.get("camt",0) or 0)
                        sg+=float(det.get("samt",0) or 0)
                    wr(wsAMD,[month_lbl,"CDNRA",ctin,nm,
                               note.get("ont_num",""),note.get("nt_num",""),note.get("nt_dt",""),
                               round(float(note.get("val",0) or 0),2),0,
                               round(tv,2),round(ig,2),round(cg,2),round(sg,2)])

        # -- Document Summary ---------------------------------
        for ds_entry in d.get("doc_issue",{}).get("doc_det",[]):
            dtype=ds_entry.get("ty","")
            for doc in ds_entry.get("docs",[]):
                wr(wsDS,[month_lbl,dtype,
                          doc.get("from",""),doc.get("to",""),
                          int(doc.get("totnum",0) or 0),
                          int(doc.get("cancel",0) or 0),
                          int(doc.get("totnum",0) or 0)-int(doc.get("cancel",0) or 0)])

        # -- Master Summary ------------------------------------
        wr(wsMS,[month_lbl,int(m["inv"]),
                  round(m["b2b_tx"],2),round(m["b2cs_tx"],2),round(m["b2cl_tx"],2),
                  round(m["exp_tx"],2),round(m["nil_tx"],2),round(m["cdn_cr"],2),round(m["cdn_dr"],2),
                  round(m["ig"],2),round(m["cg"],2),round(m["sg"],2),
                  round(m["ig"]+m["cg"]+m["sg"],2),round(m["iv"],2)])
        for k in ann: ann[k]+=m.get(k,0)

    # Annual total — use SUM formulas so Excel recalculates if data changes
    _ms_ri = ri_map.get(wsMS.title, 3)  # current row = first empty row after data
    _ms_data_start = 3                   # data starts at row 3
    _ms_data_end   = _ms_ri - 1         # last data row
    tot_row(wsMS,["ANNUAL TOTAL",
                   _fsum("B",_ms_data_start,_ms_data_end),
                   _fsum("C",_ms_data_start,_ms_data_end),
                   _fsum("D",_ms_data_start,_ms_data_end),
                   _fsum("E",_ms_data_start,_ms_data_end),
                   _fsum("F",_ms_data_start,_ms_data_end),
                   _fsum("G",_ms_data_start,_ms_data_end),
                   _fsum("H",_ms_data_start,_ms_data_end),
                   _fsum("I",_ms_data_start,_ms_data_end),
                   _fsum("J",_ms_data_start,_ms_data_end),
                   _fsum("K",_ms_data_start,_ms_data_end),
                   _fsum("L",_ms_data_start,_ms_data_end),
                   f"=SUM(J{_ms_ri}:L{_ms_ri})",
                   _fsum("N",_ms_data_start,_ms_data_end)])

    if output_path is None:
        output_path = folder/f"GSTR1_FULL_DETAIL_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(str(output_path))
    print(f"\n  ✓ Saved: {Path(output_path).name}")
    print(f"  Sheets ({len(wb.worksheets)}): {[s.title for s in wb.worksheets]}")
    return str(output_path)


if __name__ == "__main__":
    folder = sys.argv[1] if len(sys.argv)>1 else "."
    print(f"\nGSTR-1 Comprehensive Extractor v2 → {folder}\n")
    extract_gstr1_to_excel(folder)
