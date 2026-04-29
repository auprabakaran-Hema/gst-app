"""
================================================================================
  RUN ALL — Unified GST + Income Tax Automation Pipeline  v2.1
================================================================================

  PIPELINE ORDER (enforced):
  ─────────────────────────────────────────────────────────────────────────────
  Step 2+3  GST CAPTCHA entered ONCE → all clients download (GSTR-1/2B/3B/1A)
            gst_suite output → MultiYear_{ts}/AY{fy_tag}/{ClientName}/
  Step 4+5  IT portal → 26AS / AIS / TIS per PAN   (IT_OUT_DIR passed via env)
            it_suite output  → AY{AY_LABEL}_{ts}/{ClientName}/
  Step 6    Master Bridge — GST ↔ IT reconciliation
  Step 6b   GST-IT Comparison Excel (TIS / AIS template)
  Step 6c   GSTR-2B Consolidated Extractor per GSTIN
  ─────────────────────────────────────────────────────────────────────────────

  HOW TO RUN:
    python run_all.py                  ← full run (steps 2→6c)
    python run_all.py --only-gst       ← steps 2-3 only
    python run_all.py --only-it        ← steps 4-5 only
    python run_all.py --only-bridge    ← steps 6-6c only
    python run_all.py --offline        ← pick folders, run bridge (no downloads)
    python run_all.py --fy 2024-25     ← override FY
    python run_all.py --client "RAVI"  ← one client only

  KEY FIXES v2.1:
  ✓ gst_suite creates MultiYear_*/AY{fy}/ — GST_RUN auto-discovered after run
  ✓ IT_OUT_DIR env var passed to it_suite so it writes to the tracked folder
  ✓ IT_RUN auto-discovered if it_suite ignores env var and creates own folder
  ✓ Folder picker: GST expands MultiYear_* correctly; IT shows AY* as-is
    (bridge needs AY* parent, NOT client subfolders)
  ✓ Duplicate IT recon eliminated — it_suite v6 runs recon internally;
    run_all only invokes it_recon_engine.py as fallback if Excel missing
  ✓ Bridge always receives the real discovered folder paths
  ✓ Correct step order enforced with clear banners
================================================================================
"""
import os, sys, re, subprocess, argparse, logging
from pathlib import Path
from datetime import datetime

SCRIPT_DIR = Path(__file__).parent.resolve()

# ─── Change only ONE line each year ──────────────────────────────────────────
FY_LABEL   = "2025-26"
_fy_yr     = int(FY_LABEL.split("-")[0])
AY_LABEL   = f"{_fy_yr + 1}-{str(_fy_yr + 2)[2:]}"   # "2026-27"

# ─── Output base folder ───────────────────────────────────────────────────────
def _find_base_dir():
    home = Path.home()
    candidates = [
        home / "OneDrive" / "Desktop" / "OUTPUT",
        home / "OneDrive - Personal" / "Desktop" / "OUTPUT",
        *[p / "Desktop" / "OUTPUT" for p in home.glob("OneDrive*") if p.is_dir()],
        home / "Desktop" / "OUTPUT",
        home / "Downloads",
    ]
    for c in candidates:
        if c.exists():
            return c
    onedrive = next((p for p in home.glob("OneDrive*") if p.is_dir()), None)
    if onedrive:
        return onedrive / "Desktop" / "OUTPUT"
    return home / "Downloads"

BASE_DIR   = _find_base_dir()
GST_BASE   = BASE_DIR / "GST_Automation"
IT_BASE    = BASE_DIR / "IT_Automation"
RUN_TS     = datetime.now().strftime("%Y%m%d_%H%M")

# Stub paths — updated to real folders after each suite finishes
GST_RUN    = GST_BASE / f"FY{FY_LABEL}_{RUN_TS}"
IT_RUN     = IT_BASE  / f"AY{AY_LABEL}_{RUN_TS}"

VARIANCE_THRESHOLD = 5000
LOG_FILE   = BASE_DIR / f"run_all_{RUN_TS}.log"

MISSING = []
try:    import pandas as pd
except: MISSING.append("pandas")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except: MISSING.append("openpyxl")

if MISSING:
    print(f"✗ Missing: pip install {' '.join(MISSING)}")
    sys.exit(1)

# ─── Colours ─────────────────────────────────────────────────────────────────
DARK_BLUE = "1F3864"; MED_BLUE  = "2E75B6"
HDR_BG    = "1F3864"; TOT_BG    = "D6DCE4"
ALT1      = "FFFFFF"; ALT2      = "F2F2F2"
GREEN_BG  = "C6EFCE"; RED_BG    = "FFC7CE"; YELLOW_BG = "FFEB9C"
GREEN_FG  = "276221"; RED_FG    = "9C0006"
NUM_FMT   = "#,##0.00"
FY_MONTHS = ["APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC","JAN","FEB","MAR"]


# ═══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════════════════════
def _setup_log():
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)-8s | %(message)s",
        handlers=[
            logging.FileHandler(LOG_FILE, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ]
    )
    return logging.getLogger("run_all")

log = None
def _log(msg, level="info"):
    if log: getattr(log, level)(msg)
    else:   print(msg)

def _banner(title):
    _log("=" * 72)
    _log(f"  {title}")
    _log("=" * 72)


# ═══════════════════════════════════════════════════════════════════════════════
# CLIENT LOADER
# ═══════════════════════════════════════════════════════════════════════════════
def _col(df, *cands):
    def _norm(s):
        s = str(s).lower().strip()
        s = re.sub(r'[\s\n\r()/\\-]+', '_', s)
        return re.sub(r'_+', '_', s).strip('_')
    norm_cols = {_norm(c): c for c in df.columns}
    for cand in cands:
        h = norm_cols.get(_norm(cand))
        if h is not None: return h
    return None

def _clean(s):
    if s is None: return ""
    s = str(s).strip()
    return "" if s.lower() in ("nan","none","") else s

def load_clients(fy_override=None, name_filter=None):
    candidates = [
        "Client_Manager_Secure_AY2025-26.xlsx",
        "clients_manager.xlsx",
        "clients.xlsx",
        "clients.csv",
    ]
    for fname in candidates:
        fpath = SCRIPT_DIR / fname
        if not fpath.exists(): continue
        _log(f"  Loading clients from: {fname}")
        try:
            if fname.endswith(".csv"):
                df = pd.read_csv(fpath, dtype=str).fillna("")
            else:
                xl  = pd.ExcelFile(fpath, engine="openpyxl")
                sht = next(
                    (s for s in xl.sheet_names if any(k in s.lower() for k in ["client","cred","🔐"])),
                    xl.sheet_names[0])
                # Some files (e.g. Client_Manager_Secure) have a merged title in row 0
                # and the real column headers in row 2. Auto-detect the header row.
                df_raw = xl.parse(sht, header=None, dtype=str).fillna("")
                hdr_row = 0
                for _ri, _row in df_raw.iterrows():
                    _rt = " ".join(str(v).lower() for v in _row if str(v).strip())
                    if any(k in _rt for k in ["client name","gstin","pan","gst portal"]):
                        hdr_row = _ri; break
                df = xl.parse(sht, header=hdr_row, dtype=str).fillna("")
            df.columns = [re.sub(r'[\n\r]+','_', str(c)).strip().lower().replace(" ","_") for c in df.columns]
            df.columns = [re.sub(r'_+','_', c).strip('_') for c in df.columns]

            c_name = _col(df,"client_name","name","company_name","client name")
            c_pan  = _col(df,"pan","pan_no","pan_number"," pan ")
            c_gst  = _col(df,"gstin","gst_number","gstin_number","gst_no")
            c_dob  = _col(df,"dob","date_of_birth","date_of_birth_(ddmmyyyy)")
            c_itu  = _col(df,"it_username","it_user","income_tax_username","it username")
            c_itp  = _col(df,"it_password","income_tax_password","it password")
            c_gstu = _col(df,"gst_username","gst_user","gst_portal_username","gst username")
            c_gstp = _col(df,"gst_password","gst_portal_password","gst password")
            c_act  = _col(df,"active","status","active_(yes/no)","active_yes/no")
            c_fy   = _col(df,"fy","financial_year","fin_year"," fy")

            clients = []
            for _, row in df.iterrows():
                name = _clean(row.get(c_name,"")) if c_name else ""
                if not name: continue
                active = _clean(row.get(c_act,"YES")).upper() if c_act else "YES"
                if active not in ("YES","Y","1","TRUE","ACTIVE",""): continue
                if name_filter and name_filter.lower() not in name.lower(): continue
                pan    = _clean(row.get(c_pan,"")).upper() if c_pan else ""
                raw_g  = _clean(row.get(c_gst,"")) if c_gst else ""
                gstins = [g.strip() for g in re.split(r"[,;/\n]",raw_g) if g.strip()]
                fy_raw = _clean(row.get(c_fy,"")) if c_fy else ""
                fy     = fy_override or (fy_raw if re.match(r"\d{4}-\d{2,4}",fy_raw) else FY_LABEL)
                clients.append({
                    "name":         name,
                    "pan":          pan,
                    "gstin":        gstins,
                    "fy":           fy,
                    "dob":          _clean(row.get(c_dob,"")) if c_dob else "",
                    "it_username":  _clean(row.get(c_itu,"")) if c_itu else "",
                    "it_password":  _clean(row.get(c_itp,"")) if c_itp else "",
                    "gst_username": _clean(row.get(c_gstu,"")) if c_gstu else "",
                    "gst_password": _clean(row.get(c_gstp,"")) if c_gstp else "",
                })
            if clients:
                _log(f"  Loaded {len(clients)} client(s)")
                return clients
        except Exception as e:
            _log(f"  ✗ Could not read {fname}: {e}", "error")

    _log("✗ No clients file found. Create clients.xlsx with columns:", "error")
    _log("  Client Name | PAN | GSTIN | DOB | IT Username | IT Password |", "error")
    _log("  GST Username | GST Password | Active | FY", "error")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════════════
# FOLDER DISCOVERY
# ═══════════════════════════════════════════════════════════════════════════════
def _iter_gst_run_dirs():
    """
    Yield all GST run directories, newest first.
    Handles both layouts:
      run_all layout  → GST_Automation/FY2025-26_YYYYMMDD/
      gst_suite layout → GST_Automation/MultiYear_ts/AY2025_26/
    """
    if not GST_BASE.exists(): return
    for d in sorted(GST_BASE.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if not d.is_dir(): continue
        if d.name.startswith("MultiYear"):
            for sub in sorted(d.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
                if sub.is_dir(): yield sub
        else:
            yield d


def _discover_real_gst_run(fy=None):
    """
    After gst_suite finishes, find the folder that actually has output.
    gst_suite creates: GST_Automation/MultiYear_{ts}/AY{fy_tag}/
    Priority:
      1. Current GST_RUN if it has real files (run_all scenario)
      2. Newest MultiYear_*/AY{fy_tag}/ with ANNUAL_RECONCILIATION files
      3. Any GST_BASE child with ANNUAL_RECONCILIATION
    """
    fy_tag = (fy or FY_LABEL).replace("-","_")

    if GST_RUN.exists() and any(GST_RUN.rglob("ANNUAL_RECONCILIATION*.xlsx")):
        return GST_RUN

    for myd in sorted(GST_BASE.glob("MultiYear_*"), key=lambda d: d.stat().st_mtime, reverse=True):
        if not myd.is_dir(): continue
        for sub in sorted(myd.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
            if sub.is_dir() and any(sub.rglob("ANNUAL_RECONCILIATION*.xlsx")):
                return sub

    for d in sorted(GST_BASE.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if d.is_dir() and any(d.rglob("ANNUAL_RECONCILIATION*.xlsx")):
            return d

    # Fallback 2: newest MultiYear_*/AY* with GSTR2B files
    for myd in sorted(GST_BASE.glob("MultiYear_*"), key=lambda d: d.stat().st_mtime, reverse=True):
        if not myd.is_dir(): continue
        for sub in sorted(myd.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
            if sub.is_dir() and any(sub.rglob("GSTR2B*.xlsx")):
                return sub

    # Last resort: newest AY* sub of any MultiYear_* regardless of content
    for myd in sorted(GST_BASE.glob("MultiYear_*"), key=lambda d: d.stat().st_mtime, reverse=True):
        if not myd.is_dir(): continue
        for sub in sorted(myd.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
            if sub.is_dir():
                return sub

    return GST_RUN


def _discover_real_it_run():
    """
    After it_suite finishes, find the IT folder with real data.
    it_suite respects IT_OUT_DIR env var; if used correctly IT_RUN has content.
    If it ignored env var and created its own AY*_HHMM folder, discover it.

    KEY FIX (Issue 2): it_suite sometimes runs its own recon internally and
    creates a DIFFERENT AY* folder than IT_RUN. We must find the folder that
    contains the LARGEST IT_RECONCILIATION xlsx (most data), not just any folder.
    """
    if not IT_BASE.exists():
        return IT_RUN

    candidates = sorted(
        [d for d in IT_BASE.iterdir() if d.is_dir()],
        key=lambda d: d.stat().st_mtime, reverse=True
    )

    # Pass 1: find the folder whose IT_RECONCILIATION is the largest (real data)
    best_folder = None
    best_size   = 0
    for d in candidates:
        for xl in d.rglob("IT_RECONCILIATION*.xlsx"):
            try:
                sz = xl.stat().st_size
                if sz > best_size:
                    best_size   = sz
                    best_folder = d
            except: pass

    if best_folder and best_size >= 25_000:
        return best_folder

    # Pass 2: folder with any PDFs (downloads happened but recon not yet built)
    for d in candidates:
        if any(d.rglob("*.pdf")):
            return d

    # Pass 3: current IT_RUN if it exists at all
    if IT_RUN.exists():
        return IT_RUN

    return IT_RUN


def _find_gst_excel_for_client(gstins, name, fy=None):
    """Find the best GST reconciliation Excel for a client."""
    search_dirs = [GST_RUN] + list(_iter_gst_run_dirs())
    for run_dir in search_dirs:
        for gstin in gstins:
            for d in [run_dir / gstin, run_dir / name.replace(" ","_")]:
                if not d.exists(): continue
                excels = sorted(d.glob("*.xlsx"), key=lambda p: p.stat().st_size, reverse=True)
                for xl in excels:
                    if "IT_RECONCILIATION" not in xl.name.upper():
                        return xl
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2+3 — GST SUITE
# ═══════════════════════════════════════════════════════════════════════════════
def run_gst_step(clients):
    """
    Launch gst_suite_v31.py interactively.

    gst_suite has NO --gstin/--out args. It uses its own interactive main()
    that asks which returns to download, which FY range, reads clients.xlsx
    itself, and creates: GST_Automation/MultiYear_{ts}/AY{fy_tag}/{ClientName}/

    After it exits, _discover_real_gst_run() updates GST_RUN to the real folder.
    """
    global GST_RUN
    _banner("STEP 2+3 — GST Portal Download + Reconciliation Excel")
    _log("  ► Browser will open. Enter CAPTCHA ONCE — all clients processed.")
    _log(f"  ► Output: {GST_BASE}/MultiYear_*/AY{FY_LABEL.replace('-','_')}/")

    gst_script = SCRIPT_DIR / "gst_suite_v31.py"
    if not gst_script.exists():
        _log("  ✗ gst_suite_v31.py not found — skipping", "warning"); return

    GST_BASE.mkdir(parents=True, exist_ok=True)

    try:
        result = subprocess.run(
            [sys.executable, str(gst_script)],
            timeout=7200,
            capture_output=False,
        )
        if result.returncode == 0:
            _log("  ✓ gst_suite completed")
        else:
            _log(f"  ✗ gst_suite exited {result.returncode}", "warning")
    except subprocess.TimeoutExpired:
        _log("  ✗ gst_suite timed out (2 hrs)", "warning")
    except Exception as e:
        _log(f"  ✗ gst_suite error: {e}", "error")

    # Discover real output folder
    real = _discover_real_gst_run()
    if real != GST_RUN:
        _log(f"  ℹ  GST output discovered: {real}")
        GST_RUN = real
    else:
        _log(f"  ℹ  GST output: {GST_RUN}")


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4+5 — IT SUITE
# ═══════════════════════════════════════════════════════════════════════════════
def run_it_step(clients):
    """
    Launch it_suite_v6.py with IT_OUT_DIR set so it writes to our tracked folder.
    it_suite v6 checks os.environ['IT_OUT_DIR'] and uses it if the path exists.
    After it exits, verify IT_RUN has content; if not, auto-discover real folder.
    Only run it_recon_engine as fallback if it_suite didn't produce IT_RECONCILIATION.xlsx.
    """
    global IT_RUN
    _banner("STEP 4+5 — IT Portal Download + IT Reconciliation Excel")

    it_script    = SCRIPT_DIR / "it_suite_v6.py"
    recon_script = SCRIPT_DIR / "it_recon_engine.py"

    if not it_script.exists():
        _log("  ✗ it_suite_v6.py not found — skipping", "warning"); return

    # Create tracked folder and pass it to it_suite via env var
    IT_RUN.mkdir(parents=True, exist_ok=True)
    env = os.environ.copy()
    env["IT_OUT_DIR"] = str(IT_RUN)
    env["FY_LABEL"]   = FY_LABEL
    env["AY_LABEL"]   = AY_LABEL

    _log(f"  ► IT_OUT_DIR = {IT_RUN}")
    _log(f"  ► Browser(s) will open. Enter OTP once per client.")

    try:
        result = subprocess.run(
            [sys.executable, str(it_script)],
            env=env,
            timeout=5400,
            capture_output=False,
        )
        if result.returncode == 0:
            _log("  ✓ it_suite completed")
        else:
            _log(f"  ✗ it_suite exited {result.returncode}", "warning")
    except subprocess.TimeoutExpired:
        _log("  ✗ it_suite timed out (90 min)", "warning")
    except Exception as e:
        _log(f"  ✗ it_suite error: {e}", "error")

    # Discover real output folder (handles it_suite ignoring IT_OUT_DIR env var)
    real = _discover_real_it_run()
    if real != IT_RUN:
        _log(f"  ℹ  IT output discovered: {real}")
        IT_RUN = real
    else:
        _log(f"  ℹ  IT output: {IT_RUN}")

    # ── Step 5: Fallback recon ONLY if it_suite produced NO good IT_RECON at all ─
    # CRITICAL: we check the DISCOVERED IT_RUN (not the stub), so we never
    # run a second recon that overwrites the good file with blank data.
    if not recon_script.exists():
        return

    for client in clients:
        name = client["name"]
        pan  = client["pan"]
        fy   = client["fy"]
        if not pan: continue

        # Find client subfolder under DISCOVERED IT_RUN
        it_dir = IT_RUN / name.replace(" ", "_")
        if not it_dir.exists():
            # fuzzy match: first 6 chars of name
            for sub in IT_RUN.iterdir():
                if sub.is_dir() and name.upper()[:6] in sub.name.upper():
                    it_dir = sub; break

        if not it_dir.exists():
            _log(f"    ⚠  IT client folder not found for {name} — skipping fallback recon")
            continue

        # Check size of ALL IT_RECONCILIATION files for this client
        existing = list(it_dir.glob("IT_RECONCILIATION*.xlsx"))
        if any(xl.stat().st_size >= 8_000 for xl in existing):   # it_suite produces ~19KB
            _log(f"    ℹ  Good IT Recon already exists for {name} ({existing[0].stat().st_size//1024} KB) — skipping")
            continue

        # Also check parent folder (in case it_suite put it one level up)
        parent_existing = list(IT_RUN.glob("IT_RECONCILIATION*.xlsx"))
        if any(xl.stat().st_size >= 8_000 for xl in parent_existing):  # it_suite produces ~19KB
            _log(f"    ℹ  IT Recon exists at run-level for {name} — skipping")
            continue

        _log(f"    → Fallback: running it_recon_engine for {name}")
        gst_xl = _find_gst_excel_for_client(client["gstin"], name, fy)
        # Pass GST folder (parent of Excel) so engine can read turnover
        gst_folder_arg = str(Path(gst_xl).parent) if gst_xl else None
        recon_args = [
            sys.executable, str(recon_script),
            str(it_dir), name, pan,
            ",".join(client["gstin"]) if client["gstin"] else "",
            fy,
        ]
        if gst_folder_arg: recon_args += ["--gst-excel", gst_folder_arg]
        try:
            r = subprocess.run(recon_args, timeout=300, capture_output=False)
            if r.returncode == 0: _log(f"    ✓ IT Recon built for {name}")
            else: _log(f"    ✗ Recon engine failed for {pan}", "warning")
        except Exception as e:
            _log(f"    ✗ Recon error for {pan}: {e}", "error")


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 6 — MASTER BRIDGE
# ═══════════════════════════════════════════════════════════════════════════════
def run_bridge_step(clients, gst_folder_override=None, it_folder_override=None):
    _banner("STEP 6 — Master Bridge: GST ↔ IT Final Reconciliation")

    # Always re-discover real folders — the stub paths (GST_RUN / IT_RUN set at
    # startup) may not match where gst_suite / it_suite actually wrote their output.
    real_gst = Path(gst_folder_override) if gst_folder_override else _discover_real_gst_run()
    real_it  = Path(it_folder_override)  if it_folder_override  else _discover_real_it_run()

    # Update globals so subsequent steps (6b, 6c) also use the real paths
    global GST_RUN, IT_RUN
    if real_gst.exists(): GST_RUN = real_gst
    if real_it.exists():  IT_RUN  = real_it

    gst_arg = str(real_gst)
    it_arg  = str(real_it)
    _log(f"  GST folder → {gst_arg}")
    _log(f"  IT  folder → {it_arg}")

    bridge_script = SCRIPT_DIR / "master_bridge.py"
    if bridge_script.exists():
        try:
            r = subprocess.run(
                [sys.executable, str(bridge_script),
                 "--gst", gst_arg, "--it", it_arg, "--fy", FY_LABEL],
                timeout=600, capture_output=False,
            )
            if r.returncode == 0:
                _log("  ✓ master_bridge.py completed"); return
            _log("  ✗ master_bridge.py failed — running built-in bridge", "warning")
        except Exception as e:
            _log(f"  ✗ master_bridge.py error: {e} — running built-in bridge", "error")

    _log("  Running built-in bridge...")
    _builtin_master_bridge(clients)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 6b — GST-IT COMPARISON EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
def run_gst_it_comparison_step():
    _banner("STEP 6b — GST-IT Comparison Excel (TIS / AIS Template)")

    comp_script = SCRIPT_DIR / "build_gst_it_comparison.py"
    out_dir = BASE_DIR / "GST_IT_Comparison"
    out_dir.mkdir(parents=True, exist_ok=True)

    def _find_pdf(prefix):
        best_path, best_mtime = None, 0
        search_base = IT_BASE if IT_BASE.exists() else Path.home() / "Downloads"
        for p in search_base.rglob(f"{prefix}*.pdf"):
            try:
                mt = p.stat().st_mtime
                if mt > best_mtime: best_path = p; best_mtime = mt
            except: pass
        return best_path

    tis_path = _find_pdf("TIS")
    ais_path = _find_pdf("AIS")

    gst_folder = None
    # Priority 1: current GST_RUN (already discovered)
    if GST_RUN.exists() and any(GST_RUN.rglob("GSTR2B*.xlsx")):
        gst_folder = str(GST_RUN)
    # Priority 2: any MultiYear_*/AY*/ subdir that has GSTR2B files
    if not gst_folder and GST_BASE.exists():
        for myd in sorted(GST_BASE.glob("MultiYear_*"),
                          key=lambda d: d.stat().st_mtime, reverse=True):
            if not myd.is_dir(): continue
            for sub in sorted(myd.iterdir(),
                              key=lambda x: x.stat().st_mtime, reverse=True):
                if sub.is_dir() and any(sub.rglob("GSTR2B*.xlsx")):
                    gst_folder = str(sub); break
            if gst_folder: break
    # Priority 3: any direct child of GST_BASE with GSTR2B files
    if not gst_folder and GST_BASE.exists():
        for d in sorted(GST_BASE.rglob("GSTR2B*.xlsx"),
                        key=lambda p: p.stat().st_mtime, reverse=True):
            gst_folder = str(d.parent); break

    _log(f"  GST folder : {gst_folder or '(auto)'}")
    _log(f"  TIS PDF    : {tis_path or '(not found)'}")
    _log(f"  AIS PDF    : {ais_path or '(not found)'}")

    if not comp_script.exists():
        _log("  ⚠  build_gst_it_comparison.py not found — skipping", "warning"); return

    cmd = [sys.executable, str(comp_script), "--out", str(out_dir), "--fy", FY_LABEL]
    if gst_folder: cmd += ["--gst-folder", gst_folder]
    if tis_path:   cmd += ["--tis-pdf",    str(tis_path)]
    if ais_path:   cmd += ["--ais-pdf",    str(ais_path)]
    try:
        r = subprocess.run(cmd, timeout=180, capture_output=False)
        if r.returncode == 0: _log(f"  ✓ GST-IT Comparison built → {out_dir}")
        else: _log("  ✗ build_gst_it_comparison.py failed", "warning")
    except Exception as e:
        _log(f"  ✗ GST-IT Comparison error: {e}", "error")


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 6c — GSTR-2B CONSOLIDATED EXTRACTOR
# ═══════════════════════════════════════════════════════════════════════════════
def run_gstr2b_extractor_step(clients):
    _banner("STEP 6c — GSTR-2B Consolidated Extractor")

    extractor = SCRIPT_DIR / "gstr2b_extractor_v2.py"
    if not extractor.exists():
        _log("  ⚠  gstr2b_extractor_v2.py not found — skipping", "warning"); return

    processed = 0
    seen_2b_folders = set()  # prevent running extractor twice on same folder
                             # (client with multiple GSTINs stored in one ClientName/ folder)
    for client in clients:
        name   = client["name"]
        gstins = client["gstin"]
        fy     = client["fy"]

        for gstin in gstins:
            if not gstin: continue

            # Locate GSTR-2B files for this GSTIN
            gstin_dir = None

            # run_all layout: GST_RUN/GSTIN/
            if (GST_RUN / gstin).exists() and list((GST_RUN / gstin).glob("GSTR2B_*.xlsx")):
                gstin_dir = GST_RUN / gstin
            else:
                # gst_suite layout: MultiYear_*/AY*/ClientName/
                safe_name = name.replace(" ","_").replace("/","_")
                for run_dir in _iter_gst_run_dirs():
                    for sub_name in [gstin, safe_name, name]:
                        d = run_dir / sub_name
                        if d.exists() and list(d.glob("GSTR2B_*.xlsx")):
                            gstin_dir = d; break
                    if not gstin_dir:
                        for sub in run_dir.iterdir():
                            if sub.is_dir() and list(sub.glob("GSTR2B_*.xlsx")):
                                gstin_dir = sub; break
                    if gstin_dir: break

            if not gstin_dir:
                _log(f"    ⚠  No GSTR2B_*.xlsx found for {name} ({gstin}) — skipping")
                continue

            # Skip if we already ran the extractor on this physical folder
            folder_key = str(gstin_dir.resolve())
            if folder_key in seen_2b_folders:
                _log(f"    ℹ  {gstin}: folder already processed ({gstin_dir.name}) — skipping duplicate")
                continue
            seen_2b_folders.add(folder_key)

            # Name the output after the client (not GSTIN) since it's per-folder
            safe_client = name.replace(" ", "_").replace("/", "_")
            out_xl = gstin_dir / f"GSTR2B_Consolidated_Analysis_{safe_client}.xlsx"
            try:
                r = subprocess.run(
                    [sys.executable, str(extractor),
                     "--input", str(gstin_dir), "--output", str(out_xl)],
                    timeout=300, capture_output=False,
                )
                if r.returncode == 0:
                    _log(f"    ✓ 2B Consolidated: {name} ({gstin}) → {out_xl.name}")
                    processed += 1
                else:
                    _log(f"    ✗ Extractor failed for {gstin}", "warning")
            except Exception as e:
                _log(f"    ✗ Extractor error for {gstin}: {e}", "error")

    _log(f"  Step 6c done — {processed} file(s) built")


# ═══════════════════════════════════════════════════════════════════════════════
# BUILT-IN BRIDGE (fallback)
# ═══════════════════════════════════════════════════════════════════════════════
def _builtin_master_bridge(clients):
    out_xl = BASE_DIR / f"Master_Reconciliation_{FY_LABEL}_{RUN_TS}.xlsx"
    wb = Workbook()
    _build_dashboard(wb, clients)
    for client in clients:
        gst_data = _merge_gst_data(client["gstin"], client["fy"])
        it_data  = _read_it_recon(client["pan"], client["name"])
        _build_company_sheet(wb, client, gst_data, it_data)
    wb.save(out_xl)
    _log(f"\n  ✓ MASTER OUTPUT: {out_xl}")
    return out_xl


def _merge_gst_data(gstins, fy):
    merged = {"annual_turnover":0.0,"annual_purchase":0.0,"monthly":{},"gstr2b_itc":[]}
    FY_MON = {"APRIL":"APR","MAY":"MAY","JUNE":"JUN","JULY":"JUL","AUGUST":"AUG",
              "SEPTEMBER":"SEP","OCTOBER":"OCT","NOVEMBER":"NOV","DECEMBER":"DEC",
              "JANUARY":"JAN","FEBRUARY":"FEB","MARCH":"MAR"}

    for gstin in gstins:
        # Find GSTIN folder
        gstin_dir = None
        if (GST_RUN / gstin).exists():
            gstin_dir = GST_RUN / gstin
        else:
            for run_dir in _iter_gst_run_dirs():
                d = run_dir / gstin
                if d.exists(): gstin_dir = d; break
                for sub in run_dir.iterdir():
                    if sub.is_dir() and sub.name.upper().startswith(gstin[:6].upper()):
                        gstin_dir = sub; break
                if gstin_dir: break

        if not gstin_dir or not gstin_dir.exists(): continue

        for xl in gstin_dir.glob("*.xlsx"):
            if "IT_RECONCILIATION" in xl.name.upper(): continue
            try: xf = pd.ExcelFile(xl, engine="openpyxl")
            except: continue

            for sn in xf.sheet_names:
                sn_up = sn.strip().upper()
                mon_abbr = next((abbr for full,abbr in FY_MON.items()
                                 if full in sn_up or sn_up.startswith(abbr)), None)
                if not mon_abbr: continue
                try: df = xf.parse(sn, header=None, dtype=str).fillna("")
                except: continue

                for _, row in df.iterrows():
                    label = str(row.iloc[0]).lower().strip()
                    nums = []
                    for v in row.iloc[1:]:
                        try: nums.append(float(str(v).replace(",","")))
                        except: pass
                    if not nums: continue
                    m = merged["monthly"].setdefault(mon_abbr, {"r1":0.0,"r1a":0.0,"r3b":0.0})
                    if any(k in label for k in ["gstr-1 + gstr-1a","tot_r1_incl","grand total r1"]):
                        m["r1a"] = max(m["r1a"], abs(nums[0]))
                    elif any(k in label for k in ["total from gstr-1","grand total","tot_r1"]):
                        m["r1"] = max(m["r1"], abs(nums[0]))
                    elif any(k in label for k in ["3b","outward supplies"]):
                        m["r3b"] = max(m["r3b"], abs(nums[0]))

            ann_sn = next((s for s in xf.sheet_names if "annual" in s.lower()), None)
            if ann_sn:
                try:
                    adf = xf.parse(ann_sn, header=None, dtype=str).fillna("")
                    for _, row in adf.iterrows():
                        label = str(row.iloc[0]).lower()
                        nums = [float(str(v).replace(",","")) for v in row.iloc[1:]
                                if str(v).replace(",","").replace(".","").replace("-","").isdigit()]
                        if nums and any(k in label for k in ["total taxable","grand total","tot_r1"]):
                            merged["annual_turnover"] += abs(nums[0]); break
                except: pass
    return merged


def _read_it_recon(pan, name):
    result = {"tis_turnover":0.0,"ais_purchase":0.0,"tds_total":0.0,
              "advance_tax":0.0,"monthly_ais":{},"source":""}

    client_dir = IT_RUN / name.replace(" ","_")
    if not client_dir.exists() and IT_BASE.exists():
        for run_dir in sorted(IT_BASE.iterdir(), key=lambda d: d.stat().st_mtime, reverse=True):
            if not run_dir.is_dir(): continue
            cand = run_dir / name.replace(" ","_")
            if cand.exists(): client_dir = cand; break
            for sub in run_dir.iterdir():
                if sub.is_dir() and name.upper()[:6] in sub.name.upper():
                    client_dir = sub; break
            if client_dir.exists(): break

    if not client_dir.exists(): return result

    for xl in client_dir.glob("IT_RECONCILIATION*.xlsx"):
        try:
            xf = pd.ExcelFile(xl, engine="openpyxl"); result["source"] = xl.name
            for sn in xf.sheet_names:
                sn_up = sn.upper()
                if "IT_SUMMARY" in sn_up or "SUMMARY" in sn_up:
                    df = xf.parse(sn, header=None, dtype=str).fillna("")
                    for _, row in df.iterrows():
                        label = str(row.iloc[0]).lower()
                        nums  = [float(str(v).replace(",","")) for v in row.iloc[1:]
                                 if str(v).replace(",","").replace(".","").replace("-","").isdigit()]
                        if not nums: continue
                        if "tis" in label and "turnover" in label: result["tis_turnover"] = abs(nums[0])
                        elif "ais" in label and "purchase" in label: result["ais_purchase"] = abs(nums[0])
                        elif "tds" in label and "total" in label: result["tds_total"] = abs(nums[0])
                        elif "advance" in label and "tax" in label: result["advance_tax"] = abs(nums[0])
                elif "MONTHLY" in sn_up:
                    df = xf.parse(sn, header=None, dtype=str).fillna("")
                    for _, row in df.iterrows():
                        lbl = str(row.iloc[0]).upper().strip()
                        if lbl[:3] in FY_MONTHS:
                            nums = [float(str(v).replace(",","")) for v in row.iloc[1:]
                                    if str(v).replace(",","").replace(".","").replace("-","").isdigit()]
                            if nums: result["monthly_ais"][lbl[:3]] = abs(nums[0])
            break
        except Exception as e:
            _log(f"    ⚠  IT Recon read error: {e}", "warning")
    return result


# ─── Excel helpers ────────────────────────────────────────────────────────────
def _f(h):  return PatternFill("solid", fgColor=h)
def _fn(b=False,c="000000",s=9): return Font(name="Arial",bold=b,color=c,size=s)
def _bd():
    x=Side(style="thin"); return Border(left=x,right=x,top=x,bottom=x)
def _al(h="left",w=False): return Alignment(horizontal=h,vertical="center",wrap_text=w)
def _c(ws,r,col,v,bg=ALT1,bold=False,fg="000000",align="left",numfmt=None,size=9):
    cell=ws.cell(row=r,column=col,value=v)
    cell.font=_fn(bold,fg,size); cell.fill=_f(bg); cell.alignment=_al(align); cell.border=_bd()
    if numfmt and isinstance(v,(int,float)): cell.number_format=numfmt
    elif isinstance(v,(int,float)): cell.number_format=NUM_FMT
    return cell


def _build_dashboard(wb, clients):
    ws=wb.active; ws.title="Dashboard"
    ws.merge_cells("A1:I1")
    ws["A1"].value=f"GST ↔ Income Tax Master Reconciliation — FY {FY_LABEL}"
    ws["A1"].font=_fn(True,"FFFFFF",12); ws["A1"].fill=_f(DARK_BLUE)
    ws["A1"].alignment=_al("center"); ws["A1"].border=_bd(); ws.row_dimensions[1].height=28
    hdrs=[("Client Name",22),("PAN",14),("GSTINs",20),("FY",10),
          ("GST Turnover",16),("IT TIS Turnover",16),("Diff",14),("GST→IT Match",12),("Status",10)]
    for ci,(h,w) in enumerate(hdrs,1):
        c=ws.cell(row=2,column=ci,value=h)
        c.font=_fn(True,"FFFFFF",9); c.fill=_f(HDR_BG); c.alignment=_al("center"); c.border=_bd()
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[2].height=18


def _build_company_sheet(wb, client, gst_data, it_data):
    name=client["name"][:28]; pan=client["pan"]; gstins=client["gstin"]; fy=client["fy"]
    ws=wb.create_sheet(re.sub(r"[\\/*?:\[\]]","",name)[:31])
    ws.merge_cells("A1:L1")
    ws["A1"].value=f"{name} — GST ↔ IT Reconciliation  FY {fy}"
    ws["A1"].font=_fn(True,"FFFFFF",11); ws["A1"].fill=_f(DARK_BLUE)
    ws["A1"].alignment=_al("center"); ws["A1"].border=_bd(); ws.row_dimensions[1].height=26
    ri=2
    for ci,(h,v) in enumerate([("PAN",pan),("GSTINs"," / ".join(gstins)),
                                ("FY",fy),("IT Source",it_data.get("source","—"))],1):
        ws.cell(row=ri,column=ci*2-1,value=h).font=_fn(True)
        ws.cell(row=ri,column=ci*2,value=v)
    ri+=1
    ws.merge_cells(f"A{ri}:L{ri}")
    ws.cell(row=ri,column=1,value="ANNUAL SUMMARY").font=_fn(True,"FFFFFF",9)
    ws.cell(row=ri,column=1).fill=_f(MED_BLUE); ws.cell(row=ri,column=1).border=_bd(); ri+=1
    ann=[("GST GSTR-1 Turnover (B)",gst_data["annual_turnover"]),
         ("IT TIS Turnover (A)",it_data["tis_turnover"]),
         ("Difference (A−B)",round(it_data["tis_turnover"]-gst_data["annual_turnover"],2)),
         ("GST Purchases (GSTR-2B ITC)",gst_data["annual_purchase"]),
         ("IT AIS Purchases",it_data["ais_purchase"]),
         ("Purchase Difference",round(it_data["ais_purchase"]-gst_data["annual_purchase"],2)),
         ("TDS (26AS Total)",it_data["tds_total"]),
         ("Advance Tax",it_data["advance_tax"])]
    for label,val in ann:
        _c(ws,ri,1,label,bold=True); _c(ws,ri,2,val,align="right",numfmt=NUM_FMT)
        if "Diff" in label or "ifference" in label:
            ws.cell(row=ri,column=2).fill=_f(GREEN_BG if abs(val)<VARIANCE_THRESHOLD else RED_BG)
        ri+=1
    ri+=1
    ws.merge_cells(f"A{ri}:L{ri}")
    ws.cell(row=ri,column=1,value="MONTH-WISE COMPARISON").font=_fn(True,"FFFFFF",9)
    ws.cell(row=ri,column=1).fill=_f(MED_BLUE); ws.cell(row=ri,column=1).border=_bd(); ri+=1
    mh=[("Month",8),("GSTR-1 Taxable",16),("GSTR-1A Combined",16),("GSTR-3B Filed",16),
        ("AIS/TIS Turnover",16),("GST vs AIS Diff",16),("GST vs 3B Diff",16),("Status",10)]
    for ci,(h,w) in enumerate(mh,1):
        c=ws.cell(row=ri,column=ci,value=h)
        c.font=_fn(True,"FFFFFF",9); c.fill=_f(HDR_BG); c.alignment=_al("center"); c.border=_bd()
        ws.column_dimensions[get_column_letter(ci)].width=w
    ri+=1
    r1t=r1at=r3bt=aist=0.0
    for mon in FY_MONTHS:
        m=gst_data["monthly"].get(mon,{})
        r1=m.get("r1",0.0); r1a=m.get("r1a",0.0); r3b=m.get("r3b",0.0)
        ais=it_data["monthly_ais"].get(mon,0.0)
        d1=round(r1a-ais,2); d2=round(r1a-r3b,2)
        st="OK"; fbg=GREEN_BG
        if abs(d1)>VARIANCE_THRESHOLD or abs(d2)>VARIANCE_THRESHOLD: st="CHECK ⚠"; fbg=RED_BG
        elif abs(d1)>0 or abs(d2)>0: st="REVIEW"; fbg=YELLOW_BG
        rb=ALT1 if FY_MONTHS.index(mon)%2==0 else ALT2
        _c(ws,ri,1,mon,bg=rb,bold=True); _c(ws,ri,2,r1,bg=rb,align="right",numfmt=NUM_FMT)
        _c(ws,ri,3,r1a,bg=rb,align="right",numfmt=NUM_FMT); _c(ws,ri,4,r3b,bg=rb,align="right",numfmt=NUM_FMT)
        _c(ws,ri,5,ais,bg=rb,align="right",numfmt=NUM_FMT)
        _c(ws,ri,6,d1,bg=RED_BG if abs(d1)>VARIANCE_THRESHOLD else rb,align="right",numfmt=NUM_FMT)
        _c(ws,ri,7,d2,bg=RED_BG if abs(d2)>VARIANCE_THRESHOLD else rb,align="right",numfmt=NUM_FMT)
        _c(ws,ri,8,st,bg=fbg,bold=(st!="OK"),fg=RED_FG if "CHECK" in st else ("9C6500" if st=="REVIEW" else GREEN_FG))
        r1t+=r1; r1at+=r1a; r3bt+=r3b; aist+=ais; ri+=1
    tots=["TOTAL",r1t,r1at,r3bt,aist,round(r1at-aist,2),round(r1at-r3bt,2),
          "OK" if abs(r1at-aist)<VARIANCE_THRESHOLD else "CHECK ⚠"]
    for ci,v in enumerate(tots,1):
        c=ws.cell(row=ri,column=ci,value=v); c.font=_fn(True); c.fill=_f(TOT_BG); c.border=_bd()
        if isinstance(v,float): c.number_format=NUM_FMT; c.alignment=_al("right")
    ri+=2
    ws.merge_cells(f"A{ri}:L{ri}")
    ws.cell(row=ri,column=1,value="ITC vs PURCHASE RECONCILIATION").font=_fn(True,"FFFFFF",9)
    ws.cell(row=ri,column=1).fill=_f(MED_BLUE); ws.cell(row=ri,column=1).border=_bd(); ri+=1
    _c(ws,ri,1,"GSTR-2B ITC Claimed (GST)",bold=True)
    _c(ws,ri,2,gst_data.get("annual_purchase",0.0),align="right",numfmt=NUM_FMT); ri+=1
    _c(ws,ri,1,"AIS Purchases (IT Portal)",bold=True)
    _c(ws,ri,2,it_data.get("ais_purchase",0.0),align="right",numfmt=NUM_FMT); ri+=1
    diff_itc=round(gst_data.get("annual_purchase",0.0)-it_data.get("ais_purchase",0.0),2)
    _c(ws,ri,1,"Difference",bold=True)
    _c(ws,ri,2,diff_itc,align="right",numfmt=NUM_FMT,
       bg=GREEN_BG if abs(diff_itc)<VARIANCE_THRESHOLD else RED_BG)


# ═══════════════════════════════════════════════════════════════════════════════
# OFFLINE FOLDER PICKER
# ═══════════════════════════════════════════════════════════════════════════════
def _pick_existing_folder(base_path, label, expand_inner=False):
    """
    Show a numbered list of run folders.

    expand_inner=True  (GST): MultiYear_* folders are expanded to show
                               their AY* sub-folders — user picks the FY folder.
    expand_inner=False (IT):  AY* folders shown as-is — bridge needs the
                               AY* parent, NOT client children inside it.
    """
    base = Path(base_path)
    if not base.exists():
        print(f"  ✗ Base folder not found: {base}"); return None

    raw = sorted([d for d in base.iterdir() if d.is_dir()],
                 key=lambda d: d.stat().st_mtime, reverse=True)
    all_folders = []
    for f in raw:
        if expand_inner and f.name.startswith("MultiYear"):
            subs = sorted([s for s in f.iterdir() if s.is_dir()],
                          key=lambda s: s.stat().st_mtime, reverse=True)
            if subs: all_folders.extend(subs); continue
        all_folders.append(f)

    if not all_folders:
        print(f"  ✗ No folders found inside {base}"); return None

    print(f"\n  {label} — choose a folder:")
    for i, f in enumerate(all_folders, 1):
        mtime = datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
        try:    display = str(f.relative_to(base))
        except: display = f.name
        print(f"    {i:2d}.  {display}   [{mtime}]")
    try:    recent = str(all_folders[0].relative_to(base))
    except: recent = all_folders[0].name
    print(f"     0.  Use most recent ({recent})")

    while True:
        raw_in = input("  Enter number (or 0 for most recent): ").strip()
        if raw_in == "0": return all_folders[0]
        try:
            idx = int(raw_in) - 1
            if 0 <= idx < len(all_folders): return all_folders[idx]
        except ValueError: pass
        print("  Invalid choice. Try again.")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    global log, GST_RUN, IT_RUN, FY_LABEL, AY_LABEL, _fy_yr

    parser = argparse.ArgumentParser(description="Run ALL — GST + IT full pipeline v2.1")
    parser.add_argument("--only-gst",    action="store_true", help="GST steps 2-3 only")
    parser.add_argument("--only-it",     action="store_true", help="IT steps 4-5 only")
    parser.add_argument("--only-bridge", action="store_true", help="Bridge steps 6-6c only")
    parser.add_argument("--offline",     action="store_true",
                        help="Pick existing folders, run bridge (no portal downloads)")
    parser.add_argument("--gst-folder",  default=None, help="Explicit GST run folder")
    parser.add_argument("--it-folder",   default=None, help="Explicit IT run folder")
    parser.add_argument("--fy",          default=None, help="Override FY (e.g. 2024-25)")
    parser.add_argument("--client",      default=None, help="Process one client by name")
    args = parser.parse_args()

    if args.fy:
        FY_LABEL = args.fy
        _fy_yr   = int(FY_LABEL.split("-")[0])
        AY_LABEL = f"{_fy_yr+1}-{str(_fy_yr+2)[2:]}"
        GST_RUN  = GST_BASE / f"FY{FY_LABEL}_{RUN_TS}"
        IT_RUN   = IT_BASE  / f"AY{AY_LABEL}_{RUN_TS}"

    log = _setup_log()

    # ── OFFLINE / BRIDGE-ONLY ────────────────────────────────────────────────
    if args.offline or args.only_bridge:
        _banner(f"RUN ALL — OFFLINE / BRIDGE-ONLY  FY {FY_LABEL}")
        _log("  No portal downloads. Using existing files.")
        _log(f"  Script dir : {SCRIPT_DIR}")
        _log(f"  Log file   : {LOG_FILE}")

        if args.gst_folder:
            gst_offline = Path(args.gst_folder)
        else:
            print("\n  ─── SELECT GST FOLDER ───")
            print("  (MultiYear_* are expanded → pick the AY* FY sub-folder)")
            gst_offline = _pick_existing_folder(GST_BASE, "GST folder", expand_inner=True)
        if not gst_offline or not gst_offline.exists():
            _log(f"  ✗ GST folder not found: {gst_offline}", "error"); return

        if args.it_folder:
            it_offline = Path(args.it_folder)
        else:
            print("\n  ─── SELECT IT FOLDER ───")
            print("  (Select the AY* run folder — NOT a client subfolder inside it)")
            it_offline = _pick_existing_folder(IT_BASE, "IT folder", expand_inner=False)
        if not it_offline or not it_offline.exists():
            _log(f"  ✗ IT folder not found: {it_offline}", "error"); return

        # Guard: if user accidentally picked a client subfolder, go up one level
        _it_markers = ["IT_RECONCILIATION*.xlsx","26AS_*.pdf","AIS_*.pdf","TIS_*.pdf"]
        if any(list(it_offline.glob(pat)) for pat in _it_markers):
            _log(f"  ℹ  IT folder looks like client subfolder — using parent: {it_offline.parent}")
            it_offline = it_offline.parent

        # Update tracked globals
        GST_RUN = gst_offline
        IT_RUN  = it_offline
        _log(f"\n  GST folder : {gst_offline}")
        _log(f"  IT  folder : {it_offline}")

        clients = load_clients(fy_override=args.fy, name_filter=args.client)
        run_bridge_step(clients, str(gst_offline), str(it_offline))
        run_gst_it_comparison_step()
        run_gstr2b_extractor_step(clients)
        _banner("OFFLINE — ALL DONE")
        _log(f"  Log: {LOG_FILE}")
        return

    # ── NORMAL (ONLINE) MODE ─────────────────────────────────────────────────
    _banner(f"RUN ALL v2.1 — GST + Income Tax Pipeline  FY {FY_LABEL}")
    _log(f"  Script dir : {SCRIPT_DIR}")
    _log(f"  Base dir   : {BASE_DIR}")
    _log(f"  Log file   : {LOG_FILE}")
    _log(f"  Variance   : ₹{VARIANCE_THRESHOLD:,.0f}")
    _log("")
    _log("  ORDER:  Step 2+3 GST  →  Step 4+5 IT  →  Step 6 Bridge  →  6b Comparison  →  6c 2B Extract")

    clients = load_clients(fy_override=args.fy, name_filter=args.client)
    _log(f"\n  {len(clients)} client(s) loaded\n")

    run_all_flag = not (args.only_gst or args.only_it or args.only_bridge)
    do_gst    = run_all_flag or args.only_gst
    do_it     = run_all_flag or args.only_it
    do_bridge = run_all_flag or args.only_bridge

    if do_gst:
        run_gst_step(clients)        # GST_RUN updated to real folder

    if do_it:
        run_it_step(clients)         # IT_RUN updated to real folder

    if do_bridge:
        run_bridge_step(clients, str(GST_RUN), str(IT_RUN))
        run_gst_it_comparison_step()
        run_gstr2b_extractor_step(clients)

    _banner("ALL DONE")
    _log(f"  GST output : {GST_RUN}")
    _log(f"  IT  output : {IT_RUN}")
    _log(f"  Log        : {LOG_FILE}")


if __name__ == "__main__":
    main()
