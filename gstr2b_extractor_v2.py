"""
GSTR-2B Consolidated Extractor
Reads all individual GSTR-2B Excel files from a folder and produces a
multi-sheet consolidated analysis workbook.

Usage:
    python gstr2b_extractor_fixed.py --input <folder> --output <output.xlsx>

Defaults:
    --input  : current directory (.)
    --output : GSTR2B_Consolidated_Analysis.xlsx
"""

import argparse
import glob
import os
import sys
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Safe numeric helper
# ---------------------------------------------------------------------------

def _n(val):
    """Safely convert any cell value to float. Handles None, '', strings, ints, floats."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MONTH_ORDER = [
    "April", "May", "June", "July", "August", "September",
    "October", "November", "December", "January", "February", "March",
]

# Colour palette
CLR_HEADER_DARK  = "1F3864"   # dark blue  – sheet title row
CLR_HEADER_MID   = "2E75B6"   # mid blue   – column headers
CLR_SUBHDR       = "D6E4F0"   # light blue – sub-header rows
CLR_GRAND_TOTAL  = "FFF2CC"   # yellow     – grand-total row
CLR_ALT_ROW      = "EBF3FB"   # very light blue – alternating rows
CLR_WHITE        = "FFFFFF"
CLR_FONT_LIGHT   = "FFFFFF"
CLR_FONT_DARK    = "000000"

# ---------------------------------------------------------------------------
# Helpers – styling
# ---------------------------------------------------------------------------

def _font(bold=False, size=11, colour=CLR_FONT_DARK, name="Arial"):
    return Font(bold=bold, size=size, color=colour, name=name)

def _fill(hex_colour):
    return PatternFill("solid", start_color=hex_colour, fgColor=hex_colour)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _fmt_inr(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = '#,##0.00'

def _style_header_row(ws, row_num, n_cols, dark=True):
    bg = CLR_HEADER_DARK if dark else CLR_HEADER_MID
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill  = _fill(bg)
        c.font  = _font(bold=True, colour=CLR_FONT_LIGHT)
        c.alignment = _align(h="center", wrap=True)
        c.border = _border_thin()

def _style_data_row(ws, row_num, n_cols, alt=False):
    bg = CLR_ALT_ROW if alt else CLR_WHITE
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill   = _fill(bg)
        c.font   = _font()
        c.border = _border_thin()
        if col > 3:
            c.alignment = _align(h="right")

def _style_total_row(ws, row_num, n_cols):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill   = _fill(CLR_GRAND_TOTAL)
        c.font   = _font(bold=True)
        c.border = _border_thin()
        if col > 3:
            c.alignment = _align(h="right")

def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        length = max(
            (len(str(cell.value)) if cell.value else 0) for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = (
            min(max(length + 2, min_w), max_w)
        )

# ---------------------------------------------------------------------------
# Data extraction
# ---------------------------------------------------------------------------

def _get_metadata(wb):
    # Some GSTR-2B exports omit the "Read me" sheet — handle gracefully
    sheet_name = None
    for candidate in ["Read me", "Read Me", "README", "Readme", "read me"]:
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        # Fall back: try to infer month/FY from the B2B sheet header rows
        meta = {}
        try:
            ws = wb["B2B"]
            for r in ws.iter_rows(max_row=6, values_only=True):
                for cell in r:
                    if cell and "Tax Period" in str(cell):
                        # value is usually in next column — handled below
                        pass
                    if r[0] and ":" in str(r[0]):
                        parts = str(r[0]).split(":")
                        if len(parts) == 2:
                            meta[parts[0].strip()] = parts[1].strip()
        except Exception:
            pass
        return meta
    rm = wb[sheet_name]
    meta = {}
    for r in rm.iter_rows(max_row=12, values_only=True):
        if r[0]:
            meta[str(r[0]).strip()] = r[2] if len(r) > 2 else None
    return meta


def _iter_data_rows(ws, skip_rows):
    """Yield rows that have actual supplier data (col-0 looks like a GSTIN: 15 alphanumeric chars)."""
    import re
    gstin_re = re.compile(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$')
    all_rows = list(ws.iter_rows(values_only=True))
    for r in all_rows[skip_rows:]:
        val = str(r[0]).strip() if r[0] is not None else ""
        if gstin_re.match(val):
            yield r


def _detect_b2b_format(wb):
    """
    Detect whether the B2B sheet uses old format (with Rate% col) or new format (without Rate%).
    Old format (pre-2024): col8=Rate(%), col9=TaxableVal, col10=IGST ...
    New format (2024+):    col8=TaxableVal (Rate% removed), col9=IGST ...
    Returns 'old' or 'new'.
    """
    ws = wb["B2B"]
    for row in ws.iter_rows(min_row=5, max_row=6, values_only=True):
        for j, v in enumerate(row):
            if v and "Rate" in str(v) and j == 8:
                return "old"
            if v and "Taxable" in str(v) and j == 8:
                return "new"
    return "new"  # default to new if ambiguous


def _extract_b2b(wb, month, fy):
    """
    B2B sheet – auto-detects old vs new GST format.

    Old format (pre-2024, col8=Rate%):
      0 GSTIN  1 Name  2 InvNo  3 InvType  4 InvDate  5 InvValue
      6 POS  7 RCM  8 Rate%  9 TaxableVal
      10 IGST  11 CGST  12 SGST  13 Cess
      14 FilingPeriod  15 FilingDate  16 ITC  17 Reason

    New format (2024+, Rate% removed):
      0 GSTIN  1 Name  2 InvNo  3 InvType  4 InvDate  5 InvValue
      6 POS  7 RCM  8 TaxableVal
      9 IGST  10 CGST  11 SGST  12 Cess
      13 FilingPeriod  14 FilingDate  15 ITC  16 Reason
    """
    fmt = _detect_b2b_format(wb)
    rows = []
    for r in _iter_data_rows(wb["B2B"], skip_rows=6):
        if fmt == "old":
            rate = _n(r[8])
            taxable = _n(r[9])
            igst, cgst, sgst, cess = _n(r[10]), _n(r[11]), _n(r[12]), _n(r[13])
            filing_period = r[14] if len(r) > 14 else ""
            filing_date   = r[15] if len(r) > 15 else ""
            itc           = r[16] if len(r) > 16 else ""
            reason        = r[17] if len(r) > 17 else ""
        else:
            rate = 0.0
            taxable = _n(r[8])
            igst, cgst, sgst, cess = _n(r[9]), _n(r[10]), _n(r[11]), _n(r[12])
            filing_period = r[13] if len(r) > 13 else ""
            filing_date   = r[14] if len(r) > 14 else ""
            itc           = r[15] if len(r) > 15 else ""
            reason        = r[16] if len(r) > 16 else ""
        rows.append({
            "Month": month, "FY": fy, "Doc Type": "B2B Invoice",
            "Supplier GSTIN": r[0], "Supplier Name": r[1],
            "Doc Number": r[2], "Doc Sub-Type": r[3],
            "Doc Date": r[4], "Doc Value": _n(r[5]),
            "Place of Supply": r[6], "RCM": r[7], "Rate%": rate,
            "Taxable Value": taxable,
            "IGST": igst, "CGST": cgst, "SGST": sgst, "Cess": cess,
            "Total Tax": igst + cgst + sgst + cess,
            "Filing Period": filing_period,
            "Filing Date":   filing_date,
            "ITC Availability": itc,
            "Reason": reason,
        })
    return rows


def _extract_b2ba(wb, month, fy):
    """
    B2BA (amended invoices). Auto-detects old vs new GST format.

    Old format (pre-2024, col10=Rate%):
      0 OrigInvNo  1 OrigInvDate  2 GSTIN  3 Name
      4 RevInvNo  5 InvType  6 InvDate  7 InvValue  8 POS  9 RCM  10 Rate%
      11 TaxableVal  12 IGST  13 CGST  14 SGST  15 Cess
      16 Period  17 Date  18 ITC

    New format (2024+, Rate% removed):
      0 OrigInvNo  1 OrigInvDate  2 GSTIN  3 Name
      4 RevInvNo  5 InvType  6 InvDate  7 InvValue  8 POS  9 RCM
      10 TaxableVal  11 IGST  12 CGST  13 SGST  14 Cess
      (15-19 ITC reduction cols)  21 Period  22 Date  23 ITC
    """
    # Detect format via header row
    ws = wb["B2BA"]
    fmt = "new"
    for row in ws.iter_rows(min_row=6, max_row=7, values_only=True):
        for j, v in enumerate(row):
            if v and "Rate" in str(v) and j == 10:
                fmt = "old"
                break

    rows = []
    for r in _iter_data_rows(ws, skip_rows=8):
        if fmt == "old":
            rate = _n(r[10])
            taxable = _n(r[11])
            igst, cgst, sgst, cess = _n(r[12]), _n(r[13]), _n(r[14]), _n(r[15])
            filing_period = r[16] if len(r) > 16 else ""
            filing_date   = r[17] if len(r) > 17 else ""
            itc           = r[18] if len(r) > 18 else ""
        else:
            rate = 0.0
            taxable = _n(r[10])
            igst, cgst, sgst, cess = _n(r[11]), _n(r[12]), _n(r[13]), _n(r[14])
            filing_period = r[21] if len(r) > 21 else ""
            filing_date   = r[22] if len(r) > 22 else ""
            itc           = r[23] if len(r) > 23 else ""
        rows.append({
            "Month": month, "FY": fy, "Doc Type": "B2BA Amended Invoice",
            "Supplier GSTIN": r[2], "Supplier Name": r[3],
            "Doc Number": r[4], "Doc Sub-Type": r[5],
            "Doc Date": r[6], "Doc Value": _n(r[7]),
            "Place of Supply": r[8], "RCM": r[9], "Rate%": rate,
            "Taxable Value": taxable,
            "IGST": igst, "CGST": cgst, "SGST": sgst, "Cess": cess,
            "Total Tax": igst + cgst + sgst + cess,
            "Filing Period": filing_period,
            "Filing Date":   filing_date,
            "ITC Availability": itc,
            "Reason": "",
        })
    return rows


def _extract_cdnr(wb, month, fy):
    """
    B2B-CDNR – auto-detects old vs new GST format.

    Old format (pre-2024, col9=Rate%):
      0 GSTIN  1 Name  2 NoteNo  3 NoteType  4 NoteSupplyType  5 NoteDate
      6 NoteValue  7 POS  8 RCM  9 Rate%  10 TaxableVal
      11 IGST  12 CGST  13 SGST  14 Cess
      15 Period  16 Date  17 ITC  18 Reason

    New format (2024+, Rate% removed; extra ITC reduction cols):
      0 GSTIN  1 Name  2 NoteNo  3 NoteType  4 NoteSupplyType  5 NoteDate
      6 NoteValue  7 POS  8 RCM  9 TaxableVal
      10 IGST  11 CGST  12 SGST  13 Cess
      (14-19 ITC reduction/remarks cols)
      20 Period  21 Date  22 ITC  23 Reason
    """
    ws = wb["B2B-CDNR"]
    fmt = "new"
    for row in ws.iter_rows(min_row=5, max_row=6, values_only=True):
        for j, v in enumerate(row):
            if v and "Rate" in str(v) and j == 9:
                fmt = "old"
                break

    rows = []
    for r in _iter_data_rows(ws, skip_rows=6):
        if fmt == "old":
            rate = _n(r[9])
            taxable = _n(r[10])
            igst, cgst, sgst, cess = _n(r[11]), _n(r[12]), _n(r[13]), _n(r[14])
            filing_period = r[15] if len(r) > 15 else ""
            filing_date   = r[16] if len(r) > 16 else ""
            itc           = r[17] if len(r) > 17 else ""
            reason        = r[18] if len(r) > 18 else ""
        else:
            rate = 0.0
            taxable = _n(r[9])
            igst, cgst, sgst, cess = _n(r[10]), _n(r[11]), _n(r[12]), _n(r[13])
            filing_period = r[20] if len(r) > 20 else ""
            filing_date   = r[21] if len(r) > 21 else ""
            itc           = r[22] if len(r) > 22 else ""
            reason        = r[23] if len(r) > 23 else ""
        rows.append({
            "Month": month, "FY": fy, "Doc Type": "CDN (Credit/Debit Note)",
            "Supplier GSTIN": r[0], "Supplier Name": r[1],
            "Doc Number": r[2], "Doc Sub-Type": r[3],
            "Doc Date": r[5], "Doc Value": _n(r[6]),
            "Place of Supply": r[7], "RCM": r[8], "Rate%": rate,
            "Taxable Value": taxable,
            "IGST": igst, "CGST": cgst, "SGST": sgst, "Cess": cess,
            "Total Tax": igst + cgst + sgst + cess,
            "Filing Period": filing_period,
            "Filing Date":   filing_date,
            "ITC Availability": itc,
            "Reason": reason,
        })
    return rows


def _extract_cdnra(wb, month, fy):
    """
    B2B-CDNRA (amended CDN) – auto-detects old vs new GST format.

    Old format (pre-2024, col12=Rate%):
      0 origType  1 origNo  2 origDate  3 GSTIN  4 Name
      5 NoteNo  6 NoteType  7 NoteSupplyType  8 NoteDate  9 NoteValue
      10 POS  11 RCM  12 Rate%  13 TaxableVal
      14 IGST  15 CGST  16 SGST  17 Cess
      18 Period  19 Date  20 ITC

    New format (2024+, Rate% removed; extra ITC reduction cols):
      0 origType  1 origNo  2 origDate  3 GSTIN  4 Name
      5 NoteNo  6 NoteType  7 NoteSupplyType  8 NoteDate  9 NoteValue
      10 POS  11 RCM  12 TaxableVal
      13 IGST  14 CGST  15 SGST  16 Cess
      (17-22 ITC reduction/remarks cols)
      23 Period  24 Date  25 ITC
    """
    ws = wb["B2B-CDNRA"]
    fmt = "new"
    for row in ws.iter_rows(min_row=6, max_row=9, values_only=True):
        for j, v in enumerate(row):
            if v and "Rate" in str(v) and j == 12:
                fmt = "old"
                break

    rows = []
    for r in _iter_data_rows(ws, skip_rows=8):
        if fmt == "old":
            rate = _n(r[12])
            taxable = _n(r[13])
            igst, cgst, sgst, cess = _n(r[14]), _n(r[15]), _n(r[16]), _n(r[17])
            filing_period = r[18] if len(r) > 18 else ""
            filing_date   = r[19] if len(r) > 19 else ""
            itc           = r[20] if len(r) > 20 else ""
        else:
            rate = 0.0
            taxable = _n(r[12])
            igst, cgst, sgst, cess = _n(r[13]), _n(r[14]), _n(r[15]), _n(r[16])
            filing_period = r[23] if len(r) > 23 else ""
            filing_date   = r[24] if len(r) > 24 else ""
            itc           = r[25] if len(r) > 25 else ""
        rows.append({
            "Month": month, "FY": fy, "Doc Type": "CDNRA Amended CDN",
            "Supplier GSTIN": r[3], "Supplier Name": r[4],
            "Doc Number": r[5], "Doc Sub-Type": r[6],
            "Doc Date": r[8], "Doc Value": _n(r[9]),
            "Place of Supply": r[10], "RCM": r[11], "Rate%": rate,
            "Taxable Value": taxable,
            "IGST": igst, "CGST": cgst, "SGST": sgst, "Cess": cess,
            "Total Tax": igst + cgst + sgst + cess,
            "Filing Period": filing_period,
            "Filing Date":   filing_date,
            "ITC Availability": itc,
            "Reason": "",
        })
    return rows


def load_all_files(folder):
    pattern = os.path.join(folder, "GSTR2B_*.xlsx")
    files = glob.glob(pattern)
    if not files:
        # Also try with timestamp prefix (from claude.ai uploads)
        pattern = os.path.join(folder, "*GSTR2B_*.xlsx")
        files = glob.glob(pattern)

    all_records = []
    found_months = set()

    for f in files:
        try:
            wb = load_workbook(f, read_only=True)
        except Exception as e:
            print(f"  WARN: Cannot open {os.path.basename(f)}: {e}", file=sys.stderr)
            continue

        # Skip output/consolidated files that don't have the standard sheets
        if "B2B" not in wb.sheetnames:
            print(f"  SKIP: {os.path.basename(f)} — no B2B sheet (not a raw GSTR-2B file)")
            continue

        meta  = _get_metadata(wb)
        month = meta.get("Tax Period", "Unknown")
        fy    = meta.get("Financial Year", "Unknown")
        gstin = meta.get("GSTIN", "")

        # Dedup: skip if we already loaded this month (file found in 2 locations)
        month_key = f"{month}_{fy}"
        if month_key != "Unknown_Unknown" and month_key in found_months:
            print(f"  SKIP (duplicate month): {os.path.basename(f)} — {month} {fy} already loaded")
            continue
        if month_key != "Unknown_Unknown":
            found_months.add(month_key)

        print(f"  Reading {month} {fy}  ({os.path.basename(f)})")

        records = []
        records += _extract_b2b(wb, month, fy)
        records += _extract_b2ba(wb, month, fy)
        records += _extract_cdnr(wb, month, fy)
        records += _extract_cdnra(wb, month, fy)

        all_records.extend(records)
        found_months.add(month)
        wb.close()

    # Warn about missing months
    for m in MONTH_ORDER:
        if m not in found_months:
            print(f"  WARN: No file found for month '{m}' – row will be absent in Month-wise Summary",
                  file=sys.stderr)

    return all_records


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

ALLDATA_COLS = [
    "Month", "FY", "Doc Type",
    "Supplier GSTIN", "Supplier Name",
    "Doc Number", "Doc Sub-Type", "Doc Date", "Doc Value",
    "Place of Supply", "RCM", "Rate%", "Taxable Value",
    "IGST", "CGST", "SGST", "Cess", "Total Tax",
    "Filing Period", "Filing Date",
    "ITC Availability", "Reason",
]

NUM_COLS = {"Doc Value", "Taxable Value", "IGST", "CGST", "SGST", "Cess", "Total Tax"}


def build_all_data(wb, records):
    ws = wb.create_sheet("All Data")
    n = len(ALLDATA_COLS)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.cell(1, 1).value = "GSTR-2B – All Document Details"
    ws.cell(1, 1).font  = _font(bold=True, size=13, colour=CLR_FONT_LIGHT)
    ws.cell(1, 1).fill  = _fill(CLR_HEADER_DARK)
    ws.cell(1, 1).alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    # Headers
    for ci, col in enumerate(ALLDATA_COLS, 1):
        ws.cell(2, ci).value = col
    _style_header_row(ws, 2, n, dark=False)
    ws.row_dimensions[2].height = 30

    # Sort records by month order then supplier
    def sort_key(r):
        mi = MONTH_ORDER.index(r["Month"]) if r["Month"] in MONTH_ORDER else 99
        return (mi, r["Supplier Name"] or "", r["Doc Date"] or "")

    records_sorted = sorted(records, key=sort_key)

    for ri, rec in enumerate(records_sorted, 3):
        for ci, col in enumerate(ALLDATA_COLS, 1):
            cell = ws.cell(ri, ci)
            cell.value = rec.get(col, "")
            if col in NUM_COLS:
                cell.number_format = "#,##0.00"
        _style_data_row(ws, ri, n, alt=(ri % 2 == 0))

    # Freeze panes
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}2"

    # Column widths
    widths = {
        1: 12, 2: 10, 3: 22, 4: 18, 5: 28,
        6: 18, 7: 14, 8: 12, 9: 14, 10: 16,
        11: 6, 12: 7, 13: 14, 14: 14, 15: 14,
        16: 14, 17: 10, 18: 14, 19: 12, 20: 14,
        21: 14, 22: 20,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    return ws


def build_monthwise_summary(wb, records):
    ws = wb.create_sheet("Month-wise Summary")

    headers = [
        "Month", "Financial Year", "No. of Documents",
        "Taxable Value (₹)", "IGST (₹)", "CGST (₹)", "SGST (₹)",
        "Cess (₹)", "Total Tax (₹)", "Invoices", "Amount (₹)",
    ]
    n = len(headers)

    # Aggregate by (FY, Month) so multi-FY data stays separate
    agg = defaultdict(lambda: {
        "docs": 0, "taxable": 0.0, "igst": 0.0, "cgst": 0.0,
        "sgst": 0.0, "cess": 0.0, "inv": 0, "inv_amt": 0.0
    })
    all_fys = []
    for r in records:
        key = (r["FY"], r["Month"])
        agg[key]["docs"]    += 1
        agg[key]["taxable"] += r["Taxable Value"]
        agg[key]["igst"]    += r["IGST"]
        agg[key]["cgst"]    += r["CGST"]
        agg[key]["sgst"]    += r["SGST"]
        agg[key]["cess"]    += r["Cess"]
        if r["Doc Type"] in ("B2B Invoice", "B2BA Amended Invoice"):
            agg[key]["inv"]     += 1
            agg[key]["inv_amt"] += r["Doc Value"]
        if r["FY"] not in all_fys:
            all_fys.append(r["FY"])

    # Sort FYs chronologically
    all_fys.sort()

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.cell(1, 1).value = "Month-wise GSTR-2B Summary (All Document Types)"
    ws.cell(1, 1).font  = _font(bold=True, size=13, colour=CLR_FONT_LIGHT)
    ws.cell(1, 1).fill  = _fill(CLR_HEADER_DARK)
    ws.cell(1, 1).alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(headers, 1):
        ws.cell(2, ci).value = h
    _style_header_row(ws, 2, n, dark=False)
    ws.row_dimensions[2].height = 30

    ri = 3
    data_start = 3
    for fy in all_fys:
        for month in MONTH_ORDER:
            key = (fy, month)
            if key not in agg:
                continue
            a = agg[key]
            row_vals = [
                month, fy,
                a["docs"], a["taxable"], a["igst"], a["cgst"],
                a["sgst"], a["cess"], a["igst"] + a["cgst"] + a["sgst"] + a["cess"],
                a["inv"], a["inv_amt"],
            ]
            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(ri, ci)
                c.value = v
                if ci > 3:
                    c.number_format = "#,##0.00"
            _style_data_row(ws, ri, n, alt=(ri % 2 == 0))
            ri += 1

    # Grand total row
    data_end = ri - 1
    ws.cell(ri, 1).value = "GRAND TOTAL"
    ws.cell(ri, 2).value = ""
    sum_cols = {3: "C", 4: "D", 5: "E", 6: "F", 7: "G", 8: "H", 9: "I", 10: "J", 11: "K"}
    for ci, ltr in sum_cols.items():
        ws.cell(ri, ci).value = f"=SUM({ltr}{data_start}:{ltr}{data_end})"
        ws.cell(ri, ci).number_format = "#,##0.00"
    _style_total_row(ws, ri, n)

    ws.freeze_panes = "A3"

    widths = [20, 14, 18, 18, 16, 14, 14, 12, 16, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


def build_supplier_monthly(wb, records):
    ws = wb.create_sheet("Supplier Monthly")

    headers = [
        "Supplier GSTIN", "Supplier Name", "Month", "FY",
        "No. of Docs", "Taxable Value (₹)", "IGST (₹)",
        "CGST (₹)", "SGST (₹)", "Cess (₹)", "Total Tax (₹)",
    ]
    n = len(headers)

    agg = defaultdict(lambda: {
        "name": "", "docs": 0, "taxable": 0,
        "igst": 0, "cgst": 0, "sgst": 0, "cess": 0, "fy": ""
    })
    for r in records:
        key = (r["Supplier GSTIN"], r["Month"])
        d = agg[key]
        d["name"]    = r["Supplier Name"]
        d["fy"]      = r["FY"]
        d["docs"]    += 1
        d["taxable"] += r["Taxable Value"]
        d["igst"]    += r["IGST"]
        d["cgst"]    += r["CGST"]
        d["sgst"]    += r["SGST"]
        d["cess"]    += r["Cess"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.cell(1, 1).value = "Supplier-wise Monthly Summary"
    ws.cell(1, 1).font  = _font(bold=True, size=13, colour=CLR_FONT_LIGHT)
    ws.cell(1, 1).fill  = _fill(CLR_HEADER_DARK)
    ws.cell(1, 1).alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(headers, 1):
        ws.cell(2, ci).value = h
    _style_header_row(ws, 2, n, dark=False)
    ws.row_dimensions[2].height = 30

    sorted_keys = sorted(
        agg.keys(),
        key=lambda k: (k[0], MONTH_ORDER.index(k[1]) if k[1] in MONTH_ORDER else 99)
    )

    for ri, (gstin, month) in enumerate(sorted_keys, 3):
        d = agg[(gstin, month)]
        total_tax = d["igst"] + d["cgst"] + d["sgst"] + d["cess"]
        row_vals = [
            gstin, d["name"], month, d["fy"],
            d["docs"], d["taxable"], d["igst"],
            d["cgst"], d["sgst"], d["cess"], total_tax,
        ]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(ri, ci)
            c.value = v
            if ci > 5:
                c.number_format = "#,##0.00"
        _style_data_row(ws, ri, n, alt=(ri % 2 == 0))

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}2"

    widths = [18, 28, 12, 10, 10, 16, 14, 14, 14, 10, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


def build_supplier_detail(wb, records):
    ws = wb.create_sheet("Supplier Detail")

    headers = [
        "Supplier GSTIN", "Supplier Name", "Month", "FY",
        "Doc Type", "Doc Number", "Doc Sub-Type", "Doc Date",
        "Doc Value (₹)", "Taxable Value (₹)", "IGST (₹)",
        "CGST (₹)", "SGST (₹)", "Cess (₹)", "Total Tax (₹)",
        "ITC Availability", "Reason",
    ]
    n = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.cell(1, 1).value = "Supplier-wise Detailed Invoice Listing"
    ws.cell(1, 1).font  = _font(bold=True, size=13, colour=CLR_FONT_LIGHT)
    ws.cell(1, 1).fill  = _fill(CLR_HEADER_DARK)
    ws.cell(1, 1).alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(headers, 1):
        ws.cell(2, ci).value = h
    _style_header_row(ws, 2, n, dark=False)
    ws.row_dimensions[2].height = 30

    sorted_recs = sorted(
        records,
        key=lambda r: (
            r["Supplier GSTIN"] or "",
            MONTH_ORDER.index(r["Month"]) if r["Month"] in MONTH_ORDER else 99,
            r["Doc Date"] or "",
        )
    )

    for ri, r in enumerate(sorted_recs, 3):
        total_tax = r["IGST"] + r["CGST"] + r["SGST"] + r["Cess"]
        row_vals = [
            r["Supplier GSTIN"], r["Supplier Name"], r["Month"], r["FY"],
            r["Doc Type"], r["Doc Number"], r["Doc Sub-Type"], r["Doc Date"],
            r["Doc Value"], r["Taxable Value"], r["IGST"],
            r["CGST"], r["SGST"], r["Cess"], total_tax,
            r["ITC Availability"], r["Reason"],
        ]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(ri, ci)
            c.value = v
            if ci in (9, 10, 11, 12, 13, 14, 15):
                c.number_format = "#,##0.00"
        _style_data_row(ws, ri, n, alt=(ri % 2 == 0))

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}2"

    widths = [18, 28, 12, 10, 22, 18, 14, 12, 14, 14, 14, 14, 14, 10, 14, 16, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


def build_gstin_annual(wb, records):
    ws = wb.create_sheet("GSTIN Annual Summary")

    headers = [
        "Supplier GSTIN", "Supplier Name",
        "Total Docs", "Taxable Value (₹)",
        "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)", "Total Tax (₹)",
    ]
    n = len(headers)

    agg = defaultdict(lambda: {
        "name": "", "docs": 0, "taxable": 0,
        "igst": 0, "cgst": 0, "sgst": 0, "cess": 0
    })
    for r in records:
        g = r["Supplier GSTIN"]
        agg[g]["name"]    = r["Supplier Name"]
        agg[g]["docs"]    += 1
        agg[g]["taxable"] += r["Taxable Value"]
        agg[g]["igst"]    += r["IGST"]
        agg[g]["cgst"]    += r["CGST"]
        agg[g]["sgst"]    += r["SGST"]
        agg[g]["cess"]    += r["Cess"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.cell(1, 1).value = "GSTIN-wise Annual Summary"
    ws.cell(1, 1).font  = _font(bold=True, size=13, colour=CLR_FONT_LIGHT)
    ws.cell(1, 1).fill  = _fill(CLR_HEADER_DARK)
    ws.cell(1, 1).alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(headers, 1):
        ws.cell(2, ci).value = h
    _style_header_row(ws, 2, n, dark=False)
    ws.row_dimensions[2].height = 30

    sorted_gstins = sorted(agg.keys(), key=lambda g: agg[g]["taxable"], reverse=True)

    for ri, gstin in enumerate(sorted_gstins, 3):
        d = agg[gstin]
        total_tax = d["igst"] + d["cgst"] + d["sgst"] + d["cess"]
        row_vals = [
            gstin, d["name"], d["docs"], d["taxable"],
            d["igst"], d["cgst"], d["sgst"], d["cess"], total_tax,
        ]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(ri, ci)
            c.value = v
            if ci > 3:
                c.number_format = "#,##0.00"
        _style_data_row(ws, ri, n, alt=(ri % 2 == 0))

    # Grand total
    tr = len(sorted_gstins) + 3
    ws.cell(tr, 1).value = "GRAND TOTAL"
    for ci, ltr in {3: "C", 4: "D", 5: "E", 6: "F", 7: "G", 8: "H", 9: "I"}.items():
        ws.cell(tr, ci).value = f"=SUM({ltr}3:{ltr}{tr - 1})"
        ws.cell(tr, ci).number_format = "#,##0.00"
    _style_total_row(ws, tr, n)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}2"

    widths = [18, 30, 12, 18, 16, 14, 14, 10, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


def build_dashboard(wb, records):
    ws = wb.create_sheet("Dashboard", 0)

    total_docs    = len(records)
    total_taxable = sum(r["Taxable Value"] for r in records)
    total_igst    = sum(r["IGST"] for r in records)
    total_cgst    = sum(r["CGST"] for r in records)
    total_sgst    = sum(r["SGST"] for r in records)
    total_cess    = sum(r["Cess"] for r in records)
    total_tax     = total_igst + total_cgst + total_sgst + total_cess

    fys = sorted({r["FY"] for r in records})
    fy  = ", ".join(fys) if fys else ""

    # Title
    ws.merge_cells("A1:F1")
    ws.cell(1, 1).value = f"GSTR-2B Consolidated Dashboard – FY {fy}"
    ws.cell(1, 1).font  = _font(bold=True, size=14, colour=CLR_FONT_LIGHT)
    ws.cell(1, 1).fill  = _fill(CLR_HEADER_DARK)
    ws.cell(1, 1).alignment = _align(h="center")
    ws.row_dimensions[1].height = 28

    # KPI block
    kpis = [
        ("Total Documents",     total_docs,      False),
        ("Total Taxable Value", total_taxable,   True),
        ("Total IGST",          total_igst,      True),
        ("Total CGST",          total_cgst,      True),
        ("Total SGST",          total_sgst,      True),
        ("Total Cess",          total_cess,      True),
        ("Total Tax",           total_tax,       True),
        ("Files Loaded",        len({(r["FY"], r["Month"]) for r in records}), False),
    ]

    ws.cell(3, 1).value = "Summary KPIs"
    ws.cell(3, 1).font  = _font(bold=True, size=12, colour=CLR_FONT_LIGHT)
    ws.cell(3, 1).fill  = _fill(CLR_HEADER_MID)
    ws.merge_cells("A3:B3")
    ws.cell(3, 1).alignment = _align(h="center")

    for i, (label, value, is_currency) in enumerate(kpis, 4):
        cl = ws.cell(i, 1)
        cv = ws.cell(i, 2)
        cl.value = label
        cv.value = value
        cl.font  = _font(bold=True)
        cl.fill  = _fill(CLR_SUBHDR)
        cv.fill  = _fill(CLR_WHITE)
        cl.border = cv.border = _border_thin()
        cv.alignment = _align(h="right")
        if is_currency:
            cv.number_format = "₹#,##0.00"

    # For single-FY, show missing months; for multi-FY just show count
    if len(fys) == 1:
        missing = [m for m in MONTH_ORDER if m not in {r["Month"] for r in records}]
        missing_label = f"⚠  Missing months (no files found): {', '.join(missing)}" if missing else "✓  All 12 months present"
        missing_colour = "FFF2CC" if missing else "E2EFDA"
        missing_font_clr = "FF0000" if missing else "375623"
    else:
        missing_label = f"✓  Multi-FY mode: {len(fys)} financial years loaded ({', '.join(fys)})"
        missing_colour = "E2EFDA"
        missing_font_clr = "375623"

    row_notice = 4 + len(kpis) + 1
    ws.merge_cells(f"A{row_notice}:F{row_notice}")
    notice_cell = ws.cell(row_notice, 1)
    notice_cell.value = missing_label
    notice_cell.font  = Font(bold=True, color=missing_font_clr, name="Arial", size=11)
    notice_cell.fill  = _fill(missing_colour)
    notice_cell.alignment = _align(h="left")

    # Month coverage table
    row_table = row_notice + 2
    ws.cell(row_table, 1).value = "Month"
    ws.cell(row_table, 2).value = "Docs"
    ws.cell(row_table, 3).value = "Taxable Value (₹)"
    ws.cell(row_table, 4).value = "Total Tax (₹)"
    _style_header_row(ws, row_table, 4, dark=False)
    ws.row_dimensions[row_table].height = 22

    month_agg = defaultdict(lambda: {"docs": 0, "taxable": 0, "tax": 0})
    for r in records:
        month_agg[r["Month"]]["docs"]    += 1
        month_agg[r["Month"]]["taxable"] += r["Taxable Value"]
        month_agg[r["Month"]]["tax"]     += r["Total Tax"]

    for ri2, m in enumerate(MONTH_ORDER):
        rw = row_table + 1 + ri2
        d  = month_agg.get(m, {})
        ws.cell(rw, 1).value = m
        ws.cell(rw, 2).value = d.get("docs", "—")
        ws.cell(rw, 3).value = d.get("taxable", "—")
        ws.cell(rw, 4).value = d.get("tax", "—")
        for ci in range(1, 5):
            c = ws.cell(rw, ci)
            c.border = _border_thin()
            c.alignment = _align(h="right" if ci > 1 else "left")
            if ci in (3, 4) and d:
                c.number_format = "#,##0.00"
        _style_data_row(ws, rw, 4, alt=(ri2 % 2 == 0))
        if m not in month_agg:
            ws.cell(rw, 1).font = Font(color="FF0000", name="Arial", italic=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_workbook(records, output_path):
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    print("  Building Dashboard …")
    build_dashboard(wb, records)

    print("  Building All Data …")
    build_all_data(wb, records)

    print("  Building Month-wise Summary …")
    build_monthwise_summary(wb, records)

    print("  Building Supplier Monthly …")
    build_supplier_monthly(wb, records)

    print("  Building Supplier Detail …")
    build_supplier_detail(wb, records)

    print("  Building GSTIN Annual Summary …")
    build_gstin_annual(wb, records)

    wb.save(output_path)
    print(f"\n✓ Saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="GSTR-2B Consolidated Extractor")
    parser.add_argument("--input",  default=".", help="Folder containing GSTR2B_*.xlsx files")
    parser.add_argument("--output", default="GSTR2B_Consolidated_Analysis.xlsx",
                        help="Output file path")
    args = parser.parse_args()

    print(f"Scanning: {os.path.abspath(args.input)}")
    records = load_all_files(args.input)

    if not records:
        print("ERROR: No records extracted. Check that GSTR2B_*.xlsx files are in the input folder.",
              file=sys.stderr)
        sys.exit(1)

    print(f"\nTotal records extracted: {len(records)}")
    build_workbook(records, args.output)


if __name__ == "__main__":
    main()
