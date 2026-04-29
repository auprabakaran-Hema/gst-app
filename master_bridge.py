"""
================================================================================
  MASTER BRIDGE v3.0 — GST ↔ Income Tax Full Reconciliation
  
  KEY CONCEPT (your question answered):
  ─────────────────────────────────────
  GST Portal  → GSTIN (15-char) e.g. 37BENPA6909L1Z3
                BUT GSTIN chars 3-12 = PAN = BENPA6909L
  IT Portal   → PAN directly (BENPA6909L) in AIS/TIS
  
  BRIDGE LOGIC:
  ─────────────
  1. Extract PAN from GSTIN (positions 2-11, zero-indexed)
  2. One PAN can have MULTIPLE GSTINs (state registrations)
  3. Aggregate ALL GSTINs' GSTR-1 Sales, GSTR-3B Tax, GSTR-2B ITC
     → Sum by PAN, by Month → This is your "GST Side"
  4. Read AIS/TIS from IT_RECONCILIATION Excel → This is your "IT Side"
  5. Compare GST Side (by PAN+Month) vs IT Side (by PAN+Month)
     → SALES: GSTR-1 Taxable (all GSTINs) vs AIS/TIS Turnover
     → PURCHASE: GSTR-2B ITC (all GSTINs) vs AIS Purchases
     → Each Month + Full Year Total
     → Each Company (PAN)

  FOLDER STRUCTURE (auto-detected):
  ──────────────────────────────────
  ~/Downloads/GST_Automation/FY2025-26/ClientName/
      ANNUAL_RECONCILIATION_<name>_<fy>.xlsx
      GSTR2B_<Month>_<Year>.xlsx  (12 files)
  
  ~/Downloads/IT_Automation/AY2026-27/ClientName/
      IT_RECONCILIATION_<name>_FY<fy>.xlsx

  WHAT IT PRODUCES:
  ─────────────────
  1. Updates each IT_RECONCILIATION Excel:
     - Fills GSTR-1 columns in TIS_vs_GSTR_Monthly (was blank)
     - Fills GSTR-1 columns in AIS_vs_GSTR_Monthly (was blank)
     - Fills GSTR-2B ITC columns (was blank)
     - Adds PAN_Master_Comparison sheet (new — full comparison)
     - Adds ITC_vs_AIS_Purchase sheet (new)
  2. Creates MASTER_GST_IT_RECONCILIATION.xlsx (all companies)
     - DASHBOARD: all companies, annual totals, status
     - Per-company sheet: Sales + Purchase, each month, diff, flag

================================================================================
"""

import os, sys, re, shutil
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# ── Config ────────────────────────────────────────────────────────────────────
FALLBACK_FY      = "2025-26"
THRESH_OK        = 1000      # ₹ diff < this → ✓ OK
THRESH_WARN      = 50000     # ₹ diff < this → ⚠ Minor  else → ✗ CHECK
NUM_FMT          = "#,##0.00"

FY_MONTHS = ["APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC","JAN","FEB","MAR"]
MONTH_FULL = {
    "April":"APR","May":"MAY","June":"JUN","July":"JUL",
    "August":"AUG","September":"SEP","October":"OCT","November":"NOV",
    "December":"DEC","January":"JAN","February":"FEB","March":"MAR",
}
MONTH_NAME_TO_ABBR = {k.lower():v for k,v in MONTH_FULL.items()}

# ── Colours ───────────────────────────────────────────────────────────────────
C_NAVY   = "1F3864"
C_BLUE   = "2E75B6"
C_TEAL   = "1D6A72"
C_WHITE  = "FFFFFF"
C_LGRAY  = "F2F2F2"
C_DGRAY  = "D6DCE4"
C_GREEN  = "C6EFCE"; C_DKGRN = "276221"
C_AMBER  = "FFEB9C"; C_DKAMB = "9C6500"
C_RED    = "FFC7CE"; C_DKRED = "9C0006"
C_YELLOW = "FFF2CC"
C_PURPLE = "7030A0"

MISSING = []
try:    import pandas as pd
except: MISSING.append("pandas")
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except: MISSING.append("openpyxl")

if MISSING:
    print(f"✗ pip install {' '.join(MISSING)}")
    sys.exit(1)


# ════════════════════════════════════════════════════════════════════════════════
# STYLING
# ════════════════════════════════════════════════════════════════════════════════
def _f(h):  return PatternFill("solid", fgColor=h)
def _fn(bold=False, color="000000", size=9):
    return Font(name="Arial", bold=bold, color=color, size=size)
def _bd():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _al(h="left", w=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=w)

def _cell(ws, r, c, v, bg=C_WHITE, bold=False, fg="000000",
          align="left", fmt=None, size=9, span=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font      = _fn(bold, fg, size)
    cl.fill      = _f(bg)
    cl.alignment = _al(align)
    cl.border    = _bd()
    if fmt:
        cl.number_format = fmt
    elif isinstance(v, (int, float)) and v:
        cl.number_format = NUM_FMT
    if span:
        ws.merge_cells(start_row=r, start_column=c,
                       end_row=r, end_column=c+span-1)
    return cl

def _hdr(ws, cols, row=1, bg=C_NAVY):
    for ci, (h, w) in enumerate(cols, 1):
        cl = ws.cell(row=row, column=ci, value=h)
        cl.font = _fn(True, C_WHITE, 9)
        cl.fill = _f(bg)
        cl.alignment = _al("center")
        cl.border = _bd()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 20

def _title(ws, txt, nc, row=1):
    ws.merge_cells(f"A{row}:{get_column_letter(nc)}{row}")
    cl = ws.cell(row=row, column=1, value=txt)
    cl.font = _fn(True, C_WHITE, 11)
    cl.fill = _f(C_NAVY)
    cl.alignment = _al("center")
    cl.border = _bd()
    ws.row_dimensions[row].height = 24

def _note(ws, txt, nc, row=2):
    ws.merge_cells(f"A{row}:{get_column_letter(nc)}{row}")
    cl = ws.cell(row=row, column=1, value=txt)
    cl.font = _fn(False, "2F5597", 8)
    cl.fill = _f("EBF4FA")
    cl.alignment = _al("left", True)
    cl.border = _bd()
    ws.row_dimensions[row].height = 28

def _sec(ws, ri, txt, nc, bg=C_TEAL):
    ws.merge_cells(f"A{ri}:{get_column_letter(nc)}{ri}")
    cl = ws.cell(row=ri, column=1, value=txt)
    cl.font = _fn(True, C_WHITE, 9)
    cl.fill = _f(bg)
    cl.alignment = _al("left")
    cl.border = _bd()
    ws.row_dimensions[ri].height = 14
    return ri + 1

def _tot(ws, ri, vals, nc, bg=C_DGRAY, fg="000000"):
    for ci, v in enumerate(vals, 1):
        cl = ws.cell(row=ri, column=ci, value=v)
        cl.font = _fn(True, fg, 9)
        cl.fill = _f(bg)
        cl.alignment = _al("right" if isinstance(v, (int, float)) else "left")
        cl.border = _bd()
        if isinstance(v, (int, float)):
            cl.number_format = NUM_FMT
    ws.row_dimensions[ri].height = 18
    return ri + 1

def _flag_style(diff):
    """Returns (text, bg, fg)"""
    if diff is None:
        return "—", C_LGRAY, "000000"
    a = abs(diff)
    if a <= THRESH_OK:   return "✓ OK",    C_GREEN,  C_DKGRN
    if a <= THRESH_WARN: return "⚠ Minor", C_AMBER,  C_DKAMB
    return "✗ CHECK", C_RED, C_DKRED

def _diff_cell(ws, r, c, diff):
    txt, bg, fg = _flag_style(diff)
    if diff is not None:
        cl = _cell(ws, r, c, diff, bg=bg, fg=fg, align="right", fmt=NUM_FMT)
    else:
        cl = _cell(ws, r, c, "—", bg=bg, fg=fg, align="center")
    return cl

def _flag_cell(ws, r, c, diff):
    txt, bg, fg = _flag_style(diff)
    return _cell(ws, r, c, txt, bg=bg, fg=fg, bold="CHECK" in txt, align="center")


# ════════════════════════════════════════════════════════════════════════════════
# HELPER: PAN ↔ GSTIN
# ════════════════════════════════════════════════════════════════════════════════
def pan_from_gstin(gstin):
    """
    GSTIN = State(2) + PAN(10) + Entity(1) + Z + CheckDigit
    PAN   = chars at index 2..11  (0-based)
    """
    gstin = str(gstin or "").strip().upper()
    if len(gstin) == 15:
        return gstin[2:12]
    return ""

def _n(v):
    try:    return round(float(str(v or 0).replace(",", "").replace("₹", "").replace(" ", "")), 2)
    except: return 0.0

def _clean(s):
    if s is None: return ""
    s = str(s).strip()
    return "" if s.lower() in ("nan", "none", "") else s

def _month_key(raw, fy_start):
    """Convert 'April 2025', 'Apr-25', 'APR-2025' → 'APR-2025'"""
    if not raw: return None
    raw = str(raw).strip()

    # Already standard: APR-2025
    m = re.match(r"^([A-Z]{3})-(\d{4})$", raw.upper())
    if m: return raw.upper()

    # Full month name + year: April 2025
    m = re.match(r"^([A-Za-z]+)\s+(\d{4})$", raw)
    if m:
        abbr = MONTH_FULL.get(m.group(1).capitalize())
        if abbr: return f"{abbr}-{m.group(2)}"

    # Short: Apr-25 or Apr 25
    m = re.match(r"^([A-Za-z]{3,4})[- ](\d{2})$", raw)
    if m:
        abbr = m.group(1).upper()[:3]
        yr_s = m.group(2)
        yr   = str(fy_start) if abbr in {"APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"} \
               else str(fy_start + 1)
        return f"{abbr}-{yr}"
    return None

GSTIN_RE = re.compile(r"\b(\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z\d])\b")
PAN_RE   = re.compile(r"\b([A-Z]{5}\d{4}[A-Z])\b")

def _extract_gstin(s):
    m = GSTIN_RE.search(str(s or "").upper())
    return m.group(1) if m else None

def _extract_pan(s):
    m = PAN_RE.search(str(s or "").upper())
    return m.group(1) if m else None


# ════════════════════════════════════════════════════════════════════════════════
# STEP 1 — READ GST DATA  (ANNUAL_RECONCILIATION + GSTR2B files)
# Returns pan_gst_data[pan][mkey] = {r1_taxable, r3b_total, itc_igst, itc_cgst, itc_sgst, itc_total}
# Plus pan_gst_data[pan]["_gstins"] = list of GSTINs found
# Plus pan_gst_data[pan]["_annual_r1"] = sum of all 12 months r1_taxable
# ════════════════════════════════════════════════════════════════════════════════

def read_gst_folder(gst_folder, fy):
    """
    Read ALL ANNUAL_RECONCILIATION xlsx and GSTR2B xlsx in folder.
    Aggregates by PAN (extracted from GSTIN in filename/title row).
    Returns:
      {
        pan: {
          "APR-2025": {r1_taxable, r1_igst, r1_cgst, r1_sgst, r1_tax_total,
                       r3b_igst, r3b_cgst, r3b_sgst, r3b_tax_total,
                       itc_igst, itc_cgst, itc_sgst, itc_total},
          ...
          "_gstins": ["37BENPA6909L1Z3", "33BENPA6909L1ZB"],
          "_annual_r1": 10815832.0,
          "_annual_itc": 513371.0,
        }
      }
    """
    gst_folder = Path(gst_folder)
    fy_start   = int(fy.split("-")[0])
    pan_data   = defaultdict(lambda: defaultdict(lambda: {
        "r1_taxable":0.0,"r1_igst":0.0,"r1_cgst":0.0,"r1_sgst":0.0,"r1_tax_total":0.0,
        "r3b_igst":0.0,"r3b_cgst":0.0,"r3b_sgst":0.0,"r3b_tax_total":0.0,
        "itc_igst":0.0,"itc_cgst":0.0,"itc_sgst":0.0,"itc_total":0.0,
    }))
    pan_meta   = defaultdict(lambda: {"_gstins":[], "_annual_r1":0.0, "_annual_itc":0.0})

    # ── Read ANNUAL_RECONCILIATION files ──────────────────────────────────
    # rglob: finds files inside GSTIN subfolders like 33EZEPK9321K1Z6/ too
    _ann_files = sorted(gst_folder.rglob("ANNUAL_RECONCILIATION*.xlsx"))
    if not _ann_files:
        _ann_files = sorted(gst_folder.rglob("*Reconcil*.xlsx"))
    for xl_path in _ann_files:
        gstin = ""
        # 1. Try from filename
        m = GSTIN_RE.search(xl_path.name.upper())
        if m: gstin = m.group(1)

        try:
            xl = pd.ExcelFile(xl_path, engine="openpyxl")
        except Exception as e:
            print(f"    ⚠  Cannot open {xl_path.name}: {e}"); continue

        # 2. Try from title row of any sheet if filename didn't have it
        if not gstin:
            for sn in xl.sheet_names[:3]:
                try:
                    df = xl.parse(sn, nrows=3, header=None, dtype=str).fillna("")
                    for _, row in df.iterrows():
                        combined = " ".join(str(v) for v in row)
                        m2 = GSTIN_RE.search(combined.upper())
                        if m2: gstin = m2.group(1); break
                    if gstin: break
                except: pass

        pan = pan_from_gstin(gstin) if gstin else ""
        if not pan:
            # Try PAN directly from filename
            m3 = PAN_RE.search(xl_path.name.upper())
            pan = m3.group(1) if m3 else "UNKNOWN"

        if gstin and gstin not in pan_meta[pan]["_gstins"]:
            pan_meta[pan]["_gstins"].append(gstin)

        print(f"    Reading ANNUAL_RECON: {xl_path.name}")
        print(f"      GSTIN: {gstin}  →  PAN: {pan}")

        # Read Monthwise_Reconciliation
        for sn in xl.sheet_names:
            if sn.lower() not in ["monthwise_reconciliation","r1_vs_3b_recon",
                                   "gstr3b_vs_r1_recon","monthly_summary"]:
                continue
            try:
                df = xl.parse(sn, header=None, dtype=str).fillna("")
            except: continue

            for _, row in df.iterrows():
                vals = [_clean(str(v)) for v in row]
                if not vals or not vals[0]: continue
                m4 = re.match(r"^([A-Za-z]+)\s+(\d{4})$", vals[0].strip())
                if not m4: continue
                mon_full = m4.group(1).capitalize()
                yr       = m4.group(2)
                abbr     = MONTH_FULL.get(mon_full, mon_full[:3].upper())
                mkey     = f"{abbr}-{yr}"
                nums = []
                for v in vals[1:]:
                    try: nums.append(round(float(str(v).replace(",","")),2))
                    except: nums.append(0.0)
                def g(i): return nums[i] if i < len(nums) else 0.0

                # Columns (0-based from col1 onward) — matches gst_suite ANNUAL_RECONCILIATION
                # Monthwise_Reconciliation layout (R1_vs_3B_Recon sheet):
                # 0=R1_Taxable, 1=R1_IGST, 2=R1_CGST, 3=R1_SGST, 4=R1_Tax_Total
                # 5=3B_IGST, 6=3B_CGST, 7=3B_SGST, 8=3B_Tax_Total
                # 9=R1vs3B_Diff, 10=2B_ITC_IGST, 11=2B_ITC_CGST, 12=2B_ITC_SGST, 13=2B_ITC_Total
                d = pan_data[pan][mkey]
                d["r1_taxable"]    += g(0)
                d["r1_igst"]       += g(1)
                d["r1_cgst"]       += g(2)
                d["r1_sgst"]       += g(3)
                d["r1_tax_total"]  += g(4)
                d["r3b_igst"]      += g(5)
                d["r3b_cgst"]      += g(6)
                d["r3b_sgst"]      += g(7)
                d["r3b_tax_total"] += g(8)
                d["itc_igst"]      += g(10)
                d["itc_cgst"]      += g(11)
                d["itc_sgst"]      += g(12)
                # col 13 = 2B ITC Total (direct); fall back to sum of components
                itc_total_direct    = g(13)
                itc_total_calc      = g(10) + g(11) + g(12)
                d["itc_total"]     += itc_total_direct if itc_total_direct > 0 else itc_total_calc
            break  # first matching sheet

    # ── Read GSTR2B monthly files (override/supplement ITC) ──────────────
    for xl_path in sorted(gst_folder.rglob("GSTR2B_*.xlsx")):
        parts  = xl_path.stem.split("_")  # ['GSTR2B','April','2025'] or ['GSTR2B','37BENPA6909L1Z3','April','2025']
        gstin_from_file = ""
        mon_name = yr_str = ""

        # Two filename formats:
        # GSTR2B_April_2025.xlsx
        # GSTR2B_37BENPA6909L1Z3_April_2025.xlsx
        for i, p in enumerate(parts[1:], 1):
            if GSTIN_RE.match(p.upper()):
                gstin_from_file = p.upper()
            elif p.lower() in MONTH_NAME_TO_ABBR:
                mon_name = p.lower()
            elif re.match(r"^\d{4}$", p):
                yr_str = p

        if not mon_name or not yr_str: continue
        abbr = MONTH_NAME_TO_ABBR[mon_name]
        mkey = f"{abbr}-{yr_str}"

        # Determine PAN for this 2B file
        pan_2b = pan_from_gstin(gstin_from_file) if gstin_from_file else ""
        if not pan_2b:
            # Try to infer from ANNUAL_RECONCILIATION pans already loaded
            if len(pan_meta) == 1:
                pan_2b = list(pan_meta.keys())[0]
            else:
                pan_2b = "UNKNOWN"

        try:
            xl = pd.ExcelFile(xl_path, engine="openpyxl")
        except: continue

        igst = cgst = sgst = 0.0

        # Try ITC Available sheet first
        if "ITC Available" in xl.sheet_names:
            try:
                df = xl.parse("ITC Available", header=None, dtype=str).fillna("")
                for _, row in df.iterrows():
                    vals  = [_clean(str(v)) for v in row]
                    combo = " ".join(vals).lower()
                    if "4(a)(5)" in combo or "all other itc" in combo:
                        nums = [_n(v) for v in vals if _n(v) > 100]
                        if len(nums) >= 3:
                            igst, cgst, sgst = nums[0], nums[1], nums[2]
                        elif len(nums) == 2:
                            igst, cgst = nums[0], nums[1]
                        elif len(nums) == 1:
                            igst = nums[0]
                        break
            except: pass

        # Fallback: sum B2B sheet
        if igst == 0 and "B2B" in xl.sheet_names:
            try:
                df = xl.parse("B2B", header=None, dtype=str).fillna("")
                hdr_row = None
                for i in range(min(12, len(df))):
                    txt = " ".join(str(v).lower() for v in df.iloc[i])
                    if "integrated" in txt or "igst" in txt:
                        hdr_row = i; break
                if hdr_row is not None:
                    ci_ig = ci_cg = ci_sg = None
                    for ci, v in enumerate(df.iloc[hdr_row]):
                        vl = str(v).lower()
                        if "integrated" in vl: ci_ig = ci
                        elif "central" in vl:  ci_cg = ci
                        elif "state" in vl:    ci_sg = ci
                    for i in range(hdr_row + 2, len(df)):
                        r = df.iloc[i]
                        if ci_ig is not None: igst += _n(r.iloc[ci_ig])
                        if ci_cg is not None: cgst += _n(r.iloc[ci_cg])
                        if ci_sg is not None: sgst += _n(r.iloc[ci_sg])
            except: pass

        if igst + cgst + sgst > 0:
            d = pan_data[pan_2b][mkey]
            # Only override if better data
            if igst + cgst + sgst > d["itc_total"]:
                d["itc_igst"]  = igst
                d["itc_cgst"]  = cgst
                d["itc_sgst"]  = sgst
                d["itc_total"] = igst + cgst + sgst

    # ── Read GSTR-1 from ZIP/JSON files (when ANNUAL_RECONCILIATION not available) ──
    import zipfile, json as _json
    for xl_path in sorted(gst_folder.rglob("GSTR1_*.zip")):
        parts = xl_path.stem.split("_")  # ['GSTR1','April','2025'] or ['GSTR1','GSTIN','April','2025']
        gstin_from_file = ""
        mon_name = yr_str = ""
        for p in parts[1:]:
            if GSTIN_RE.match(p.upper()):
                gstin_from_file = p.upper()
            elif p.lower() in MONTH_NAME_TO_ABBR:
                mon_name = p.lower()
            elif re.match(r"^\d{4}$", p):
                yr_str = p
        if not mon_name or not yr_str:
            continue
        abbr = MONTH_NAME_TO_ABBR[mon_name]
        mkey = f"{abbr}-{yr_str}"
        pan_z = pan_from_gstin(gstin_from_file) if gstin_from_file else ""
        if not pan_z:
            # Infer from parent folder name (GSTIN folder)
            parent_gstin = xl_path.parent.name.upper()
            if GSTIN_RE.match(parent_gstin):
                pan_z = pan_from_gstin(parent_gstin)
                if pan_z and parent_gstin not in pan_meta[pan_z]["_gstins"]:
                    pan_meta[pan_z]["_gstins"].append(parent_gstin)
        if not pan_z:
            # Try to get GSTIN from TaxLiability Excel in same folder
            for _tl in xl_path.parent.glob("TaxLiability_*.xlsx"):
                try:
                    import openpyxl as _oox
                    _twb = _oox.load_workbook(str(_tl), read_only=True, data_only=True)
                    for _tsn in _twb.sheetnames[:3]:
                        _tws = _twb[_tsn]
                        for _tr in _tws.iter_rows(min_row=1, max_row=10, values_only=True):
                            for _tv in _tr:
                                if _tv:
                                    _gm = GSTIN_RE.search(str(_tv).upper())
                                    if _gm:
                                        _found_gstin = _gm.group(1)
                                        pan_z = pan_from_gstin(_found_gstin)
                                        if pan_z and _found_gstin not in pan_meta[pan_z]["_gstins"]:
                                            pan_meta[pan_z]["_gstins"].append(_found_gstin)
                                        break
                            if pan_z: break
                        _twb.close()
                        if pan_z: break
                except: pass
                if pan_z: break
        if not pan_z:
            pan_z = list(pan_meta.keys())[0] if len(pan_meta) == 1 else "UNKNOWN"
        # Only fill if r1_taxable still zero (ANNUAL_RECONCILIATION not loaded)
        if pan_data[pan_z][mkey]["r1_taxable"] != 0.0:
            continue
        try:
            taxable = 0.0
            with zipfile.ZipFile(str(xl_path)) as z:
                for name in z.namelist():
                    if not name.endswith(".json"):
                        continue
                    data = _json.loads(z.read(name))
                    for inv in data.get("b2b", []):
                        for itm in inv.get("inv", []):
                            for it in itm.get("itms", []):
                                taxable += float(it.get("itm_det", {}).get("txval", 0))
                    for itm in data.get("b2cs", []):
                        taxable += float(itm.get("txval", 0))
                    for itm in data.get("b2cl", []):
                        for inv2 in itm.get("inv", []):
                            taxable += float(inv2.get("txval", 0))
                    for itm in data.get("exp", []):
                        for inv2 in itm.get("exp", []):
                            taxable += float(inv2.get("txval", 0))
                    break  # first JSON in zip
            if taxable > 0:
                pan_data[pan_z][mkey]["r1_taxable"] = round(taxable, 2)
        except Exception as _e:
            print(f"    ⚠  Could not read {xl_path.name}: {_e}")

    # ── Read GSTR-3B from TaxLiability Excel (when ANNUAL_RECONCILIATION not available) ──
    for xl_path in sorted(gst_folder.rglob("TaxLiability_*.xlsx")):
        # Determine FY from filename e.g. TaxLiability_2025_26.xlsx
        fy_tag = xl_path.stem.replace("TaxLiability_","").replace("_","-")  # "2025-26"
        if fy_tag != fy:
            continue
        # Determine PAN/GSTIN from parent folder
        parent = xl_path.parent.name.upper()
        pan_tl = ""
        if GSTIN_RE.match(parent):
            pan_tl = pan_from_gstin(parent)
            if pan_tl and parent not in pan_meta[pan_tl]["_gstins"]:
                pan_meta[pan_tl]["_gstins"].append(parent)
        if not pan_tl:
            pan_tl = list(pan_meta.keys())[0] if len(pan_meta) == 1 else "UNKNOWN"
        try:
            xl = pd.ExcelFile(str(xl_path), engine="openpyxl")
        except:
            continue
        # Read "Tax Liability Summary" or "Tax liability" sheet
        for sn in xl.sheet_names:
            if sn.lower() in ["tax liability summary", "tax liability", "tax_liability_summary"]:
                try:
                    df = xl.parse(sn, header=None, dtype=str).fillna("")
                except:
                    continue
                for _, row in df.iterrows():
                    vals = [_clean(str(v)) for v in row]
                    if not vals or not vals[0]: continue
                    # Match rows like "Apr-25", "May-25" etc.
                    m5 = re.match(r"^([A-Za-z]{3})[-]?(\d{2})$", vals[0].strip())
                    if not m5: continue
                    mon_abbr = m5.group(1).upper()
                    yr_2d    = m5.group(2)
                    yr_full  = ("20" + yr_2d) if int(yr_2d) < 50 else ("19" + yr_2d)
                    mkey = f"{mon_abbr}-{yr_full}"
                    # Columns: IGST(1) CGST(2) SGST(3) CESS(4) TOTAL(5) | IGST_3B(6) CGST_3B(7) SGST_3B(8) CESS_3B(9) TOTAL_3B(10)
                    nums = []
                    for v in vals[1:]:
                        try: nums.append(round(float(str(v).replace(",","")), 2))
                        except: nums.append(0.0)
                    def gn(i): return nums[i] if i < len(nums) else 0.0
                    d = pan_data[pan_tl][mkey]
                    # Only fill 3B if not already set
                    if d["r3b_tax_total"] == 0.0:
                        d["r3b_igst"]      = gn(5)   # paid as per GSTR-3B IGST
                        d["r3b_cgst"]      = gn(6)
                        d["r3b_sgst"]      = gn(7)
                        d["r3b_tax_total"] = gn(9)   # GSTR-3B total paid
                break

    # ── Compute annual totals ─────────────────────────────────────────────
    for pan in pan_data:
        monthly_vals = {k: v for k, v in pan_data[pan].items()
                        if k != "_meta" and re.match(r"[A-Z]{3}-\d{4}", k)}
        pan_meta[pan]["_annual_r1"]  = sum(v["r1_taxable"]  for v in monthly_vals.values())
        pan_meta[pan]["_annual_itc"] = sum(v["itc_total"]   for v in monthly_vals.values())
        pan_meta[pan]["_annual_3b"]  = sum(v["r3b_tax_total"] for v in monthly_vals.values())
        print(f"    PAN {pan}: {len(monthly_vals)} months | "
              f"R1 ₹{pan_meta[pan]['_annual_r1']:,.0f} | "
              f"ITC ₹{pan_meta[pan]['_annual_itc']:,.0f} | "
              f"GSTINs: {pan_meta[pan]['_gstins']}")

    # ── Merge UNKNOWN pan into correct PAN if only one real PAN exists ──────
    if "UNKNOWN" in pan_data and len([p for p in pan_data if p != "UNKNOWN"]) == 1:
        real_pan = [p for p in pan_data if p != "UNKNOWN"][0]
        for mkey, d in pan_data["UNKNOWN"].items():
            if not re.match(r"[A-Z]{3}-\d{4}", mkey): continue
            td = pan_data[real_pan][mkey]
            # Merge: take max of each field (UNKNOWN has ITC from GSTR2B, real_pan has R1/3B)
            for k in d:
                if d[k] > td[k]:
                    td[k] = d[k]
        del pan_data["UNKNOWN"]
        if "UNKNOWN" in pan_meta: del pan_meta["UNKNOWN"]
        # Recompute annual totals for merged PAN
        mv = {k:v for k,v in pan_data[real_pan].items() if re.match(r"[A-Z]{3}-\d{4}",k)}
        pan_meta[real_pan]["_annual_r1"]  = sum(v["r1_taxable"]    for v in mv.values())
        pan_meta[real_pan]["_annual_itc"] = sum(v["itc_total"]     for v in mv.values())
        pan_meta[real_pan]["_annual_3b"]  = sum(v["r3b_tax_total"] for v in mv.values())
        print(f"    Merged UNKNOWN → {real_pan}: R1 ₹{pan_meta[real_pan]['_annual_r1']:,.0f} ",
              f"ITC ₹{pan_meta[real_pan]['_annual_itc']:,.0f} ",
              f"3B ₹{pan_meta[real_pan]['_annual_3b']:,.0f}")

    return dict(pan_data), dict(pan_meta)


# ════════════════════════════════════════════════════════════════════════════════
# STEP 2 — READ IT DATA from IT_RECONCILIATION Excel
# Returns it_data = {
#   "pan": "BENPA6909L",
#   "name": "...",
#   "fy": "2025-26",
#   "gstins": ["37BENPA6909L1Z3","33BENPA6909L1ZB"],
#   "ais_monthly": { gstin: {mkey: {ais_sales, r1_sales, r3b_sales, ais_pur, r3b_itc}} },
#   "tis_monthly": { gstin: {mkey: {ais_total, ais_taxable, r1_taxable, r3b_filed, diff}} },
#   "pan_ais_monthly": { mkey: {ais_sales, ais_pur} },  ← SUM all GSTINs
#   "pan_tis_monthly": { mkey: {tis_sales} },            ← SUM all GSTINs
#   "annual_ais_sales": float, "annual_tis_sales": float,
#   "annual_ais_pur": float,  "annual_tis_pur": float,
#   "tds": float,
# }
# ════════════════════════════════════════════════════════════════════════════════

def read_it_recon(it_recon_path):
    try:
        wb = load_workbook(it_recon_path, data_only=True)
    except Exception as e:
        print(f"    ✗ Cannot open IT Recon: {e}"); return None

    result = {
        "pan":"","name":"","fy":"","gstins":[],
        "ais_monthly":{},"tis_monthly":{},
        "pan_ais_monthly":defaultdict(lambda:{"ais_sales":0.0,"ais_pur":0.0}),
        "pan_tis_monthly":defaultdict(lambda:{"tis_sales":0.0,"tis_taxable":0.0}),
        "annual_ais_sales":0.0,"annual_tis_sales":0.0,
        "annual_ais_pur":0.0,"annual_tis_pur":0.0,"tds":0.0,
    }

    # IT_Summary — get PAN, name, GSTINs, AIS/TIS totals
    if "IT_Summary" in wb.sheetnames:
        ws = wb["IT_Summary"]
        for r in range(1, ws.max_row + 1):
            c1  = ws.cell(r, 1).value
            c2  = ws.cell(r, 2).value
            c5  = ws.cell(r, 5).value  # sometimes notes col
            lbl = str(c1 or "").strip().lower()

            if not result["name"] and "assessee" in lbl or "company" in lbl or "name" in lbl:
                v = str(c2 or c5 or "").strip()
                if v and len(v) > 3: result["name"] = v

            if not result["pan"]:
                # Try col2 for PAN
                p = _extract_pan(str(c2 or ""))
                if not p: p = _extract_pan(str(c5 or ""))
                if p: result["pan"] = p

            g = _extract_gstin(str(c2 or "") + " " + str(c5 or ""))
            if g and g not in result["gstins"]:
                result["gstins"].append(g)
                if not result["pan"]:
                    result["pan"] = pan_from_gstin(g)

            if "tis" in lbl and "turnover" in lbl:
                result["annual_tis_sales"] = _n(c2)
            if "ais" in lbl and "turnover" in lbl and "gstin" not in lbl:
                result["annual_ais_sales"] = _n(c2)
            if "tis" in lbl and "purchase" in lbl:
                result["annual_tis_pur"] = _n(c2)
            if "ais" in lbl and "purchase" in lbl:
                result["annual_ais_pur"] = _n(c2)
            if "tds" in lbl and "total" in lbl:
                result["tds"] = _n(c2)
            if "financial year" in lbl or "fy" == lbl:
                fy_v = str(c2 or "").strip()
                if re.match(r"\d{4}-\d{2}", fy_v):
                    result["fy"] = fy_v

    # AIS_vs_GSTR_Monthly → per GSTIN, per month
    if "AIS_vs_GSTR_Monthly" in wb.sheetnames:
        ws = wb["AIS_vs_GSTR_Monthly"]
        cur_gstin = None
        for r in range(4, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            c2 = ws.cell(r, 2).value
            if c1 and not c2:
                g = _extract_gstin(str(c1))
                if g:
                    cur_gstin = g
                    result["ais_monthly"].setdefault(cur_gstin, {})
                    if g not in result["gstins"]:
                        result["gstins"].append(g)
                    if not result["pan"]:
                        result["pan"] = pan_from_gstin(g)
                continue
            if not c2 or not cur_gstin: continue
            mkey = str(c2).strip().upper()
            if not re.match(r"[A-Z]{3}-\d{4}", mkey): continue

            ais_s = _n(ws.cell(r, 3).value)
            ais_p = _n(ws.cell(r, 7).value)
            result["ais_monthly"][cur_gstin][mkey] = {
                "ais_sales":  ais_s,
                "r1_sales":   _n(ws.cell(r, 4).value),
                "r3b_sales":  _n(ws.cell(r, 6).value),
                "ais_pur":    ais_p,
                "r3b_itc":    _n(ws.cell(r, 9).value),
                "sales_diff": _n(ws.cell(r, 10).value),
                "status":     str(ws.cell(r, 12).value or ""),
            }
            # Aggregate to PAN level
            result["pan_ais_monthly"][mkey]["ais_sales"] += ais_s
            result["pan_ais_monthly"][mkey]["ais_pur"]   += ais_p

    # TIS_vs_GSTR_Monthly → per GSTIN, per month
    if "TIS_vs_GSTR_Monthly" in wb.sheetnames:
        ws = wb["TIS_vs_GSTR_Monthly"]
        cur_gstin = None
        for r in range(4, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            c2 = ws.cell(r, 2).value
            if c1 and not c2:
                g = _extract_gstin(str(c1))
                if g:
                    cur_gstin = g
                    result["tis_monthly"].setdefault(cur_gstin, {})
                    if g not in result["gstins"]:
                        result["gstins"].append(g)
                continue
            if not c2 or not cur_gstin: continue
            mkey = str(c2).strip().upper()
            if not re.match(r"[A-Z]{3}-\d{4}", mkey): continue

            tis_s  = _n(ws.cell(r, 3).value)
            tis_tx = _n(ws.cell(r, 4).value)
            result["tis_monthly"][cur_gstin][mkey] = {
                "ais_total":   tis_s,
                "ais_taxable": tis_tx,
                "r1_taxable":  _n(ws.cell(r, 5).value),
                "r3b_filed":   _n(ws.cell(r, 7).value),
                "diff":        _n(ws.cell(r, 8).value),
                "status":      str(ws.cell(r, 9).value or ""),
            }
            # Aggregate to PAN level (use taxable as best proxy for sales)
            result["pan_tis_monthly"][mkey]["tis_sales"]   += tis_s
            result["pan_tis_monthly"][mkey]["tis_taxable"] += tis_tx

    wb.close()

    # Fix: if AIS monthly sum > annual, update annual
    pan_ais_ann = sum(v["ais_sales"] for v in result["pan_ais_monthly"].values())
    if pan_ais_ann > result["annual_ais_sales"]:
        result["annual_ais_sales"] = pan_ais_ann

    print(f"    IT Recon: PAN={result['pan']}  Name={result['name']}")
    print(f"      GSTINs: {result['gstins']}")
    print(f"      AIS Turnover (PAN): ₹{result['annual_ais_sales']:,.0f}  "
          f"AIS Purchases: ₹{result['annual_ais_pur']:,.0f}")
    return result


# ════════════════════════════════════════════════════════════════════════════════
# STEP 3 — INJECT GST DATA INTO IT_RECONCILIATION EXCEL
# This fills blank columns and adds PAN_Master_Comparison sheet
# ════════════════════════════════════════════════════════════════════════════════

def inject_into_it_recon(it_recon_path, pan_gst, pan_meta_single, it_data, fy):
    """
    pan_gst          : {mkey: {r1_taxable, r3b_tax_total, itc_igst, ...}}  ← PAN-aggregated GST data
    pan_meta_single  : {"_gstins":[...], "_annual_r1":..., ...}
    it_data          : result from read_it_recon()
    """
    try:
        wb = load_workbook(it_recon_path)
    except Exception as e:
        print(f"    ✗ Cannot open {Path(it_recon_path).name}: {e}"); return False

    fy_start = int(fy.split("-")[0])
    pan = it_data.get("pan","")
    name = it_data.get("name","")
    changed = False

    # ── Fill TIS_vs_GSTR_Monthly ──────────────────────────────────────────
    if "TIS_vs_GSTR_Monthly" in wb.sheetnames:
        ws = wb["TIS_vs_GSTR_Monthly"]
        # Add header for new columns if needed
        for col, lbl in [(5,"GSTR-1 Taxable ₹"),(7,"GSTR-3B Filed ₹"),
                         (8,"Diff (R1-AIS) ₹"),(11,"⚠ Flag")]:
            if not _clean(str(ws.cell(3, col).value)):
                cl = ws.cell(3, col, lbl)
                cl.font = _fn(True, C_WHITE, 9)
                cl.fill = _f(C_NAVY)
                cl.alignment = _al("center")
                cl.border = _bd()

        cur_gstin = None
        for r in range(4, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            c2 = ws.cell(r, 2).value
            if c1 and not c2:
                cur_gstin = _extract_gstin(str(c1))
                continue
            if not c2: continue
            mkey = str(c2).strip().upper()
            if not re.match(r"[A-Z]{3}-\d{4}", mkey): continue

            # We use PAN-level GST data (sum of all GSTINs)
            gd   = pan_gst.get(mkey, {})
            r1   = gd.get("r1_taxable", 0.0)
            r3b  = gd.get("r3b_tax_total", 0.0)
            ais  = _n(ws.cell(r, 4).value)  # AIS Taxable from IT portal

            if r1:
                cl = ws.cell(r, 5, r1)
                cl.number_format = NUM_FMT
                cl.fill = _f("EBF3FB")
                cl.alignment = _al("right")
                cl.border = _bd()
                changed = True
            if r3b:
                cl = ws.cell(r, 7, r3b)
                cl.number_format = NUM_FMT
                cl.fill = _f(C_YELLOW)
                cl.alignment = _al("right")
                cl.border = _bd()
                changed = True
            if r1:
                diff = r1 - ais
                ws.cell(r, 8, diff).number_format = NUM_FMT
                ws.cell(r, 8).fill = _f(C_GREEN if abs(diff)<THRESH_OK else
                                         C_AMBER if abs(diff)<THRESH_WARN else C_RED)
                ws.cell(r, 8).border = _bd()
                txt, bg, fg = _flag_style(diff)
                fc = ws.cell(r, 11, txt)
                fc.fill = _f(bg); fc.font = _fn("CHECK" in txt, fg, 9)
                fc.alignment = _al("center"); fc.border = _bd()
                changed = True

    # ── Fill AIS_vs_GSTR_Monthly ──────────────────────────────────────────
    if "AIS_vs_GSTR_Monthly" in wb.sheetnames:
        ws = wb["AIS_vs_GSTR_Monthly"]
        for col, lbl in [(4,"GSTR-1 Sales ₹"),(6,"GSTR-3B Sales ₹"),
                         (9,"GSTR-2B ITC ₹"),(10,"Sales Diff ₹"),
                         (11,"Purchase Diff ₹"),(13,"Sales Flag"),(14,"Purch Flag")]:
            if not _clean(str(ws.cell(3, col).value)):
                cl = ws.cell(3, col, lbl)
                cl.font = _fn(True, C_WHITE, 9)
                cl.fill = _f(C_NAVY)
                cl.alignment = _al("center")
                cl.border = _bd()

        cur_gstin = None
        for r in range(4, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            c2 = ws.cell(r, 2).value
            if c1 and not c2:
                cur_gstin = _extract_gstin(str(c1))
                continue
            if not c2: continue
            mkey = str(c2).strip().upper()
            if not re.match(r"[A-Z]{3}-\d{4}", mkey): continue

            gd    = pan_gst.get(mkey, {})
            r1    = gd.get("r1_taxable", 0.0)
            r3b   = gd.get("r3b_tax_total", 0.0)
            itc   = gd.get("itc_total", 0.0)
            ais_s = _n(ws.cell(r, 3).value)   # AIS sales
            ais_p = _n(ws.cell(r, 7).value)   # AIS purchases

            for col, val in [(4, r1), (6, r3b), (9, itc)]:
                if val:
                    cl = ws.cell(r, col, val)
                    cl.number_format = NUM_FMT
                    cl.fill = _f("EBF3FB")
                    cl.alignment = _al("right")
                    cl.border = _bd()
                    changed = True

            if r1 and ais_s:
                sd = r1 - ais_s
                cl = ws.cell(r, 10, sd)
                cl.number_format = NUM_FMT
                cl.fill = _f(C_GREEN if abs(sd)<THRESH_OK else
                              C_AMBER if abs(sd)<THRESH_WARN else C_RED)
                cl.border = _bd()
                txt, bg, fg = _flag_style(sd)
                fc = ws.cell(r, 13, txt)
                fc.fill = _f(bg); fc.font = _fn(False, fg, 9)
                fc.alignment = _al("center"); fc.border = _bd()
                changed = True

            if itc and ais_p:
                pd_ = itc - ais_p
                cl = ws.cell(r, 11, pd_)
                cl.number_format = NUM_FMT
                cl.fill = _f(C_GREEN if abs(pd_)<THRESH_OK else
                              C_AMBER if abs(pd_)<THRESH_WARN else C_RED)
                cl.border = _bd()
                txt, bg, fg = _flag_style(pd_)
                fc = ws.cell(r, 14, txt)
                fc.fill = _f(bg); fc.font = _fn(False, fg, 9)
                fc.alignment = _al("center"); fc.border = _bd()
                changed = True

    # ── Update IT_Summary turnover ────────────────────────────────────────
    if "IT_Summary" in wb.sheetnames:
        ann_r1 = pan_meta_single.get("_annual_r1", 0.0)
        if ann_r1:
            ws = wb["IT_Summary"]
            for r in range(1, ws.max_row + 1):
                lbl = str(ws.cell(r, 1).value or "").lower()
                if any(k in lbl for k in ["tis gst turnover","ais gst turnover",
                                           "gstr-1 turnover","gst turnover"]):
                    tis_val = _n(ws.cell(r, 2).value)
                    diff    = ann_r1 - tis_val
                    ws.cell(r, 3, ann_r1).number_format = NUM_FMT
                    ws.cell(r, 4, diff).number_format   = NUM_FMT
                    txt, bg, fg = _flag_style(diff)
                    ws.cell(r, 5, txt)
                    ws.cell(r, 6, f"GSTR-1 total (all GSTINs under PAN {pan})")
            changed = True

    # ── Add / replace PAN_Master_Comparison sheet ─────────────────────────
    if "PAN_Master_Comparison" in wb.sheetnames:
        del wb["PAN_Master_Comparison"]
    ws_pm = wb.create_sheet("PAN_Master_Comparison", 0)
    ws_pm.sheet_view.showGridLines = False
    ws_pm.freeze_panes = "C5"
    _build_pan_master_sheet(ws_pm, pan_gst, pan_meta_single, it_data, fy)
    changed = True

    # ── Add / replace ITC_vs_AIS_Purchase sheet ───────────────────────────
    if "ITC_vs_AIS_Purchase" in wb.sheetnames:
        del wb["ITC_vs_AIS_Purchase"]
    ws_itc = wb.create_sheet("ITC_vs_AIS_Purchase")
    ws_itc.sheet_view.showGridLines = False
    _build_itc_sheet(ws_itc, pan_gst, it_data, fy)
    changed = True

    if changed:
        try:
            wb.save(it_recon_path)
            print(f"    ✓ IT Recon updated: {Path(it_recon_path).name}")
            return True
        except Exception as e:
            print(f"    ✗ Save error: {e}"); return False
    return False


# ════════════════════════════════════════════════════════════════════════════════
# BUILD PAN_Master_Comparison SHEET
# THE CORE OUTPUT — Shows full comparison per GSTIN per month + PAN totals
# SALES:    GSTR-1 (GST Portal, all GSTINs) vs AIS/TIS (IT Portal, by PAN)
# PURCHASE: GSTR-2B ITC (GST Portal) vs AIS Purchases (IT Portal)
# ════════════════════════════════════════════════════════════════════════════════

def _build_pan_master_sheet(ws, pan_gst, pan_meta, it_data, fy):
    fy_start = int(fy.split("-")[0])
    pan  = it_data.get("pan", "")
    name = it_data.get("name", "")
    gstins_in_it = it_data.get("gstins", [])
    gstins_in_gst= pan_meta.get("_gstins", [])
    all_gstins   = sorted(set(gstins_in_it + gstins_in_gst))

    NC = 17
    COLS = [
        ("GSTIN",           20),  # 1
        ("Month",           10),  # 2
        # GST Portal (from GSTR-1/3B/2B)
        ("GSTR-1 Taxable ₹",17), # 3  — Sales declared in GSTR-1
        ("GSTR-3B Tax ₹",   15), # 4  — Tax filed in GSTR-3B
        ("GSTR-2B ITC ₹",   15), # 5  — ITC claimed (all GSTINs combined)
        # IT Portal (from AIS/TIS)
        ("AIS Sales ₹",     16), # 6  — IT portal sees your sales
        ("TIS Sales ₹",     16), # 7  — Taxpayer-confirmed in TIS
        ("AIS Purchases ₹", 16), # 8  — IT portal sees your purchases
        # Comparisons
        ("Sales Diff ₹",    15), # 9  — GSTR-1 minus AIS Sales
        ("Sales Flag",      10), # 10
        ("ITC−AIS Diff ₹",  15), # 11 — GSTR-2B ITC minus AIS Purchases
        ("ITC Flag",        10), # 12
        # Extra context
        ("GSTR-1 Invoices", 10), # 13
        ("GSTR-1A Amend ₹", 14), # 14
        ("PAN",             12), # 15
        ("Source",          10), # 16
        ("Action Required", 36), # 17
    ]

    ri = 1
    _title(ws,
        f"PAN MASTER COMPARISON — {name} ({pan}) — FY {fy}  "
        f"[GST Portal (GSTIN-wise) ↔ IT Portal (PAN-wise)]", NC, row=ri); ri += 1
    _note(ws,
        f"HOW PAN BRIDGE WORKS: GSTIN {' + '.join(all_gstins[:3])} → "
        f"PAN {pan} (chars 3-12 of GSTIN = PAN)  |  "
        f"GSTR-1 = sum of all GSTINs under this PAN each month  |  "
        f"AIS/TIS = IT portal aggregates ALL GSTINs → reports at PAN level  |  "
        f"Threshold: ✓ OK<₹{THRESH_OK:,}  ⚠ Minor<₹{THRESH_WARN:,}  ✗ CHECK≥₹{THRESH_WARN:,}",
        NC, row=ri); ri += 1

    ws.row_dimensions[ri].height = 28
    # Sub-header explaining the two sides
    ws.merge_cells(f"A{ri}:{get_column_letter(NC)}{ri}")
    cl = ws.cell(ri, 1,
        "◀ GST PORTAL SIDE ▶  GSTR-1 Taxable  |  GSTR-3B Filed Tax  |  GSTR-2B ITC"
        "        ↔        "
        "◀ IT PORTAL SIDE ▶  AIS GST Turnover  |  TIS Confirmed Sales  |  AIS Purchases"
        "        →        "
        "◀ DIFFERENCE & FLAG ▶  Red=Action Needed  |  Amber=Minor  |  Green=OK")
    cl.font = _fn(True, "2F5597", 8)
    cl.fill = _f("DDEEFF")
    cl.alignment = _al("center", True)
    cl.border = _bd()
    ri += 1

    _hdr(ws, COLS, row=ri); ri += 1

    fy_months_keys = [
        f"{'APR' if m not in ('JAN','FEB','MAR') else m}-{fy_start if m not in ('JAN','FEB','MAR') else fy_start+1}"
        for m in FY_MONTHS
    ]
    # Normalise: APR-2025 format
    fy_month_keys = []
    for m in FY_MONTHS:
        yr = fy_start if m not in ("JAN","FEB","MAR") else fy_start+1
        fy_month_keys.append(f"{m}-{yr}")

    grand = {k:0.0 for k in ["r1","r3b","itc","ais_s","tis_s","ais_p",
                               "sales_diff","itc_diff"]}

    # ── PAN TOTALS SECTION (the key comparison — all GSTINs combined) ─────
    ri = _sec(ws, ri, f"SECTION A — PAN TOTAL (All GSTINs Combined) — {pan}  "
              f"← This is the main comparison: GST portal sum vs IT portal PAN-level", NC, bg=C_NAVY)

    pan_totals = {k:0.0 for k in grand}
    for mkey in fy_month_keys:
        gd    = pan_gst.get(mkey, {})
        r1    = gd.get("r1_taxable", 0.0)
        r3b   = gd.get("r3b_tax_total", 0.0)
        itc   = gd.get("itc_total", 0.0)
        ais_s = it_data["pan_ais_monthly"].get(mkey, {}).get("ais_sales", 0.0)
        tis_s = it_data["pan_tis_monthly"].get(mkey, {}).get("tis_sales", 0.0)
        ais_p = it_data["pan_ais_monthly"].get(mkey, {}).get("ais_pur", 0.0)
        sd    = (r1 - ais_s) if (r1 or ais_s) else None
        pd_   = (itc - ais_p) if (itc or ais_p) else None

        st, sbg, sfg = _flag_style(sd)
        pt, pbg, pfg = _flag_style(pd_)

        action = ""
        if sd is not None and abs(sd) > THRESH_OK:
            if ais_s == 0:
                action = "AIS NIL for this month — attach GSTR-1 filing proof; declare in ITR"
            elif sd > 0:
                action = "GSTR-1 > AIS — IT portal timing lag; attach GSTR-1 acknowledgement"
            else:
                action = "AIS > GSTR-1 — check GSTR-1A amendments / other GSTIN / advance"

        bg = C_LGRAY if ri % 2 == 0 else C_WHITE
        _cell(ws, ri, 1, f"PAN: {pan}", bg=C_NAVY, fg=C_WHITE, bold=True, align="center")
        _cell(ws, ri, 2, mkey, bg=bg, bold=True, align="center")
        _cell(ws, ri, 3, r1   or None, bg=bg, align="right")
        _cell(ws, ri, 4, r3b  or None, bg=bg, align="right")
        _cell(ws, ri, 5, itc  or None, bg=bg, align="right")
        _cell(ws, ri, 6, ais_s or None, bg=bg, align="right")
        _cell(ws, ri, 7, tis_s or None, bg=bg, align="right")
        _cell(ws, ri, 8, ais_p or None, bg=bg, align="right")
        _diff_cell(ws, ri, 9, sd)
        _flag_cell(ws, ri, 10, sd)
        _diff_cell(ws, ri, 11, pd_)
        _flag_cell(ws, ri, 12, pd_)
        _cell(ws, ri, 13, None, bg=bg)
        _cell(ws, ri, 14, None, bg=bg)
        _cell(ws, ri, 15, pan, bg=bg, align="center")
        _cell(ws, ri, 16, "PAN-AGG", bg=bg, align="center")
        _cell(ws, ri, 17, action, bg=sbg if action else bg,
              fg=sfg if action else "000000", align="left")
        ws.row_dimensions[ri].height = 15; ri += 1

        for k, v in [("r1",r1),("r3b",r3b),("itc",itc),("ais_s",ais_s),
                     ("tis_s",tis_s),("ais_p",ais_p)]:
            pan_totals[k] += v or 0
        if sd is not None: pan_totals["sales_diff"] += sd
        if pd_ is not None: pan_totals["itc_diff"] += pd_

    # PAN annual total row
    sd_a  = pan_totals["r1"] - pan_totals["ais_s"]
    pd_a  = pan_totals["itc"] - pan_totals["ais_p"]
    ri = _tot(ws, ri,
        [f"PAN ANNUAL TOTAL — {pan}","12 months",
         pan_totals["r1"], pan_totals["r3b"], pan_totals["itc"],
         pan_totals["ais_s"], pan_totals["tis_s"], pan_totals["ais_p"],
         sd_a, _flag_style(sd_a)[0],
         pd_a, _flag_style(pd_a)[0],
         "","",pan,"",
         f"Annual: GSTR-1 ₹{pan_totals['r1']:,.0f} vs AIS ₹{pan_totals['ais_s']:,.0f} | "
         f"Diff ₹{sd_a:,.0f}"],
        NC, bg=C_NAVY, fg=C_WHITE)
    ri += 1

    # ── GSTIN-WISE SECTION (drill-down) ──────────────────────────────────
    ri = _sec(ws, ri,
        f"SECTION B — GSTIN-WISE DRILL-DOWN  "
        f"(Same data split by GSTIN for filing reference — "
        f"{len(all_gstins)} GSTIN(s) under PAN {pan})", NC, bg=C_TEAL)

    for gstin in all_gstins:
        gstin_pan = pan_from_gstin(gstin)
        ri = _sec(ws, ri, f"GSTIN: {gstin}  (PAN: {gstin_pan})", NC, bg=C_BLUE)

        ais_gstin   = it_data["ais_monthly"].get(gstin, {})
        tis_gstin   = it_data["tis_monthly"].get(gstin, {})
        gstin_totals= {k:0.0 for k in grand}

        for mkey in fy_month_keys:
            # GST data: use PAN-level (all GSTINs combined)
            # — because GSTR-1 and ITC are at entity level not GSTIN level
            # For GSTIN-wise we show what IT portal says for THIS GSTIN
            gd    = pan_gst.get(mkey, {})
            r1_pan= gd.get("r1_taxable", 0.0)     # PAN-level GSTR-1
            r3b   = gd.get("r3b_tax_total", 0.0)
            itc   = gd.get("itc_total", 0.0)

            # IT portal data for THIS specific GSTIN
            ag    = ais_gstin.get(mkey, {})
            tg    = tis_gstin.get(mkey, {})
            ais_s = ag.get("ais_sales", tg.get("ais_total", 0.0))
            tis_s = tg.get("ais_total", ag.get("ais_sales", 0.0))
            ais_p = ag.get("ais_pur", 0.0)

            # For GSTIN-level comparison, compare GSTIN's AIS vs PAN GSTR-1
            sd  = (r1_pan - ais_s) if (r1_pan or ais_s) else None
            pd_ = (itc    - ais_p) if (itc    or ais_p) else None

            bg = C_LGRAY if ri % 2 == 0 else C_WHITE
            _cell(ws, ri, 1, gstin, bg=bg, align="center")
            _cell(ws, ri, 2, mkey,  bg=bg, align="center")
            _cell(ws, ri, 3, r1_pan or None, bg=bg, align="right")
            _cell(ws, ri, 4, r3b    or None, bg=bg, align="right")
            _cell(ws, ri, 5, itc    or None, bg=bg, align="right")
            _cell(ws, ri, 6, ais_s  or None, bg=bg, align="right")
            _cell(ws, ri, 7, tis_s  or None, bg=bg, align="right")
            _cell(ws, ri, 8, ais_p  or None, bg=bg, align="right")
            _diff_cell(ws, ri, 9,  sd)
            _flag_cell(ws, ri, 10, sd)
            _diff_cell(ws, ri, 11, pd_)
            _flag_cell(ws, ri, 12, pd_)
            _cell(ws, ri, 13, None, bg=bg)
            _cell(ws, ri, 14, None, bg=bg)
            _cell(ws, ri, 15, gstin_pan, bg=bg, align="center")
            _cell(ws, ri, 16, "GSTIN", bg=bg, align="center")
            _cell(ws, ri, 17, ag.get("status",""), bg=bg, align="left")
            ws.row_dimensions[ri].height = 15; ri += 1

            for k, v in [("r1",r1_pan),("r3b",r3b),("itc",itc),
                         ("ais_s",ais_s),("tis_s",tis_s),("ais_p",ais_p)]:
                gstin_totals[k] += v or 0

        # GSTIN subtotal
        gs_d = gstin_totals["r1"] - gstin_totals["ais_s"]
        gp_d = gstin_totals["itc"] - gstin_totals["ais_p"]
        ri = _tot(ws, ri,
            [f"Subtotal — {gstin}", "12 months",
             gstin_totals["r1"], gstin_totals["r3b"], gstin_totals["itc"],
             gstin_totals["ais_s"], gstin_totals["tis_s"], gstin_totals["ais_p"],
             gs_d, _flag_style(gs_d)[0],
             gp_d, _flag_style(gp_d)[0],
             "","",gstin_pan,"",""],
            NC, bg=C_DGRAY)
        ri += 1

    # ── SECTION C — VARIANCE SUMMARY ─────────────────────────────────────
    ri = _sec(ws, ri, "SECTION C — VARIANCE MONTHS (Action Items Before ITR Filing)", NC, bg=C_PURPLE)
    # Reprint only the months with differences
    had_variance = False
    for mkey in fy_month_keys:
        gd    = pan_gst.get(mkey, {})
        r1    = gd.get("r1_taxable", 0.0)
        itc   = gd.get("itc_total", 0.0)
        ais_s = it_data["pan_ais_monthly"].get(mkey, {}).get("ais_sales", 0.0)
        ais_p = it_data["pan_ais_monthly"].get(mkey, {}).get("ais_pur", 0.0)
        sd    = (r1 - ais_s) if (r1 or ais_s) else None
        pd_   = (itc - ais_p) if (itc or ais_p) else None

        if (sd is not None and abs(sd) > THRESH_OK) or \
           (pd_ is not None and abs(pd_) > THRESH_OK):
            had_variance = True
            bg = C_RED if (sd and abs(sd) > THRESH_WARN) else C_AMBER
            if ais_s == 0:
                action = "AIS NIL — File GSTR-1 proof; income must be declared in ITR regardless"
            elif sd and sd > 0:
                action = "GSTR-1 > AIS — Likely IT portal timing delay; provide GSTR-1 filing acknowledgement"
            else:
                action = "AIS > GSTR-1 — Check GSTR-1A amendments / other GSTIN sales / advance receipts"

            _cell(ws, ri, 1, f"PAN: {pan}", bg=C_NAVY, fg=C_WHITE, bold=True, align="center")
            _cell(ws, ri, 2, mkey, bg=bg, bold=True, align="center")
            _cell(ws, ri, 3, r1   or None, bg=bg, align="right")
            _cell(ws, ri, 4, None, bg=bg)
            _cell(ws, ri, 5, itc  or None, bg=bg, align="right")
            _cell(ws, ri, 6, ais_s or None, bg=bg, align="right")
            _cell(ws, ri, 7, None, bg=bg)
            _cell(ws, ri, 8, ais_p or None, bg=bg, align="right")
            _diff_cell(ws, ri, 9, sd)
            _flag_cell(ws, ri, 10, sd)
            _diff_cell(ws, ri, 11, pd_)
            _flag_cell(ws, ri, 12, pd_)
            _cell(ws, ri, 13, None, bg=bg)
            _cell(ws, ri, 14, None, bg=bg)
            _cell(ws, ri, 15, pan, bg=bg, align="center")
            _cell(ws, ri, 16, "ALERT", bg=bg, align="center")
            _cell(ws, ri, 17, action, bg=bg, align="left")
            ws.row_dimensions[ri].height = 15; ri += 1

    if not had_variance:
        ri = _sec(ws, ri, "✓ ALL MONTHS WITHIN TOLERANCE — No variances to explain!", NC, bg=C_DKGRN)

    ws.sheet_properties.tabColor = C_NAVY
    print(f"    ✓ PAN_Master_Comparison built: {ri} rows")


# ════════════════════════════════════════════════════════════════════════════════
# BUILD ITC_vs_AIS_Purchase SHEET
# ════════════════════════════════════════════════════════════════════════════════

def _build_itc_sheet(ws, pan_gst, it_data, fy):
    fy_start = int(fy.split("-")[0])
    pan  = it_data.get("pan","")
    name = it_data.get("name","")

    NC = 9
    COLS = [("Month",12),("GSTR-2B ITC IGST ₹",18),("GSTR-2B ITC CGST ₹",18),
            ("GSTR-2B ITC SGST ₹",18),("Total ITC ₹",18),
            ("AIS Purchases ₹",18),("Diff (ITC−AIS) ₹",18),("Flag",10),("Notes",36)]

    _title(ws, f"GSTR-2B ITC vs AIS Purchases — {name} ({pan}) — FY {fy}", NC)
    _note(ws,
        "ITC (GSTR-2B) = Tax paid by you on purchases — what you claim as credit  |  "
        "AIS Purchases = Total purchase VALUE as reported by your SUPPLIERS in their GSTR-1  |  "
        "⚠ ITC ≠ Purchase Value — ITC is only the GST portion; AIS is full value  |  "
        "Large diff is EXPECTED — explain at ITR filing with purchase register",
        NC, row=2)
    _hdr(ws, COLS, row=3)
    ws.freeze_panes = "A4"

    ann_ais_pur = it_data.get("annual_ais_pur", 0.0)
    ri = 4
    grand_igst = grand_cgst = grand_sgst = grand_itc = 0.0

    fy_month_keys = []
    for m in FY_MONTHS:
        yr = fy_start if m not in ("JAN","FEB","MAR") else fy_start+1
        fy_month_keys.append(f"{m}-{yr}")

    for i, mkey in enumerate(fy_month_keys):
        gd   = pan_gst.get(mkey, {})
        igst = gd.get("itc_igst", 0.0)
        cgst = gd.get("itc_cgst", 0.0)
        sgst = gd.get("itc_sgst", 0.0)
        itc  = gd.get("itc_total", igst+cgst+sgst)
        # AIS monthly purchase not available from IT portal — show annual total in last row
        bg = C_LGRAY if i % 2 == 0 else C_WHITE
        _cell(ws, ri, 1, mkey,  bg=bg, bold=True, align="center")
        _cell(ws, ri, 2, igst or None, bg=bg, align="right")
        _cell(ws, ri, 3, cgst or None, bg=bg, align="right")
        _cell(ws, ri, 4, sgst or None, bg=bg, align="right")
        _cell(ws, ri, 5, itc  or None, bg=bg, align="right")
        _cell(ws, ri, 6, None, bg=bg, align="right")  # AIS no monthly breakdown
        _cell(ws, ri, 7, None, bg=bg)
        _cell(ws, ri, 8, "—",  bg=bg, align="center")
        _cell(ws, ri, 9, "AIS has no monthly purchase breakdown" if itc else "", bg=bg)
        ws.row_dimensions[ri].height = 15; ri += 1
        grand_igst += igst; grand_cgst += cgst
        grand_sgst += sgst; grand_itc  += itc

    # Annual total with AIS comparison
    diff = grand_itc - ann_ais_pur if ann_ais_pur else None
    txt, bg, fg = _flag_style(diff)
    ri = _tot(ws, ri,
        ["ANNUAL TOTAL", grand_igst, grand_cgst, grand_sgst, grand_itc,
         ann_ais_pur, diff, txt,
         f"ITC ₹{grand_itc:,.0f} vs AIS Purch ₹{ann_ais_pur:,.0f}"],
        NC, bg=C_NAVY, fg=C_WHITE)
    ri += 1
    _note(ws,
        f"ℹ  AIS Purchases ₹{ann_ais_pur:,.0f} = What ALL your suppliers declared in GSTR-1 "
        f"(reported to IT dept)  |  Your GSTR-2B ITC ₹{grand_itc:,.0f} = Tax you claimed on purchases  |  "
        f"Diff ₹{(grand_itc-ann_ais_pur):,.0f} is expected — not all purchases are ITC-eligible. "
        f"Explain with purchase register at ITR filing.",
        NC, row=ri)
    ws.sheet_properties.tabColor = C_PURPLE


# ════════════════════════════════════════════════════════════════════════════════
# STEP 4 — MASTER DASHBOARD EXCEL (all companies)
# ════════════════════════════════════════════════════════════════════════════════

def build_master_excel(all_results, output_folder):
    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]

    ws_dash = wb.create_sheet("DASHBOARD")
    ws_dash.sheet_view.showGridLines = False
    _build_dashboard(ws_dash, all_results)

    for res in all_results:
        sname = re.sub(r'[\\/:"*?<>|]', "_", res["name"])[:28]
        ws_co = wb.create_sheet(sname)
        ws_co.sheet_view.showGridLines = False
        _build_company_sheet(ws_co, res)

    ts  = datetime.now().strftime("%Y%m%d_%H%M")
    out = output_folder / f"MASTER_GST_IT_RECONCILIATION_{ts}.xlsx"
    wb.save(str(out))
    print(f"\n  ✓ Master Excel: {out.name}  ({out.stat().st_size//1024} KB)")
    return str(out)


def _build_dashboard(ws, results):
    NC = 15
    _title(ws,
        f"GST ↔ Income Tax Master Dashboard — {datetime.now().strftime('%d-%b-%Y %H:%M')}  "
        f"[PAN-based comparison: GSTIN→PAN→AIS/TIS]", NC)
    COLS = [
        ("Client Name",28),("PAN",12),("GSTINs",28),("FY",8),
        ("GSTR-1 Turnover ₹",20),("AIS Turnover ₹",18),("Sales Diff ₹",16),("Sales Flag",11),
        ("GSTR-2B ITC ₹",17),("AIS Purchases ₹",17),("ITC Diff ₹",16),("Purch Flag",11),
        ("IT Updated",10),("GST Data",10),("Notes",34),
    ]
    _hdr(ws, COLS, row=2); ws.freeze_panes = "A3"

    ri = 3
    for res in results:
        bg = C_LGRAY if ri % 2 == 0 else C_WHITE
        r1   = res.get("annual_r1",0); ais   = res.get("annual_ais",0)
        itc  = res.get("annual_itc",0); ais_p = res.get("annual_ais_pur",0)
        sd   = r1-ais   if (r1 and ais)   else None
        pd_  = itc-ais_p if (itc and ais_p) else None
        st, sbg, sfg = _flag_style(sd)
        pt, pbg, pfg = _flag_style(pd_)

        row_data = [
            res.get("name",""), res.get("pan",""),
            "  |  ".join(res.get("gstins",[])),
            res.get("fy",""),
            r1, ais, sd if sd is not None else "", st,
            itc, ais_p, pd_ if pd_ is not None else "", pt,
            "✓" if res.get("it_ok") else "✗",
            "✓" if res.get("gst_ok") else "✗",
            res.get("notes",""),
        ]
        for ci, v in enumerate(row_data, 1):
            cbg = bg
            if ci == 8:  cbg = sbg
            elif ci == 12: cbg = pbg
            elif ci == 13: cbg = C_GREEN if "✓" in str(v) else C_RED
            elif ci == 14: cbg = C_GREEN if "✓" in str(v) else C_AMBER
            _cell(ws, ri, ci, v, bg=cbg,
                  align="right" if isinstance(v, float) else
                  "center" if ci in (8,12,13,14) else "left",
                  fg=sfg if ci==8 else pfg if ci==12 else "000000")
        ws.row_dimensions[ri].height = 18; ri += 1

    _tot(ws, ri,
        [f"GRAND TOTAL — {len(results)} companies","","","",
         sum(r.get("annual_r1",0)  for r in results),
         sum(r.get("annual_ais",0) for r in results),
         "","","","","","","","",""],
        NC, bg=C_NAVY, fg=C_WHITE)
    ws.sheet_properties.tabColor = C_NAVY


def _build_company_sheet(ws, res):
    fy       = res.get("fy", FALLBACK_FY)
    fy_start = int(fy.split("-")[0])
    name     = res.get("name","")
    pan      = res.get("pan","")

    NC = 12
    COLS = [
        ("Month",12),
        ("GSTR-1 Taxable ₹",20),("GSTR-3B Tax ₹",18),("GSTR-2B ITC ₹",18),
        ("AIS Sales ₹",18),("TIS Sales ₹",16),
        ("Sales Diff ₹",16),("Sales Flag",11),
        ("AIS Purchases ₹",17),("ITC−AIS Diff ₹",16),("Purch Flag",11),
        ("Action Required",44),
    ]
    _title(ws,
        f"GST ↔ IT Month-wise — {name} ({pan}) — FY {fy}  "
        f"[GSTR-1+2B (GST Portal) vs AIS/TIS (IT Portal) via PAN {pan}]", NC)
    _hdr(ws, COLS, row=2); ws.freeze_panes = "A3"; ri = 3

    monthly = res.get("monthly", {})
    tot = {k:0.0 for k in ["r1","r3b","itc","ais_s","tis_s","ais_p"]}

    for mon in FY_MONTHS:
        yr   = fy_start if mon not in ("JAN","FEB","MAR") else fy_start + 1
        mkey = f"{mon}-{yr}"
        md   = monthly.get(mkey, {})
        r1   = md.get("r1_taxable", 0.0)
        r3b  = md.get("r3b_tax_total", 0.0)
        itc  = md.get("itc_total", 0.0)
        ais_s= md.get("ais_sales", 0.0)
        tis_s= md.get("tis_sales", 0.0)
        ais_p= md.get("ais_pur",   0.0)
        sd   = (r1-ais_s)   if (r1 or ais_s)   else None
        pd_  = (itc-ais_p)  if (itc or ais_p)  else None

        action = ""
        if sd is not None and abs(sd) > THRESH_OK:
            if ais_s == 0: action = "AIS NIL — verify GSTR-3B filed; declare income in ITR"
            elif sd > 0:   action = "GSTR-1 > AIS — timing lag; attach acknowledgement"
            else:           action = "AIS > GSTR-1 — check GSTR-1A / other GSTIN / advance"

        bg = C_LGRAY if ri % 2 == 0 else C_WHITE
        _cell(ws, ri, 1, mkey, bg=bg, bold=True, align="center")
        _cell(ws, ri, 2, r1   or None, bg=bg, align="right")
        _cell(ws, ri, 3, r3b  or None, bg=bg, align="right")
        _cell(ws, ri, 4, itc  or None, bg=bg, align="right")
        _cell(ws, ri, 5, ais_s or None, bg=bg, align="right")
        _cell(ws, ri, 6, tis_s or None, bg=bg, align="right")
        _diff_cell(ws, ri, 7, sd)
        _flag_cell(ws, ri, 8, sd)
        _cell(ws, ri, 9, ais_p or None, bg=bg, align="right")
        _diff_cell(ws, ri, 10, pd_)
        _flag_cell(ws, ri, 11, pd_)
        st, sbg, _ = _flag_style(sd)
        _cell(ws, ri, 12, action, bg=sbg if action else bg, align="left")
        ws.row_dimensions[ri].height = 15; ri += 1

        for k, v in [("r1",r1),("r3b",r3b),("itc",itc),
                     ("ais_s",ais_s),("tis_s",tis_s),("ais_p",ais_p)]:
            tot[k] += v

    sd_t  = tot["r1"] - tot["ais_s"]
    pd_t  = tot["itc"] - tot["ais_p"]
    st, _, _ = _flag_style(sd_t)
    pt, _, _ = _flag_style(pd_t)
    _tot(ws, ri,
        ["ANNUAL TOTAL",tot["r1"],tot["r3b"],tot["itc"],
         tot["ais_s"],tot["tis_s"],
         sd_t, st, tot["ais_p"], pd_t, pt,
         f"R1 ₹{tot['r1']:,.0f} vs AIS ₹{tot['ais_s']:,.0f} | Diff ₹{sd_t:,.0f}"],
        NC, bg=C_NAVY, fg=C_WHITE)
    ws.sheet_properties.tabColor = C_BLUE


# ════════════════════════════════════════════════════════════════════════════════
# FOLDER HELPERS
# ════════════════════════════════════════════════════════════════════════════════

def find_latest_run_folder(base, prefix):
    """
    Find the newest run folder under base that starts with prefix.
    SPECIAL CASE: If base/MultiYear_*/ exists (gst_suite layout), expand it
    to its AY* sub-folders and return the newest one with real content.
    This fixes Issue 4: gst_suite writes to MultiYear_*/AY2025_26/ but
    master_bridge previously only looked for FY2025-26_* direct children.
    """
    base = Path(base)
    if not base.exists(): return None

    all_candidates = []

    for d in base.iterdir():
        if not d.is_dir(): continue
        if d.name.upper().startswith("MULTIYEAR"):
            # Expand: add each AY* subfolder
            for sub in d.iterdir():
                if sub.is_dir():
                    all_candidates.append(sub)
        elif d.name.upper().startswith(prefix.upper()):
            all_candidates.append(d)

    if not all_candidates:
        return None

    # Prefer folder with real data (ANNUAL_RECONCILIATION or GSTR2B files)
    def _score(d):
        has_recon  = any(d.rglob("ANNUAL_RECONCILIATION*.xlsx"))
        has_2b     = any(d.rglob("GSTR2B*.xlsx"))
        has_it     = any(d.rglob("IT_RECONCILIATION*.xlsx"))
        has_pdf    = any(d.rglob("*.pdf"))
        return (has_recon or has_it, has_2b or has_pdf, d.stat().st_mtime)

    all_candidates.sort(key=_score, reverse=True)
    return all_candidates[0]

def _is_client_folder(folder):
    folder = Path(folder)
    for pat in ["ANNUAL_RECONCILIATION*.xlsx","GSTR2B_*.xlsx",
                "IT_RECONCILIATION*.xlsx","26AS_*.pdf","AIS_*.pdf"]:
        if any(folder.glob(pat)): return True
    return False

def load_clients(script_dir, fallback_fy):
    """Read Client_Manager_Secure_AY*.xlsx or clients.xlsx"""
    for fname_pat in ["Client_Manager_Secure_AY*.xlsx",
                      "Client_Manager*.xlsx","clients.xlsx","clients_manager.xlsx"]:
        for fpath in sorted(Path(script_dir).glob(fname_pat)):
            try:
                xl = pd.ExcelFile(fpath, engine="openpyxl")
                sheet = next((s for s in xl.sheet_names
                              if any(k in s.lower()
                                     for k in ["client","credential","🔐"])),
                             xl.sheet_names[0])
                df = xl.parse(sheet, dtype=str).fillna("")
                df.columns = [re.sub(r"\s+","_",str(c).strip().lower())
                              for c in df.columns]

                def gc(*cands):
                    for c in cands:
                        if c in df.columns: return c
                    return None

                c_name  = gc("client_name","name","company_name")
                c_pan   = gc("_pan_","pan","pan_no")
                c_gstin = gc("gstin\n(15_digits)","gstin","gst_number")
                c_act   = gc("active\n(yes/no)","active","status")
                c_fy    = gc("_fy","fy","financial_year")

                clients = []
                for _, row in df.iterrows():
                    name = str(row.get(c_name,"") if c_name else "").strip()
                    if not name or name.lower() in ("nan","none",""): continue
                    act = str(row.get(c_act,"YES") if c_act else "YES").upper()
                    if act not in ("YES","Y","1","TRUE","ACTIVE",""): continue
                    pan_raw = str(row.get(c_pan,"") if c_pan else "").strip().upper()
                    gstin_raw = str(row.get(c_gstin,"") if c_gstin else "").strip()
                    gstins = [g.strip() for g in re.split(r"[,;/\n]", gstin_raw)
                              if GSTIN_RE.match(g.strip().upper())]
                    # Derive PAN from GSTIN if not given
                    if not PAN_RE.match(pan_raw) and gstins:
                        pan_raw = pan_from_gstin(gstins[0])
                    fy_raw = str(row.get(c_fy,"") if c_fy else "").strip()
                    fy = fy_raw if re.match(r"\d{4}-\d{2}", fy_raw) else fallback_fy
                    clients.append({"name":name,"pan":pan_raw,
                                    "gstin_list":gstins,"fy":fy})

                if clients:
                    print(f"  Loaded {len(clients)} client(s) from {fpath.name}")
                    return clients
            except Exception as e:
                print(f"  ⚠  {fpath.name}: {e}")
    return []


# ════════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════════

def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--gst",  default=None, help="GST output folder")
    ap.add_argument("--it",   default=None, help="IT output folder")
    ap.add_argument("--fy",   default=None, help="FY e.g. 2025-26")
    ap.add_argument("--out",  default=None, help="Output folder for Master Excel")
    args = ap.parse_args()

    script_dir = Path(os.path.dirname(os.path.abspath(__file__)))
    # Auto-detect output base: OUTPUT folder on OneDrive if present, else Downloads
    def _find_output_base_mb():
        home = Path(os.path.expanduser("~"))
        candidates = [
            home / "OneDrive" / "Desktop" / "OUTPUT",
            home / "OneDrive - Personal" / "Desktop" / "OUTPUT",
            home / "Desktop" / "OUTPUT",
            home / "Downloads",
        ]
        try:
            candidates[2:2] = [p / "Desktop" / "OUTPUT"
                                for p in home.glob("OneDrive*") if p.is_dir()]
        except Exception:
            pass
        for c in candidates:
            if c.exists():
                return c
        return home / "Downloads"
    home_dl = _find_output_base_mb()

    gst_base = Path(args.gst) if args.gst else \
               (find_latest_run_folder(home_dl/"GST_Automation","FY") or
                find_latest_run_folder(home_dl/"GST_Automation","AY") or
                home_dl/"GST_Automation")

    it_base_raw = Path(args.it) if args.it else \
                  (find_latest_run_folder(home_dl/"IT_Automation","AY") or
                   home_dl/"IT_Automation")

    # KEY FIX (Issue 2): it_suite may create a DIFFERENT AY* folder than the one
    # run_all expected. Always pick the folder containing the LARGEST (most-data)
    # IT_RECONCILIATION file across all AY* sibling folders.
    if not args.it and (home_dl/"IT_Automation").exists():
        best_it_folder = None
        best_it_size   = 0
        for d in (home_dl/"IT_Automation").iterdir():
            if not d.is_dir(): continue
            for xl in d.rglob("IT_RECONCILIATION*.xlsx"):
                try:
                    sz = xl.stat().st_size
                    if sz > best_it_size:
                        best_it_size   = sz
                        best_it_folder = d
                except: pass
        if best_it_folder and best_it_size >= 25_000:
            it_base = best_it_folder
        else:
            it_base = it_base_raw
    else:
        it_base = it_base_raw

    print("\n" + "="*72)
    print("  MASTER BRIDGE v3.0 — GST ↔ Income Tax Reconciliation")
    print("  PAN Bridge: GSTIN→PAN = same PAN as AIS/TIS in IT Portal")
    print("="*72)
    print(f"  GST folder : {gst_base}")
    print(f"  IT  folder : {it_base}")

    if not gst_base.exists():
        print(f"\n  ✗ GST folder not found: {gst_base}"); sys.exit(1)
    if not it_base.exists():
        print(f"\n  ✗ IT  folder not found: {it_base}");  sys.exit(1)

    global_fy = args.fy or FALLBACK_FY
    clients   = load_clients(script_dir, global_fy)

    gst_nested = not _is_client_folder(gst_base)
    it_nested  = not _is_client_folder(it_base)

    # If it_base IS already a client folder (contains 26AS/AIS/IT_RECON directly),
    # remember it so we can assign it directly to each matching client below.
    it_base_is_client = _is_client_folder(it_base)

    if not clients:
        # Auto-detect: each subfolder = one client
        if gst_nested:
            clients = [{"name": d.name, "pan":"","gstin_list":[],"fy":global_fy,
                        "_gst_dir":d, "_it_dir":None}
                       for d in sorted(gst_base.iterdir())
                       if d.is_dir() and not d.name.startswith(".")]
        else:
            clients = [{"name": gst_base.name, "pan":"","gstin_list":[],
                        "fy":global_fy,"_gst_dir":gst_base,"_it_dir":it_base}]

    ts     = datetime.now().strftime("%Y%m%d_%H%M")
    outdir = Path(args.out) if args.out else home_dl/"GST_IT_Bridge"/f"Run_{ts}"
    outdir.mkdir(parents=True, exist_ok=True)

    print(f"  Output     : {outdir}")
    print(f"  Clients    : {len(clients)}")
    print(f"  Thresholds : ✓ OK<₹{THRESH_OK:,}  ⚠ Minor<₹{THRESH_WARN:,}  ✗ CHECK≥\n")

    all_results = []

    for idx, client in enumerate(clients, 1):
        name   = client["name"]
        pan    = client.get("pan","")
        gstins = client.get("gstin_list",[])
        fy     = client.get("fy") or global_fy

        print(f"[{idx}/{len(clients)}] {name}  (PAN: {pan}  GSTINs: {gstins}  FY: {fy})")

        # ── Resolve GST folder ──────────────────────────────────────────
        gst_cdir = client.get("_gst_dir")
        if not gst_cdir:
            safe = re.sub(r'[\\/:"*?<>|]','_', name).replace(' ', '_')
            gst_cdir = gst_base / safe
            if not gst_cdir.exists():
                for d in gst_base.iterdir():
                    if not d.is_dir(): continue
                    du = d.name.upper()
                    if (pan and pan in du) or \
                       any(g.upper() in du for g in gstins) or \
                       safe.upper()[:6] in du:
                        gst_cdir = d; break
                else:
                    gst_cdir = gst_base  # fallback flat
        gst_cdir = Path(gst_cdir)
        # If still not found, try underscore↔space normalisation
        if not gst_cdir.exists() and gst_cdir != gst_base:
            for d in gst_base.iterdir():
                if d.is_dir() and d.name.upper().replace("_"," ") == safe.upper().replace("_"," "):
                    gst_cdir = d; break

        # ── Resolve IT folder ─────────────────────────────────────────────
        # Strategy (tried in order):
        #   1. Explicit _it_dir from caller
        #   2. Exact folder name match  (ARUN_ENTERPRISES)
        #   3. Space↔underscore normalised match
        #   4. PAN found anywhere in folder name
        #   5. Any GSTIN found anywhere in folder name
        #   6. First 6 chars of safe name (prefix match)
        #   7. Folder contains IT_RECONCILIATION*.xlsx with matching PAN inside
        #   8. Fallback: gst_cdir (single-folder layout)
        it_cdir = client.get("_it_dir")
        # Pass 0: if it_base itself is a client folder, use it directly
        if not it_cdir and it_base_is_client:
            it_cdir = it_base
        if not it_cdir:
            safe_it = re.sub(r'[\\/:"*?<>|]', '_', name).replace(' ', '_')
            it_cdir = it_base / safe_it
            if not it_cdir.exists():
                it_cdir = None
                _norm = lambda s: s.upper().replace("_"," ").replace("-"," ")
                _it_candidates = [d for d in it_base.iterdir() if d.is_dir()]
                # pass 1: exact normalised name
                for d in _it_candidates:
                    if _norm(d.name) == _norm(safe_it):
                        it_cdir = d; break
                # pass 2: PAN in folder name
                if not it_cdir and pan:
                    for d in _it_candidates:
                        if pan in d.name.upper():
                            it_cdir = d; break
                # pass 3: any GSTIN in folder name
                if not it_cdir:
                    for d in _it_candidates:
                        if any(g.upper() in d.name.upper() for g in gstins):
                            it_cdir = d; break
                # pass 4: first 6 chars of name
                if not it_cdir and len(safe_it) >= 6:
                    pfx = _norm(safe_it)[:6]
                    for d in _it_candidates:
                        if _norm(d.name).startswith(pfx):
                            it_cdir = d; break
                # pass 5: scan inside each subfolder for IT_RECONCILIATION
                #          with matching PAN in the file
                if not it_cdir and pan:
                    for d in _it_candidates:
                        for itf in d.rglob("IT_RECONCILIATION*.xlsx"):
                            if pan in itf.name.upper() or pan in d.name.upper():
                                it_cdir = d; break
                        if it_cdir: break
                # pass 6: fallback to gst_cdir (single combined folder layout)
                if not it_cdir:
                    it_cdir = gst_cdir
        it_cdir = Path(it_cdir)

        print(f"  GST folder: {gst_cdir}")
        print(f"  IT  folder: {it_cdir}")

        # ── Read GST data ───────────────────────────────────────────────
        pan_gst_all, pan_meta_all = read_gst_folder(gst_cdir, fy)

        # ── Authoritative PAN resolution ─────────────────────────────────
        # Priority:
        #   1. clients.xlsx PAN  (from IT portal — most reliable)
        #   2. PAN derived from any GSTIN in the client's list that matches IT PAN
        #   3. First PAN found in GST data
        #
        # Problem scenario: client has GSTINs 33EZEPK9321K1Z6 AND 33BENPA6909L1ZB
        #   → two different PANs in GST data; IT PAN = BENPA6909L
        #   → bridge must merge ALL GSTINs' turnover under BENPA6909L

        it_recon_path = None
        for f in sorted(it_cdir.rglob("IT_RECONCILIATION*.xlsx")):
            it_recon_path = f; break
        if not it_recon_path:
            for f in sorted(it_cdir.rglob("*.xlsx")):
                if "IT_RECON" in f.name.upper():
                    it_recon_path = f; break

        it_data = None
        if it_recon_path:
            it_data = read_it_recon(str(it_recon_path))
        else:
            print(f"  ⚠  IT_RECONCILIATION not found in {it_cdir}")

        it_pan = (it_data.get("pan") or "").strip().upper() if it_data else ""

        # Step 1: derive expected PAN from each known GSTIN
        gstin_to_pan = {}
        for g in gstins:
            if len(g) >= 12:
                gstin_to_pan[g] = g[2:12].upper()

        # Step 2: pick authoritative PAN
        #   — prefer clients.xlsx PAN if it appears in GST data OR in GSTIN list
        #   — if not, prefer the IT portal PAN
        #   — fall back to first GST PAN
        auth_pan = ""
        if pan and pan in pan_meta_all:
            auth_pan = pan                          # exact match in GST data
        elif it_pan and it_pan in pan_meta_all:
            auth_pan = it_pan                       # IT PAN found in GST data
        elif pan and pan in gstin_to_pan.values():
            auth_pan = pan                          # GSTIN confirms clients.xlsx PAN
        elif it_pan and it_pan in gstin_to_pan.values():
            auth_pan = it_pan                       # GSTIN confirms IT PAN
        elif pan_meta_all:
            auth_pan = list(pan_meta_all.keys())[0] # fallback: first GST PAN
        pan = auth_pan or pan or it_pan

        if it_pan and auth_pan and it_pan != auth_pan:
            print(f"  ℹ  PAN resolution: clients.xlsx/GSTIN→{auth_pan}  "
                  f"IT portal→{it_pan}  Using {auth_pan} as master key")

        # Step 3: MERGE all GST PANs that belong to this client.
        # If client has GSTINs under multiple PANs (e.g. EZEPK9321K + BENPA6909L)
        # we sum all their monthly data into the auth_pan bucket.
        pan_gst  = {}
        pan_meta = {}
        if pan_meta_all:
            # Collect every PAN that appears in the client's GSTIN list
            client_pans = set(gstin_to_pan.values()) | {auth_pan}
            if it_pan:
                client_pans.add(it_pan)
            # KEY FIX: If the GST folder only has ONE client's data (single
            # ANNUAL_RECONCILIATION file), include ALL PANs found in that folder.
            # This handles the case where clients.xlsx GSTIN was wrong (e.g.
            # 33EZEPK9321K1Z6) but the actual ANNUAL_RECONCILIATION was built
            # under that GSTIN — we should still merge that data for this client.
            if len(pan_meta_all) == 1:
                client_pans.update(pan_meta_all.keys())
            # Also include any PAN whose GSTINs overlap with our known GSTINs
            for gst_pan, gst_meta in pan_meta_all.items():
                folder_gstins = set(g.upper() for g in gst_meta.get("_gstins", []))
                known_gstins  = set(g.upper() for g in gstins)
                if folder_gstins & known_gstins:   # any overlap
                    client_pans.add(gst_pan)

            for cpan in client_pans:
                if cpan not in pan_gst_all:
                    continue
                for mkey, mdata in pan_gst_all[cpan].items():
                    if mkey not in pan_gst:
                        pan_gst[mkey] = dict(mdata)
                    else:
                        # Sum numeric fields
                        for k, v in mdata.items():
                            if isinstance(v, (int, float)):
                                pan_gst[mkey][k] = pan_gst[mkey].get(k, 0.0) + v

                cm = pan_meta_all[cpan]
                if not pan_meta:
                    pan_meta = dict(cm)
                else:
                    pan_meta["_annual_r1"]  = pan_meta.get("_annual_r1",  0.0) + cm.get("_annual_r1",  0.0)
                    pan_meta["_annual_itc"] = pan_meta.get("_annual_itc", 0.0) + cm.get("_annual_itc", 0.0)
                    pan_meta["_gstins"]     = list(set(pan_meta.get("_gstins",[]) + cm.get("_gstins",[])))

            if not pan_gst:
                # Fallback: just use whatever GST data we have
                pan_gst  = dict(pan_gst_all.get(pan, {}))
                pan_meta = pan_meta_all.get(pan, {})

        all_gstins = pan_meta.get("_gstins", gstins)

        # ── Align IT data PAN to authoritative PAN ───────────────────────
        if it_data and not it_data.get("pan"):
            it_data["pan"] = pan
        if it_data and not it_data.get("name"):
            it_data["name"] = name
        if it_data:
            it_data["_it_pan"] = it_data.get("pan", pan)
            it_data["pan"]     = pan    # always use resolved auth_pan for joins

        # ── Inject GST data into IT Recon Excel ────────────────────────
        it_ok = False
        if it_recon_path and it_data:
            it_ok = inject_into_it_recon(
                str(it_recon_path), pan_gst, pan_meta, it_data, fy)

        # ── Merge monthly data for Master Excel ─────────────────────────
        fy_start = int(fy.split("-")[0])
        fy_month_keys = [
            f"{'APR' if m not in ('JAN','FEB','MAR') else m}-"
            f"{fy_start if m not in ('JAN','FEB','MAR') else fy_start+1}"
            for m in FY_MONTHS
        ]
        merged_monthly = {}
        for m in FY_MONTHS:
            yr   = fy_start if m not in ("JAN","FEB","MAR") else fy_start+1
            mkey = f"{m}-{yr}"
            gd   = pan_gst.get(mkey, {})
            ais_m= it_data["pan_ais_monthly"].get(mkey, {}) if it_data else {}
            tis_m= it_data["pan_tis_monthly"].get(mkey, {}) if it_data else {}
            merged_monthly[mkey] = {
                "r1_taxable":    gd.get("r1_taxable", 0.0),
                "r3b_tax_total": gd.get("r3b_tax_total", 0.0),
                "itc_total":     gd.get("itc_total", 0.0),
                "ais_sales":     ais_m.get("ais_sales", 0.0),
                "tis_sales":     tis_m.get("tis_sales", 0.0),
                "ais_pur":       ais_m.get("ais_pur", 0.0),
            }

        annual_r1  = pan_meta.get("_annual_r1", 0.0)
        annual_itc = pan_meta.get("_annual_itc", 0.0)
        annual_ais = it_data.get("annual_ais_sales", 0.0) if it_data else 0.0
        annual_pur = it_data.get("annual_ais_pur",   0.0) if it_data else 0.0

        all_results.append({
            "name":        name,
            "pan":         pan,
            "gstins":      all_gstins,
            "fy":          fy,
            "annual_r1":   annual_r1,
            "annual_ais":  annual_ais,
            "annual_itc":  annual_itc,
            "annual_ais_pur": annual_pur,
            "monthly":     merged_monthly,
            "it_ok":       it_ok,
            "gst_ok":      bool(pan_gst),
            "notes":       f"GSTINs: {','.join(all_gstins)}" if all_gstins else "No GST data",
        })
        print()

    print("  Building Master Reconciliation Excel...")
    master = build_master_excel(all_results, outdir)

    ok_it  = sum(1 for r in all_results if r["it_ok"])
    ok_gst = sum(1 for r in all_results if r["gst_ok"])

    print("\n" + "="*72)
    print("  MASTER BRIDGE v3.0 — COMPLETE")
    print("="*72)
    print(f"  Clients processed  : {len(all_results)}")
    print(f"  IT Recons updated  : {ok_it}/{len(all_results)}")
    print(f"  GST data found     : {ok_gst}/{len(all_results)}")
    print(f"  Master Excel       : {Path(master).name}")
    print(f"  Output folder      : {outdir}")
    print("="*72)
    print()
    print("  KEY — HOW PAN BRIDGE WORKS:")
    print("  GSTIN 37BENPA6909L1Z3  →  PAN = chars[2:12] = BENPA6909L")
    print("  GSTIN 33BENPA6909L1ZB  →  PAN = chars[2:12] = BENPA6909L  (same PAN!)")
    print("  AIS/TIS in IT portal shows PAN BENPA6909L → matches BOTH GSTINs")
    print("  All GSTR-1 sales are summed across all GSTINs → compared to AIS/TIS")
    print("="*72)


if __name__ == "__main__":
    main()
