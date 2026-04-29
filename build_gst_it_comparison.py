"""
GST_IT_COMPARISON_BUILDER — v2.0  (FULLY AUTOMATIC)
=====================================================
AUTO-READS all data — no manual entry needed:
  ✅ GSTR-2B  → Reads all GSTR2B_*.xlsx monthly files automatically
  ✅ AIS PDF  → Reads AIS_*.pdf and extracts purchases by supplier & month
  ✅ TIS PDF  → Reads TIS_*.pdf and extracts accepted GST purchase values

Creates TWO Excel workbooks:
  FILE 1: GSTR2B_EXTRACT.xlsx
    Sheet 1 — 2B_Raw_Input     : All GSTR-2B data (auto-imported from 12 monthly files)
    Sheet 2 — 2B_By_Supplier   : Supplier totals (all months) for TIS comparison
    Sheet 3 — 2B_MonthWise     : Supplier × Month totals for AIS comparison

  FILE 2: TIS_AIS_COMPARISON.xlsx
    Sheet 1 — TIS_vs_2B        : 2B totals vs TIS accepted values (auto-filled)
    Sheet 2 — AIS_vs_2B        : Month-wise 2B vs AIS (auto-filled)
    Sheet 3 — Summary_Dashboard: Annual reconciliation summary

Usage:
  python build_gst_it_comparison.py
  python build_gst_it_comparison.py --gst-folder "C:/Downloads/FT6/OUTPUT/GST_Automation/..."
  python build_gst_it_comparison.py --out "C:/MyOutput"
"""

import os, re, glob, pathlib, warnings
from datetime import datetime
from collections import defaultdict

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings("ignore")

# ── Colours ─────────────────────────────────────────────────────────
NAVY    = "1F3864"; BLUE  = "2E75B6"; TEAL  = "1D6A72"
WHITE   = "FFFFFF"; LGRAY = "F2F2F2"; DGRAY = "D6DCE4"
GREEN   = "C6EFCE"; DKGRN = "276221"
AMBER   = "FFEB9C"; DKAMB = "9C6500"
RED_BG  = "FFC7CE"; DKRED = "9C0006"
YELLOW  = "FFF2CC"; ORANGE= "FCE4D6"
ALT1    = "FFFFFF"; ALT2  = "F2F2F2"
INPUT_BG= "EBF3FB"; AUTO_BG= "E2EFDA"   # green = auto-filled
NUM_FMT = "#,##0.00"; INT_FMT = "#,##0"

MONTHS_FY   = ["Apr-2025","May-2025","Jun-2025","Jul-2025","Aug-2025","Sep-2025",
                "Oct-2025","Nov-2025","Dec-2025","Jan-2026","Feb-2026","Mar-2026"]
MONTHS_SHORT= [("Apr","2025"),("May","2025"),("Jun","2025"),("Jul","2025"),
               ("Aug","2025"),("Sep","2025"),("Oct","2025"),("Nov","2025"),
               ("Dec","2025"),("Jan","2026"),("Feb","2026"),("Mar","2026")]

# Month name → abbreviation map (for PDF parsing)
MON_MAP = {
    "january":"Jan","february":"Feb","march":"Mar","april":"Apr",
    "may":"May","june":"Jun","july":"Jul","august":"Aug",
    "september":"Sep","october":"Oct","november":"Nov","december":"Dec",
    "jan":"Jan","feb":"Feb","mar":"Mar","apr":"Apr",
    "jun":"Jun","jul":"Jul","aug":"Aug","sep":"Sep",
    "oct":"Oct","nov":"Nov","dec":"Dec",
}

# ── Style helpers ────────────────────────────────────────────────────
def _f(h):  return PatternFill("solid", fgColor=h)
def _fn(b=False, c="000000", s=9): return Font(name="Arial", bold=b, color=c, size=s)
def _bd():
    x = Side(style="thin"); return Border(left=x, right=x, top=x, bottom=x)
def _al(h="left", w=False): return Alignment(horizontal=h, vertical="center", wrap_text=w)

def _c(ws, r, col, v, bg=ALT1, bold=False, fg="000000", align="left", numfmt=None, size=9):
    c = ws.cell(row=r, column=col, value=v)
    c.font = _fn(bold, fg, size)
    c.fill = _f(bg)
    wrap = (align == "left" and isinstance(v, str) and len(str(v)) > 30)
    c.alignment = _al(align, wrap)
    c.border = _bd()
    if numfmt and (isinstance(v, (int, float)) or (isinstance(v, str) and v.startswith("="))):
        c.number_format = numfmt
    return c

def _title(ws, txt, nc, bg=NAVY, size=11):
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    c = ws["A1"]
    c.value = txt
    c.font  = _fn(True, WHITE, size)
    c.fill  = _f(bg)
    c.alignment = _al("center")
    c.border = _bd()
    ws.row_dimensions[1].height = 28

def _hdr(ws, cols, row=2, bg=NAVY):
    for ci, (h, w) in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _fn(True, WHITE, 9)
        c.fill = _f(bg)
        c.alignment = _al("center", w=True)
        c.border = _bd()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 22

def _sep(ws, r, lbl, nc, bg=BLUE):
    ws.merge_cells(f"A{r}:{get_column_letter(nc)}{r}")
    c = ws.cell(row=r, column=1, value=lbl)
    c.font = _fn(True, WHITE, 9)
    c.fill = _f(bg)
    c.alignment = _al("left")
    c.border = _bd()
    ws.row_dimensions[r].height = 18

def _info(ws, r, nc, txt, bg="E2EFDA", fc="375623"):
    ws.merge_cells(f"A{r}:{get_column_letter(nc)}{r}")
    c = ws.cell(row=r, column=1, value=txt)
    c.font = _fn(False, fc, 9)
    c.fill = _f(bg)
    c.alignment = _al("left", w=True)
    c.border = _bd()
    ws.row_dimensions[r].height = 30


# ═══════════════════════════════════════════════════════════════════
# DATA EXTRACTION ENGINES
# ═══════════════════════════════════════════════════════════════════

class GSTRData:
    """Container for all extracted data."""
    def __init__(self):
        # rows: list of dict {gstin, name, month, taxable, igst, cgst, sgst, source}
        self.b2b_rows     = []
        # tis_data: {PAN: {name, accepted}}   — TIS uses PAN not GSTIN
        self.tis_data     = {}
        # ais_data: {supplier_GSTIN: {name, pan, months: {mon_str: amount}}}
        self.ais_data     = {}
        self.warnings     = []

    @staticmethod
    def pan_from_gstin(gstin):
        """Extract 10-char PAN from 15-char GSTIN (chars 3–12, 0-indexed 2:12)."""
        g = str(gstin).strip().upper()
        return g[2:12] if len(g) == 15 else ""

    def get_tis(self, gstin):
        """Lookup TIS data by GSTIN — matches via PAN extracted from GSTIN."""
        pan = self.pan_from_gstin(gstin)
        return self.tis_data.get(pan) or self.tis_data.get(gstin) or {}

    def get_ais_months(self, gstin):
        """Lookup AIS month-data by GSTIN, with fallback to PAN-based lookup."""
        # Direct GSTIN match
        if gstin in self.ais_data:
            return self.ais_data[gstin].get("months", {})
        # Try matching any AIS entry whose PAN matches this GSTIN's PAN
        pan = self.pan_from_gstin(gstin)
        if pan:
            for k, v in self.ais_data.items():
                if v.get("pan") == pan:
                    return v.get("months", {})
        return {}

    # ── Derived aggregates ──────────────────────────────────────────

    def supplier_totals(self):
        """Returns dict: gstin → {name, taxable, igst, cgst, sgst, total}"""
        totals = defaultdict(lambda: {"name":"", "taxable":0.0, "igst":0.0, "cgst":0.0, "sgst":0.0})
        for row in self.b2b_rows:
            g = row["gstin"]
            totals[g]["name"]    = row["name"]
            totals[g]["taxable"] += row.get("taxable", 0) or 0
            totals[g]["igst"]    += row.get("igst",    0) or 0
            totals[g]["cgst"]    += row.get("cgst",    0) or 0
            totals[g]["sgst"]    += row.get("sgst",    0) or 0
        for g, d in totals.items():
            d["total"] = d["igst"] + d["cgst"] + d["sgst"]
        return dict(totals)

    def supplier_month_totals(self):
        """Returns dict: (gstin, month_str) → {name, taxable, igst, cgst, sgst, total}"""
        totals = defaultdict(lambda: {"name":"", "taxable":0.0, "igst":0.0, "cgst":0.0, "sgst":0.0})
        for row in self.b2b_rows:
            key = (row["gstin"], row["month"])
            totals[key]["name"]    = row["name"]
            totals[key]["taxable"] += row.get("taxable", 0) or 0
            totals[key]["igst"]    += row.get("igst",    0) or 0
            totals[key]["cgst"]    += row.get("cgst",    0) or 0
            totals[key]["sgst"]    += row.get("sgst",    0) or 0
        for k, d in totals.items():
            d["total"] = d["igst"] + d["cgst"] + d["sgst"]
        return dict(totals)

    def all_suppliers(self):
        """Returns sorted list of (gstin, name)."""
        seen = {}
        for row in self.b2b_rows:
            if row["gstin"] not in seen:
                seen[row["gstin"]] = row["name"]
        return sorted(seen.items(), key=lambda x: x[1])


# ── GSTR-2B Excel reader ────────────────────────────────────────────

def _safe_float(v):
    if v is None: return 0.0
    try:    return float(str(v).replace(",","").strip()) if str(v).strip() else 0.0
    except: return 0.0

GSTIN_RE = re.compile(r'\b\d{2}[A-Z]{5}\d{4}[A-Z]{1}[A-Z\d]{1}[Z]{1}[A-Z\d]{1}\b')

def _month_from_filename(fname):
    """Guess month string like 'Apr-2025' from filename like GSTR2B_April_2025.xlsx"""
    fname_low = fname.lower()
    for full, abbr in [
        ("april","Apr"),("may","May"),("june","Jun"),("july","Jul"),
        ("august","Aug"),("september","Sep"),("october","Oct"),
        ("november","Nov"),("december","Dec"),("january","Jan"),
        ("february","Feb"),("march","Mar"),
    ]:
        if full in fname_low:
            for yr in ["2025","2026"]:
                if yr in fname:
                    return f"{abbr}-{yr}"
            # guess year from FY
            return f"{abbr}-2025"
    return None

def _find_header_row(ws, keywords=("gstin","supplier","trade","party","taxable")):
    """Find the row index containing column headers."""
    for row in ws.iter_rows(min_row=1, max_row=30):
        vals = " ".join(str(c.value or "").lower() for c in row)
        if sum(1 for k in keywords if k in vals) >= 2:
            return row[0].row
    return None

def read_gstr2b_excel(filepath, month_str):
    """
    Read one GSTR-2B Excel file and return list of row dicts.
    Handles both GSTIN portal format (B2B section) and simple tabular formats.
    """
    rows = []
    fname = os.path.basename(filepath)
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return rows, [f"Cannot open {fname}: {e}"]
    
    warnings_out = []
    for shname in wb.sheetnames:
        ws = wb[shname]
        # Skip sheets that are clearly not B2B data
        if any(x in shname.lower() for x in ["cdnr","cdna","impg","impgs","ecom","b2cs","nil","hsn","docs"]):
            continue

        # Find header row
        hdr_row = _find_header_row(ws)
        if hdr_row is None:
            continue

        # Map column names → column index
        hdr_cells = list(ws.iter_rows(min_row=hdr_row, max_row=hdr_row))[0]
        col_map = {}
        for c in hdr_cells:
            val = str(c.value or "").lower().strip()
            if not val: continue
            if any(x in val for x in ["gstin","gst in"]): col_map["gstin"] = c.column
            elif any(x in val for x in ["trade name","supplier name","party name","legal name","name"]): col_map.setdefault("name", c.column)
            elif "taxable" in val or "taxable value" in val: col_map["taxable"] = c.column
            elif val in ("igst","igst amount","integrated tax"): col_map["igst"] = c.column
            elif val in ("cgst","cgst amount","central tax"): col_map["cgst"] = c.column
            elif val in ("sgst","sgst amount","state/ut tax","utgst"): col_map["sgst"] = c.column

        if "gstin" not in col_map:
            continue

        # Read data rows
        found_any = False
        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            gstin_val = str(row[col_map["gstin"] - 1] or "").strip().upper()
            if not GSTIN_RE.match(gstin_val):
                continue
            name  = str(row[col_map["name"]    - 1] or "").strip() if "name"    in col_map else gstin_val
            taxable = _safe_float(row[col_map["taxable"] - 1]) if "taxable" in col_map else 0.0
            igst    = _safe_float(row[col_map["igst"]    - 1]) if "igst"    in col_map else 0.0
            cgst    = _safe_float(row[col_map["cgst"]    - 1]) if "cgst"    in col_map else 0.0
            sgst    = _safe_float(row[col_map["sgst"]    - 1]) if "sgst"    in col_map else 0.0
            rows.append({
                "gstin":   gstin_val,
                "name":    name,
                "month":   month_str,
                "taxable": taxable,
                "igst":    igst,
                "cgst":    cgst,
                "sgst":    sgst,
                "source":  fname,
            })
            found_any = True

        if not found_any:
            # Try reading the sheet as a GST portal export where GSTIN appears
            # in column A-style (grouped under a GSTIN header cell)
            # and subsequent rows have invoice details
            current_gstin = None
            current_name  = None
            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row: continue
                first = str(row[0] or "").strip().upper()
                if GSTIN_RE.match(first):
                    current_gstin = first
                    # Name is usually next cell or same row col B
                    current_name = str(row[1] or "").strip() if len(row) > 1 else first
                    continue
                if current_gstin and any(isinstance(c, (int, float)) for c in row):
                    # Invoice detail row — accumulate into one row per gstin×month
                    vals = [_safe_float(c) for c in row]
                    # Heuristic: pick 4 largest numbers as taxable, igst, cgst, sgst
                    # This is imprecise but handles diverse portal layouts
                    nums = sorted([v for v in vals if v > 0], reverse=True)
                    if len(nums) >= 1:
                        rows.append({
                            "gstin":   current_gstin,
                            "name":    current_name or current_gstin,
                            "month":   month_str,
                            "taxable": nums[0] if len(nums)>0 else 0,
                            "igst":    nums[1] if len(nums)>1 else 0,
                            "cgst":    nums[2] if len(nums)>2 else 0,
                            "sgst":    nums[3] if len(nums)>3 else 0,
                            "source":  fname,
                        })
    wb.close()
    return rows, warnings_out


def find_and_read_gstr2b_files(search_paths, exclude_paths=None):
    """
    Search for GSTR2B_*.xlsx files in all given paths.
    Returns GSTRData with b2b_rows filled.
    Deduplicates: (1) by normalized absolute path, (2) by basename, (3) by month string.
    exclude_paths: iterable of directory paths to skip entirely.
    """
    data = GSTRData()

    # Build exclusion set
    _excl_norm = set()
    if exclude_paths:
        for ep in exclude_paths:
            try: _excl_norm.add(os.path.normcase(os.path.normpath(str(ep))))
            except: pass

    def _is_excl(fp):
        try:
            np = os.path.normcase(os.path.normpath(fp))
            for ex in _excl_norm:
                if np == ex or np.startswith(ex + os.sep):
                    return True
        except: pass
        return False

    all_files = []
    for sp in search_paths:
        if not sp or not os.path.isdir(sp): continue
        if _is_excl(sp): continue
        # Use only the precise GSTR2B pattern (avoid over-broad "*2B*" which picks up unrelated files)
        for pattern in ["GSTR2B*.xlsx", "GSTR-2B*.xlsx", "gstr2b*.xlsx"]:
            for f in glob.glob(os.path.join(sp, "**", pattern), recursive=True):
                if not _is_excl(os.path.dirname(f)):
                    all_files.append(f)
            for f in glob.glob(os.path.join(sp, pattern)):
                if not _is_excl(os.path.dirname(f)):
                    all_files.append(f)

    # Step 1 — deduplicate by normalised absolute path
    seen_real = set()
    deduped = []
    for f in all_files:
        try: nf = os.path.normcase(os.path.normpath(os.path.abspath(f)))
        except: nf = f
        if nf not in seen_real:
            seen_real.add(nf)
            deduped.append(f)
    all_files = deduped

    # Step 2 — skip temp files, output files, and consolidated analysis files
    SKIP_EXACT = {"gstr2b_extract.xlsx", "tis_ais_comparison.xlsx"}
    seen_names = {}
    for f in sorted(all_files):
        bname = os.path.basename(f)
        bname_low = bname.lower()
        if bname.startswith("~$"):
            continue  # Excel temp file
        if bname_low in SKIP_EXACT:
            data.warnings.append(f"  ⚠ Skipping output file: {bname}")
            continue
        # Skip consolidated analysis files — only raw GSTR2B_MonthName_YYYY.xlsx are valid
        import re as _re2
        if not _re2.match(r"gstr2b_[a-z]+_\d{4}\.xlsx$", bname_low):
            data.warnings.append(f"  ⚠ Skipping non-raw file (not GSTR2B_Month_YYYY): {bname}")
            continue
        # Keep only one file per basename — prefer the file in the most recently modified folder
        if bname_low not in seen_names:
            seen_names[bname_low] = f
        else:
            # Replace if this file is in a newer folder (later mtime = current run's client)
            try:
                existing_folder_mtime = os.path.getmtime(os.path.dirname(seen_names[bname_low]))
                this_folder_mtime = os.path.getmtime(os.path.dirname(f))
                if this_folder_mtime > existing_folder_mtime:
                    seen_names[bname_low] = f
            except: pass
    all_files = sorted(seen_names.values())

    if not all_files:
        data.warnings.append("⚠ No GSTR2B Excel files found — searched: " + ", ".join(str(p) for p in search_paths if p))
        return data

    print(f"  Found {len(all_files)} GSTR-2B Excel file(s)")
    seen_months = set()
    for fpath in all_files:
        fname = os.path.basename(fpath)
        month_str = _month_from_filename(fname)
        if not month_str:
            data.warnings.append(f"  ⚠ Cannot determine month for: {fname} — skipped")
            continue
        if month_str in seen_months:
            data.warnings.append(f"  ⚠ Duplicate month {month_str} in {fname} — skipped")
            continue
        seen_months.add(month_str)
        print(f"    Reading {fname} → {month_str}")
        rows, warns = read_gstr2b_excel(fpath, month_str)
        data.b2b_rows.extend(rows)
        data.warnings.extend(warns)
        print(f"      → {len(rows)} supplier-invoice rows extracted")

    print(f"  Total B2B rows: {len(data.b2b_rows)}")
    return data


# ── TIS PDF reader ──────────────────────────────────────────────────

def read_tis_pdf(pdf_path):
    """
    Parse TIS PDF (IT portal) — GST Purchases section.
    TIS uses PAN (10-char) NOT GSTIN as the supplier identifier.
    Tries multiple parsing strategies to handle varied PDF layouts.
    Returns dict: {pan: {name, accepted}}
    """
    PAN_RE2  = re.compile(r'\b([A-Z]{5}[0-9]{4}[A-Z])\b')
    NUM_RE2  = re.compile(r'([\d,]+(?:\.\d+)?)')
    result   = {}
    warns    = []
    if not pdf_path or not os.path.exists(pdf_path):
        return result, [f"TIS PDF not found: {pdf_path}"]

    print(f"  Reading TIS PDF: {os.path.basename(pdf_path)}")
    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            lines = [l.strip() for l in full_text.split("\n") if l.strip()]

        # ── PASS 1: section-aware scan (standard TIS layout) ─────────
        #   Looks for "GST Purchases" header, then lines with "Purchase from"
        in_gst_pur = False
        for line in lines:
            ll = line.lower()
            # Section start triggers
            if any(k in ll for k in ["gst purchases", "gst purchase", "purchases (gst)"]):
                in_gst_pur = True
            # Section end triggers
            if in_gst_pur and any(k in ll for k in [
                "income from salary", "tds on salary", "interest income",
                "dividend", "capital gains", "other income"
            ]):
                in_gst_pur = False

            if in_gst_pur and "purchase from" in ll:
                pm = PAN_RE2.search(line)
                if not pm:
                    continue
                pan = pm.group(1)
                # Supplier name: text before PAN parenthesis (if any), or before first number
                before_pan = line[:line.upper().find(pan)].strip()
                # Remove leading serial + keywords
                name_clean = re.sub(
                    r'^\d+\s*(Other\s+)?Purchases?\s+reported\s+(under\s+GSTR-1\s+of\s+seller\s+)?',
                    '', before_pan, flags=re.IGNORECASE).strip()
                # Remove trailing "Purchase from" if it crept in
                name_clean = re.sub(r'\s*Purchase from.*$', '', name_clean,
                                    flags=re.IGNORECASE).strip()
                after = line[line.lower().index("purchase from"):]
                nums = []
                for n in NUM_RE2.findall(after):
                    try:
                        v = float(n.replace(",", ""))
                        if v > 0 and not (v < 10000 and v == int(v) and len(str(int(v))) == 4):
                            nums.append(v)
                    except: pass
                if not nums:
                    continue
                accepted = nums[-1]
                if pan not in result or accepted > 0:
                    result[pan] = {"name": name_clean or pan, "accepted": accepted}

        # ── PASS 2: global PAN + "Purchase from" scan (no section needed) ─
        #   Handles PDFs where section header is missing or different
        if not result:
            for line in lines:
                if "purchase from" not in line.lower():
                    continue
                pm = PAN_RE2.search(line)
                if not pm:
                    continue
                pan = pm.group(1)
                # Skip if this PAN looks like a GSTIN embedded (unlikely in TIS but safe)
                after = line[line.lower().index("purchase from"):]
                nums = []
                for n in NUM_RE2.findall(after):
                    try:
                        v = float(n.replace(",", ""))
                        if v > 0 and not (v < 10000 and v == int(v) and len(str(int(v))) == 4):
                            nums.append(v)
                    except: pass
                if not nums:
                    continue
                before_pan = line[:line.upper().find(pan)].strip()
                name_clean = re.sub(
                    r'^\d+\s*(Other\s+)?Purchases?\s+reported\s+(under\s+GSTR-1\s+of\s+seller\s+)?',
                    '', before_pan, flags=re.IGNORECASE).strip()
                name_clean = re.sub(r'\s*Purchase from.*$', '', name_clean,
                                    flags=re.IGNORECASE).strip()
                accepted = nums[-1]
                if pan not in result or accepted > 0:
                    result[pan] = {"name": name_clean or pan, "accepted": accepted}

        # ── PASS 3: PAN + numeric pattern (no "purchase from" keyword) ──
        #   Handles highly condensed TIS PDFs
        if not result:
            # Look for lines that match: <serial> <name> (<PAN>) <numbers>
            SUPPLIER_RE = re.compile(
                r'\d+\s+(.+?)\s+\(([A-Z]{5}[0-9]{4}[A-Z])\)\s+([\d,\.]+(?:\s+[\d,\.]+)*)',
                re.IGNORECASE
            )
            for line in lines:
                m = SUPPLIER_RE.search(line)
                if not m:
                    continue
                name_raw, pan, nums_str = m.group(1), m.group(2), m.group(3)
                name_clean = re.sub(
                    r'(Other\s+)?Purchases?\s+reported\s+(under\s+GSTR-1\s+of\s+seller\s+)?',
                    '', name_raw, flags=re.IGNORECASE).strip()
                nums = []
                for n in NUM_RE2.findall(nums_str):
                    try:
                        v = float(n.replace(",", ""))
                        if v > 0 and not (v < 10000 and v == int(v) and len(str(int(v))) <= 4):
                            nums.append(v)
                    except: pass
                if nums:
                    accepted = nums[-1]
                    if pan not in result or accepted > 0:
                        result[pan] = {"name": name_clean or pan, "accepted": accepted}

    except Exception as e:
        warns.append(f"TIS PDF parse error: {e}")

    print(f"    → {len(result)} supplier(s) found in TIS")
    if not result:
        warns.append(
            "⚠ TIS PDF: Could not extract supplier data — "
            "PDF may be scanned/image-based or use an unsupported layout. "
            "Please enter TIS values manually in the blue cells."
        )
    return result, warns


# ── AIS PDF reader ──────────────────────────────────────────────────

def read_ais_pdf(pdf_path):
    """
    Parse AIS PDF — EXC-GSTR1(P) section (Purchases reported under GSTR-1 of seller).
    Each supplier block starts with an EXC-GSTR1(P) header line containing PAN:
      e.g.  "3 EXC-GSTR1(P)  Purchases reported under GSTR-1 of seller  SRI RK ASSOCIATES (ADRFS6163J)  8  32,41,528"
    Detail rows: BUYER_GSTIN  SUPPLIER_NAME (SUPPLIER_GSTIN)  MON-YYYY  AMOUNT  Active
    Returns dict keyed by supplier GSTIN:
      {gstin: {name, pan, months: {mon_str: amount}}}
    Also creates a parallel PAN-keyed dict for TIS matching.
    """
    MON_MAP2  = {'JAN':'Jan','FEB':'Feb','MAR':'Mar','APR':'Apr','MAY':'May','JUN':'Jun',
                 'JUL':'Jul','AUG':'Aug','SEP':'Sep','OCT':'Oct','NOV':'Nov','DEC':'Dec'}
    PAN_RE2   = re.compile(r'\(([A-Z]{5}[0-9]{4}[A-Z])\)')
    GSTIN_RE2 = re.compile(r'\b(\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]Z[A-Z0-9])\b')
    MON_RE2   = re.compile(r'\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)-(20\d\d)\b')
    NUM_RE2   = re.compile(r'([\d,]+(?:\.\d+)?)')

    result  = defaultdict(lambda: {"name": "", "pan": "", "months": defaultdict(float)})
    warns   = []
    if not pdf_path or not os.path.exists(pdf_path):
        return result, [f"AIS PDF not found: {pdf_path}"]

    print(f"  Reading AIS PDF: {os.path.basename(pdf_path)}")
    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_pages_text = []
            all_tables     = []
            for page in pdf.pages:
                t = page.extract_text() or ""
                all_pages_text.append(t)
                tbls = page.extract_tables() or []
                all_tables.extend(tbls)

        full_text = "\n".join(all_pages_text)

        # ── PASS 1: Text-mode line scanner (handles plain-text AIS PDFs) ──────
        # State machine: detect EXC-GSTR1/EXC-GSTR1(P) header → collect detail rows
        current_pan   = None
        current_name  = None
        current_gstin = None

        lines = [l.strip() for l in full_text.split("\n") if l.strip()]
        for line in lines:
            # ── Detect new supplier block header ─────────────────────────
            # AIS uses both "EXC-GSTR1(P)" and "EXC-GSTR1" (without (P))
            # Both signal: Purchases reported under GSTR-1 of seller
            is_header = (
                ("EXC-GSTR1" in line and
                 any(k in line for k in ["Purchase", "GSTR-1 of seller", "Purchases"]))
                or
                ("EXC-GSTR1(P)" in line)
            )
            if is_header:
                pm = PAN_RE2.search(line)
                if pm:
                    current_pan   = pm.group(1)
                    m2 = re.search(r'of seller\s+(.+?)\s+\([A-Z]{5}\d{4}[A-Z]\)', line)
                    current_name  = m2.group(1).strip() if m2 else current_pan
                    current_gstin = None
                else:
                    # Header found but no PAN yet — wait for next line with PAN
                    current_pan = None
                continue

            # Reset on any other EXC-/SFT-/TDS- code that is NOT GSTR1 purchase
            if re.search(r'\b(EXC-GSTR3B|EXC-GSTR2|SFT-\d{3}|TDS-\d{3}|TCS-\d{3})\b', line):
                # Don't reset on EXC-GSTR1 variants (handled above)
                if not re.search(r'EXC-GSTR1', line):
                    current_pan = None
                continue

            # ── Collect detail row for current supplier ───────────────────
            if not current_pan:
                continue

            gstins_on_line = GSTIN_RE2.findall(line)
            mon_m          = MON_RE2.search(line)
            if not (gstins_on_line and mon_m):
                continue

            # ── Identify supplier GSTIN (generic — no hardcoded PANs) ────
            # Strategy 1: GSTIN inside parentheses → always the supplier
            sup_gstin = None
            paren_gstin = re.search(r'\((\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]Z[A-Z0-9])\)', line)
            if paren_gstin:
                sup_gstin = paren_gstin.group(1)
            else:
                # Strategy 2: GSTIN whose embedded PAN matches current_pan
                for g in gstins_on_line:
                    g_pan = g[2:12] if len(g) == 15 else ""
                    if g_pan == current_pan:
                        sup_gstin = g
                        break
                # Strategy 3: second GSTIN on line (first = buyer, second = supplier)
                if not sup_gstin and len(gstins_on_line) >= 2:
                    sup_gstin = gstins_on_line[1]
                # Strategy 4: only one GSTIN, use it
                if not sup_gstin:
                    sup_gstin = gstins_on_line[0]

            if not sup_gstin:
                sup_gstin = current_pan + "_pan_only"
            current_gstin = sup_gstin

            mon_str  = f"{MON_MAP2[mon_m.group(1)]}-{mon_m.group(2)}"
            raw_nums = NUM_RE2.findall(line)
            nums = []
            for n in raw_nums:
                try:
                    v = float(n.replace(",", ""))
                    if v > 0 and not (v < 100 and v == int(v)) and v not in (2024.0, 2025.0, 2026.0):
                        nums.append(v)
                except:
                    pass
            if nums:
                amt = nums[-1]
                result[current_gstin]["name"]  = current_name or current_gstin
                result[current_gstin]["pan"]   = current_pan
                result[current_gstin]["months"][mon_str] += amt

        # ── PASS 2: Table-mode scanner (recovers rows that text parser misses) ─
        # pdfplumber tables can capture rows the line scanner misses (merged cells etc.)
        seen_pairs = {(g, mon)
                      for g, d in result.items()
                      for mon in d["months"]}

        in_pur_tbl = False
        tbl_pan    = None
        tbl_name   = None
        for table in all_tables:
            if not table: continue
            for row in table:
                if not row or not any(row): continue
                vals     = [str(c).strip() if c else "" for c in row]
                flat     = " ".join(vals)
                flat_low = flat.lower()

                # Detect header row
                if "EXC-GSTR1" in flat and any(k in flat for k in ["Purchase","GSTR-1 of seller"]):
                    pm = PAN_RE2.search(flat)
                    if pm:
                        tbl_pan  = pm.group(1)
                        m2 = re.search(r'of seller\s+(.+?)\s+\([A-Z]{5}\d{4}[A-Z]\)', flat)
                        tbl_name = m2.group(1).strip() if m2 else tbl_pan
                        in_pur_tbl = True
                    continue

                if not in_pur_tbl or not tbl_pan:
                    continue

                # Detect non-purchase code → end section
                if re.search(r'\b(EXC-GSTR3B|SFT-|TDS-|TCS-)\b', flat) and "EXC-GSTR1" not in flat:
                    in_pur_tbl = False; tbl_pan = None; continue

                gstins_tbl = GSTIN_RE2.findall(flat)
                mon_m_tbl  = MON_RE2.search(flat)
                if not (gstins_tbl and mon_m_tbl):
                    continue

                # Same supplier identification strategy
                sup_g = None
                pg = re.search(r'\((\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]Z[A-Z0-9])\)', flat)
                if pg:
                    sup_g = pg.group(1)
                else:
                    for g in gstins_tbl:
                        if g[2:12] == tbl_pan:
                            sup_g = g; break
                    if not sup_g and len(gstins_tbl) >= 2:
                        sup_g = gstins_tbl[1]
                    if not sup_g:
                        sup_g = gstins_tbl[0]

                mon_str_tbl = f"{MON_MAP2[mon_m_tbl.group(1)]}-{mon_m_tbl.group(2)}"
                pair = (sup_g, mon_str_tbl)
                if pair in seen_pairs:
                    continue
                seen_pairs.add(pair)

                nums_tbl = []
                for n in NUM_RE2.findall(flat):
                    try:
                        v = float(n.replace(",", ""))
                        if v > 0 and not (v < 100 and v == int(v)) and v not in (2024.0, 2025.0, 2026.0):
                            nums_tbl.append(v)
                    except: pass
                if nums_tbl:
                    result[sup_g]["name"]  = tbl_name or sup_g
                    result[sup_g]["pan"]   = tbl_pan
                    result[sup_g]["months"][mon_str_tbl] += nums_tbl[-1]

    except Exception as e:
        warns.append(f"AIS PDF parse error: {e}")

    # Convert to plain dict with regular dicts
    final = {}
    for gstin, d in result.items():
        final[gstin] = {
            "name":   d["name"],
            "pan":    d["pan"],
            "months": dict(d["months"])
        }

    total_months = sum(len(d["months"]) for d in final.values())
    print(f"    → {len(final)} supplier(s), {total_months} supplier-month entries found in AIS")
    if not final:
        warns.append("⚠ AIS PDF: Could not extract supplier data — please check PDF format")
    return final, warns


def _parse_period_to_month(text):
    """
    Convert return period text to month string like 'Apr-2025'.
    Handles: APR-2025, April 2025, 04/2025, 04-2025, Apr2025, etc.
    """
    text = str(text).strip()
    if not text: return None
    
    # Pattern: APR-2025 or APR 2025
    m = re.search(r'([A-Za-z]{3,9})[-\s](\d{4})', text)
    if m:
        mon_word = m.group(1).lower()[:3]
        yr = m.group(2)
        abbr = MON_MAP.get(mon_word)
        if abbr: return f"{abbr}-{yr}"
    
    # Pattern: 04/2025 or 04-2025
    m = re.search(r'(\d{2})[-/](\d{4})', text)
    if m:
        mon_num = int(m.group(1))
        yr      = m.group(2)
        mon_abbrs = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        if 1 <= mon_num <= 12:
            return f"{mon_abbrs[mon_num-1]}-{yr}"
    
    return None


# ── Find PDF and GST Automation folder ──────────────────────────────

def find_files(extra_gst_folder=None, exclude_dirs=None):
    """
    Auto-discover all needed files.
    Returns: (gstr2b_dirs, tis_pdf_path, ais_pdf_path)
    exclude_dirs: set/list of normalized absolute paths to skip (e.g. output folder).
    """
    home = pathlib.Path.home()
    downloads = home / "Downloads"

    # Build exclude set (normalised so comparisons work on Windows/Linux)
    _excl = set()
    if exclude_dirs:
        for d in exclude_dirs:
            try: _excl.add(os.path.normcase(os.path.normpath(str(d))))
            except: pass

    def _is_excluded(p):
        try:
            np = os.path.normcase(os.path.normpath(str(p)))
            for ex in _excl:
                if np == ex or np.startswith(ex + os.sep):
                    return True
        except: pass
        return False

    # Common search roots
    search_roots = [downloads]
    try:
        search_roots += list(home.glob("OneDrive*"))
        search_roots += [home / "Desktop", home / "Documents"]
    except: pass
    if extra_gst_folder:
        search_roots.insert(0, pathlib.Path(extra_gst_folder))

    # Find GSTR-2B Excel dirs — use ONLY precise GSTR2B_Month_Year.xlsx pattern to
    # avoid picking up unrelated files that happen to contain "2B"
    gstr2b_dirs_raw = []
    for root in search_roots:
        if not root.exists(): continue
        # Direct pattern in root
        if list(root.glob("GSTR2B*.xlsx")) or list(root.glob("GSTR-2B*.xlsx")):
            gstr2b_dirs_raw.append(str(root))
        # Recurse (only precise name patterns)
        for pat in ["GSTR2B*.xlsx", "GSTR-2B*.xlsx"]:
            for p in root.rglob(pat):
                if _is_excluded(p.parent): continue
                d = str(p.parent)
                gstr2b_dirs_raw.append(d)

    # Also search for FT6 / GST_Automation style folders by name
    for root in search_roots:
        if not root.exists(): continue
        for p in root.rglob("GST_Automation"):
            if p.is_dir() and not _is_excluded(p):
                gstr2b_dirs_raw.append(str(p))
        for pat in ["FT6", "GST*Automation*"]:
            for p in root.glob(pat):
                if p.is_dir() and not _is_excluded(p):
                    gstr2b_dirs_raw.append(str(p))

    # Deduplicate directories using normalised path comparison
    seen_dirs = set()
    gstr2b_dirs = []
    for d in gstr2b_dirs_raw:
        nd = os.path.normcase(os.path.normpath(d))
        if nd not in seen_dirs and not _is_excluded(d):
            seen_dirs.add(nd)
            gstr2b_dirs.append(d)

    # Find TIS and AIS PDFs (most recent wins)
    tis_pdf = None
    ais_pdf = None
    tis_mtime = 0
    ais_mtime = 0
    for root in search_roots:
        if not root.exists(): continue
        for p in root.rglob("TIS*.pdf"):
            try:
                mt = p.stat().st_mtime
                if mt > tis_mtime:
                    tis_pdf = str(p); tis_mtime = mt
            except: pass
        for p in root.rglob("AIS*.pdf"):
            try:
                mt = p.stat().st_mtime
                if mt > ais_mtime:
                    ais_pdf = str(p); ais_mtime = mt
            except: pass

    return gstr2b_dirs, tis_pdf, ais_pdf


# ═══════════════════════════════════════════════════════════════════
# FILE 1: GSTR2B_EXTRACT.xlsx  (AUTO-POPULATED)
# ═══════════════════════════════════════════════════════════════════

def build_2b_extract(out_dir, gdata: GSTRData):
    wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]

    sup_totals      = gdata.supplier_totals()      # gstin → {...}
    sup_mon_totals  = gdata.supplier_month_totals() # (gstin,mon) → {...}
    all_suppliers   = gdata.all_suppliers()          # [(gstin,name), ...]

    # ── SHEET 1: 2B_Raw_Input (auto-filled) ─────────────────────
    ws1 = wb.create_sheet("2B_Raw_Input")
    ws1.sheet_view.showGridLines = False

    COLS1 = [
        ("Supplier GSTIN",    22), ("Supplier Name",     36),
        ("Month",             12), ("Taxable Value ₹",   18),
        ("IGST ₹",            14), ("CGST ₹",            14),
        ("SGST ₹",            14), ("Total GST ₹",       16),
        ("Source File",       24), ("Remarks",           28),
    ]
    NC1 = len(COLS1)
    _title(ws1, "GSTR-2B RAW DATA  —  AUTO-IMPORTED FROM MONTHLY GSTR-2B EXCEL FILES", NC1)
    _info(ws1, 2, NC1,
          f"✅ AUTO-IMPORTED: {len(gdata.b2b_rows)} rows from {len(set(r['source'] for r in gdata.b2b_rows))} GSTR-2B file(s).  "
          f"Suppliers: {len(all_suppliers)}  |  Months: {len(set(r['month'] for r in gdata.b2b_rows))}  "
          f"|  Last run: {datetime.now().strftime('%d-%b-%Y %H:%M')}")
    _hdr(ws1, COLS1, row=3, bg=NAVY)
    ws1.freeze_panes = "A4"

    r = 4
    for row in sorted(gdata.b2b_rows, key=lambda x: (x["name"], MONTHS_FY.index(x["month"]) if x["month"] in MONTHS_FY else 99)):
        bg = ALT1 if r % 2 == 0 else ALT2
        total_gst = (row.get("igst") or 0) + (row.get("cgst") or 0) + (row.get("sgst") or 0)
        _c(ws1, r, 1, row["gstin"],           AUTO_BG, align="left")
        _c(ws1, r, 2, row["name"],            AUTO_BG, align="left")
        _c(ws1, r, 3, row["month"],           AUTO_BG, align="center")
        _c(ws1, r, 4, row.get("taxable",0),   AUTO_BG, align="right", numfmt=NUM_FMT)
        _c(ws1, r, 5, row.get("igst",0),      AUTO_BG, align="right", numfmt=NUM_FMT)
        _c(ws1, r, 6, row.get("cgst",0),      AUTO_BG, align="right", numfmt=NUM_FMT)
        _c(ws1, r, 7, row.get("sgst",0),      AUTO_BG, align="right", numfmt=NUM_FMT)
        _c(ws1, r, 8, total_gst,              bg,      align="right", numfmt=NUM_FMT)
        _c(ws1, r, 9, row.get("source",""),   bg,      align="left")
        _c(ws1, r,10, "",                      bg)
        r += 1

    if not gdata.b2b_rows:
        _info(ws1, 4, NC1, "⚠ No GSTR-2B data found. Place GSTR2B_*.xlsx files in ~/Downloads and re-run.", "FFC7CE", "9C0006")

    # ── SHEET 2: 2B_By_Supplier (auto-computed) ──────────────────
    ws2 = wb.create_sheet("2B_By_Supplier")
    ws2.sheet_view.showGridLines = False

    COLS2 = [
        ("Supplier GSTIN",    22), ("Supplier Name",     36),
        ("Taxable Value ₹",   18), ("IGST ₹",            14),
        ("CGST ₹",            14), ("SGST ₹",            14),
        ("Total GST ₹",       16), ("TIS Accepted ₹",    18),
        ("Difference ₹",      16), ("Status",            14),
        ("Remarks",           28),
    ]
    NC2 = len(COLS2)
    _title(ws2, "TRADE NAME TOTALS — ALL MONTHS  |  TIS COMPARISON  |  AY 2026-27", NC2)
    _info(ws2, 2, NC2,
          "✅ AUTO-COMPUTED from 2B_Raw_Input.  "
          "TIS Accepted (col H) is auto-filled from TIS PDF if available.  "
          "Green = ✓ Match  |  Amber = ⚠ Minor Variance  |  Red = ✗ CHECK")
    _sep(ws2, 3, "SUPPLIER TOTALS  (All 12 months combined)", NC2, bg=TEAL)
    _hdr(ws2, COLS2, row=4, bg=NAVY)
    ws2.freeze_panes = "A5"

    r = 5
    for gstin, name in all_suppliers:
        d   = sup_totals.get(gstin, {})
        tis = gdata.get_tis(gstin)
        bg  = ALT1 if r % 2 == 0 else ALT2

        taxable = d.get("taxable", 0)
        igst    = d.get("igst",    0)
        cgst    = d.get("cgst",    0)
        sgst    = d.get("sgst",    0)
        total   = d.get("total",   0)
        tis_val = tis.get("accepted", 0)
        # TIS "Accepted" is taxable purchase amount — diff against Taxable Value
        diff    = taxable - tis_val

        _c(ws2, r, 1, gstin,   AUTO_BG, align="left")
        _c(ws2, r, 2, name,    AUTO_BG, align="left")
        _c(ws2, r, 3, taxable, AUTO_BG, align="right", numfmt=NUM_FMT)
        _c(ws2, r, 4, igst,    AUTO_BG, align="right", numfmt=NUM_FMT)
        _c(ws2, r, 5, cgst,    AUTO_BG, align="right", numfmt=NUM_FMT)
        _c(ws2, r, 6, sgst,    AUTO_BG, align="right", numfmt=NUM_FMT)
        _c(ws2, r, 7, total,   AUTO_BG, align="right", numfmt=NUM_FMT)

        tis_bg = AUTO_BG if tis_val > 0 else INPUT_BG  # Green if auto-filled, blue if needs input
        _c(ws2, r, 8, tis_val, tis_bg,  align="right", numfmt=NUM_FMT)
        _c(ws2, r, 9, diff,    bg,      align="right", numfmt=NUM_FMT)

        if tis_val == 0:    status = "Enter TIS →"
        elif abs(diff) < 1000:  status = "✓ Match"
        elif abs(diff) < 50000: status = "⚠ Minor Var"
        else:               status = "✗ CHECK"
        status_bg = (GREEN if "Match" in status else AMBER if "Minor" in status else RED_BG if "CHECK" in status else ALT1)
        _c(ws2, r, 10, status,  status_bg, align="center")
        _c(ws2, r, 11, tis.get("name","") if tis_val>0 else "", bg)
        r += 1

    # Grand Total
    if all_suppliers:
        tot_r = r
        _c(ws2, tot_r, 1, "GRAND TOTAL", DGRAY, bold=True)
        _c(ws2, tot_r, 2, "", DGRAY)
        for ci in range(3, 10):
            col_l = get_column_letter(ci)
            _c(ws2, tot_r, ci, f"=SUM({col_l}5:{col_l}{tot_r-1})",
               DGRAY, bold=True, align="right", numfmt=NUM_FMT)
        for ci in [10, 11]:
            _c(ws2, tot_r, ci, "", DGRAY)

    # ── SHEET 3: 2B_MonthWise (auto-computed) ────────────────────
    ws3 = wb.create_sheet("2B_MonthWise")
    ws3.sheet_view.showGridLines = False

    COLS3 = [
        ("Supplier GSTIN",    22), ("Supplier Name",     36),
        ("Month",             12), ("Taxable Value ₹",   18),
        ("IGST ₹",            14), ("CGST ₹",            14),
        ("SGST ₹",            14), ("Total GST ₹",       16),
        ("AIS Reported ₹",    18), ("Difference ₹",      16),
        ("AIS Status",        14), ("Return Period",     14),
        ("Remarks",           28),
    ]
    NC3 = len(COLS3)
    _title(ws3, "TRADE NAME TOTALS — MONTH-WISE  |  AIS COMPARISON  |  AY 2026-27", NC3)
    _info(ws3, 2, NC3,
          "✅ AUTO-COMPUTED from 2B_Raw_Input.  "
          "AIS Reported (col I) is auto-filled from AIS PDF if available.  "
          "Green = ✓ Match  |  Amber = ⚠ Minor  |  Red = ✗ CHECK  |  — = No data both sides")
    _hdr(ws3, COLS3, row=3, bg=NAVY)
    ws3.freeze_panes = "A4"

    r = 4
    for gstin, name in all_suppliers:
        ais_months = gdata.get_ais_months(gstin)

        # ── Pre-check: collect only months that have data on at least one side ──
        data_rows = []
        for mon_abbr, yr in MONTHS_SHORT:
            mon_str = f"{mon_abbr}-{yr}"
            key = (gstin, mon_str)
            d   = sup_mon_totals.get(key, {})
            taxable = d.get("taxable", 0) or 0
            igst    = d.get("igst",    0) or 0
            cgst    = d.get("cgst",    0) or 0
            sgst    = d.get("sgst",    0) or 0
            total   = d.get("total",   0) or 0
            ais_val = ais_months.get(mon_str, 0) or 0
            # Only include row if GSTR-2B taxable OR AIS has a non-zero value
            if taxable == 0 and ais_val == 0:
                continue
            data_rows.append((mon_str, mon_abbr, yr, taxable, igst, cgst, sgst, total, ais_val))

        if not data_rows:
            continue   # skip supplier entirely if no data in any month

        _sep(ws3, r, f"  {name}  |  GSTIN: {gstin}", NC3, bg=BLUE)
        r += 1

        for (mon_str, mon_abbr, yr, taxable, igst, cgst, sgst, total, ais_val) in data_rows:
            bg  = ALT1 if r % 2 == 0 else ALT2

            # Difference: compare Taxable Value vs AIS (both are purchase values, not GST)
            diff = taxable - ais_val

            _c(ws3, r, 1, gstin,   AUTO_BG, align="left")
            _c(ws3, r, 2, name,    AUTO_BG, align="left")
            _c(ws3, r, 3, mon_str, AUTO_BG, align="center")
            _c(ws3, r, 4, taxable, AUTO_BG, align="right", numfmt=NUM_FMT)
            _c(ws3, r, 5, igst,    AUTO_BG, align="right", numfmt=NUM_FMT)
            _c(ws3, r, 6, cgst,    AUTO_BG, align="right", numfmt=NUM_FMT)
            _c(ws3, r, 7, sgst,    AUTO_BG, align="right", numfmt=NUM_FMT)
            _c(ws3, r, 8, total,   AUTO_BG, align="right", numfmt=NUM_FMT)

            ais_bg = AUTO_BG if ais_val > 0 else INPUT_BG
            _c(ws3, r, 9,  ais_val, ais_bg, align="right", numfmt=NUM_FMT)
            _c(ws3, r, 10, diff,    bg,     align="right", numfmt=NUM_FMT)

            if taxable == 0 and ais_val == 0:  status = "—"
            elif ais_val == 0 and taxable > 0: status = "Not in AIS"
            elif taxable == 0 and ais_val > 0: status = "Only in AIS"
            elif abs(diff) < 1:                status = "✓ Match"
            elif abs(diff) < 5000:             status = "⚠ Minor"
            else:                              status = "✗ CHECK"
            sbg = (GREEN  if "Match"    in status else
                   AMBER  if "Minor"    in status else
                   RED_BG if "CHECK"    in status else
                   YELLOW if "Only"     in status else ALT1)
            _c(ws3, r, 11, status,  sbg, align="center")
            _c(ws3, r, 12, f"{mon_abbr.upper()}-{yr}", bg, align="center")
            _c(ws3, r, 13, "", bg)
            r += 1

    out_path = os.path.join(out_dir, "GSTR2B_EXTRACT.xlsx")
    wb.save(out_path)
    return out_path


# ═══════════════════════════════════════════════════════════════════
# FILE 2: TIS_AIS_COMPARISON.xlsx  (AUTO-POPULATED)
# ═══════════════════════════════════════════════════════════════════

def build_tis_ais_comparison(out_dir, gdata: GSTRData):
    wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]

    sup_totals     = gdata.supplier_totals()
    sup_mon_totals = gdata.supplier_month_totals()
    all_suppliers  = gdata.all_suppliers()

    # ── SHEET 1: TIS_vs_2B ───────────────────────────────────────
    ws1 = wb.create_sheet("TIS_vs_2B")
    ws1.sheet_view.showGridLines = False

    COLS1 = [
        ("Supplier GSTIN",        22), ("Party Name",            36),
        ("2B Taxable Value ₹",    18), ("2B IGST ₹",             14),
        ("2B CGST ₹",             14), ("2B SGST ₹",             14),
        ("2B Total GST ₹",        16), ("TIS Accepted ₹",        18),
        ("Difference (2B-TIS) ₹", 18), ("% Variance",            12),
        ("Status",                14), ("Action Needed",         32),
    ]
    NC1 = len(COLS1)
    _title(ws1, "TIS vs GSTR-2B COMPARISON  |  All Months  |  AY 2026-27", NC1)

    tis_count = len([g for g in all_suppliers if gdata.get_tis(g[0]).get("accepted",0) > 0])
    _info(ws1, 2, NC1,
          f"✅ AUTO-FILLED: 2B amounts from GSTR-2B files.  TIS Accepted from TIS PDF ({tis_count} supplier(s) matched).  "
          f"Green = ✓ Match (<₹1K diff)  |  Amber = ⚠ Minor (<₹50K)  |  Red = ✗ CHECK  "
          f"|  Blue cells = TIS value not found — enter manually")
    _hdr(ws1, COLS1, row=3, bg=NAVY)
    ws1.freeze_panes = "C4"

    r = 4
    for gstin, name in all_suppliers:
        d   = sup_totals.get(gstin, {})
        tis = gdata.get_tis(gstin)
        bg  = ALT1 if r % 2 == 0 else ALT2

        taxable = d.get("taxable", 0); igst = d.get("igst",0)
        cgst    = d.get("cgst",   0);  sgst = d.get("sgst",0)
        total   = d.get("total",  0)
        tis_val = tis.get("accepted", 0)
        # TIS "Accepted" is the taxable purchase amount — compare against Taxable Value, not GST
        diff    = taxable - tis_val
        pct     = (diff / tis_val * 100) if tis_val else None

        _c(ws1, r, 1, gstin,   AUTO_BG, align="left")
        _c(ws1, r, 2, name,    AUTO_BG, align="left")
        _c(ws1, r, 3, taxable, AUTO_BG, align="right", numfmt=NUM_FMT)
        _c(ws1, r, 4, igst,    AUTO_BG, align="right", numfmt=NUM_FMT)
        _c(ws1, r, 5, cgst,    AUTO_BG, align="right", numfmt=NUM_FMT)
        _c(ws1, r, 6, sgst,    AUTO_BG, align="right", numfmt=NUM_FMT)
        _c(ws1, r, 7, total,   AUTO_BG, align="right", numfmt=NUM_FMT)

        tis_bg = AUTO_BG if tis_val > 0 else INPUT_BG
        _c(ws1, r, 8, tis_val, tis_bg, align="right", numfmt=NUM_FMT)
        _c(ws1, r, 9, diff,    bg,     align="right", numfmt=NUM_FMT)
        _c(ws1, r,10, f"{pct:.2f}%" if pct is not None else "N/A", bg, align="center")

        if tis_val == 0:         status = "Enter TIS →"; sbg = INPUT_BG
        elif abs(diff) < 1000:   status = "✓ Match";     sbg = GREEN
        elif abs(diff) < 50000:  status = "⚠ Minor Var"; sbg = AMBER
        else:                    status = "✗ CHECK";      sbg = RED_BG
        _c(ws1, r, 11, status, sbg, align="center")

        if tis_val == 0: action = "TIS PDF not matched — enter manually"
        elif abs(diff) < 1000: action = "No action needed"
        elif diff > 0: action = "2B higher than TIS — verify invoices"
        else: action = "TIS higher than 2B — check missing invoices"
        _c(ws1, r, 12, action, bg, align="left")
        r += 1

    # Grand Total
    tot_r = r
    _c(ws1, tot_r, 1, "GRAND TOTAL", DGRAY, bold=True)
    _c(ws1, tot_r, 2, "", DGRAY)
    for ci in range(3, 10):
        col_l = get_column_letter(ci)
        _c(ws1, tot_r, ci, f"=SUM({col_l}4:{col_l}{tot_r-1})", DGRAY, bold=True, align="right", numfmt=NUM_FMT)
    _c(ws1, tot_r, 10, "", DGRAY, bold=True)
    _c(ws1, tot_r, 11, f'=IF(ABS(I{tot_r})<1000,"✓ Match",IF(ABS(I{tot_r})<50000,"⚠ Minor","✗ CHECK"))',
       DGRAY, bold=True, align="center")
    _c(ws1, tot_r, 12, "", DGRAY)

    # ── SHEET 2: AIS_vs_2B ───────────────────────────────────────
    ws2 = wb.create_sheet("AIS_vs_2B")
    ws2.sheet_view.showGridLines = False

    COLS2 = [
        ("Supplier GSTIN",        22), ("Party Name",            36),
        ("Return Period",         14), ("2B Purchase ₹",         18),
        ("2B Total GST ₹",        16), ("AIS Reported ₹",        18),
        ("Difference (2B-AIS) ₹", 18), ("AIS Status",            14),
        ("Match Status",          16), ("Remarks",               32),
    ]
    NC2 = len(COLS2)
    _title(ws2, "AIS vs GSTR-2B COMPARISON  |  Month-Wise  |  AY 2026-27", NC2)

    ais_count = len(gdata.ais_data)
    _info(ws2, 2, NC2,
          f"✅ AUTO-FILLED: 2B monthly amounts from GSTR-2B files.  AIS data from AIS PDF ({ais_count} supplier(s) matched).  "
          f"AIS Return Period = filing month shown in AIS (may lag 1-2 months vs actual purchase month).  "
          f"Blue cells = AIS value not found — enter manually if needed")
    _hdr(ws2, COLS2, row=3, bg=NAVY)
    ws2.freeze_panes = "C4"

    r = 4
    for gstin, name in all_suppliers:
        ais_months = gdata.get_ais_months(gstin)

        # Pre-collect only months with data on at least one side
        data_rows = []
        for mon_abbr, yr in MONTHS_SHORT:
            mon_str = f"{mon_abbr}-{yr}"
            key = (gstin, mon_str)
            d   = sup_mon_totals.get(key, {})
            taxable = d.get("taxable", 0) or 0
            total   = d.get("total",   0) or 0
            ais_val = ais_months.get(mon_str, 0) or 0
            if taxable == 0 and ais_val == 0:
                continue
            data_rows.append((mon_str, mon_abbr, yr, taxable, total, ais_val))

        if not data_rows:
            continue   # skip supplier block entirely

        _sep(ws2, r, f"  {name}  |  {gstin}", NC2, bg=BLUE)
        r += 1

        for (mon_str, mon_abbr, yr, taxable, total, ais_val) in data_rows:
            bg  = ALT1 if r % 2 == 0 else ALT2

            # Difference: Taxable Value vs AIS Reported (both represent purchase amount)
            diff = taxable - ais_val

            ais_period = f"{mon_abbr.upper()}-{yr}"
            _c(ws2, r, 1, gstin,      AUTO_BG, align="left")
            _c(ws2, r, 2, name,       AUTO_BG, align="left")
            _c(ws2, r, 3, ais_period, AUTO_BG, align="center")
            _c(ws2, r, 4, taxable,    AUTO_BG, align="right", numfmt=NUM_FMT)
            _c(ws2, r, 5, total,      AUTO_BG, align="right", numfmt=NUM_FMT)

            ais_bg = AUTO_BG if ais_val > 0 else INPUT_BG
            _c(ws2, r, 6, ais_val, ais_bg, align="right", numfmt=NUM_FMT)
            _c(ws2, r, 7, diff,    bg,     align="right", numfmt=NUM_FMT)

            if taxable == 0 and ais_val == 0:  ais_status = "Active"; mstatus = "— No data"
            elif ais_val == 0 and taxable > 0: ais_status = "Active"; mstatus = "Not in AIS"
            elif taxable == 0 and ais_val > 0: ais_status = "Active"; mstatus = "Only in AIS"
            elif abs(diff) < 1:                ais_status = "Active"; mstatus = "✓ Match"
            elif abs(diff) < 5000:             ais_status = "Active"; mstatus = "⚠ Minor"
            else:                              ais_status = "Active"; mstatus = "✗ CHECK"
            sbg = (GREEN  if "Match"  in mstatus else
                   AMBER  if "Minor"  in mstatus else
                   RED_BG if "CHECK"  in mstatus else
                   YELLOW if "Only"   in mstatus else ALT1)
            _c(ws2, r, 8, ais_status, bg,  align="center")
            _c(ws2, r, 9, mstatus,    sbg, align="center")
            _c(ws2, r,10, "",         bg)
            r += 1

    # ── SHEET 3: Summary_Dashboard ───────────────────────────────
    ws3 = wb.create_sheet("Summary_Dashboard")
    ws3.sheet_view.showGridLines = False

    NC3 = 8
    _title(ws3, "RECONCILIATION DASHBOARD  —  GSTR-2B vs TIS / AIS  |  AY 2026-27", NC3)

    ws3.column_dimensions["A"].width = 46
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 24
    ws3.column_dimensions["D"].width = 20
    ws3.column_dimensions["E"].width = 14
    ws3.column_dimensions["F"].width = 36
    ws3.column_dimensions["G"].width = 18
    ws3.column_dimensions["H"].width = 18

    ri = 2
    def hdr_row():
        nonlocal ri
        for ci, (h, _) in enumerate([
            ("Particulars",""),("GSTR-2B Amount ₹",""),("TIS / AIS Amount ₹",""),
            ("Difference ₹",""),("Status",""),("Notes / Action",""),("",""),("","")
        ], 1):
            c = ws3.cell(row=ri, column=ci, value=h)
            c.font = _fn(True, WHITE, 9); c.fill = _f(NAVY)
            c.alignment = _al("center", w=True); c.border = _bd()
        ws3.row_dimensions[ri].height = 20
        ri += 1

    def dash_sep(lbl, bg=NAVY):
        nonlocal ri; _sep(ws3, ri, lbl, NC3, bg=bg); ri += 1

    def dash_row(label, val2b=None, val_tis_ais=None, bold=False, bg=None, note="", fmt=NUM_FMT):
        nonlocal ri
        bgu = bg or (ALT2 if ri % 2 == 0 else ALT1)
        diff = None
        if isinstance(val2b, (int,float)) and isinstance(val_tis_ais, (int,float)):
            diff = val2b - val_tis_ais
        if diff is not None:
            if abs(diff) < 1000:    status = "✓ Match";    sbg = GREEN
            elif abs(diff) < 50000: status = "⚠ Minor";    sbg = AMBER
            else:                   status = "✗ CHECK";     sbg = RED_BG
        else: status = ""; sbg = bgu

        _c(ws3, ri, 1, label,       bgu,  bold=bold)
        _c(ws3, ri, 2, val2b,       bgu,  bold=bold, align="right", numfmt=fmt if isinstance(val2b,(int,float)) else None)
        _c(ws3, ri, 3, val_tis_ais, bgu,  bold=bold, align="right", numfmt=fmt if isinstance(val_tis_ais,(int,float)) else None)
        _c(ws3, ri, 4, diff,        bgu,  bold=bold, align="right", numfmt=fmt if isinstance(diff,(int,float)) else None)
        _c(ws3, ri, 5, status,      sbg,  bold=bold, align="center")
        _c(ws3, ri, 6, note,        bgu,  align="left")
        _c(ws3, ri, 7, "",          bgu)
        _c(ws3, ri, 8, "",          bgu)
        ws3.row_dimensions[ri].height = 16
        ri += 1

    hdr_row()

    # Compute overall totals
    total_2b_taxable = sum(d["taxable"] for d in sup_totals.values())
    total_2b_gst     = sum(d["total"]   for d in sup_totals.values())
    total_tis        = sum(d["accepted"] for d in gdata.tis_data.values())
    total_ais        = sum(
        sum(months.values())
        for d in gdata.ais_data.values()
        for months in [d.get("months", {})]
    )

    dash_sep("SECTION A — OVERALL PURCHASE RECONCILIATION (GSTR-2B vs TIS)", NAVY)
    dash_row("Total 2B Taxable Purchases (All Suppliers)", total_2b_taxable, None,
             note="Auto-computed from GSTR-2B Excel files")
    dash_row("Total 2B GST (IGST+CGST+SGST)", total_2b_gst, total_tis,
             bold=True,
             note=f"TIS from PDF ({len(gdata.tis_data)} suppliers). Enter manually if 0.")
    dash_row("No. of Suppliers in 2B", len(all_suppliers), None,
             note="Unique suppliers in all GSTR-2B files", fmt=INT_FMT)
    dash_row("No. of Suppliers in TIS (matched)", None, len(gdata.tis_data),
             note="Auto-matched from TIS PDF by GSTIN", fmt=INT_FMT)
    ri += 1

    dash_sep("SECTION B — MONTH-WISE RECONCILIATION (GSTR-2B vs AIS)", NAVY)
    for mon_abbr, yr in MONTHS_SHORT:
        mon_str = f"{mon_abbr}-{yr}"
        mon_2b  = sum(d.get("total",0) for (g,m), d in sup_mon_totals.items() if m == mon_str)
        mon_ais = sum(
            d.get("months", {}).get(mon_str, 0)
            for d in gdata.ais_data.values()
        )
        dash_row(f"  {mon_str}  —  Total GST", mon_2b, mon_ais if mon_ais > 0 else None,
                 note="AIS from PDF" if mon_ais > 0 else "AIS not found — enter manually")
    ri += 1

    dash_sep("SECTION C — SUMMARY STATISTICS", TEAL)
    matched_sup = len([g for g, n in all_suppliers if gdata.get_tis(g) != {}])
    unmatched   = len(all_suppliers) - matched_sup
    dash_row("Suppliers with TIS match", matched_sup, None, fmt=INT_FMT,
             note="Matched by GSTIN from TIS PDF")
    dash_row("Suppliers without TIS data", unmatched, None, fmt=INT_FMT,
             note="Enter TIS values manually in TIS_vs_2B sheet")
    dash_row("Suppliers with AIS data", len(gdata.ais_data), None, fmt=INT_FMT,
             note="Matched by GSTIN from AIS PDF")
    ri += 1

    dash_sep("SECTION D — DATA SOURCES USED", TEAL)
    gstr2b_files = sorted(set(r["source"] for r in gdata.b2b_rows))
    dash_row(f"GSTR-2B Files Read", len(gstr2b_files), None, fmt=INT_FMT,
             note=", ".join(gstr2b_files[:3]) + ("..." if len(gstr2b_files) > 3 else ""))
    dash_row("TIS PDF", None, None,
             note="Auto-read" if gdata.tis_data else "Not found — place TIS_*.pdf in Downloads")
    dash_row("AIS PDF", None, None,
             note="Auto-read" if gdata.ais_data else "Not found — place AIS_*.pdf in Downloads")
    ri += 1

    if gdata.warnings:
        dash_sep("SECTION E — WARNINGS", RED_BG)
        for w in gdata.warnings[:10]:
            _c(ws3, ri, 1, w, "FFF2CC"); ws3.merge_cells(f"A{ri}:H{ri}")
            ws3.row_dimensions[ri].height = 15; ri += 1

    out_path = os.path.join(out_dir, "TIS_AIS_COMPARISON.xlsx")
    wb.save(out_path)
    return out_path


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="GST-IT Comparison Builder v2.0 — FULLY AUTOMATIC")
    parser.add_argument("--fy",         default="2025-26", help="Financial Year e.g. 2025-26")
    parser.add_argument("--out",        default=None,      help="Output folder")
    parser.add_argument("--gst-folder", default=None,      help="Extra folder to search for GSTR-2B files")
    parser.add_argument("--tis-pdf",    default=None,      help="Direct path to TIS PDF")
    parser.add_argument("--ais-pdf",    default=None,      help="Direct path to AIS PDF")
    args = parser.parse_args()

    # ── Find output folder ────────────────────────────────────────
    def _find_output_base():
        home = pathlib.Path.home()
        candidates = [
            home / "OneDrive" / "Desktop" / "OUTPUT",
            home / "OneDrive - Personal" / "Desktop" / "OUTPUT",
            home / "Desktop" / "OUTPUT",
            home / "Downloads",
        ]
        try:
            candidates[2:2] = [p / "Desktop" / "OUTPUT"
                                for p in home.glob("OneDrive*") if p.is_dir()]
        except: pass
        for c in candidates:
            if c.exists():
                return str(c)
        return os.path.expanduser("~/Downloads")

    out_dir = args.out or os.path.join(_find_output_base(), "GST_IT_Comparison")
    os.makedirs(out_dir, exist_ok=True)

    print("=" * 60)
    print(f"  GST-IT COMPARISON BUILDER v2.0 — FULLY AUTOMATIC")
    print(f"  FY: {args.fy}   |   Output: {out_dir}")
    print("=" * 60)

    # ── Step 1: Discover files ────────────────────────────────────
    print("\n[1/4] Searching for GSTR-2B, TIS, AIS files...")
    # Exclude the output folder so GSTR2B_EXTRACT.xlsx from a prior run is never re-read
    gstr2b_dirs, auto_tis, auto_ais = find_files(
        args.gst_folder, exclude_dirs=[out_dir]
    )
    tis_pdf = args.tis_pdf or auto_tis
    ais_pdf = args.ais_pdf or auto_ais

    if gstr2b_dirs:
        print(f"  GSTR-2B search folders: {len(gstr2b_dirs)}")
    if tis_pdf:
        print(f"  TIS PDF: {tis_pdf}")
    else:
        print("  TIS PDF: ⚠ NOT FOUND — place TIS_*.pdf in ~/Downloads")
    if ais_pdf:
        print(f"  AIS PDF: {ais_pdf}")
    else:
        print("  AIS PDF: ⚠ NOT FOUND — place AIS_*.pdf in ~/Downloads")

    # ── Step 2: Read GSTR-2B ─────────────────────────────────────
    print("\n[2/4] Reading GSTR-2B Excel files...")
    gdata = find_and_read_gstr2b_files(gstr2b_dirs, exclude_paths=[out_dir])

    # ── Step 3: Read TIS + AIS PDFs ──────────────────────────────
    print("\n[3/4] Reading TIS and AIS PDFs...")
    if tis_pdf:
        tis_result, tis_warns = read_tis_pdf(tis_pdf)
        gdata.tis_data = tis_result
        gdata.warnings.extend(tis_warns)
    else:
        gdata.warnings.append("TIS PDF not found")

    if ais_pdf:
        ais_result, ais_warns = read_ais_pdf(ais_pdf)
        gdata.ais_data = ais_result
        gdata.warnings.extend(ais_warns)
    else:
        gdata.warnings.append("AIS PDF not found")

    # ── Step 4: Build Excel files ─────────────────────────────────
    print("\n[4/4] Building Excel files...")

    print("  Building GSTR2B_EXTRACT.xlsx ...")
    p1 = build_2b_extract(out_dir, gdata)
    print(f"  ✓ Saved: {p1}")

    print("  Building TIS_AIS_COMPARISON.xlsx ...")
    p2 = build_tis_ais_comparison(out_dir, gdata)
    print(f"  ✓ Saved: {p2}")

    print("\n" + "=" * 60)
    print("  ✅ DONE")
    print(f"  Suppliers imported   : {len(gdata.all_suppliers())}")
    print(f"  B2B rows imported    : {len(gdata.b2b_rows)}")
    print(f"  TIS suppliers matched: {len(gdata.tis_data)}")
    print(f"  AIS suppliers matched: {len(gdata.ais_data)}")
    if gdata.warnings:
        print(f"  Warnings ({len(gdata.warnings)}):")
        for w in gdata.warnings:
            print(f"    {w}")
    print(f"\n  Output folder: {out_dir}")
    print("=" * 60)
