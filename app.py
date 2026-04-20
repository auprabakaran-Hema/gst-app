"""
GST Reconciliation Web App — v7 with Phased Auto Download
===================================================
• Fully free — no license, no restrictions
• Scripts (gst_suite_final.py, gstr1_extract.py) never exposed to users
• 4 tabs: Reconciliation | GSTR-1 Detail | Download Status | Auto Download
• NEW v7: Phased download flow (Generate → 2B → 3B → Collect)
• NEW v7: ⚡ Fast Case — GSTR-2B + GSTR-3B only (~10 min, no generate)
• NEW v7: Retry T1/T2/T3 with backoff per failed download
• NEW v7: Session keep-alive (Phase 2+3 prevent 20-min logout)
• NEW v7: Not Found page auto-reload handler
• NEW v7: Batch control (6 months per batch, avoids portal overload)
• NEW v7: Case selector with tab/time estimates for each option
• Render.com ready — binds to $PORT
"""

import os, sys, json, zipfile, re, time, shutil, uuid, threading, asyncio
from pathlib import Path
from datetime import datetime
from flask import Flask, request, jsonify, send_file, render_template_string, abort
import tempfile, platform

# ── HTTP long-poll bridge (no WebSocket needed) ───────────────────
WEBSOCKET_AVAILABLE = True   # always True — uses plain HTTP polling

# ── Directories ───────────────────────────────────────────────────
def _get_app_dir(subfolder):
    if platform.system() == "Windows":
        base = Path(os.path.expanduser("~")) / "Downloads" / "GST_WebApp"
    else:
        base = Path(tempfile.gettempdir()) / "gst_webapp"
    d = base / subfolder
    d.mkdir(parents=True, exist_ok=True)
    return d

UPLOAD_DIR  = _get_app_dir("uploads")
OUTPUT_DIR  = _get_app_dir("outputs")
FEEDBACK_FILE = Path(__file__).parent / "feedback.json"
ALLOWED_EXT = {".zip", ".xlsx", ".xls", ".pdf", ".json"}
MAX_FILE_MB = 100
JOB_TTL_S   = 7200

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_FILE_MB * 1024 * 1024

# (No WebSocket init needed — using HTTP long-poll)

jobs      = {}
jobs_lock = threading.Lock()

# ── HTTP Bridge State ─────────────────────────────────────────────
# cmd_queue  : server → PC  (commands waiting to be picked up)
# resp_queue : PC → server  (responses from PC browser)
import queue as _queue
_cmd_queue  = _queue.Queue()
_resp_queue = _queue.Queue()
_bridge_last_seen = 0       # epoch seconds of last PC poll
_bridge_lock = threading.Lock()

def _bridge_connected():
    """Server-side mode — always connected"""
    return True

# ── Rate limiting ─────────────────────────────────────────────────
_rate = {}
_rate_lock = threading.Lock()

def _check_rate(ip, limit=30, window=60):
    now = time.time()
    with _rate_lock:
        hits = [t for t in _rate.get(ip, []) if now - t < window]
        if len(hits) >= limit: return False
        hits.append(now); _rate[ip] = hits
    return True

def rate_limit(limit=30, window=60):
    from functools import wraps
    def dec(f):
        @wraps(f)
        def wrapped(*a, **kw):
            ip = request.remote_addr or "unknown"
            if not _check_rate(ip, limit, window):
                from flask import Response
                return Response('{"error":"Too many requests. Wait 1 minute."}',
                    status=429, mimetype="application/json")
            return f(*a, **kw)
        return wrapped
    return dec

# ── Feedback store ────────────────────────────────────────────────
def _load_feedback():
    try:
        if FEEDBACK_FILE.exists():
            return json.loads(FEEDBACK_FILE.read_text(encoding="utf-8"))
    except: pass
    return []

def _save_feedback(fb_list):
    try:
        FEEDBACK_FILE.write_text(json.dumps(fb_list, ensure_ascii=False, indent=2), encoding="utf-8")
    except: pass

# ── Helpers ───────────────────────────────────────────────────────
def _cleanup_old_jobs():
    try:
        now = time.time()
        for d in [UPLOAD_DIR, OUTPUT_DIR]:
            for sub in d.iterdir():
                if sub.is_dir() and (now - sub.stat().st_mtime) > JOB_TTL_S:
                    shutil.rmtree(str(sub), ignore_errors=True)
    except: pass

def _cleanup_uploads(job_id):
    try:
        up = UPLOAD_DIR / job_id
        if up.exists(): shutil.rmtree(str(up), ignore_errors=True)
    except: pass

def _find_engine(name):
    for loc in [
        Path(__file__).parent / name,
        Path(os.getcwd()) / name,
        Path(os.path.expanduser("~")) / "Desktop" / name,
    ]:
        if loc.exists(): return loc
    return None

MONTHS_MAP = {
    "april":"April","may":"May","june":"June","july":"July","august":"August",
    "september":"September","october":"October","november":"November",
    "december":"December","january":"January","february":"February","march":"March",
    "04":"April","05":"May","06":"June","07":"July","08":"August",
    "09":"September","10":"October","11":"November","12":"December",
    "01":"January","02":"February","03":"March",
}

def _fy_months(fy):
    s = int(fy.split("-")[0]); e = s + 1
    return {
        "April":str(s),"May":str(s),"June":str(s),"July":str(s),
        "August":str(s),"September":str(s),"October":str(s),"November":str(s),
        "December":str(s),"January":str(e),"February":str(e),"March":str(e),
    }

def _detect_month(fpath, FY_MONTHS):
    name = Path(fpath).stem.lower()
    for part in re.split(r'[_\-\s]', name):
        if part in MONTHS_MAP:
            mon = MONTHS_MAP[part]
            return mon, FY_MONTHS.get(mon, list(FY_MONTHS.values())[0])
    try:
        with zipfile.ZipFile(fpath) as z:
            for jn in z.namelist():
                if jn.endswith(".json"):
                    with z.open(jn) as jf:
                        d = json.load(jf)
                        fp = re.sub(r'[^0-9]','', d.get("fp",""))
                        if len(fp) == 6:
                            mon = MONTHS_MAP.get(fp[:2])
                            if mon: return mon, fp[2:]
    except: pass
    return None, None

# ── Block script access ───────────────────────────────────────────
@app.before_request
def block_scripts():
    p = request.path.lower()
    if p.endswith((".py", ".pyc")) or "gst_suite" in p or "gstr1_extract" in p or "gstr2b_extract" in p:
        abort(403)

# ═══════════════════════════════════════════════════════════════════
# HTML
# ═══════════════════════════════════════════════════════════════════
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta http-equiv="X-UA-Compatible" content="IE=edge">
<meta name="color-scheme" content="dark">
<title>GST Reconciliation Portal — Free</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#0a0e1a;--surf:#111827;--surf2:#1a2235;--bdr:#1e3050;
  --accent:#00e5ff;--accent2:#7c3aed;--grn:#00e676;--org:#ff6d00;
  --red:#ff1744;--txt:#e8edf5;--muted:#6b7fa3;
  --mono:'IBM Plex Mono',monospace;--sans:'Syne',sans-serif;
}
body{background:var(--bg);color:var(--txt);font-family:var(--sans);min-height:100vh;overflow-x:hidden}
body::before{content:'';position:fixed;inset:0;
  background-image:linear-gradient(rgba(0,229,255,.04) 1px,transparent 1px),
  linear-gradient(90deg,rgba(0,229,255,.04) 1px,transparent 1px);
  background-size:40px 40px;pointer-events:none;z-index:0}
.wrap{max-width:1000px;margin:0 auto;padding:2rem 1.5rem;position:relative;z-index:1}

/* Header */
header{text-align:center;padding:2rem 0 1.25rem}
.logo{display:inline-flex;align-items:center;gap:.7rem;margin-bottom:.8rem}
.logo-icon{width:46px;height:46px;background:linear-gradient(135deg,var(--accent),var(--accent2));
  border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.4rem}
.logo-text{font-size:1rem;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
  background:linear-gradient(135deg,var(--accent),var(--accent2));
  -webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent}
h1{font-size:clamp(1.5rem,3.2vw,2.2rem);font-weight:800;letter-spacing:-.02em;line-height:1.1}
h1 span{background:linear-gradient(135deg,var(--accent),var(--accent2));
  -webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent}
.sub{color:var(--muted);font-size:.82rem;margin-top:.35rem;font-family:var(--mono)}
.badges{display:flex;gap:.5rem;justify-content:center;flex-wrap:wrap;margin-top:.6rem}
.badge{display:inline-flex;align-items:center;gap:.3rem;padding:.28rem .8rem;border-radius:100px;
  font-size:.7rem;font-weight:700;font-family:var(--mono)}
.badge-grn{background:rgba(0,230,118,.15);color:var(--grn);border:1px solid rgba(0,230,118,.4)}
.badge-blue{background:rgba(0,229,255,.1);color:var(--accent);border:1px solid rgba(0,229,255,.3)}
.badge-purple{background:rgba(124,58,237,.15);color:#a78bfa;border:1px solid rgba(124,58,237,.3)}
.badge-orange{background:rgba(255,109,0,.15);color:var(--org);border:1px solid rgba(255,109,0,.4)}

/* Tabs */
.tabs{display:flex;gap:.25rem;border-bottom:2px solid var(--bdr);margin-bottom:1.1rem;overflow-x:auto}
.tb{padding:.55rem 1.1rem;background:none;border:none;color:var(--muted);
  font-family:var(--sans);font-size:.78rem;font-weight:700;cursor:pointer;
  border-bottom:2px solid transparent;margin-bottom:-2px;transition:all .2s;
  text-transform:uppercase;letter-spacing:.06em;white-space:nowrap}
.tb:hover{color:var(--txt)}.tb.active{color:var(--accent);border-bottom-color:var(--accent)}
.tp{display:none}.tp.active{display:block}

/* Cards */
.card{background:var(--surf);border:1px solid var(--bdr);border-radius:13px;
  padding:1.4rem;margin-bottom:1rem;transition:border-color .2s}
.card:hover{border-color:rgba(0,229,255,.15)}
.ct{font-size:.8rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;
  color:var(--accent);margin-bottom:.9rem;display:flex;align-items:center;gap:.45rem}
.ct::before{content:'';width:3px;height:1em;background:var(--accent);border-radius:2px}

/* Form */
.fg2{display:grid;grid-template-columns:1fr 1fr;gap:.75rem}
@media(max-width:600px){.fg2{grid-template-columns:1fr}}
.fg{display:flex;flex-direction:column;gap:.3rem}
label{font-size:.68rem;font-weight:600;letter-spacing:.06em;text-transform:uppercase;color:var(--muted)}
input[type=text],input[type=password],textarea,select{
  background:var(--surf2);border:1px solid var(--bdr);border-radius:7px;
  padding:.52rem .78rem;color:var(--txt);font-family:var(--mono);font-size:.82rem;
  transition:border-color .2s;width:100%}
textarea{resize:vertical;min-height:90px;line-height:1.55}
input:focus,textarea:focus,select:focus{outline:none;border-color:var(--accent)}
input::placeholder,textarea::placeholder{color:var(--muted)}
select option{background:var(--surf)}

/* Drop zones */
.dg{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:.65rem;margin-top:.45rem}
.dz{background:var(--surf2);border:2px dashed var(--bdr);border-radius:10px;
  padding:1rem .65rem;text-align:center;cursor:pointer;transition:all .2s;
  position:relative;min-height:100px;display:flex;flex-direction:column;
  align-items:center;justify-content:center;gap:.3rem}
.dz:hover,.dz.drag-over{border-color:var(--accent);background:rgba(0,229,255,.04)}
.dz.has-files{border-color:var(--grn);border-style:solid;background:rgba(0,230,118,.04)}
.dz-ic{font-size:1.6rem;line-height:1}
.dz-lb{font-size:.64rem;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:var(--muted)}
.dz-ht{font-size:.6rem;color:var(--muted);font-family:var(--mono)}
.dz-cn{font-size:.64rem;color:var(--grn);font-weight:600;font-family:var(--mono)}
.dz input[type=file]{position:absolute;inset:0;opacity:0;cursor:pointer}

/* Buttons */
.btn{width:100%;padding:.8rem;background:linear-gradient(135deg,var(--accent),var(--accent2));
  border:none;border-radius:10px;color:#000;font-family:var(--sans);font-size:.88rem;
  font-weight:800;letter-spacing:.05em;text-transform:uppercase;cursor:pointer;
  transition:transform .15s,box-shadow .15s;margin-top:.3rem}
.btn:hover{transform:translateY(-2px);box-shadow:0 8px 24px rgba(0,229,255,.25)}
.btn:disabled{opacity:.4;cursor:not-allowed;transform:none}
.btn-sec{width:100%;padding:.65rem;background:var(--surf2);border:1px solid var(--accent);
  border-radius:9px;color:var(--accent);font-family:var(--sans);font-size:.82rem;
  font-weight:700;cursor:pointer;transition:all .15s;margin-top:.4rem}
.btn-sec:hover{background:rgba(0,229,255,.08)}
.btn-orange{width:100%;padding:.8rem;background:linear-gradient(135deg,var(--org),#ff9100);
  border:none;border-radius:10px;color:#000;font-family:var(--sans);font-size:.88rem;
  font-weight:800;letter-spacing:.05em;text-transform:uppercase;cursor:pointer;
  transition:transform .15s,box-shadow .15s;margin-top:.3rem}
.btn-orange:hover{transform:translateY(-2px);box-shadow:0 8px 24px rgba(255,109,0,.25)}
.btn-orange:disabled{opacity:.4;cursor:not-allowed;transform:none}

/* Progress */
.pw{display:none}
.pb-w{background:var(--surf2);border-radius:100px;height:5px;overflow:hidden;margin:.65rem 0}
.pb{height:100%;background:linear-gradient(90deg,var(--accent),var(--accent2));
  border-radius:100px;transition:width .4s;width:0%}
.lb{background:#000;border:1px solid var(--bdr);border-radius:7px;
  padding:.75rem;font-family:var(--mono);font-size:.7rem;height:160px;overflow-y:auto;
  color:#aaffcc;line-height:1.7}
.lb .err{color:#ff6b6b}.lb .info{color:var(--accent)}
.lb .ok{color:var(--grn)}.lb .warn{color:var(--org)}

/* Downloads */
.dw{display:none}
.dl-g{display:grid;grid-template-columns:repeat(auto-fill,minmax(185px,1fr));gap:.65rem;margin-top:.65rem}
.dlc{background:var(--surf2);border:1px solid var(--bdr);border-radius:10px;
  padding:.9rem;display:flex;flex-direction:column;gap:.5rem}
.dl-n{font-size:.72rem;font-weight:600;color:var(--txt)}
.dl-s{font-size:.64rem;color:var(--muted);font-family:var(--mono)}
.btn-dl{padding:.38rem .8rem;background:var(--surf);border:1px solid var(--accent);
  border-radius:5px;color:var(--accent);font-family:var(--mono);font-size:.72rem;
  cursor:pointer;text-decoration:none;display:inline-block;transition:background .15s}
.btn-dl:hover{background:rgba(0,229,255,.1)}

/* Status badge */
.sbg{display:inline-flex;align-items:center;gap:.25rem;padding:.2rem .55rem;
  border-radius:100px;font-size:.64rem;font-weight:700;font-family:var(--mono)}
.s-p{background:rgba(255,109,0,.15);color:var(--org);border:1px solid rgba(255,109,0,.4)}
.s-d{background:rgba(0,230,118,.15);color:var(--grn);border:1px solid rgba(0,230,118,.4)}
.s-e{background:rgba(255,23,68,.15);color:var(--red);border:1px solid rgba(255,23,68,.4)}
.s-w{background:rgba(0,229,255,.15);color:var(--accent);border:1px solid rgba(0,229,255,.4)}
.pulse{animation:pulse 1.2s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.4}}
@keyframes fadeIn{from{opacity:0;transform:translateY(4px)}to{opacity:1;transform:translateY(0)}}

/* Status table */
.dst{width:100%;border-collapse:collapse;font-size:.69rem;font-family:var(--mono);margin-top:.55rem}
.dst th{background:var(--surf2);color:var(--muted);font-size:.6rem;font-weight:700;
  text-transform:uppercase;letter-spacing:.05em;padding:.4rem .5rem;
  border:1px solid var(--bdr);text-align:center}
.dst th:first-child{text-align:left}
.dst td{padding:.36rem .5rem;border:1px solid var(--bdr);text-align:center}
.dst tr:nth-child(even) td{background:rgba(255,255,255,.016)}
.dst td:first-child{text-align:left;color:var(--txt);font-weight:600}
.c-ok{color:var(--grn);font-weight:700}.c-fl{color:var(--red);font-weight:700}
.c-pd{color:var(--org)}.c-sk{color:var(--muted)}

/* Pills */
.pills{display:flex;flex-wrap:wrap;gap:.35rem;margin-bottom:.75rem}
.pill{padding:.22rem .6rem;background:var(--surf2);border:1px solid var(--bdr);
  border-radius:100px;font-size:.64rem;color:var(--muted);font-family:var(--mono)}

/* Info box */
.info-box{background:rgba(0,229,255,.05);border:1px solid rgba(0,229,255,.18);
  border-radius:9px;padding:.85rem 1rem;margin-bottom:.9rem;
  font-size:.78rem;color:var(--muted);line-height:1.65}
.info-box strong{color:var(--txt)}
.info-box.warn{background:rgba(255,109,0,.05);border-color:rgba(255,109,0,.18)}
.info-box.success{background:rgba(0,230,118,.05);border-color:rgba(0,230,118,.18)}

/* Connection status */
.conn-status{display:flex;align-items:center;gap:.5rem;padding:.5rem .75rem;
  background:var(--surf2);border:1px solid var(--bdr);border-radius:8px;
  font-size:.72rem;font-family:var(--mono);margin-bottom:1rem}
.conn-dot{width:8px;height:8px;border-radius:50%;background:var(--red)}
.conn-dot.online{background:var(--grn);box-shadow:0 0 8px var(--grn)}
.conn-dot.connecting{background:var(--org);animation:pulse 1s infinite}

/* Feedback section */
.fb-card{background:var(--surf);border:1px solid var(--bdr);border-radius:13px;
  padding:1.4rem;margin-top:2rem;margin-bottom:2rem}
.fb-list{margin-top:1rem;display:flex;flex-direction:column;gap:.65rem;max-height:380px;overflow-y:auto}
.fb-item{background:var(--surf2);border:1px solid var(--bdr);border-radius:9px;padding:.85rem}
.fb-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:.4rem;flex-wrap:wrap;gap:.3rem}
.fb-name{font-size:.8rem;font-weight:700;color:var(--txt)}
.fb-type{font-size:.65rem;font-family:var(--mono);padding:.2rem .55rem;border-radius:100px}
.fb-bug{background:rgba(255,23,68,.12);color:var(--red);border:1px solid rgba(255,23,68,.3)}
.fb-sugg{background:rgba(0,229,255,.1);color:var(--accent);border:1px solid rgba(0,229,255,.25)}
.fb-praise{background:rgba(0,230,118,.1);color:var(--grn);border:1px solid rgba(0,230,118,.25)}
.fb-other{background:rgba(255,109,0,.1);color:var(--org);border:1px solid rgba(255,109,0,.25)}
.fb-time{font-size:.62rem;color:var(--muted);font-family:var(--mono)}
.fb-msg{font-size:.8rem;color:var(--muted);line-height:1.6;margin-top:.25rem}
.stars{color:#ffd700;font-size:.85rem;letter-spacing:.05rem}
.no-fb{text-align:center;color:var(--muted);font-size:.8rem;padding:1.5rem;font-family:var(--mono)}

/* Footer */
footer{text-align:center;padding:1.5rem 0 2rem;color:var(--muted);font-size:.72rem;font-family:var(--mono)}
footer a{color:var(--accent);text-decoration:none}
</style>
</head>
<body>
<div class="wrap">

<header>
  <div class="logo">
    <div class="logo-icon">₹</div>
    <div class="logo-text">GST Recon</div>
  </div>
  <h1>Annual GST <span>Reconciliation Portal</span></h1>
  <p class="sub">Upload your returns → Get reconciliation Excel in seconds</p>
  <div class="badges">
    <span class="badge badge-grn">⭐ 100% Free</span>
    <span class="badge badge-blue">📊 Full Reconciliation</span>
    <span class="badge badge-purple">🔒 Your files stay private</span>
    <span class="badge badge-orange">🌐 Auto Download NEW</span>
  </div>
  <div id="global-dl-bar" style="display:none;margin-top:.8rem;padding:.55rem 1rem;
       background:rgba(0,230,118,.1);border:1px solid rgba(0,230,118,.35);border-radius:9px;
       align-items:center;gap:.75rem;flex-wrap:wrap;justify-content:center">
    <span style="font-size:.74rem;font-weight:700;color:var(--grn);font-family:var(--mono)">
      ✅ Files ready across tabs
    </span>
    <button onclick="globalDownloadAll()"
            style="padding:.42rem 1.1rem;background:linear-gradient(135deg,var(--grn),#00c853);
                   border:none;border-radius:7px;color:#000;font-family:var(--sans);font-size:.78rem;
                   font-weight:800;cursor:pointer;letter-spacing:.04em;white-space:nowrap">
      ⬇ Download All Files (All Tabs)
    </button>
    <span id="global-dl-count" style="font-size:.68rem;color:var(--muted);font-family:var(--mono)"></span>
  </div>
</header>

<!-- TABS -->
<div class="tabs">
  <button class="tb active" onclick="switchTab('recon',event)">📊 Reconciliation</button>
  <button class="tb" onclick="switchTab('gstr1',event)">📋 GSTR-1 Detail</button>
  <button class="tb" onclick="switchTab('gstr2b',event)">🏦 GSTR-2B Detail</button>
  <button class="tb" onclick="switchTab('dlstatus',event)">🔄 Download Status</button>
  <button class="tb" onclick="switchTab('autodl',event)">🌐 Auto Download</button>
  <button class="tb" onclick="switchTab('bulk',event)">📋 Bulk Download</button>
  <button class="tb" onclick="switchTab('itbulk',event)">📋 IT Bulk Download</button>
  <button class="tb" onclick="switchTab('itrecon',event)">🏦 Income Tax</button>
</div>

<!-- ══ TAB 1: RECONCILIATION ══ -->
<div class="tp active" id="tab-recon">

<div class="info-box">
  <strong>How it works:</strong> Download your GST return files using <strong>RUN_ME.bat</strong> on your PC,
  then upload them here. The portal generates a full
  <strong>Annual Reconciliation Excel</strong> (7 sheets: Summary, GSTR-1 Sales, GSTR-2B ITC,
  GSTR-2A Purchases, GSTR-3B Status, R1 vs 3B Recon, Tax Liability) plus a separate
  <strong>GSTR-1 Full Detail Excel</strong> (13 sheets) — instantly.
  <br><br>
  <strong>Files are auto-deleted after 2 hours. Nothing is stored permanently.</strong>
</div>

<form id="recon-form">
<div class="card">
  <div class="ct">Client Details</div>
  <div class="fg2">
    <div class="fg"><label>GSTIN *</label>
      <input type="text" id="r-gstin" placeholder="33ABCDE1234F1ZX" maxlength="15" required></div>
    <div class="fg"><label>Company Name *</label>
      <input type="text" id="r-name" placeholder="ABC Traders" required></div>
    <div class="fg"><label>Financial Year</label>
      <select id="r-fy">
        <option value="2026-27">2026-27</option>
        <option value="2025-26" selected>2025-26</option>
        <option value="2024-25">2024-25</option>
        <option value="2023-24">2023-24</option>
        <option value="2022-23">2022-23</option>
        <option value="2021-22">2021-22</option>
        <option value="2020-21">2020-21</option>
      </select></div>
    <div class="fg"><label>State (optional)</label>
      <input type="text" id="r-state" placeholder="Tamil Nadu"></div>
  </div>
</div>

<div class="card">
  <div class="ct">Upload Return Files</div>
  <div class="dg">
    <div class="dz" id="zone-r1">
      <div class="dz-ic">📋</div><div class="dz-lb">GSTR-1</div>
      <div class="dz-ht">ZIP files (12 months)</div>
      <div class="dz-cn" id="cnt-r1">No files</div>
      <input type="file" multiple accept=".zip,.json" data-zone="r1" onchange="updateZone('r1',this)">
    </div>
    <div class="dz" id="zone-r1a">
      <div class="dz-ic">📋</div><div class="dz-lb">GSTR-1A</div>
      <div class="dz-ht">ZIP files (amendments)</div>
      <div class="dz-cn" id="cnt-r1a">No files</div>
      <input type="file" multiple accept=".zip,.json" data-zone="r1a" onchange="updateZone('r1a',this)">
    </div>
    <div class="dz" id="zone-r2b">
      <div class="dz-ic">🏦</div><div class="dz-lb">GSTR-2B</div>
      <div class="dz-ht">Excel (.xlsx)</div>
      <div class="dz-cn" id="cnt-r2b">No files</div>
      <input type="file" multiple accept=".xlsx,.xls,.zip" data-zone="r2b" onchange="updateZone('r2b',this)">
    </div>
    <div class="dz" id="zone-r2a">
      <div class="dz-ic">📊</div><div class="dz-lb">GSTR-2A</div>
      <div class="dz-ht">ZIP or Excel</div>
      <div class="dz-cn" id="cnt-r2a">No files</div>
      <input type="file" multiple accept=".zip,.xlsx" data-zone="r2a" onchange="updateZone('r2a',this)">
    </div>
    <div class="dz" id="zone-r3b">
      <div class="dz-ic">📄</div><div class="dz-lb">GSTR-3B</div>
      <div class="dz-ht">PDF files</div>
      <div class="dz-cn" id="cnt-r3b">No files</div>
      <input type="file" multiple accept=".pdf" data-zone="r3b" onchange="updateZone('r3b',this)">
    </div>
    <div class="dz" id="zone-cust">
      <div class="dz-ic">👥</div><div class="dz-lb">Customer Names</div>
      <div class="dz-ht">GSTIN → Name Excel</div>
      <div class="dz-cn" id="cnt-cust">No file</div>
      <input type="file" accept=".xlsx,.xls" data-zone="cust" onchange="updateZone('cust',this)">
    </div>
    <div class="dz" id="zone-taxlib">
      <div class="dz-ic">📑</div><div class="dz-lb">Tax Liability</div>
      <div class="dz-ht">Portal Excel export</div>
      <div class="dz-cn" id="cnt-taxlib">No file</div>
      <input type="file" accept=".xlsx,.xls" data-zone="taxlib" onchange="updateZone('taxlib',this)">
    </div>
  </div>
</div>
<div class="card" style="display:flex;gap:.65rem;align-items:stretch;flex-wrap:wrap">
  <button type="submit" class="btn" id="r-submit" style="flex:1;margin-top:0">Generate Reconciliation + GSTR-1 Detail →</button>
  <button type="button" onclick="resetRecon()" id="r-reset"
          style="flex:0 0 auto;padding:.8rem 1.4rem;background:var(--surf2);
                 border:1px solid var(--red);border-radius:10px;color:var(--red);
                 font-family:var(--sans);font-size:.82rem;font-weight:700;
                 cursor:pointer;transition:all .15s;white-space:nowrap;margin-top:0"
          title="Clear all files and reset the form">
    🔄 Reset
  </button>
</div>
</form>

<div class="card pw" id="r-pw">
  <div class="ct">Processing <span class="sbg s-p pulse" id="r-badge">Running</span></div>
  <div class="pb-w"><div class="pb" id="r-pb"></div></div>
  <div class="lb" id="r-lb"></div>
</div>
<div class="card dw" id="r-dw">
  <div class="ct">Downloads Ready
    <button id="r-dl-all-btn" onclick="showDownloads._dlAll && showDownloads._dlAll()"
            style="display:none;margin-left:auto;padding:.3rem .9rem;background:linear-gradient(135deg,var(--grn),#00c853);
                   border:none;border-radius:7px;color:#000;font-family:var(--mono);font-size:.7rem;
                   font-weight:800;cursor:pointer;letter-spacing:.04em">
      ⬇ Download All
    </button>
  </div>
  <div class="dl-g" id="r-dlg"></div>
  <p style="color:var(--muted);font-size:.66rem;margin-top:.65rem;font-family:var(--mono)">
    ⏳ Files deleted automatically after 2 hours. Download before closing.
  </p>
</div>
</div><!-- /tab-recon -->

<!-- ══ TAB 2: GSTR-1 DETAIL ══ -->
<div class="tp" id="tab-gstr1">
<div class="card">
  <div class="ct">GSTR-1 Comprehensive Extraction — 13 Sheets</div>
  <div class="pills">
    <span class="pill">B2B Invoices</span><span class="pill">B2B Item Detail</span>
    <span class="pill">HSN Summary</span><span class="pill">B2CS / B2CL</span>
    <span class="pill">Credit Notes</span><span class="pill">Debit Notes</span>
    <span class="pill">Exports</span><span class="pill">Nil Rated</span>
    <span class="pill">GSTR-1A Amendments</span><span class="pill">Document Summary</span>
    <span class="pill">Master Summary</span>
  </div>
  <p style="color:var(--muted);font-size:.78rem;line-height:1.6">
    Upload all 12 GSTR-1 ZIP files. Customer names are auto-looked up from GSTR-2B/2A files.
    Add <strong style="color:var(--txt)">customer_names.xlsx</strong> (GSTIN + Name columns) for manual lookup.
  </p>
</div>
<form id="g1-form">
<div class="card">
  <div class="ct">Client Details</div>
  <div class="fg2">
    <div class="fg"><label>GSTIN *</label>
      <input type="text" id="g1-gstin" placeholder="33ABCDE1234F1ZX" maxlength="15" required></div>
    <div class="fg"><label>Company Name *</label>
      <input type="text" id="g1-name" placeholder="ABC Traders" required></div>
    <div class="fg"><label>Financial Year</label>
      <select id="g1-fy">
        <option value="2026-27">2026-27</option>
        <option value="2025-26" selected>2025-26</option>
        <option value="2024-25">2024-25</option>
        <option value="2023-24">2023-24</option>
        <option value="2022-23">2022-23</option>
        <option value="2021-22">2021-22</option>
        <option value="2020-21">2020-21</option>
      </select></div>
  </div>
</div>
<div class="card">
  <div class="ct">Upload Files</div>
  <div class="dg">
    <div class="dz" id="zone-g1r1">
      <div class="dz-ic">📋</div><div class="dz-lb">GSTR-1 ZIPs</div>
      <div class="dz-ht">All 12 months</div>
      <div class="dz-cn" id="cnt-g1r1">No files</div>
      <input type="file" multiple accept=".zip" data-zone="g1r1" onchange="updateZone('g1r1',this)">
    </div>
    <div class="dz" id="zone-g1r2b">
      <div class="dz-ic">🏦</div><div class="dz-lb">GSTR-2B / 2A</div>
      <div class="dz-ht">For name lookup</div>
      <div class="dz-cn" id="cnt-g1r2b">No files</div>
      <input type="file" multiple accept=".xlsx,.zip" data-zone="g1r2b" onchange="updateZone('g1r2b',this)">
    </div>
    <div class="dz" id="zone-g1cust">
      <div class="dz-ic">👥</div><div class="dz-lb">Customer Names</div>
      <div class="dz-ht">GSTIN → Name Excel</div>
      <div class="dz-cn" id="cnt-g1cust">No file</div>
      <input type="file" accept=".xlsx" data-zone="g1cust" onchange="updateZone('g1cust',this)">
    </div>
  </div>
</div>
<div class="card">
  <button type="submit" class="btn" id="g1-submit">Generate GSTR-1 Full Detail Excel →</button>
</div>
</form>
<div class="card pw" id="g1-pw">
  <div class="ct">Extracting <span class="sbg s-p pulse" id="g1-badge">Running</span></div>
  <div class="pb-w"><div class="pb" id="g1-pb"></div></div>
  <div class="lb" id="g1-lb"></div>
</div>
<div class="card dw" id="g1-dw">
  <div class="ct">GSTR-1 Detail Ready
    <button id="g1-dl-all-btn" onclick="if(window._g1DlAll)window._g1DlAll()"
            style="display:none;margin-left:auto;padding:.3rem .9rem;background:linear-gradient(135deg,var(--grn),#00c853);
                   border:none;border-radius:7px;color:#000;font-family:var(--mono);font-size:.7rem;
                   font-weight:800;cursor:pointer;letter-spacing:.04em">
      ⬇ Download All
    </button>
  </div>
  <div class="dl-g" id="g1-dlg"></div>
  <p style="color:var(--muted);font-size:.66rem;margin-top:.65rem;font-family:var(--mono)">
    ⏳ Files deleted automatically after 2 hours. Download before closing.
  </p>
</div>
</div><!-- /tab-gstr1 -->

<!-- ══ TAB: GSTR-2B DETAIL ══ -->
<div class="tp" id="tab-gstr2b">
<div class="card">
  <div class="ct">GSTR-2B Full-Year Extraction — 5 Sheets</div>
  <div class="pills">
    <span class="pill">B2B Invoices (All Months)</span>
    <span class="pill">Credit / Debit Notes</span>
    <span class="pill">ITC Ineligible</span>
    <span class="pill">Supplier Summary</span>
    <span class="pill">Annual Summary</span>
  </div>
  <p style="color:var(--muted);font-size:.78rem;line-height:1.6">
    Upload all GSTR-2B Excel files (one per month — portal downloads as <strong style="color:var(--txt)">.xlsx</strong>).
    All months are merged into <strong style="color:var(--txt)">one combined Excel</strong> with no month-name column in data rows.
    Supplier names, ITC eligibility and totals are extracted automatically.
  </p>
</div>
<form id="g2b-form">
<div class="card">
  <div class="ct">Client Details</div>
  <div class="fg2">
    <div class="fg"><label>GSTIN *</label>
      <input type="text" id="g2b-gstin" placeholder="33ABCDE1234F1ZX" maxlength="15" required></div>
    <div class="fg"><label>Company Name *</label>
      <input type="text" id="g2b-name" placeholder="ABC Traders" required></div>
    <div class="fg"><label>Financial Year</label>
      <select id="g2b-fy">
        <option value="2026-27">2026-27</option>
        <option value="2025-26" selected>2025-26</option>
        <option value="2024-25">2024-25</option>
        <option value="2023-24">2023-24</option>
        <option value="2022-23">2022-23</option>
        <option value="2021-22">2021-22</option>
        <option value="2020-21">2020-21</option>
      </select></div>
  </div>
</div>
<div class="card">
  <div class="ct">Upload GSTR-2B Excel Files</div>
  <div class="dg">
    <div class="dz" id="zone-g2b-files">
      <div class="dz-ic">🏦</div><div class="dz-lb">GSTR-2B Excel</div>
      <div class="dz-ht">All 12 months (.xlsx or .zip)</div>
      <div class="dz-cn" id="cnt-g2b-files">No files</div>
      <input type="file" multiple accept=".xlsx,.xls,.zip" data-zone="g2b-files" onchange="updateZone('g2b-files',this)">
    </div>
  </div>
  <p style="color:var(--muted);font-size:.72rem;margin-top:.6rem;font-family:var(--mono)">
    Expected file names:  GSTR2B_April_2025.xlsx · GSTR2B_May_2025.xlsx · ... or portal-format names.
  </p>
</div>
<div class="card">
  <button type="submit" class="btn" id="g2b-submit">Generate GSTR-2B Full-Year Excel →</button>
</div>
</form>
<div class="card pw" id="g2b-pw">
  <div class="ct">Extracting <span class="sbg s-p pulse" id="g2b-badge">Running</span></div>
  <div class="pb-w"><div class="pb" id="g2b-pb"></div></div>
  <div class="lb" id="g2b-lb"></div>
</div>
<div class="card dw" id="g2b-dw">
  <div class="ct">GSTR-2B Detail Ready
    <button id="g2b-dl-all-btn" onclick="if(window._g2bDlAll)window._g2bDlAll()"
            style="display:none;margin-left:auto;padding:.3rem .9rem;background:linear-gradient(135deg,var(--grn),#00c853);
                   border:none;border-radius:7px;color:#000;font-family:var(--mono);font-size:.7rem;
                   font-weight:800;cursor:pointer;letter-spacing:.04em">
      ⬇ Download All
    </button>
  </div>
  <div class="dl-g" id="g2b-dlg"></div>
  <p style="color:var(--muted);font-size:.66rem;margin-top:.65rem;font-family:var(--mono)">
    ⏳ Files deleted automatically after 2 hours. Download before closing.
  </p>
</div>
</div><!-- /tab-gstr2b -->
<div class="tp" id="tab-dlstatus">
<div class="card">
  <div class="ct">Download Status — 12 Months × 5 Returns</div>
  <p style="color:var(--muted);font-size:.78rem;line-height:1.6;margin-bottom:.8rem">
    After running <strong style="color:var(--txt)">RUN_ME.bat</strong> on your PC,
    upload the <strong style="color:var(--txt)">MASTER_REPORT_*.xlsx</strong> file here
    to see which returns downloaded successfully. Or paste a live Job ID from Tab 1.
  </p>
  <div class="fg2" style="margin-bottom:.75rem">
    <div class="fg">
      <label>Live Job ID (from Reconciliation tab)</label>
      <input type="text" id="ds-jid" placeholder="e.g. a3f2c9b1">
    </div>
    <div class="fg">
      <label>Upload MASTER_REPORT Excel</label>
      <div class="dz" id="zone-master"
           style="min-height:60px;flex-direction:row;padding:.6rem .75rem;gap:.65rem">
        <div class="dz-ic" style="font-size:1.2rem">📊</div>
        <div style="text-align:left">
          <div class="dz-lb">MASTER_REPORT*.xlsx</div>
          <div class="dz-cn" id="cnt-master">No file</div>
        </div>
        <input type="file" accept=".xlsx" data-zone="master" onchange="updateZone('master',this)">
      </div>
    </div>
  </div>
  <button class="btn" style="margin-top:0" onclick="loadDlStatus()">Load Status →</button>
</div>
<div class="card" id="ds-result" style="display:none">
  <div class="ct">Status — <span id="ds-title">—</span></div>
  <div style="overflow-x:auto">
    <table class="dst">
      <thead><tr>
        <th style="text-align:left">Month</th>
        <th>GSTR-1</th><th>GSTR-1A</th><th>GSTR-2B</th><th>GSTR-2A</th><th>GSTR-3B</th>
        <th>Summary</th>
      </tr></thead>
      <tbody id="ds-tb"></tbody>
    </table>
  </div>
  <div id="ds-sum" style="margin-top:.55rem;font-size:.7rem;font-family:var(--mono);color:var(--muted)"></div>
</div>
<div class="card pw" id="ds-pw">
  <div class="ct">Job Progress <span class="sbg s-p pulse" id="ds-badge">Running</span></div>
  <div class="pb-w"><div class="pb" id="ds-pb"></div></div>
  <div class="lb" id="ds-lb"></div>
</div>
</div><!-- /tab-dlstatus -->

<!-- ══ TAB 4: AUTO DOWNLOAD ══ -->
<div class="tp" id="tab-autodl">

<div class="card">
  <div class="ct">🌐 Auto Download from GST Portal</div>
  <div class="pills">
    <span class="pill">GSTR-1</span><span class="pill">GSTR-1A</span>
    <span class="pill">GSTR-2B</span><span class="pill">GSTR-2A</span><span class="pill">GSTR-3B</span>
  </div>
  <div class="info-box" style="margin-top:.7rem">
    <strong>How it works:</strong> Login to GST portal in your browser → copy your session token → paste it here. The server uses that token to download all your returns automatically. <strong style="color:var(--txt)">No CAPTCHA issues. No software needed.</strong>
  </div>
</div>

<!-- Step 1: How it works — PHASED DOWNLOAD -->
<div class="card" id="ad-step1">
  <div class="ct">How It Works — Phased Download (v11)</div>
  <div style="font-size:.78rem;color:var(--muted);line-height:1.9">
    1. Fill in your GSTIN, Company Name, Username &amp; Password below<br>
    2. Select your <strong style="color:var(--accent)">Download Case</strong> — choose what you need today<br>
    3. Click <strong style="color:var(--org)">🚀 Start Auto Download</strong><br>
    4. A CAPTCHA screenshot appears here — type it and click <strong style="color:var(--accent)">Submit</strong><br>
    5. Script follows <strong style="color:var(--txt)">Phased Download flow</strong> automatically (see phases below)<br>
    6. Files appear below — also auto-loaded into the <strong style="color:var(--txt)">Reconciliation tab</strong>
  </div>

  <!-- Phased Download Flow Display -->
  <div style="margin-top:1rem;display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:.5rem">
    <div style="background:rgba(0,229,255,.07);border:1px solid rgba(0,229,255,.2);border-radius:8px;padding:.7rem;text-align:center">
      <div style="font-size:.65rem;font-weight:700;letter-spacing:.08em;color:var(--accent);margin-bottom:.3rem">PHASE 1</div>
      <div style="font-size:.72rem;color:var(--txt)">Click <strong>GENERATE</strong></div>
      <div style="font-size:.65rem;color:var(--muted);margin-top:.2rem">GSTR-1 + GSTR-2A<br>All 12 months</div>
    </div>
    <div style="background:rgba(0,230,118,.07);border:1px solid rgba(0,230,118,.2);border-radius:8px;padding:.7rem;text-align:center">
      <div style="font-size:.65rem;font-weight:700;letter-spacing:.08em;color:var(--grn);margin-bottom:.3rem">PHASE 2</div>
      <div style="font-size:.72rem;color:var(--txt)">Direct <strong>Download</strong></div>
      <div style="font-size:.65rem;color:var(--muted);margin-top:.2rem">GSTR-2B<br>Keeps session alive ✓</div>
    </div>
    <div style="background:rgba(0,230,118,.07);border:1px solid rgba(0,230,118,.2);border-radius:8px;padding:.7rem;text-align:center">
      <div style="font-size:.65rem;font-weight:700;letter-spacing:.08em;color:var(--grn);margin-bottom:.3rem">PHASE 3</div>
      <div style="font-size:.72rem;color:var(--txt)">Direct <strong>Download</strong></div>
      <div style="font-size:.65rem;color:var(--muted);margin-top:.2rem">GSTR-3B<br>Keeps session alive ✓</div>
    </div>
    <div style="background:rgba(0,229,255,.07);border:1px solid rgba(0,229,255,.2);border-radius:8px;padding:.7rem;text-align:center">
      <div style="font-size:.65rem;font-weight:700;letter-spacing:.08em;color:var(--accent);margin-bottom:.3rem">PHASE 4</div>
      <div style="font-size:.72rem;color:var(--txt)">Click <strong>DOWNLOAD LINK</strong></div>
      <div style="font-size:.65rem;color:var(--muted);margin-top:.2rem">GSTR-1 + GSTR-2A<br>Now ready ✓</div>
    </div>
  </div>

  <!-- Retry + Session info -->
  <div style="margin-top:.8rem;display:grid;grid-template-columns:1fr 1fr;gap:.5rem">
    <div style="background:rgba(255,23,68,.07);border:1px solid rgba(255,23,68,.2);border-radius:8px;padding:.65rem">
      <div style="font-size:.65rem;font-weight:700;color:var(--red);margin-bottom:.3rem">🔴 RETRY LOGIC</div>
      <div style="font-size:.67rem;color:var(--muted);line-height:1.7">
        T1 → fail → wait 30s<br>T2 → fail → wait 60s<br>T3 → fail → reported in log<br>
        <span style="color:var(--txt)">3 retries per month per return</span>
      </div>
    </div>
    <div style="background:rgba(0,230,118,.07);border:1px solid rgba(0,230,118,.2);border-radius:8px;padding:.65rem">
      <div style="font-size:.65rem;font-weight:700;color:var(--grn);margin-bottom:.3rem">✅ SESSION TIPS</div>
      <div style="font-size:.67rem;color:var(--muted);line-height:1.7">
        Session expires: 20 min idle<br>Phase 2&3 keep session alive<br>Not Found → auto F5 reload<br>
        <span style="color:var(--txt)">Batch: 6 months at a time</span>
      </div>
    </div>
  </div>

  <div class="info-box" style="margin-top:.8rem;font-size:.72rem">
    Your password is used only for this session and is never stored.
  </div>
</div>

<!-- Step 2: Enter details -->
<form id="ad-form">
<div class="card">
  <div class="ct">Enter Your GST Portal Details</div>
  <div class="fg2">
    <div class="fg"><label>GSTIN *</label>
      <input type="text" id="ad-gstin" placeholder="33ABCDE1234F1ZX" maxlength="15" required></div>
    <div class="fg"><label>Company Name *</label>
      <input type="text" id="ad-name" placeholder="ABC Traders" required></div>
    <div class="fg"><label>Username *</label>
      <input type="text" id="ad-username" placeholder="Your GST portal username" required></div>
    <div class="fg"><label>Password *</label>
      <input type="password" id="ad-password" placeholder="Your GST portal password" required></div>
    <div class="fg"><label>Financial Year</label>
      <select id="ad-fy">
        <option value="2026-27">2026-27</option>
        <option value="2025-26" selected>2025-26</option>
        <option value="2024-25">2024-25</option>
        <option value="2023-24">2023-24</option>
        <option value="2022-23">2022-23</option>
        <option value="2021-22">2021-22</option>
        <option value="2020-21">2020-21</option>
      </select></div>
    <div class="fg"><label>Download Case — Select What You Need Today</label>
      <select id="ad-returns" onchange="updateCaseInfo()">
        <optgroup label="── FULL SUITE ──────────────────────────">
          <option value="all">📋 Full Suite — GSTR-1 + 1A + 2B + 2A + 3B  (~20 min, phased)</option>
        </optgroup>
        <optgroup label="── ⚡ FAST CASES (No generate, direct download) ──">
          <option value="gstr2b_3b">⚡ FAST — GSTR-2B + GSTR-3B only  (~10 min)</option>
          <option value="gstr2b">⚡ FAST — GSTR-2B only  (~5 min)</option>
          <option value="gstr3b">⚡ FAST — GSTR-3B only  (~5 min)</option>
        </optgroup>
        <optgroup label="── INDIVIDUAL RETURNS ───────────────────">
          <option value="gstr1">GSTR-1 only  (Generate first)</option>
          <option value="gstr1a">GSTR-1A only  (If available)</option>
          <option value="gstr2a">GSTR-2A only  (Generate first)</option>
        </optgroup>
        <optgroup label="── COMBO CASES ─────────────────────────">
          <option value="gstr1_2b_3b">GSTR-1 + GSTR-2B + GSTR-3B  (No 2A, phased)</option>
          <option value="gstr1_1a_2b_3b">GSTR-1 + GSTR-1A + GSTR-2B + GSTR-3B  (No 2A, phased)</option>
          <option value="gstr1_2b_2a">GSTR-1 + GSTR-2B + GSTR-2A  (No 3B, phased)</option>
        </optgroup>
      </select>
      <!-- Case info badge shown dynamically -->
      <div id="ad-case-info" style="margin-top:.4rem;font-size:.68rem;color:var(--muted);
           font-family:var(--mono);padding:.3rem .6rem;background:var(--surf2);
           border-radius:5px;border:1px solid var(--bdr)">
        📋 Full Suite | 48 tabs | ~20 min | Phased: Generate → 2B → 3B → Collect
      </div>
    </div>
  </div>
</div>
<div class="card">
  <button type="submit" class="btn-orange" id="ad-submit">🚀 Start Auto Download</button>
</div>
</form>

<script>
// Case info descriptions matching our discussion
const CASE_INFO = {
  "all":            "📋 Full Suite | 48 tabs | ~20 min | Phase1: Generate(1,2A) → Phase2: 2B → Phase3: 3B → Phase4: Collect(1,2A)",
  "gstr2b_3b":      "⚡ FAST | 24 tabs | ~10 min | No generate needed — direct download both simultaneously",
  "gstr2b":         "⚡ FAST | 12 tabs | ~5 min  | No generate needed — direct download GSTR-2B",
  "gstr3b":         "⚡ FAST | 12 tabs | ~5 min  | No generate needed — direct download GSTR-3B (PDF)",
  "gstr1":          "📄 GSTR-1 | 12 tabs | ~20 min | Phase1: Generate → Phase4: Collect download link",
  "gstr1a":         "📄 GSTR-1A | 12 tabs | ~20 min | Only appears when supplier amends GSTR-1",
  "gstr2a":         "📄 GSTR-2A | 12 tabs | ~20 min | Phase1: Generate → Phase4: Collect download link",
  "gstr1_2b_3b":    "📋 Combo | 36 tabs | ~15 min | Phase1: Generate(1) → Phase2: 2B → Phase3: 3B → Phase4: Collect(1)",
  "gstr1_1a_2b_3b": "📋 Combo | 48 tabs | ~20 min | Phase1: Generate(1,1A) → Phase2: 2B → Phase3: 3B → Phase4: Collect",
  "gstr1_2b_2a":    "📋 Combo | 36 tabs | ~20 min | Phase1: Generate(1,2A) → Phase2: 2B → Phase4: Collect(1,2A)",
};
function updateCaseInfo(){
  const v = document.getElementById('ad-returns').value;
  const el = document.getElementById('ad-case-info');
  if(el) el.textContent = CASE_INFO[v] || '';
}
</script>

<!-- Progress & logs -->
<div class="card pw" id="ad-pw">
  <div class="ct">Progress <span class="sbg s-p pulse" id="ad-badge">Running</span>
    <a id="ad-screenshot-link" href="#" target="_blank"
       style="display:none;margin-left:.8rem;font-size:.7rem;padding:.2rem .6rem;
              background:rgba(0,229,255,.12);border:1px solid var(--accent);border-radius:5px;
              color:var(--accent);text-decoration:none;font-family:var(--mono)">
      🖥 View Live Screenshot
    </a>
  </div>
  <div class="pb-w"><div class="pb" id="ad-pb"></div></div>
  <div class="lb" id="ad-lb"></div>

  <!-- Live downloaded files tracker — shown inside progress card as files arrive -->
  <div id="ad-live-files" style="display:none;margin-top:.85rem">
    <div style="font-size:.7rem;font-weight:700;text-transform:uppercase;letter-spacing:.07em;
                color:var(--grn);margin-bottom:.5rem;display:flex;align-items:center;gap:.45rem">
      <span style="width:7px;height:7px;border-radius:50%;background:var(--grn);
                   display:inline-block;box-shadow:0 0 6px var(--grn);animation:pulse 1.2s infinite"></span>
      Downloaded So Far
      <span id="ad-live-count" style="color:var(--muted);font-weight:400;font-family:var(--mono)"></span>
    </div>
    <div id="ad-live-grid"
         style="display:grid;grid-template-columns:repeat(auto-fill,minmax(170px,1fr));gap:.5rem"></div>
  </div>
</div>

<!-- CAPTCHA card (shown when server browser needs CAPTCHA input) -->
<div class="card" id="ad-captcha-card" style="display:none">
  <div class="ct">🔐 CAPTCHA Required — Type Below</div>
  <div class="info-box" style="margin-bottom:.8rem;font-size:.75rem">
    The server has opened GST portal and needs you to solve the CAPTCHA.<br>
    <strong style="color:var(--txt)">Type exactly what you see in the image, then click Submit.</strong>
  </div>
  <div style="text-align:center;margin-bottom:.8rem">
    <img id="ad-captcha-img" src="" alt="CAPTCHA"
         style="max-width:100%;border-radius:8px;border:2px solid var(--accent);
                background:#fff;padding:4px;cursor:pointer"
         title="Click to refresh screenshot"
         onclick="refreshAdCaptcha()">
    <div style="font-size:.68rem;color:var(--muted);margin-top:.35rem;font-family:var(--mono)">
      Click image to refresh screenshot
    </div>
  </div>
  <div style="display:flex;gap:.5rem;align-items:center">
    <input type="text" id="ad-captcha-input"
           placeholder="Type CAPTCHA letters here"
           style="flex:1;font-size:.85rem;letter-spacing:.15em;text-transform:uppercase"
           onkeydown="if(event.key==='Enter')submitAdCaptcha()">
    <button onclick="submitAdCaptcha()"
            style="padding:.55rem 1.2rem;background:linear-gradient(135deg,var(--accent),var(--accent2));
                   border:none;border-radius:8px;color:#000;font-weight:800;font-size:.82rem;
                   cursor:pointer;white-space:nowrap">
      Submit →
    </button>
  </div>
  <div id="ad-captcha-err" style="color:var(--red);font-size:.72rem;margin-top:.4rem;font-family:var(--mono)"></div>
</div>

<!-- Re-login card (shown if token expires mid-download) -->
<div class="card" id="ad-relogin-card" style="display:none">
  <div class="ct">🔄 Token Expired — Re-login Required</div>
  <div class="info-box warn" style="margin-bottom:.75rem;font-size:.74rem">
    Your GST portal session expired during download.<br>
    <strong>Option A (Easy):</strong> <a href="https://services.gst.gov.in/services/login" target="_blank"
       style="color:var(--accent);font-weight:700">Login to GST Portal again →</a>
    then click the <strong style="color:#a78bfa">🔖 GST Token Capture</strong> bookmark — token will be sent automatically<br>
    <strong>Option B (Manual):</strong> Press F12 → Application → Cookies → copy fresh <strong>AuthToken</strong> → paste below
  </div>
  <div class="fg">
    <label>Fresh AuthToken *</label>
    <input type="text" id="ad-relogin-token" placeholder="Paste new AuthToken here" style="font-size:.72rem">
  </div>
  <button class="btn" onclick="submitAdRelogin()" id="ad-relogin-btn" style="margin-top:.5rem">
    Submit New Token →
  </button>
</div>

<!-- Download results -->
<div class="card dw" id="ad-dw">
  <div class="ct">✅ Downloaded Files
    <span id="ad-file-count" style="font-size:.7rem;color:var(--muted);font-family:var(--mono);margin-left:.4rem"></span>
  </div>
  <div class="dl-g" id="ad-dlg"></div>
  <div style="margin-top:1rem;padding:.85rem;background:rgba(0,229,255,.04);border:1px solid rgba(0,229,255,.15);border-radius:9px">
    <div style="font-size:.75rem;font-weight:700;color:var(--accent);margin-bottom:.5rem;text-transform:uppercase;letter-spacing:.05em">
      📤 Send to Reconciliation Tab
    </div>
    <p style="font-size:.72rem;color:var(--muted);line-height:1.6;margin-bottom:.65rem">
      Click below to automatically load all downloaded return files (GSTR-1, 2B, 2A, 3B) directly into the Reconciliation tab — no manual upload needed.
    </p>
    <div style="display:flex;gap:.6rem;flex-wrap:wrap;align-items:center">
      <button onclick="adTransferToRecon()" id="ad-transfer-btn"
              style="padding:.65rem 1.4rem;background:linear-gradient(135deg,var(--grn),#00c853);
                     border:none;border-radius:9px;color:#000;font-weight:800;font-size:.85rem;
                     cursor:pointer;transition:transform .15s;letter-spacing:.03em"
              onmouseover="this.style.transform='translateY(-2px)'"
              onmouseout="this.style.transform=''">
        📤 Load into Reconciliation Tab →
      </button>
      <span id="ad-transfer-status" style="font-size:.7rem;color:var(--muted);font-family:var(--mono)"></span>
    </div>
  </div>
  <p style="color:var(--muted);font-size:.68rem;margin-top:.65rem;font-family:var(--mono)">
    ⏳ Files deleted after 2 hours. Download ZIP before closing.
  </p>
</div>

<!-- Failure Screenshots Panel (auto download tab) -->
<div class="card" id="ad-fail-shots" style="display:none">
  <div class="ct" style="color:var(--red)">
    📸 Failure Screenshots
    <span id="ad-fail-count" style="font-size:.7rem;color:var(--muted);font-family:var(--mono);margin-left:.4rem"></span>
    <button onclick="refreshFailShots()"
            style="margin-left:auto;padding:.22rem .65rem;background:var(--surf2);
                   border:1px solid var(--bdr);border-radius:6px;color:var(--muted);
                   font-family:var(--mono);font-size:.65rem;cursor:pointer">
      🔄 Refresh
    </button>
  </div>
  <p style="color:var(--muted);font-size:.74rem;line-height:1.65;margin-bottom:.85rem">
    These screenshots were captured <strong style="color:var(--txt)">automatically</strong>
    at each point where a download failed or a button was not found —
    so you can see exactly what the GST portal showed at that moment.
    Share these with support if downloads are not working.
  </p>
  <div id="ad-fail-grid" style="display:flex;flex-direction:column;gap:1.1rem"></div>
</div>

</div><!-- /tab-autodl -->

<!-- ══ TAB 5: BULK DOWNLOAD ══ -->
<div class="tp" id="tab-bulk">

<div class="card">
  <div class="ct">📋 Bulk Download — Multiple Companies</div>
  <div class="pills">
    <span class="pill">Multiple GSTINs</span><span class="pill">One by one</span>
    <span class="pill">CAPTCHA per company</span><span class="pill">Auto ZIP output</span>
  </div>
  <div class="info-box" style="margin-top:.75rem">
    <strong>How it works:</strong><br>
    1. Download the Excel template below → fill in your company list<br>
    2. Upload it here → Click Start Bulk Download<br>
    3. For each company, login to GST portal → copy AuthToken → paste here → Download proceeds<br>
    4. All files are saved and available to download as a ZIP<br><br>
    <strong style="color:var(--grn)">✅ No limit on number of companies. Each uses its own login session.</strong>
  </div>
  <a href="/api/bulk-template" class="btn-dl"
     style="display:inline-block;padding:.55rem 1.1rem;margin-top:.5rem">
    ⬇ Download Excel Template
  </a>
</div>

<div class="card">
  <div class="ct">Upload Company List</div>
  <div class="fg2">
    <div class="fg">
      <label>Company List Excel *</label>
      <div class="dz" id="zone-bulk" style="min-height:70px;flex-direction:row;padding:.6rem .75rem;gap:.65rem">
        <div class="dz-ic" style="font-size:1.2rem">📊</div>
        <div style="text-align:left">
          <div class="dz-lb">companies.xlsx</div>
          <div class="dz-cn" id="cnt-bulk">No file</div>
        </div>
        <input type="file" accept=".xlsx,.xls" data-zone="bulk" onchange="updateZone('bulk',this)">
      </div>
    </div>
    <div class="fg">
      <label>Financial Year</label>
      <select id="bulk-fy">
        <option value="2026-27">2026-27</option>
        <option value="2025-26" selected>2025-26</option>
        <option value="2024-25">2024-25</option>
        <option value="2023-24">2023-24</option>
        <option value="2022-23">2022-23</option>
        <option value="2021-22">2021-22</option>
        <option value="2020-21">2020-21</option>
      </select>
    </div>
    <div class="fg">
      <label>Returns to Download</label>
      <select id="bulk-returns">
        <option value="all">All Returns (GSTR-1, 1A, 2B, 2A, 3B)</option>
        <option value="gstr1">GSTR-1 Only</option>
        <option value="gstr2b">GSTR-2B Only</option>
        <option value="gstr2a">GSTR-2A Only</option>
        <option value="gstr3b">GSTR-3B Only</option>
        <option value="gstr1_2b_3b">GSTR-1 + GSTR-2B + GSTR-3B (No 2A, phased)</option>
        <option value="gstr1_1a_2b_3b">GSTR-1 + GSTR-1A + GSTR-2B + GSTR-3B (No 2A, phased)</option>
        <option value="gstr2b_3b">⚡ FAST — GSTR-2B + GSTR-3B only (~10 min)</option>
        <option value="gstr1_2b_2a">GSTR-1 + GSTR-2B + GSTR-2A (No 3B, phased)</option>
      </select>
    </div>
  </div>
  <button class="btn-orange" onclick="startBulk()" id="bulk-submit" style="margin-top:.5rem">
    🚀 Start Bulk Download
  </button>
</div>

<!-- Per-company CAPTCHA card (shown one at a time) -->
<div class="card" id="bulk-captcha-card" style="display:none">
  <div class="ct">🔐 Login Required — <span id="bulk-co-name" style="color:var(--accent)"></span></div>
  <div class="info-box" style="margin-bottom:.8rem;font-size:.75rem">
    <strong>Step 1:</strong>
    <a href="https://services.gst.gov.in/services/login" target="_blank"
       style="color:var(--accent);font-weight:700">Open GST Portal Login →</a>
    login with username &amp; password + CAPTCHA<br>
    <strong>Step 2:</strong> Press F12 → Application → Cookies → copy <strong>AuthToken</strong><br>
    <strong>Step 3:</strong> Paste it below and click Submit
  </div>
  <div class="fg2">
    <div class="fg">
      <label>GSTIN</label>
      <input type="text" id="bulk-cap-gstin" readonly style="opacity:.6">
    </div>
    <div class="fg">
      <label>Username</label>
      <input type="text" id="bulk-cap-user" readonly style="opacity:.6">
    </div>
  </div>
  <div class="fg" style="margin-top:.6rem">
    <label>AuthToken from GST Portal Cookies *</label>
    <input type="text" id="bulk-token-input" placeholder="Paste AuthToken here"
           style="font-size:.72rem">
  </div>
  <button class="btn" onclick="submitBulkToken()" id="bulk-token-btn" style="margin-top:.5rem">
    Submit Token &amp; Download →
  </button>
  <div id="bulk-token-err" style="color:var(--red);font-size:.72rem;margin-top:.4rem;font-family:var(--mono)"></div>
</div>

<!-- Progress -->
<div class="card" id="bulk-pw" style="display:none">
  <div class="ct">Bulk Progress <span class="sbg s-p pulse" id="bulk-badge">Running</span>
    <span id="bulk-counter" style="font-size:.7rem;color:var(--muted);font-family:var(--mono);margin-left:.5rem"></span>
  </div>
  <div class="pb-w"><div class="pb" id="bulk-pb"></div></div>
  <div class="lb" id="bulk-lb"></div>
</div>

<!-- Results -->
<div class="card" id="bulk-dw" style="display:none">
  <div class="ct">✅ Bulk Download Complete</div>
  <div class="dl-g" id="bulk-dlg"></div>
  <p style="color:var(--muted);font-size:.66rem;margin-top:.65rem;font-family:var(--mono)">
    ⏳ Files deleted after 2 hours. Download the ZIP and upload each company's files to Reconciliation tab.
  </p>
</div>

</div><!-- /tab-bulk -->

<!-- ══ TAB: IT BULK DOWNLOAD ══ -->
<div class="tp" id="tab-itbulk">

<div class="card">
  <div class="ct">📋 IT Bulk Download — Multiple Clients</div>
  <div class="pills">
    <span class="pill">Form 26AS</span><span class="pill">AIS PDF</span>
    <span class="pill">TIS PDF</span><span class="pill">IT Recon Excel</span>
    <span class="pill">One-time OTP per client</span>
  </div>
  <div class="info-box" style="margin-top:.75rem">
    <strong>How it works:</strong><br>
    1. Download the Excel template → fill in client PAN, IT password, and (optional) GSTIN<br>
    2. Upload it here → click Start IT Bulk Download<br>
    3. For each client, server shows OTP/CAPTCHA screenshot → you enter it → download proceeds automatically<br>
    4. After first OTP per client, "Remember Device" is auto-clicked — future runs are fully automatic<br>
    5. All 26AS, AIS, TIS PDFs + IT Recon Excels are zipped for download<br><br>
    <strong style="color:var(--grn)">✅ No limit on clients. Each client uses its own login session.</strong>
  </div>
  <a href="/api/it-bulk-template" class="btn-dl"
     style="display:inline-block;padding:.55rem 1.1rem;margin-top:.5rem">
    ⬇ Download Client Template (Excel)
  </a>
</div>

<div class="card">
  <div class="ct">Upload Client List</div>
  <div class="fg2">
    <div class="fg">
      <label>Client List Excel *</label>
      <div class="dz" id="zone-itbulk" style="min-height:70px;flex-direction:row;padding:.6rem .75rem;gap:.65rem">
        <div class="dz-ic" style="font-size:1.2rem">📊</div>
        <div style="text-align:left">
          <div class="dz-lb">it_clients.xlsx</div>
          <div class="dz-cn" id="cnt-itbulk">No file</div>
        </div>
        <input type="file" accept=".xlsx,.xls" data-zone="itbulk" onchange="updateZone('itbulk',this)">
      </div>
    </div>
    <div class="fg">
      <label>Financial Year</label>
      <select id="itbulk-fy">
        <option value="2026-27">2026-27</option>
        <option value="2025-26" selected>2025-26</option>
        <option value="2024-25">2024-25</option>
        <option value="2023-24">2023-24</option>
      </select>
    </div>
    <div class="fg">
      <label>Download Options</label>
      <select id="itbulk-mode">
        <option value="all">All — 26AS + AIS + TIS + IT Recon Excel</option>
        <option value="pdfs">PDFs Only — 26AS + AIS + TIS</option>
        <option value="26as">26AS Only</option>
        <option value="ais_tis">AIS + TIS Only</option>
        <option value="recon">IT Recon Excel Only (from uploaded PDFs)</option>
      </select>
    </div>
  </div>
  <button class="btn-orange" onclick="startITBulk()" id="itbulk-submit" style="margin-top:.5rem">
    🚀 Start IT Bulk Download
  </button>
</div>

<!-- Per-client OTP card -->
<div class="card" id="itbulk-otp-card" style="display:none">
  <div class="ct">📱 OTP / Login Required — <span id="itbulk-client-name" style="color:var(--accent)"></span></div>
  <div class="info-box" style="margin-bottom:.8rem;font-size:.75rem">
    The IT Portal screenshot is shown below for this client.<br>
    <strong>If OTP was sent to mobile/email:</strong> Enter it below and click Submit.<br>
    <strong>If no OTP is needed:</strong> Type <strong style="color:var(--accent)">SKIP</strong> and click Submit.<br>
    <strong>After first OTP:</strong> "Remember Device" is auto-clicked — future runs skip OTP ✓
  </div>
  <div class="fg2" style="margin-bottom:.65rem;opacity:.6">
    <div class="fg">
      <label>PAN</label>
      <input type="text" id="itbulk-cap-pan" readonly>
    </div>
  </div>
  <div style="text-align:center;margin-bottom:.8rem">
    <img id="itbulk-otp-img" src="" alt="IT Portal Screenshot"
         style="max-width:100%;border-radius:8px;border:2px solid #a78bfa;background:#fff;padding:4px;cursor:pointer"
         onclick="itBulkRefreshShot()">
    <div style="font-size:.68rem;color:var(--muted);margin-top:.35rem;font-family:var(--mono)">Click to refresh screenshot</div>
  </div>
  <div style="display:flex;gap:.5rem;align-items:center">
    <input type="text" id="itbulk-otp-input" placeholder="Enter OTP, CAPTCHA, or type SKIP"
           style="flex:1;font-size:.85rem;letter-spacing:.1em"
           onkeydown="if(event.key==='Enter')submitITBulkOTP()">
    <button onclick="submitITBulkOTP()"
            style="padding:.55rem 1.2rem;background:linear-gradient(135deg,#a78bfa,var(--accent2));
                   border:none;border-radius:8px;color:#fff;font-weight:800;font-size:.82rem;cursor:pointer">
      Submit →
    </button>
  </div>
  <div id="itbulk-otp-err" style="color:var(--red);font-size:.72rem;margin-top:.4rem;font-family:var(--mono)"></div>
</div>

<!-- Progress -->
<div class="card" id="itbulk-pw" style="display:none">
  <div class="ct">IT Bulk Progress <span class="sbg s-p pulse" id="itbulk-badge">Running</span>
    <span id="itbulk-counter" style="font-size:.7rem;color:var(--muted);font-family:var(--mono);margin-left:.5rem"></span>
  </div>
  <div class="pb-w"><div class="pb" id="itbulk-pb"></div></div>
  <div class="lb" id="itbulk-lb"></div>
</div>

<!-- Results -->
<div class="card" id="itbulk-dw" style="display:none">
  <div class="ct">✅ IT Bulk Download Complete</div>
  <div class="dl-g" id="itbulk-dlg"></div>
  <p style="color:var(--muted);font-size:.66rem;margin-top:.65rem;font-family:var(--mono)">
    ⏳ Files deleted after 2 hours. Download the ZIP to save all client files.
  </p>
</div>

</div><!-- /tab-itbulk -->

<!-- ══ TAB 6: INCOME TAX RECONCILIATION ══ -->
<div class="tp" id="tab-itrecon">

<div class="info-box">
  <strong>Income Tax Reconciliation — What this does:</strong>
  Upload your <strong>26AS PDF</strong>, optionally the <strong>AIS PDF</strong> and <strong>TIS PDF</strong>
  (all from IT Portal), and optionally the <strong>GST Recon Excel</strong> from Tab 1 for auto-matched turnover.
  <br><br>
  The portal generates an Excel with 9 sheets:
  <strong>IT_Summary</strong> (key figures — TIS vs GSTR-1 turnover, TDS summary, advance tax) |
  <strong>TDS_26AS_Detail</strong> (all deductors Part A/A1/A2/B/C with every transaction line) |
  <strong>TIS_vs_GSTR_Annual</strong> (TIS categories vs GSTR-1, GSTR-1A & GSTR-3B annual totals) |
  <strong>TIS_vs_GSTR_Monthly</strong> (12 months × GSTIN, APR→MAR, with GSTR-1A & 3B) |
  <strong>Purchase_Detail</strong> (supplier-wise purchase from AIS/TIS with totals) |
  <strong>AIS_vs_Turnover</strong> (full reconciliation with blank rows for manual adjustments) |
  <strong>Advance_Tax_Challan</strong> (Part C challans + quarter-wise summary) |
  <strong>AIS_vs_GSTR_Monthly</strong> (12 months — GSTR-1, GSTR-1A, GSTR-3B Sales & Purchases) |
  <strong>IT_Filing_Checklist</strong> (40-item ITR verification checklist with auto-detected data).
  <br><br>
  💡 <strong>For best results upload all three PDFs</strong> — 26AS (TDS details) + AIS (purchase/income breakdown) + TIS (confirmed turnover figures used by IT dept). TIS data is the most important for ITR reconciliation.
  <br>
  <strong>Files are auto-deleted after 2 hours. Nothing stored permanently.</strong>
</div>

<form id="it-form">
<div class="card">
  <div class="ct">Company Details</div>
  <div class="fg2">
    <div class="fg"><label>Company Name *</label>
      <input type="text" id="it-name" placeholder="ABC Traders Pvt Ltd" required></div>
    <div class="fg"><label>PAN *</label>
      <input type="text" id="it-pan" placeholder="ABCDE1234F" maxlength="10"
             style="text-transform:uppercase" required></div>
    <div class="fg"><label>GSTIN (linked to PAN)</label>
      <input type="text" id="it-gstin" placeholder="33ABCDE1234F1ZX" maxlength="15"
             style="text-transform:uppercase"></div>
    <div class="fg"><label>Financial Year</label>
      <select id="it-fy">
        <option value="2026-27">2026-27</option>
        <option value="2025-26" selected>2025-26</option>
        <option value="2024-25">2024-25</option>
        <option value="2023-24">2023-24</option>
        <option value="2022-23">2022-23</option>
        <option value="2021-22">2021-22</option>
        <option value="2020-21">2020-21</option>
      </select>
    </div>
    <div class="fg"><label>ITR Form Type</label>
      <select id="it-itr-form">
        <option value="ITR-3">ITR-3 (Business/Profession with GST)</option>
        <option value="ITR-6">ITR-6 (Companies other than Sec 11)</option>
        <option value="ITR-5">ITR-5 (Firms / LLP / AOP / BOI)</option>
        <option value="ITR-4">ITR-4 (Sugam — Presumptive Income)</option>
        <option value="ITR-1">ITR-1 (Sahaj — Salary + Small Business)</option>
        <option value="ITR-2">ITR-2 (Capital Gains, no business)</option>
        <option value="ITR-7">ITR-7 (Trusts / Section 11)</option>
      </select>
    </div>
    <div class="fg"><label>Entity Type</label>
      <select id="it-entity">
        <option value="company">Private Limited Company</option>
        <option value="llp">LLP / Partnership Firm</option>
        <option value="proprietorship">Proprietorship</option>
        <option value="huf">HUF</option>
        <option value="trust">Trust / Society</option>
        <option value="individual">Individual</option>
      </select>
    </div>
  </div>
</div>

<div class="card">
  <div class="ct">Upload Files</div>
  <div class="dg">
    <div class="dz" id="zone-it26as">
      <div class="dz-ic">📄</div>
      <div class="dz-lb">Form 26AS</div>
      <div class="dz-ht">PDF from IT Portal</div>
      <div class="dz-cn" id="cnt-it26as">No file</div>
      <input type="file" accept=".pdf" data-zone="it26as" onchange="updateZone('it26as',this)">
    </div>
    <div class="dz" id="zone-itais">
      <div class="dz-ic">📊</div>
      <div class="dz-lb">AIS PDF</div>
      <div class="dz-ht">Annual Info Statement</div>
      <div class="dz-cn" id="cnt-itais">No file (optional)</div>
      <input type="file" accept=".pdf" data-zone="itais" onchange="updateZone('itais',this)">
    </div>
    <div class="dz" id="zone-ittis">
      <div class="dz-ic">📑</div>
      <div class="dz-lb">TIS PDF</div>
      <div class="dz-ht">Taxpayer Info Summary</div>
      <div class="dz-cn" id="cnt-ittis">No file (optional)</div>
      <input type="file" accept=".pdf" data-zone="ittis" onchange="updateZone('ittis',this)">
    </div>
    <div class="dz" id="zone-itgst">
      <div class="dz-ic">📋</div>
      <div class="dz-lb">GST Recon Excel</div>
      <div class="dz-ht">Output from Tab 1 (optional)</div>
      <div class="dz-cn" id="cnt-itgst">No file (optional)</div>
      <input type="file" accept=".xlsx,.xls" data-zone="itgst" onchange="updateZone('itgst',this)">
    </div>
  </div>
  <div class="info-box" style="margin-top:.75rem;font-size:.74rem">
    <strong>How to download 26AS:</strong>
    IT Portal (incometax.gov.in) → Login → e-File → Income Tax Returns → View Form 26AS →
    Select Assessment Year → Export to PDF<br>
    <strong>How to download AIS:</strong>
    IT Portal → Services → Annual Information Statement (AIS) → Download PDF<br>
    <strong>How to download TIS:</strong>
    IT Portal → Services → Annual Information Statement (AIS) → Switch to TIS tab → Download PDF
  </div>
</div>

<div class="card" style="display:flex;gap:.65rem;align-items:stretch;flex-wrap:wrap">
  <button type="submit" class="btn" id="it-submit" style="flex:1;margin-top:0">
    Generate IT Reconciliation Excel →
  </button>
  <button type="button" onclick="resetIT()" id="it-reset"
          style="flex:0 0 auto;padding:.8rem 1.4rem;background:var(--surf2);
                 border:1px solid var(--red);border-radius:10px;color:var(--red);
                 font-family:var(--sans);font-size:.82rem;font-weight:700;
                 cursor:pointer;transition:all .15s;white-space:nowrap;margin-top:0"
          title="Clear all files and reset">
    🔄 Reset
  </button>
</div>
</form>

<!-- IT Auto Download Section -->
<div class="card" style="border-color:rgba(124,58,237,.35)">
  <div class="ct" style="color:#a78bfa">🌐 Auto Download from IT Portal</div>
  <div class="info-box" style="margin-top:.6rem;font-size:.76rem">
    <strong>How it works:</strong> Enter your IT portal credentials below → click <strong style="color:var(--org)">Start IT Auto Download</strong> →
    OTP/CAPTCHA screenshot appears → enter it → server downloads <strong>26AS, AIS & TIS</strong> PDFs automatically.
    Files are auto-loaded into the upload zones above.
    <br><br>
    <strong>Your password is used only for this session and never stored.</strong>
  </div>
  <div class="fg2" style="margin-top:.75rem">
    <div class="fg"><label>IT Portal Username (PAN) *</label>
      <input type="text" id="it-ad-user" placeholder="ABCDE1234F" maxlength="10" style="text-transform:uppercase"></div>
    <div class="fg"><label>IT Portal Password *</label>
      <input type="password" id="it-ad-pass" placeholder="Your IT portal password"></div>
  </div>
  <button class="btn-orange" id="it-ad-btn" style="margin-top:.7rem" onclick="startITAutoDownload()">
    🌐 Start IT Auto Download (26AS + AIS + TIS)
  </button>
</div>

<!-- IT Auto Download Progress -->
<div class="card pw" id="it-ad-pw" style="display:none">
  <div class="ct">IT Portal Download <span class="sbg s-p pulse" id="it-ad-badge">Running</span>
    <a id="it-ad-ss-link" href="#" target="_blank" style="display:none;margin-left:.8rem;font-size:.7rem;padding:.2rem .6rem;
       background:rgba(0,229,255,.12);border:1px solid var(--accent);border-radius:5px;
       color:var(--accent);text-decoration:none;font-family:var(--mono)">🖥 View Screenshot</a>
  </div>
  <div class="pb-w"><div class="pb" id="it-ad-pb"></div></div>
  <div class="lb" id="it-ad-lb"></div>
</div>

<!-- IT Auto Download OTP/CAPTCHA card -->
<div class="card" id="it-ad-captcha-card" style="display:none">
  <div class="ct">🔐 OTP / CAPTCHA Required</div>
  <div class="info-box" style="margin-bottom:.8rem;font-size:.75rem">
    The IT Portal screenshot is shown below.<br>
    <strong>If OTP was sent to your phone/email:</strong> Enter it below and click Submit.<br>
    <strong>If no OTP needed:</strong> Type <strong style="color:var(--accent)">SKIP</strong> and click Submit.<br>
    <strong>If CAPTCHA shown:</strong> Type the CAPTCHA characters and click Submit.
  </div>
  <div style="text-align:center;margin-bottom:.8rem">
    <img id="it-ad-captcha-img" src="" alt="IT Portal Screenshot"
         style="max-width:100%;border-radius:8px;border:2px solid #a78bfa;background:#fff;padding:4px;cursor:pointer"
         onclick="itAdRefreshShot()">
    <div style="font-size:.68rem;color:var(--muted);margin-top:.35rem;font-family:var(--mono)">Click image to refresh</div>
  </div>
  <div style="display:flex;gap:.5rem;align-items:center">
    <input type="text" id="it-ad-captcha-input" placeholder="Enter OTP, CAPTCHA, or type SKIP"
           style="flex:1;font-size:.85rem;letter-spacing:.1em"
           onkeydown="if(event.key==='Enter')itAdSubmit()">
    <button onclick="itAdSubmit()"
            style="padding:.55rem 1.2rem;background:linear-gradient(135deg,#a78bfa,var(--accent2));
                   border:none;border-radius:8px;color:#fff;font-weight:800;font-size:.82rem;cursor:pointer">
      Submit →
    </button>
  </div>
  <div id="it-ad-captcha-err" style="color:var(--red);font-size:.72rem;margin-top:.4rem;font-family:var(--mono)"></div>
</div>

<!-- IT Auto Download results -->
<div class="card dw" id="it-ad-dw" style="display:none">
  <div class="ct">✅ IT Portal Downloads Ready</div>
  <div class="dl-g" id="it-ad-dlg"></div>
  <div style="margin-top:1rem;padding:.85rem;background:rgba(124,58,237,.05);border:1px solid rgba(124,58,237,.2);border-radius:9px">
    <div style="font-size:.75rem;font-weight:700;color:#a78bfa;margin-bottom:.5rem">📤 Load into Upload Zones Above</div>
    <p style="font-size:.72rem;color:var(--muted);margin-bottom:.65rem;line-height:1.6">
      Click to automatically load downloaded PDFs into the upload zones (26AS → 26AS zone, AIS → AIS zone, TIS → TIS zone).
    </p>
    <button onclick="itAdTransfer()"
            style="padding:.65rem 1.4rem;background:linear-gradient(135deg,#a78bfa,var(--accent2));
                   border:none;border-radius:9px;color:#fff;font-weight:800;font-size:.85rem;cursor:pointer">
      📤 Load PDFs into Upload Zones →
    </button>
    <span id="it-ad-transfer-status" style="font-size:.7rem;color:var(--muted);font-family:var(--mono);margin-left:.75rem"></span>
  </div>
  <p style="color:var(--muted);font-size:.66rem;margin-top:.65rem;font-family:var(--mono)">
    ⏳ Files deleted automatically after 2 hours. Download before closing.
  </p>
</div>


  <div class="ct">Processing <span class="sbg s-p pulse" id="it-badge">Running</span></div>
  <div class="pb-w"><div class="pb" id="it-pb"></div></div>
  <div class="lb" id="it-lb"></div>
</div>

<!-- Downloads -->
<div class="card dw" id="it-dw">
  <div class="ct">✅ IT Reconciliation Ready</div>
  <div class="dl-g" id="it-dlg"></div>
  <p style="color:var(--muted);font-size:.66rem;margin-top:.65rem;font-family:var(--mono)">
    ⏳ Files deleted automatically after 2 hours. Download before closing.
  </p>
</div>

</div><!-- /tab-itrecon -->


<!-- ══ FEEDBACK SECTION ══ -->
<div class="fb-card" id="feedback-section">
  <div class="ct">💬 Feedback &amp; Comments</div>
  <p style="color:var(--muted);font-size:.78rem;line-height:1.6;margin-bottom:1rem">
    This portal is in <strong style="color:var(--txt)">public beta</strong> — completely free.
    Found a bug? Have a suggestion? We'd love to hear from you.
    Your feedback helps improve this tool for all GST practitioners in India.
  </p>

  <form id="fb-form">
  <div class="fg2" style="margin-bottom:.65rem">
    <div class="fg">
      <label>Your Name</label>
      <input type="text" id="fb-name" placeholder="CA Rajesh Kumar (optional)">
    </div>
    <div class="fg">
      <label>Feedback Type</label>
      <select id="fb-type">
        <option value="bug">🐛 Bug Report</option>
        <option value="suggestion" selected>💡 Suggestion</option>
        <option value="praise">👍 Works Great!</option>
        <option value="other">💬 Other</option>
      </select>
    </div>
  </div>
  <div class="fg" style="margin-bottom:.65rem">
    <label>Rating</label>
    <div style="display:flex;gap:.5rem;margin-top:.25rem" id="star-row">
      <span class="star-btn" data-val="1" onclick="setRating(1)" style="font-size:1.4rem;cursor:pointer;opacity:.4">★</span>
      <span class="star-btn" data-val="2" onclick="setRating(2)" style="font-size:1.4rem;cursor:pointer;opacity:.4">★</span>
      <span class="star-btn" data-val="3" onclick="setRating(3)" style="font-size:1.4rem;cursor:pointer;opacity:.4">★</span>
      <span class="star-btn" data-val="4" onclick="setRating(4)" style="font-size:1.4rem;cursor:pointer;opacity:.4">★</span>
      <span class="star-btn" data-val="5" onclick="setRating(5)" style="font-size:1.4rem;cursor:pointer;opacity:.4">★</span>
      <span id="rating-lbl" style="font-size:.72rem;color:var(--muted);font-family:var(--mono);align-self:center;margin-left:.35rem"></span>
    </div>
  </div>
  <div class="fg" style="margin-bottom:.75rem">
    <label>Your Comment *</label>
    <textarea id="fb-msg" placeholder="e.g. GSTR-2B extraction is not matching for April month..." required></textarea>
  </div>
  <button type="submit" class="btn-sec" id="fb-submit">Submit Feedback →</button>
  <div id="fb-status" style="font-size:.72rem;margin-top:.45rem;font-family:var(--mono)"></div>
  </form>

  <!-- Existing feedback -->
  <div style="margin-top:1.4rem">
    <div style="font-size:.78rem;font-weight:700;color:var(--accent);text-transform:uppercase;
                letter-spacing:.06em;margin-bottom:.65rem">
      Recent Comments (<span id="fb-count">0</span>)
    </div>
    <div class="fb-list" id="fb-list">
      <div class="no-fb">No comments yet. Be the first!</div>
    </div>
  </div>
</div>

<footer>
  GST Reconciliation Portal — Public Beta &nbsp;|&nbsp;
  Built for Indian CA firms &nbsp;|&nbsp;
  100% Free · No data stored permanently<br>
  <span style="color:rgba(107,127,163,.5)">Scripts run on server only — your code is never shared</span>
</footer>

</div><!-- /wrap -->

<script>
// ── Tab switching ─────────────────────────────────────────────────
function switchTab(name, e){
  if(e) e.preventDefault();
  document.querySelectorAll('.tb').forEach(b=>b.classList.remove('active'));
  document.querySelectorAll('.tp').forEach(p=>p.classList.remove('active'));
  // Always find the button by matching the name in its onclick attribute
  document.querySelectorAll('.tb').forEach(b=>{
    const oc = b.getAttribute('onclick') || '';
    if(oc.includes("'"+name+"'") || oc.includes('"'+name+'"'))
      b.classList.add('active');
  });
  const tabEl = document.getElementById('tab-'+name);
  if(tabEl) tabEl.classList.add('active');
  
  // Check connection status when switching to autodl tab
  if(name==='autodl') checkBrowserConnection();
}

// ── File zones ────────────────────────────────────────────────────
const zoneFiles={};
function updateZone(zone, inp){
  const files=Array.from(inp.files);
  zoneFiles[zone]=files;
  const cn=document.getElementById('cnt-'+zone);
  const el=document.getElementById('zone-'+zone);
  if(cn) cn.textContent=files.length?files.length+' file'+(files.length>1?'s':'')+' selected':'No files';
  if(el) el.classList.toggle('has-files',files.length>0);
}
document.querySelectorAll('.dz').forEach(z=>{
  z.addEventListener('dragover',e=>{e.preventDefault();z.classList.add('drag-over');});
  z.addEventListener('dragleave',()=>z.classList.remove('drag-over'));
  z.addEventListener('drop',e=>{
    e.preventDefault();z.classList.remove('drag-over');
    const inp=z.querySelector('input[type=file]');if(!inp) return;
    try{const dt=new DataTransfer();[...e.dataTransfer.files].forEach(f=>dt.items.add(f));inp.files=dt.files;}catch(_){}
    updateZone(inp.dataset.zone,inp);
  });
});

// ── Recon form ────────────────────────────────────────────────────
document.getElementById('recon-form').addEventListener('submit', async e=>{
  e.preventDefault();
  const gstin=document.getElementById('r-gstin').value.trim().toUpperCase();
  const cname=document.getElementById('r-name').value.trim();
  const fy=document.getElementById('r-fy').value.trim()||'2025-26';
  if(!gstin||gstin.length!==15){alert('Enter a valid 15-character GSTIN');return;}
  if(!cname){alert('Enter company name');return;}
  const hasFiles=['r1','r1a','r2b','r2a','r3b','cust','taxlib'].some(z=>(zoneFiles[z]||[]).length>0);
  if(!hasFiles){alert('Upload at least one return file');return;}
  const fd=new FormData();
  fd.append('gstin',gstin);fd.append('client_name',cname);
  fd.append('fy',fy);fd.append('mode','recon');
  ['r1','r1a','r2b','r2a','r3b','cust','taxlib'].forEach(z=>
    (zoneFiles[z]||[]).forEach(f=>fd.append('files_'+z,f)));
  await startJob(fd,'r','Generate Reconciliation + GSTR-1 Detail →');
});

// ── Recon reset ───────────────────────────────────────────────────
function resetRecon(){
  // Clear text inputs
  document.getElementById('r-gstin').value='';
  document.getElementById('r-name').value='';
  document.getElementById('r-fy').value='2025-26';
  document.getElementById('r-state').value='';

  // Clear all file zones for reconciliation tab
  ['r1','r1a','r2b','r2a','r3b','cust','taxlib'].forEach(zone=>{
    zoneFiles[zone]=[];
    const cnt=document.getElementById('cnt-'+zone);
    const el=document.getElementById('zone-'+zone);
    if(cnt) cnt.textContent='No file'+(zone==='r1'||zone==='r1a'||zone==='r2b'||zone==='r2a'||zone==='r3b'?'s':'');
    if(el){
      el.classList.remove('has-files');
      const inp=el.querySelector('input[type=file]');
      if(inp) inp.value='';
    }
  });

  // Hide progress & download panels
  const pw=document.getElementById('r-pw');
  const dw=document.getElementById('r-dw');
  if(pw) pw.style.display='none';
  if(dw) dw.style.display='none';

  // Re-enable submit button
  const btn=document.getElementById('r-submit');
  if(btn){btn.disabled=false;btn.textContent='Generate Reconciliation + GSTR-1 Detail →';}

  // Clear log & progress bar
  const lb=document.getElementById('r-lb');
  const pb=document.getElementById('r-pb');
  if(lb) lb.innerHTML='';
  if(pb) pb.style.width='0%';

  // Reset badge
  setBadge('r','p','Running');
}

// ── GSTR-1 form ───────────────────────────────────────────────────
document.getElementById('g1-form').addEventListener('submit', async e=>{
  e.preventDefault();
  const gstin=document.getElementById('g1-gstin').value.trim().toUpperCase();
  const cname=document.getElementById('g1-name').value.trim();
  const fy=document.getElementById('g1-fy').value.trim()||'2025-26';
  if(!gstin||gstin.length!==15){alert('Enter a valid 15-character GSTIN');return;}
  if(!cname){alert('Enter company name');return;}
  if(!(zoneFiles['g1r1']||[]).length){alert('Upload at least one GSTR-1 ZIP');return;}
  const fd=new FormData();
  fd.append('gstin',gstin);fd.append('client_name',cname);
  fd.append('fy',fy);fd.append('mode','gstr1only');
  (zoneFiles['g1r1']||[]).forEach(f=>fd.append('files_r1',f));
  (zoneFiles['g1r2b']||[]).forEach(f=>fd.append('files_r2b',f));
  (zoneFiles['g1cust']||[]).forEach(f=>fd.append('files_cust',f));
  await startJob(fd,'g1','Generate GSTR-1 Full Detail Excel →');
});

// ── GSTR-2B Full-Year Detail form ─────────────────────────────
document.getElementById('g2b-form').addEventListener('submit', async e=>{
  e.preventDefault();
  const gstin=document.getElementById('g2b-gstin').value.trim().toUpperCase();
  const cname=document.getElementById('g2b-name').value.trim();
  const fy=document.getElementById('g2b-fy').value.trim()||'2025-26';
  if(!gstin||gstin.length!==15){alert('Enter a valid 15-character GSTIN');return;}
  if(!cname){alert('Enter company name');return;}
  if(!(zoneFiles['g2b-files']||[]).length){alert('Upload at least one GSTR-2B Excel file');return;}
  const fd=new FormData();
  fd.append('gstin',gstin);fd.append('client_name',cname);
  fd.append('fy',fy);fd.append('mode','gstr2bonly');
  (zoneFiles['g2b-files']||[]).forEach(f=>fd.append('files_r2b',f));
  await startJob(fd,'g2b','Generate GSTR-2B Full-Year Excel →');
});

// ── Generic job runner ────────────────────────────────────────────
async function startJob(fd, pfx, btnLbl){
  document.getElementById(pfx+'-pw').style.display='block';
  const dw=document.getElementById(pfx+'-dw');if(dw)dw.style.display='none';
  document.getElementById(pfx+'-lb').innerHTML='';
  document.getElementById(pfx+'-pb').style.width='0%';
  const btn=document.getElementById(pfx+'-submit');
  btn.disabled=true;btn.textContent='Uploading...';
  try{
    const res=await fetch('/api/upload',{method:'POST',body:fd});
    let d;try{d=await res.json();}catch(_){throw new Error('Server error (HTTP '+res.status+'). Try again.');}
    if(!d.job_id) throw new Error(d.error||'Upload failed');
    addLog(pfx,'info','Files uploaded. Processing started...');
    btn.textContent='Processing...';
    document.getElementById('ds-jid').value=d.job_id;
    pollJob(d.job_id,pfx,btnLbl);
  }catch(err){
    addLog(pfx,'err','Error: '+err.message);
    setBadge(pfx,'e','Failed');
    btn.disabled=false;btn.textContent=btnLbl;
  }
}

async function pollJob(jid,pfx,btnLbl){
  try{
    const res=await fetch('/api/job/'+jid);
    let d;try{d=await res.json();}catch(_){setTimeout(()=>pollJob(jid,pfx,btnLbl),3000);return;}
    if(d.logs) d.logs.forEach(l=>addLog(pfx,l.type,l.msg));
    if(d.progress!==undefined)
      document.getElementById(pfx+'-pb').style.width=d.progress+'%';
    if(d.dl_status&&Object.keys(d.dl_status).length) renderDlStatus(d.dl_status,jid);
    if(d.status==='done'){
      setBadge(pfx,'d','Complete');
      document.getElementById(pfx+'-pb').style.width='100%';
      const btn=document.getElementById(pfx+'-submit');
      btn.disabled=false;btn.textContent=btnLbl;
      showDownloads(pfx,jid,d.files);
      return;
    }
    if(d.status==='error'){
      addLog(pfx,'err','Error: '+(d.error||'Unknown error'));
      setBadge(pfx,'e','Failed');
      const btn=document.getElementById(pfx+'-submit');
      btn.disabled=false;btn.textContent=btnLbl;
      return;
    }
    setTimeout(()=>pollJob(jid,pfx,btnLbl),1500);
  }catch(err){setTimeout(()=>pollJob(jid,pfx,btnLbl),3000);}
}

function addLog(pfx,type,msg){
  const b=document.getElementById(pfx+'-lb');if(!b) return;
  const l=document.createElement('div');l.className=type;
  l.textContent='['+new Date().toLocaleTimeString()+'] '+msg;
  b.appendChild(l);b.scrollTop=b.scrollHeight;
}
function setBadge(pfx,type,label){
  const b=document.getElementById(pfx+'-badge');if(!b) return;
  b.className='sbg s-'+type;b.textContent=label;
  if(type!=='p') b.classList.remove('pulse');
}
function _autoTriggerDownload(url, filename){
  // Silently trigger browser download without any user click
  const a = document.createElement('a');
  a.href = url; a.download = filename; a.style.display = 'none';
  document.body.appendChild(a); a.click();
  setTimeout(() => document.body.removeChild(a), 1000);
}

function _downloadAllFiles(files, jid, apiBase){
  // Download all files one by one with small delay to avoid browser blocking
  files.forEach((f, i) => {
    setTimeout(() => {
      _autoTriggerDownload(`${apiBase}/${jid}/${encodeURIComponent(f.name)}`, f.name);
    }, i * 600);
  });
}

function showDownloads(pfx,jid,files){
  const sec=document.getElementById(pfx+'-dw');
  const grid=document.getElementById(pfx+'-dlg');
  if(!sec||!grid) return;
  sec.style.display='block';grid.innerHTML='';
  const ICONS={'ANNUAL':'📊','GSTR3BR1':'📋','GSTR3BR2A':'📈','GSTR1_FULL':'📑',
               'RECONCIL':'📊','SUMMARY':'📊','R1_VS':'📋','TAX_LI':'📑'};

  // ── "Download All" button for this tab ──
  if(files && files.length > 1){
    const allBtn = document.createElement('button');
    allBtn.className = 'btn-sec';
    allBtn.style.cssText = 'margin-bottom:.75rem;width:auto;padding:.5rem 1.2rem;';
    allBtn.innerHTML = `⬇ Download All (${files.length} files)`;
    allBtn.onclick = () => _downloadAllFiles(files, jid, '/api/download');
    grid.appendChild(allBtn);
    // Wire the header Download All button for recon tab
    const hdrBtn = document.getElementById('r-dl-all-btn');
    if(hdrBtn && pfx==='r'){
      hdrBtn.style.display='inline-block';
      showDownloads._dlAll = () => _downloadAllFiles(files, jid, '/api/download');
    }
    // Wire g1 Download All button
    const g1HdrBtn = document.getElementById('g1-dl-all-btn');
    if(g1HdrBtn && pfx==='g1'){
      g1HdrBtn.style.display='inline-block';
      window._g1DlAll = () => _downloadAllFiles(files, jid, '/api/download');
    }
    // Wire g2b Download All button
    const g2bHdrBtn = document.getElementById('g2b-dl-all-btn');
    if(g2bHdrBtn && pfx==='g2b'){
      g2bHdrBtn.style.display='inline-block';
      window._g2bDlAll = () => _downloadAllFiles(files, jid, '/api/download');
    }
  }

  files.forEach(f=>{
    const icon=Object.entries(ICONS).find(([k])=>f.name.toUpperCase().includes(k))?.[1]||'📁';
    const c=document.createElement('div');c.className='dlc';
    c.innerHTML=`<div style="font-size:1.4rem">${icon}</div>
      <div class="dl-n">${f.name}</div><div class="dl-s">${f.size}</div>
      <a href="/api/download/${jid}/${encodeURIComponent(f.name)}" class="btn-dl" download>Download ↓</a>`;
    grid.appendChild(c);
  });

  // ── GSTR-2B: auto-trigger download immediately (Level 1 — no Phase 2 needed) ──
  const gstr2bFiles = files.filter(f => f.name.toUpperCase().includes('GSTR2B') || f.name.toUpperCase().includes('GSTR-2B'));
  gstr2bFiles.forEach((f, i) => {
    setTimeout(() => {
      _autoTriggerDownload(`/api/download/${jid}/${encodeURIComponent(f.name)}`, f.name);
    }, i * 700);
  });

  // Register in global download registry
  _registerFilesForGlobalDl(pfx, jid, files, '/api/download');
}

// ── Auto Download ────────────────────────────────────────────────
let _adJobId=null;
function checkBrowserConnection(){}

document.getElementById('ad-form').addEventListener('submit',async e=>{
  e.preventDefault();
  const gstin=document.getElementById('ad-gstin').value.trim().toUpperCase();
  const cname=document.getElementById('ad-name').value.trim();
  const username=document.getElementById('ad-username').value.trim();
  const password=document.getElementById('ad-password').value;
  const token='';
  const fy=document.getElementById('ad-fy').value;
  const returns=document.getElementById('ad-returns').value;
  if(!gstin||gstin.length!==15){alert('Enter valid 15-char GSTIN');return;}
  if(!cname){alert('Enter company name');return;}
  if(!username){alert('Enter username');return;}
  if(!password){alert('Enter your GST portal password');return;}
  document.getElementById('ad-pw').style.display='block';
  document.getElementById('ad-dw').style.display='none';
  document.getElementById('ad-lb').innerHTML='';
  document.getElementById('ad-pb').style.width='0%';
  setBadge('ad','p','Running');
  // Reset failure screenshots panel for new job
  _failShotsRendered = 0;
  document.getElementById('ad-fail-shots').style.display='none';
  document.getElementById('ad-fail-grid').innerHTML='';
  document.getElementById('ad-fail-count').textContent='';
  // Reset live file tracker
  const livePanel = document.getElementById('ad-live-files');
  if(livePanel){ livePanel.style.display='none'; }
  const liveGrid = document.getElementById('ad-live-grid');
  if(liveGrid){ liveGrid.innerHTML=''; }
  const liveCount = document.getElementById('ad-live-count');
  if(liveCount){ liveCount.textContent=''; }
  const btn=document.getElementById('ad-submit');
  btn.disabled=true;btn.textContent='Starting…';
  addLog('ad','info','Starting browser on server — GST portal login in progress...');
  try{
    const res=await fetch('/api/auto-download',{method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({gstin,client_name:cname,username,password,token,fy,returns})});
    let d;try{d=await res.json();}catch(_){
      addLog('ad','err','Server error — try again');setBadge('ad','e','Failed');
      btn.disabled=false;btn.textContent='🚀 Start Auto Download';return;}
    if(d.error){addLog('ad','err',d.error);setBadge('ad','e','Failed');
      btn.disabled=false;btn.textContent='🚀 Start Auto Download';return;}
    _adJobId=d.job_id;
    // Show live screenshot link
    const ssLink=document.getElementById('ad-screenshot-link');
    if(ssLink){ssLink.href='/api/debug-screenshot/'+d.job_id;ssLink.style.display='inline';}
    btn.textContent='Running…';
    _adPoll(_adJobId);
  }catch(err){addLog('ad','err','Network error: '+err.message);setBadge('ad','e','Failed');
    btn.disabled=false;btn.textContent='🚀 Start Auto Download';}
});

async function _adPoll(jid){
  try{
    const r=await fetch('/api/job/'+jid);
    let d;try{d=await r.json();}catch(_){setTimeout(()=>_adPoll(jid),3000);return;}
    if(d.logs)d.logs.forEach(l=>addLog('ad',l.type,l.msg));
    if(d.progress!=null)document.getElementById('ad-pb').style.width=d.progress+'%';

    const captchaCard = document.getElementById('ad-captcha-card');
    const reloginCard = document.getElementById('ad-relogin-card');

    // Show CAPTCHA card when server has a screenshot waiting
    if(captchaCard){
      const wasHidden = captchaCard.style.display === 'none' || captchaCard.style.display === '';
      if(d.captcha_needed && d.captcha_img){
        // Update image
        document.getElementById('ad-captcha-img').src = 'data:image/png;base64,' + d.captcha_img;
        captchaCard.style.display = 'block';
        if(wasHidden){
          captchaCard.scrollIntoView({behavior:'smooth', block:'nearest'});
          document.getElementById('ad-captcha-input').focus();
        }
      } else {
        captchaCard.style.display = 'none';
        document.getElementById('ad-captcha-input').value = '';
      }
    }

    if(reloginCard){
      const wasHidden2 = reloginCard.style.display === 'none' || reloginCard.style.display === '';
      if(d.captcha_needed && !d.captcha_img){
        reloginCard.style.display = 'block';
        if(wasHidden2) reloginCard.scrollIntoView({behavior:'smooth', block:'nearest'});
      } else {
        reloginCard.style.display = 'none';
      }
    }

    // Show files as they arrive during download (live update)
    if(d.files && d.files.length){
      // Always keep _adLastJobId/_adLastFiles up to date so transfer works mid-run
      _adLastJobId = jid;
      _adLastFiles = d.files;

      // ── Live tracker inside the progress card ───────────────────
      const livePanel = document.getElementById('ad-live-files');
      const liveGrid  = document.getElementById('ad-live-grid');
      const liveCount = document.getElementById('ad-live-count');
      if(livePanel && liveGrid){
        livePanel.style.display = 'block';
        // Only render newly added files (append, don't rebuild)
        const rendered = liveGrid.children.length;
        const incoming = d.files.filter(f => !f.name.startsWith('GST_Downloads')); // skip master zip
        if(incoming.length > rendered){
          const newFiles = incoming.slice(rendered);
          newFiles.forEach(f => {
            const icon = f.name.endsWith('.pdf') ? '📄' : f.name.includes('GSTR3B') ? '📄'
                       : f.name.endsWith('.zip') || f.name.endsWith('.json') ? '🗜' : '📊';
            const retType = f.name.split('_')[0] || '';
            const month   = f.name.split('_')[1] || '';
            const label   = retType + (month ? ' · ' + month : '');
            const chip = document.createElement('div');
            chip.style.cssText = `background:rgba(0,230,118,.07);border:1px solid rgba(0,230,118,.25);
              border-radius:8px;padding:.45rem .65rem;display:flex;align-items:center;gap:.45rem;
              animation:fadeIn .4s ease`;
            chip.innerHTML = `<span style="font-size:1.1rem;line-height:1">${icon}</span>
              <div style="min-width:0">
                <div style="font-size:.68rem;font-weight:700;color:var(--grn);font-family:var(--mono);
                            white-space:nowrap;overflow:hidden;text-overflow:ellipsis" title="${f.name}">${label}</div>
                <div style="font-size:.6rem;color:var(--muted);font-family:var(--mono)">${f.size||''}</div>
              </div>
              <a href="/api/dl-file/${jid}/${encodeURIComponent(f.name)}" download
                 style="margin-left:auto;font-size:.65rem;color:var(--accent);text-decoration:none;
                        white-space:nowrap;font-family:var(--mono)" title="Download ${f.name}">⬇</a>`;
            liveGrid.appendChild(chip);
          });
          if(liveCount) liveCount.textContent = `(${incoming.length} file${incoming.length>1?'s':''})`;
        }
      }

      // ── Full downloads card below (shown when done) ─────────────
      const grid = document.getElementById('ad-dlg');
      const sec  = document.getElementById('ad-dw');
      if(sec) sec.style.display = 'block';
      if(grid && grid.children.length !== d.files.length){
        grid.innerHTML = '';
        // Download All button
        if(d.files.length > 1){
          const allBtn = document.createElement('button');
          allBtn.className = 'btn-sec';
          allBtn.style.cssText = 'margin-bottom:.75rem;width:auto;padding:.5rem 1.2rem;';
          allBtn.innerHTML = `⬇ Download All (${d.files.length} files)`;
          allBtn.onclick = () => _downloadAllFiles(d.files, jid, '/api/dl-file');
          grid.appendChild(allBtn);
        }
        d.files.forEach(f => {
          const icon = f.name.endsWith('.pdf') ? '📄' : f.name.endsWith('.zip') ? '🗜' : '📊';
          const c = document.createElement('div'); c.className = 'dlc';
          c.innerHTML = `<div style="font-size:1.4rem">${icon}</div>
            <div class="dl-n">${f.name}</div>
            <div class="dl-s">${f.size||''}</div>
            <a href="/api/dl-file/${jid}/${encodeURIComponent(f.name)}" class="btn-dl" download>⬇ Download</a>`;
          grid.appendChild(c);
        });
        // Show transfer button as soon as first files arrive
        const transferBtn = document.querySelector('[onclick="adTransferToRecon()"]');
        if(transferBtn) transferBtn.style.display='inline-block';
      }
    }

    // Live-update failure screenshots during polling
    if(d.failure_screenshots && d.failure_screenshots.length > _failShotsRendered){
      _renderFailShots(d.failure_screenshots);
    }

    if(d.status==='done'){
      setBadge('ad','d','Complete');
      document.getElementById('ad-pb').style.width='100%';
      document.getElementById('ad-submit').disabled=false;
      document.getElementById('ad-submit').textContent='🚀 Start Auto Download';
      if(captchaCard) captchaCard.style.display='none';
      if(reloginCard) reloginCard.style.display='none';
      // Stop pulsing dot on live tracker and update label
      const livePanel = document.getElementById('ad-live-files');
      if(livePanel){
        const dot = livePanel.querySelector('span[style*="border-radius:50%"]');
        if(dot){ dot.style.animation='none'; dot.style.background='var(--grn)'; }
        const lbl = livePanel.querySelector('div[style*="color:var(--grn)"]');
        if(lbl){ const t=lbl.firstChild; if(t&&t.nodeType===3) t.textContent=''; }
        const hdr = livePanel.querySelector('div[style*="color:var(--grn)"]');
        if(hdr) hdr.innerHTML = hdr.innerHTML.replace('Downloaded So Far','✅ Download Complete');
      }
      // Auto-fill Download Status job ID
      const dsJid = document.getElementById('ds-jid');
      if(dsJid) dsJid.value = jid;
      _adShowFiles(jid, d.files);
      _registerFilesForGlobalDl('autodl', jid, d.files, '/api/dl-file');
      // Final refresh of failure screenshots
      refreshFailShots();
      return;
    }
    if(d.status==='error'){
      addLog('ad','err',d.error||'Unknown error');setBadge('ad','e','Failed');
      document.getElementById('ad-submit').disabled=false;
      document.getElementById('ad-submit').textContent='🚀 Start Auto Download';
      if(captchaCard) captchaCard.style.display='none';
      if(reloginCard) reloginCard.style.display='none';
      return;
    }
    setTimeout(()=>_adPoll(jid),1500);
  }catch(e){setTimeout(()=>_adPoll(jid),3000);}
}

async function submitAdCaptcha(){
  const text=document.getElementById('ad-captcha-input').value.trim();
  if(!text){document.getElementById('ad-captcha-err').textContent='Please type the CAPTCHA first';return;}
  document.getElementById('ad-captcha-err').textContent='';
  try{
    const res=await fetch(`/api/captcha-submit/${_adJobId}`,{
      method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({captcha:text})});
    const d=await res.json();
    if(d.ok){
      document.getElementById('ad-captcha-input').value='';
      addLog('ad','ok','CAPTCHA submitted — logging in...');
    } else {
      document.getElementById('ad-captcha-err').textContent='Error: '+(d.error||'Failed');
    }
  }catch(err){document.getElementById('ad-captcha-err').textContent='Network error: '+err.message;}
}

async function refreshAdCaptcha(){
  if(!_adJobId)return;
  try{
    const res=await fetch(`/api/captcha-refresh/${_adJobId}`,{method:'POST'});
    const d=await res.json();
    if(d.img)document.getElementById('ad-captcha-img').src='data:image/png;base64,'+d.img;
  }catch(e){}
}

async function submitAdRelogin(){
  const token=document.getElementById('ad-relogin-token').value.trim();
  if(!token){alert('Paste the new AuthToken');return;}
  const btn=document.getElementById('ad-relogin-btn');
  btn.disabled=true;btn.textContent='Submitting…';
  try{
    const res=await fetch(`/api/captcha-submit/${_adJobId}`,{
      method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({captcha:token})});
    const d=await res.json();
    if(d.ok){
      document.getElementById('ad-relogin-card').style.display='none';
      document.getElementById('ad-relogin-token').value='';
      addLog('ad','ok','Re-login submitted — resuming download…');
    } else {
      alert('Error: '+(d.error||'Failed'));
    }
  }catch(err){alert('Network error: '+err.message);}
  btn.disabled=false;btn.textContent='Submit New Token →';
}

function showBmInstr(browser){
  document.querySelectorAll('.bm-instr').forEach(el=>el.style.display='none');
  const el=document.getElementById('bm-instr-'+browser);
  if(el)el.style.display='block';
}
function bookmarkletClick(e){e.preventDefault();}


// Store job files globally for transfer
let _adLastJobId = '';
let _adLastFiles = [];

function _adShowFiles(jid, files){
  _adLastJobId = jid;
  _adLastFiles = files || [];
  const sec = document.getElementById('ad-dw'), grid = document.getElementById('ad-dlg');
  sec.style.display = 'block'; grid.innerHTML = '';
  // Update file count badge
  const countEl = document.getElementById('ad-file-count');
  const individual = (files||[]).filter(f=>!f.name.startsWith('GST_Downloads'));
  if(countEl) countEl.textContent = individual.length ? `(${individual.length} return file${individual.length>1?'s':''} + ZIP)` : '';
  if(!files || !files.length){
    grid.innerHTML = '<p style="color:var(--muted);font-size:.8rem">No files downloaded. Check logs above.</p>';
    return;
  }
  // ── "Download All" button ──
  if(files.length > 0){
    const allBtn = document.createElement('button');
    allBtn.className = 'btn-sec';
    allBtn.style.cssText = 'margin-bottom:.75rem;width:auto;padding:.5rem 1.2rem;';
    allBtn.innerHTML = `⬇ Download All (${files.length} file${files.length>1?'s':''})`;
    allBtn.onclick = () => _downloadAllFiles(files, jid, '/api/dl-file');
    grid.appendChild(allBtn);
  }
  files.forEach(f => {
    const icon = f.name.endsWith('.pdf') ? '📄' : f.name.endsWith('.zip') ? '🗜' : '📊';
    const c = document.createElement('div'); c.className = 'dlc';
    c.innerHTML = `<div style="font-size:1.4rem">${icon}</div>
      <div class="dl-n">${f.name}</div>
      <div class="dl-s">${f.size || ''}</div>
      <a href="/api/dl-file/${jid}/${encodeURIComponent(f.name)}" class="btn-dl" download>⬇ Download</a>`;
    grid.appendChild(c);
  });

  // Also update Download Status tab automatically
  const dlStatus = {};
  files.forEach(f => {
    const m = f.name.match(/^(GSTR[^_]+)_([A-Za-z]+)_(\d{4})/);
    if(m){
      const rt = m[1].replace('-',''), mon = m[2];
      dlStatus[`${mon}_${rt}`] = 'OK';
    }
  });
  if(Object.keys(dlStatus).length) renderDlStatus(dlStatus, jid);
}

async function adTransferToRecon(){
  if(!_adLastJobId || !_adLastFiles.length){
    alert('No downloaded files to transfer. Run Auto Download first.'); return;
  }

  // Zone mapping: return type prefix → drop zone id
  const zoneMap = {
    'GSTR1':  'r1',
    'GSTR1A': 'r1a',
    'GSTR2B': 'r2b',
    'GSTR2A': 'r2a',
    'GSTR3B': 'r3b',
  };

  // Filter only individual return files (skip master ZIP)
  const toTransfer = _adLastFiles.filter(f => {
    if(!f.name) return false;
    const n = f.name.toUpperCase();
    // Skip the master ZIP bundle
    if(n.startsWith('GST_DOWNLOADS') && n.endsWith('.ZIP')) return false;
    // Must match a known return type prefix
    return Object.keys(zoneMap).some(rt => n.startsWith(rt + '_'));
  });

  if(!toTransfer.length){
    alert('No individual return files found to transfer. Only the ZIP bundle was downloaded — please download files individually first.');
    return;
  }

  // Switch to reconciliation tab first
  switchTab('recon', null);

  // Show progress in recon log
  document.getElementById('r-pw').style.display='block';
  addLog('r','info',`📥 Transferring ${toTransfer.length} file(s) from Auto Download...`);

  const fetched = {r1:[], r1a:[], r2b:[], r2a:[], r3b:[]};
  let fetchOk = 0, fetchFail = 0;

  for(const f of toTransfer){
    // Determine zone from filename prefix (case-insensitive)
    const upper = f.name.toUpperCase();
    let zone = null;
    for(const [rt, zn] of Object.entries(zoneMap)){
      if(upper.startsWith(rt + '_')){ zone = zn; break; }
    }
    if(!zone){ addLog('r','warn',`⚠ Skipped (unknown type): ${f.name}`); continue; }

    try{
      addLog('r','info',`  Fetching ${f.name}...`);
      const resp = await fetch(`/api/dl-file/${_adLastJobId}/${encodeURIComponent(f.name)}`);
      if(!resp.ok){
        addLog('r','err',`  ✗ HTTP ${resp.status} for ${f.name}`);
        fetchFail++;
        continue;
      }
      const blob = await resp.blob();
      if(blob.size < 100){
        addLog('r','warn',`  ⚠ ${f.name} looks empty (${blob.size} bytes) — skipping`);
        fetchFail++;
        continue;
      }
      const mimeMap = {'.pdf':'application/pdf','.zip':'application/zip','.xlsx':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet','.json':'application/json'};
      const ext = f.name.slice(f.name.lastIndexOf('.')).toLowerCase();
      const mime = mimeMap[ext] || blob.type || 'application/octet-stream';
      fetched[zone].push(new File([blob], f.name, {type: mime}));
      fetchOk++;
    }catch(err){
      addLog('r','err',`  ✗ Fetch error for ${f.name}: ${err.message}`);
      fetchFail++;
    }
  }

  // Load into zone files and update UI
  let total = 0;
  for(const [zone, files] of Object.entries(fetched)){
    if(!files.length) continue;
    zoneFiles[zone] = (zoneFiles[zone] || []).concat(files);
    const cnt = document.getElementById('cnt-'+zone);
    const el  = document.getElementById('zone-'+zone);
    if(cnt){
      const n = zoneFiles[zone].length;
      cnt.textContent = n + ' file' + (n>1?'s':'') + ' selected';
    }
    if(el) el.classList.add('has-files');
    total += files.length;
  }

  document.getElementById('r-pw').style.display='none';

  // Auto-fill company name and GSTIN from auto-download tab
  const adName  = document.getElementById('ad-name');
  const adGstin = document.getElementById('ad-gstin');
  const adFy    = document.getElementById('ad-fy');
  if(adName  && adName.value)  { const el=document.getElementById('r-name');  if(el && !el.value) el.value=adName.value;  }
  if(adGstin && adGstin.value) { const el=document.getElementById('r-gstin'); if(el && !el.value) el.value=adGstin.value.trim().toUpperCase(); }
  if(adFy    && adFy.value)    { const el=document.getElementById('r-fy');    if(el) el.value=adFy.value; }

  if(total > 0){
    addLog('r','ok',`✅ ${total} file(s) loaded! ${fetchFail?fetchFail+' skipped — ':''} GSTIN & Company auto-filled. Click Generate when ready.`);
    // Update transfer button status
    const ts = document.getElementById('ad-transfer-status');
    if(ts){ ts.textContent = `✅ ${total} file(s) loaded into Reconciliation tab`; ts.style.color='var(--grn)'; }
    // Scroll to top of reconciliation tab so user sees the filled fields
    window.scrollTo({top:0, behavior:'smooth'});
  } else {
    addLog('r','err','❌ No files transferred. Server may still be copying files — wait 10 seconds and try again, or download the ZIP and upload manually.');
    const ts = document.getElementById('ad-transfer-status');
    if(ts){ ts.textContent='❌ Transfer failed — try again in 10s'; ts.style.color='var(--red)'; }
  }
}



const MONS=['April','May','June','July','August','September',
            'October','November','December','January','February','March'];
const RETS=['GSTR1','GSTR1A','GSTR2B','GSTR2A','GSTR3B'];

function renderDlStatus(st,jid){
  document.getElementById('ds-result').style.display='block';
  document.getElementById('ds-title').textContent=jid||'—';
  const tb=document.getElementById('ds-tb');tb.innerHTML='';
  let ok=0,fl=0,pd=0;
  MONS.forEach(m=>{
    const tr=document.createElement('tr');
    let rok=0,rfl=0;
    let td=`<td>${m}</td>`;
    RETS.forEach(r=>{
      const v=(st[m+'_'+r]||'SKIP').toUpperCase();
      let cls,txt;
      if(v==='OK'||v==='DONE'){cls='c-ok';txt='✓ OK';rok++;ok++;}
      else if(['TILE_FAIL','NOT_FOUND','TILE_NOT_FOUND','GEN_FAIL','ERR','FAIL'].some(x=>v.includes(x)))
        {cls='c-fl';txt='✗';rfl++;fl++;}
      else if(v==='TRIGGERED'||v==='PENDING')
        {cls='c-pd';txt='⋯';pd++;}
      else{cls='c-sk';txt='—';}
      td+=`<td class="${cls}">${txt}</td>`;
    });
    const rs=rfl>0?`<span style="color:var(--red)">${rfl} failed</span>`:
              rok===5?`<span style="color:var(--grn)">All OK</span>`:
              rok>0?`<span style="color:var(--org)">${rok}/5 OK</span>`:
              `<span style="color:var(--muted)">—</span>`;
    td+=`<td>${rs}</td>`;tr.innerHTML=td;tb.appendChild(tr);
  });
  document.getElementById('ds-sum').innerHTML=
    `<strong style="color:var(--grn)">${ok} ✓ OK</strong> &nbsp; `+
    `<strong style="color:var(--red)">${fl} ✗ Failed</strong> &nbsp; `+
    `<strong style="color:var(--org)">${pd} ⋯ Pending</strong> &nbsp; `+
    `out of ${MONS.length*RETS.length} expected`;
}

async function loadDlStatus(){
  const jid=document.getElementById('ds-jid').value.trim();
  if(jid){
    document.getElementById('ds-pw').style.display='block';
    try{
      const res=await fetch('/api/job/'+jid);
      const d=await res.json();
      if(d.error){alert('Job not found: '+jid);return;}
      if(d.logs) d.logs.forEach(l=>addLog('ds',l.type,l.msg));
      if(d.progress!==undefined) document.getElementById('ds-pb').style.width=d.progress+'%';
      const st=d.dl_status&&Object.keys(d.dl_status).length
        ?d.dl_status:buildStFromFiles(d.files||[]);
      renderDlStatus(st,jid);
      setBadge('ds',d.status==='done'?'d':d.status==='error'?'e':'p',
        d.status==='done'?'Complete':d.status==='error'?'Failed':'Running');
    }catch(e){alert('Error: '+e.message);}
    return;
  }
  const mf=(zoneFiles['master']||[]);
  if(!mf.length){alert('Enter a Job ID or upload a Master Report Excel');return;}
  const fd=new FormData();mf.forEach(f=>fd.append('master_file',f));
  try{
    const res=await fetch('/api/parse_master',{method:'POST',body:fd});
    const d=await res.json();
    if(d.dl_status) renderDlStatus(d.dl_status,'Master Report');
    else alert('Parse error: '+(d.error||'Unknown'));
  }catch(e){alert('Error: '+e.message);}
}

function buildStFromFiles(files){
  const st={};
  MONS.forEach(m=>RETS.forEach(r=>{
    st[m+'_'+r]=files.some(f=>f.name.includes(r)&&f.name.includes(m))?'OK':'SKIP';
  }));
  return st;
}

// ── Feedback / Comments ───────────────────────────────────────────
let currentRating=0;
function setRating(val){
  currentRating=val;
  const LABELS={1:'Poor',2:'Fair',3:'Good',4:'Very Good',5:'Excellent'};
  document.querySelectorAll('.star-btn').forEach(s=>{
    s.style.opacity=parseInt(s.dataset.val)<=val?'1':'.35';
    s.style.color=parseInt(s.dataset.val)<=val?'#ffd700':'var(--muted)';
  });
  document.getElementById('rating-lbl').textContent=LABELS[val]||'';
}

document.getElementById('fb-form').addEventListener('submit', async e=>{
  e.preventDefault();
  const name=document.getElementById('fb-name').value.trim();
  const type=document.getElementById('fb-type').value;
  const msg=document.getElementById('fb-msg').value.trim();
  const stat=document.getElementById('fb-status');
  if(!msg){stat.textContent='Please write a comment.';stat.style.color='var(--red)';return;}
  const btn=document.getElementById('fb-submit');
  btn.disabled=true;btn.textContent='Submitting...';
  stat.textContent='';
  try{
    const res=await fetch('/api/feedback',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({name,type,msg,rating:currentRating})
    });
    const d=await res.json();
    if(d.success){
      stat.textContent='✓ Thank you! Your feedback was recorded.';
      stat.style.color='var(--grn)';
      document.getElementById('fb-msg').value='';
      document.getElementById('fb-name').value='';
      currentRating=0;
      document.querySelectorAll('.star-btn').forEach(s=>{s.style.opacity='.35';s.style.color='var(--muted)';});
      document.getElementById('rating-lbl').textContent='';
      loadFeedback();
    } else {
      stat.textContent='Error: '+(d.error||'Unknown');
      stat.style.color='var(--red)';
    }
  }catch(err){
    stat.textContent='Network error. Please try again.';
    stat.style.color='var(--red)';
  }
  btn.disabled=false;btn.textContent='Submit Feedback →';
});

async function loadFeedback(){
  try{
    const res=await fetch('/api/feedback');
    const d=await res.json();
    const list=document.getElementById('fb-list');
    const count=document.getElementById('fb-count');
    count.textContent=d.count||0;
    if(!d.items||!d.items.length){
      list.innerHTML='<div class="no-fb">No comments yet. Be the first!</div>';
      return;
    }
    list.innerHTML='';
    d.items.forEach(fb=>{
      const TCLS={bug:'fb-bug',suggestion:'fb-sugg',praise:'fb-praise',other:'fb-other'};
      const TLBL={bug:'🐛 Bug',suggestion:'💡 Suggestion',praise:'👍 Praise',other:'💬 Other'};
      const stars=fb.rating?'★'.repeat(fb.rating)+'☆'.repeat(5-fb.rating):'';
      const div=document.createElement('div');div.className='fb-item';
      div.innerHTML=`
        <div class="fb-header">
          <div style="display:flex;align-items:center;gap:.5rem;flex-wrap:wrap">
            <span class="fb-name">${escHtml(fb.name||'Anonymous')}</span>
            <span class="fb-type ${TCLS[fb.type]||'fb-other'}">${TLBL[fb.type]||'Other'}</span>
            ${stars?`<span class="stars" title="${fb.rating}/5 stars">${stars}</span>`:''}
          </div>
          <span class="fb-time">${escHtml(fb.time||'')}</span>
        </div>
        <div class="fb-msg">${escHtml(fb.msg)}</div>`;
      list.appendChild(div);
    });
  }catch(e){console.error('Feedback load error:',e);}
}

function escHtml(s){
  return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;')
    .replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

// Load feedback on page open
loadFeedback();

// ── Bulk Download ─────────────────────────────────────────────────
let _bulkJobId = null;
let _bulkPollTimer = null;

async function startBulk(){
  const files = zoneFiles['bulk'] || [];
  if(!files.length){ alert('Upload a company list Excel first'); return; }
  const fy      = document.getElementById('bulk-fy').value;
  const returns = document.getElementById('bulk-returns').value;
  const fd = new FormData();
  files.forEach(f => fd.append('companies_file', f));
  fd.append('fy', fy);
  fd.append('returns', returns);

  document.getElementById('bulk-pw').style.display = 'block';
  document.getElementById('bulk-dw').style.display = 'none';
  document.getElementById('bulk-captcha-card').style.display = 'none';
  document.getElementById('bulk-lb').innerHTML = '';
  document.getElementById('bulk-pb').style.width = '0%';
  setBadge('bulk','p','Running');
  const btn = document.getElementById('bulk-submit');
  btn.disabled = true; btn.textContent = 'Starting…';
  addLog('bulk','info','Uploading company list…');
  try{
    const res = await fetch('/api/bulk-start', {method:'POST', body:fd});
    const d   = await res.json();
    if(d.error){ addLog('bulk','err',d.error); setBadge('bulk','e','Failed'); btn.disabled=false; btn.textContent='🚀 Start Bulk Download'; return; }
    _bulkJobId = d.job_id;
    addLog('bulk','ok',`Loaded ${d.total} companies. Starting downloads…`);
    _bulkPoll(_bulkJobId);
  }catch(err){
    addLog('bulk','err','Network error: '+err.message);
    setBadge('bulk','e','Failed');
    btn.disabled=false; btn.textContent='🚀 Start Bulk Download';
  }
}

async function _bulkPoll(jid){
  try{
    const r = await fetch('/api/job/'+jid);
    let d; try{ d = await r.json(); }catch(_){ _bulkPollTimer=setTimeout(()=>_bulkPoll(jid),3000); return; }
    if(d.logs) d.logs.forEach(l=>addLog('bulk',l.type,l.msg));
    if(d.progress!=null) document.getElementById('bulk-pb').style.width=d.progress+'%';
    if(d.counter) document.getElementById('bulk-counter').textContent=d.counter;

    // Company needs token
    if(d.captcha_needed && d.captcha_company){
      _showBulkTokenCard(d.captcha_company);
    } else {
      document.getElementById('bulk-captcha-card').style.display='none';
    }

    if(d.status==='done'){
      setBadge('bulk','d','Complete');
      document.getElementById('bulk-pb').style.width='100%';
      document.getElementById('bulk-captcha-card').style.display='none';
      document.getElementById('bulk-submit').disabled=false;
      document.getElementById('bulk-submit').textContent='🚀 Start Bulk Download';
      _bulkShowFiles(jid, d.files);
      _registerFilesForGlobalDl('bulk', jid, d.files, '/api/dl-file');
      return;
    }
    if(d.status==='error'){
      setBadge('bulk','e','Failed');
      document.getElementById('bulk-captcha-card').style.display='none';
      document.getElementById('bulk-submit').disabled=false;
      document.getElementById('bulk-submit').textContent='🚀 Start Bulk Download';
      return;
    }
    _bulkPollTimer = setTimeout(()=>_bulkPoll(jid), 1500);
  }catch(e){ _bulkPollTimer=setTimeout(()=>_bulkPoll(jid),3000); }
}

function _showBulkTokenCard(company){
  document.getElementById('bulk-captcha-card').style.display='block';
  document.getElementById('bulk-co-name').textContent = company.name || company.gstin;
  document.getElementById('bulk-cap-gstin').value = company.gstin || '';
  document.getElementById('bulk-cap-user').value  = company.username || '';
  document.getElementById('bulk-token-input').value = '';
  document.getElementById('bulk-token-err').textContent = '';
  document.getElementById('bulk-token-btn').disabled = false;
  document.getElementById('bulk-token-btn').textContent = 'Submit Token & Download →';
  // Scroll into view
  document.getElementById('bulk-captcha-card').scrollIntoView({behavior:'smooth',block:'center'});
}

async function submitBulkToken(){
  const token = document.getElementById('bulk-token-input').value.trim();
  if(!token){ document.getElementById('bulk-token-err').textContent='Please paste the AuthToken'; return; }
  const btn = document.getElementById('bulk-token-btn');
  btn.disabled=true; btn.textContent='Submitting…';
  document.getElementById('bulk-token-err').textContent='';
  try{
    const res = await fetch(`/api/bulk-token/${_bulkJobId}`, {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({token})
    });
    const d = await res.json();
    if(!d.ok){
      document.getElementById('bulk-token-err').textContent = d.error||'Failed';
      btn.disabled=false; btn.textContent='Submit Token & Download →';
      return;
    }
    document.getElementById('bulk-captcha-card').style.display='none';
    addLog('bulk','ok','Token submitted — downloading…');
  }catch(err){
    document.getElementById('bulk-token-err').textContent='Network error: '+err.message;
    btn.disabled=false; btn.textContent='Submit Token & Download →';
  }
}

function _bulkShowFiles(jid, files){
  const sec=document.getElementById('bulk-dw'), grid=document.getElementById('bulk-dlg');
  sec.style.display='block'; grid.innerHTML='';
  if(!files||!files.length){ grid.innerHTML='<p style="color:var(--muted);font-size:.8rem">No files downloaded.</p>'; return; }
  // ── "Download All" button ──
  if(files.length > 0){
    const allBtn = document.createElement('button');
    allBtn.className = 'btn-sec';
    allBtn.style.cssText = 'margin-bottom:.75rem;width:auto;padding:.5rem 1.2rem;';
    allBtn.innerHTML = `⬇ Download All (${files.length} file${files.length>1?'s':''})`;
    allBtn.onclick = () => _downloadAllFiles(files, jid, '/api/dl-file');
    grid.appendChild(allBtn);
  }
  files.forEach(f=>{
    const c=document.createElement('div'); c.className='dlc';
    c.innerHTML=`<div style="font-size:1.4rem">📥</div>
      <div class="dl-n">${f.name}</div><div class="dl-s">${f.size||''}</div>
      <a href="/api/dl-file/${jid}/${encodeURIComponent(f.name)}" class="btn-dl" download>⬇ Download</a>`;
    grid.appendChild(c);
  });
}

// ── Failure Screenshots ──────────────────────────────────────────
let _failShotsRendered = 0;

async function refreshFailShots(){
  if(!_adJobId) return;
  try{
    const res = await fetch(`/api/failure-screenshots/${_adJobId}`);
    const d   = await res.json();
    if(d.error || !d.screenshots) return;
    _renderFailShots(d.screenshots);
  }catch(e){ console.warn('Failure shots fetch error:', e); }
}

function _renderFailShots(shots){
  const panel = document.getElementById('ad-fail-shots');
  const grid  = document.getElementById('ad-fail-grid');
  const count = document.getElementById('ad-fail-count');
  if(!panel || !grid) return;

  if(!shots || shots.length === 0){
    panel.style.display = 'none';
    _failShotsRendered = 0;
    return;
  }

  panel.style.display = 'block';
  count.textContent = `(${shots.length} screenshot${shots.length>1?'s':''})`;

  // Only add newly arrived screenshots (don't re-render existing ones)
  const newShots = shots.slice(_failShotsRendered);
  newShots.forEach((shot, i) => {
    const idx = _failShotsRendered + i + 1;
    const card = document.createElement('div');
    card.style.cssText = 'background:var(--surf2);border:1px solid rgba(255,23,68,.25);border-radius:10px;padding:.85rem;';

    // Header row
    const hdr = document.createElement('div');
    hdr.style.cssText = 'display:flex;align-items:center;gap:.6rem;margin-bottom:.6rem;flex-wrap:wrap;';
    hdr.innerHTML = `
      <span style="background:rgba(255,23,68,.15);color:var(--red);border:1px solid rgba(255,23,68,.3);
                   border-radius:100px;font-size:.62rem;font-weight:700;padding:.18rem .55rem;
                   font-family:var(--mono)">#${idx}</span>
      <span style="font-size:.78rem;font-weight:700;color:var(--txt);flex:1">${escHtml(shot.label)}</span>
      <span style="font-size:.65rem;color:var(--muted);font-family:var(--mono)">⏰ ${escHtml(shot.ts)}</span>`;
    card.appendChild(hdr);

    // Screenshot image
    const img = document.createElement('img');
    img.src = 'data:image/png;base64,' + shot.img_b64;
    img.alt = shot.label;
    img.style.cssText = 'width:100%;border-radius:7px;border:1px solid var(--bdr);cursor:zoom-in;display:block;';
    img.title = 'Click to open full size';
    img.onclick = () => {
      const win = window.open();
      win.document.write(`<html><head><title>${escHtml(shot.label)}</title>
        <style>body{margin:0;background:#111;display:flex;flex-direction:column;align-items:center;padding:1rem}
        img{max-width:100%;border-radius:8px}
        p{color:#aaa;font-family:monospace;font-size:.8rem;margin:.5rem 0}</style></head>
        <body><p>📸 ${escHtml(shot.label)} — ${escHtml(shot.ts)}</p>
        <img src="data:image/png;base64,${shot.img_b64}"></body></html>`);
    };
    card.appendChild(img);

    // Download link
    const dlRow = document.createElement('div');
    dlRow.style.cssText = 'margin-top:.5rem;display:flex;gap:.5rem;align-items:center;flex-wrap:wrap;';
    const dlBtn = document.createElement('a');
    dlBtn.href = 'data:image/png;base64,' + shot.img_b64;
    dlBtn.download = `failure_${idx}_${shot.label.replace(/[^a-zA-Z0-9]/g,'_').slice(0,40)}.png`;
    dlBtn.className = 'btn-dl';
    dlBtn.style.cssText = 'font-size:.68rem;padding:.3rem .75rem;';
    dlBtn.textContent = '⬇ Download PNG';
    dlRow.appendChild(dlBtn);
    const hint = document.createElement('span');
    hint.style.cssText = 'font-size:.65rem;color:var(--muted);font-family:var(--mono);';
    hint.textContent = 'Click image to open full size • Download to save for reference';
    dlRow.appendChild(hint);
    card.appendChild(dlRow);

    grid.appendChild(card);
  });

  _failShotsRendered = shots.length;
}

// ── Self-Ping to Keep Render App Alive ────────────────────────────
// Pings every 4 minutes while page is open (prevents Render free tier sleep)
(function keepAlive(){
  const PING_INTERVAL = 4 * 60 * 1000; // 4 minutes
  
  async function ping(){
    try{
      const res = await fetch('/health');
      if(res.ok){
        console.log('[KeepAlive] ✓ Ping successful at', new Date().toLocaleTimeString());
      } else {
        console.log('[KeepAlive] ⚠ Ping failed:', res.status);
      }
    }catch(e){
      console.log('[KeepAlive] ⚠ Ping error:', e.message);
    }
  }
  
  // Ping immediately on load
  ping();
  
  // Then ping every 4 minutes
  setInterval(ping, PING_INTERVAL);
  
  console.log('[KeepAlive] Self-ping started - app will stay awake while this tab is open');
})();

// ── Income Tax Reconciliation ─────────────────────────────────────
let _itJobId = null;
let _itPollTimer = null;

document.getElementById('it-form').addEventListener('submit', async function(e){
  e.preventDefault();
  const name      = document.getElementById('it-name').value.trim();
  const pan       = document.getElementById('it-pan').value.trim().toUpperCase();
  const gstin     = document.getElementById('it-gstin').value.trim().toUpperCase();
  const fy        = document.getElementById('it-fy').value;
  const itrForm   = document.getElementById('it-itr-form').value;
  const entityType= document.getElementById('it-entity').value;

  if(!name){ alert('Please enter company name'); return; }
  if(!pan || pan.length !== 10){ alert('PAN must be 10 characters (e.g. ABCDE1234F)'); return; }

  const zones = ['it26as','itais','ittis','itgst'];
  const hasFile = zones.some(z => (zoneFiles[z]||[]).length > 0);
  if(!hasFile){ alert('Please upload at least the 26AS PDF'); return; }

  const btn = document.getElementById('it-submit');
  btn.disabled = true; btn.textContent = 'Uploading…';

  const fd = new FormData();
  fd.append('company_name', name);
  fd.append('pan',  pan);
  fd.append('gstin', gstin);
  fd.append('fy',   fy);
  fd.append('itr_form', itrForm);
  fd.append('entity_type', entityType);

  zones.forEach(z => {
    (zoneFiles[z]||[]).forEach(f => fd.append(`files_${z}`, f));
  });

  try{
    const res = await fetch('/api/it-upload', {method:'POST', body:fd});
    const d   = await res.json();
    if(d.error){ alert('Error: '+d.error); btn.disabled=false; btn.textContent='Generate IT Reconciliation Excel →'; return; }
    _itJobId = d.job_id;
    document.getElementById('it-pw').style.display = 'block';
    document.getElementById('it-dw').style.display = 'none';
    setBadge('it','p','Running');
    btn.textContent = 'Processing…';
    _itPoll(_itJobId);
  }catch(err){
    alert('Network error: '+err.message);
    btn.disabled=false; btn.textContent='Generate IT Reconciliation Excel →';
  }
});

async function _itPoll(jid){
  try{
    const res = await fetch(`/api/it-job/${jid}`);
    const d   = await res.json();
    if(d.error){ return; }

    // Update progress bar
    document.getElementById('it-pb').style.width = (d.progress||0)+'%';

    // Stream logs
    if(d.logs && d.logs.length){
      const lb = document.getElementById('it-lb');
      d.logs.forEach(l => {
        const sp = document.createElement('span');
        sp.className = l.type || 'info';
        sp.textContent = l.msg + '\n';
        lb.appendChild(sp);
      });
      lb.scrollTop = lb.scrollHeight;
    }

    if(d.status === 'done'){
      setBadge('it','d','Complete');
      document.getElementById('it-pb').style.width = '100%';
      document.getElementById('it-submit').disabled = false;
      document.getElementById('it-submit').textContent = 'Generate IT Reconciliation Excel →';
      _itShowFiles(jid, d.files);
      return;
    }
    if(d.status === 'error'){
      setBadge('it','e','Failed');
      document.getElementById('it-submit').disabled = false;
      document.getElementById('it-submit').textContent = 'Generate IT Reconciliation Excel →';
      return;
    }
    _itPollTimer = setTimeout(() => _itPoll(jid), 1500);
  }catch(e){ _itPollTimer = setTimeout(() => _itPoll(jid), 3000); }
}

function _itShowFiles(jid, files){
  const sec  = document.getElementById('it-dw');
  const grid = document.getElementById('it-dlg');
  sec.style.display = 'block';
  grid.innerHTML = '';
  if(!files || !files.length){
    grid.innerHTML = '<p style="color:var(--muted);font-size:.8rem">No files generated.</p>';
    return;
  }
  // ── "Download All" button ──
  if(files.length > 0){
    const allBtn = document.createElement('button');
    allBtn.className = 'btn-sec';
    allBtn.style.cssText = 'margin-bottom:.75rem;width:auto;padding:.5rem 1.2rem;';
    allBtn.innerHTML = `⬇ Download All (${files.length} file${files.length>1?'s':''})`;
    allBtn.onclick = () => _downloadAllFiles(files, jid, '/api/it-dl');
    grid.appendChild(allBtn);
  }
  files.forEach(f => {
    const c = document.createElement('div'); c.className = 'dlc';
    c.innerHTML = `<div style="font-size:1.4rem">🏦</div>
      <div class="dl-n">${f.name}</div>
      <div class="dl-s">${f.size||''}</div>
      <a href="/api/it-dl/${jid}/${encodeURIComponent(f.name)}" class="btn-dl" download>⬇ Download</a>`;
    grid.appendChild(c);
  });
  // Auto-trigger all IT reconciliation downloads immediately
  files.forEach((f, i) => {
    setTimeout(() => {
      _autoTriggerDownload(`/api/it-dl/${jid}/${encodeURIComponent(f.name)}`, f.name);
    }, i * 700);
  });
  // Register in global download registry so global "Download All" picks it up
  _registerFilesForGlobalDl('itrecon', jid, files, '/api/it-dl');
}

function resetIT(){
  if(_itPollTimer) clearTimeout(_itPollTimer);
  _itJobId = null;
  document.getElementById('it-form').reset();
  ['it26as','itais','ittis','itgst'].forEach(z => {
    zoneFiles[z] = [];
    const el = document.getElementById('zone-'+z);
    if(el){ el.classList.remove('has-files'); const inp=el.querySelector('input[type=file]'); if(inp) inp.value=''; }
    const cnt = document.getElementById('cnt-'+z);
    if(cnt) cnt.textContent = z==='it26as'?'No file':'No file (optional)';
  });
  document.getElementById('it-pw').style.display = 'none';
  document.getElementById('it-dw').style.display = 'none';
  document.getElementById('it-lb').innerHTML = '';
  document.getElementById('it-pb').style.width = '0%';
  const btn = document.getElementById('it-submit');
  btn.disabled = false; btn.textContent = 'Generate IT Reconciliation Excel →';
}

// ── Global Download Registry ──────────────────────────────────────
// Each tab registers its ready files here so global "Download All" works
const _globalDlRegistry = {};

function _registerFilesForGlobalDl(tabKey, jid, files, apiBase){
  if(!files || !files.length) return;
  _globalDlRegistry[tabKey] = {jid, files, apiBase};
  // Show global download bar
  const bar = document.getElementById('global-dl-bar');
  if(bar){ bar.style.display='flex'; }
  _updateGlobalDlCount();
}

function _updateGlobalDlCount(){
  let total = 0;
  Object.values(_globalDlRegistry).forEach(r => total += (r.files||[]).length);
  const el = document.getElementById('global-dl-count');
  if(el) el.textContent = `${total} file${total!==1?'s':''} across ${Object.keys(_globalDlRegistry).length} tab${Object.keys(_globalDlRegistry).length!==1?'s':''}`;
}

function globalDownloadAll(){
  let delay = 0;
  Object.values(_globalDlRegistry).forEach(r => {
    (r.files||[]).forEach(f => {
      const url = `${r.apiBase}/${r.jid}/${encodeURIComponent(f.name)}`;
      setTimeout(() => _autoTriggerDownload(url, f.name), delay);
      delay += 650;
    });
  });
}

// Patch showDownloads, _itShowFiles, _adShowFiles, _bulkShowFiles
// to also register in global registry
const _origShowDownloads = showDownloads;
// We'll wrap registration inside each show function call via pollJob / _itPoll / _adPoll


// ── IT Portal Auto Download ──────────────────────────────────────
let _itAdJobId = null;
let _itAdFiles = [];

async function startITAutoDownload(){
  const pan   = (document.getElementById('it-pan').value.trim() || document.getElementById('it-ad-user').value.trim()).toUpperCase();
  const name  = document.getElementById('it-name').value.trim();
  const user  = document.getElementById('it-ad-user').value.trim().toUpperCase() || pan;
  const pass  = document.getElementById('it-ad-pass').value;
  const fy    = document.getElementById('it-fy').value;

  if(!pan || pan.length !== 10){ alert('Enter PAN (10 chars) in Company Details or IT Portal Username field'); return; }
  if(!name){ alert('Enter Company Name above'); return; }
  if(!pass){ alert('Enter IT Portal Password'); return; }

  const btn = document.getElementById('it-ad-btn');
  btn.disabled = true; btn.textContent = 'Starting…';

  document.getElementById('it-ad-pw').style.display = 'block';
  document.getElementById('it-ad-dw').style.display = 'none';
  document.getElementById('it-ad-captcha-card').style.display = 'none';
  document.getElementById('it-ad-lb').innerHTML = '';
  document.getElementById('it-ad-pb').style.width = '0%';

  try{
    const res = await fetch('/api/it-auto-download', {
      method:'POST',headers:{'Content-Type':'application/json'},
      body: JSON.stringify({pan, company_name:name, username:user, password:pass, fy})
    });
    const d = await res.json();
    if(d.error){ alert('Error: '+d.error); btn.disabled=false; btn.textContent='🌐 Start IT Auto Download (26AS + AIS + TIS)'; return; }
    _itAdJobId = d.job_id;
    const ssLink = document.getElementById('it-ad-ss-link');
    if(ssLink){ ssLink.href='/api/debug-screenshot/'+d.job_id; ssLink.style.display='inline'; }
    btn.textContent = 'Running…';
    _itAdPoll(_itAdJobId);
  }catch(err){
    alert('Network error: '+err.message);
    btn.disabled=false; btn.textContent='🌐 Start IT Auto Download (26AS + AIS + TIS)';
  }
}

function _itAdAddLog(type, msg){
  const b = document.getElementById('it-ad-lb'); if(!b) return;
  const l = document.createElement('div'); l.className = type;
  l.textContent = '['+new Date().toLocaleTimeString()+'] '+msg;
  b.appendChild(l); b.scrollTop = b.scrollHeight;
}

async function _itAdPoll(jid){
  try{
    const r = await fetch('/api/job/'+jid);
    const d = await r.json();
    if(d.logs) d.logs.forEach(l => _itAdAddLog(l.type, l.msg));
    if(d.progress != null) document.getElementById('it-ad-pb').style.width = d.progress+'%';

    // Handle CAPTCHA/OTP card
    const cc = document.getElementById('it-ad-captcha-card');
    if(d.captcha_needed && d.captcha_img){
      document.getElementById('it-ad-captcha-img').src = 'data:image/png;base64,'+d.captcha_img;
      if(cc.style.display === 'none') {
        cc.style.display = 'block';
        cc.scrollIntoView({behavior:'smooth', block:'nearest'});
        document.getElementById('it-ad-captcha-input').focus();
      }
    } else {
      if(cc) cc.style.display = 'none';
    }

    if(d.status === 'done'){
      document.getElementById('it-ad-pb').style.width = '100%';
      document.getElementById('it-ad-badge').className = 'sbg s-d';
      document.getElementById('it-ad-badge').textContent = 'Complete';
      document.getElementById('it-ad-btn').disabled = false;
      document.getElementById('it-ad-btn').textContent = '🌐 Start IT Auto Download (26AS + AIS + TIS)';
      if(cc) cc.style.display = 'none';
      _itAdFiles = d.files || [];
      _itAdShowFiles(jid, d.files);
      return;
    }
    if(d.status === 'error'){
      document.getElementById('it-ad-badge').className = 'sbg s-e';
      document.getElementById('it-ad-badge').textContent = 'Failed';
      document.getElementById('it-ad-btn').disabled = false;
      document.getElementById('it-ad-btn').textContent = '🌐 Start IT Auto Download (26AS + AIS + TIS)';
      if(cc) cc.style.display = 'none';
      return;
    }
    // Show live files during progress
    if(d.files && d.files.length){
      _itAdFiles = d.files;
      _itAdShowFiles(jid, d.files);
    }
    setTimeout(() => _itAdPoll(jid), 1500);
  }catch(e){ setTimeout(() => _itAdPoll(jid), 3000); }
}

async function itAdSubmit(){
  const text = document.getElementById('it-ad-captcha-input').value.trim();
  if(!text){ document.getElementById('it-ad-captcha-err').textContent = 'Please type a value first'; return; }
  document.getElementById('it-ad-captcha-err').textContent = '';
  try{
    const res = await fetch(`/api/captcha-submit/${_itAdJobId}`, {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({captcha: text})
    });
    const d = await res.json();
    if(d.ok){
      document.getElementById('it-ad-captcha-input').value = '';
      _itAdAddLog('ok', 'Input submitted — continuing...');
    } else {
      document.getElementById('it-ad-captcha-err').textContent = 'Error: '+(d.error||'Failed');
    }
  }catch(err){ document.getElementById('it-ad-captcha-err').textContent = 'Network error: '+err.message; }
}

async function itAdRefreshShot(){
  if(!_itAdJobId) return;
  try{
    const res = await fetch(`/api/captcha-refresh/${_itAdJobId}`, {method:'POST'});
    const d = await res.json();
    if(d.img) document.getElementById('it-ad-captcha-img').src = 'data:image/png;base64,'+d.img;
  }catch(e){}
}

function _itAdShowFiles(jid, files){
  const sec  = document.getElementById('it-ad-dw');
  const grid = document.getElementById('it-ad-dlg');
  if(!sec || !grid || !files || !files.length) return;
  sec.style.display = 'block';
  grid.innerHTML = '';
  files.forEach(f => {
    const icon = f.name.endsWith('.pdf') ? '📄' : f.name.endsWith('.zip') ? '🗜' : '📊';
    const c = document.createElement('div'); c.className = 'dlc';
    c.innerHTML = `<div style="font-size:1.4rem">${icon}</div>
      <div class="dl-n">${f.name}</div>
      <div class="dl-s">${f.size||''}</div>
      <a href="/api/it-dl/${jid}/${encodeURIComponent(f.name)}" class="btn-dl" download>⬇ Download</a>`;
    grid.appendChild(c);
  });
  _registerFilesForGlobalDl('it-autodl', jid, files, '/api/it-dl');
}

async function itAdTransfer(){
  if(!_itAdJobId || !_itAdFiles.length){ alert('No downloaded files to transfer'); return; }
  const zoneMap = {'26AS': 'it26as', 'AIS': 'itais', 'TIS': 'ittis'};
  const toTransfer = _itAdFiles.filter(f => !f.name.endsWith('.zip'));
  let ok = 0;
  for(const f of toTransfer){
    let zone = null;
    const upper = f.name.toUpperCase();
    for(const [key, zn] of Object.entries(zoneMap)){
      if(upper.includes(key)){ zone = zn; break; }
    }
    if(!zone) continue;
    try{
      const resp = await fetch(`/api/it-dl/${_itAdJobId}/${encodeURIComponent(f.name)}`);
      if(!resp.ok) continue;
      const blob = await resp.blob();
      const file = new File([blob], f.name, {type:'application/pdf'});
      zoneFiles[zone] = [file];
      const cnt = document.getElementById('cnt-'+zone);
      const el  = document.getElementById('zone-'+zone);
      if(cnt) cnt.textContent = '1 file selected';
      if(el)  el.classList.add('has-files');
      ok++;
    }catch(e){ console.warn('Transfer error:', e); }
  }
  const ts = document.getElementById('it-ad-transfer-status');
  if(ok > 0){
    if(ts) { ts.textContent = `✅ ${ok} PDF(s) loaded into zones`; ts.style.color='var(--grn)'; }
  } else {
    if(ts) { ts.textContent = '❌ Transfer failed — try again'; ts.style.color='var(--red)'; }
  }
}

// ── IT Bulk Download ─────────────────────────────────────────────
let _itBulkJobId = null;

async function startITBulk(){
  const files = zoneFiles['itbulk'] || [];
  if(!files.length){ alert('Upload an IT client list Excel first'); return; }
  const fy   = document.getElementById('itbulk-fy').value;
  const mode = document.getElementById('itbulk-mode').value;
  const fd = new FormData();
  files.forEach(f => fd.append('clients_file', f));
  fd.append('fy', fy);
  fd.append('mode', mode);
  document.getElementById('itbulk-pw').style.display = 'block';
  document.getElementById('itbulk-dw').style.display = 'none';
  document.getElementById('itbulk-otp-card').style.display = 'none';
  document.getElementById('itbulk-lb').innerHTML = '';
  document.getElementById('itbulk-pb').style.width = '0%';
  const btn = document.getElementById('itbulk-submit');
  btn.disabled = true; btn.textContent = 'Starting…';
  try{
    const res = await fetch('/api/it-bulk-start', {method:'POST', body:fd});
    const d   = await res.json();
    if(d.error){ alert('Error: '+d.error); btn.disabled=false; btn.textContent='🚀 Start IT Bulk Download'; return; }
    _itBulkJobId = d.job_id;
    _itBulkAddLog('ok', `Loaded ${d.total} clients. Starting…`);
    _itBulkPoll(_itBulkJobId);
  }catch(err){
    alert('Network error: '+err.message);
    btn.disabled=false; btn.textContent='🚀 Start IT Bulk Download';
  }
}
function _itBulkAddLog(type, msg){
  const b = document.getElementById('itbulk-lb'); if(!b) return;
  const l = document.createElement('div'); l.className = type;
  l.textContent = '['+new Date().toLocaleTimeString()+'] '+msg;
  b.appendChild(l); b.scrollTop = b.scrollHeight;
}
async function _itBulkPoll(jid){
  try{
    const r = await fetch('/api/job/'+jid);
    const d = await r.json();
    if(d.logs) d.logs.forEach(l => _itBulkAddLog(l.type, l.msg));
    if(d.progress != null) document.getElementById('itbulk-pb').style.width = d.progress+'%';
    if(d.counter) document.getElementById('itbulk-counter').textContent = d.counter;
    const oc = document.getElementById('itbulk-otp-card');
    if(d.captcha_needed && d.captcha_img){
      document.getElementById('itbulk-otp-img').src = 'data:image/png;base64,'+d.captcha_img;
      if(d.captcha_company){
        document.getElementById('itbulk-client-name').textContent = d.captcha_company.name || d.captcha_company.pan || '';
        document.getElementById('itbulk-cap-pan').value = d.captcha_company.pan || '';
      }
      if(oc.style.display === 'none'){
        oc.style.display='block';
        oc.scrollIntoView({behavior:'smooth',block:'center'});
        document.getElementById('itbulk-otp-input').focus();
      }
    } else { if(oc) oc.style.display='none'; }
    if(d.status==='done'){
      document.getElementById('itbulk-badge').className='sbg s-d';
      document.getElementById('itbulk-badge').textContent='Complete';
      document.getElementById('itbulk-pb').style.width='100%';
      document.getElementById('itbulk-submit').disabled=false;
      document.getElementById('itbulk-submit').textContent='🚀 Start IT Bulk Download';
      if(oc) oc.style.display='none';
      _itBulkShowFiles(jid, d.files);
      _registerFilesForGlobalDl('itbulk', jid, d.files||[], '/api/it-dl');
      return;
    }
    if(d.status==='error'){
      document.getElementById('itbulk-badge').className='sbg s-e';
      document.getElementById('itbulk-badge').textContent='Failed';
      document.getElementById('itbulk-submit').disabled=false;
      document.getElementById('itbulk-submit').textContent='🚀 Start IT Bulk Download';
      if(oc) oc.style.display='none';
      return;
    }
    setTimeout(()=>_itBulkPoll(jid),1500);
  }catch(e){ setTimeout(()=>_itBulkPoll(jid),3000); }
}
async function submitITBulkOTP(){
  const text = document.getElementById('itbulk-otp-input').value.trim();
  if(!text){ document.getElementById('itbulk-otp-err').textContent='Enter a value or type SKIP'; return; }
  document.getElementById('itbulk-otp-err').textContent='';
  try{
    const res = await fetch(`/api/it-bulk-otp/${_itBulkJobId}`,{
      method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({otp:text})
    });
    const d = await res.json();
    if(d.ok){
      document.getElementById('itbulk-otp-input').value='';
      _itBulkAddLog('ok','OTP submitted — continuing…');
      document.getElementById('itbulk-otp-card').style.display='none';
    } else {
      document.getElementById('itbulk-otp-err').textContent=d.error||'Failed';
    }
  }catch(err){ document.getElementById('itbulk-otp-err').textContent='Network error: '+err.message; }
}
async function itBulkRefreshShot(){
  if(!_itBulkJobId) return;
  try{
    const res=await fetch(`/api/captcha-refresh/${_itBulkJobId}`,{method:'POST'});
    const d=await res.json();
    if(d.img) document.getElementById('itbulk-otp-img').src='data:image/png;base64,'+d.img;
  }catch(e){}
}
function _itBulkShowFiles(jid, files){
  const sec=document.getElementById('itbulk-dw'),grid=document.getElementById('itbulk-dlg');
  sec.style.display='block'; grid.innerHTML='';
  if(!files||!files.length){ grid.innerHTML='<p style="color:var(--muted);font-size:.8rem">No files downloaded.</p>'; return; }
  if(files.length>1){
    const ab=document.createElement('button');
    ab.className='btn-sec'; ab.style.cssText='margin-bottom:.75rem;width:auto;padding:.5rem 1.2rem;';
    ab.innerHTML=`⬇ Download All (${files.length} files)`;
    ab.onclick=()=>_downloadAllFiles(files,jid,'/api/it-dl');
    grid.appendChild(ab);
  }
  files.forEach(f=>{
    const icon=f.name.endsWith('.pdf')?'📄':f.name.endsWith('.zip')?'🗜':'📊';
    const c=document.createElement('div'); c.className='dlc';
    c.innerHTML=`<div style="font-size:1.4rem">${icon}</div>
      <div class="dl-n">${f.name}</div><div class="dl-s">${f.size||''}</div>
      <a href="/api/it-dl/${jid}/${encodeURIComponent(f.name)}" class="btn-dl" download>⬇ Download</a>`;
    grid.appendChild(c);
  });
}

</script>
</body>
</html>"""


# ═══════════════════════════════════════════════════════════════════
# RECONCILIATION WORKERS
# ═══════════════════════════════════════════════════════════════════

def run_reconciliation(job_id):
    def log(msg, t="info"):
        with jobs_lock: jobs[job_id]["logs"].append({"type":t,"msg":msg})
    def prog(p):
        with jobs_lock: jobs[job_id]["progress"] = p
    def set_dl(k, v):
        with jobs_lock: jobs[job_id]["dl_status"][k] = v
    try:
        job         = jobs[job_id]
        gstin       = job["gstin"]
        client_name = job["client_name"]
        fy          = job["fy"]
        job_dir     = Path(job["job_dir"])
        out_dir     = Path(job["out_dir"])
        saved       = job["saved"]
        FY_MONTHS   = _fy_months(fy)

        log(f"Starting: {client_name} ({gstin}) FY {fy}")
        prog(5)

        # GSTR-1
        for fpath in saved.get("r1", []):
            mon, yr = _detect_month(fpath, FY_MONTHS)
            if mon:
                dest = job_dir / f"GSTR1_{mon}_{yr}.zip"
                if not dest.exists():
                    try: Path(fpath).rename(dest)
                    except: shutil.copy2(fpath, str(dest))
                log(f"  GSTR-1: {mon} {yr}"); set_dl(f"{mon}_GSTR1", "OK")
            else:
                log(f"  ⚠ Month not detected: {Path(fpath).name}", "warn")

        # GSTR-1A
        for fpath in saved.get("r1a", []):
            mon, yr = _detect_month(fpath, FY_MONTHS)
            if mon:
                dest = job_dir / f"GSTR1A_{mon}_{yr}.zip"
                if not dest.exists():
                    try: Path(fpath).rename(dest)
                    except: shutil.copy2(fpath, str(dest))
                log(f"  GSTR-1A: {mon} {yr}"); set_dl(f"{mon}_GSTR1A", "OK")
            else:
                log(f"  ⚠ GSTR-1A month not detected: {Path(fpath).name}", "warn")

        # GSTR-2B
        for fpath in saved.get("r2b", []):
            mon, yr = _detect_month(fpath, FY_MONTHS)
            if mon:
                dest = job_dir / f"GSTR2B_{mon}_{yr}.xlsx"
                if not dest.exists():
                    try: Path(fpath).rename(dest)
                    except: shutil.copy2(fpath, str(dest))
                log(f"  GSTR-2B: {mon} {yr}"); set_dl(f"{mon}_GSTR2B", "OK")

        # GSTR-2A
        for fpath in saved.get("r2a", []):
            mon, yr = _detect_month(fpath, FY_MONTHS)
            if mon:
                ext = Path(fpath).suffix.lower()
                dest = job_dir / f"GSTR2A_{mon}_{yr}{ext}"
                if not dest.exists():
                    try: Path(fpath).rename(dest)
                    except: shutil.copy2(fpath, str(dest))
                log(f"  GSTR-2A: {mon} {yr}"); set_dl(f"{mon}_GSTR2A", "OK")

        # GSTR-3B
        for fpath in saved.get("r3b", []):
            mon, yr = _detect_month(fpath, FY_MONTHS)
            if mon:
                dest = job_dir / f"GSTR3B_{mon}_{yr}.pdf"
                if not dest.exists():
                    try: Path(fpath).rename(dest)
                    except: shutil.copy2(fpath, str(dest))
                log(f"  GSTR-3B: {mon} {yr}"); set_dl(f"{mon}_GSTR3B", "OK")

        # Customer names
        for fpath in saved.get("cust", []):
            dest = job_dir / "customer_names.xlsx"
            if not dest.exists():
                try: Path(fpath).rename(dest)
                except: shutil.copy2(fpath, str(dest))
            log("  Customer names loaded"); break

        # Tax liability
        for fpath in saved.get("taxlib", []):
            dest = job_dir / f"TAX_LIABILITY_{Path(fpath).name}"
            if not dest.exists():
                try: Path(fpath).rename(dest)
                except: shutil.copy2(fpath, str(dest))
            log(f"  Tax Liability: {Path(dest).name}"); break

        prog(25)

        suite_path = _find_engine("gst_suite_final.py")
        if not suite_path:
            raise FileNotFoundError("gst_suite_final.py not found on server.")

        log("Loading reconciliation engine...")
        import importlib.util as _ilu, logging as _lg
        spec = _ilu.spec_from_file_location("gst_suite", str(suite_path))
        gst  = _ilu.module_from_spec(spec)
        spec.loader.exec_module(gst)

        s = int(fy.split("-")[0]); e = s + 1
        gst.FY_LABEL = fy
        gst.MONTHS = [
            ("April","04",str(s)),("May","05",str(s)),("June","06",str(s)),
            ("July","07",str(s)),("August","08",str(s)),("September","09",str(s)),
            ("October","10",str(s)),("November","11",str(s)),("December","12",str(s)),
            ("January","01",str(e)),("February","02",str(e)),("March","03",str(e)),
        ]

        _log = _lg.getLogger(f"gst_{job_id}")
        _log.setLevel(_lg.DEBUG)
        class WL(_lg.Handler):
            def emit(self, r):
                log(self.format(r), "err" if r.levelno >= _lg.WARNING else "info")
        _log.addHandler(WL())

        prog(30)
        log("Running annual reconciliation (1-2 minutes)...")
        # Try with_formulas=True first; fall back gracefully
        try:
            gst.write_annual_reconciliation(str(job_dir), client_name, gstin, _log, with_formulas=True)
        except TypeError:
            gst.write_annual_reconciliation(str(job_dir), client_name, gstin, _log)
        prog(65)
        log("  ✓ Annual reconciliation complete", "ok")

        extract_path = _find_engine("gstr1_extract.py")
        gstr1_zips   = list(job_dir.glob("GSTR1_*.zip"))

        if extract_path and gstr1_zips:
            log(f"Running GSTR-1 detail extraction ({len(gstr1_zips)} months)...")
            try:
                spec2 = _ilu.spec_from_file_location("gstr1_extract", str(extract_path))
                gstr1 = _ilu.module_from_spec(spec2)
                spec2.loader.exec_module(gstr1)
                out_xl = job_dir / f"GSTR1_FULL_DETAIL_{client_name.replace(' ','_')}.xlsx"
                try:
                    gstr1.extract_gstr1_to_excel(str(job_dir), str(out_xl), with_formulas=True)
                except TypeError:
                    gstr1.extract_gstr1_to_excel(str(job_dir), str(out_xl))
                log(f"  ✓ GSTR-1 detail: {out_xl.name}", "ok")
            except Exception as ex:
                log(f"  ⚠ GSTR-1 extraction error: {ex}", "warn")
        elif not extract_path:
            log("⚠ gstr1_extract.py not on server — GSTR-1 detail skipped", "warn")

        prog(85)
        log("Collecting output files...")
        output_files = []
        for fp in sorted(job_dir.glob("*.xlsx")):
            dest_fp = out_dir / fp.name
            shutil.copy2(str(fp), str(dest_fp))
            sz = dest_fp.stat().st_size // 1024
            output_files.append({"name": fp.name, "size": f"{sz} KB"})
            log(f"  ✓ {fp.name} ({sz} KB)", "ok")

        if not output_files:
            raise RuntimeError("No Excel output generated.")

        prog(100)
        log(f"Done! {len(output_files)} file(s) ready to download.", "ok")
        with jobs_lock:
            jobs[job_id]["status"] = "done"
            jobs[job_id]["files"]  = output_files
        _cleanup_uploads(job_id)

    except Exception as exc:
        import traceback
        log(f"Error: {exc}", "err")
        for line in traceback.format_exc().split("\n"):
            if line.strip(): log(f"  {line}", "err")
        with jobs_lock:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["error"]  = str(exc)
        _cleanup_uploads(job_id)


def run_gstr1_only(job_id):
    def log(msg, t="info"):
        with jobs_lock: jobs[job_id]["logs"].append({"type":t,"msg":msg})
    def prog(p):
        with jobs_lock: jobs[job_id]["progress"] = p
    try:
        job         = jobs[job_id]
        client_name = job["client_name"]
        fy          = job["fy"]
        job_dir     = Path(job["job_dir"])
        out_dir     = Path(job["out_dir"])
        saved       = job["saved"]
        FY_MONTHS   = _fy_months(fy)

        log(f"GSTR-1 Detail: {client_name} FY {fy}")
        prog(5)

        for fpath in saved.get("r1", []):
            mon, yr = _detect_month(fpath, FY_MONTHS)
            if mon:
                dest = job_dir / f"GSTR1_{mon}_{yr}.zip"
                if not dest.exists():
                    try: Path(fpath).rename(dest)
                    except: shutil.copy2(fpath, str(dest))
                log(f"  GSTR-1: {mon} {yr}")
            else:
                log(f"  ⚠ Month not detected: {Path(fpath).name}", "warn")

        for fpath in saved.get("r2b", []) + saved.get("r2a", []):
            dest = job_dir / Path(fpath).name
            if not dest.exists():
                try: Path(fpath).rename(dest)
                except: shutil.copy2(fpath, str(dest))

        for fpath in saved.get("cust", []):
            dest = job_dir / "customer_names.xlsx"
            if not dest.exists():
                try: Path(fpath).rename(dest)
                except: shutil.copy2(fpath, str(dest))
            break

        prog(20)
        gstr1_zips = list(job_dir.glob("GSTR1_*.zip"))
        if not gstr1_zips:
            raise RuntimeError("No GSTR1_*.zip files found.")

        extract_path = _find_engine("gstr1_extract.py")
        if not extract_path:
            raise FileNotFoundError("gstr1_extract.py not found on server.")

        log(f"Extracting {len(gstr1_zips)} month(s)...")
        import importlib.util as _ilu
        spec = _ilu.spec_from_file_location("gstr1_extract", str(extract_path))
        gstr1_mod = _ilu.module_from_spec(spec)
        spec.loader.exec_module(gstr1_mod)

        prog(30)
        out_xl = job_dir / f"GSTR1_FULL_DETAIL_{client_name.replace(' ','_')}.xlsx"
        gstr1_mod.extract_gstr1_to_excel(str(job_dir), str(out_xl))
        prog(90)

        output_files = []
        for fp in sorted(job_dir.glob("*.xlsx")):
            dest_fp = out_dir / fp.name
            shutil.copy2(str(fp), str(dest_fp))
            sz = dest_fp.stat().st_size // 1024
            output_files.append({"name": fp.name, "size": f"{sz} KB"})
            log(f"  ✓ {fp.name} ({sz} KB)", "ok")

        if not output_files:
            raise RuntimeError("No output files generated.")

        prog(100)
        log(f"Done! {len(output_files)} file(s) ready.", "ok")
        with jobs_lock:
            jobs[job_id]["status"] = "done"
            jobs[job_id]["files"]  = output_files
        _cleanup_uploads(job_id)

    except Exception as exc:
        import traceback
        log(f"Error: {exc}", "err")
        for line in traceback.format_exc().split("\n"):
            if line.strip(): log(f"  {line}", "err")
        with jobs_lock:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["error"]  = str(exc)
        _cleanup_uploads(job_id)

def run_gstr2b_only(job_id):
    """
    GSTR-2B Full-Year Detail extraction job.
    Reads all GSTR2B_*.xlsx files uploaded and produces one combined Excel
    (no month column in data rows).
    """
    def log(msg, t="info"):
        with jobs_lock: jobs[job_id]["logs"].append({"type": t, "msg": msg})
    def prog(p):
        with jobs_lock: jobs[job_id]["progress"] = p
    try:
        job         = jobs[job_id]
        client_name = job["client_name"]
        fy          = job["fy"]
        job_dir     = Path(job["job_dir"])
        out_dir     = Path(job["out_dir"])
        saved       = job["saved"]

        log(f"GSTR-2B Full-Year Extraction: {client_name}  FY {fy}")
        prog(5)

        # ── Copy uploaded GSTR-2B files into job_dir with standard names ──
        # Try to detect month from filename and rename; if not detectable keep original name.
        FY_MONTHS = _fy_months(fy)
        for fpath in saved.get("r2b", []):
            fp   = Path(fpath)
            mon, yr = _detect_month(fpath, FY_MONTHS)
            ext  = fp.suffix.lower()
            if mon and yr:
                dest_name = f"GSTR2B_{mon}_{yr}{ext}"
            else:
                dest_name = fp.name   # keep original; gstr2b_extract will still scan it
            dest = job_dir / dest_name
            if not dest.exists():
                try:   fp.rename(dest)
                except: shutil.copy2(str(fp), str(dest))
            if mon:
                log(f"  GSTR-2B: {mon} {yr}  →  {dest_name}")
            else:
                log(f"  File kept as-is (month not detected): {dest_name}", "warn")

        prog(20)

        # ── Load gstr2b_extract engine ────────────────────────────────────
        extract_path = _find_engine("gstr2b_extract.py")
        if not extract_path:
            raise FileNotFoundError(
                "gstr2b_extract.py not found on server. "
                "Place it in the same folder as app.py."
            )

        log("Running GSTR-2B extractor...")
        import importlib.util as _ilu
        spec = _ilu.spec_from_file_location("gstr2b_extract", str(extract_path))
        g2b_mod = _ilu.module_from_spec(spec)
        spec.loader.exec_module(g2b_mod)

        prog(35)
        safe_name = client_name.replace(" ", "_").replace("/", "_")
        out_xl    = job_dir / f"GSTR2B_FULL_YEAR_{safe_name}.xlsx"
        result    = g2b_mod.extract_gstr2b_to_excel(str(job_dir), str(out_xl))
        prog(90)

        if not result or not Path(result).exists():
            raise RuntimeError(
                "Extractor returned no output. "
                "Ensure GSTR-2B files are valid portal Excel downloads."
            )

        # ── Collect all generated xlsx files ──────────────────────────────
        output_files = []
        for fp in sorted(job_dir.glob("*.xlsx")):
            dest_fp = out_dir / fp.name
            shutil.copy2(str(fp), str(dest_fp))
            sz = dest_fp.stat().st_size // 1024
            output_files.append({"name": fp.name, "size": f"{sz} KB"})
            log(f"  ✓ {fp.name}  ({sz} KB)", "ok")

        if not output_files:
            raise RuntimeError("No output Excel generated.")

        prog(100)
        log(f"Done!  {len(output_files)} file(s) ready to download.", "ok")
        with jobs_lock:
            jobs[job_id]["status"] = "done"
            jobs[job_id]["files"]  = output_files
        _cleanup_uploads(job_id)

    except Exception as exc:
        import traceback
        log(f"Error: {exc}", "err")
        for line in traceback.format_exc().split("\n"):
            if line.strip(): log(f"  {line}", "err")
        with jobs_lock:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["error"]  = str(exc)
        _cleanup_uploads(job_id)


# ── Routes ────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/api/upload", methods=["POST"])
@rate_limit(limit=20, window=60)
def api_upload():
    _cleanup_old_jobs()
    gstin       = request.form.get("gstin","").strip().upper()
    client_name = request.form.get("client_name","").strip()
    fy          = request.form.get("fy","2025-26").strip() or "2025-26"
    mode        = request.form.get("mode","recon")

    if not gstin or len(gstin) != 15:
        return jsonify(error="Invalid GSTIN (must be 15 characters)"), 400
    if not client_name:
        return jsonify(error="Company name is required"), 400

    job_id  = str(uuid.uuid4())[:8]
    job_dir = UPLOAD_DIR / job_id
    out_dir = OUTPUT_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    out_dir.mkdir(parents=True, exist_ok=True)

    saved = {k: [] for k in ("r1","r1a","r2b","r2a","r3b","cust","taxlib")}
    for zone in saved:
        for fobj in request.files.getlist(f"files_{zone}"):
            if not fobj.filename: continue
            from werkzeug.utils import secure_filename
            safe = secure_filename(fobj.filename) or f"upload_{zone}_{uuid.uuid4().hex[:6]}"
            if Path(safe).suffix.lower() not in ALLOWED_EXT: continue
            dest = job_dir / safe
            fobj.save(str(dest))
            saved[zone].append(str(dest))

    with jobs_lock:
        jobs[job_id] = {
            "status":"queued","progress":0,"logs":[],"files":[],
            "error":None,"gstin":gstin,"client_name":client_name,
            "fy":fy,"job_dir":str(job_dir),"out_dir":str(out_dir),
            "saved":saved,"mode":mode,"dl_status":{},
        }

    if mode == "gstr1only":
        target = run_gstr1_only
    elif mode == "gstr2bonly":
        target = run_gstr2b_only
    else:
        target = run_reconciliation
    threading.Thread(target=target, args=(job_id,), daemon=True).start()
    return jsonify(job_id=job_id)

@app.route("/api/job/<job_id>")
@rate_limit(limit=120, window=60)
def api_job(job_id):
    with jobs_lock:
        job = jobs.get(job_id)
        if not job:
            return jsonify(error="Job not found"), 404
        new_logs = job["logs"][:]
        job["logs"] = []
        # Return failure screenshots (without img data in poll — just label+ts for count)
        # Full images are fetched via /api/failure-screenshots/<job_id>
        fail_shots = job.get("failure_screenshots", [])
        fail_shots_light = [{"label": s["label"], "ts": s["ts"], "img_b64": s["img_b64"]} for s in fail_shots]
        return jsonify(
            status=job["status"], progress=job["progress"],
            logs=new_logs, files=job["files"], error=job["error"],
            dl_status=job.get("dl_status",{}),
            captcha_needed=job.get("captcha_needed", False),
            captcha_img=job.get("captcha_img", None),
            captcha_company=job.get("captcha_company", None),
            counter=job.get("counter",""),
            failure_screenshots=fail_shots_light,
        )

@app.route("/api/download/<job_id>/<filename>")
@rate_limit(limit=30, window=60)
def api_download(job_id, filename):
    if not re.match(r'^[\w\-. ()]+\.(xlsx|pdf|zip)$', filename):
        abort(400)
    fpath = OUTPUT_DIR / job_id / filename
    if not fpath.exists() or not fpath.is_file():
        abort(404)
    return send_file(str(fpath), as_attachment=True, download_name=filename)

@app.route("/api/parse_master", methods=["POST"])
@rate_limit(limit=10, window=60)
def api_parse_master():
    fobj = request.files.get("master_file")
    if not fobj:
        return jsonify(error="No file"), 400
    tmp = Path(tempfile.gettempdir()) / f"master_{uuid.uuid4().hex[:8]}.xlsx"
    try:
        fobj.save(str(tmp))
        import openpyxl
        wb = openpyxl.load_workbook(str(tmp), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify(error="Empty file"), 400
        headers = [str(c or "").strip().upper() for c in rows[0]]
        col = {h: i for i, h in enumerate(headers)}
        def _c(*names):
            for n in names:
                if n in col: return col[n]
            return -1
        MC  = _c("MONTH","PERIOD","MONTH YEAR")
        R1  = _c("GSTR-1","GSTR1","R1")
        R1A = _c("GSTR-1A","GSTR1A","R1A")
        R2B = _c("GSTR-2B","GSTR2B","R2B")
        R2A = _c("GSTR-2A","GSTR2A","R2A")
        R3B = _c("GSTR-3B","GSTR3B","R3B")
        dl_status = {}
        for row in rows[1:]:
            try:
                if MC < 0: continue
                raw = str(row[MC] or "").strip()
                if not raw or raw.lower() in ("none","nan",""): continue
                mon = raw.split()[0]
                def _st(i):
                    if i < 0 or i >= len(row): return "SKIP"
                    return str(row[i] or "SKIP").strip().upper()
                dl_status[f"{mon}_GSTR1"]  = _st(R1)
                dl_status[f"{mon}_GSTR1A"] = _st(R1A)
                dl_status[f"{mon}_GSTR2B"] = _st(R2B)
                dl_status[f"{mon}_GSTR2A"] = _st(R2A)
                dl_status[f"{mon}_GSTR3B"] = _st(R3B)
            except: continue
        wb.close()
        return jsonify(dl_status=dl_status)
    except Exception as e:
        return jsonify(error=str(e)), 500
    finally:
        try: tmp.unlink(missing_ok=True)
        except: pass

# ── Feedback API ──────────────────────────────────────────────────
@app.route("/api/feedback", methods=["GET","POST"])
@rate_limit(limit=20, window=60)
def api_feedback():
    if request.method == "GET":
        fb = _load_feedback()
        return jsonify(count=len(fb), items=list(reversed(fb[-50:])))

    data = request.get_json(silent=True) or {}
    msg  = str(data.get("msg","")).strip()[:1000]
    if not msg:
        return jsonify(error="Comment cannot be empty"), 400

    name   = str(data.get("name","Anonymous")).strip()[:60] or "Anonymous"
    ftype  = data.get("type","other")
    if ftype not in ("bug","suggestion","praise","other"): ftype = "other"
    rating = int(data.get("rating",0))
    if rating not in range(0,6): rating = 0

    entry = {
        "name":   name,
        "type":   ftype,
        "msg":    msg,
        "rating": rating,
        "time":   datetime.now().strftime("%d-%b-%Y %H:%M"),
    }
    fb = _load_feedback()
    fb.append(entry)
    if len(fb) > 500: fb = fb[-500:]
    _save_feedback(fb)
    return jsonify(success=True)

# ═══════════════════════════════════════════════════════════════════
# SERVER-SIDE AUTO DOWNLOAD  — No bridge, no OTP, no paid service
# User types CAPTCHA once in the web UI. Server does everything else.
# ═══════════════════════════════════════════════════════════════════
import queue as _queue, base64 as _b64

# Per-job state: captcha_q gets the text user types; screenshot holds b64 PNG
_sessions: dict = {}
_sess_lock = threading.Lock()

# ── Health ────────────────────────────────────────────────────────
@app.route("/health")
def health():
    return jsonify(status="ok")

# ── Browser status — always ready (server runs browser) ───────────
@app.route("/api/browser-status")
def browser_status():
    return jsonify(connected=True, mode="server")

@app.route("/api/debug-screenshot/<job_id>")
def debug_screenshot(job_id):
    """Returns the latest screenshot as an HTML page for easy viewing."""
    with jobs_lock:
        job = jobs.get(job_id)
    img = (job or {}).get("captcha_img") or ""
    with _sess_lock:
        s = _sessions.get(job_id, {})
    img = img or s.get("screenshot") or ""
    if not img:
        return "<h2>No screenshot available yet</h2><p>Run the download first, then refresh.</p>"
    return f"""<!DOCTYPE html><html><head>
    <title>Debug Screenshot - Job {job_id}</title>
    <meta http-equiv="refresh" content="3">
    <style>body{{background:#111;color:#eee;font-family:monospace;padding:20px}}
    img{{max-width:100%;border:2px solid #0f0;border-radius:8px}}</style>
    </head><body>
    <h3>🖥 Server Browser Screenshot — Job {job_id}</h3>
    <p style="color:#0f0">Auto-refreshes every 3 seconds</p>
    <img src="data:image/png;base64,{img}">
    </body></html>"""



# ── Failure screenshots endpoint ─────────────────────────────────
@app.route("/api/failure-screenshots/<job_id>")
def failure_screenshots(job_id):
    """Returns list of failure screenshots captured during auto download."""
    with jobs_lock:
        job = jobs.get(job_id)
    if not job:
        return jsonify(error="job not found"), 404
    shots = job.get("failure_screenshots", [])
    # Return label, ts, and img_b64 for each
    return jsonify(count=len(shots), screenshots=[
        {"label": s["label"], "ts": s["ts"], "img_b64": s["img_b64"]}
        for s in shots
    ])

# ── CAPTCHA image endpoint ────────────────────────────────────────
@app.route("/api/captcha-img/<job_id>")
def captcha_img(job_id):
    with _sess_lock:
        s = _sessions.get(job_id, {})
    img = s.get("screenshot")
    if not img:
        return jsonify(error="not ready"), 404
    return jsonify(img=img)

# ── CAPTCHA refresh ───────────────────────────────────────────────
@app.route("/api/captcha-refresh/<job_id>", methods=["POST"])
def captcha_refresh(job_id):
    with _sess_lock:
        s = _sessions.get(job_id, {})
    if not s:
        return jsonify(error="no session"), 404
    s.get("refresh_event", threading.Event()).set()
    # wait up to 6s for new screenshot
    for _ in range(60):
        time.sleep(0.1)
        img = s.get("screenshot")
        if img:
            return jsonify(img=img)
    return jsonify(error="timeout"), 504

# ── CAPTCHA submit ────────────────────────────────────────────────
@app.route("/api/captcha-submit/<job_id>", methods=["POST"])
def captcha_submit(job_id):
    text = (request.get_json(silent=True) or {}).get("captcha","").strip()
    if not text:
        return jsonify(ok=False, error="empty")
    with _sess_lock:
        s = _sessions.get(job_id)
    if not s:
        return jsonify(ok=False, error="no session")
    s["captcha_q"].put(text)
    return jsonify(ok=True)

# ── Receive token from bookmarklet ───────────────────────────────
@app.route("/api/receive-token/<job_id>", methods=["POST","OPTIONS"])
def receive_token(job_id):
    # CORS — bookmarklet runs on gst.gov.in, needs cross-origin POST
    if request.method == "OPTIONS":
        from flask import Response
        r = Response("", 204)
        r.headers["Access-Control-Allow-Origin"]  = "*"
        r.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
        r.headers["Access-Control-Allow-Headers"] = "Content-Type"
        return r

    data  = request.get_json(silent=True) or {}
    token = str(data.get("token","")).strip()
    if not token:
        from flask import Response
        r = Response('{"ok":false,"error":"empty token"}', 400, mimetype="application/json")
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r

    # Put token into the job's captcha_q so the background thread picks it up
    with _sess_lock:
        s = _sessions.get(job_id)
    if s:
        # Clear stale tokens first, then add new one
        while not s["captcha_q"].empty():
            try: s["captcha_q"].get_nowait()
            except: pass
        s["captcha_q"].put(token)
        # Update job state so UI knows token arrived
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["captcha_needed"] = False
                jobs[job_id]["captcha_img"]    = None
        from flask import Response
        r = Response('{"ok":true,"msg":"Token received — download resuming!"}',
                     200, mimetype="application/json")
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r

    # Job not found — but still return ok (bookmarklet shouldn't see errors)
    from flask import Response
    r = Response('{"ok":false,"error":"job not found"}', 404, mimetype="application/json")
    r.headers["Access-Control-Allow-Origin"] = "*"
    return r


@app.route("/api/dl-file/<job_id>/<filename>")
def dl_file(job_id, filename):
    filename = Path(filename).name
    fp = OUTPUT_DIR / job_id / filename
    if not fp.exists():
        abort(404)
    return send_file(str(fp), as_attachment=True, download_name=filename)

# ── Start Auto Download ───────────────────────────────────────────
@app.route("/api/auto-download", methods=["POST"])
@rate_limit(limit=10, window=60)
def api_auto_download():
    d = request.get_json(silent=True) or {}
    gstin       = d.get("gstin","").strip().upper()
    client_name = d.get("client_name","").strip()
    username    = d.get("username","").strip()
    password    = d.get("password","")
    token       = d.get("token","").strip()
    fy          = d.get("fy","2025-26")
    returns     = d.get("returns","all")

    if not gstin or len(gstin) != 15:
        return jsonify(error="Invalid GSTIN"), 400
    if not client_name:
        return jsonify(error="Company name required"), 400
    if not username:
        return jsonify(error="Username required"), 400
    # Accept either token (direct paste) or no token (bookmarklet will send later)
    # password field kept for backwards compatibility

    job_id = str(uuid.uuid4())[:8]
    out_dir = OUTPUT_DIR / job_id
    out_dir.mkdir(parents=True, exist_ok=True)

    with jobs_lock:
        jobs[job_id] = {
            "status": "running", "progress": 0,
            "logs": [{"type":"info","msg":"Starting..."}],
            "files": [], "error": None,
            "captcha_needed": False, "captcha_img": None,
            "out_dir": str(out_dir),
            "failure_screenshots": [],   # list of {label, img_b64, ts}
        }

    sess = {
        "captcha_q":     _queue.Queue(),
        "refresh_event": threading.Event(),
        "screenshot":    None,
    }
    with _sess_lock:
        _sessions[job_id] = sess

    def run_bg():
        try:
            _auto_download(job_id, gstin, client_name,
                           username, password, fy, returns, sess, token=token)
        except Exception as _bg_exc:
            import traceback as _tb
            _msg = str(_bg_exc)
            with jobs_lock:
                if job_id in jobs:
                    jobs[job_id]["status"] = "error"
                    jobs[job_id]["error"]  = _msg
                    jobs[job_id]["logs"].append({"type":"err","msg":f"Fatal: {_msg}"})
                    for _l in _tb.format_exc().split("\n"):
                        if _l.strip():
                            jobs[job_id]["logs"].append({"type":"err","msg":f"  {_l}"})
        finally:
            with _sess_lock:
                _sessions.pop(job_id, None)

    threading.Thread(target=run_bg, daemon=True).start()
    return jsonify(job_id=job_id)


def _auto_download(job_id, gstin, client_name,
                    username, password, fy, returns, sess, token=""):
    """
    Selenium-based download — exactly follows gst_suite_final.py flow:
      1. Open www.gst.gov.in → LOGIN button → fill username + password
      2. Show CAPTCHA screenshot → user types CAPTCHA in web UI → click LOGIN
      3. Services → Returns → Returns Dashboard
      4. For each month: select FY/Quarter/Period → SEARCH → click tile DOWNLOAD
         GSTR-3B  : PDF downloads directly
         GSTR-1/1A: GENERATE JSON → wait → download link
         GSTR-2B/2A: GENERATE EXCEL → wait → download link
      5. ZIP all files → done
    """
    import base64, tempfile, shutil as _shutil

    def log(msg, t="info"):
        print(f"[{job_id}] {msg}")
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["logs"].append({"type": t, "msg": msg})

    def prog(p):
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["progress"] = p

    def show_captcha(img_b64):
        sess["screenshot"] = img_b64
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["captcha_needed"] = True
                jobs[job_id]["captcha_img"]    = img_b64

    def clear_captcha():
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["captcha_needed"] = False
                jobs[job_id]["captcha_img"]    = None

    def save_failure_screenshot(label):
        """Capture current browser state and store as a failure screenshot for reference."""
        try:
            img_b64 = _screenshot_b64()
            if not img_b64:
                return
            entry = {
                "label": label,
                "img_b64": img_b64,
                "ts": datetime.now().strftime("%H:%M:%S"),
            }
            with jobs_lock:
                if job_id in jobs:
                    jobs[job_id].setdefault("failure_screenshots", []).append(entry)
            log(f"  📸 Failure screenshot saved: {label}", "warn")
        except Exception as _se:
            log(f"  ⚠ Could not save failure screenshot: {_se}", "warn")

    def wait_for_captcha_input():
        """Block until user submits CAPTCHA text from web UI. Returns the text."""
        while not sess["captcha_q"].empty():
            try: sess["captcha_q"].get_nowait()
            except: pass
        log("⏳ CAPTCHA screenshot shown — please type the CAPTCHA in the box above and click Submit")
        try:
            return sess["captcha_q"].get(timeout=600)
        except _queue.Empty:
            raise RuntimeError("CAPTCHA timeout — no input received in 10 minutes")

    # ── FY and months setup ──────────────────────────────────────────
    fy_start = int(fy.split("-")[0])
    MONTHS_LIST = [
        ("April","04",str(fy_start)),    ("May","05",str(fy_start)),
        ("June","06",str(fy_start)),     ("July","07",str(fy_start)),
        ("August","08",str(fy_start)),   ("September","09",str(fy_start)),
        ("October","10",str(fy_start)),  ("November","11",str(fy_start)),
        ("December","12",str(fy_start)), ("January","01",str(fy_start+1)),
        ("February","02",str(fy_start+1)),("March","03",str(fy_start+1)),
    ]
    QUARTER_MAP_LOCAL = {
        "April":"Quarter 1 (Apr - Jun)","May":"Quarter 1 (Apr - Jun)","June":"Quarter 1 (Apr - Jun)",
        "July":"Quarter 2 (Jul - Sep)","August":"Quarter 2 (Jul - Sep)","September":"Quarter 2 (Jul - Sep)",
        "October":"Quarter 3 (Oct - Dec)","November":"Quarter 3 (Oct - Dec)","December":"Quarter 3 (Oct - Dec)",
        "January":"Quarter 4 (Jan - Mar)","February":"Quarter 4 (Jan - Mar)","March":"Quarter 4 (Jan - Mar)",
    }

    out_dir = Path(jobs[job_id]["out_dir"])
    dl_dir  = out_dir / "browser_downloads"
    dl_dir.mkdir(parents=True, exist_ok=True)
    downloaded = []
    driver = None

    # ── Selenium imports ────────────────────────────────────────────
    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait, Select
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.chrome.options import Options as ChromeOptions
        from selenium.webdriver.chrome.service import Service as ChromeService
        import selenium.common.exceptions as SeEx
    except ImportError:
        raise RuntimeError("Selenium not installed on server. Run: pip install selenium")

    # ── Helper functions (mirrors gst_suite_final.py) ───────────────
    def _try_click(xpaths, timeout=8):
        for xp in xpaths:
            try:
                el = WebDriverWait(driver, timeout).until(
                    EC.element_to_be_clickable((By.XPATH, xp)))
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                time.sleep(0.3)
                try: el.click()
                except: driver.execute_script("arguments[0].click();", el)
                return True
            except: continue
        return False

    def _human_type(by, val, text):
        try:
            el = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((by, val)))
            driver.execute_script("arguments[0].scrollIntoView(true);", el)
            time.sleep(0.3); el.click(); time.sleep(0.2)
            el.clear(); time.sleep(0.2)
            for ch in str(text): el.send_keys(ch); time.sleep(0.03)
            time.sleep(0.3)
            return True
        except Exception as e:
            log(f"  Type failed {val}: {e}", "warn")
            return False

    def _is_session_lost():
        try:
            url = driver.current_url.lower()
            if "accessdenied" in url: return True
            if "login" in url and "fowelcome" not in url and "gst.gov.in" in url: return True
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            for phrase in ["session expired","you are not logged in","please login again","access denied"]:
                if phrase in body: return True
        except: pass
        return False

    def _screenshot_b64():
        try: return base64.b64encode(driver.get_screenshot_as_png()).decode()
        except: return None

    def _do_login():
        """Navigate to GST portal, fill creds, show CAPTCHA screenshot, wait for user input."""
        log("🌐 Opening www.gst.gov.in ...")
        driver.get("https://www.gst.gov.in")
        time.sleep(2)

        log("  Clicking LOGIN button...")
        _try_click([
            "//a[normalize-space()='LOGIN']",
            "//a[normalize-space()='Login']",
            "//button[normalize-space()='LOGIN']",
            "//a[contains(@href,'login')]",
        ])
        # Smart wait for login page to load
        try:
            WebDriverWait(driver, 6).until(
                lambda d: len(d.find_elements(By.CSS_SELECTOR, "input[type=text],input[type=password]")) > 0
            )
        except: time.sleep(2)
        log(f"  Login page: {driver.current_url}")

        log(f"  Filling username: {username}")
        filled = False
        from selenium.webdriver.common.by import By as _By
        for by, val in [
            (_By.ID,"username"),(_By.NAME,"username"),
            (_By.ID,"user_name"),(_By.NAME,"user_name"),
            (_By.CSS_SELECTOR,"input[placeholder*='sername']"),
            (_By.CSS_SELECTOR,"input[type='text']:not([readonly])"),
        ]:
            if _human_type(by, val, username):
                filled = True; break
        if not filled:
            raise RuntimeError("Cannot find username field on GST portal login page")

        time.sleep(2)
        log("  Filling password...")
        filled = False
        for by, val in [
            (_By.ID,"user_pass"),(_By.NAME,"user_pass"),
            (_By.ID,"password"),(_By.NAME,"password"),
            (_By.CSS_SELECTOR,"input[type='password']"),
        ]:
            if _human_type(by, val, password):
                filled = True; break
        if not filled:
            raise RuntimeError("Cannot find password field on GST portal login page")

        time.sleep(2)

        # Show CAPTCHA screenshot to user
        log("📸 Taking CAPTCHA screenshot — please type the CAPTCHA in the box below...")
        img = _screenshot_b64()
        show_captcha(img)

        # Wait for user to type CAPTCHA in web UI
        captcha_text = wait_for_captcha_input()
        clear_captcha()
        log(f"  CAPTCHA received: {'*' * len(captcha_text)}")

        # Fill CAPTCHA in browser
        log("  Filling CAPTCHA in browser...")
        captcha_filled = False
        for by, val in [
            (_By.ID,"captcha"),(_By.NAME,"captcha"),
            (_By.ID,"imgCaptcha"),(_By.NAME,"imgCaptcha"),
            (_By.CSS_SELECTOR,"input[placeholder*='aptcha']"),
            (_By.CSS_SELECTOR,"input[placeholder*='APTCHA']"),
            (_By.XPATH,"//input[@id='captcha' or @name='captcha' or contains(@placeholder,'aptcha')]"),
        ]:
            try:
                if _human_type(by, val, captcha_text):
                    captcha_filled = True; break
            except: continue

        if not captcha_filled:
            log("  ⚠ Could not auto-fill CAPTCHA field — please type it manually in the browser", "warn")
            # Give user 30 seconds to fill it manually
            time.sleep(30)

        time.sleep(2)

        # Click LOGIN button
        log("  Clicking LOGIN button...")
        _try_click([
            "//button[@id='btnlogin']",
            "//button[normalize-space()='LOGIN']",
            "//button[normalize-space()='Login']",
            "//button[@type='submit']",
            "//input[@type='submit']",
        ])
        # Smart wait: wait for URL change after login (up to 8s)
        try:
            WebDriverWait(driver, 8).until(
                lambda d: "login" not in d.current_url.lower() or "fowelcome" in d.current_url.lower()
            )
        except: pass
        time.sleep(2)

        # OTP check
        try:
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            if "otp" in body and ("enter" in body or "verify" in body):
                log("📱 OTP required — please enter OTP in the browser...")
                img = _screenshot_b64()
                show_captcha(img)
                otp = wait_for_captcha_input()
                clear_captcha()
                for by, val in [
                    (By.ID,"otp"),(By.NAME,"otp"),
                    (By.CSS_SELECTOR,"input[placeholder*='OTP']"),
                    (By.CSS_SELECTOR,"input[placeholder*='otp']"),
                ]:
                    try:
                        if _human_type(by, val, otp): break
                    except: continue
                _try_click(["//button[contains(text(),'VERIFY')]","//button[contains(text(),'Submit')]","//button[@type='submit']"])
                time.sleep(4)
        except: pass

        try:
            cur = driver.current_url.lower()
        except Exception as _sess_err:
            raise RuntimeError(
                "Chrome session crashed while waiting for CAPTCHA/OTP. "
                "This was caused by --single-process on Windows (now fixed). "
                "Please click 'Start Download' again to retry."
            ) from _sess_err
        log(f"  Post-login URL: {driver.current_url}")
        if "accessdenied" in cur or ("login" in cur and "fowelcome" not in cur):
            raise RuntimeError("Login failed — wrong username/password/CAPTCHA. Please try again.")
        log("  ✅ Login successful!", "ok")
        return True

    def _go_to_dashboard():
        """Navigate to Returns Dashboard by clicking Services→Returns→Returns Dashboard.
        Never uses direct URLs — always follows menu clicks to avoid Access Denied."""
        cur = driver.current_url
        if "return.gst.gov.in" in cur and "dashboard" in cur:
            return True

        if _is_session_lost():
            log("  ⚠ Session lost — re-logging in...", "warn")
            _do_login()

        log("  Navigating: Services → Returns → Returns Dashboard")

        for attempt in range(3):
            log(f"  Nav attempt {attempt+1} from: {driver.current_url}")

            # Step 1: Click Services (also hover to open dropdown)
            try:
                from selenium.webdriver.common.action_chains import ActionChains
                svc_el = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[normalize-space(text())='Services']")))
                ActionChains(driver).move_to_element(svc_el).click(svc_el).perform()
                log("  Services clicked ✓")
            except:
                _try_click([
                    "//a[normalize-space(text())='Services']",
                    "//nav//a[normalize-space()='Services']",
                ])
            time.sleep(1)

            # Step 2: Click Returns in dropdown (hover first to keep dropdown open)
            try:
                ret_el = WebDriverWait(driver, 8).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[normalize-space(text())='Returns']")))
                ActionChains(driver).move_to_element(ret_el).click(ret_el).perform()
                log("  Returns clicked ✓")
            except:
                _try_click([
                    "//a[normalize-space(text())='Returns']",
                    "//*[contains(@class,'dropdown-menu')]//a[normalize-space()='Returns']",
                    "//*[contains(@class,'open')]//a[normalize-space()='Returns']",
                ])
            time.sleep(1)

            # Step 3: Click Returns Dashboard — XPath first, then full page scan
            clicked = _try_click([
                "//a[contains(normalize-space(text()),'Returns Dashboard')]",
            ])
            if not clicked:
                # Scan ALL links on page (same as local script "scan" approach)
                for el in driver.find_elements(By.TAG_NAME, "a"):
                    try:
                        if "Returns Dashboard" in (el.text or "") and el.is_displayed():
                            driver.execute_script("arguments[0].click();", el)
                            log("  Returns Dashboard clicked via scan ✓")
                            clicked = True
                            break
                    except: continue
            # Smart wait: wait for URL to change to dashboard or timeout after 8s
            try:
                WebDriverWait(driver, 8).until(
                    lambda d: "return.gst.gov.in" in d.current_url and "dashboard" in d.current_url
                )
            except: pass
            time.sleep(1)

            final = driver.current_url
            log(f"  URL after nav attempt {attempt+1}: {final}")

            if "accessdenied" in final.lower():
                log("  Access Denied — re-logging in...", "warn")
                _do_login()
                continue

            if "return.gst.gov.in" in final and "dashboard" in final:
                log("  ✅ Returns Dashboard loaded", "ok")
                return True

            # Still on wrong page — log what's visible to help diagnose
            try:
                links = [(a.text.strip(), a.get_attribute("href") or "")
                         for a in driver.find_elements(By.TAG_NAME, "a")
                         if a.is_displayed() and a.text.strip()]
                log(f"  Links on page: {links[:10]}", "info")
            except: pass

        raise RuntimeError(f"Could not reach Returns Dashboard. Last URL: {driver.current_url}")

    def _select_and_search(month_name):
        """Select FY, Quarter, Period then click SEARCH (mirrors select_and_search)"""
        log(f"  Setting: FY={fy}  Quarter={QUARTER_MAP_LOCAL.get(month_name,'')}  Period={month_name}")
        time.sleep(1)

        all_sels = driver.find_elements(By.TAG_NAME, "select")
        # FY
        for sel_el in all_sels:
            try:
                s = Select(sel_el)
                opts = [o.text.strip() for o in s.options]
                if any("-" in o and len(o) <= 9 for o in opts):
                    for opt in s.options:
                        if fy in opt.text:
                            s.select_by_visible_text(opt.text)
                            log(f"  FY: {opt.text} ✓")
                            break
                    break
            except: continue
        time.sleep(0.3)

        all_sels = driver.find_elements(By.TAG_NAME, "select")
        # Quarter
        qtr = QUARTER_MAP_LOCAL.get(month_name, "")
        for sel_el in all_sels:
            try:
                s = Select(sel_el)
                opts = [o.text.strip() for o in s.options]
                if any("quarter" in o.lower() for o in opts):
                    for opt in s.options:
                        if qtr[:9].lower() in opt.text.lower():
                            s.select_by_visible_text(opt.text)
                            log(f"  Quarter: {opt.text} ✓")
                            break
                    break
            except: continue
        time.sleep(0.3)

        all_sels = driver.find_elements(By.TAG_NAME, "select")
        # Period/Month
        month_names_lower = ["january","february","march","april","may","june",
                             "july","august","september","october","november","december"]
        for sel_el in all_sels:
            try:
                s = Select(sel_el)
                opts = [o.text.strip() for o in s.options]
                if any(m in " ".join(opts).lower() for m in month_names_lower):
                    for opt in s.options:
                        if month_name.lower() in opt.text.lower():
                            s.select_by_visible_text(opt.text)
                            log(f"  Period: {opt.text} ✓")
                            break
                    break
            except: continue
        time.sleep(0.3)

        # SEARCH
        clicked = _try_click([
            "//button[normalize-space()='SEARCH']",
            "//button[normalize-space()='Search']",
            "//button[contains(text(),'SEARCH')]",
            "//input[@value='SEARCH']",
        ])
        if not clicked:
            driver.execute_script("""
                var btns=document.querySelectorAll('button,input[type=submit]');
                for(var i=0;i<btns.length;i++){
                    if((btns[i].innerText||btns[i].value||'').toUpperCase().includes('SEARCH')){
                        btns[i].click(); break;
                    }
                }
            """)
        # Smart wait: wait for GSTR tiles to appear (max 8s)
        try:
            WebDriverWait(driver, 8).until(
                lambda d: any(t in d.find_element(By.TAG_NAME, "body").text
                              for t in ["GSTR-1","GSTR-2","GSTR-3","GSTR1","GSTR2","GSTR3"])
            )
        except: time.sleep(2)
        log(f"  Tiles loaded ✓")

    def _click_tile_download(tile_name):
        """Find tile and click its DOWNLOAD button"""
        log(f"  Finding {tile_name} tile DOWNLOAD button...")
        time.sleep(0.5)

        name_variants = {
            "GSTR1":  ["GSTR1","GSTR-1","GSTR 1","gstr1","Gstr1"],
            "GSTR1A": ["GSTR1A","GSTR-1A","GSTR 1A","gstr1a"],
            "GSTR2B": ["GSTR2B","GSTR-2B","GSTR 2B","gstr2b"],
            "GSTR2A": ["GSTR2A","GSTR-2A","GSTR 2A","gstr2a"],
            "GSTR3B": ["GSTR3B","GSTR-3B","GSTR 3B","gstr3b"],
        }
        variants = name_variants.get(tile_name.upper().replace("-",""), [tile_name])

        # Strategy 1: find subtitle text → walk up to container → find DOWNLOAD button inside
        for variant in variants:
            try:
                subtitle_els = driver.find_elements(By.XPATH,
                    f"//*[normalize-space(text())='{variant}' or "
                    f"contains(normalize-space(text()),'{variant}')]")
                for subtitle_el in subtitle_els:
                    if not subtitle_el.is_displayed(): continue
                    parent = subtitle_el
                    for level in range(8):
                        try:
                            parent = driver.execute_script("return arguments[0].parentElement;", parent)
                            if parent is None: break
                            btns = parent.find_elements(By.XPATH,
                                ".//*[contains(translate(normalize-space(.),'download','DOWNLOAD'),'DOWNLOAD') "
                                "and (self::button or self::a)]")
                            for btn in btns:
                                if btn.is_displayed():
                                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                                    time.sleep(0.4)
                                    driver.execute_script("arguments[0].click();", btn)
                                    log(f"  ✅ {tile_name} DOWNLOAD clicked (strategy 1, level {level})", "ok")
                                    return True
                        except: break
            except: continue

        # Strategy 2: scan all DOWNLOAD buttons, find the one near the tile title
        try:
            all_dl_btns = driver.find_elements(By.XPATH,
                "//*[contains(translate(normalize-space(text()),'download','DOWNLOAD'),'DOWNLOAD') "
                "and (self::button or self::a) and not(contains(text(),'GENERATE'))]")
            for btn in all_dl_btns:
                if not btn.is_displayed(): continue
                # Check if any ancestor contains the tile name
                try:
                    parent = btn
                    for _ in range(10):
                        parent = driver.execute_script("return arguments[0].parentElement;", parent)
                        if parent is None: break
                        ptext = (driver.execute_script("return arguments[0].innerText;", parent) or "").upper()
                        tile_key = tile_name.upper().replace("-","")
                        if tile_key in ptext.replace("-","").replace(" ",""):
                            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                            time.sleep(0.4)
                            driver.execute_script("arguments[0].click();", btn)
                            log(f"  ✅ {tile_name} DOWNLOAD clicked (strategy 2)", "ok")
                            return True
                except: continue
        except Exception as e:
            log(f"  Strategy 2 error: {e}", "warn")

        log(f"  ⚠ {tile_name} DOWNLOAD tile not found — check page dump above", "warn")
        return False

    def _get_latest_file(extensions):
        files = []
        for ext in extensions:
            files.extend(dl_dir.glob(f"*{ext}"))
        if not files: return None
        return max(files, key=lambda f: f.stat().st_mtime)

    def _rename_latest(save_name, extensions):
        try:
            f = _get_latest_file(extensions)
            if f:
                dest = dl_dir / save_name
                if not dest.exists():
                    f.rename(dest)
                log(f"  ✅ Saved: {save_name}", "ok")
                return True
        except Exception as e:
            log(f"  Rename failed: {e}", "warn")
        return False

    def _generate_and_download(save_name, gen_xpaths, dl_extensions, max_wait=120):
        """Click GENERATE → immediately check for auto-download → poll for download link.
        GSTR-2B / GSTR-3B Excel downloads start instantly when Generate is clicked.
        GSTR-1 / GSTR-2A need portal to generate (may take 30s-2min)."""
        import time as _t
        start_time = _t.time()
        time.sleep(3)
        log(f"  Generate page: {driver.current_url}")

        # List files already in dl_dir before clicking
        before_files = {str(f): f.stat().st_mtime for f in dl_dir.iterdir()
                        if f.suffix.lower() in [e.lower() for e in dl_extensions]}

        gen_clicked = _try_click(gen_xpaths, timeout=10)
        if gen_clicked:
            log(f"  GENERATE clicked — checking for instant download...")
        else:
            log(f"  ⚠ GENERATE button not found — checking for existing link...", "warn")

        # ── Immediate download check (GSTR-2B Excel / GSTR-3B Excel download instantly) ──
        for _chk in range(25):   # up to 5 seconds
            time.sleep(0.2)
            for f in dl_dir.iterdir():
                if f.suffix.lower() not in [e.lower() for e in dl_extensions]:
                    continue
                # New file or grown file?
                prev_mtime = before_files.get(str(f))
                if prev_mtime is None or f.stat().st_mtime > prev_mtime + 0.1:
                    if f.stat().st_size > 500 and not f.name.endswith(".crdownload"):
                        log(f"  ⚡ Instant download detected: {f.name}")
                        time.sleep(1)  # let it finish
                        if _rename_latest(save_name, dl_extensions):
                            return True

        # ── Polling: wait for download link to appear (GSTR-1 / GSTR-2A which need portal generation) ──
        DOWNLOAD_XP = [
            "//a[contains(text(),'Click here to download')]",
            "//a[contains(text(),'click here to download')]",
            "//a[contains(text(),'File 1')]",
            "//a[contains(text(),'File 2')]",
            "//a[contains(@href,'.xlsx')]",
            "//a[contains(@href,'.zip')]",
            "//a[contains(@href,'filedownload')]",
            "//a[contains(@href,'download') and string-length(@href) > 50]",
            "//button[contains(text(),'Download') or contains(text(),'DOWNLOAD')]",
        ]

        elapsed = 0
        while elapsed < max_wait:
            for xp in DOWNLOAD_XP:
                try:
                    els = driver.find_elements(By.XPATH, xp)
                    for el in els:
                        if el.is_displayed():
                            href = el.get_attribute("href") or ""
                            if len(href) > 20:
                                txt = el.text.strip() or href[:50]
                                log(f"  Download link found: '{txt}'")
                                log(f"  Downloading: {save_name}")
                                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                                time.sleep(0.5)
                                driver.execute_script("arguments[0].click();", el)
                                time.sleep(2)
                                if _rename_latest(save_name, dl_extensions):
                                    return True
                except: continue

            # Also check dl_dir for any new file that may have arrived
            for f in dl_dir.iterdir():
                if f.suffix.lower() not in [e.lower() for e in dl_extensions]:
                    continue
                prev_mtime = before_files.get(str(f))
                if (prev_mtime is None or f.stat().st_mtime > prev_mtime + 0.1) and \
                   f.stat().st_size > 500 and not f.name.endswith(".crdownload"):
                    log(f"  📥 File appeared in downloads: {f.name}")
                    time.sleep(1)
                    if _rename_latest(save_name, dl_extensions):
                        return True

            elapsed += 3
            time.sleep(3)

            if elapsed % 30 == 0:
                log(f"  Still waiting... ({elapsed}s) — refreshing page")
                try:
                    driver.refresh()
                    time.sleep(2)
                    if _is_session_lost():
                        log("  Session lost during wait — re-logging in...", "warn")
                        _do_login()
                        _go_to_dashboard()
                        _select_and_search(current_month[0])
                        _click_tile_download(current_tile[0])
                        time.sleep(2)
                except: pass

        log(f"  ⚠ No download link found for {save_name} after {max_wait}s", "warn")
        return False

    # Mutable refs for session recovery inside _generate_and_download
    current_month = [""]
    current_tile  = [""]

    GENERATE_JSON_XP = [
        "//button[contains(text(),'GENERATE JSON FILE TO DOWNLOAD')]",
        "//button[contains(text(),'GENERATE JSON')]",
        "//button[contains(text(),'Generate JSON')]",
        "//a[contains(text(),'GENERATE JSON')]",
    ]
    GENERATE_EXCEL_XP = [
        "//button[contains(text(),'GENERATE EXCEL FILE TO DOWNLOAD')]",
        "//button[contains(text(),'GENERATE EXCEL')]",
        "//button[contains(text(),'Generate Excel')]",
        "//a[contains(text(),'GENERATE EXCEL')]",
        # fallback to JSON if no Excel button
        "//button[contains(text(),'GENERATE JSON FILE TO DOWNLOAD')]",
        "//button[contains(text(),'Generate JSON')]",
    ]

    returns_set = set()
    if returns in ("all","gstr1"):  returns_set.add("GSTR1")
    if returns in ("all","gstr1a"): returns_set.add("GSTR1A")
    if returns in ("all","gstr2b"): returns_set.add("GSTR2B")
    if returns in ("all","gstr2a"): returns_set.add("GSTR2A")
    if returns in ("all","gstr3b"): returns_set.add("GSTR3B")
    if returns == "gstr1":  returns_set = {"GSTR1"}
    if returns == "gstr1a": returns_set = {"GSTR1A"}
    if returns == "gstr2b": returns_set = {"GSTR2B"}
    if returns == "gstr2a": returns_set = {"GSTR2A"}
    if returns == "gstr3b": returns_set = {"GSTR3B"}
    # Extra combo options (no GSTR-2A)
    if returns == "gstr1_2b_3b":    returns_set = {"GSTR1", "GSTR2B", "GSTR3B"}
    if returns == "gstr1_1a_2b_3b": returns_set = {"GSTR1", "GSTR1A", "GSTR2B", "GSTR3B"}

    # ── Setup browser ────────────────────────────────────────────────
    try:
        opts = ChromeOptions()
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--window-size=1280,900")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        # Speed flags for Render/Linux server only — these crash Chrome on Windows
        if platform.system() != "Windows":
            opts.add_argument("--no-zygote")
            opts.add_argument("--single-process")
        opts.add_argument("--disable-setuid-sandbox")
        opts.add_argument("--disable-software-rasterizer")
        opts.add_argument("--disable-background-networking")
        opts.add_argument("--disable-default-apps")
        opts.add_argument("--disable-sync")
        opts.add_argument("--metrics-recording-only")
        opts.add_argument("--mute-audio")
        opts.add_argument("--no-first-run")
        opts.add_argument("--safebrowsing-disable-auto-update")
        opts.add_argument("--disable-background-timer-throttling")
        opts.add_argument("--disable-renderer-backgrounding")
        opts.add_argument("--disable-backgrounding-occluded-windows")
        opts.add_argument("--memory-pressure-off")
        opts.add_argument(
            "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        opts.add_experimental_option("prefs", {
            "download.default_directory": str(dl_dir),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        })
        opts.add_experimental_option("excludeSwitches", ["enable-automation","enable-logging"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.page_load_strategy = "eager"

        try:
            import shutil as _sh
            _IS_SERVER = bool(os.environ.get("RENDER") or os.environ.get("HEADLESS"))
            # Prefer explicit env vars set in Dockerfile/render.yaml
            _cb  = (os.environ.get("CHROME_BIN")
                    or _sh.which("chromium") or _sh.which("chromium-browser")
                    or _sh.which("google-chrome"))
            _cd  = (os.environ.get("CHROMEDRIVER_PATH")
                    or _sh.which("chromedriver"))
            log(f"  Chrome binary : {_cb}", "info")
            log(f"  Chromedriver  : {_cd}", "info")
            if _IS_SERVER and _cb:
                opts.binary_location = _cb
                svc = ChromeService(executable_path=_cd) if _cd else ChromeService()
                driver = webdriver.Chrome(service=svc, options=opts)
            else:
                from webdriver_manager.chrome import ChromeDriverManager
                driver = webdriver.Chrome(
                    service=ChromeService(ChromeDriverManager().install()), options=opts)
        except Exception as _ce:
            log(f"  ✗ Chrome failed to start: {_ce}", "err")
            raise

        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
        })
        log("✅ Headless Chrome started", "ok")
        prog(5)

        # ── Step 1: Login ────────────────────────────────────────────
        if not password and not token:
            # Token mode — inject cookie instead of full login
            log("  Using session token to authenticate...")
            driver.get("https://services.gst.gov.in/services/login")
            time.sleep(2)
            driver.add_cookie({"name": "AuthToken", "value": token, "domain": ".gst.gov.in"})
            driver.add_cookie({"name": "token",     "value": token, "domain": ".gst.gov.in"})
            driver.get("https://services.gst.gov.in/services/auth/api/profile")
            time.sleep(3)
        else:
            _do_login()

        prog(15)

        # ══════════════════════════════════════════════════════════════
        # Step 2: FAST DOWNLOAD LOOP
        # ── KEY FIXES ──
        # GSTR-2B: Portal ALWAYS shows a generate page after tile click.
        #          We ALWAYS click GENERATE EXCEL first, then poll for the
        #          instant download — never rely on direct file arriving
        #          without clicking Generate. This is why only April worked
        #          before (Generate was only clicked as a fallback).
        # GSTR-3B: PDF downloads directly — use tight 0.5s polling loop
        #          instead of a fixed 6s sleep → 3-4x faster.
        # SPEED:   Removed all fixed sleeps; use smart waits throughout.
        # ══════════════════════════════════════════════════════════════
        triggered = {}
        total_months = len(MONTHS_LIST)

        # ── Identify case type for phased flow ───────────────────
        _DIRECT_ONLY_TYPES = {"GSTR2B", "GSTR3B"}
        _is_fast_case = returns_set.issubset(_DIRECT_ONLY_TYPES)
        _generate_types = [r for r in ("GSTR1","GSTR1A","GSTR2A") if r in returns_set]
        _direct_types   = [r for r in ("GSTR2B","GSTR3B")         if r in returns_set]

        log(f"\n{'═'*55}")
        log(f"📋 DOWNLOAD CASE    : {returns}  →  {sorted(returns_set)}")
        if _is_fast_case:
            log(f"⚡ TYPE             : FAST — direct download, no generate needed")
            log(f"   Tabs needed     : {len(returns_set) * 12} tabs | Est time: ~{len(returns_set)*5} min")
        else:
            log(f"📋 TYPE             : PHASED download flow")
            log(f"   Phase 1         : Generate {_generate_types} — all {total_months} months")
            log(f"   Phase 2+3       : Direct download {_direct_types} — keeps session alive")
            log(f"   Phase 4         : Collect generate links {_generate_types}")
        log(f"   Retry policy    : T1 → 30s wait → T2 → 60s wait → T3 → report fail")
        log(f"   Batch size      : 6 months per batch (avoids portal Not Found)")
        log(f"   Session timeout : Downloading 2B+3B in Phase 2+3 prevents 20-min logout")
        log(f"{'═'*55}\n")

        log(f"\n📋 Downloading {total_months} months × {', '.join(sorted(returns_set))} ...")

        def _fast_wait_file(extensions, before_snap, timeout=90):
            """Poll dl_dir every 0.5s for a new complete file. Much faster than fixed sleeps."""
            deadline = time.time() + timeout
            while time.time() < deadline:
                time.sleep(0.5)
                for f in dl_dir.iterdir():
                    if f.suffix.lower() not in extensions: continue
                    if f.name.endswith((".crdownload", ".tmp")): continue
                    prev = before_snap.get(str(f))
                    if (prev is None or f.stat().st_mtime > prev + 0.1) and f.stat().st_size > 500:
                        time.sleep(0.8)   # let chrome finish flushing
                        return f
            return None

        def _snap_dl_dir(extensions):
            return {str(f): f.stat().st_mtime for f in dl_dir.iterdir()
                    if f.suffix.lower() in extensions}

        def _save_file(src_f, save_name, month_key, return_key):
            """Copy downloaded file to out_dir and register it."""
            dest = out_dir / save_name
            try: _shutil.copy2(str(src_f), str(dest))
            except: pass
            sz = src_f.stat().st_size // 1024
            triggered[f"{month_key}_{return_key}"] = "OK"
            downloaded.append({"name": save_name, "size": f"{sz} KB"})
            with jobs_lock:
                if job_id in jobs: jobs[job_id]["files"] = list(downloaded)
            log(f"  ✅ Saved: {save_name} ({sz} KB)", "ok")

        # ══════════════════════════════════════════════════════════════
        # PHASED DOWNLOAD — replaces old per-month sequential loop
        #
        # OLD: for each month → do all returns  (mixed, slow)
        # NEW: for each return → do all months  (phased, fast)
        #
        # Phase 1 : Trigger GENERATE for GSTR-1, GSTR-1A, GSTR-2A
        #           ALL 12 months quickly — click & move on, no waiting
        # Phase 2 : Download GSTR-2B ALL 12 months  (INSTANT — no wait)
        # Phase 3 : Download GSTR-3B ALL 12 months  (INSTANT — no wait)
        # Phase 4 : Collect DOWNLOAD LINKS for GSTR-1, GSTR-1A, GSTR-2A
        #           (now ready — portal generated during Phase 2+3)
        #
        # FAST CASE (2B only / 3B only / 2B+3B):
        #   Skip Phase 1 entirely → do Phase 2+3 only → done in ~5-10 min
        # ══════════════════════════════════════════════════════════════

        def _do_month_return(month_name, month_num, year, rtype, action):
            """
            Navigate to dashboard, search for month, perform action for rtype.
            action: 'generate' | 'direct_2b' | 'direct_3b' | 'collect'
            Returns 'OK' | 'TILE_FAIL' | 'GEN_FAIL' | 'NOT_FOUND' | 'ERR:...'
            """
            key = f"{month_name}_{year}"
            save_name = {
                "GSTR1":  f"GSTR1_{month_name}_{year}.zip",
                "GSTR1A": f"GSTR1A_{month_name}_{year}.zip",
                "GSTR2A": f"GSTR2A_{month_name}_{year}.zip",
                "GSTR2B": f"GSTR2B_{month_name}_{year}.xlsx",
                "GSTR3B": f"GSTR3B_{month_name}_{year}.pdf",
            }[rtype]

            # Skip if already downloaded
            if (dl_dir / save_name).exists() or (out_dir / save_name).exists():
                log(f"  ✓ {rtype} {month_name} already downloaded — skip", "ok")
                return "OK"

            _go_to_dashboard()
            _select_and_search(month_name)
            current_tile[0] = rtype
            current_month[0] = month_name

            if not _click_tile_download(rtype):
                save_failure_screenshot(f"{rtype} {month_name} {year} — Tile Not Found")
                return "TILE_FAIL"

            # ── Generate trigger (GSTR-1, GSTR-1A, GSTR-2A) ──────────
            if action == "generate":
                time.sleep(0.8)
                xp = GENERATE_JSON_XP if rtype in ("GSTR1","GSTR1A") else GENERATE_EXCEL_XP
                if _try_click(xp, timeout=8):
                    log(f"  ✓ GENERATE clicked {rtype} {month_name}")
                    time.sleep(0.5)
                    return "TRIGGERED"
                else:
                    save_failure_screenshot(f"{rtype} {month_name} {year} — Generate button not found")
                    return "GEN_FAIL"

            # ── GSTR-2B instant download ───────────────────────────────
            elif action == "direct_2b":
                time.sleep(0.8)
                snap = _snap_dl_dir({".xlsx", ".zip", ".json"})
                gen_ok = _try_click(GENERATE_EXCEL_XP, timeout=8)
                if not gen_ok:
                    save_failure_screenshot(f"GSTR2B {month_name} {year} — Generate Excel button not found")
                new_f = _fast_wait_file({".xlsx", ".zip", ".json"}, snap, timeout=60)
                if new_f:
                    dest = dl_dir / save_name
                    if str(new_f) != str(dest):
                        try: new_f.rename(dest)
                        except: _shutil.copy2(str(new_f), str(dest))
                    _save_file(dest, save_name, key, "GSTR2B")
                    return "OK"
                else:
                    save_failure_screenshot(f"GSTR2B {month_name} {year} — File not received after Generate")
                    return "NOT_FOUND"

            # ── GSTR-3B direct PDF download ────────────────────────────
            elif action == "direct_3b":
                snap = _snap_dl_dir({".pdf"})
                new_f = _fast_wait_file({".pdf"}, snap, timeout=30)
                if new_f:
                    dest = dl_dir / save_name
                    if str(new_f) != str(dest):
                        try: new_f.rename(dest)
                        except: _shutil.copy2(str(new_f), str(dest))
                    _save_file(dest, save_name, key, "GSTR3B")
                    return "OK"
                else:
                    save_failure_screenshot(f"GSTR3B {month_name} {year} — PDF not received in 30s")
                    return "NOT_FOUND"

            # ── Collect generate link (Phase 4) ────────────────────────
            elif action == "collect":
                time.sleep(1)
                xp = GENERATE_JSON_XP if rtype in ("GSTR1","GSTR1A") else GENERATE_EXCEL_XP
                exts = [".zip",".json"] if rtype in ("GSTR1","GSTR1A") else [".zip",".xlsx"]
                if _generate_and_download(save_name, xp, exts, max_wait=120):
                    src_f = dl_dir / save_name
                    try: _shutil.copy2(str(src_f), str(out_dir / save_name))
                    except: pass
                    sz = (out_dir / save_name).stat().st_size // 1024
                    downloaded.append({"name": save_name, "size": f"{sz} KB"})
                    with jobs_lock:
                        if job_id in jobs: jobs[job_id]["files"] = list(downloaded)
                    return "OK"
                else:
                    save_failure_screenshot(f"{rtype} {month_name} {year} — Download link not found (Phase 4)")
                    return "NOT_FOUND"

            return "UNKNOWN"

        def _run_phase(phase_label, rtype_list, action, retry_waits=(0,30,60)):
            """
            Run one phase: for each return in rtype_list, loop all 12 months.
            Retries T1 → T2 → T3 on failure.
            """
            if not rtype_list:
                return
            log(f"\n{'═'*55}")
            log(f"  {phase_label}")
            log(f"  Returns: {rtype_list}  |  Action: {action}  |  Retry: T1→T2→T3")
            log(f"{'═'*55}")
            for rtype in rtype_list:
                log(f"\n  ── {rtype} — all 12 months ──")
                for midx, (month_name, month_num, year) in enumerate(MONTHS_LIST):
                    key = f"{month_name}_{year}"
                    current_month[0] = month_name
                    prog(15 + int((midx) / total_months * 20))
                    result = "FAIL"
                    for attempt, wait in enumerate(retry_waits, 1):
                        if wait > 0:
                            log(f"    [T{attempt}] Waiting {wait}s before retry {rtype} {month_name}...")
                            time.sleep(wait)
                        try:
                            result = _do_month_return(month_name, month_num, year, rtype, action)
                            if result in ("OK","TRIGGERED"):
                                if attempt > 1:
                                    log(f"    [T{attempt}] ✓ Success on retry: {rtype} {month_name}", "ok")
                                break
                            else:
                                log(f"    [T{attempt}] ✗ {result}: {rtype} {month_name}", "warn")
                        except Exception as e:
                            result = f"ERR:{e}"
                            log(f"    [T{attempt}] ✗ Error: {e}", "warn")
                            save_failure_screenshot(f"{rtype} {month_name} {year} T{attempt}: {str(e)[:50]}")
                    triggered[f"{key}_{rtype}"] = result
                    status = "✅" if result in ("OK","TRIGGERED") else "❌"
                    log(f"    {status} {rtype} {month_name} {year}: {result}")

        # ── PHASE 1: Trigger GENERATE — GSTR-1, GSTR-1A, GSTR-2A ─────────
        _generate_types = [r for r in ("GSTR1","GSTR1A","GSTR2A") if r in returns_set]
        _direct_2b      = "GSTR2B" in returns_set
        _direct_3b      = "GSTR3B" in returns_set

        if _generate_types:
            _run_phase(
                "PHASE 1 — Click GENERATE (all months, no waiting — click & move on)",
                _generate_types, "generate"
            )
            prog(35)
            log("\n  ✅ Phase 1 complete — portal generating files in background (~15-20 min)")
            log("  ▶ Moving to Phase 2+3 (instant downloads) to keep session alive...")

        # ── PHASE 2: GSTR-2B — INSTANT download all 12 months ─────────────
        if _direct_2b:
            _run_phase(
                "PHASE 2 — GSTR-2B Direct Download (INSTANT — no generate wait needed)",
                ["GSTR2B"], "direct_2b"
            )
            prog(55)

        # ── PHASE 3: GSTR-3B — INSTANT download all 12 months ─────────────
        if _direct_3b:
            _run_phase(
                "PHASE 3 — GSTR-3B Direct Download (INSTANT PDF — no generate wait needed)",
                ["GSTR3B"], "direct_3b"
            )
            prog(70)

        # ── PHASE 4: Collect GENERATE links — GSTR-1, GSTR-1A, GSTR-2A ───
        # By now Phase 2+3 kept session alive and portal has finished generating
        need_phase4 = any(
            triggered.get(f"{mn}_{yr}_{rt}") == "TRIGGERED"
            for mn, mm, yr in MONTHS_LIST
            for rt in ("GSTR1","GSTR1A","GSTR2A")
        )
        if need_phase4:
            _collect_types = [r for r in ("GSTR1","GSTR1A","GSTR2A")
                              if r in returns_set and any(
                                  triggered.get(f"{mn}_{yr}_{r}") == "TRIGGERED"
                                  for mn, mm, yr in MONTHS_LIST)]
            _run_phase(
                "PHASE 4 — Collect Download Links (GSTR-1, GSTR-1A, GSTR-2A now ready)",
                _collect_types, "collect"
            )

        prog(97)

        # ── Step 4: Copy all downloaded files to job output + ZIP ───────
        log("\n📦 Packaging files...")
        import zipfile as _zf
        zip_name = f"GST_Downloads_{client_name.replace(' ','_')}_{fy}.zip"
        zip_path = out_dir / zip_name
        with _zf.ZipFile(str(zip_path), "w", _zf.ZIP_DEFLATED) as zf:
            for item in downloaded:
                src = dl_dir / item["name"]
                if src.exists():
                    dest_in_job = out_dir / item["name"]
                    _shutil.copy2(str(src), str(dest_in_job))
                    zf.write(str(src), item["name"])
        if zip_path.exists():
            sz = zip_path.stat().st_size // 1024
            downloaded.insert(0, {"name": zip_name, "size": f"{sz} KB"})
            log(f"✅ ZIP created: {zip_name} ({sz} KB)", "ok")

        prog(100)
        n = len([d for d in downloaded if not d["name"].endswith(".zip")]) if len(downloaded) > 1 else 0
        log(f"✅ Complete! {n} file(s) downloaded. Click ZIP to save all.", "ok")

        with jobs_lock:
            jobs[job_id]["status"] = "done"
            jobs[job_id]["files"]  = downloaded

    except Exception as exc:
        import traceback
        log(f"❌ Error: {exc}", "err")
        for ln in traceback.format_exc().split("\n"):
            if ln.strip(): log(f"  {ln}", "err")
        with jobs_lock:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["error"]  = str(exc)
    finally:
        if driver:
            try: driver.quit()
            except: pass


# ═══════════════════════════════════════════════════════════════════
# BULK DOWNLOAD — multiple companies from Excel list
# ═══════════════════════════════════════════════════════════════════

@app.route("/api/bulk-template")
def bulk_template():
    """Return a sample Excel template for bulk download."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Companies"
    headers = ["COMPANY NAME", "GSTIN", "USERNAME", "NOTES"]
    widths  = [30, 20, 20, 30]
    for i,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1a2235")
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64+i)].width = w
    examples = [
        ["ABC Traders", "33ABCDE1234F1ZX", "abctraders_gst", "Example row"],
        ["XYZ Pvt Ltd", "29XYZAB5678G2ZY", "xyz_gst_login",  ""],
    ]
    for row in examples:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bulk_companies_template.xlsx"})


@app.route("/api/bulk-start", methods=["POST"])
@rate_limit(limit=5, window=60)
def api_bulk_start():
    fobj = request.files.get("companies_file")
    if not fobj:
        return jsonify(error="No file uploaded"), 400
    fy      = request.form.get("fy","2025-26")
    returns = request.form.get("returns","all")

    # Parse the Excel
    import io, openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(fobj.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return jsonify(error=f"Cannot read Excel: {e}"), 400

    if not rows or len(rows) < 2:
        return jsonify(error="Excel is empty or has only headers"), 400

    headers = [str(c or "").strip().upper() for c in rows[0]]
    def _col(*names):
        for n in names:
            if n in headers: return headers.index(n)
        return -1

    ci_name = _col("COMPANY NAME","NAME","COMPANY")
    ci_gst  = _col("GSTIN","GSTIN NO","GST")
    ci_user = _col("USERNAME","USER","LOGIN","USER NAME")

    if ci_gst < 0:
        return jsonify(error="Column 'GSTIN' not found in Excel"), 400

    companies = []
    for row in rows[1:]:
        gstin = str(row[ci_gst] or "").strip().upper() if ci_gst >= 0 else ""
        if not gstin or len(gstin) != 15:
            continue
        companies.append({
            "name":     str(row[ci_name] or "").strip() if ci_name >= 0 else gstin,
            "gstin":    gstin,
            "username": str(row[ci_user] or "").strip() if ci_user >= 0 else "",
        })

    if not companies:
        return jsonify(error="No valid GSTINs found in Excel"), 400

    job_id  = str(uuid.uuid4())[:8]
    out_dir = OUTPUT_DIR / job_id
    out_dir.mkdir(parents=True, exist_ok=True)

    with jobs_lock:
        jobs[job_id] = {
            "status":"running","progress":0,
            "logs":[{"type":"info","msg":f"Loaded {len(companies)} companies. Starting…"}],
            "files":[],"error":None,
            "captcha_needed":False,"captcha_img":None,"captcha_company":None,
            "out_dir":str(out_dir),"counter":"",
        }

    sess = {"token_q": _queue.Queue(), "screenshot": None, "refresh_event": threading.Event()}
    with _sess_lock:
        _sessions[job_id] = sess

    def _run():
        try:
            _bulk_worker(job_id, companies, fy, returns, sess, out_dir)
        except Exception as exc:
            import traceback as _tb
            with jobs_lock:
                if job_id in jobs:
                    jobs[job_id]["status"] = "error"
                    jobs[job_id]["error"]  = str(exc)
                    for ln in _tb.format_exc().split("\n"):
                        if ln.strip():
                            jobs[job_id]["logs"].append({"type":"err","msg":ln})
        finally:
            with _sess_lock:
                _sessions.pop(job_id, None)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify(job_id=job_id, total=len(companies))


@app.route("/api/bulk-token/<job_id>", methods=["POST"])
def api_bulk_token(job_id):
    """User submits token for a company during bulk download."""
    token = (request.get_json(silent=True) or {}).get("token","").strip()
    if not token:
        return jsonify(ok=False, error="Empty token")
    with _sess_lock:
        sess = _sessions.get(job_id)
    if not sess:
        return jsonify(ok=False, error="No active session")
    sess["token_q"].put(token)
    # Clear captcha_needed immediately so UI hides the card
    with jobs_lock:
        if job_id in jobs:
            jobs[job_id]["captcha_needed"]  = False
            jobs[job_id]["captcha_company"] = None
    return jsonify(ok=True)


def _bulk_worker(job_id, companies, fy, returns, sess, out_dir):
    """Process each company one by one, requesting a token for each."""
    import requests as _req

    def log(msg, t="info"):
        print(f"[BULK {job_id}] {msg}")
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["logs"].append({"type":t,"msg":msg})

    def prog(p):
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["progress"] = p

    def set_counter(i, total):
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["counter"] = f"Company {i}/{total}"

    total = len(companies)
    all_files = []
    out_path  = Path(out_dir)

    for idx, company in enumerate(companies, 1):
        set_counter(idx, total)
        name     = company["name"]
        gstin    = company["gstin"]
        username = company["username"]
        log(f"━━━ [{idx}/{total}] {name} ({gstin}) ━━━", "info")
        prog(int((idx-1)/total*100))

        # ── Ask user for token ──────────────────────────────────────
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["captcha_needed"]  = True
                jobs[job_id]["captcha_company"] = {"name":name,"gstin":gstin,"username":username}

        log(f"  Waiting for AuthToken from user for {name}…")

        # Wait up to 15 minutes for token
        try:
            token = sess["token_q"].get(timeout=900)
        except _queue.Empty:
            log(f"  ⏱ Timeout waiting for token — skipping {name}", "warn")
            continue

        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["captcha_needed"]  = False
                jobs[job_id]["captcha_company"] = None

        log(f"  Token received — starting download for {name}…", "ok")

        # ── Download returns for this company ───────────────────────
        company_dir = out_path / gstin
        company_dir.mkdir(exist_ok=True)

        S = _req.Session()
        S.headers.update({
            "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept":          "application/json, text/plain, */*",
            "Accept-Language": "en-IN,en-US;q=0.9,en;q=0.8",
            "Authorization":   f"Bearer {token}",
        })
        S.cookies.set("AuthToken", token, domain=".gst.gov.in")
        S.cookies.set("token",     token, domain=".gst.gov.in")

        # Verify token
        try:
            vr = S.get(f"https://services.gst.gov.in/services/api/search/taxpayerDetails?gstin={gstin}",
                       timeout=15)
            if vr.status_code in (401, 403):
                log(f"  ⚠ Token rejected for {name} — skipping", "warn")
                continue
        except Exception as ve:
            log(f"  Token check warning: {ve} — continuing", "warn")

        fy_start = int(fy.split("-")[0])
        MONTHS = [
            ("April","04",str(fy_start)),    ("May","05",str(fy_start)),
            ("June","06",str(fy_start)),     ("July","07",str(fy_start)),
            ("August","08",str(fy_start)),   ("September","09",str(fy_start)),
            ("October","10",str(fy_start)),  ("November","11",str(fy_start)),
            ("December","12",str(fy_start)), ("January","01",str(fy_start+1)),
            ("February","02",str(fy_start+1)),("March","03",str(fy_start+1)),
        ]
        BASE_RET = "https://return.gst.gov.in/returns/auth"
        co_files = []

        def dl_one(ret_type, mon_name, mon_num, mon_yr):
            period = f"{mon_num}{mon_yr}"
            ext    = {"gstr1":".zip","gstr1a":".zip","gstr2b":".xlsx",
                      "gstr2a":".xlsx","gstr3b":".pdf"}.get(ret_type,".zip")
            fname  = f"{gstin}_{ret_type.upper()}_{mon_name}_{mon_yr}{ext}"
            fpath  = company_dir / fname
            urls   = [
                f"{BASE_RET}/{ret_type}/download?gstin={gstin}&ret_period={period}&action_type=download",
                f"{BASE_RET}/{ret_type}?action=download&gstin={gstin}&ret_period={period}",
                f"https://return.gst.gov.in/returns/api/{ret_type}/{gstin}/{period}/download",
            ]
            for url in urls:
                try:
                    r = S.get(url, timeout=60)
                    if r.status_code == 200 and len(r.content) > 500:
                        # Check if token expired (portal returns JSON error instead of file)
                        if r.headers.get("content-type","").startswith("application/json"):
                            ec = r.json().get("errorCode","")
                            if ec in ("AUTH4033","AUTH4035","SWEB_9000"):
                                log(f"  ⚠ Token expired mid-download for {name}", "warn")
                                return "TOKEN_EXPIRED"
                        fpath.write_bytes(r.content)
                        sz = fpath.stat().st_size // 1024
                        co_files.append({"name":fname,"size":f"{sz} KB"})
                        log(f"  ✓ {ret_type.upper()} {mon_name} {mon_yr} ({sz} KB)", "ok")
                        return "OK"
                except Exception:
                    pass
            log(f"  – {ret_type.upper()} {mon_name} {mon_yr}: not available", "warn")
            return "SKIP"

        ret_types = []
        if returns in ("all","gstr1"):  ret_types.append("gstr1")
        if returns in ("all","gstr1a"): ret_types.append("gstr1a")
        if returns in ("all","gstr2b"): ret_types.append("gstr2b")
        if returns in ("all","gstr2a"): ret_types.append("gstr2a")
        if returns in ("all","gstr3b"): ret_types.append("gstr3b")
        # Combo options (no GSTR-2A)
        if returns == "gstr1_2b_3b":    ret_types = ["gstr1","gstr2b","gstr3b"]
        if returns == "gstr1_1a_2b_3b": ret_types = ["gstr1","gstr1a","gstr2b","gstr3b"]

        token_expired = False
        for rt in ret_types:
            if token_expired: break
            log(f"  ── {rt.upper()} ──")
            for mn, mm, my in MONTHS:
                result = dl_one(rt, mn, mm, my)
                if result == "TOKEN_EXPIRED":
                    token_expired = True
                    # Ask user for a fresh token
                    log(f"  🔄 Token expired — requesting new token for {name}…", "warn")
                    with jobs_lock:
                        if job_id in jobs:
                            jobs[job_id]["captcha_needed"]  = True
                            jobs[job_id]["captcha_company"] = {
                                "name": f"{name} (RE-LOGIN)", "gstin": gstin, "username": username}
                    # Clear token queue and wait
                    while not sess["token_q"].empty():
                        try: sess["token_q"].get_nowait()
                        except: pass
                    try:
                        new_token = sess["token_q"].get(timeout=600)
                        S.headers["Authorization"] = f"Bearer {new_token}"
                        S.cookies.set("AuthToken", new_token, domain=".gst.gov.in")
                        S.cookies.set("token",     new_token, domain=".gst.gov.in")
                        with jobs_lock:
                            if job_id in jobs:
                                jobs[job_id]["captcha_needed"]  = False
                                jobs[job_id]["captcha_company"] = None
                        token_expired = False
                        log(f"  ✅ New token received — resuming {name}…", "ok")
                        # Retry this month
                        dl_one(rt, mn, mm, my)
                    except _queue.Empty:
                        log(f"  ⏱ Re-login timeout — stopping {name}", "warn")
                        break

        all_files.extend(co_files)
        log(f"  ✅ {name}: {len(co_files)} file(s) downloaded", "ok")

        # Update running file list
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["files"] = list(all_files)

    # ── Create ZIP of everything ────────────────────────────────────
    prog(98)
    if all_files:
        zip_name = f"BULK_DOWNLOAD_{fy}.zip"
        zip_path = out_path / zip_name
        import zipfile as _zf
        with _zf.ZipFile(str(zip_path), "w", _zf.ZIP_DEFLATED) as zf:
            for f in all_files:
                # find the file
                for sub in out_path.rglob(f["name"]):
                    zf.write(str(sub), f["name"])
                    break
        sz = zip_path.stat().st_size // 1024
        all_files.insert(0, {"name": zip_name, "size": f"{sz} KB"})
        log(f"✅ ZIP created: {zip_name} ({sz} KB)", "ok")

    prog(100)
    log(f"✅ Bulk complete — {len(companies)} companies, {len(all_files)-1} total files.", "ok")
    with jobs_lock:
        if job_id in jobs:
            jobs[job_id]["status"] = "done"
            jobs[job_id]["files"]  = all_files
            jobs[job_id]["captcha_needed"]  = False
            jobs[job_id]["captcha_company"] = None


# ── Startup ───────────────────────────────────────────────────────

# ═══════════════════════════════════════════════════════════════════
# INCOME TAX RECONCILIATION — Upload + Worker + Job + Download
# Same pattern as GST reconciliation routes above
# ═══════════════════════════════════════════════════════════════════

def run_it_reconciliation(job_id):
    def log(msg, t="info"):
        with jobs_lock:
            jobs[job_id]["logs"].append({"type": t, "msg": msg})
    def prog(p):
        with jobs_lock:
            jobs[job_id]["progress"] = p

    try:
        job          = jobs[job_id]
        company_name = job["company_name"]
        pan          = job["pan"]
        gstin        = job["gstin"]
        fy           = job["fy"]
        itr_form     = job.get("itr_form", "ITR-3")
        entity_type  = job.get("entity_type", "company")
        job_dir      = Path(job["job_dir"])
        out_dir      = Path(job["out_dir"])
        saved        = job["saved"]

        log(f"Starting IT Reconciliation: {company_name} ({pan}) FY {fy} | {itr_form} | {entity_type}")
        prog(5)

        # -- Move uploaded files into job_dir with engine-recognisable names --
        pdf_found = {}
        for zone, dest_prefix in [("it26as","26AS"), ("itais","AIS"), ("ittis","TIS"), ("itgst","GST_RECON")]:
            for fpath in saved.get(zone, []):
                ext  = Path(fpath).suffix.lower()
                dest = job_dir / f"{dest_prefix}{ext}"
                if not dest.exists():
                    try:    Path(fpath).rename(dest)
                    except: shutil.copy2(fpath, str(dest))
                log(f"  ✓ {dest_prefix}: {dest.name}")
                pdf_found[zone] = dest.name

        if "it26as" not in pdf_found:
            log("  ⚠ 26AS PDF not uploaded — results will be limited", "warn")
        if "ittis" not in pdf_found:
            log("  ℹ TIS PDF not uploaded — AIS data will be used for turnover figures", "info")
        if "itais" not in pdf_found:
            log("  ℹ AIS PDF not uploaded — purchase/income breakdown will be limited", "info")

        prog(20)

        # -- Load it_recon_engine.py --
        engine_path = _find_engine("it_recon_engine.py")
        if not engine_path:
            raise FileNotFoundError(
                "it_recon_engine.py not found. "
                "Place it in the same folder as app.py.")

        log("Loading IT reconciliation engine...")
        import importlib.util as _ilu
        spec = _ilu.spec_from_file_location("it_recon", str(engine_path))
        it   = _ilu.module_from_spec(spec)
        spec.loader.exec_module(it)
        prog(30)

        log(f"Parsing PDFs and generating IT Reconciliation for {itr_form}...")
        # Try with all kwargs; fall back gracefully if engine doesn't support them
        call_kwargs = {"log": log, "itr_form": itr_form, "entity_type": entity_type}
        try:
            out_xl = it.write_it_reconciliation(str(job_dir), company_name, pan, gstin, fy, **call_kwargs)
        except TypeError:
            # Engine may not support itr_form/entity_type yet — try without them
            try:
                out_xl = it.write_it_reconciliation(str(job_dir), company_name, pan, gstin, fy, log=log)
            except TypeError:
                out_xl = it.write_it_reconciliation(str(job_dir), company_name, pan, gstin, fy)
        prog(85)

        # -- Collect outputs (search both job_dir and out_dir) --------
        output_files = []
        seen = set()
        for search_dir in [job_dir, out_dir]:
            for fp in sorted(search_dir.glob("IT_RECONCILIATION_*.xlsx")):
                if fp.name in seen:
                    continue
                seen.add(fp.name)
                dest_fp = out_dir / fp.name
                if fp.parent != out_dir:
                    shutil.copy2(str(fp), str(dest_fp))
                sz = dest_fp.stat().st_size // 1024
                output_files.append({"name": fp.name, "size": f"{sz} KB"})
                log(f"  ✓ {fp.name} ({sz} KB)", "ok")

        if not output_files:
            raise RuntimeError("No IT Reconciliation Excel generated. "
                               "Check that 26AS PDF was uploaded correctly.")

        prog(100)
        log(f"Done! {len(output_files)} file(s) ready to download.", "ok")
        with jobs_lock:
            jobs[job_id]["status"] = "done"
            jobs[job_id]["files"]  = output_files
        _cleanup_uploads(job_id)

    except Exception as exc:
        import traceback
        log(f"Error: {exc}", "err")
        for line in traceback.format_exc().split("\n"):
            if line.strip(): log(f"  {line}", "err")
        with jobs_lock:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["error"]  = str(exc)
        _cleanup_uploads(job_id)


@app.route("/api/it-upload", methods=["POST"])
@rate_limit(limit=60, window=60)
def api_it_upload():
    _cleanup_old_jobs()
    company_name = request.form.get("company_name","").strip()
    pan          = request.form.get("pan","").strip().upper()
    gstin        = request.form.get("gstin","").strip().upper()
    fy           = request.form.get("fy","2025-26").strip() or "2025-26"
    itr_form     = request.form.get("itr_form","ITR-3").strip()
    entity_type  = request.form.get("entity_type","company").strip()

    if not company_name:
        return jsonify(error="Company name is required"), 400
    if not pan or len(pan) != 10:
        return jsonify(error="PAN must be 10 characters (e.g. ABCDE1234F)"), 400

    job_id  = str(uuid.uuid4())[:8]
    job_dir = UPLOAD_DIR / job_id
    out_dir = OUTPUT_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    out_dir.mkdir(parents=True, exist_ok=True)

    saved = {k: [] for k in ("it26as", "itais", "ittis", "itgst")}
    for zone in saved:
        for fobj in request.files.getlist(f"files_{zone}"):
            if not fobj.filename: continue
            from werkzeug.utils import secure_filename
            safe = secure_filename(fobj.filename) or f"upload_{zone}_{uuid.uuid4().hex[:6]}"
            if Path(safe).suffix.lower() not in ALLOWED_EXT: continue
            dest = job_dir / safe
            fobj.save(str(dest))
            saved[zone].append(str(dest))

    with jobs_lock:
        jobs[job_id] = {
            "status":       "queued",
            "progress":     0,
            "logs":         [],
            "files":        [],
            "error":        None,
            "company_name": company_name,
            "pan":          pan,
            "gstin":        gstin,
            "fy":           fy,
            "itr_form":     itr_form,
            "entity_type":  entity_type,
            "job_dir":      str(job_dir),
            "out_dir":      str(out_dir),
            "saved":        saved,
        }

    threading.Thread(target=run_it_reconciliation, args=(job_id,), daemon=True).start()
    return jsonify(job_id=job_id)


@app.route("/api/it-job/<job_id>")
@rate_limit(limit=120, window=60)
def api_it_job(job_id):
    with jobs_lock:
        job = jobs.get(job_id)
        if not job:
            return jsonify(error="Job not found"), 404
        new_logs = job["logs"][:]
        job["logs"] = []
        return jsonify(
            status   = job["status"],
            progress = job["progress"],
            logs     = new_logs,
            files    = job["files"],
            error    = job["error"],
        )


@app.route("/api/it-dl/<job_id>/<filename>")
@rate_limit(limit=30, window=60)
def api_it_dl(job_id, filename):
    filename = Path(filename).name
    if not re.match(r'^[\w\-. ()]+\.(xlsx|pdf|zip)$', filename):
        abort(400)
    fp = OUTPUT_DIR / job_id / filename
    if not fp.exists() or not fp.is_file():
        abort(404)
    return send_file(str(fp), as_attachment=True, download_name=filename)



# ═══════════════════════════════════════════════════════════════════
# INCOME TAX PORTAL AUTO DOWNLOAD — Selenium-based
# Downloads 26AS, AIS, TIS automatically from incometax.gov.in
# ═══════════════════════════════════════════════════════════════════

@app.route("/api/it-auto-download", methods=["POST"])
@rate_limit(limit=5, window=60)
def api_it_auto_download():
    d = request.get_json(silent=True) or {}
    pan          = d.get("pan","").strip().upper()
    company_name = d.get("company_name","").strip()
    username     = d.get("username","").strip()
    password     = d.get("password","")
    fy           = d.get("fy","2025-26")

    if not pan or len(pan) != 10:
        return jsonify(error="PAN must be 10 characters"), 400
    if not company_name:
        return jsonify(error="Company name required"), 400
    if not username:
        return jsonify(error="IT Portal username (PAN) required"), 400

    job_id  = str(uuid.uuid4())[:8]
    out_dir = OUTPUT_DIR / job_id
    out_dir.mkdir(parents=True, exist_ok=True)

    with jobs_lock:
        jobs[job_id] = {
            "status": "running", "progress": 0,
            "logs": [{"type":"info","msg":"Starting IT Portal auto-download..."}],
            "files": [], "error": None,
            "captcha_needed": False, "captcha_img": None,
            "out_dir": str(out_dir),
            "failure_screenshots": [],
        }

    sess = {"captcha_q": _queue.Queue(), "refresh_event": threading.Event(), "screenshot": None}
    with _sess_lock:
        _sessions[job_id] = sess

    def run_bg():
        try:
            _it_auto_download(job_id, pan, company_name, username, password, fy, sess)
        except Exception as _exc:
            import traceback as _tb
            with jobs_lock:
                if job_id in jobs:
                    jobs[job_id]["status"] = "error"
                    jobs[job_id]["error"]  = str(_exc)
                    for _l in _tb.format_exc().split("\n"):
                        if _l.strip():
                            jobs[job_id]["logs"].append({"type":"err","msg":f"  {_l}"})
        finally:
            with _sess_lock:
                _sessions.pop(job_id, None)

    threading.Thread(target=run_bg, daemon=True).start()
    return jsonify(job_id=job_id)


def _it_auto_download(job_id, pan, company_name, username, password, fy, sess):
    """
    Server-side IT Portal auto-downloader.
    Mirrors it_suite.py exactly:
      1. incometax.gov.in → Login button → PAN → Continue → Password → Login/Submit
      2. OTP handling (show screenshot → user types OTP)
      3. Remember Device → auto-click YES
      4. e-File → Income Tax Returns → View AIS → TIS download → AIS download
      5. e-File → Income Tax Returns → View Form 26AS → TRACES → Export as PDF
      6. ZIP all files
    """
    import base64, shutil as _shutil

    # ── Assessment year from FY ──────────────────────────────────────────
    fy_start = int(fy.split("-")[0])
    AY_LABEL  = f"{fy_start+1}-{str(fy_start+2)[-2:]}"   # e.g. "2025-26"
    IT_PORTAL = "https://www.incometax.gov.in/iec/foportal"
    PAGE_WAIT, SHORT_WAIT = 8, 3

    def log(msg, t="info"):
        print(f"[IT {job_id}] {msg}")
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["logs"].append({"type": t, "msg": msg})

    def prog(p):
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["progress"] = p

    def show_captcha(img_b64):
        sess["screenshot"] = img_b64
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["captcha_needed"] = True
                jobs[job_id]["captcha_img"]    = img_b64

    def clear_captcha():
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["captcha_needed"] = False
                jobs[job_id]["captcha_img"]    = None

    def wait_user_input(prompt_msg):
        while not sess["captcha_q"].empty():
            try: sess["captcha_q"].get_nowait()
            except: pass
        log(f"⏳ {prompt_msg}")
        try:
            return sess["captcha_q"].get(timeout=600)
        except _queue.Empty:
            raise RuntimeError("Timeout — no input received in 10 minutes")

    def save_fail_shot(label):
        try:
            img_b64 = base64.b64encode(driver.get_screenshot_as_png()).decode()
            with jobs_lock:
                if job_id in jobs:
                    jobs[job_id].setdefault("failure_screenshots", []).append({
                        "label": label, "img_b64": img_b64,
                        "ts": datetime.now().strftime("%H:%M:%S"),
                    })
            log(f"  📸 Screenshot saved: {label}", "warn")
        except: pass

    def _sshot_b64():
        try: return base64.b64encode(driver.get_screenshot_as_png()).decode()
        except: return None

    out_dir = Path(jobs[job_id]["out_dir"])
    dl_dir  = out_dir / "it_downloads"
    dl_dir.mkdir(parents=True, exist_ok=True)
    downloaded = []
    driver = None

    # ── Selenium setup ────────────────────────────────────────────────────
    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait, Select
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.chrome.options import Options as ChromeOptions
        from selenium.webdriver.chrome.service import Service as ChromeService
    except ImportError:
        raise RuntimeError("Selenium not installed. Run: pip install selenium")

    # ── Helper: try_click ─────────────────────────────────────────────────
    def _click(xpaths, timeout=8):
        for xp in xpaths:
            try:
                el = WebDriverWait(driver, timeout).until(
                    EC.element_to_be_clickable((By.XPATH, xp)))
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                time.sleep(0.3)
                try: el.click()
                except: driver.execute_script("arguments[0].click();", el)
                return True
            except: continue
        return False

    # ── Helper: human_type ────────────────────────────────────────────────
    def _type(by, val, text, timeout=10):
        try:
            el = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, val)))
            driver.execute_script("arguments[0].scrollIntoView(true);", el)
            time.sleep(0.3); el.click(); time.sleep(0.2)
            el.clear(); time.sleep(0.2)
            for ch in str(text): el.send_keys(ch); time.sleep(0.04)
            time.sleep(0.3)
            # JS fallback if field is still empty
            if not (el.get_attribute("value") or "").strip():
                driver.execute_script(
                    "arguments[0].value=arguments[1];"
                    "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));"
                    "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
                    el, text)
            return True
        except: return False

    # ── Helper: _dismiss_popup ────────────────────────────────────────
    def _dismiss_popup():
        """
        Dismiss the IT portal 'Reload? You will be logged out.' popup.
        Handles:  1) native browser confirm/alert dialog
                  2) Angular Material dialog with No / Stay / Keep me signed in button
        Safe to call from background thread. Returns True if a popup was dismissed.
        """
        try:
            # Form 1: native browser dialog (window.confirm not yet suppressed)
            try:
                alert = driver.switch_to.alert
                log(f"  ⚠ Native dialog: '{alert.text[:80]}' — dismissing (No)...", "warn")
                alert.dismiss()
                time.sleep(0.5)
                return True
            except Exception:
                pass

            # Form 2 & 3: Angular / HTML modal 'No' / 'Stay' button
            no_xpaths = [
                "//button[normalize-space()='No']",
                "//button[normalize-space()='NO']",
                "//button[contains(normalize-space(),'No')]",
                "//button[normalize-space()='Stay']",
                "//button[normalize-space()='STAY']",
                "//button[contains(normalize-space(),'Stay')]",
                "//button[contains(normalize-space(),'Keep me')]",
                "//button[contains(normalize-space(),'Remain')]",
                "//*[@role='dialog']//button[contains(normalize-space(),'No')]",
                "//*[@role='dialog']//button[contains(normalize-space(),'Stay')]",
                "//mat-dialog-container//button[contains(normalize-space(),'No')]",
                "//mat-dialog-container//button[contains(normalize-space(),'Stay')]",
                "//*[contains(@class,'modal')]//button[contains(normalize-space(),'No')]",
                "//*[contains(@class,'dialog')]//button[contains(normalize-space(),'No')]",
                "//button[contains(normalize-space(),'Cancel')]",
            ]
            clicked = _click(no_xpaths, timeout=2)
            if clicked:
                log("  IT portal popup dismissed ✓")
                time.sleep(0.8)
            return clicked
        except Exception:
            return False

    # ── SessionWatchdog: background thread polls every 4s ─────────────
    import threading as _threading

    class _SessionWatchdog:
        """
        Continuously dismisses the IT portal's reload/logout popup in the
        background so it can never freeze the browser or expire the session,
        regardless of when it fires (during downloads, OTP wait, navigation).
        """
        POLL = 4   # seconds between checks

        def __init__(self):
            self._stop = _threading.Event()
            self._t    = _threading.Thread(target=self._run, daemon=True,
                                           name="AppSessionWatchdog")

        def start(self):
            self._stop.clear()
            self._t.start()
            log("  SessionWatchdog started ✓")

        def stop(self):
            self._stop.set()
            self._t.join(timeout=self.POLL + 2)

        def _run(self):
            while not self._stop.wait(self.POLL):
                try:
                    _dismiss_popup()
                except Exception:
                    pass  # browser mid-navigation; retry next cycle

    watchdog = _SessionWatchdog()

    # ── Helper: wait_for_new_file ─────────────────────────────────────────
    def _wait_new_file(extensions, before_set, timeout=120):
        deadline = time.time() + timeout
        while time.time() < deadline:
            time.sleep(2)
            for f in dl_dir.iterdir():
                if (f not in before_set
                        and f.suffix.lower() in extensions
                        and not f.name.endswith((".crdownload",".tmp"))
                        and f.stat().st_size > 5000):
                    time.sleep(1)  # let download finish
                    return f
        return None

    # ── Browser launch ────────────────────────────────────────────────────
    try:
        opts = ChromeOptions()
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--window-size=1280,900")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        # Speed flags for Render/Linux server only — these crash Chrome on Windows
        if platform.system() != "Windows":
            opts.add_argument("--no-zygote")
            opts.add_argument("--single-process")
        opts.add_argument("--disable-setuid-sandbox")
        opts.add_argument("--disable-software-rasterizer")
        opts.add_argument("--disable-background-networking")
        opts.add_argument("--disable-default-apps")
        opts.add_argument("--disable-sync")
        opts.add_argument("--metrics-recording-only")
        opts.add_argument("--mute-audio")
        opts.add_argument("--no-first-run")
        opts.add_argument("--safebrowsing-disable-auto-update")
        opts.add_argument("--disable-background-timer-throttling")
        opts.add_argument("--disable-renderer-backgrounding")
        opts.add_argument("--disable-backgrounding-occluded-windows")
        opts.add_argument("--memory-pressure-off")
        opts.add_argument(
            "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        opts.add_experimental_option("prefs", {
            "download.default_directory":       str(dl_dir),
            "download.prompt_for_download":     False,
            "download.directory_upgrade":       True,
            "plugins.always_open_pdf_externally": True,
            "safebrowsing.enabled":             True,
            "credentials_enable_service":       False,
            "profile.password_manager_enabled": False,
        })
        opts.add_experimental_option("excludeSwitches", ["enable-automation","enable-logging"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.page_load_strategy = "eager"

        try:
            import shutil as _sh2
            _IS_SERVER2 = bool(os.environ.get("RENDER") or os.environ.get("HEADLESS"))
            _cb2 = (os.environ.get("CHROME_BIN")
                    or _sh2.which("chromium") or _sh2.which("chromium-browser")
                    or _sh2.which("google-chrome"))
            _cd2 = (os.environ.get("CHROMEDRIVER_PATH")
                    or _sh2.which("chromedriver"))
            if _IS_SERVER2 and _cb2:
                opts.binary_location = _cb2
                svc2 = ChromeService(executable_path=_cd2) if _cd2 else ChromeService()
                driver = webdriver.Chrome(service=svc2, options=opts)
            else:
                from webdriver_manager.chrome import ChromeDriverManager
                driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=opts)
        except:
            driver = webdriver.Chrome(options=opts)

        _STEALTH_JS = """
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'languages', {get: () => ['en-IN', 'en']});
            Object.defineProperty(navigator, 'plugins',   {get: () => [1, 2, 3, 4, 5]});
            window.chrome = { runtime: {}, loadTimes: function(){}, csi: function(){}, app: {} };
            // ── Suppress IT portal reload/logout dialogs ──────────────────────
            // The IT portal fires window.confirm('Reload? You will be logged out.')
            // on every Angular route change. Auto-clicking 'No' (return false)
            // prevents the browser from freezing and prevents session expiry.
            window.confirm      = function(msg){ console.log('[CONFIRM suppressed] ' + msg); return false; };
            window.alert        = function(msg){ console.log('[ALERT suppressed]   ' + msg); };
            window.onbeforeunload = null;
            // Re-apply on every future hash navigation (Angular SPA guard)
            window.addEventListener('hashchange', function(){
                window.confirm = function(msg){ return false; };
                window.alert   = function(msg){ };
                window.onbeforeunload = null;
            });
        """
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": _STEALTH_JS})
        # Allow multiple file downloads without per-file permission prompts
        driver.execute_cdp_cmd("Browser.setDownloadBehavior", {
            "behavior": "allow",
            "downloadPath": str(dl_dir),
            "eventsEnabled": True,
        })
        log("✅ Headless Chrome started for IT Portal", "ok")
        prog(5)

        # ════════════════════════════════════════════════════════════════
        # STEP 1: LOGIN  (mirrors it_login() from it_suite.py exactly)
        # ════════════════════════════════════════════════════════════════
        log("🌐 Opening incometax.gov.in ...")
        driver.get(IT_PORTAL)
        time.sleep(PAGE_WAIT)

        log("  Clicking Login button on portal home...")
        _click([
            "//a[normalize-space()='Login']",
            "//button[normalize-space()='Login']",
            "//a[contains(@href,'login')]",
            "//span[normalize-space()='Login']",
        ])
        time.sleep(PAGE_WAIT)
        log(f"  Login page URL: {driver.current_url}")

        log(f"  Entering PAN/User ID: {username}")

        # ── Wait for Angular app to render input fields ────────────────────
        for _ in range(20):
            inputs = driver.find_elements(By.CSS_SELECTOR, "input:not([type='hidden'])")
            if inputs:
                log(f"  Login form ready — {len(inputs)} input(s) visible")
                break
            time.sleep(2)
        time.sleep(1)
        log(f"  Login page URL after wait: {driver.current_url}")

        def _type_mat(by, val, text, timeout=8):
            """
            Angular-Material-compatible type function.
            Fires input/change/keyup events so Angular picks up the value.
            Falls back to plain send_keys if CDP insertText fails.
            """
            try:
                el = WebDriverWait(driver, timeout).until(
                    EC.presence_of_element_located((by, val)))
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center',inline:'center'});", el)
                time.sleep(0.4)
                driver.execute_script(
                    "arguments[0].click(); arguments[0].focus();", el)
                time.sleep(0.3)
                # Clear existing value
                driver.execute_script(
                    "arguments[0].value = '';"
                    "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", el)
                time.sleep(0.2)
                # Insert text via CDP (works for Angular Material)
                try:
                    driver.execute_cdp_cmd("Input.insertText", {"text": str(text)})
                except Exception:
                    for ch in str(text): el.send_keys(ch); time.sleep(0.04)
                time.sleep(0.3)
                # Fire Angular events
                driver.execute_script(
                    "arguments[0].dispatchEvent(new Event('input',  {bubbles:true}));"
                    "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));"
                    "arguments[0].dispatchEvent(new KeyboardEvent('keyup',{bubbles:true}));",
                    el)
                time.sleep(0.2)
                actual = (el.get_attribute("value") or "").strip().upper()
                expected = str(text).strip().upper()
                if actual == expected or len(actual) > 0:
                    log(f"    ✓ Typed via _type_mat: '{actual}'  [{val}]")
                    return True
                # Last resort: plain send_keys
                el.clear()
                for ch in str(text): el.send_keys(ch); time.sleep(0.04)
                time.sleep(0.2)
                actual = (el.get_attribute("value") or "").strip()
                return bool(actual)
            except Exception as e:
                log(f"    _type_mat fail [{val}]: {e}", "warn")
                return False

        filled = False
        for by, val in [
            (By.ID,           "panAdhaarUserId"),                              # IT portal real field ID (April 2026)
            (By.NAME,         "panAdhaarUserId"),
            (By.CSS_SELECTOR, "#panAdhaarUserId"),
            (By.CSS_SELECTOR, "input[name='panAdhaarUserId']"),
            (By.CSS_SELECTOR, "input[formcontrolname='userId']"),
            (By.CSS_SELECTOR, "input[placeholder*='PAN']"),
            (By.CSS_SELECTOR, "input[placeholder*='Aadhaar']"),
            (By.CSS_SELECTOR, "input[placeholder*='User ID']"),
            (By.CSS_SELECTOR, "input.mat-mdc-input-element:not([type='password'])"),
            (By.CSS_SELECTOR, "input[type='text']:not([readonly]):not([disabled])"),
        ]:
            if _type_mat(by, val, username):
                filled = True
                log(f"  PAN typed via selector: {val}")
                break

        if not filled:
            # Wait 15s more for slow portal and retry primary selectors
            log("  PAN field not found yet — waiting 15s for slow portal load...")
            time.sleep(15)
            filled = (
                _type_mat(By.ID, "panAdhaarUserId", username) or
                _type_mat(By.CSS_SELECTOR, "input[formcontrolname='userId']", username) or
                _type_mat(By.CSS_SELECTOR, "input[type='text']:not([readonly])", username)
            )

        if not filled:
            raise RuntimeError("Cannot find User ID / PAN field on IT portal login page")

        time.sleep(1)
        # Some portals have a 2-step: PAN first → Continue → Password
        _click([
            "//button[normalize-space()='Continue']",
            "//input[@value='Continue']",
            "//button[contains(text(),'Continue')]",
        ])
        time.sleep(SHORT_WAIT)

        log("  Entering password...")
        filled = False
        for by, val in [
            (By.CSS_SELECTOR, "input[type='password']"),
            (By.CSS_SELECTOR, "input[formcontrolname='password']"),
            (By.CSS_SELECTOR, "input[formcontrolname='currentPassword']"),
            (By.ID,   "password"),
            (By.NAME, "password"),
            (By.ID,   "current-password"),
            (By.CSS_SELECTOR, "input[placeholder*='assword']"),
        ]:
            if _type_mat(by, val, password):
                filled = True; break
        if not filled:
            raise RuntimeError("Cannot find password field on IT portal login page")

        time.sleep(1)

        # ── Tick "Secure Access Message" checkbox (IT portal password page requires this) ──
        log("  Ticking 'Secure Access Message' checkbox (if present)...")
        checkbox_clicked = False
        for xp in [
            "//input[@type='checkbox']",
            "//input[contains(@id,'secure')]",
            "//input[contains(@name,'secure')]",
            "//input[contains(@id,'confirm')]",
            "//*[contains(text(),'secure access')]/preceding-sibling::input[@type='checkbox']",
            "//*[contains(text(),'secure access')]/..//input[@type='checkbox']",
            "//*[contains(@class,'checkbox')]//input[@type='checkbox']",
        ]:
            try:
                cb = WebDriverWait(driver, 4).until(
                    EC.presence_of_element_located((By.XPATH, xp)))
                if not cb.is_selected():
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'});", cb)
                    time.sleep(0.3)
                    try:   cb.click()
                    except: driver.execute_script("arguments[0].click();", cb)
                log(f"  Checkbox ticked ✓")
                checkbox_clicked = True
                break
            except: continue
        if not checkbox_clicked:
            try:
                driver.execute_script("""
                    var cbs = document.querySelectorAll('input[type="checkbox"]');
                    cbs.forEach(function(cb){
                        if(!cb.checked){
                            cb.click();
                            cb.dispatchEvent(new Event('change',{bubbles:true}));
                        }
                    });
                """)
                checkbox_clicked = True
                log("  Checkbox ticked via JS bulk-click ✓")
            except: pass
        if not checkbox_clicked:
            log("  Checkbox not found — portal may still accept login", "warn")
        time.sleep(1)

        log("  Clicking Login submit...")
        _click([
            "//button[@type='submit']",
            "//button[normalize-space()='Login']",
            "//button[normalize-space()='Sign In']",
            "//input[@type='submit']",
            "//button[contains(text(),'Login')]",
            "//button[contains(text(),'Sign in')]",
        ])
        time.sleep(PAGE_WAIT + 2)

        # ── OTP handling ──────────────────────────────────────────────────
        try:
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            otp_present = "otp" in body and ("enter" in body or "verify" in body or "sent" in body)
            if otp_present:
                log("📱 OTP screen detected — showing screenshot for input...")
                img = _sshot_b64()
                show_captcha(img)
                otp_text = wait_user_input(
                    "OTP was sent to your registered mobile/email. "
                    "Enter the OTP below and click Submit. "
                    "Type SKIP if no OTP is required."
                )
                clear_captcha()

                if otp_text.upper() != "SKIP":
                    log(f"  Entering OTP...")
                    otp_filled = False
                    for by, val in [
                        (By.ID,   "otp"),
                        (By.NAME, "otp"),
                        (By.CSS_SELECTOR, "input[placeholder*='OTP']"),
                        (By.CSS_SELECTOR, "input[placeholder*='otp']"),
                        (By.CSS_SELECTOR, "input[type='tel']"),
                        (By.CSS_SELECTOR, "input[maxlength='6']"),
                    ]:
                        if _type(by, val, otp_text):
                            otp_filled = True; break

                    log("  Clicking Validate/Submit for OTP...")
                    _click([
                        "//button[contains(text(),'Validate')]",
                        "//button[normalize-space()='Submit']",
                        "//button[normalize-space()='Verify']",
                        "//button[@type='submit']",
                    ])
                    time.sleep(PAGE_WAIT)
                    log("  OTP submitted ✓", "ok")
        except Exception as e:
            log(f"  OTP check skipped: {e}", "warn")

        # ── Remember Device handling (auto-click YES for permanent unlock) ─
        time.sleep(SHORT_WAIT)
        try:
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            remember_present = any(w in body for w in [
                "remember", "register this device", "trust this device",
                "don't ask again", "secure access", "add device"
            ])
            if remember_present:
                log("  'Remember Device' prompt detected — clicking YES...")
                _click([
                    "//button[normalize-space()='Yes']",
                    "//button[normalize-space()='YES']",
                    "//button[contains(text(),'Yes')]",
                    "//button[contains(text(),'Register')]",
                    "//button[contains(text(),'Trust')]",
                    "//a[normalize-space()='Yes']",
                    "//input[@value='Yes']",
                ])
                # Wait for the portal to complete device registration and reach dashboard.
                # The Angular app transitions: Remember-Device → #/login/password → #/dashboard
                # This can take up to 30s on a slow connection — DO NOT navigate away early!
                log("  Waiting for portal to complete device registration (up to 60s)...")
                for _rd_wait in range(30):
                    time.sleep(2)
                    _rd_url = driver.current_url.lower()
                    if "dashboard" in _rd_url and "sessionexpire" not in _rd_url:
                        log("  ✅ Device registered — dashboard reached!", "ok")
                        break
                    if _rd_wait % 5 == 0:
                        log(f"  Still waiting for dashboard ({_rd_wait*2}s)… URL: {driver.current_url}")
                else:
                    log("  ✅ Device registered — future logins will skip OTP", "ok")
        except: pass

        # ── Verify login success ───────────────────────────────────────────
        cur = driver.current_url.lower()
        log(f"  Post-login URL: {driver.current_url}")

        # CRITICAL: If still on login/password page, this is the Angular mid-flow state
        # (#/login/password is NORMAL while the portal is setting JWT cookies after OTP).
        # Wait up to 60s watching for #/dashboard.  DO NOT navigate away — that kills the
        # JWT handshake and permanently breaks the session (causes the infinite redirect loop).
        # Only treat it as a real failure if we land on #/sessionExpire (not #/login/password).
        if "login" in cur or "password" in cur:
            log("  Still on login/password page — waiting for Angular to reach dashboard (up to 60s)...")
            for _lw in range(30):
                time.sleep(2)
                cur = driver.current_url.lower()
                if "dashboard" in cur and "sessionexpire" not in cur:
                    log(f"  ✅ Portal reached dashboard: {driver.current_url}", "ok")
                    break
                if "sessionexpire" in cur:
                    log("  ⚠ Session expired URL during post-login wait — will recover later", "warn")
                    break
                if _lw % 5 == 0 and _lw > 0:
                    log(f"  Still waiting… ({_lw*2}s) URL: {driver.current_url}")
            else:
                log("  Portal still on login after 60s — showing screenshot for manual input", "warn")
            cur = driver.current_url.lower()

        login_ok = any([
            "dashboard" in cur and "sessionexpire" not in cur,
            "profile"   in cur,
            "myaccount" in cur,
            ("home" in cur and "foportal" in cur),
            ("iec/foportal" in cur and "login" not in cur and "sessionexpire" not in cur),
        ])
        if not login_ok:
            # Show screenshot and let user verify / continue
            img = _sshot_b64()
            show_captcha(img)
            resp = wait_user_input(
                "Login may need an extra step. Check the screenshot — "
                "enter any additional OTP/CAPTCHA, or type SKIP to continue."
            )
            clear_captcha()
            if resp.upper() != "SKIP":
                for by, val in [
                    (By.CSS_SELECTOR, "input[type='tel']"),
                    (By.CSS_SELECTOR, "input[type='text']:not([readonly])"),
                    (By.ID, "otp"),
                ]:
                    if _type(by, val, resp): break
                _click(["//button[@type='submit']", "//button[contains(text(),'Continu')]"])
                time.sleep(PAGE_WAIT + 5)
                # After user action: wait again for dashboard
                for _uw in range(20):
                    time.sleep(2)
                    cur = driver.current_url.lower()
                    if "dashboard" in cur and "sessionexpire" not in cur:
                        log(f"  ✅ Dashboard after user action: {driver.current_url}", "ok")
                        break
            else:
                # User typed SKIP — wait for the portal to settle naturally
                time.sleep(SHORT_WAIT + 5)

        # ── Navigate to dashboard and confirm session is live ─────────────
        # IMPORTANT: After OTP + Remember-Device, the portal's Angular app may still
        # be at #/login/password while it writes JWT cookies to sessionStorage.
        # We must WAIT (not navigate) until it reaches #/dashboard on its own.
        # Navigating away prematurely kills the cookie exchange → infinite redirect.
        log("  Navigating to dashboard to establish session...")

        def _is_session_dead(url):
            """Returns True only if portal is stuck at a genuine dead-session URL."""
            u = url.lower()
            return "sessionexpire" in u

        def _is_mid_login(url):
            """Returns True if portal is in the normal Angular mid-flow state (not a dead session)."""
            u = url.lower()
            return "password" in u or ("/login" in u and "sessionexpire" not in u)

        def _recover_session():
            """
            Full session recovery when #/sessionExpire is detected.
            Strategy:
              1. Clear stale Angular session flags from storage.
              2. Navigate to base portal URL (no hash) to let the server re-issue cookies.
              3. Wait up to 30s for Angular to auto-auth from existing cookie.
              4. Force navigate to dashboard.
            Returns True if dashboard is reached.
            """
            log("  🔄 Session expired URL detected — attempting recovery...", "warn")

            # Step 0: Clear any Angular-set 'sessionExpired' flags before navigating
            try:
                driver.execute_script("""
                    try { sessionStorage.removeItem('sessionExpired'); } catch(e){}
                    try { localStorage.removeItem('sessionExpired');   } catch(e){}
                    window.dispatchEvent(new Event('focus'));
                    document.dispatchEvent(new Event('visibilitychange'));
                """)
                time.sleep(1)
            except Exception:
                pass

            # Step 1: Go to the base portal URL (no hash) — forces full cookie refresh
            driver.get("https://eportal.incometax.gov.in/iec/foservices/")
            time.sleep(PAGE_WAIT + 2)
            _u = driver.current_url.lower()
            log(f"  After base-URL nav: {driver.current_url}")

            # Step 2: Wait up to 20s for Angular to auto-auth and reach dashboard
            if "login" in _u or "sessionexpire" in _u:
                log("  Portal shows login/expire after base nav — waiting for auto re-auth...", "warn")
                for _ar in range(10):
                    time.sleep(2)
                    _u = driver.current_url.lower()
                    if "dashboard" in _u and "sessionexpire" not in _u:
                        log(f"  ✅ Auto re-auth succeeded: {driver.current_url}", "ok")
                        return True
                    if _ar == 5:
                        # Nudge the Angular app: clear flags again mid-wait
                        try:
                            driver.execute_script("""
                                try { sessionStorage.removeItem('sessionExpired'); } catch(e){}
                                try { localStorage.removeItem('sessionExpired');   } catch(e){}
                                window.dispatchEvent(new Event('focus'));
                            """)
                        except Exception:
                            pass

            # Step 3: Force navigate to dashboard via JS replace (bypasses Angular guard hash-check)
            log("  Forcing dashboard via JS location.replace...", "warn")
            try:
                driver.execute_script(
                    "window.location.replace('https://eportal.incometax.gov.in/iec/foservices/#/dashboard');"
                )
                time.sleep(PAGE_WAIT + 3)
            except Exception:
                driver.get("https://eportal.incometax.gov.in/iec/foservices/#/dashboard")
                time.sleep(PAGE_WAIT + 3)

            _uf = driver.current_url.lower()
            if "dashboard" in _uf and "sessionexpire" not in _uf:
                log(f"  ✅ Session recovered — dashboard: {driver.current_url}", "ok")
                return True

            log(f"  ⚠ Recovery failed — current URL: {driver.current_url}", "warn")
            return False

        # ── Dashboard wait loop ───────────────────────────────────────────
        # Watch for dashboard arrival. Only act on #/sessionExpire (true dead session).
        # #/login/password is a NORMAL mid-flow state — do NOT navigate away from it.
        _dashboard_ready = False
        _session_recovery_done = False

        for _wait_i in range(40):          # up to 80 seconds
            _cur2 = driver.current_url.lower()

            if "dashboard" in _cur2 and not _is_session_dead(_cur2):
                _dashboard_ready = True
                log(f"  ✅ Dashboard loaded: {driver.current_url}", "ok")
                break

            if _is_session_dead(_cur2):
                # True dead session → try recovery (only once)
                if not _session_recovery_done:
                    _session_recovery_done = True
                    log(f"  Session redirected ({_wait_i+1}/40) — running recovery...", "warn")
                    if _recover_session():
                        _dashboard_ready = True
                        break
                else:
                    # Second time hitting sessionExpire — recovery already failed; just try nav
                    log(f"  Session redirected ({_wait_i+1}/40) — retrying nav...", "warn")
                    driver.get("https://eportal.incometax.gov.in/iec/foservices/#/dashboard")

            elif _is_mid_login(_cur2):
                # Normal Angular mid-flow (#/login/password) — just wait, DO NOT navigate
                if _wait_i % 5 == 0:
                    log(f"  Still on login page ({_wait_i+1}/40) — waiting for portal…")

            else:
                # Some intermediate URL (e.g. /#/) — navigate to dashboard gently
                if _wait_i % 5 == 0:
                    log(f"  Intermediate URL ({_wait_i+1}/40): {driver.current_url}")
                    driver.get("https://eportal.incometax.gov.in/iec/foservices/#/dashboard")

            time.sleep(2)

        if not _dashboard_ready:
            log("  ⚠ Dashboard did not load after 80s — attempting final forced recovery...", "warn")
            if not _session_recovery_done:
                if _recover_session():
                    _dashboard_ready = True
            if not _dashboard_ready:
                # Last resort: try clearing storage + full page reload
                try:
                    driver.execute_script("""
                        try { sessionStorage.removeItem('sessionExpired'); } catch(e){}
                        try { localStorage.removeItem('sessionExpired');   } catch(e){}
                    """)
                    driver.get("https://eportal.incometax.gov.in/iec/foservices/#/dashboard")
                    time.sleep(PAGE_WAIT + 5)
                    _cur_final = driver.current_url.lower()
                    if "dashboard" in _cur_final and "sessionexpire" not in _cur_final:
                        log(f"  ✅ Dashboard via final attempt: {driver.current_url}", "ok")
                        _dashboard_ready = True
                except Exception:
                    pass
            if not _dashboard_ready:
                log(f"  ⚠ Still not on dashboard — URL: {driver.current_url}", "warn")
                log("  ⚠ Proceeding anyway — AIS/TIS/26AS may still work if auth cookie is valid", "warn")

        if _dashboard_ready:
            # Extra wait for Angular to fully render widgets and auth tokens after dashboard load
            time.sleep(SHORT_WAIT + 2)

        log("  ✅ Login successful!", "ok")
        prog(20)

        # ── Start background popup watchdog ───────────────────────────
        # Session is now established. The watchdog polls every 4s and
        # dismisses any reload/logout popup the moment it appears —
        # preventing session expiry during downloads, waits, navigation.
        watchdog.start()

        # ════════════════════════════════════════════════════════════════
        # STEP 2: DOWNLOAD AIS + TIS
        # FIX: The AIS link is directly on the dashboard navbar — NOT
        # under e-File menu. Clicking it opens insight.gov.in in a NEW TAB.
        # We must switch to that tab and use the modal popup approach.
        # Direct URL (#/dashboard/ais-tis) kills Angular session online.
        # ════════════════════════════════════════════════════════════════
        log("\n📑 Downloading AIS + TIS (Annual Information Statement / Taxpayer Info Summary)...")

        def _navigate_to_ais_tab():
            """
            Click the 'AIS' link directly on the IT portal dashboard navbar.
            This opens insight.gov.in in a NEW TAB — we switch to it and wait
            for the 'Download AIS/TIS' button to appear (JWT redirect takes ~10s).
            Returns the AIS tab handle, or None on failure.
            NEVER uses direct URL navigation — that kills Angular session.
            """
            # ── Pre-check: ensure we are actually on the dashboard ────────────
            _cur_before_ais = driver.current_url.lower()
            if "sessionexpire" in _cur_before_ais or ("dashboard" not in _cur_before_ais and "login" in _cur_before_ais):
                log("  Session expired before AIS nav — recovering...", "warn")
                if not _recover_session():
                    log("  ⚠ Session recovery failed for AIS — cannot navigate", "warn")
                    return None
                # Extra wait after recovery
                time.sleep(SHORT_WAIT + 2)

            original_handles = set(driver.window_handles)
            log("  Step 1: clicking 'AIS' link on dashboard navbar...")
            clicked = _click([
                "//a[normalize-space()='AIS']",
                "//li/a[normalize-space()='AIS']",
                "//nav//a[normalize-space()='AIS']",
                "//button[normalize-space()='AIS']",
                "//*[normalize-space()='AIS']",
                "//a[contains(normalize-space(),'Annual Information Statement')]",
                "//*[contains(@class,'ais')]//a",
            ])
            if not clicked:
                # JS full-page scan as last resort (no direct URL)
                try:
                    result = driver.execute_script("""
                        var tags = ['a','button','span','li','div'];
                        for(var tag of tags){
                            var els = document.querySelectorAll(tag);
                            for(var el of els){
                                var t = (el.innerText||el.textContent||'').trim();
                                if(t === 'AIS' && el.offsetParent !== null){
                                    el.scrollIntoView({block:'center'});
                                    el.click(); return 'clicked exact: AIS';
                                }
                            }
                        }
                        for(var tag of tags){
                            var els = document.querySelectorAll(tag);
                            for(var el of els){
                                var t = (el.innerText||el.textContent||'').trim();
                                if(t.includes('Annual Information Statement') && el.offsetParent !== null){
                                    el.scrollIntoView({block:'center'});
                                    el.click(); return 'clicked partial: ' + t;
                                }
                            }
                        }
                        return 'not found';
                    """)
                    log(f"  JS scan result: {result}")
                    clicked = result and "not found" not in str(result)
                except Exception as _je:
                    log(f"  JS AIS scan error: {_je}", "warn")

            if not clicked:
                log("  ⚠ AIS link not found on dashboard — cannot navigate to AIS", "warn")
                return None

            # Dismiss any popup that fires on AIS click
            time.sleep(2)
            _dismiss_popup()

            # Wait for the new AIS tab (insight.gov.in) to open
            log("  Waiting for AIS tab to open (insight.gov.in)...")
            ais_handle = None
            for _ in range(25):
                new_handles = set(driver.window_handles) - original_handles
                if new_handles:
                    ais_handle = list(new_handles)[0]
                    break
                _dismiss_popup()
                time.sleep(1)

            if ais_handle:
                driver.switch_to.window(ais_handle)
                log(f"  Switched to AIS tab: {driver.current_url}")
                # Fire focus/visibility events so the AIS Angular app initialises
                try:
                    driver.execute_script("""
                        Object.defineProperty(document,'visibilityState',{get:()=>'visible',configurable:true});
                        Object.defineProperty(document,'hidden',{get:()=>false,configurable:true});
                        document.dispatchEvent(new Event('visibilitychange'));
                        window.dispatchEvent(new Event('focus'));
                        window.dispatchEvent(new Event('pageshow'));
                    """)
                except Exception:
                    pass
            else:
                log(f"  AIS on same tab: {driver.current_url}")

            # Wait for JWT redirect to complete (up to 30s)
            log("  Waiting for AIS JWT redirect to complete...")
            for _jwt_i in range(15):
                _ju = (driver.current_url or "").lower()
                if "access?param" not in _ju and "insight.gov.in" in _ju:
                    log(f"  JWT complete — AIS dashboard: {driver.current_url}")
                    break
                time.sleep(2)

            # Poll for 'Download AIS/TIS' button (up to 90s)
            log("  Polling for 'Download AIS/TIS' button (up to 90s)...")
            _dl_btn_found = False
            _dl_deadline = time.time() + 90
            while time.time() < _dl_deadline:
                try:
                    _btns = driver.find_elements(By.XPATH,
                        "//*[contains(normalize-space(),'Download AIS') or "
                        "contains(normalize-space(),'Download AIS/TIS')]")
                    _visible = [b for b in _btns if b.is_displayed()]
                    if _visible:
                        log(f"  'Download AIS/TIS' button visible ✓ — page loaded")
                        _dl_btn_found = True
                        break
                except Exception:
                    pass
                time.sleep(3)
            if not _dl_btn_found:
                log("  ⚠ 'Download AIS/TIS' button did not appear within 90s", "warn")

            return ais_handle

        def _open_ais_download_popup():
            """Open the Download AIS/TIS popup and wait for 3 Download buttons."""
            log("  Opening 'Download AIS/TIS (F.Y.)' popup...")
            opened = _click([
                "//button[contains(normalize-space(),'Download AIS/TIS (F.Y.')]",
                "//a[contains(normalize-space(),'Download AIS/TIS (F.Y.')]",
                "//button[contains(normalize-space(),'Download AIS/TIS (FY')]",
                "//button[contains(normalize-space(),'Download AIS/TIS')]",
                "//a[contains(normalize-space(),'Download AIS/TIS')]",
                "//button[contains(normalize-space(),'Download AIS')]",
                "//*[contains(normalize-space(),'Download AIS/TIS')]",
            ])
            if not opened:
                log("  ⚠ 'Download AIS/TIS' button not found", "warn")
                return False
            # Wait for popup with exactly 3 Download buttons
            log("  Waiting for popup with 3 Download buttons...")
            for _at in range(25):
                try:
                    _eb = driver.find_elements(By.XPATH, "//button[normalize-space()='Download']")
                    _vb = [b for b in _eb if b.is_displayed()]
                    log(f"  Attempt {_at+1}: {len(_vb)} 'Download' button(s) visible")
                    if len(_vb) >= 3:
                        log("  Popup loaded — 3 Download buttons visible ✓")
                        time.sleep(0.3)
                        return True
                except Exception:
                    pass
                time.sleep(1)
            log("  ⚠ Popup did not show 3 buttons — proceeding anyway", "warn")
            return True

        def _click_popup_dl_btn(target):
            """
            Click Download button for 'ais' (index 0) or 'tis' (index 2) in the popup modal.
            Uses JS modal-scoped approach identical to it_suite.py.
            """
            log(f"  Modal-scoped click for {target.upper()} PDF Download button...")
            try:
                result = driver.execute_script("""
                    var target = arguments[0];
                    var modalSels = ['[role="dialog"]','.modal','.dialog','.popup',
                        '.cdk-overlay-pane','.cdk-dialog-container',
                        '[class*="modal"]','[class*="dialog"]','[class*="popup"]',
                        '[class*="overlay-container"]','[class*="download"]'];
                    var modal = null;
                    for(var sel of modalSels){
                        var els = document.querySelectorAll(sel);
                        for(var el of els){
                            var btns = el.querySelectorAll('button');
                            var dl = Array.from(btns).filter(function(b){
                                return (b.innerText||b.textContent||'').trim()==='Download'
                                       && b.offsetParent!==null;
                            });
                            if(dl.length>=3){ modal=el; break; }
                        }
                        if(modal) break;
                    }
                    if(!modal) return 'modal_not_found';
                    var rows = modal.querySelectorAll('tr,li,.row,[class*="row"],[class*="item"],[class*="list"]');
                    if(rows.length===0) rows=Array.from(modal.children);
                    var aisKw=['annual information','ais'];
                    var tisKw=['taxpayer information summary','tis'];
                    var skipKw=['json','utility'];
                    for(var row of rows){
                        var txt=(row.innerText||row.textContent||'').toLowerCase();
                        var btn=row.querySelector('button');
                        if(!btn||(btn.innerText||btn.textContent||'').trim()!=='Download') continue;
                        if(!btn.offsetParent) continue;
                        if(target==='ais'){
                            var isAis=aisKw.some(function(k){return txt.includes(k);});
                            var isJson=skipKw.some(function(k){return txt.includes(k);});
                            var isTis=tisKw.some(function(k){return txt.includes(k);});
                            if(isAis&&!isJson&&!isTis){btn.scrollIntoView({block:'center'});btn.click();return 'clicked_ais_by_text';}
                        } else {
                            var isTisRow=tisKw.some(function(k){return txt.includes(k);});
                            var isJsonRow=skipKw.some(function(k){return txt.includes(k);});
                            if(isTisRow&&!isJsonRow){btn.scrollIntoView({block:'center'});btn.click();return 'clicked_tis_by_text';}
                        }
                    }
                    var allDl=Array.from(modal.querySelectorAll('button')).filter(function(b){
                        return (b.innerText||b.textContent||'').trim()==='Download'&&b.offsetParent!==null;
                    });
                    var idx=(target==='ais')?0:2;
                    if(allDl.length>idx){allDl[idx].scrollIntoView({block:'center'});allDl[idx].click();
                        return 'clicked_by_index_'+idx+' (total='+allDl.length+')';}
                    return 'not_found (modal buttons='+allDl.length+')';
                """, target)
                log(f"  Modal click result: {result}")
                return result and "not_found" not in str(result) and "modal_not_found" not in str(result)
            except Exception as _e:
                log(f"  Modal click error: {_e}", "warn")
                return False

        def _is_popup_still_open():
            try:
                btns = driver.find_elements(By.XPATH, "//button[normalize-space()='Download']")
                return len([b for b in btns if b.is_displayed()]) >= 3
            except Exception:
                return False

        # ── Navigate to AIS (opens insight.gov.in in new tab) ─────────────
        _it_tab = driver.window_handles[0]  # remember IT portal tab
        _ais_tab_handle = _navigate_to_ais_tab()

        log(f"  AIS page URL: {driver.current_url}")
        img = _sshot_b64()
        show_captcha(img)
        time.sleep(2)
        clear_captcha()

        # ── Open the Download AIS/TIS popup ───────────────────────────────
        _popup_ok = _open_ais_download_popup()

        # ── Download AIS PDF (popup button index 0) ────────────────────────
        log("\n📊 Downloading AIS (Annual Information Statement)...")
        before_ais = set(dl_dir.iterdir())
        ais_clicked = _click_popup_dl_btn("ais")
        if ais_clicked:
            time.sleep(SHORT_WAIT)
            log("  Waiting for AIS PDF download...")
            new_f = _wait_new_file({".pdf"}, before_ais, timeout=120)
            if new_f:
                ais_name = f"AIS_{pan}_AY{AY_LABEL.replace('-','_')}.pdf"
                dest = out_dir / ais_name
                _shutil.copy2(str(new_f), str(dest))
                sz = dest.stat().st_size // 1024
                downloaded.append({"name": ais_name, "size": f"{sz} KB"})
                log(f"  ✅ AIS saved: {ais_name} ({sz} KB)", "ok")
                with jobs_lock:
                    if job_id in jobs: jobs[job_id]["files"] = list(downloaded)
            else:
                log("  ⚠ AIS PDF not downloaded within 120s", "warn")
                save_fail_shot("AIS — PDF not downloaded")
        else:
            log("  ⚠ AIS Download button not found in popup", "warn")
            save_fail_shot("AIS — Download button not found")

        before_after_ais = set(dl_dir.iterdir())
        time.sleep(SHORT_WAIT)

        # ── Download TIS PDF (popup button index 2) ────────────────────────
        log("\n📑 Downloading TIS (Taxpayer Information Summary)...")
        popup_still_open = _is_popup_still_open()
        log(f"  Popup still open after AIS? {popup_still_open}")
        if not popup_still_open:
            log("  Popup closed after AIS — reopening for TIS...")
            _open_ais_download_popup()

        before_tis = set(dl_dir.iterdir())
        tis_clicked = _click_popup_dl_btn("tis")
        if tis_clicked:
            time.sleep(SHORT_WAIT)
            log("  Waiting for TIS PDF download...")
            new_f = _wait_new_file({".pdf"}, before_after_ais, timeout=120)
            if new_f:
                tis_name = f"TIS_{pan}_AY{AY_LABEL.replace('-','_')}.pdf"
                dest = out_dir / tis_name
                _shutil.copy2(str(new_f), str(dest))
                sz = dest.stat().st_size // 1024
                downloaded.append({"name": tis_name, "size": f"{sz} KB"})
                log(f"  ✅ TIS saved: {tis_name} ({sz} KB)", "ok")
                with jobs_lock:
                    if job_id in jobs: jobs[job_id]["files"] = list(downloaded)
            else:
                log("  ⚠ TIS PDF not downloaded within 120s", "warn")
                save_fail_shot("TIS — PDF not downloaded")
        else:
            log("  ⚠ TIS Download button not found in popup", "warn")
            save_fail_shot("TIS — Download button not found")

        prog(60)

        # ════════════════════════════════════════════════════════════════
        # STEP 3: DOWNLOAD FORM 26AS
        # FIX: After AIS/TIS we are on insight.gov.in tab — must switch
        # back to IT portal tab first. Then use e-File menu (no direct URL).
        # TRACES opens in a new tab — we switch to it and handle download.
        # ════════════════════════════════════════════════════════════════
        log("\n📄 Downloading Form 26AS...")

        # ── Switch back to IT portal tab (close AIS insight.gov.in tab) ──
        log("  Switching back to IT portal tab...")
        try:
            for _h in list(driver.window_handles):
                try:
                    driver.switch_to.window(_h)
                    _hurl = (driver.current_url or "").lower()
                    if "eportal.incometax" in _hurl or "incometax.gov" in _hurl:
                        log(f"  Switched to IT portal tab: {driver.current_url}")
                        break
                except Exception:
                    continue
            else:
                # If none matched, close all extra tabs and use first handle
                _all_h = driver.window_handles
                driver.switch_to.window(_all_h[0])
                for _eh in _all_h[1:]:
                    try:
                        driver.switch_to.window(_eh)
                        driver.close()
                    except Exception:
                        pass
                driver.switch_to.window(_all_h[0])
                log(f"  Switched to first tab: {driver.current_url}")
        except Exception as _sw_err:
            log(f"  Tab switch warning: {_sw_err}", "warn")

        time.sleep(SHORT_WAIT)
        _dismiss_popup()

        # ── Verify IT portal session is still alive on dashboard ──────────
        _26as_pre_url = driver.current_url.lower()
        if _is_session_dead(_26as_pre_url) or "dashboard" not in _26as_pre_url:
            log("  ⚠ Not on dashboard — navigating back...", "warn")
            _26as_recovered = False
            if "sessionexpire" in _26as_pre_url:
                log("  Session expired before 26AS — running session recovery...", "warn")
                _26as_recovered = _recover_session()
            if not _26as_recovered:
                driver.get("https://eportal.incometax.gov.in/iec/foservices/#/dashboard")
                for _26_w in range(20):
                    time.sleep(2)
                    _26as_pre_url = driver.current_url.lower()
                    if "dashboard" in _26as_pre_url and not _is_session_dead(_26as_pre_url):
                        log("  ✅ Dashboard restored for 26AS", "ok")
                        _26as_recovered = True
                        break
                    if "sessionexpire" in _26as_pre_url and _26_w == 5:
                        if _recover_session():
                            _26as_recovered = True
                            break
            if not _26as_recovered:
                log("  ⚠ Could not restore dashboard — 26AS may fail", "warn")
        time.sleep(SHORT_WAIT)
        _dismiss_popup()

        # ── Navigate via e-File menu (NO direct URL fallback) ─────────────
        log("  Nav: e-File → Income Tax Returns → View Form 26AS")
        _before_26as_handles = set(driver.window_handles)

        # Step 1: Click e-File
        from selenium.webdriver.common.action_chains import ActionChains as _AC
        efile_clicked = _click([
            "//a[normalize-space()='e-File']",
            "//span[normalize-space()='e-File']",
            "//li[normalize-space()='e-File']",
            "//button[normalize-space()='e-File']",
            "//*[normalize-space()='e-File']",
            "//nav//*[contains(normalize-space(),'e-File')]",
        ])
        if not efile_clicked:
            try:
                driver.execute_script("""
                    var tags=['a','span','li','button','div'];
                    for(var tag of tags){ var els=document.querySelectorAll(tag);
                    for(var el of els){ var t=(el.innerText||el.textContent||'').trim();
                    if(t==='e-File'){el.click();return;} }}
                """)
                log("  e-File clicked via JS ✓")
            except Exception as _eje:
                log(f"  e-File click failed: {_eje}", "warn")
        time.sleep(SHORT_WAIT)

        # Step 2: Hover over Income Tax Returns
        try:
            _itr_el = None
            for _xp in [
                "//*[normalize-space()='Income Tax Returns']",
                "//*[contains(normalize-space(),'Income Tax Returns')]",
                "//a[contains(normalize-space(),'Income Tax Return')]",
            ]:
                try:
                    _itr_el = WebDriverWait(driver, 5).until(
                        EC.visibility_of_element_located((By.XPATH, _xp)))
                    if _itr_el.is_displayed():
                        break
                except Exception:
                    continue
            if _itr_el:
                _AC(driver).move_to_element(_itr_el).perform()
                time.sleep(2)
                log("  Hovered 'Income Tax Returns' ✓")
        except Exception as _hov_e:
            log(f"  Hover warning: {_hov_e}", "warn")

        # Step 3: Click View Form 26AS
        f26_clicked = _click([
            "//*[normalize-space()='View Form 26AS']",
            "//a[normalize-space()='View Form 26AS']",
            "//span[normalize-space()='View Form 26AS']",
            "//*[contains(normalize-space(),'View Form 26AS')]",
            "//a[contains(normalize-space(),'26AS')]",
            "//*[contains(normalize-space(),'26AS')]",
            "//a[contains(normalize-space(),'View Tax Credit')]",
        ])
        if not f26_clicked:
            # JS scan — still no direct URL
            try:
                _r26 = driver.execute_script("""
                    var tags=['a','button','span','li','div','p'];
                    for(var tag of tags){ var els=document.querySelectorAll(tag);
                    for(var el of els){ var t=(el.innerText||el.textContent||'').trim();
                    if((t==='View Form 26AS'||t.includes('26AS'))&&el.offsetParent!==null){
                        el.scrollIntoView({block:'center'}); el.click();
                        return 'clicked: '+t; } }}
                    return 'not found';
                """)
                log(f"  JS 26AS scan: {_r26}")
                if _r26 and "not found" not in str(_r26):
                    f26_clicked = True
            except Exception as _j26e:
                log(f"  JS 26AS scan error: {_j26e}", "warn")
        log(f"  'View Form 26AS' clicked: {f26_clicked}")
        time.sleep(PAGE_WAIT)

        # Confirm any redirect popup
        _click([
            "//button[normalize-space()='Confirm']",
            "//button[normalize-space()='OK']",
            "//button[normalize-space()='Proceed']",
            "//button[contains(text(),'Confirm')]",
            "//button[contains(text(),'Continue')]",
        ])
        time.sleep(PAGE_WAIT + 2)
        log(f"  After 26AS nav URL: {driver.current_url}")

        # Handle TRACES window (opens in new tab — same as local run)
        _new_26as_handles = set(driver.window_handles) - _before_26as_handles
        if _new_26as_handles:
            driver.switch_to.window(list(_new_26as_handles)[0])
            log(f"  Switched to TRACES window: {driver.current_url}")
            time.sleep(PAGE_WAIT)
        elif len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[-1])
            log(f"  Switched to last window: {driver.current_url}")
            time.sleep(PAGE_WAIT)

        # ── On TRACES: Select AY ───────────────────────────────────────────
        try:
            selects = driver.find_elements(By.TAG_NAME, "select")
            for sel_el in selects:
                try:
                    s = Select(sel_el)
                    for opt in s.options:
                        if AY_LABEL in opt.text or str(fy_start+1) in opt.text:
                            s.select_by_visible_text(opt.text)
                            log(f"  Selected AY: {opt.text} ✓")
                            break
                except: continue
            time.sleep(SHORT_WAIT)
        except: pass

        # ── Click View Tax Credit (Form 26AS) ──────────────────────────────
        _click([
            "//input[@value='View Tax Credit (Form 26AS)']",
            "//button[contains(text(),'View Tax Credit')]",
            "//a[contains(text(),'View Tax Credit')]",
            "//input[contains(@value,'26AS')]",
            "//button[contains(text(),'26AS')]",
        ])
        time.sleep(PAGE_WAIT + 2)

        # ── Export as PDF ──────────────────────────────────────────────────
        before_26as = set(dl_dir.iterdir())
        _click([
            "//a[contains(text(),'Export as PDF')]",
            "//button[contains(text(),'Export as PDF')]",
            "//a[contains(text(),'Download')]",
            "//button[contains(text(),'Download')]",
            "//input[@value='Download']",
            "//a[@title='Download PDF']",
            "//a[contains(@href,'.pdf')]",
        ])
        time.sleep(SHORT_WAIT)
        # Confirm any dialog
        _click([
            "//button[normalize-space()='OK']",
            "//button[normalize-space()='Confirm']",
        ])

        new_f = _wait_new_file({".pdf"}, before_26as, timeout=90)
        if new_f:
            form_name = f"26AS_{pan}_AY{AY_LABEL.replace('-','_')}.pdf"
            dest = out_dir / form_name
            _shutil.copy2(str(new_f), str(dest))
            sz = dest.stat().st_size // 1024
            downloaded.append({"name": form_name, "size": f"{sz} KB"})
            log(f"  ✅ Form 26AS saved: {form_name} ({sz} KB)", "ok")
            with jobs_lock:
                if job_id in jobs: jobs[job_id]["files"] = list(downloaded)
        else:
            log("  ⚠ 26AS PDF not downloaded within 90s", "warn")
            save_fail_shot("26AS — PDF not downloaded from TRACES")

        prog(90)

        # ── ZIP all downloaded files ───────────────────────────────────────
        if downloaded:
            import zipfile as _zf
            zip_name = f"IT_Downloads_{company_name.replace(' ','_')}_{fy}.zip"
            zip_path = out_dir / zip_name
            with _zf.ZipFile(str(zip_path), "w", _zf.ZIP_DEFLATED) as zf:
                for item in downloaded:
                    fp = out_dir / item["name"]
                    if fp.exists(): zf.write(str(fp), item["name"])
            sz = zip_path.stat().st_size // 1024
            downloaded.insert(0, {"name": zip_name, "size": f"{sz} KB"})
            log(f"✅ ZIP created: {zip_name} ({sz} KB)", "ok")

        prog(100)
        n = len([d for d in downloaded if not d["name"].endswith(".zip")])
        if n > 0:
            log(f"✅ Complete! {n} file(s) downloaded from IT Portal.", "ok")
            _final_status = "done"
        else:
            log("⚠ Completed but 0 PDF files were downloaded — login may have failed or portal blocked.", "warn")
            _final_status = "done_empty"

        with jobs_lock:
            jobs[job_id]["status"] = _final_status
            jobs[job_id]["files"]  = downloaded

    except Exception as exc:
        import traceback
        log(f"❌ Error: {exc}", "err")
        for ln in traceback.format_exc().split("\n"):
            if ln.strip(): log(f"  {ln}", "err")
        with jobs_lock:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["error"]  = str(exc)
    finally:
        # Stop watchdog FIRST so it doesn't access a closed driver
        try: watchdog.stop()
        except: pass
        if driver:
            try: driver.quit()
            except: pass



# ═══════════════════════════════════════════════════════════════════
# IT BULK DOWNLOAD — Multiple clients from Excel list
# ═══════════════════════════════════════════════════════════════════

@app.route("/api/it-bulk-template")
def it_bulk_template():
    """Return a sample Excel template for IT bulk download."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "IT Clients"
    headers = ["COMPANY NAME", "PAN", "GSTIN (optional)", "IT PASSWORD", "IT_ACTIVE"]
    widths  = [32, 14, 22, 20, 12]
    for i,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64+i)].width = w
    examples = [
        ["ABC Traders Pvt Ltd", "AABCT1234C", "33AABCT1234C1ZX", "Password@123", "YES"],
        ["XYZ Enterprises",     "XYZAB5678E", "29XYZAB5678E2ZY", "MyPass456",     "YES"],
    ]
    note = ws.cell(row=1, column=6, value="NOTE: PAN is used as IT Portal User ID. GSTIN is optional (for IT Recon linking).")
    note.font = Font(italic=True, color="9C6500", size=8)
    ws.column_dimensions["F"].width = 70
    for row in examples:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=it_clients_template.xlsx"})


@app.route("/api/it-bulk-start", methods=["POST"])
@rate_limit(limit=5, window=60)
def api_it_bulk_start():
    fobj = request.files.get("clients_file")
    if not fobj:
        return jsonify(error="No file uploaded"), 400
    fy   = request.form.get("fy", "2025-26")
    mode = request.form.get("mode", "all")

    import io, openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(fobj.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return jsonify(error=f"Cannot read Excel: {e}"), 400

    if not rows or len(rows) < 2:
        return jsonify(error="Excel is empty or has only headers"), 400

    headers = [str(c or "").strip().upper() for c in rows[0]]
    def _col(*names):
        for n in names:
            if n in headers: return headers.index(n)
        return -1

    ci_name = _col("COMPANY NAME","NAME","CLIENT NAME","COMPANY")
    ci_pan  = _col("PAN")
    ci_gst  = _col("GSTIN","GSTIN (OPTIONAL)","GST")
    ci_pass = _col("IT PASSWORD","IT_PASSWORD","PASSWORD")
    ci_act  = _col("IT_ACTIVE","ACTIVE")

    if ci_pan < 0:
        return jsonify(error="Column 'PAN' not found in Excel"), 400
    if ci_pass < 0:
        return jsonify(error="Column 'IT PASSWORD' not found in Excel"), 400

    clients = []
    for row in rows[1:]:
        pan  = str(row[ci_pan] or "").strip().upper() if ci_pan >= 0 else ""
        if not pan or len(pan) != 10: continue
        active = str(row[ci_act] or "YES").strip().upper() if ci_act >= 0 else "YES"
        if active == "NO": continue
        it_pass = str(row[ci_pass] or "").strip() if ci_pass >= 0 else ""
        if not it_pass: continue
        clients.append({
            "name":    str(row[ci_name] or pan).strip() if ci_name >= 0 else pan,
            "pan":     pan,
            "gstin":   str(row[ci_gst] or "").strip() if ci_gst >= 0 else "",
            "it_pass": it_pass,
        })

    if not clients:
        return jsonify(error="No valid clients found — check PAN (10 chars) and IT PASSWORD columns"), 400

    job_id  = str(uuid.uuid4())[:8]
    out_dir = OUTPUT_DIR / job_id
    out_dir.mkdir(parents=True, exist_ok=True)

    with jobs_lock:
        jobs[job_id] = {
            "status":"running","progress":0,
            "logs":[{"type":"info","msg":f"Loaded {len(clients)} clients. Starting IT downloads…"}],
            "files":[],"error":None,
            "captcha_needed":False,"captcha_img":None,"captcha_company":None,
            "out_dir":str(out_dir),"counter":"",
            "failure_screenshots":[],
        }

    sess = {"captcha_q": _queue.Queue(), "screenshot": None, "refresh_event": threading.Event()}
    with _sess_lock:
        _sessions[job_id] = sess

    def _run():
        try:
            _it_bulk_worker(job_id, clients, fy, mode, sess, out_dir)
        except Exception as exc:
            import traceback as _tb
            with jobs_lock:
                if job_id in jobs:
                    jobs[job_id]["status"] = "error"
                    jobs[job_id]["error"]  = str(exc)
                    for ln in _tb.format_exc().split("\n"):
                        if ln.strip():
                            jobs[job_id]["logs"].append({"type":"err","msg":ln})
        finally:
            with _sess_lock:
                _sessions.pop(job_id, None)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify(job_id=job_id, total=len(clients))


@app.route("/api/it-bulk-otp/<job_id>", methods=["POST"])
def api_it_bulk_otp(job_id):
    """User submits OTP/CAPTCHA for a client during IT bulk download."""
    otp = (request.get_json(silent=True) or {}).get("otp","").strip()
    if not otp:
        return jsonify(ok=False, error="Empty input")
    with _sess_lock:
        sess = _sessions.get(job_id)
    if not sess:
        return jsonify(ok=False, error="No active session")
    sess["captcha_q"].put(otp)
    with jobs_lock:
        if job_id in jobs:
            jobs[job_id]["captcha_needed"] = False
            jobs[job_id]["captcha_img"]    = None
    return jsonify(ok=True)


def _it_bulk_worker(job_id, clients, fy, mode, sess, out_dir):
    """Process each IT client one by one."""
    import base64, shutil as _shutil

    def log(msg, t="info"):
        print(f"[IT-BULK {job_id}] {msg}")
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["logs"].append({"type":t,"msg":msg})

    def prog(p):
        with jobs_lock:
            if job_id in jobs: jobs[job_id]["progress"] = p

    def set_counter(i, total):
        with jobs_lock:
            if job_id in jobs: jobs[job_id]["counter"] = f"Client {i}/{total}"

    def show_captcha_for_client(img_b64, company):
        sess["screenshot"] = img_b64
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["captcha_needed"]  = True
                jobs[job_id]["captcha_img"]     = img_b64
                jobs[job_id]["captcha_company"] = company

    def clear_captcha():
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["captcha_needed"]  = False
                jobs[job_id]["captcha_img"]     = None
                jobs[job_id]["captcha_company"] = None

    def wait_otp(prompt):
        while not sess["captcha_q"].empty():
            try: sess["captcha_q"].get_nowait()
            except: pass
        log(f"⏳ {prompt}")
        try:
            return sess["captcha_q"].get(timeout=900)
        except _queue.Empty:
            raise RuntimeError("OTP timeout — no input received in 15 minutes")

    total    = len(clients)
    all_files = []
    out_path  = Path(out_dir)

    for idx, client in enumerate(clients, 1):
        set_counter(idx, total)
        name    = client["name"]
        pan     = client["pan"]
        gstin   = client.get("gstin","")
        it_pass = client["it_pass"]
        log(f"━━━ [{idx}/{total}] {name} (PAN: {pan}) ━━━")
        prog(int((idx-1)/total*100))

        client_dir = out_path / pan
        client_dir.mkdir(exist_ok=True)

        # Run the IT auto-downloader for this client
        # We create a mini job state just for the sub-session
        sub_job_id = f"{job_id}_{pan}"
        with jobs_lock:
            jobs[sub_job_id] = {
                "status":"running","progress":0,"logs":[],
                "files":[],"error":None,
                "captcha_needed":False,"captcha_img":None,
                "out_dir":str(client_dir),"failure_screenshots":[],
            }

        # Shared sess: route captcha/OTP from sub-job to main job UI
        sub_sess = {
            "captcha_q":     sess["captcha_q"],   # shared queue!
            "screenshot":    None,
            "refresh_event": sess["refresh_event"],
        }
        # Intercept captcha_needed updates so main job shows them
        import threading as _th
        import queue as _q

        # Monkey-patch: after _it_auto_download sets captcha on sub_job,
        # mirror it to the parent job with client info
        _orig_q = sub_sess["captcha_q"]

        class _MirrorSess(dict):
            """Wrapper that mirrors show_captcha calls to parent job."""
            def __setitem__(self, k, v):
                super().__setitem__(k, v)
                if k == "screenshot" and v:
                    with jobs_lock:
                        if job_id in jobs:
                            jobs[job_id]["captcha_needed"] = True
                            jobs[job_id]["captcha_img"]    = v
                            jobs[job_id]["captcha_company"] = {"name":name,"pan":pan}

        mirror_sess = _MirrorSess(sub_sess)

        try:
            if mode in ("all","pdfs","26as","ais_tis"):
                _it_auto_download(sub_job_id, pan, name, pan, it_pass, fy, mirror_sess)
            clear_captcha()

            # Collect downloaded files into main output
            sub_job = jobs.get(sub_job_id,{})
            co_files = sub_job.get("files",[])
            # Real PDFs = non-ZIP files; status "done" (not "done_empty") + ≥1 PDF = true success
            real_files = [f for f in co_files if not f["name"].endswith(".zip")]
            download_ok = (sub_job.get("status") == "done") and len(real_files) > 0
            if download_ok:
                # dl_dir used by _it_auto_download is <out_dir>/it_downloads
                sub_out_base = Path(sub_job.get("out_dir", str(client_dir)))
                sub_dl_dir   = sub_out_base / "it_downloads"
                for f in co_files:
                    fname = f["name"]
                    # Search in client_dir, then it_downloads subdir, then sub_out root
                    candidate = None
                    for search_dir in [client_dir, sub_dl_dir, sub_out_base]:
                        c = search_dir / fname
                        if c.exists():
                            candidate = c
                            break
                    if candidate:
                        dest_name = f"{pan}_{fname}" if not fname.startswith(pan) else fname
                        dest = out_path / dest_name
                        try: shutil.copy2(str(candidate), str(dest))
                        except: pass
                        sz = dest.stat().st_size // 1024 if dest.exists() else 0
                        all_files.append({"name": dest_name, "size": f"{sz} KB"})
                        log(f"  ✓ {dest_name} ({sz} KB)", "ok")
                    else:
                        log(f"  ⚠ File not found anywhere: {fname}", "warn")
                log(f"  ✅ {name}: {len(real_files)} PDF(s) downloaded", "ok")
            else:
                log(f"  ⚠ {name}: download may have failed — check failure screenshots", "warn")

            # Run IT Recon if requested
            # Only skip recon when mode is purely "pdfs"/"26as"/"ais_tis" with no downloads.
            # When mode is "recon" (offline), always run. When mode is "all", only run if
            # downloads succeeded OR if user explicitly chose recon-only mode.
            _recon_mode_only = (mode == "recon")
            _should_recon = _recon_mode_only or (mode == "all" and download_ok)
            if mode in ("all","recon") and _should_recon:
                engine_path = _find_engine("it_recon_engine.py")
                if engine_path:
                    log(f"  Running IT Recon for {name}...")
                    import importlib.util as _ilu
                    spec = _ilu.spec_from_file_location("it_recon_bulk", str(engine_path))
                    it_mod = _ilu.module_from_spec(spec)
                    spec.loader.exec_module(it_mod)
                    # Build filename exactly as write_it_reconciliation does:
                    # re.sub(r'[\\/:*?"<>|]', "_", name)  — spaces are kept
                    import re as _re
                    safe_name = _re.sub(r'[\\/:*?"<>|]', "_", name)
                    fy_safe = fy.replace("-", "_")
                    recon_xl = client_dir / f"IT_RECONCILIATION_{safe_name}_FY{fy_safe}.xlsx"
                    try:
                        rp = it_mod.write_it_reconciliation(
                            str(client_dir), name, pan, gstin, fy, log=lambda m,t="info": log(f"    {m}",t))
                        # Use the path returned by the engine when available
                        if rp:
                            recon_xl = Path(rp)
                        if recon_xl.exists():
                            dest = out_path / recon_xl.name
                            shutil.copy2(str(recon_xl), str(dest))
                            sz = dest.stat().st_size // 1024
                            all_files.append({"name": recon_xl.name, "size": f"{sz} KB"})
                            log(f"  ✅ IT Recon: {recon_xl.name}", "ok")
                        else:
                            log(f"  ⚠ IT Recon file not found: {recon_xl.name}", "warn")
                    except Exception as recon_err:
                        log(f"  ⚠ IT Recon error for {name}: {recon_err}", "warn")

        except Exception as e:
            log(f"  ✗ Error for {name}: {e}", "err")
            clear_captcha()
        finally:
            # Clean up sub-job
            with jobs_lock:
                jobs.pop(sub_job_id, None)

        # Update running file list
        with jobs_lock:
            if job_id in jobs: jobs[job_id]["files"] = list(all_files)

    # ── Create master ZIP ──────────────────────────────────────────
    prog(98)
    if all_files:
        import zipfile as _zf
        zip_name = f"IT_BULK_{fy}_{datetime.now().strftime('%Y%m%d')}.zip"
        zip_path = out_path / zip_name
        with _zf.ZipFile(str(zip_path), "w", _zf.ZIP_DEFLATED) as zf:
            for f in all_files:
                fp = out_path / f["name"]
                if fp.exists(): zf.write(str(fp), f["name"])
        sz = zip_path.stat().st_size // 1024
        all_files.insert(0, {"name": zip_name, "size": f"{sz} KB"})
        log(f"✅ ZIP created: {zip_name} ({sz} KB)", "ok")

    prog(100)
    file_count = len(all_files) - 1 if all_files and all_files[0]["name"].endswith(".zip") else len(all_files)
    log(f"✅ IT Bulk complete — {total} clients processed, {file_count} file(s).", "ok")
    with jobs_lock:
        if job_id in jobs:
            jobs[job_id]["status"] = "done"
            jobs[job_id]["files"]  = all_files
            jobs[job_id]["captcha_needed"]  = False
            jobs[job_id]["captcha_company"] = None


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"\n  ============================================================")
    print(f"   GST Reconciliation Portal v6 — with Auto Download")
    print(f"  ============================================================")
    print(f"   Upload dir    : {UPLOAD_DIR}")
    print(f"   Output dir    : {OUTPUT_DIR}")
    print(f"   Feedback file : {FEEDBACK_FILE}")
    suite = _find_engine("gst_suite_final.py")
    ext   = _find_engine("gstr1_extract.py")
    print(f"   Suite engine  : {suite or '⚠ NOT FOUND'}")
    print(f"   GSTR-1 engine : {ext   or '⚠ NOT FOUND'}")
    print(f"   WebSocket     : {'✅ Enabled' if WEBSOCKET_AVAILABLE else '⚠ Not available'}")
    print(f"\n   Open: http://localhost:{port}")
    print(f"  ============================================================\n")
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
