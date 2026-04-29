"""
================================================================================
  GSTIN NAME CACHE  v1.0
  ========================
  Shared module used by gstr1_fy_v5.py, gst_suite_v31.py, etc.

  HOW IT WORKS:
  ─────────────
  1. Keeps a local cache file  →  gstin_name_cache.json  (same folder as script)
  2. On first run: fetches missing names from GST portal (free, no key)
  3. On later runs: reads from cache instantly — NO portal call needed
  4. Also seeds from CustomerMaster.xlsx if present (existing manual data preserved)
  5. Auto-saves new names back to cache after every run

  USAGE IN OTHER SCRIPTS:
  ────────────────────────
  from gstin_name_cache import GSTINNameCache

  cache = GSTINNameCache()                    # loads cache + CustomerMaster
  name  = cache.get("33AABCT1234C1ZX")        # instant if cached, else fetches
  names = cache.get_bulk(list_of_gstins)      # bulk fetch with progress
  cache.save()                                # persist new names to disk

================================================================================
"""

import json, ssl, time, re
from pathlib import Path
from datetime import datetime

# ── Try requests, fall back to urllib ─────────────────────────────────────────
try:
    import requests as _req
    _USE_REQUESTS = True
except ImportError:
    import urllib.request, urllib.error
    _USE_REQUESTS = False

# ── Cache file location (same folder as this script) ─────────────────────────
_CACHE_FILE       = Path(__file__).parent / "gstin_name_cache.json"
_CUSTOMER_MASTER  = Path(__file__).parent / "CustomerMaster.xlsx"
_DELAY            = 0.3    # seconds between portal calls (polite rate limit)

# ── GST portal endpoints (tried in order) ─────────────────────────────────────
_ENDPOINTS = [
    "https://services.gst.gov.in/services/api/search/gstin?gstin={gstin}",
    "https://www.gst.gov.in/util/rest/toolkit/searchTax?gstin={gstin}",
]
_HEADERS = {
    "User-Agent":      ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/124.0.0.0 Safari/537.36"),
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "en-IN,en-GB;q=0.9,en;q=0.8",
    "Referer":         "https://services.gst.gov.in/services/searchtp",
    "Origin":          "https://services.gst.gov.in",
}


# ══════════════════════════════════════════════════════════════════════════════
class GSTINNameCache:
    """
    GSTIN → Name lookup with persistent local cache.

    Parameters
    ----------
    cache_file      : path to JSON cache  (default: gstin_name_cache.json)
    customer_master : path to xlsx seed   (default: CustomerMaster.xlsx)
    log_fn          : callable(msg) for progress output (default: print)
    auto_fetch      : fetch missing names from portal automatically (default: True)
    """

    def __init__(self,
                 cache_file=None,
                 customer_master=None,
                 log_fn=None,
                 auto_fetch=True):
        self._cache_path = Path(cache_file or _CACHE_FILE)
        self._cm_path    = Path(customer_master or _CUSTOMER_MASTER)
        self._log        = log_fn or (lambda m: None)   # silent by default
        self._auto_fetch = auto_fetch
        self._data       = {}     # gstin → {legal_name, trade_name, fetched_at}
        self._dirty      = False  # True if new names were added (need save)

        self._load_cache()
        self._seed_from_customer_master()

    # ── Public API ─────────────────────────────────────────────────────────────

    def get(self, gstin, fallback=""):
        """
        Return the best display name for a GSTIN.
        Fetches from portal if not in cache (and auto_fetch is True).
        Returns fallback string if lookup fails.
        """
        gstin = self._clean(gstin)
        if not gstin:
            return fallback

        if gstin in self._data:
            return self._best_name(self._data[gstin], fallback)

        if self._auto_fetch:
            rec = self._fetch_one(gstin)
            if rec:
                self._data[gstin] = rec
                self._dirty = True
                return self._best_name(rec, fallback)

        return fallback

    def get_bulk(self, gstins, show_progress=True):
        """
        Fetch names for a list of GSTINs.
        Returns dict {gstin: name}.
        Only hits the portal for GSTINs not already in cache.
        """
        gstins  = list(dict.fromkeys(self._clean(g) for g in gstins if self._clean(g)))
        missing = [g for g in gstins if g not in self._data]

        if missing and self._auto_fetch:
            self._log(f"  📡 Fetching {len(missing)} new GSTIN name(s) from GST portal…")
            for i, gstin in enumerate(missing, 1):
                if show_progress:
                    self._log(f"    [{i}/{len(missing)}] {gstin} …")
                rec = self._fetch_one(gstin)
                if rec:
                    self._data[gstin] = rec
                    self._dirty = True
                    if show_progress:
                        self._log(f"      ✓ {self._best_name(rec)}")
                else:
                    if show_progress:
                        self._log(f"      ⚠ Not found")
                if i < len(missing):
                    time.sleep(_DELAY)

        return {g: self._best_name(self._data.get(g, {})) for g in gstins}

    def set_manual(self, gstin, legal_name, trade_name=""):
        """Manually add / override a GSTIN name (won't be overwritten by portal fetch)."""
        gstin = self._clean(gstin)
        if gstin:
            self._data[gstin] = {
                "legal_name": legal_name.strip(),
                "trade_name": trade_name.strip(),
                "source":     "manual",
                "fetched_at": datetime.now().isoformat(),
            }
            self._dirty = True

    def save(self):
        """Persist cache to disk (only writes if new data was added)."""
        if not self._dirty:
            return
        try:
            self._cache_path.write_text(
                json.dumps(self._data, ensure_ascii=False, indent=2),
                encoding="utf-8"
            )
            self._log(f"  💾 GSTIN name cache saved ({len(self._data)} entries) → {self._cache_path.name}")
            self._dirty = False
        except Exception as e:
            self._log(f"  ⚠ Cache save failed: {e}")

    def stats(self):
        """Return (total_cached, fetched_from_portal, manual, from_master)."""
        sources = [r.get("source","") for r in self._data.values()]
        return {
            "total":   len(self._data),
            "portal":  sources.count("portal"),
            "manual":  sources.count("manual"),
            "master":  sources.count("customer_master"),
        }

    def export_excel(self, out_path=None):
        """Export entire cache to Excel for reference / editing."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self._log("  ⚠ openpyxl not installed — cannot export Excel")
            return None

        out = Path(out_path or (self._cache_path.parent / "GSTIN_Name_Master.xlsx"))
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "GSTIN Master"

        hdrs = ["GSTIN", "PAN", "Legal Name", "Trade Name", "Source", "Fetched At"]
        wids = [20, 13, 45, 35, 15, 20]
        for ci, (h, w) in enumerate(zip(hdrs, wids), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Calibri")
            c.fill = PatternFill("solid", fgColor="1F3864")
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[chr(64+ci)].width = w

        for ri, (gstin, rec) in enumerate(sorted(self._data.items()), 2):
            alt = "F2F2F2" if ri % 2 == 0 else "FFFFFF"
            vals = [
                gstin,
                gstin[2:12] if len(gstin)==15 else "",
                rec.get("legal_name",""),
                rec.get("trade_name",""),
                rec.get("source",""),
                rec.get("fetched_at","")[:10] if rec.get("fetched_at") else "",
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.fill = PatternFill("solid", fgColor=alt)
                c.font = Font(name="Calibri", size=10)

        wb.save(str(out))
        self._log(f"  📊 Exported {len(self._data)} entries → {out.name}")
        return out

    # ── Private ───────────────────────────────────────────────────────────────

    def _clean(self, g):
        return re.sub(r'[^A-Z0-9]', '', str(g or "").strip().upper())

    def _best_name(self, rec, fallback=""):
        """Return trade name if available, else legal name, else fallback."""
        if not rec:
            return fallback
        return (rec.get("trade_name") or rec.get("legal_name") or fallback).strip()

    def _load_cache(self):
        """Load existing cache JSON from disk."""
        if self._cache_path.exists():
            try:
                self._data = json.loads(
                    self._cache_path.read_text(encoding="utf-8")
                )
                self._log(f"  📂 Loaded {len(self._data)} cached GSTIN names")
            except Exception as e:
                self._log(f"  ⚠ Cache load failed: {e}")
                self._data = {}

    def _seed_from_customer_master(self):
        """
        Read CustomerMaster.xlsx and seed cache with those names.
        Preserves portal-fetched data — manual/master entries take priority
        only when no portal data exists for that GSTIN.
        """
        if not self._cm_path.exists():
            return
        try:
            import openpyxl
            wb  = openpyxl.load_workbook(str(self._cm_path), read_only=True, data_only=True)
            ws  = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not rows or len(rows) < 2:
                return

            hdrs = [str(c or "").strip().upper() for c in rows[0]]

            def _col(*names):
                for n in names:
                    if n in hdrs: return hdrs.index(n)
                return -1

            ci_gstin = _col("GSTIN/UIN","GSTIN","GST NO","GSTIN NO")
            ci_name  = _col("PARTICULARS","NAME","COMPANY NAME","TRADE NAME","LEGAL NAME")

            if ci_gstin == -1:
                return

            added = 0
            for row in rows[1:]:
                def _v(ci): return str(row[ci] or "").strip() if ci != -1 and ci < len(row) else ""
                gstin = self._clean(_v(ci_gstin))
                name  = _v(ci_name) if ci_name != -1 else ""
                if len(gstin) == 15 and name:
                    if gstin not in self._data:   # don't overwrite portal data
                        self._data[gstin] = {
                            "legal_name": name,
                            "trade_name": "",
                            "source":     "customer_master",
                            "fetched_at": datetime.now().isoformat(),
                        }
                        self._dirty = True
                        added += 1

            if added:
                self._log(f"  📋 Seeded {added} names from CustomerMaster.xlsx")

        except ImportError:
            pass   # openpyxl not available — skip silently
        except Exception as e:
            self._log(f"  ⚠ CustomerMaster seed failed: {e}")

    def _fetch_one(self, gstin):
        """
        Fetch a single GSTIN from the GST portal.
        Returns a record dict or None on failure.
        """
        if len(gstin) != 15:
            return None

        last_err = None
        for endpoint_tpl in _ENDPOINTS:
            url = endpoint_tpl.format(gstin=gstin)
            try:
                raw = self._http_get(url)
                info = raw.get("taxpayerInfo") or raw.get("data") or raw

                def _g(*keys):
                    for k in keys:
                        v = info.get(k, "")
                        if v: return str(v).strip()
                    return ""

                legal = _g("lgnm", "legalName", "tradeName")
                trade = _g("tradeNam", "tradeName")
                if legal or trade:
                    return {
                        "legal_name": legal,
                        "trade_name": trade if trade != legal else "",
                        "status":     _g("sts", "status"),
                        "source":     "portal",
                        "fetched_at": datetime.now().isoformat(),
                    }
            except Exception as e:
                last_err = e
                continue

        return None   # all endpoints failed

    def _http_get(self, url):
        """HTTP GET → parsed JSON. Works with or without requests."""
        if _USE_REQUESTS:
            r = _req.get(url, headers=_HEADERS, timeout=12, verify=False)
            r.raise_for_status()
            return r.json()
        else:
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            req = urllib.request.Request(url, headers=_HEADERS)
            with urllib.request.urlopen(req, timeout=12, context=ctx) as resp:
                return json.loads(resp.read().decode("utf-8"))
