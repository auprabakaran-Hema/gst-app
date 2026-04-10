"""
GST Reconciliation Suite - Annual GST Reconciliation Engine
Processes GSTR-1, GSTR-2B, GSTR-2A, and GSTR-3B files
"""
import json
import zipfile
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import openpyxl
FY_LABEL = "2025-26"
FY_MONTHS = []
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import pdfplumber


# Financial Year Configuration
FY_LABEL = "2025-26"
MONTHS = [
    ("April", "04", "2025"), ("May", "05", "2025"), ("June", "06", "2025"),
    ("July", "07", "2025"), ("August", "08", "2025"), ("September", "09", "2025"),
    ("October", "10", "2025"), ("November", "11", "2025"), ("December", "12", "2025"),
    ("January", "01", "2026"), ("February", "02", "2026"), ("March", "03", "2026"),
]


def write_annual_reconciliation(input_dir: str, client_name: str, gstin: str, logger=None):
    """
    Generate annual reconciliation report from GST files
    
    Args:
        input_dir: Directory containing GST files
        client_name: Client/Company name
        gstin: GSTIN number
        logger: Optional logger object
    """
    def log(msg):
        if logger:
            logger.info(msg)
        print(msg)
    
    input_path = Path(input_dir)
    
    log(f"Starting reconciliation for {client_name} ({gstin})")
    
    # Load all data
    gstr1_data = load_gstr1_files(input_path)
    gstr2b_data = load_gstr2b_files(input_path)
    gstr2a_data = load_gstr2a_files(input_path)
    gstr3b_data = load_gstr3b_files(input_path)
    
    log(f"Loaded: GSTR-1: {len(gstr1_data)} invoices, GSTR-2B: {len(gstr2b_data)} entries")
    
    # Create reconciliation workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Annual Summary
    create_annual_summary(wb, "Annual Summary", client_name, gstin, 
                         gstr1_data, gstr2b_data, gstr2a_data, gstr3b_data)
    
    # Sheet 2: GSTR-1 vs GSTR-3B
    create_gstr1_vs_3b(wb, "GSTR1 vs 3B", gstr1_data, gstr3b_data)
    
    # Sheet 3: GSTR-1 vs GSTR-2B
    create_gstr1_vs_2b(wb, "GSTR1 vs 2B", gstr1_data, gstr2b_data)
    
    # Sheet 4: GSTR-2A vs GSTR-2B
    create_2a_vs_2b(wb, "2A vs 2B", gstr2a_data, gstr2b_data)
    
    # Sheet 5: Missing in GSTR-2B
    create_missing_in_2b(wb, "Missing in 2B", gstr1_data, gstr2b_data)
    
    # Sheet 6: Missing in GSTR-1
    create_missing_in_gstr1(wb, "Missing in GSTR1", gstr2b_data, gstr1_data)
    
    # Sheet 7: Monthly Summary
    create_monthly_summary(wb, "Monthly Summary", gstr1_data, gstr2b_data, gstr3b_data)
    
    # Save workbook
    output_file = input_path / f"ANNUAL_RECON_{client_name.replace(' ', '_')}.xlsx"
    wb.save(output_file)
    log(f"Reconciliation report saved: {output_file}")

    # ── Save GSTR3B vs R1 Reconciled Summary ───────────────────────
    try:
        wb_r1 = Workbook(); wb_r1.remove(wb_r1.active)
        for q_name, q_months in [
            ("Q1 - APR-JUN", FY_MONTHS[:3]),
            ("Q2 - JUL-SEP", FY_MONTHS[3:6]),
            ("Q3 - OCT-DEC", FY_MONTHS[6:9]),
            ("Q4 - JAN-MAR", FY_MONTHS[9:12]),
            ("Annual - APR-MAR", FY_MONTHS),
        ]:
            ws_r1 = wb_r1.create_sheet(q_name)
            # Title
            ws_r1.merge_cells("A1:H1")
            ws_r1["A1"] = f"{FY_LABEL}  {client_name}  GSTIN: {gstin}  GSTR-3B Vs GSTR-1 RECONCILED"
            ws_r1["A1"].font = Font(bold=True, size=10)
            # Headers row 4
            for ci, hdr in enumerate(["",""]+[m for m in q_months]+
                                      (["Quarter"] if len(q_months)<12 else ["Annual"]),1):
                ws_r1.cell(row=4,column=ci+2,value=hdr).font=Font(bold=True,size=9)
            # Sub-headers row 5
            for ci in range(3, 3+len(q_months)+1):
                ws_r1.cell(row=5,column=ci,value="Taxable").font=Font(bold=True,size=8)
            # Data rows
            ROWS = [
                ("GSTR-3B", "3.1(a) - Outward taxable (other than zero/nil/exempt)", "3b_31a"),
                ("GSTR-3B", "3.1(b) - Outward taxable (zero rated)", "3b_zero"),
                ("GSTR-3B", "3.1(c) - Other outward (nil rated, exempted)", "3b_nil"),
                ("GSTR-3B", "3.1(e) - Non GST outward supplies", "zero"),
                (None, "Total from GSTR-3B (A)", "tot_3b"),
                (None, None, None),
                ("GSTR-1", "B2B Supplies", "r1_b2b"),
                ("GSTR-1", "B2C Small Supplies", "r1_b2cs"),
                ("GSTR-1", "Nil Rated / Exempt Supplies", "r1_nil"),
                ("GSTR-1", "Credit/Debit Notes (Net)", "r1_cdn"),
                (None, "Total from GSTR-1 (B)", "tot_r1"),
                (None, None, None),
                (None, "Difference (A - B)", "diff"),
            ]
            row = 6
            for sect, part, key in ROWS:
                if key is None: row+=1; continue
                ws_r1.cell(row=row,column=1,value=sect or "").font=Font(size=9)
                ws_r1.cell(row=row,column=2,value=part or "").font=Font(size=9)
                q_tot = 0.0
                for ci, mon in enumerate(q_months):
                    r1_tx = sum(i.get("taxable_value",0) for i in gstr1_data if i.get("month_year")==mon)
                    r1_tax = sum(i.get("igst",0)+i.get("cgst",0)+i.get("sgst",0) for i in gstr1_data if i.get("month_year")==mon)
                    b3_tx = sum(d.get("taxable_value",0) for d in gstr3b_data if d.get("month_year")==mon)
                    b3_tax = sum(d.get("igst",0)+d.get("cgst",0)+d.get("sgst",0) for d in gstr3b_data if d.get("month_year")==mon)
                    nil_v = sum(i.get("taxable_value",0) for i in gstr1_data
                                if i.get("month_year")==mon and i.get("supply_type","").upper() in ("NIL","EXEMPT","NILL"))
                    b2b_tx = sum(i.get("taxable_value",0) for i in gstr1_data
                                 if i.get("month_year")==mon and i.get("supply_type","").upper() not in ("NIL","EXEMPT","NILL","B2CS"))
                    b2cs_tx= sum(i.get("taxable_value",0) for i in gstr1_data
                                 if i.get("month_year")==mon and "B2CS" in i.get("supply_type","").upper())
                    if   key=="3b_31a":  v=round(b3_tx,2)
                    elif key=="3b_zero": v=0.0
                    elif key=="3b_nil":  v=round(nil_v,2)
                    elif key=="tot_3b":  v=round(b3_tx,2)
                    elif key=="r1_b2b":  v=round(b2b_tx,2)
                    elif key=="r1_b2cs": v=round(b2cs_tx,2)
                    elif key=="r1_nil":  v=round(nil_v,2)
                    elif key=="r1_cdn":  v=0.0
                    elif key=="tot_r1":  v=round(r1_tx,2)
                    elif key=="diff":    v=round(b3_tx-r1_tx,2)
                    else: v=0.0
                    cv=ws_r1.cell(row=row,column=3+ci,value=v)
                    cv.number_format="#,##0.00"
                    if key in ("tot_3b","tot_r1","diff"): cv.font=Font(bold=True,size=9)
                    else: cv.font=Font(size=9)
                    q_tot+=v if key!="diff" else 0
                    if key=="diff":
                        cv.fill=PatternFill("solid",fgColor=("C6EFCE" if v==0 else ("FFC7CE" if abs(v)>100 else "FFEB9C")))
                row+=1
            ws_r1.column_dimensions["A"].width=20; ws_r1.column_dimensions["B"].width=55
            for ci in range(3, 3+len(q_months)+2):
                ws_r1.column_dimensions[get_column_letter(ci)].width=15

        p_r1 = input_path / f"{gstin}_GSTR3BR1_RECONCILED_Summary_{FY_LABEL}.xlsx"
        wb_r1.save(str(p_r1)); log(f"GSTR3B-R1 Reconciled saved: {p_r1.name}")
    except Exception as _e:
        log(f"Warning: GSTR3BR1 workbook error: {_e}")

    # ── Save GSTR3B vs 2A Reconciled Summary ───────────────────────
    try:
        wb_r2a = Workbook(); wb_r2a.remove(wb_r2a.active)
        for q_name, q_months in [
            ("Q1 - APR-JUN", FY_MONTHS[:3]),
            ("Q2 - JUL-SEP", FY_MONTHS[3:6]),
            ("Q3 - OCT-DEC", FY_MONTHS[6:9]),
            ("Q4 - JAN-MAR", FY_MONTHS[9:12]),
            ("Annual - APR-MAR", FY_MONTHS),
        ]:
            ws_r2a = wb_r2a.create_sheet(q_name)
            ws_r2a.merge_cells("A1:H1")
            ws_r2a["A1"] = f"{FY_LABEL}  {client_name}  GSTIN: {gstin}  GSTR-3B Vs GSTR-2A RECONCILED"
            ws_r2a["A1"].font = Font(bold=True, size=10)
            ROWS_2A = [
                ("GSTR-3B ITC", "4A(5) All Other ITC", "3b_itc"),
                ("GSTR-3B ITC", "4C Net ITC Available", "3b_net_itc"),
                (None, "Total ITC from GSTR-3B (A)", "tot_3b"),
                (None, None, None),
                ("GSTR-2A ITC", "Total ITC from GSTR-2A", "r2a_itc"),
                (None, "Difference (2A − 3B)", "diff"),
            ]
            row = 6
            for sect, part, key in ROWS_2A:
                if key is None: row+=1; continue
                ws_r2a.cell(row=row,column=1,value=sect or "").font=Font(size=9)
                ws_r2a.cell(row=row,column=2,value=part or "").font=Font(size=9)
                for ci, mon in enumerate(q_months):
                    b3_itc = sum(d.get("cgst",0)+d.get("sgst",0)+d.get("igst",0) for d in gstr3b_data if d.get("month_year")==mon)
                    r2a_itc= sum(d.get("cgst",0)+d.get("sgst",0)+d.get("igst",0) for d in gstr2a_data if d.get("month_year")==mon)
                    if   key in ("3b_itc","3b_net_itc","tot_3b"): v=round(b3_itc,2)
                    elif key=="r2a_itc": v=round(r2a_itc,2)
                    elif key=="diff":    v=round(r2a_itc-b3_itc,2)
                    else: v=0.0
                    cv=ws_r2a.cell(row=row,column=3+ci,value=v)
                    cv.number_format="#,##0.00"
                    if key in ("tot_3b","diff"): cv.font=Font(bold=True,size=9)
                    else: cv.font=Font(size=9)
                    if key=="diff":
                        cv.fill=PatternFill("solid",fgColor=("C6EFCE" if abs(v)<100 else "FFC7CE"))
                row+=1
            ws_r2a.column_dimensions["A"].width=20; ws_r2a.column_dimensions["B"].width=45
            for ci in range(3, 3+len(q_months)+2):
                ws_r2a.column_dimensions[get_column_letter(ci)].width=15

        p_r2a = input_path / f"{gstin}_GSTR3BR2A_RECONCILED_Summary_{FY_LABEL}.xlsx"
        wb_r2a.save(str(p_r2a)); log(f"GSTR3B-R2A Reconciled saved: {p_r2a.name}")
    except Exception as _e:
        log(f"Warning: GSTR3BR2A workbook error: {_e}")

    return output_file


def load_gstr1_files(input_path: Path) -> list:
    """Load all GSTR-1 invoice data from ZIP files"""
    invoices = []
    zip_files = list(input_path.glob("GSTR1_*.zip"))
    
    for zip_file in zip_files:
        month_year = _extract_month_year(zip_file.name)
        try:
            with zipfile.ZipFile(zip_file, 'r') as zf:
                for json_name in zf.namelist():
                    if json_name.endswith('.json'):
                        with zf.open(json_name) as jf:
                            data = json.load(jf)
                            if 'b2b' in data:
                                for b2b in data['b2b']:
                                    ctin = b2b.get('ctin', '')
                                    for inv in b2b.get('inv', []):
                                        inv_data = _extract_invoice_data(inv, ctin, month_year)
                                        invoices.append(inv_data)
        except Exception as e:
            print(f"Error loading {zip_file}: {e}")
    
    return invoices


def load_gstr2b_files(input_path: Path) -> list:
    """Load GSTR-2B data from Excel files"""
    entries = []
    excel_files = list(input_path.glob("GSTR2B_*.xlsx")) + list(input_path.glob("GSTR2B_*.xls"))
    
    for excel_file in excel_files:
        month_year = _extract_month_year(excel_file.name)
        try:
            df = pd.read_excel(excel_file)
            for _, row in df.iterrows():
                entries.append({
                    'Month': month_year,
                    'GSTIN': str(row.get('GSTIN of Supplier', row.get('GSTIN', ''))),
                    'Invoice No': str(row.get('Invoice Number', row.get('Invoice No', ''))),
                    'Invoice Date': str(row.get('Invoice Date', '')),
                    'Taxable Value': float(row.get('Taxable Value', 0) or 0),
                    'IGST': float(row.get('IGST', 0) or 0),
                    'CGST': float(row.get('CGST', 0) or 0),
                    'SGST': float(row.get('SGST', 0) or 0),
                    'Total Tax': float(row.get('Total Tax', 0) or 0)
                })
        except Exception as e:
            print(f"Error loading {excel_file}: {e}")
    
    return entries


def load_gstr2a_files(input_path: Path) -> list:
    """Load GSTR-2A data from Excel or ZIP files"""
    entries = []
    files = list(input_path.glob("GSTR2A_*.xlsx")) + list(input_path.glob("GSTR2A_*.xls"))
    
    for file in files:
        month_year = _extract_month_year(file.name)
        try:
            df = pd.read_excel(file)
            for _, row in df.iterrows():
                entries.append({
                    'Month': month_year,
                    'GSTIN': str(row.get('GSTIN', '')),
                    'Invoice No': str(row.get('Invoice No', '')),
                    'Invoice Date': str(row.get('Invoice Date', '')),
                    'Taxable Value': float(row.get('Taxable Value', 0) or 0),
                    'IGST': float(row.get('IGST', 0) or 0),
                    'CGST': float(row.get('CGST', 0) or 0),
                    'SGST': float(row.get('SGST', 0) or 0)
                })
        except Exception as e:
            print(f"Error loading {file}: {e}")
    
    return entries


def load_gstr3b_files(input_path: Path) -> list:
    """Load GSTR-3B summary data from PDF files"""
    entries = []
    pdf_files = list(input_path.glob("GSTR3B_*.pdf"))
    
    for pdf_file in pdf_files:
        month_year = _extract_month_year(pdf_file.name)
        try:
            with pdfplumber.open(pdf_file) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() or ""
                
                # Extract summary data from PDF text
                entries.append({
                    'Month': month_year,
                    'Raw Text': text[:1000]  # Store first 1000 chars for analysis
                })
        except Exception as e:
            print(f"Error loading {pdf_file}: {e}")
    
    return entries


def _extract_invoice_data(inv, ctin, month_year):
    """Extract relevant data from invoice JSON"""
    items = inv.get('itms', [])
    taxable_value = sum(float(it.get('itm_det', {}).get('txval', 0)) for it in items)
    igst = sum(float(it.get('itm_det', {}).get('iamt', 0)) for it in items)
    cgst = sum(float(it.get('itm_det', {}).get('camt', 0)) for it in items)
    sgst = sum(float(it.get('itm_det', {}).get('samt', 0)) for it in items)
    
    return {
        'Month': month_year,
        'GSTIN': ctin,
        'Invoice No': inv.get('inum', ''),
        'Invoice Date': inv.get('idt', ''),
        'Invoice Value': float(inv.get('val', 0)),
        'Taxable Value': taxable_value,
        'IGST': igst,
        'CGST': cgst,
        'SGST': sgst,
        'Total Tax': igst + cgst + sgst
    }


def _extract_month_year(filename):
    """Extract month and year from filename"""
    months = {
        "january": "January", "february": "February", "march": "March",
        "april": "April", "may": "May", "june": "June",
        "july": "July", "august": "August", "september": "September",
        "october": "October", "november": "November", "december": "December"
    }
    
    name_lower = filename.lower()
    for month_key, month_name in months.items():
        if month_key in name_lower:
            year_match = re.search(r'20\d{2}', filename)
            if year_match:
                return f"{month_name} {year_match.group()}"
            return month_name
    
    return "Unknown"


def create_annual_summary(wb, title, client_name, gstin, gstr1_data, gstr2b_data, gstr2a_data, gstr3b_data):
    """Create annual summary sheet"""
    ws = wb.create_sheet(title)
    
    # Title
    ws['A1'] = f"ANNUAL GST RECONCILIATION - FY {FY_LABEL}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Client: {client_name}"
    ws['A3'] = f"GSTIN: {gstin}"
    
    # Summary data
    row = 5
    headers = ["Description", "GSTR-1", "GSTR-2B", "GSTR-2A", "GSTR-3B", "Difference"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Calculate totals
    gstr1_total = sum(inv.get('Taxable Value', 0) for inv in gstr1_data)
    gstr2b_total = sum(inv.get('Taxable Value', 0) for inv in gstr2b_data)
    gstr2a_total = sum(inv.get('Taxable Value', 0) for inv in gstr2a_data)
    
    row = 6
    ws.cell(row=row, column=1, value="Total Taxable Value")
    ws.cell(row=row, column=2, value=gstr1_total)
    ws.cell(row=row, column=3, value=gstr2b_total)
    ws.cell(row=row, column=4, value=gstr2a_total)
    ws.cell(row=row, column=6, value=gstr1_total - gstr2b_total)
    
    # Tax totals
    gstr1_tax = sum(inv.get('Total Tax', 0) for inv in gstr1_data)
    gstr2b_tax = sum(inv.get('Total Tax', 0) for inv in gstr2b_data)
    
    row = 7
    ws.cell(row=row, column=1, value="Total Tax")
    ws.cell(row=row, column=2, value=gstr1_tax)
    ws.cell(row=row, column=3, value=gstr2b_tax)
    ws.cell(row=row, column=6, value=gstr1_tax - gstr2b_tax)
    
    # Invoice counts
    row = 8
    ws.cell(row=row, column=1, value="Invoice Count")
    ws.cell(row=row, column=2, value=len(gstr1_data))
    ws.cell(row=row, column=3, value=len(gstr2b_data))
    ws.cell(row=row, column=4, value=len(gstr2a_data))
    
    # Format numbers
    for row in range(6, 9):
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'


def create_gstr1_vs_3b(wb, title, gstr1_data, gstr3b_data):
    """Create GSTR-1 vs GSTR-3B comparison sheet"""
    ws = wb.create_sheet(title)
    
    headers = ["Month", "GSTR-1 Taxable", "GSTR-1 Tax", "GSTR-3B Taxable", "GSTR-3B Tax", "Difference"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Group by month
    gstr1_by_month = defaultdict(lambda: {'taxable': 0, 'tax': 0})
    for inv in gstr1_data:
        month = inv.get('Month', 'Unknown')
        gstr1_by_month[month]['taxable'] += inv.get('Taxable Value', 0)
        gstr1_by_month[month]['tax'] += inv.get('Total Tax', 0)
    
    row = 2
    for month in [m[0] for m in MONTHS]:
        gstr1 = gstr1_by_month.get(month, {'taxable': 0, 'tax': 0})
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=gstr1['taxable'])
        ws.cell(row=row, column=3, value=gstr1['tax'])
        ws.cell(row=row, column=6, value=gstr1['tax'])
        
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        row += 1


def create_gstr1_vs_2b(wb, title, gstr1_data, gstr2b_data):
    """Create GSTR-1 vs GSTR-2B comparison sheet"""
    ws = wb.create_sheet(title)
    
    headers = ["Month", "GSTR-1 Taxable", "GSTR-1 Tax", "GSTR-2B Taxable", "GSTR-2B Tax", "Difference"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Group by month
    gstr1_by_month = defaultdict(lambda: {'taxable': 0, 'tax': 0})
    for inv in gstr1_data:
        month = inv.get('Month', 'Unknown')
        gstr1_by_month[month]['taxable'] += inv.get('Taxable Value', 0)
        gstr1_by_month[month]['tax'] += inv.get('Total Tax', 0)
    
    gstr2b_by_month = defaultdict(lambda: {'taxable': 0, 'tax': 0})
    for inv in gstr2b_data:
        month = inv.get('Month', 'Unknown')
        gstr2b_by_month[month]['taxable'] += inv.get('Taxable Value', 0)
        gstr2b_by_month[month]['tax'] += inv.get('Total Tax', 0)
    
    row = 2
    for month in [m[0] for m in MONTHS]:
        gstr1 = gstr1_by_month.get(month, {'taxable': 0, 'tax': 0})
        gstr2b = gstr2b_by_month.get(month, {'taxable': 0, 'tax': 0})
        
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=gstr1['taxable'])
        ws.cell(row=row, column=3, value=gstr1['tax'])
        ws.cell(row=row, column=4, value=gstr2b['taxable'])
        ws.cell(row=row, column=5, value=gstr2b['tax'])
        ws.cell(row=row, column=6, value=gstr1['tax'] - gstr2b['tax'])
        
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        row += 1


def create_2a_vs_2b(wb, title, gstr2a_data, gstr2b_data):
    """Create GSTR-2A vs GSTR-2B comparison sheet"""
    ws = wb.create_sheet(title)
    
    headers = ["Month", "GSTR-2A Taxable", "GSTR-2A Tax", "GSTR-2B Taxable", "GSTR-2B Tax", "Difference"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Group by month
    gstr2a_by_month = defaultdict(lambda: {'taxable': 0, 'tax': 0})
    for inv in gstr2a_data:
        month = inv.get('Month', 'Unknown')
        gstr2a_by_month[month]['taxable'] += inv.get('Taxable Value', 0)
        gstr2a_by_month[month]['tax'] += inv.get('IGST', 0) + inv.get('CGST', 0) + inv.get('SGST', 0)
    
    gstr2b_by_month = defaultdict(lambda: {'taxable': 0, 'tax': 0})
    for inv in gstr2b_data:
        month = inv.get('Month', 'Unknown')
        gstr2b_by_month[month]['taxable'] += inv.get('Taxable Value', 0)
        gstr2b_by_month[month]['tax'] += inv.get('Total Tax', 0)
    
    row = 2
    for month in [m[0] for m in MONTHS]:
        gstr2a = gstr2a_by_month.get(month, {'taxable': 0, 'tax': 0})
        gstr2b = gstr2b_by_month.get(month, {'taxable': 0, 'tax': 0})
        
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=gstr2a['taxable'])
        ws.cell(row=row, column=3, value=gstr2a['tax'])
        ws.cell(row=row, column=4, value=gstr2b['taxable'])
        ws.cell(row=row, column=5, value=gstr2b['tax'])
        ws.cell(row=row, column=6, value=gstr2a['tax'] - gstr2b['tax'])
        
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        row += 1


def create_missing_in_2b(wb, title, gstr1_data, gstr2b_data):
    """Create sheet showing invoices in GSTR-1 but missing in GSTR-2B"""
    ws = wb.create_sheet(title)
    
    headers = ["Month", "GSTIN", "Invoice No", "Invoice Date", "Taxable Value", "IGST", "CGST", "SGST", "Total Tax"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Create set of GSTR-2B invoices for lookup
    gstr2b_keys = set()
    for inv in gstr2b_data:
        key = f"{inv.get('GSTIN', '')}|{inv.get('Invoice No', '')}"
        gstr2b_keys.add(key)
    
    # Find missing invoices
    row = 2
    for inv in gstr1_data:
        key = f"{inv.get('GSTIN', '')}|{inv.get('Invoice No', '')}"
        if key not in gstr2b_keys:
            ws.cell(row=row, column=1, value=inv.get('Month', ''))
            ws.cell(row=row, column=2, value=inv.get('GSTIN', ''))
            ws.cell(row=row, column=3, value=inv.get('Invoice No', ''))
            ws.cell(row=row, column=4, value=inv.get('Invoice Date', ''))
            ws.cell(row=row, column=5, value=inv.get('Taxable Value', 0))
            ws.cell(row=row, column=6, value=inv.get('IGST', 0))
            ws.cell(row=row, column=7, value=inv.get('CGST', 0))
            ws.cell(row=row, column=8, value=inv.get('SGST', 0))
            ws.cell(row=row, column=9, value=inv.get('Total Tax', 0))
            
            for col in range(5, 10):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '#,##0.00'
            row += 1


def create_missing_in_gstr1(wb, title, gstr2b_data, gstr1_data):
    """Create sheet showing invoices in GSTR-2B but missing in GSTR-1"""
    ws = wb.create_sheet(title)
    
    headers = ["Month", "GSTIN", "Invoice No", "Invoice Date", "Taxable Value", "IGST", "CGST", "SGST", "Total Tax"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Create set of GSTR-1 invoices for lookup
    gstr1_keys = set()
    for inv in gstr1_data:
        key = f"{inv.get('GSTIN', '')}|{inv.get('Invoice No', '')}"
        gstr1_keys.add(key)
    
    # Find missing invoices
    row = 2
    for inv in gstr2b_data:
        key = f"{inv.get('GSTIN', '')}|{inv.get('Invoice No', '')}"
        if key not in gstr1_keys:
            ws.cell(row=row, column=1, value=inv.get('Month', ''))
            ws.cell(row=row, column=2, value=inv.get('GSTIN', ''))
            ws.cell(row=row, column=3, value=inv.get('Invoice No', ''))
            ws.cell(row=row, column=4, value=inv.get('Invoice Date', ''))
            ws.cell(row=row, column=5, value=inv.get('Taxable Value', 0))
            ws.cell(row=row, column=6, value=inv.get('IGST', 0))
            ws.cell(row=row, column=7, value=inv.get('CGST', 0))
            ws.cell(row=row, column=8, value=inv.get('SGST', 0))
            ws.cell(row=row, column=9, value=inv.get('Total Tax', 0))
            
            for col in range(5, 10):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '#,##0.00'
            row += 1


def create_monthly_summary(wb, title, gstr1_data, gstr2b_data, gstr3b_data):
    """Create monthly summary sheet"""
    ws = wb.create_sheet(title)
    
    headers = ["Month", "GSTR-1 Count", "GSTR-1 Value", "GSTR-2B Count", "GSTR-2B Value", 
               "GSTR-3B Count", "GSTR-3B Value", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Group data by month
    gstr1_by_month = defaultdict(lambda: {'count': 0, 'value': 0})
    for inv in gstr1_data:
        month = inv.get('Month', 'Unknown')
        gstr1_by_month[month]['count'] += 1
        gstr1_by_month[month]['value'] += inv.get('Taxable Value', 0)
    
    gstr2b_by_month = defaultdict(lambda: {'count': 0, 'value': 0})
    for inv in gstr2b_data:
        month = inv.get('Month', 'Unknown')
        gstr2b_by_month[month]['count'] += 1
        gstr2b_by_month[month]['value'] += inv.get('Taxable Value', 0)
    
    row = 2
    for month in [m[0] for m in MONTHS]:
        gstr1 = gstr1_by_month.get(month, {'count': 0, 'value': 0})
        gstr2b = gstr2b_by_month.get(month, {'count': 0, 'value': 0})
        
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=gstr1['count'])
        ws.cell(row=row, column=3, value=gstr1['value'])
        ws.cell(row=row, column=4, value=gstr2b['count'])
        ws.cell(row=row, column=5, value=gstr2b['value'])
        
        # Status based on comparison
        diff = abs(gstr1['value'] - gstr2b['value'])
        if diff < 1:
            status = "MATCHED"
            status_color = "00FF00"
        elif diff < 1000:
            status = "MINOR DIFF"
            status_color = "FFFF00"
        else:
            status = "REVIEW"
            status_color = "FF0000"
        
        status_cell = ws.cell(row=row, column=8, value=status)
        status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
        
        for col in [3, 5]:
            cell = ws.cell(row=row, column=col)
            cell.number_format = '#,##0.00'
        
        row += 1


# For direct execution
if __name__ == "__main__":
    import sys
    if len(sys.argv) >= 4:
        write_annual_reconciliation(sys.argv[1], sys.argv[2], sys.argv[3])
    else:
        print("Usage: python gst_suite_final.py <input_dir> <client_name> <gstin>")
