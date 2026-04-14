"""
Income Tax Reconciliation Engine — v4 (Production)
====================================================
Built from actual PDF structure of IT portal documents.

Sheets:
  1. IT_Summary          — All key figures (TDS, Advance Tax, Turnover)
  2. TDS_26AS_Detail     — All Parts A, A1, A2, B, C from 26AS (transaction-wise)
  3. TIS_vs_GSTR_Annual  — TIS categories vs GSTR-1 Sales + Purchases (ANNUAL totals)
  4. TIS_vs_GSTR_Monthly — TIS GST Turnover vs AIS month-wise per GSTIN
  5. Purchase_Detail     — Supplier-wise purchase from AIS (all 81 suppliers)
  6. AIS_vs_Turnover     — AIS/TIS turnover vs GST recon reconciliation
  7. Advance_Tax_Challan — Part C challans + quarter-wise summary

Usage:
  python it_recon_engine.py <folder> <company> <pan> <gstin> <fy> [--gst <amount>]

  Expects in <folder>:
    - One PDF with "26as" in name  → Form 26AS from TRACES
    - One PDF with "ais" in name   → AIS from IT portal
    - One PDF with "tis" in name   → TIS from IT portal (optional, AIS used if absent)
    - Optional: GST Recon Excel from gst_suite_final.py
"""

import re, sys
from pathlib import Path
from datetime import datetime

MISSING = []
try:    import pdfplumber
except: MISSING.append("pdfplumber")
try:    import pandas as pd
except: MISSING.append("pandas")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except: MISSING.append("openpyxl")

if MISSING:
    print(f"Missing: pip install {' '.join(MISSING)}"); sys.exit(1)

# ── Colours ────────────────────────────────────────────────────────────────
DARK_BLUE="1F3864"; MED_BLUE="2E75B6"; SEC_BG="2E75B6"
HDR_BG="1F3864"; HDR_FG="FFFFFF"; TOT_BG="D6DCE4"
ALT1="FFFFFF"; ALT2="F2F2F2"
GREEN_BG="C6EFCE"; RED_BG="FFC7CE"; YELLOW_BG="FFEB9C"
ORANGE_BG="FCE4D6"; PURPLE="7030A0"
YELLOW_FG="9C6500"; RED_FG="9C0006"; GREEN_FG="276221"
NUM_FMT="#,##0.00"

def _f(h):  return PatternFill("solid", fgColor=h)
def _fn(b=False,c="000000",s=9): return Font(name="Arial",bold=b,color=c,size=s)
def _bd():
    x=Side(style="thin"); return Border(left=x,right=x,top=x,bottom=x)
def _al(h="left",w=False): return Alignment(horizontal=h,vertical="center",wrap_text=w)

def _is_formula(v): return isinstance(v,str) and v.startswith("=")

def _c(ws,r,col,v,bg=ALT1,bold=False,fg="000000",align="left",numfmt=None,size=9):
    c=ws.cell(row=r,column=col,value=v)
    c.font=_fn(bold,fg,size); c.fill=_f(bg); c.alignment=_al(align); c.border=_bd()
    # Apply number_format to numeric values AND formula cells (formulas evaluate to numbers)
    if numfmt and (isinstance(v,(int,float)) or _is_formula(v)):
        c.number_format=numfmt
    elif _is_formula(v) and not numfmt:
        c.number_format=NUM_FMT
    return c

def _fc(ws,r,col,formula,bg=ALT1,bold=False,fg="000000",numfmt=None,size=9):
    """Write an Excel formula cell with right-aligned numeric formatting."""
    return _c(ws,r,col,formula,bg,bold,fg,"right",numfmt or NUM_FMT,size)

def _fsum(col_letter,row_start,row_end):
    """=SUM(B3:B14)"""
    return f"=SUM({col_letter}{row_start}:{col_letter}{row_end})"

def _fdiff(col_a,col_b,row):
    """=B3-C3"""
    return f"={col_a}{row}-{col_b}{row}"

def _fsum_multi(cols,row_start,row_end):
    """=SUM(B3:B14)+SUM(C3:C14) etc."""
    return "+".join(f"SUM({c}{row_start}:{c}{row_end})" for c in cols)
if False: _fsum_multi=_fsum_multi  # keep linter happy

def _title(ws,txt,nc):
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    c=ws["A1"]; c.value=txt
    c.font=_fn(True,"FFFFFF",11); c.fill=_f(DARK_BLUE)
    c.alignment=_al("center"); c.border=_bd()
    ws.row_dimensions[1].height=26

def _hdr(ws,cols,row=2,bg=HDR_BG):
    for ci,(h,w) in enumerate(cols,1):
        c=ws.cell(row=row,column=ci,value=h)
        c.font=_fn(True,HDR_FG,9); c.fill=_f(bg)
        c.alignment=_al("center"); c.border=_bd()
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[row].height=20

def _tot(ws,ri,vals,bg=TOT_BG,fgc="000000"):
    for ci,v in enumerate(vals,1):
        c=ws.cell(row=ri,column=ci,value=v)
        c.font=_fn(True,fgc,9); c.fill=_f(bg)
        is_num = isinstance(v,(int,float)) or _is_formula(v)
        c.alignment=_al("right" if is_num else "left")
        c.border=_bd()
        if is_num: c.number_format=NUM_FMT
    ws.row_dimensions[ri].height=18

def _sep(ws,ri,txt,nc,bg=SEC_BG):
    ws.merge_cells(f"A{ri}:{get_column_letter(nc)}{ri}")
    c=ws.cell(row=ri,column=1,value=txt)
    c.font=_fn(True,"FFFFFF",9); c.fill=_f(bg)
    c.alignment=_al("left"); c.border=_bd()
    ws.row_dimensions[ri].height=14

def _note(ws,ri,txt,nc,bg=YELLOW_BG,fg=YELLOW_FG):
    ws.merge_cells(f"A{ri}:{get_column_letter(nc)}{ri}")
    c=ws.cell(row=ri,column=1,value=txt)
    c.font=_fn(False,fg,8); c.fill=_f(bg)
    c.alignment=_al("left"); c.border=_bd()
    ws.row_dimensions[ri].height=13

def _n(v):
    try: return round(float(str(v or 0).replace(",","")),2)
    except: return 0.0

def _clean(s):
    if s is None: return ""
    s=str(s).strip()
    return "" if s.lower() in ("nan","none","") else s

# ── Month ordering ─────────────────────────────────────────────────────────
MONTH_ORDER = {
    "APR":1,"MAY":2,"JUN":3,"JUL":4,"AUG":5,"SEP":6,
    "OCT":7,"NOV":8,"DEC":9,"JAN":10,"FEB":11,"MAR":12,
    "APRIL":1,"MAY":2,"JUNE":3,"JULY":4,"AUGUST":5,"SEPTEMBER":6,
    "OCTOBER":7,"NOVEMBER":8,"DECEMBER":9,"JANUARY":10,"FEBRUARY":11,"MARCH":12,
}
FY_MONTHS_ORDER = ["APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC","JAN","FEB","MAR"]

def _mon_key(period_str):
    """APR-2024 → sortable key (year, month_order)"""
    p = str(period_str).upper().strip()
    m = re.match(r"([A-Z]+)[- ](\d{4})", p)
    if m:
        mon = m.group(1)[:3]; yr = int(m.group(2))
        order = MONTH_ORDER.get(mon, 99)
        # FY sort: APR(1)..MAR(12) across years
        adj_yr = yr if order >= 1 and order <= 9 else yr + 1
        return (adj_yr, order)
    return (9999, 99)


# ═══════════════════════════════════════════════════════════════════════════
# TIS PARSER — reads the TIS PDF summary table
# ═══════════════════════════════════════════════════════════════════════════

def parse_tis_pdf(pdf_path, log=None):
    """
    Parse TIS PDF. Returns:
    {
      "header": {"pan","name","fy","ay"},
      "categories": {
          "rent_received": {"processed": 420000, "accepted": 420000},
          "interest_savings": {...},
          "business_receipts": {...},
          "gst_turnover": {...},
          "gst_purchases": {...},
          "business_expenses": {...},
          "cash_deposits": {...},
          "cash_withdrawals": {...},
      },
      "gst_turnover_detail": [{"gstin","source","amount","accepted"},...],
      "gst_purchase_detail": [{"supplier_name","pan","amount","accepted"},...],
      "business_receipt_detail": [{"deductor","section","amount"},...],
    }
    """
    def _log(m):
        if log: log(f"    {m}")
        else: print(f"    {m}")

    result = {
        "header": {},
        "categories": {},
        "gst_turnover_detail": [],
        "gst_purchase_detail": [],
        "business_receipt_detail": [],
    }

    if not Path(pdf_path).exists():
        _log(f"TIS not found: {pdf_path}"); return result

    CATEGORY_MAP = {
        "rent received":            "rent_received",
        "interest from savings":    "interest_savings",
        "business receipts":        "business_receipts",
        "gst turnover":             "gst_turnover",
        "gst purchases":            "gst_purchases",
        "business expenses":        "business_expenses",
        "cash deposits":            "cash_deposits",
        "cash withdrawals":         "cash_withdrawals",
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_tables = []
            for p in pdf.pages:
                tbls = p.extract_tables()
                if tbls: all_tables.extend(tbls)
    except Exception as e:
        _log(f"TIS pdfplumber error: {e}"); return result

    for table in all_tables:
        if not table or len(table) < 1: continue
        hdr = [_clean(c) for c in (table[0] or [])]
        hdr_text = " ".join(hdr).lower()

        # ── Main TIS summary table (has "SR. NO." + "INFORMATION CATEGORY") ──
        if "information category" in hdr_text and "processed" in hdr_text:
            for row in table[1:]:
                if not row: continue
                vals = [_clean(c) for c in row]
                # row: [sr_no, category, processed, accepted]
                # find category text and the two number columns
                cat_text = ""
                nums = []
                for v in vals:
                    if v and not v.isdigit():
                        lo = v.lower()
                        for k, mapped in CATEGORY_MAP.items():
                            if k in lo:
                                cat_text = mapped; break
                    n = _n(v)
                    if n > 0: nums.append(n)
                if cat_text and nums:
                    # nums[0] is often SR.NO (1,2,3...) — skip small integers
                    real_nums = [n for n in nums if n > 10]
                    if real_nums:
                        result["categories"][cat_text] = {
                            "processed": real_nums[0],
                            "accepted":  real_nums[1] if len(real_nums) > 1 else real_nums[0],
                        }

        # ── GST Turnover detail (GSTIN + Total Turnover per GSTIN) ──
        elif "gstin" in hdr_text and ("total turnover" in hdr_text or "return period" in hdr_text):
            for row in table[1:]:
                if not row: continue
                vals = [_clean(c) for c in row]
                # Skip sub-header rows
                if vals[0] in ("SR. NO.", "Sr. No.", ""): continue
                gstin_v = ""
                source_v = ""
                amount_v = 0.0
                # Find GSTIN pattern
                for v in vals:
                    m = re.search(r"\b(\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z\d])\b", v)
                    if m: gstin_v = m.group(1)
                nums = [_n(v) for v in vals if _n(v) > 0]
                if nums:
                    amount_v = nums[-1]  # last positive = processed/accepted
                # Source = the full row text if GSTIN info is there
                for v in vals:
                    if "KASIREDDY" in v or "BENPA" in v or "GSTIN" not in v.upper():
                        if v and len(v) > 5 and v not in (gstin_v,):
                            source_v = v[:60]; break
                if amount_v > 0 and gstin_v:
                    result["gst_turnover_detail"].append({
                        "gstin": gstin_v, "source": source_v,
                        "amount": amount_v, "accepted": amount_v
                    })

        # ── GST Purchase detail (supplier name + amount) ──
        elif ("purchase from supplier" in hdr_text or
              ("supplier" in hdr_text and "gstin" in hdr_text)):
            for row in table[1:]:
                if not row: continue
                vals = [_clean(c) for c in row]
                if not any(vals): continue
                # Skip header repetitions
                if any(k in _clean(vals[0]).upper() for k in
                       ["SR.", "S.NO", "GSTIN", "SUPPLIER"]): continue
                nums = [_n(v) for v in vals if _n(v) > 0]
                if not nums: continue
                # Supplier name — the long text field
                sup_name = ""
                pan_v = ""
                for v in vals:
                    # PAN pattern in brackets like (AABFU4898D)
                    pm = re.search(r"\(([A-Z]{5}\d{4}[A-Z])\)", v)
                    if pm: pan_v = pm.group(1)
                    if len(_clean(v)) > 8 and not re.match(r"^\d", _clean(v)):
                        sup_name = _clean(v)[:60]
                result["gst_purchase_detail"].append({
                    "supplier_name": sup_name,
                    "pan": pan_v,
                    "amount": nums[-1],
                    "accepted": nums[-1],
                })

        # ── Business receipt detail ──
        elif "information description" in hdr_text and "information source" in hdr_text:
            for row in table[1:]:
                if not row: continue
                vals = [_clean(c) for c in row]
                if not any(vals): continue
                nums = [_n(v) for v in vals if _n(v) > 0]
                if not nums: continue
                desc = ""
                for v in vals:
                    if any(k in v.lower() for k in
                           ["receipt","contract","perquisit","rent","benefit"]):
                        desc = v[:80]; break
                # Info source
                source = ""
                for v in vals:
                    if "LIMITED" in v or "CORPORATION" in v or "PRIVATE" in v:
                        source = v[:60]; break
                # Section
                sec = ""
                for v in vals:
                    m = re.search(r"(194[A-Z0-9]+|192)", v)
                    if m: sec = m.group(1); break
                if nums:
                    result["business_receipt_detail"].append({
                        "description": desc, "source": source,
                        "section": sec, "amount": nums[-1],
                    })

    # ── Extract header ──────────────────────────────────────────────
    for table in all_tables:
        if not table: continue
        hdr_text = " ".join(_clean(c) for c in (table[0] or [])).lower()
        for row in table[:3]:
            if not row: continue
            for cell in row:
                cv = _clean(cell)
                pan_m = re.search(r"\b([A-Z]{5}\d{4}[A-Z])\b", cv)
                if pan_m: result["header"]["pan"] = pan_m.group(1)
                if "KASIREDDY" in cv or ("ARUN" in cv and "KUMAR" in cv):
                    result["header"]["name"] = cv[:80]
                fy_m = re.search(r"(\d{4}-\d{2,4})", cv)
                if fy_m and "financial" not in cv.lower():
                    result["header"]["fy"] = fy_m.group(1)
                if "financial year" in cv.lower():
                    fm = re.search(r"(\d{4}-\d{2,4})", cv)
                    if fm: result["header"]["fy"] = fm.group(1)

    _log(f"TIS: categories={list(result['categories'].keys())} "
         f"gst_turnover_detail={len(result['gst_turnover_detail'])} "
         f"purchases={len(result['gst_purchase_detail'])}")
    return result


# ═══════════════════════════════════════════════════════════════════════════
# AIS PARSER — reads AIS PDF (month-wise GSTIN turnover + supplier purchases)
# ═══════════════════════════════════════════════════════════════════════════

def parse_ais_pdf(pdf_path, log=None):
    """
    Parse AIS PDF. Returns:
    {
      "header": {"pan","name","fy","ay"},
      "gst_turnover_monthly": {
          "37BENPA6909L1Z3": [
              {"period":"APR-2024","total_turnover":999,"taxable_turnover":999,"status":"Active"},
              ...
          ],
          "33BENPA6909L1ZB": [...],
      },
      "gst_purchases_by_supplier": [
          {"supplier_name","pan","gstin_buyer","period","amount","status"},
          ...
      ],
      "tds_income": [{"code","description","source","amount","quarter_details":[...]}],
      "interest": [{"bank","account","amount"}],
      "cash_deposits": {"total":0,"detail":[]},
      "refunds": [{"ay","amount","date"}],
      "tax_paid": [{"ay","amount","bsr","date","challan"}],
      "summary": {"total_turnover_37":0,"total_turnover_33":0,"total_purchases":0,
                  "total_tds_income":0,"total_interest":0,"total_cash_deposit":0}
    }
    """
    def _log(m):
        if log: log(f"    {m}")
        else: print(f"    {m}")

    result = {
        "header": {},
        "gst_turnover_monthly": {},
        "gst_purchases_by_supplier": [],
        "tds_income": [],
        "interest": [],
        "cash_deposits": {"total": 0.0, "detail": []},
        "refunds": [],
        "tax_paid": [],
        "summary": {
            "total_turnover": 0.0, "total_purchases": 0.0,
            "total_tds_income": 0.0, "total_interest": 0.0,
            "total_cash_deposit": 0.0,
        }
    }

    if not Path(pdf_path).exists():
        _log(f"AIS not found: {pdf_path}"); return result

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
            all_tables = []
            for p in pdf.pages:
                tbls = p.extract_tables()
                if tbls: all_tables.extend(tbls)
    except Exception as e:
        _log(f"AIS pdfplumber error: {e}"); return result

    # ── Header from text ───────────────────────────────────────────
    for line in all_text.split("\n")[:30]:
        m = re.search(r"\b([A-Z]{5}\d{4}[A-Z])\b", line)
        if m: result["header"]["pan"] = m.group(1)
        if "KASIREDDY" in line or ("ARUN" in line and "KUMAR" in line):
            result["header"]["name"] = line.strip()[:80]

    # ── AIS mixed-format table parser ──────────────────────────────
    # AIS tables contain: info-code row, then sub-header row, then data rows
    # All mixed in a single table. We parse row-by-row with state tracking.

    current_info_code = ""
    current_info_source = ""
    in_monthly_mode = False   # True after seeing GSTIN/RETURN PERIOD sub-header
    in_purchase_mode = False  # True after seeing SUPPLIER NAME sub-header
    in_interest_mode = False
    current_bank = ""

    for table in all_tables:
        if not table or len(table) < 2: continue

        for row in table:
            if not row or not any(row): continue
            vals = [_clean(c) for c in row]
            flat_vals = [v for v in vals if v]
            flat_lower = " ".join(flat_vals).lower()

            # ── Skip footer/header noise ────────────────────────
            if "download id" in flat_lower or "generation date" in flat_lower: continue
            if "pan name assessment" in flat_lower: continue

            # ── Detect info code rows (EXC-GSTR3B, SFT-016, TDS-194C etc.) ──
            if len(flat_vals) >= 2 and flat_vals[0].isdigit():
                code_val = flat_vals[1] if len(flat_vals) > 1 else ""
                if re.match(r"EXC-|SFT-|TDS-|TCS-", code_val):
                    current_info_code = code_val
                    in_monthly_mode = False
                    in_purchase_mode = False
                    in_interest_mode = False
                    # Extract source/description
                    for v in flat_vals[3:]:
                        if len(v) > 8 and not re.match(r"[\d,]+$", v):
                            current_info_source = v[:80]; break
                    # For interest, track bank name
                    if "SFT-016" in code_val:
                        current_bank = current_info_source
                        in_interest_mode = True
                    elif "EXC-GSTR3B" in code_val:
                        in_monthly_mode = True
                    elif "EXC-GSTR1" in code_val:
                        in_purchase_mode = True
                    # TDS income summary
                    elif re.match(r"TDS-|TCS-", code_val):
                        nums=[_n(v) for v in flat_vals if _n(v)>0]
                        amt=nums[-1] if nums else 0.0
                        desc=""
                        for v in flat_vals:
                            if len(v)>10 and not re.match(r"TDS-|TCS-|EXC-",v) and v!=flat_vals[0]:
                                desc=v[:80]; break
                        result["tds_income"].append({
                            "code":code_val,"description":desc,
                            "source":current_info_source,"amount":amt,
                        })
                        result["summary"]["total_tds_income"]+=amt
                    continue

            # ── Sub-header detection ─────────────────────────────
            if "GSTIN" in flat_vals and "supplier name" in flat_lower:
                in_purchase_mode = True
                in_monthly_mode = False; in_interest_mode = False
                continue
            if "GSTIN" in flat_vals and "RETURN PERIOD" in flat_vals and "supplier name" not in flat_lower:
                in_monthly_mode = True
                in_purchase_mode = False; in_interest_mode = False
                continue
            if "REPORTED ON" in flat_vals and "ACCOUNT NUMBER" in flat_vals:
                in_interest_mode = True
                in_monthly_mode = False; in_purchase_mode = False
                continue
            if "ACCOUNT NUMBER" in flat_vals and "ACCOUNT TYPE" in flat_vals:
                continue  # skip sub-headers for cash deposit detail

            # ── Monthly turnover data rows ───────────────────────
            if in_monthly_mode and not in_purchase_mode:
                # Row format: [sr_no, gstin, period, None, total_tv, None, taxable_tv, None, status, None]
                gstin_v=""; period_v=""; total_tv=0.0; taxable_tv=0.0; status_v=""
                for v in vals:
                    m=re.search(r"\b(\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z\d])\b",v)
                    if m and not gstin_v: gstin_v=m.group(1)
                    if re.match(r"[A-Z]{3}-\d{4}$",v): period_v=v
                    if v in ("Active","Inactive"): status_v=v
                nums=[_n(v) for v in vals if _n(v)!=0]
                real_nums=[n for n in nums if n>100]  # skip SR.NO
                if real_nums: total_tv=real_nums[0]; taxable_tv=real_nums[1] if len(real_nums)>1 else real_nums[0]
                if gstin_v and period_v:
                    if gstin_v not in result["gst_turnover_monthly"]:
                        result["gst_turnover_monthly"][gstin_v]=[]
                    result["gst_turnover_monthly"][gstin_v].append({
                        "period":period_v,"total_turnover":total_tv,
                        "taxable_turnover":taxable_tv,"status":status_v,
                    })
                continue

            # ── Purchase data rows ────────────────────────────────
            if in_purchase_mode:
                # Row: [sr_no, buyer_gstin, "SUPPLIER (GSTIN)", period, amount, status]
                buyer_gstin=""; sup_name=""; sup_gstin=""; period_v=""; amount_v=0.0; status_v=""
                for v in vals:
                    m=re.search(r"\b(\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z\d])\b",v)
                    if m and not buyer_gstin: buyer_gstin=m.group(1)
                    if re.match(r"[A-Z]{3}-\d{4}$",v): period_v=v
                    if v in ("Active","Inactive"): status_v=v
                    # Supplier name + GSTIN in same cell: "NAME (GSTIN)"
                    pm=re.search(r"^(.+?)\s*\((\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z\d])\)",v)
                    if pm: sup_name=pm.group(1).strip()[:50]; sup_gstin=pm.group(2)
                nums=[_n(v) for v in vals if _n(v)>0]
                amount_v=nums[-1] if nums else 0.0
                if buyer_gstin and period_v and amount_v>0:
                    sup_pan=sup_gstin[2:12] if len(sup_gstin)>=12 else ""
                    result["gst_purchases_by_supplier"].append({
                        "buyer_gstin":buyer_gstin,"supplier_name":sup_name,
                        "supplier_gstin":sup_gstin,"supplier_pan":sup_pan,
                        "period":period_v,"amount":amount_v,"status":status_v,
                    })
                    if status_v=="Active":
                        result["summary"]["total_purchases"]+=amount_v
                continue

            # ── Interest data rows ────────────────────────────────
            if in_interest_mode:
                nums=[_n(v) for v in vals if _n(v)>0]
                acct=""
                for v in vals:
                    if re.match(r"\d{8,}$",v): acct=v; break
                if nums and acct:
                    result["interest"].append({
                        "bank":current_bank[:40],"account":acct,"amount":nums[-1],
                    })
                    result["summary"]["total_interest"]+=nums[-1]
                continue

            # ── Cash deposit rows ─────────────────────────────────
            if "SFT-003" in current_info_code and "cash deposit" in flat_lower:
                nums=[_n(v) for v in vals if _n(v)>0]
                if nums:
                    result["cash_deposits"]["total"]+=nums[-1]
                    result["summary"]["total_cash_deposit"]+=nums[-1]
                continue

            # ── Tax paid (Part B3) ────────────────────────────────
            # Rows: [sr_no, ay, major_head, minor_head, tax, 0, 0, 0, total, bsr, date, challan, ...]
            if len(flat_vals)>=5 and flat_vals[0].isdigit():
                # Check if it's a tax payment row (has year like 2016-17)
                ay_m=re.search(r"(\d{4}-\d{2})",flat_lower)
                bsr_m=re.search(r"\b(\d{7})\b"," ".join(vals))
                date_m=re.search(r"(\d{2}/\d{2}/\d{4})"," ".join(vals))
                if ay_m and (bsr_m or date_m):
                    nums=[_n(v) for v in vals if _n(v)>0 and _n(v)<1e8]
                    if nums:
                        result["tax_paid"].append({
                            "ay":ay_m.group(1),"amount":nums[0],
                            "bsr":bsr_m.group(1) if bsr_m else "",
                            "date":date_m.group(1) if date_m else "",
                            "challan":"",
                        })
                continue

            # ── Refunds ───────────────────────────────────────────
            if "ecs" in flat_lower and "refund" in flat_lower:
                nums=[_n(v) for v in vals if _n(v)>0]
                date_m=re.search(r"(\d{2}/\d{2}/\d{4})"," ".join(vals))
                ay_m=re.search(r"(\d{4}-\d{2})"," ".join(vals))
                if nums:
                    result["refunds"].append({
                        "ay":ay_m.group(1) if ay_m else "",
                        "amount":nums[0],
                        "date":date_m.group(1) if date_m else "",
                    })
                continue

    # ── Compute totals ─────────────────────────────────────────────
    for gstin, months in result["gst_turnover_monthly"].items():
        total = sum(m["total_turnover"] for m in months
                    if m.get("status","") == "Active")
        result["summary"]["total_turnover"] = (
            result["summary"].get("total_turnover", 0) + total)

    _log(f"AIS: GSTINs={list(result['gst_turnover_monthly'].keys())} "
         f"purchases={len(result['gst_purchases_by_supplier'])} "
         f"tds_entries={len(result['tds_income'])}")
    return result


# ═══════════════════════════════════════════════════════════════════════════
# 26AS PARSER
# ═══════════════════════════════════════════════════════════════════════════

def parse_26as_pdf(pdf_path, log=None):
    """
    Parse Form 26AS from TRACES.
    Returns deductor-wise TDS with transaction detail.
    """
    def _log(m):
        if log: log(f"    {m}")
        else: print(f"    {m}")

    result = {
        "header": {},
        "deductors": [],   # Part I TDS entries
        "tcs": [],         # Part VI TCS entries
        "summary": {"total_tds": 0.0, "total_tcs": 0.0}
    }

    if not Path(pdf_path).exists():
        _log(f"26AS not found: {pdf_path}"); return result

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_tables = []
            for p in pdf.pages:
                tbls = p.extract_tables()
                if tbls: all_tables.extend(tbls)
    except Exception as e:
        _log(f"26AS pdfplumber error: {e}"); return result

    current_deductor = None

    for table in all_tables:
        if not table or len(table) < 1: continue
        hdr = [_clean(c) for c in (table[0] or [])]
        hdr_text = " ".join(hdr).lower()

        # ── Deductor summary row (Name + TAN + Total) ──
        if ("name of deductor" in hdr_text and
            ("total amount paid" in hdr_text or "total tds" in hdr_text)):
            for row in table[1:]:
                if not row: continue
                vals = [_clean(c) for c in row]
                if not any(vals): continue
                # Check if it's a deductor row (not sub-header)
                if vals[0] and vals[0].isdigit():
                    name = ""
                    tan = ""
                    # Only treat as deductor row if it has a TAN code
                    has_tan = any(re.match(r"[A-Z]{4}\d{5}[A-Z]$", v) for v in vals if v)
                    if has_tan:
                        for v in vals:
                            if re.match(r"[A-Z]{4}\d{5}[A-Z]$", v): tan = v
                            elif len(v) > 5 and not re.match(r"[\d,\.]+$", v) and v != vals[0]:
                                if not name and not re.match(r"[A-Z]{4}\d{5}[A-Z]$", v):
                                    name = v[:60]
                        nums = [_n(v) for v in vals if _n(v) > 0]
                        total_paid = nums[0] if nums else 0.0
                        total_deducted = nums[1] if len(nums) > 1 else 0.0
                        total_deposited = nums[2] if len(nums) > 2 else (nums[1] if len(nums) > 1 else 0.0)
                        current_deductor = {
                            "name": name, "tan": tan,
                            "total_paid": total_paid,
                            "total_deducted": total_deducted,
                            "total_deposited": total_deposited,
                            "transactions": [],
                        }
                        result["deductors"].append(current_deductor)
                        result["summary"]["total_tds"] += total_deposited

        # ── Transaction detail rows (Section + Date + Amount) ──
        elif ("section" in hdr_text and "transaction date" in hdr_text and
              "tds deposited" in hdr_text):
            for row in table:
                if not row: continue
                vals = [_clean(c) for c in row]
                if not any(vals): continue
                # Skip header rows
                if any(k in vals[0] for k in ["Sr.", "Section", "No.", ""]): continue
                sec = vals[0] if vals else ""
                if not sec or not re.match(r"19\d", sec): continue
                date_v = vals[1] if len(vals) > 1 else ""
                status = vals[2] if len(vals) > 2 else ""
                nums = [_n(v) for v in vals if _n(v) != 0]
                if nums and current_deductor is not None:
                    current_deductor["transactions"].append({
                        "section": sec,
                        "date": date_v,
                        "status": status,
                        "amount_paid": _n(nums[0]) if nums else 0.0,
                        "tds_deducted": _n(nums[1]) if len(nums) > 1 else 0.0,
                        "tds_deposited": _n(nums[2]) if len(nums) > 2 else (_n(nums[1]) if len(nums) > 1 else 0.0),
                    })

        # ── TCS (Part VI) ──
        elif ("name of collector" in hdr_text and
              ("total tax collected" in hdr_text or "tcs deposited" in hdr_text)):
            for row in table[1:]:
                if not row: continue
                vals = [_clean(c) for c in row]
                if not any(vals) or not vals[0].isdigit(): continue
                name = ""
                tan = ""
                for v in vals:
                    if re.match(r"[A-Z]{4}\d{5}[A-Z]$", v): tan = v
                    elif len(v) > 5 and not re.match(r"[\d,\.]+$", v) and v != vals[0]:
                        if not name: name = v[:60]
                nums = [_n(v) for v in vals if _n(v) > 0]
                tcs_dep = nums[2] if len(nums) >= 3 else (nums[0] if nums else 0.0)
                result["tcs"].append({
                    "name": name, "tan": tan,
                    "amount_received": nums[0] if nums else 0.0,
                    "tcs_collected": nums[1] if len(nums) > 1 else 0.0,
                    "tcs_deposited": tcs_dep,
                })
                result["summary"]["total_tcs"] += tcs_dep

    _log(f"26AS: deductors={len(result['deductors'])} "
         f"TDS=₹{result['summary']['total_tds']:,.2f} "
         f"TCS=₹{result['summary']['total_tcs']:,.2f}")
    return result


# ═══════════════════════════════════════════════════════════════════════════
# GST TURNOVER READER (from GST recon Excel)
# ═══════════════════════════════════════════════════════════════════════════

def _read_gst_turnover(job_dir, log=None):
    def _log(m):
        if log: log(f"  {m}")
        else: print(f"  {m}")

    ROW_LABELS=[
        "total sales (b2b + b2cs) taxable value",
        "total taxable turnover","net taxable turnover",
        "gross turnover","aggregate turnover",
        "total outward supply","total sales taxable",
    ]
    COL_KWDS=[
        "taxable value","taxable turnover","gross sale","aggregate turnover",
        "outward supply","net sales","total sales","turnover","gross turnover",
    ]

    def _pri(p):
        n=p.name.lower()
        if "annual_reconciliation" in n: return 0
        if "gst_recon" in n: return 1
        if "reconciliation" in n: return 2
        return 3

    for xl in sorted([f for f in list(Path(job_dir).glob("*.xlsx"))+list(Path(job_dir).glob("*.xls")) if "IT_RECONCILIATION" not in f.name],
                     key=_pri):
        try:
            xf=pd.ExcelFile(xl,engine="openpyxl")
            PREF=["annual_summary","summary_report","summary","gstr1_sales"]
            sheets=sorted(xf.sheet_names,
                key=lambda s:next((i for i,p in enumerate(PREF) if p in s.lower()),99))
            for sn in sheets:
                try:
                    df=xf.parse(sn,header=None,dtype=str)
                    if df.empty or df.shape[1]<2: continue
                    for _,row in df.iterrows():
                        ca=_clean(str(row.iloc[0])).lower()
                        if any(kw in ca for kw in ROW_LABELS):
                            rn=[_n(v) for v in row.iloc[1:] if _n(v)>0]
                            if rn:
                                _log(f"GST Turnover ₹{rn[-1]:,.2f} from {xl.name}→{sn}")
                                return rn[-1], f"{xl.name}→{sn}"
                    for hi in range(min(15,len(df))):
                        hdr=[_clean(str(v)).lower() for v in df.iloc[hi]]
                        mcols=[i for i,h in enumerate(hdr) if any(k in h for k in COL_KWDS)]
                        if mcols:
                            data=df.iloc[hi+1:]
                            for ci in mcols:
                                cv=pd.to_numeric(data.iloc[:,ci],errors="coerce").dropna()
                                cv=cv[cv>0]
                                if len(cv)>0:
                                    v=float(cv.sum())
                                    _log(f"GST Turnover ₹{v:,.2f} col-scan {xl.name}→{sn}")
                                    return v, f"{xl.name}→{sn}"
                except: continue
        except: continue

    _log("GST turnover not found in Excel — enter manually or use --gst")
    return 0.0,"Not found"


# ═══════════════════════════════════════════════════════════════════════════
# MAIN EXCEL WRITER
# ═══════════════════════════════════════════════════════════════════════════

def write_it_reconciliation(job_dir, company_name, pan, gstin, fy,
                             log=None, gst_turnover_override=None):
    def _log(msg,t="info"):
        if log: log(msg,t)
        else: print(f"  {msg}")

    job_dir = Path(job_dir)
    _log(f"Starting IT Reconciliation: {company_name} ({pan}) FY {fy}")

    # ── Find PDFs ──────────────────────────────────────────────────
    pdf_26as=pdf_ais=pdf_tis=None
    for p in job_dir.glob("*.pdf"):
        nm=p.name.lower()
        if any(k in nm for k in ["26as","form26","26_as","26-as","traces"]): pdf_26as=p
        elif "tis" in nm: pdf_tis=p
        elif "ais" in nm: pdf_ais=p

    # Fallback: assign by size (largest usually = AIS, medium = TIS, smallest = 26AS)
    all_pdfs=sorted(job_dir.glob("*.pdf"),key=lambda p:p.stat().st_size,reverse=True)
    if not pdf_26as and not pdf_ais and not pdf_tis and len(all_pdfs)>=1:
        # With 3 PDFs: biggest=AIS, middle=TIS, smallest=26AS
        if len(all_pdfs)>=3:
            pdf_ais=all_pdfs[0]; pdf_tis=all_pdfs[1]; pdf_26as=all_pdfs[2]
        elif len(all_pdfs)==2:
            pdf_ais=all_pdfs[0]; pdf_26as=all_pdfs[1]
        else:
            pdf_26as=all_pdfs[0]
    elif not pdf_26as and all_pdfs:
        remaining=[p for p in all_pdfs if p not in (pdf_ais,pdf_tis)]
        if remaining: pdf_26as=remaining[0]; _log(f"Using {pdf_26as.name} as 26AS","warn")

    # ── Parse ──────────────────────────────────────────────────────
    _log(f"Parsing 26AS: {pdf_26as.name if pdf_26as else 'NOT FOUND'}")
    data_26as = parse_26as_pdf(str(pdf_26as),log=log) if pdf_26as else _empty_26as()

    _log(f"Parsing AIS: {pdf_ais.name if pdf_ais else 'NOT FOUND'}")
    data_ais = parse_ais_pdf(str(pdf_ais),log=log) if pdf_ais else _empty_ais()

    _log(f"Parsing TIS: {pdf_tis.name if pdf_tis else 'NOT FOUND (using AIS)'}")
    data_tis = parse_tis_pdf(str(pdf_tis),log=log) if pdf_tis else _empty_tis()

    # ── GST Turnover ───────────────────────────────────────────────
    if gst_turnover_override and gst_turnover_override > 0:
        gst_turnover = float(gst_turnover_override)
        gst_source = f"Manual: ₹{gst_turnover:,.2f}"
    else:
        gst_turnover, gst_source = _read_gst_turnover(job_dir, log=log)

    # ── Derived totals ─────────────────────────────────────────────
    s26 = data_26as["summary"]
    tis_cats = data_tis.get("categories", {})

    tis_gst_tv  = tis_cats.get("gst_turnover",  {}).get("accepted", 0.0)
    tis_gst_pur = tis_cats.get("gst_purchases", {}).get("accepted", 0.0)
    tis_bus_rec = tis_cats.get("business_receipts",{}).get("accepted",0.0)
    tis_rent    = tis_cats.get("rent_received",  {}).get("accepted", 0.0)
    tis_interest= tis_cats.get("interest_savings",{}).get("accepted",0.0)
    tis_cash_dep= tis_cats.get("cash_deposits",  {}).get("accepted", 0.0)

    ais_tv_total = sum(
        sum(m["total_turnover"] for m in months if m.get("status","")=="Active")
        for months in data_ais["gst_turnover_monthly"].values()
    )
    ais_pur_total= data_ais["summary"]["total_purchases"]

    total_tds = (sum(d["total_deposited"] for d in data_26as["deductors"]) +
                 sum(t["tcs_deposited"]   for t in data_26as["tcs"]))

    # ── Build workbook ─────────────────────────────────────────────
    import shutil
    wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 1 — IT_Summary
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws1=wb.create_sheet("IT_Summary"); ws1.sheet_view.showGridLines=False
    C1=[("Particulars",42),("TIS / 26AS Amount ₹",20),("GSTR-1 Amount ₹",20),
        ("Difference ₹",16),("Status",16),("Remarks",32)]
    _title(ws1,f"Income Tax Reconciliation Summary — {company_name} ({pan}) — FY {fy}",len(C1))
    _hdr(ws1,C1); ws1.freeze_panes="A3"; ri=3

    def s1r(label,tis_v,gst_v=None,rmk="",bold=False,bg=None):
        nonlocal ri
        bgu=bg or (ALT2 if ri%2==0 else ALT1)
        diff=_n(gst_v-tis_v) if gst_v is not None else None
        if bold and diff is not None:
            dbg=(GREEN_BG if abs(diff)<1000 else YELLOW_BG if abs(diff)<50000 else RED_BG)
            st=("✓ Match" if abs(diff)<1000 else "⚠ Minor" if abs(diff)<50000 else "✗ Check")
        else: dbg=bgu; st=""
        _c(ws1,ri,1,label,bgu,bold=bold)
        _c(ws1,ri,2,_n(tis_v) if isinstance(tis_v,(int,float)) else (tis_v or ""),
           bgu,bold=bold,align="right",numfmt=NUM_FMT if isinstance(tis_v,(int,float)) else None)
        _c(ws1,ri,3,_n(gst_v) if gst_v is not None and isinstance(gst_v,(int,float)) else "",
           bgu,bold=bold,align="right",numfmt=NUM_FMT if isinstance(gst_v,(int,float)) else None)
        _c(ws1,ri,4,_n(diff) if bold and diff is not None else "",
           dbg if bold else bgu,bold=bold,align="right",
           numfmt=NUM_FMT if bold and diff is not None else None)
        _c(ws1,ri,5,st,dbg if bold else bgu,bold=bold)
        _c(ws1,ri,6,rmk,bgu)
        ws1.row_dimensions[ri].height=15; ri+=1

    def s1sep(lbl,c=SEC_BG):
        nonlocal ri; _sep(ws1,ri,lbl,len(C1),bg=c); ri+=1

    s1sep("COMPANY IDENTIFICATION")
    s1r("Company Name",company_name); s1r("PAN",pan)
    s1r("GSTIN(s)",gstin); s1r("Financial Year",fy)
    s1r("Assessment Year",_fy_to_ay(fy))
    s1r("Report Generated",datetime.now().strftime("%d-%b-%Y %H:%M")); ri+=1

    s1sep("TURNOVER RECONCILIATION — TIS vs GSTR-1","375623")
    s1r("TIS GST Turnover (confirmed by taxpayer)",tis_gst_tv,gst_turnover,
        gst_source,bold=True)
    s1r("AIS GST Turnover (processed by system)",ais_tv_total,gst_turnover,"",bold=True)
    # GSTIN-wise breakdown
    for gstin_v, months in data_ais["gst_turnover_monthly"].items():
        gtotal=sum(m["total_turnover"] for m in months if m.get("status","")=="Active")
        s1r(f"  GSTIN {gstin_v}",gtotal,None,
            f"{len(months)} months active")
    ri+=1

    s1sep("PURCHASES — TIS vs AIS","2E75B6")
    s1r("TIS GST Purchases (confirmed)",tis_gst_pur,None,
        "Reported by sellers in their GSTR-1")
    s1r("AIS GST Purchases (processed)",ais_pur_total,None,
        f"{len(data_ais['gst_purchases_by_supplier'])} supplier transactions")
    ri+=1

    s1sep("OTHER INCOME — From TIS","843C0C")
    s1r("Business Receipts (Sec 194C/R)",tis_bus_rec,None,
        f"{len(data_tis.get('business_receipt_detail',[]))} entries")
    s1r("Rent Received (Sec 194I(b))",tis_rent,None,"")
    s1r("Interest from Savings Bank",tis_interest,None,"")
    s1r("Cash Deposits (SFT-003)",tis_cash_dep,None,""); ri+=1

    s1sep("TDS CREDIT — From 26AS","1F3864")
    for d in data_26as["deductors"]:
        sec=""
        if d["transactions"]:
            secs=list(set(t["section"] for t in d["transactions"] if t.get("section")))
            sec=", ".join(secs[:3])
        s1r(f"  {d['name'][:40]} ({d['tan']})",
            d["total_deposited"],None,f"Sec {sec}" if sec else "")
    for t in data_26as["tcs"]:
        s1r(f"  TCS: {t['name'][:40]} ({t['tan']})",
            t["tcs_deposited"],None,"TCS")
    s1r("TOTAL TDS + TCS CREDIT",total_tds,None,
        "Claimable in ITR",bold=True,bg=TOT_BG)

    # Refunds
    if data_ais["refunds"]:
        ri+=1; s1sep("REFUNDS — From AIS Part B4")
        for ref in data_ais["refunds"]:
            s1r(f"  AY {ref['ay']} — {ref['date']}",ref["amount"],None,"ECS refund")
    ws1.sheet_properties.tabColor=DARK_BLUE

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 2 — TDS_26AS_Detail
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws2=wb.create_sheet("TDS_26AS_Detail"); ws2.sheet_view.showGridLines=False
    C2=[("Deductor Name",34),("TAN",12),("Section",10),("Transaction Date",16),
        ("Status",8),("Amount Paid ₹",16),("TDS Deducted ₹",16),("TDS Deposited ₹",16),("Remarks",20)]
    _title(ws2,f"26AS TDS / TCS Detail — {company_name} ({pan}) — FY {fy}",len(C2))
    _hdr(ws2,C2); ws2.freeze_panes="A3"; ri2=3

    for d in data_26as["deductors"]:
        # Deductor header
        _sep(ws2,ri2,f"{d['name']}  |  TAN: {d['tan']}  |  Total Deposited ₹{d['total_deposited']:,.2f}",
             len(C2),bg=MED_BLUE); ri2+=1
        blk_start=ri2
        for tx in d["transactions"]:
            bg2=ALT2 if ri2%2==0 else ALT1
            sbg=(GREEN_BG if tx["status"]=="F" else
                 RED_BG   if tx["status"] in ("U","Z") else
                 YELLOW_BG if tx["status"]=="O" else bg2)
            _c(ws2,ri2,1,d["name"][:34],bg2)
            _c(ws2,ri2,2,d["tan"],bg2)
            _c(ws2,ri2,3,tx["section"],bg2)
            _c(ws2,ri2,4,tx["date"],bg2)
            _c(ws2,ri2,5,tx["status"],sbg,align="center")
            _c(ws2,ri2,6,_n(tx["amount_paid"]),bg2,align="right",numfmt=NUM_FMT)
            _c(ws2,ri2,7,_n(tx["tds_deducted"]),bg2,align="right",numfmt=NUM_FMT)
            _c(ws2,ri2,8,_n(tx["tds_deposited"]),bg2,align="right",numfmt=NUM_FMT)
            _c(ws2,ri2,9,"",bg2)
            ws2.row_dimensions[ri2].height=14; ri2+=1
        blk_end=ri2-1
        # Use SUM formulas instead of hard-coded Python totals
        _tot(ws2,ri2,[f"Subtotal — {d['name'][:25]}","","","","",
                      _fsum("F",blk_start,blk_end),
                      _fsum("G",blk_start,blk_end),
                      _fsum("H",blk_start,blk_end),""]);ri2+=2

    # TCS
    if data_26as["tcs"]:
        _sep(ws2,ri2,"PART VI — Tax Collected at Source (TCS)",len(C2),bg="375623"); ri2+=1
        tcs_start=ri2
        for t in data_26as["tcs"]:
            bg2=ALT2 if ri2%2==0 else ALT1
            _c(ws2,ri2,1,t["name"][:34],bg2); _c(ws2,ri2,2,t["tan"],bg2)
            _c(ws2,ri2,3,"206CE",bg2); _c(ws2,ri2,4,"",bg2); _c(ws2,ri2,5,"F",GREEN_BG)
            _c(ws2,ri2,6,_n(t["amount_received"]),bg2,align="right",numfmt=NUM_FMT)
            _c(ws2,ri2,7,_n(t["tcs_collected"]),bg2,align="right",numfmt=NUM_FMT)
            _c(ws2,ri2,8,_n(t["tcs_deposited"]),bg2,align="right",numfmt=NUM_FMT)
            _c(ws2,ri2,9,"TCS",bg2)
            ws2.row_dimensions[ri2].height=14; ri2+=1

    # Grand Total: SUM entire H column from row 3
    _tot(ws2,ri2,["GRAND TOTAL","","","","","",
                  "",_fsum("H",3,ri2-1),""],bg=DARK_BLUE,fgc="FFFFFF")
    ws2.sheet_properties.tabColor="2E75B6"

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 3 — TIS_vs_GSTR_Annual  (ALL CATEGORIES — ANNUAL)
    #           NOW includes GSTR-1, GSTR-1A and GSTR-3B columns
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws3=wb.create_sheet("TIS_vs_GSTR_Annual"); ws3.sheet_view.showGridLines=False
    C3=[("Information Category",36),("TIS Processed ₹",18),("TIS Accepted ₹",18),
        ("GSTR-1 Sales ₹",18),("GSTR-1A Amended ₹",18),("GSTR-3B Filed ₹",18),
        ("Difference ₹",16),("Status",16),("Remarks",30)]
    _title(ws3,f"TIS vs GSTR-1/1A/3B — Annual Summary — {company_name} ({pan}) — FY {fy}",len(C3))
    ws3.merge_cells(f"A2:{get_column_letter(len(C3))}2")
    sh3=ws3["A2"]
    sh3.value=("TIS = Taxpayer Information Summary | GSTR-1 = Outward supply (B2B+B2CS) | "
               "GSTR-1A = Amendments | GSTR-3B = Filed liability | Diff = GSTR-1 minus TIS Accepted")
    sh3.font=_fn(False,"000000",8); sh3.fill=_f(YELLOW_BG)
    sh3.alignment=_al("left"); sh3.border=_bd()
    ws3.row_dimensions[2].height=14
    _hdr(ws3,C3,row=3); ws3.freeze_panes="A4"; ri3=4

    def s3r(label,processed,accepted,gstr1_v=None,gstr1a_v=None,gstr3b_v=None,rmk="",bold=False,bg=None):
        nonlocal ri3
        bgu=bg or (ALT2 if ri3%2==0 else ALT1)
        # Diff = GSTR-1 minus TIS Accepted
        diff=_n(gstr1_v-accepted) if gstr1_v is not None and isinstance(gstr1_v,(int,float)) else None
        if bold and diff is not None:
            dbg=(GREEN_BG if abs(diff)<1000 else YELLOW_BG if abs(diff)<50000 else RED_BG)
            st=("✓ Match" if abs(diff)<1000 else "⚠ Minor" if abs(diff)<50000 else "✗ Check")
        else: dbg=bgu; st=""
        _c(ws3,ri3,1,label,bgu,bold=bold)
        _c(ws3,ri3,2,_n(processed),bgu,align="right",numfmt=NUM_FMT)
        _c(ws3,ri3,3,_n(accepted), bgu,bold=bold,align="right",numfmt=NUM_FMT)
        _c(ws3,ri3,4,_n(gstr1_v)  if gstr1_v  is not None else "",
           bgu,align="right",numfmt=NUM_FMT if gstr1_v  is not None else None)
        _c(ws3,ri3,5,_n(gstr1a_v) if gstr1a_v is not None else "",
           "EBF3FB" if gstr1a_v else bgu,align="right",numfmt=NUM_FMT if gstr1a_v is not None else None)
        _c(ws3,ri3,6,_n(gstr3b_v) if gstr3b_v is not None else "",
           "FFF2CC" if gstr3b_v else bgu,align="right",numfmt=NUM_FMT if gstr3b_v is not None else None)
        # Excel formula: Diff = GSTR-1 (col D) minus TIS Accepted (col C)
        if bold and gstr1_v is not None:
            _c(ws3,ri3,7,f"=D{ri3}-C{ri3}",dbg,bold=bold,align="right",numfmt=NUM_FMT)
        else:
            _c(ws3,ri3,7,"",dbg if bold else bgu)
        _c(ws3,ri3,8,st,dbg if bold else bgu,bold=bold)
        _c(ws3,ri3,9,rmk,bgu)
        ws3.row_dimensions[ri3].height=15; ri3+=1

    def s3sep(lbl,c=SEC_BG): nonlocal ri3; _sep(ws3,ri3,lbl,len(C3),bg=c); ri3+=1

    s3sep("SALES — TIS GST Turnover vs GSTR-1 / GSTR-1A / GSTR-3B","375623")
    tv=tis_cats.get("gst_turnover",{})
    # GSTR-3B turnover = gst_turnover from recon excel; GSTR-1A = 0 unless parsed separately
    s3r("GST Turnover (Annual)",tv.get("processed",0),tv.get("accepted",0),
        gst_turnover,None,gst_turnover,gst_source,bold=True)
    # Per-GSTIN from AIS
    for gstin_v,months in data_ais["gst_turnover_monthly"].items():
        gtotal=sum(m["total_turnover"] for m in months if m.get("status","")=="Active")
        s3r(f"  {gstin_v}",gtotal,gtotal,None,None,None,f"{len(months)} months")
    ri3+=1

    s3sep("PURCHASES — TIS GST Purchases vs AIS (Suppliers)","2E75B6")
    pv=tis_cats.get("gst_purchases",{})
    s3r("GST Purchases (reported by sellers in GSTR-1/GSTR-3B)",
        pv.get("processed",0),pv.get("accepted",0),ais_pur_total,None,None,
        f"{len(data_ais['gst_purchases_by_supplier'])} transactions",bold=True)

    # Top suppliers from AIS
    sup_totals={}
    for row in data_ais["gst_purchases_by_supplier"]:
        if row.get("status","")=="Active":
            key=row["supplier_name"] or row["supplier_gstin"]
            sup_totals[key]=sup_totals.get(key,0)+row["amount"]
    for sup,amt in sorted(sup_totals.items(),key=lambda x:-x[1])[:20]:
        s3r(f"  {sup[:40]}",amt,amt,None,None,None,"")
    if len(sup_totals)>20:
        others=sum(v for k,v in sup_totals.items()
                   if k not in list(sup_totals.keys())[:20])
        s3r(f"  ...{len(sup_totals)-20} more suppliers",others,others,None,None,None,"")
    ri3+=1

    s3sep("CREDIT / DEBIT NOTES (CDNR) — Including Amendments","843C0C")
    # Credit notes reduce taxable turnover; debit notes increase it
    # These values come from AIS or GST recon if available
    cdnr_credit=_n(data_ais["summary"].get("total_cdnr_credit",0))
    cdnr_debit =_n(data_ais["summary"].get("total_cdnr_debit",0))
    cdnr_amend =_n(data_ais["summary"].get("total_cdnr_amended",0))
    s3r("Credit Notes — CDNR (reduces outward supply)",0,cdnr_credit,cdnr_credit,None,None,
        "Registered buyers — GSTR-1 CDNR Section")
    s3r("Debit Notes — CDNR (increases outward supply)",0,cdnr_debit,cdnr_debit,None,None,
        "Registered buyers — GSTR-1 CDNR Section")
    s3r("Amendments — CDNRA (amended CDN)",0,cdnr_amend,cdnr_amend,None,None,
        "GSTR-1A Amended Credit/Debit Notes")
    ri3+=1

    s3sep("OTHER INCOME — TIS Categories","843C0C")
    for cat_key, label, rmk in [
        ("business_receipts","Business Receipts (TDS/TCS from clients)","Sec 194C, 194R"),
        ("rent_received",    "Rent Received",                          "Sec 194I(b)"),
        ("interest_savings", "Interest from Savings Bank",             "SFT-016"),
        ("business_expenses","Business Expenses (TCS paid)",           "Sec 206C"),
        ("cash_deposits",    "Cash Deposits (SFT-003)",                "Current account"),
        ("cash_withdrawals", "Cash Withdrawals",                       "Current account"),
    ]:
        cv=tis_cats.get(cat_key,{})
        if cv.get("processed",0)>0 or cv.get("accepted",0)>0:
            s3r(label,cv.get("processed",0),cv.get("accepted",0),None,None,None,rmk)
    ws3.sheet_properties.tabColor="375623"

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 4 — TIS_vs_GSTR_Monthly (12 months × 2 GSTINs)
    #           AIS vs GSTR-1, GSTR-1A, GSTR-3B in FY order APR→MAR
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws4=wb.create_sheet("TIS_vs_GSTR_Monthly"); ws4.sheet_view.showGridLines=False
    C4=[("GSTIN",22),("Month-Year",14),
        ("AIS Total Turnover ₹",20),("AIS Taxable Turnover ₹",20),
        ("GSTR-1 Taxable ₹",18),("GSTR-1A Amended ₹",18),("GSTR-3B Filed ₹",18),
        ("Diff (R1−AIS) ₹",16),("Status",10),("Remarks",20)]
    _title(ws4,f"Month-wise GST Turnover — AIS vs GSTR-1/1A/3B — {company_name} ({pan}) — FY {fy}",len(C4))
    ws4.merge_cells(f"A2:{get_column_letter(len(C4))}2")
    sh4=ws4["A2"]
    sh4.value=("AIS = Month-wise GSTR-3B turnover reported to IT dept  |  "
               "GSTR-1 = Outward supply (upload GST Recon Excel to auto-fill)  |  "
               "GSTR-1A = Amendments (auto-filled if available)  |  "
               "GSTR-3B = Filed liability from PDF  |  Diff = GSTR-1 minus AIS Taxable")
    sh4.font=_fn(False,"000000",8); sh4.fill=_f(YELLOW_BG)
    sh4.alignment=_al("left"); sh4.border=_bd()
    ws4.row_dimensions[2].height=14
    _hdr(ws4,C4,row=3); ws4.freeze_panes="A4"; ri4=4

    FY_MONTHS_ALL=["APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC","JAN","FEB","MAR"]
    fy_start=int(fy.split("-")[0])

    for gstin_v,months in sorted(data_ais["gst_turnover_monthly"].items()):
        _sep(ws4,ri4,f"GSTIN: {gstin_v} — Same PAN: {pan}",len(C4),bg=MED_BLUE); ri4+=1
        # Build month-keyed dict
        mon_data={}
        for m in months:
            p=m["period"].upper()  # e.g. APR-2024
            mon_data[p]=m
        # Print in FY order APR→MAR
        gstin_total_tv=gstin_total_taxable=0.0
        gstin_r1_total=gstin_r1a_total=gstin_3b_total=0.0
        for mon_abbr in FY_MONTHS_ALL:
            yr=fy_start if mon_abbr not in ("JAN","FEB","MAR") else fy_start+1
            key=f"{mon_abbr}-{yr}"
            m=mon_data.get(key,{})
            ttv=_n(m.get("total_turnover",0))
            tatv=_n(m.get("taxable_turnover",0))
            status=m.get("status","")
            # GSTR-1 / 1A / 3B values — blank if not parsed from GST recon
            r1_v  = _n(m.get("gstr1_taxable",0))
            r1a_v = _n(m.get("gstr1a_amended",0))
            r3b_v = _n(m.get("gstr3b_filed",0))
            diff_v = _n(r1_v - tatv) if r1_v else None
            bg4=(ALT2 if ri4%2==0 else ALT1) if m else ALT2
            diff_bg=(GREEN_BG if diff_v is not None and abs(diff_v)<1000
                     else YELLOW_BG if diff_v is not None and abs(diff_v)<50000
                     else RED_BG   if diff_v is not None
                     else bg4)
            _c(ws4,ri4,1,gstin_v,bg4)
            _c(ws4,ri4,2,key,bg4)
            _c(ws4,ri4,3,ttv  if m else 0.0,bg4,align="right",numfmt=NUM_FMT)
            _c(ws4,ri4,4,tatv if m else 0.0,bg4,align="right",numfmt=NUM_FMT)
            _c(ws4,ri4,5,r1_v  if r1_v  else "",bg4,align="right",
               numfmt=NUM_FMT if r1_v else None)
            _c(ws4,ri4,6,r1a_v if r1a_v else "",
               "EBF3FB" if r1a_v else bg4,align="right",
               numfmt=NUM_FMT if r1a_v else None)
            _c(ws4,ri4,7,r3b_v if r3b_v else "",
               "FFF2CC" if r3b_v else bg4,align="right",
               numfmt=NUM_FMT if r3b_v else None)
            _c(ws4,ri4,8,_n(diff_v) if diff_v is not None else "",
               diff_bg,align="right",numfmt=NUM_FMT if diff_v is not None else None)
            _c(ws4,ri4,9,status if m else "NIL",
               GREEN_BG if status=="Active" else (YELLOW_BG if m else ALT2))
            _c(ws4,ri4,10,"" if m else "Not in AIS",bg4)
            ws4.row_dimensions[ri4].height=14; ri4+=1
            if m and status=="Active":
                gstin_total_tv+=ttv; gstin_total_taxable+=tatv
                gstin_r1_total+=r1_v; gstin_r1a_total+=r1a_v; gstin_3b_total+=r3b_v
        blk4_start=ri4-12; blk4_end=ri4-1  # 12 months of data
        _tot(ws4,ri4,[f"Total — {gstin_v}","",
                      _fsum("C",blk4_start,blk4_end),
                      _fsum("D",blk4_start,blk4_end),
                      _fsum("E",blk4_start,blk4_end),
                      _fsum("F",blk4_start,blk4_end),
                      _fsum("G",blk4_start,blk4_end),
                      f"=E{ri4}-D{ri4}","",""]); ri4+=2

    # Combined all GSTINs total
    all_months_total_tv=sum(
        sum(_n(m["total_turnover"]) for m in months if m.get("status","")=="Active")
        for months in data_ais["gst_turnover_monthly"].values()
    )
    all_months_total_taxable=sum(
        sum(_n(m["taxable_turnover"]) for m in months if m.get("status","")=="Active")
        for months in data_ais["gst_turnover_monthly"].values()
    )
    _tot(ws4,ri4,["GRAND TOTAL — ALL GSTINs (Same PAN)","",
                  _fsum("C",4,ri4-1),_fsum("D",4,ri4-1),
                  _fsum("E",4,ri4-1),_fsum("F",4,ri4-1),
                  _fsum("G",4,ri4-1),f"=E{ri4}-D{ri4}","",""],
         bg=DARK_BLUE,fgc="FFFFFF")
    ws4.sheet_properties.tabColor="1F3864"

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 5 — Purchase_Detail (Supplier-wise from AIS — Month-wise)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws5=wb.create_sheet("Purchase_Detail"); ws5.sheet_view.showGridLines=False
    C5=[("Buyer GSTIN",22),("Supplier Name",36),("Supplier GSTIN",22),
        ("Supplier PAN",12),("Month-Year",14),("Purchase ₹",18),("Status",10)]
    _title(ws5,f"Purchase Detail — Supplier-wise Month-wise (AIS) — {company_name} — FY {fy}",len(C5))
    _hdr(ws5,C5); ws5.freeze_panes="A3"; ri5=3

    # Group by supplier
    sup_map={}
    for row in data_ais["gst_purchases_by_supplier"]:
        key=(row["buyer_gstin"],row["supplier_name"],row["supplier_gstin"])
        if key not in sup_map: sup_map[key]=[]
        sup_map[key].append(row)

    sup_grand_total=0.0
    for (bgstin,sname,sgstin), rows in sorted(
            sup_map.items(),
            key=lambda x: -sum(r["amount"] for r in x[1] if r.get("status","")=="Active")):
        sup_total=sum(r["amount"] for r in rows if r.get("status","")=="Active")
        _sep(ws5,ri5,
             f"{sname}  |  {sgstin}  |  Total Active: ₹{sup_total:,.2f}",
             len(C5),bg=MED_BLUE); ri5+=1
        for row in sorted(rows,key=lambda r:_mon_key(r["period"])):
            bg5=ALT2 if ri5%2==0 else ALT1
            sbg=GREEN_BG if row.get("status","")=="Active" else YELLOW_BG
            _c(ws5,ri5,1,row["buyer_gstin"],bg5)
            _c(ws5,ri5,2,row["supplier_name"][:36],bg5)
            _c(ws5,ri5,3,row["supplier_gstin"],bg5)
            _c(ws5,ri5,4,row["supplier_pan"],bg5)
            _c(ws5,ri5,5,row["period"],bg5)
            _c(ws5,ri5,6,_n(row["amount"]),bg5,align="right",numfmt=NUM_FMT)
            _c(ws5,ri5,7,row.get("status",""),sbg,align="center")
            ws5.row_dimensions[ri5].height=14; ri5+=1
        _tot(ws5,ri5,[bgstin,f"Subtotal: {sname[:30]}","","","",
                      _n(sup_total),""]); ri5+=1
        sup_grand_total+=sup_total

    _tot(ws5,ri5,["GRAND TOTAL","All Suppliers","","","",
                  _n(sup_grand_total),""],
         bg=DARK_BLUE,fgc="FFFFFF")
    ws5.sheet_properties.tabColor="843C0C"

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 6 — AIS_vs_Turnover (GSTR-1 vs AIS/TIS + reconciling items)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws6=wb.create_sheet("AIS_vs_Turnover"); ws6.sheet_view.showGridLines=False
    C6=[("Particulars",40),("AIS / TIS Amount ₹",20),("GSTR-1 Amount ₹",20),
        ("Difference ₹",18),("Status",16),("Remarks",32)]
    _title(ws6,f"AIS vs GSTR-1 Turnover Reconciliation — {company_name} — FY {fy}",len(C6))
    ws6.merge_cells(f"A2:{get_column_letter(len(C6))}2")
    sh6=ws6["A2"]
    sh6.value=("Col2=AIS/TIS confirmed amount | Col3=GSTR-1 from Recon Excel | "
               "Diff=GSTR-1 minus AIS | Green=<₹1,000 | Yellow=<₹50,000 | Red=needs explanation")
    sh6.font=_fn(False,"000000",8); sh6.fill=_f(YELLOW_BG)
    sh6.alignment=_al("left"); sh6.border=_bd(); ws6.row_dimensions[2].height=14
    _hdr(ws6,C6,row=3); ws6.freeze_panes="A4"; ri6=4

    def s6r(label,av,gv=None,rmk="",bold=False,fbg=None):
        nonlocal ri6
        bg6=fbg or (ALT2 if ri6%2==0 else ALT1)
        diff=_n(gv-av) if gv is not None and isinstance(gv,(int,float)) else None
        if bold and diff is not None:
            dbg=(GREEN_BG if abs(diff)<1000 else YELLOW_BG if abs(diff)<50000 else RED_BG)
            st=("✓ Match" if abs(diff)<1000 else "⚠ Minor" if abs(diff)<50000 else "✗ Check")
        else: dbg=bg6; st=""
        _c(ws6,ri6,1,label,bg6,bold=bold)
        _c(ws6,ri6,2,_n(av),bg6,align="right",numfmt=NUM_FMT)
        _c(ws6,ri6,3,_n(gv) if gv is not None else "",bg6,align="right",
           numfmt=NUM_FMT if gv is not None else None)
        _c(ws6,ri6,4,_n(diff) if bold and diff is not None else "",
           dbg if bold else bg6,bold=bold,align="right",
           numfmt=NUM_FMT if bold and diff is not None else None)
        _c(ws6,ri6,5,st,dbg if bold else bg6,bold=bold)
        _c(ws6,ri6,6,rmk,bg6)
        ws6.row_dimensions[ri6].height=15; ri6+=1

    def s6sep(lbl,c=SEC_BG): nonlocal ri6; _sep(ws6,ri6,lbl,len(C6),bg=c); ri6+=1

    s6sep("TURNOVER — AIS/TIS vs GSTR-1","375623")
    s6r("TIS GST Turnover (confirmed by taxpayer)",tis_gst_tv,gst_turnover,gst_source,bold=True)
    s6r("AIS GST Turnover (processed by system)",ais_tv_total,gst_turnover,"",bold=True)
    for gstin_v,months in data_ais["gst_turnover_monthly"].items():
        gt=sum(_n(m["total_turnover"]) for m in months if m.get("status","")=="Active")
        s6r(f"  {gstin_v}",gt,None,f"{len(months)} months active")
    ri6+=1

    s6sep("COMMON RECONCILING ITEMS (fill manually)")
    for lbl,rmk in [
        ("Exempt/Nil-rated supplies","e.g. SEZ, nil-rated, non-GST"),
        ("Export supplies without payment","Under LUT"),
        ("Advance receipts timing difference","GST earlier, AIS/books later"),
        ("RCM inward supplies","Not in GSTR-1 outward"),
        ("Credit notes (CDNR)","Reduces GSTR-1, may not reduce AIS"),
        ("Other differences","Specify reason"),
    ]:
        s6r(lbl,0.0,0.0,rmk)
    ri6+=1

    s6sep("OTHER TIS INCOME — For ITR Cross-check","843C0C")
    s6r("Business Receipts (Sec 194C/R/194C)",tis_bus_rec,None,
        f"{len(data_tis.get('business_receipt_detail',[]))} entries in TIS")
    for br in data_tis.get("business_receipt_detail",[]):
        s6r(f"  {br.get('source','')[:40]}",_n(br.get("amount",0)),None,
            f"Sec {br.get('section','')}")
    s6r("Rent Received (Sec 194I(b))",tis_rent,None,"")
    s6r("Interest from Savings Bank",tis_interest,None,"")
    s6r("Cash Deposits (SFT-003)",tis_cash_dep,None,"")
    ws6.sheet_properties.tabColor="375623"

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 7 — Advance_Tax_Challan
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws7=wb.create_sheet("Advance_Tax_Challan"); ws7.sheet_view.showGridLines=False
    C7=[("S.No",5),("Assessment Year",14),("BSR Code",12),("Date of Deposit",16),
        ("Challan No",14),("Tax ₹",16),("Surcharge ₹",14),("Total ₹",16),("Remarks",24)]
    _title(ws7,f"Tax Paid — Part B3 (AIS) — {company_name} ({pan}) — FY {fy}",len(C7))
    _hdr(ws7,C7); ws7.freeze_panes="A3"; ri7=3

    total_paid_amt=0.0
    for sno,tp in enumerate(data_ais.get("tax_paid",[]),1):
        bg7=ALT2 if ri7%2==0 else ALT1
        amt=_n(tp.get("amount",0))
        _c(ws7,ri7,1,sno,bg7,align="center")
        _c(ws7,ri7,2,tp.get("ay",""),bg7)
        _c(ws7,ri7,3,tp.get("bsr",""),bg7)
        _c(ws7,ri7,4,tp.get("date",""),bg7)
        _c(ws7,ri7,5,tp.get("challan",""),bg7)
        _c(ws7,ri7,6,amt,bg7,align="right",numfmt=NUM_FMT)
        _c(ws7,ri7,7,0.0,bg7,align="right",numfmt=NUM_FMT)
        _c(ws7,ri7,8,amt,ORANGE_BG,align="right",numfmt=NUM_FMT)
        _c(ws7,ri7,9,"Regular Assessment / Outstanding Demand",bg7)
        ws7.row_dimensions[ri7].height=15; ri7+=1
        total_paid_amt+=amt

    if not data_ais.get("tax_paid"):
        ws7.merge_cells(f"A{ri7}:{get_column_letter(len(C7))}{ri7}")
        c7e=ws7.cell(row=ri7,column=1,value="No advance tax challans found in AIS Part B3")
        c7e.font=_fn(False,YELLOW_FG,9); c7e.fill=_f(YELLOW_BG)
        c7e.alignment=_al("center"); c7e.border=_bd(); ri7+=1

    # Refunds section
    if data_ais["refunds"]:
        ri7+=1; _sep(ws7,ri7,"REFUNDS — AIS Part B4",len(C7),bg="375623"); ri7+=1
        for ref in data_ais["refunds"]:
            bg7=ALT2 if ri7%2==0 else ALT1
            _c(ws7,ri7,1,"",bg7); _c(ws7,ri7,2,ref.get("ay",""),bg7)
            _c(ws7,ri7,3,"",bg7); _c(ws7,ri7,4,ref.get("date",""),bg7)
            _c(ws7,ri7,5,"",bg7)
            _c(ws7,ri7,6,_n(ref.get("amount",0)),GREEN_BG,align="right",numfmt=NUM_FMT)
            _c(ws7,ri7,7,0.0,bg7); _c(ws7,ri7,8,_n(ref.get("amount",0)),GREEN_BG,align="right",numfmt=NUM_FMT)
            _c(ws7,ri7,9,"ECS Refund",bg7)
            ws7.row_dimensions[ri7].height=15; ri7+=1

    ws7.sheet_properties.tabColor=PURPLE

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 8 — AIS_vs_GSTR_Monthly
    #   12 months × 2 GSTINs (same PAN) in FY order APR→MAR
    #   Columns: GSTR-1, GSTR-1A, GSTR-3B (Sales + Purchase)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws8=wb.create_sheet("AIS_vs_GSTR_Monthly"); ws8.sheet_view.showGridLines=False
    C8=[("GSTIN",22),("Month-Year",14),
        ("AIS GST Turnover ₹",20),
        ("GSTR-1 Sales ₹",18),("GSTR-1A Amended Sales ₹",20),("GSTR-3B Sales ₹",18),
        ("AIS Purchases ₹",18),("GSTR-1 Purchase ₹",18),("GSTR-3B ITC ₹",18),
        ("Sales Diff ₹",16),("Purchase Diff ₹",16),("Status",12)]
    _title(ws8,
           f"AIS vs GSTR Monthly — {company_name} ({pan}) — FY {fy} — Both GSTINs (Same PAN)",
           len(C8))
    ws8.merge_cells(f"A2:{get_column_letter(len(C8))}2")
    sh8=ws8["A2"]
    sh8.value=("Both GSTINs belong to same PAN | AIS = IT portal data | "
               "GSTR-1 = Outward supply (B2B+B2CS) | GSTR-1A = Amendments | "
               "GSTR-3B = Filed return | Sales Diff = GSTR-1 minus AIS Turnover | "
               "Green=<₹1000 | Yellow=<₹50000 | Red=review needed | APR→MAR FY order")
    sh8.font=_fn(False,"000000",8); sh8.fill=_f(YELLOW_BG)
    sh8.alignment=_al("left"); sh8.border=_bd()
    ws8.row_dimensions[2].height=14
    _hdr(ws8,C8,row=3); ws8.freeze_panes="A4"; ri8=4

    FY_MONTHS_8=["APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC","JAN","FEB","MAR"]
    fy_start8=int(fy.split("-")[0])

    # Grand totals across all GSTINs
    g8_ais_tv=g8_r1_sales=g8_r1a_sales=g8_3b_sales=0.0
    g8_ais_pur=g8_r1_pur=g8_3b_itc=0.0

    for gstin_v,months in sorted(data_ais["gst_turnover_monthly"].items()):
        _sep(ws8,ri8,f"GSTIN: {gstin_v}  (PAN: {pan})",len(C8),bg=MED_BLUE); ri8+=1
        mon_data8={}
        for m in months:
            p=m["period"].upper()
            mon_data8[p]=m

        # Build purchase map for this GSTIN
        pur_by_month8={}
        for row in data_ais.get("gst_purchases_by_supplier",[]):
            if row.get("buyer_gstin","").upper()==gstin_v.upper() and row.get("status","")=="Active":
                pm=re.sub(r"[- ]+","-",row.get("period","").upper().strip())[:7]
                pur_by_month8[pm]=pur_by_month8.get(pm,0)+row["amount"]

        st8_ais=st8_r1=st8_r1a=st8_3b=st8_pur=st8_r1pur=st8_itc=0.0

        for mon_abbr in FY_MONTHS_8:
            yr8=fy_start8 if mon_abbr not in ("JAN","FEB","MAR") else fy_start8+1
            key8=f"{mon_abbr}-{yr8}"
            m8=mon_data8.get(key8,{})
            ais_tv=_n(m8.get("total_turnover",0))
            r1_v  =_n(m8.get("gstr1_taxable",0))
            r1a_v =_n(m8.get("gstr1a_amended",0))
            r3b_v =_n(m8.get("gstr3b_filed",0))
            ais_pur=_n(pur_by_month8.get(key8,0))
            r1_pur =0.0   # purchase from GSTR-1 (if provided via supplier invoices)
            itc_v  =0.0   # ITC from GSTR-3B
            sdiff=_n(r1_v-ais_tv) if r1_v else None
            pdiff=_n(r1_pur-ais_pur) if r1_pur else None

            sdiff_bg=(GREEN_BG if sdiff is not None and abs(sdiff)<1000
                      else YELLOW_BG if sdiff is not None and abs(sdiff)<50000
                      else RED_BG   if sdiff is not None else ALT2)
            pdiff_bg=(GREEN_BG if pdiff is not None and abs(pdiff)<1000
                      else YELLOW_BG if pdiff is not None and abs(pdiff)<50000
                      else RED_BG   if pdiff is not None else ALT2)
            status8=m8.get("status","NIL" if not m8 else "")
            sbg8=(GREEN_BG if status8=="Active" else
                  YELLOW_BG if m8 else ALT2)
            bg8=(ALT2 if ri8%2==0 else ALT1) if m8 else ALT2

            _c(ws8,ri8,1,gstin_v,bg8)
            _c(ws8,ri8,2,key8,bg8)
            _c(ws8,ri8,3,ais_tv if m8 else 0.0,bg8,align="right",numfmt=NUM_FMT)
            _c(ws8,ri8,4,r1_v  if r1_v  else "",bg8,align="right",
               numfmt=NUM_FMT if r1_v else None)
            _c(ws8,ri8,5,r1a_v if r1a_v else "",
               "EBF3FB" if r1a_v else bg8,align="right",
               numfmt=NUM_FMT if r1a_v else None)
            _c(ws8,ri8,6,r3b_v if r3b_v else "",
               "FFF2CC" if r3b_v else bg8,align="right",
               numfmt=NUM_FMT if r3b_v else None)
            _c(ws8,ri8,7,ais_pur if ais_pur else "",bg8,align="right",
               numfmt=NUM_FMT if ais_pur else None)
            _c(ws8,ri8,8,"",bg8,align="right")   # GSTR-1 purchase — manual or from GST recon
            _c(ws8,ri8,9,"",bg8,align="right")   # GSTR-3B ITC — manual or from PDF
            _c(ws8,ri8,10,_n(sdiff) if sdiff is not None else "",
               sdiff_bg,align="right",numfmt=NUM_FMT if sdiff is not None else None)
            _c(ws8,ri8,11,_n(pdiff) if pdiff is not None else "",
               pdiff_bg,align="right",numfmt=NUM_FMT if pdiff is not None else None)
            _c(ws8,ri8,12,status8,sbg8)
            ws8.row_dimensions[ri8].height=14; ri8+=1

            if m8 and status8=="Active":
                st8_ais+=ais_tv; st8_r1+=r1_v; st8_r1a+=r1a_v; st8_3b+=r3b_v
                st8_pur+=ais_pur

        _tot(ws8,ri8,[f"Subtotal — {gstin_v}","",
                      _n(st8_ais),_n(st8_r1) if st8_r1 else "",
                      _n(st8_r1a) if st8_r1a else "",
                      _n(st8_3b)  if st8_3b  else "",
                      _n(st8_pur) if st8_pur else "","","","","",""]); ri8+=2
        g8_ais_tv+=st8_ais; g8_r1_sales+=st8_r1; g8_r1a_sales+=st8_r1a
        g8_3b_sales+=st8_3b; g8_ais_pur+=st8_pur

    _tot(ws8,ri8,[f"GRAND TOTAL — PAN {pan}","",
                  _n(g8_ais_tv),_n(g8_r1_sales) if g8_r1_sales else "",
                  _n(g8_r1a_sales) if g8_r1a_sales else "",
                  _n(g8_3b_sales)  if g8_3b_sales  else "",
                  _n(g8_ais_pur)   if g8_ais_pur   else "","","","","",""],
         bg=DARK_BLUE,fgc="FFFFFF")
    ws8.sheet_properties.tabColor="843C0C"

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # SHEET 9 — IT_Filing_Checklist
    #   Important activities for ITR filing verification
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws9=wb.create_sheet("IT_Filing_Checklist"); ws9.sheet_view.showGridLines=False
    C9=[("S.No",5),("Activity / Verification Step",50),("Source",24),
        ("Status",14),("Remarks",36),("Priority",12)]
    _title(ws9,f"Income Tax Filing Checklist — {company_name} ({pan}) — FY {fy} (AY {_fy_to_ay(fy)})",len(C9))
    ws9.merge_cells(f"A2:{get_column_letter(len(C9))}2")
    sh9=ws9["A2"]
    sh9.value=("Comprehensive checklist for ITR verification before filing. "
               "Status: ✓ Done | ⚠ Pending | ✗ Not Applicable. Update manually before filing.")
    sh9.font=_fn(False,"000000",8); sh9.fill=_f(YELLOW_BG)
    sh9.alignment=_al("left"); sh9.border=_bd()
    ws9.row_dimensions[2].height=14
    _hdr(ws9,C9,row=3); ws9.freeze_panes="A4"; ri9=4

    CHECKLIST=[
      # (section_label, None) or (sno, activity, source, priority)
      ("SECTION A — INCOME VERIFICATION",None,None,None,None),
      (1,"Reconcile AIS GST Turnover with GSTR-1 filed amounts","AIS + GSTR-1 JSON","⚠ Pending","Compare Sheet TIS_vs_GSTR_Monthly","HIGH"),
      (2,"Reconcile TIS Turnover with GSTR-3B filed amounts","TIS + GSTR-3B PDFs","⚠ Pending","Compare Sheet TIS_vs_GSTR_Annual","HIGH"),
      (3,"Verify GSTR-1A amendments included in final turnover","GSTR-1A ZIPs","⚠ Pending","Check AIS_vs_GSTR_Monthly sheet","HIGH"),
      (4,"Match 26AS TDS with Income/Sales invoices (Sec 194C/R)","26AS Part I","⚠ Pending","Sheet TDS_26AS_Detail","HIGH"),
      (5,"Verify Interest income from banks (Sec 194A)","26AS + AIS","⚠ Pending","Cross-check with bank statements","MEDIUM"),
      (6,"Verify Rent income reported (Sec 194I)","TIS + 26AS","⚠ Pending","Reconcile with rental agreements","MEDIUM"),
      (7,"Confirm Cash Deposits match books (SFT-003)","AIS SFT","⚠ Pending","Red flag if unexplained","HIGH"),
      (8,"Check Business Receipts match TDS certificates","TIS + 26AS","⚠ Pending","Sec 194C contract receipts","HIGH"),

      ("SECTION B — GST RECONCILIATION",None,None,None,None),
      (9,"GSTR-1 vs GSTR-3B — Output tax difference nil/explained","GSTR-1 + 3B","⚠ Pending","Sheet GSTR3B_vs_R1_Recon","HIGH"),
      (10,"GSTR-2B ITC claimed in 3B vs actual 2B amount","GSTR-2B + 3B","⚠ Pending","Check excess ITC","HIGH"),
      (11,"Credit Notes (CDNR) properly disclosed in GSTR-1","GSTR-1 JSON","⚠ Pending","Credit notes reduce turnover","MEDIUM"),
      (12,"Debit Notes (CDNR-D) added back to turnover","GSTR-1 JSON","⚠ Pending","Debit notes increase turnover","MEDIUM"),
      (13,"GSTR-1A amendments accounted in final turnover","GSTR-1A ZIPs","⚠ Pending","Check AIS_vs_GSTR_Monthly","MEDIUM"),
      (14,"Reverse Charge Mechanism (RCM) liability deposited","GSTR-3B Table 3.2","⚠ Pending","Sec 9(3) / 9(4)","MEDIUM"),
      (15,"Annual aggregate turnover correct for GST composition","All GSTINs","⚠ Pending","Both GSTINs same PAN combined","HIGH"),

      ("SECTION C — TDS / TAX CREDIT",None,None,None,None),
      (16,"All TDS certificates (Form 16/16A) received from deductors","26AS Part I","⚠ Pending","Match with each deductor","HIGH"),
      (17,"TCS credits verified (if applicable)","26AS Part VI","⚠ Pending","Sec 206C collections","MEDIUM"),
      (18,"Advance Tax challans match AIS Part B3","AIS Part B3","⚠ Pending","Sheet Advance_Tax_Challan","HIGH"),
      (19,"Self-Assessment Tax paid for current AY","Challan 280","⚠ Pending","Pay before ITR filing","HIGH"),
      (20,"Previous year refund applied/received (AIS B4)","AIS Part B4","⚠ Pending","Check refund sheet","MEDIUM"),

      ("SECTION D — DEDUCTIONS & EXEMPTIONS",None,None,None,None),
      (21,"Sec 80C deductions (LIC, PF, ELSS, PPF etc.) documented","Investment proofs","⚠ Pending","Max ₹1,50,000","HIGH"),
      (22,"Sec 80D Health Insurance premium paid","Premium receipts","⚠ Pending","Self/Family/Parents","MEDIUM"),
      (23,"Sec 80G Donations — 50%/100% eligible receipts","Donation receipts","⚠ Pending","Verify NGO registration","LOW"),
      (24,"Sec 80GGB/GGC Political party contributions","Bank proofs","⚠ Pending","Only non-cash","LOW"),
      (25,"Home loan interest Sec 24(b) certificate","Bank certificate","⚠ Pending","Max ₹2L for SOP","MEDIUM"),
      (26,"Business expenses deductible under Sec 37","Books of accounts","⚠ Pending","Maintain vouchers","HIGH"),
      (27,"Depreciation on fixed assets (Schedule FA)","Fixed asset register","⚠ Pending","Check WDV rates","MEDIUM"),

      ("SECTION E — ITR FORM & FILING",None,None,None,None),
      (28,"Confirm correct ITR form (ITR-3/4 for business)","IT Portal","⚠ Pending","Check turnover / presumptive","HIGH"),
      (29,"P&L account and Balance Sheet prepared (if ITR-3)","Books of accounts","⚠ Pending","Needed for Tax Audit if > ₹1Cr","HIGH"),
      (30,"Tax Audit (Sec 44AB) requirement checked","Turnover + profit","⚠ Pending","If turnover > ₹1Cr (non-digital)","HIGH"),
      (31,"GST Portal Annual Return GSTR-9 filed","GST Portal","⚠ Pending","Due date 31-Dec","HIGH"),
      (32,"GSTR-9C Reconciliation Statement filed (if needed)","GST Portal","⚠ Pending","If turnover > ₹5Cr","MEDIUM"),
      (33,"ITR filed with e-verification (Aadhaar OTP / Net banking)","IT Portal","⚠ Pending","Within 30 days of filing","HIGH"),
      (34,"Acknowledgment (ITR-V) saved and verified","IT Portal","⚠ Pending","Keep for 8 years","HIGH"),

      ("SECTION F — COMPLIANCE & NOTICES",None,None,None,None),
      (35,"Any outstanding demand from previous AY cleared","IT Portal Demand","⚠ Pending","Check 'e-Proceedings' tab","HIGH"),
      (36,"Respond to any AIS/TIS mismatch notices","IT Portal","⚠ Pending","Rectify or explain differences","HIGH"),
      (37,"Form 67 filed for foreign tax credit (if applicable)","IT Portal","⚠ Pending","DTAA relief claim","LOW"),
      (38,"Capital Gains (Schedule CG) computed if applicable","Sale deed / DMAT","⚠ Pending","LTCG / STCG rates","MEDIUM"),
      (39,"Foreign Assets / Foreign Income disclosed (Schedule FA/FSI)","Bank statements","⚠ Pending","Mandatory if any foreign asset","MEDIUM"),
      (40,"Director Identification Number (DIN) details if director","MCA Portal","⚠ Pending","Disclose in ITR-3","LOW"),
    ]

    sno_counter=0
    for item in CHECKLIST:
        if item[1] is None:
            # Section separator
            _sep(ws9,ri9,item[0],len(C9),bg=SEC_BG); ri9+=1
        else:
            sno,activity,source,status,remarks,priority=item
            bg9=(ALT2 if ri9%2==0 else ALT1)
            sbg={"HIGH":RED_BG,"MEDIUM":YELLOW_BG,"LOW":"E2EFDA"}.get(priority,bg9)
            sbg_status=(GREEN_BG if "Done" in status else
                        YELLOW_BG if "Pending" in status else ALT2)
            _c(ws9,ri9,1,sno,bg9,align="center")
            _c(ws9,ri9,2,activity,bg9)
            _c(ws9,ri9,3,source,bg9,align="center")
            _c(ws9,ri9,4,status,sbg_status,align="center")
            _c(ws9,ri9,5,remarks,bg9)
            _c(ws9,ri9,6,priority,sbg,align="center",bold=(priority=="HIGH"))
            ws9.row_dimensions[ri9].height=16; ri9+=1

    # Summary box
    ri9+=1
    _sep(ws9,ri9,"AUTO-DETECTED FROM UPLOADED DOCUMENTS",len(C9),bg=DARK_BLUE); ri9+=1
    auto_items=[
        (f"26AS Deductors found",f"{len(data_26as['deductors'])} deductors  |  Total TDS ₹{total_tds:,.2f}"),
        (f"AIS Turnover GSTINs",f"{len(data_ais['gst_turnover_monthly'])} GSTINs  |  Total Turnover ₹{ais_tv_total:,.2f}"),
        (f"Supplier Purchases (AIS)",f"{len(data_ais['gst_purchases_by_supplier'])} transactions  |  Total ₹{ais_pur_total:,.2f}"),
        (f"TIS GST Turnover",f"Processed ₹{tis_gst_tv:,.2f}  |  Accepted ₹{tis_gst_tv:,.2f}"),
        (f"Advance Tax Challans",f"{len(data_ais.get('tax_paid',[]))} entries"),
        (f"Refunds (AIS B4)",f"{len(data_ais.get('refunds',[]))} entries"),
        (f"TCS entries (26AS)",f"{len(data_26as.get('tcs',[]))} collectors"),
        (f"GSTR-1 Turnover (from GST Recon Excel)",f"₹{gst_turnover:,.2f}  |  Source: {gst_source}"),
    ]
    for lbl,val in auto_items:
        bg9=(ALT2 if ri9%2==0 else ALT1)
        _c(ws9,ri9,1,"→",bg9,align="center")
        _c(ws9,ri9,2,lbl,"DEEAF1",bold=True)
        ws9.merge_cells(f"C{ri9}:{get_column_letter(len(C9))}{ri9}")
        _c(ws9,ri9,3,val,bg9)
        ws9.row_dimensions[ri9].height=15; ri9+=1
    ws9.sheet_properties.tabColor="C00000"

    # ── Save ──────────────────────────────────────────────────────
    safe=re.sub(r'[\\/:*?"<>|]',"_",company_name)
    out=job_dir/f"IT_RECONCILIATION_{safe}_FY{fy.replace('-','_')}.xlsx"
    wb.save(str(out))
    sz=out.stat().st_size//1024
    _log(f"✓ Saved: {out.name} ({sz} KB)")
    _log(f"Sheets: {[s.title for s in wb.worksheets]}")
    return str(out)


# ── Empty data helpers ─────────────────────────────────────────────────────
def _empty_26as():
    return {"header":{},"deductors":[],"tcs":[],
            "summary":{"total_tds":0.0,"total_tcs":0.0}}

def _empty_ais():
    return {"header":{},"gst_turnover_monthly":{},"gst_purchases_by_supplier":[],
            "tds_income":[],"interest":[],"cash_deposits":{"total":0,"detail":[]},
            "refunds":[],"tax_paid":[],
            "summary":{"total_turnover":0.0,"total_purchases":0.0,
                       "total_tds_income":0.0,"total_interest":0.0,"total_cash_deposit":0.0,
                       "total_cdnr_credit":0.0,"total_cdnr_debit":0.0,"total_cdnr_amended":0.0}}

def _empty_tis():
    return {"header":{},"categories":{},"gst_turnover_detail":[],
            "gst_purchase_detail":[],"business_receipt_detail":[]}

def _fy_to_ay(fy):
    try: y=int(fy.split("-")[0]); return f"{y+1}-{str(y+2)[2:]}"
    except: return ""


# ── Standalone ─────────────────────────────────────────────────────────────
if __name__=="__main__":
    folder=sys.argv[1] if len(sys.argv)>1 else "."
    name  =sys.argv[2] if len(sys.argv)>2 else "Test Company"
    pan   =sys.argv[3] if len(sys.argv)>3 else "AAAAA0000A"
    gstin =sys.argv[4] if len(sys.argv)>4 else "33AAAAA0000A1ZX"
    fy    =sys.argv[5] if len(sys.argv)>5 else "2024-25"
    gst_ov=None
    if "--gst" in sys.argv:
        idx=sys.argv.index("--gst")
        if idx+1<len(sys.argv):
            try: gst_ov=float(sys.argv[idx+1].replace(",",""))
            except: pass
    print(f"\nIT Recon Engine v4 → {folder}\n")
    out=write_it_reconciliation(folder,name,pan,gstin,fy,gst_turnover_override=gst_ov)
    print(f"Output: {out}")
